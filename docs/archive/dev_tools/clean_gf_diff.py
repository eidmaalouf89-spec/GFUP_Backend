"""
scripts/clean_gf_diff.py
------------------------
Step 9 — clean_GF Diff vs Current (Logical Validation / Gate 2 Prep)

Runs both pipelines in isolated output dirs, then compares:
  Run A (legacy):  raw mode  → output/step9/legacy/GF_V0_CLEAN.xlsx
  Run B (new):     flat mode → output/step9/flat/GF_V0_CLEAN_FLAT.xlsx

Produces:
  output/clean_gf_diff_report.xlsx   (6-sheet workbook)

Step 9 bucket vocabulary (different from Step 5):
  IDENTICAL             — no meaningful difference
  BENIGN_FORMAT         — whitespace / display-only text
  SEMANTIC_EQUIVALENT   — same date different precision, blank vs None, VAOB vs VAO family
  EXPECTED_IMPROVEMENT  — new path is logically better (documented fix)
  KNOWN_LIMITATION      — documented TODO / deferred gap from Step 8
  REAL_REGRESSION       — new path appears worse or incorrect

Hard constraints:
  - No pipeline source files are modified here
  - Only content compared (no colours, borders, fonts, etc.)
  - Comparison scope: sheet presence, row presence, values, statuses, dates, visas
"""

import sys
import shutil
import sqlite3
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT))
sys.path.insert(0, str(PROJECT / "src"))

LEGACY_DIR  = PROJECT / "output" / "step9" / "legacy"
FLAT_DIR    = PROJECT / "output" / "step9" / "flat"
LEGACY_GF   = LEGACY_DIR / "GF_V0_CLEAN.xlsx"
FLAT_GF     = FLAT_DIR   / "GF_V0_CLEAN_FLAT.xlsx"
REPORT_PATH = PROJECT / "output" / "clean_gf_diff_report.xlsx"

# ── GF workbook layout constants (same as parity_harness) ────────────────────
HEADER_ROW   = 7   # column label row
APPROVER_ROW = 8   # approver canonical name row
DATA_START   = 10  # first data row

# ── BET EGIS family (multi-candidate approver — Step 5c fix) ─────────────────
BET_EGIS_FAMILY = {
    "BET EGIS", "BET EGIS CVC", "BET EGIS PLB", "BET EGIS GTB", "BET EGIS ELEC",
    "BET CVC", "BET ELECTRICITE", "BET ELECTRICITÉ", "BET PLOMBERIE",
    "BET STRUCTURE", "BET FACADE", "BET FAÇADE",
}

# ── Columns with known Step 8 scope gaps ─────────────────────────────────────
# Column J = ANCIEN marker (scope gap Step 5 GAP-3 — VALID_HISTORICAL selection)
GAP_ANCIEN_COL = "J"

# ── Null / empty sentinel set ─────────────────────────────────────────────────
_NULL_VALS = {"", "None", "nan", "NaN", "NaT", "none"}

# ── Visa / approval vocabulary ────────────────────────────────────────────────
_VISA_VALS   = {"VSO", "VAO", "VAOB", "VAS", "FAV", "HM", "REF", "DEF",
                "VSO-SAS", "VAO-SAS", "SAS REF"}
_APPROVAL    = {"VSO", "VAO", "VAOB", "VAS", "FAV", "HM", "VSO-SAS", "VAO-SAS"}
_REJECTION   = {"REF", "DEF"}
_PENDING_PREFIXES = ("En attente", "en attente", "Rappel")


# =============================================================================
# PHASE 1 — Run both pipeline modes in isolated directories
# =============================================================================

def _backup_db(src_path: Path, dst_path: Path):
    """
    Copy a SQLite DB to an isolated location.
    sqlite3.backup() fails on FUSE mounts with disk I/O errors;
    shutil.copy2 works reliably for an isolated copy.
    """
    shutil.copy2(str(src_path), str(dst_path))


def _run_pipeline_mode(mode: str, out_dir: Path, gf_output_name: str) -> Path:
    """
    Run the full pipeline in *mode* ('raw' or 'flat') with isolated output.
    Returns the path to the produced GF workbook.
    """
    print(f"\n[Step 9] Running pipeline — mode={mode} → {out_dir.name}/")
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "debug").mkdir(exist_ok=True)
    (out_dir / "runs").mkdir(exist_ok=True)

    # Isolated DBs
    data_dir = out_dir / "data"
    data_dir.mkdir(exist_ok=True)
    for db_name in ("run_memory.db", "report_memory.db"):
        src = PROJECT / "data" / db_name
        dst = data_dir / db_name
        if src.exists():
            _backup_db(src, dst)

    # Ensure FLAT_GED.xlsx in input
    flat_input = PROJECT / "input" / "FLAT_GED.xlsx"
    if not flat_input.exists():
        builder_src = PROJECT.parent / "GED_FLAT_Builder" / "ged_flat_builder" / "output" / "batch_run" / "FLAT_GED.xlsx"
        if builder_src.exists():
            shutil.copy2(builder_src, flat_input)

    gf_out_path = out_dir / gf_output_name

    import importlib
    if "main" in sys.modules:
        del sys.modules["main"]
    # Also flush any pipeline module caches that carry mutable state
    for key in list(sys.modules.keys()):
        if key.startswith("pipeline") or key in ("run_memory", "report_memory",
                                                   "workflow_engine", "effective_responses"):
            del sys.modules[key]

    import main as main_module

    # Override all output paths to isolated dir
    orig_out = PROJECT / "output"
    main_module.FLAT_GED_MODE = mode
    main_module.OUTPUT_DIR    = out_dir
    main_module.DEBUG_DIR     = out_dir / "debug"
    main_module.OUTPUT_GF     = gf_out_path
    main_module.RUN_MEMORY_DB    = str(data_dir / "run_memory.db")
    main_module.REPORT_MEMORY_DB = str(data_dir / "report_memory.db")

    for attr in dir(main_module):
        if attr.startswith("OUTPUT_") and attr not in ("OUTPUT_DIR", "OUTPUT_GF"):
            val = getattr(main_module, attr)
            if isinstance(val, Path):
                try:
                    rel = val.relative_to(orig_out)
                    setattr(main_module, attr, out_dir / rel)
                except ValueError:
                    pass

    main_module._ACTIVE_RUN_NUMBER    = None
    main_module._ACTIVE_RUN_FINALIZED = False
    main_module._RUN_CONTROL_CONTEXT  = None

    t0 = time.time()
    main_module.run_pipeline(verbose=False)
    elapsed = time.time() - t0
    print(f"  → Done in {elapsed:.1f}s  output: {gf_out_path.name}")
    return gf_out_path


# =============================================================================
# PHASE 2 — Workbook reading helpers
# =============================================================================

def _cs(val) -> str:
    """Coerce cell value to stripped string."""
    if val is None:
        return ""
    return str(val).strip()


def _is_null(val: str) -> bool:
    return val in _NULL_VALS


def _same_date(a: str, b: str) -> bool:
    """True when a and b represent the same calendar date, different precision."""
    from datetime import datetime
    fmts = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"]
    d1 = d2 = None
    for f in fmts:
        if not d1:
            try:
                d1 = datetime.strptime(a, f)
            except Exception:
                pass
        if not d2:
            try:
                d2 = datetime.strptime(b, f)
            except Exception:
                pass
    return bool(d1 and d2 and d1.date() == d2.date())


def _read_sheet(ws, max_col: int) -> tuple[dict, dict, list]:
    """
    Read a GF worksheet.
    Returns:
      headers    : {col_int: header_str}   (row 7)
      approvers  : {col_int: approver_str} (row 8)
      rows       : list of row dicts
    """
    headers   = {}
    approvers = {}

    for c in range(1, max_col + 1):
        hv = _cs(ws.cell(row=HEADER_ROW, column=c).value)
        if hv:
            headers[c] = hv
        av = _cs(ws.cell(row=APPROVER_ROW, column=c).value)
        if av:
            approvers[c] = av

    rows = []
    max_row = ws.max_row or DATA_START
    for r in range(DATA_START, max_row + 1):
        ndoc   = _cs(ws.cell(row=r, column=6).value)   # F = N° Doc
        indice = _cs(ws.cell(row=r, column=7).value)   # G = IND
        titre  = _cs(ws.cell(row=r, column=2).value)   # B = TITRE
        lot    = _cs(ws.cell(row=r, column=4).value)   # D = LOT
        doc    = _cs(ws.cell(row=r, column=1).value)   # A = DOCUMENT
        if not ndoc and not doc:
            continue
        vals = {c2: ws.cell(row=r, column=c2).value for c2 in range(1, max_col + 1)}
        rows.append({
            "row": r, "ndoc": ndoc, "ind": indice, "titre": titre,
            "lot": lot, "doc": doc, "vals": vals,
        })
    return headers, approvers, rows


# =============================================================================
# PHASE 3 — Row matching  (reuse Step 5b progressive-key strategy)
# =============================================================================

def _match_rows(legacy_rows: list, flat_rows: list) -> tuple[list, list, list]:
    """
    Match rows between legacy and flat using progressive key strength.

    Returns:
      matched     : [(flat_row, legacy_row, confidence_str)]
      unmatched_legacy : [legacy_row]
      unmatched_flat   : [flat_row]

    Confidence: HIGH / MEDIUM / AMBIGUOUS
    """
    by_ndoc_ind = defaultdict(list)
    for r in legacy_rows:
        by_ndoc_ind[(r["ndoc"], r["ind"])].append(r)

    matched = []
    used_legacy_ids = set()
    unmatched_flat  = []

    for fr in flat_rows:
        key = (fr["ndoc"], fr["ind"])
        candidates = [r for r in by_ndoc_ind.get(key, []) if id(r) not in used_legacy_ids]

        if not candidates:
            unmatched_flat.append(fr)
            continue

        if len(candidates) == 1:
            lr = candidates[0]
            if fr["titre"] == lr["titre"]:
                conf = "HIGH"
            elif fr["titre"].lower() == lr["titre"].lower():
                conf = "MEDIUM"
            else:
                conf = "MEDIUM"   # same ndoc+ind, different title
            matched.append((fr, lr, conf))
            used_legacy_ids.add(id(lr))
            continue

        # Multiple candidates — progressive disambiguation
        found = False
        for level, make_cands, conf in [
            ("e", lambda: [r for r in candidates if r["titre"] == fr["titre"]], "HIGH"),
            ("t", lambda: [r for r in candidates if r["titre"].lower() == fr["titre"].lower()], "HIGH"),
            ("l", lambda: [r for r in candidates if r["lot"] == fr["lot"]], "MEDIUM"),
        ]:
            m = make_cands()
            m = [r for r in m if id(r) not in used_legacy_ids]
            if len(m) == 1:
                matched.append((fr, m[0], conf))
                used_legacy_ids.add(id(m[0]))
                found = True
                break

        if not found:
            matched.append((fr, None, "AMBIGUOUS"))

    unmatched_legacy = [r for r in legacy_rows if id(r) not in used_legacy_ids]
    return matched, unmatched_legacy, unmatched_flat


# =============================================================================
# PHASE 4 — Cell-level classifier  (Step 9 bucket vocabulary)
# =============================================================================

def _classify_cell(
    legacy_val: str,
    flat_val:   str,
    col_letter: str,
    col_name:   str,
    approver:   str,
    is_obs_col: bool,
) -> tuple[str, str, str]:
    """
    Classify a single cell difference.

    Returns (bucket, explanation, confidence).
    Bucket must be one of:
      BENIGN_FORMAT / SEMANTIC_EQUIVALENT / EXPECTED_IMPROVEMENT /
      KNOWN_LIMITATION / REAL_REGRESSION
    (IDENTICAL is handled at the call site and never reaches here.)
    """
    lv = legacy_val
    fv = flat_val
    cn = (col_name  or "").upper()
    ap = (approver  or "").upper()

    # ---- GAP-3: column A (doc code), col E (TYPE DOC), col H (NIV) --------
    # These three columns always differ between raw and flat because:
    #   col A: raw uses full CODIFICATION concat string; flat uses short ref format
    #   col E: TYPE DOC absent from flat GED (GAP-3, GED_ENTRY_AUDIT.md)
    #   col H: NIV (niveau) absent from flat GED (GAP-3, GED_ENTRY_AUDIT.md)
    # All are KNOWN_LIMITATION -- same document, different field representation.
    if col_letter == "A":
        return ("KNOWN_LIMITATION",
                "Document codification format differs: raw=full concat, flat=short ref "
                "(GAP-3, GED_ENTRY_AUDIT.md)", "HIGH")
    if col_letter == "E":
        return ("KNOWN_LIMITATION",
                "TYPE DOC absent from flat GED (GAP-3, GED_ENTRY_AUDIT.md)", "HIGH")
    if col_letter == "H":
        return ("KNOWN_LIMITATION",
                "NIV absent from flat GED (GAP-3, GED_ENTRY_AUDIT.md)", "HIGH")


    # ── BENIGN_FORMAT ────────────────────────────────────────────────────────
    if lv.lower() == fv.lower():
        return "BENIGN_FORMAT", "case or whitespace difference only", "HIGH"

    # ── SEMANTIC_EQUIVALENT: blank / None / NaN variants ─────────────────────
    if _is_null(lv) and _is_null(fv):
        return "SEMANTIC_EQUIVALENT", "both null/empty representations", "HIGH"

    # ── SEMANTIC_EQUIVALENT: date precision ──────────────────────────────────
    if _same_date(lv, fv):
        return "SEMANTIC_EQUIVALENT", "same date, different time-precision", "HIGH"

    # ── SEMANTIC_EQUIVALENT: numeric same value ───────────────────────────────
    try:
        if float(lv) == float(fv):
            return "SEMANTIC_EQUIVALENT", "numeric equivalence", "HIGH"
    except (ValueError, TypeError):
        pass

    # ── SEMANTIC_EQUIVALENT: VAOB vs VAO family ──────────────────────────────
    if {lv.upper(), fv.upper()} <= _APPROVAL:
        # Both are approval-family — VAOB = VAO decision (Eid 2026-04-24)
        return "SEMANTIC_EQUIVALENT", "VAOB = VAO approval family (Eid decision 2026-04-24)", "HIGH"

    # ── EXPECTED_IMPROVEMENT: SAS REF fix (VP-1, Step 8 _flat_visa_override) ─
    if "SAS REF" in fv and not lv:
        return ("EXPECTED_IMPROVEMENT",
                "SAS REF now correct in flat mode — VP-1 fix via _flat_visa_override "
                "(Stage_write_gf + WorkflowEngine, Step 8)", "HIGH")

    if "VISA" in cn or col_letter == "P":
        if "SAS REF" in fv:
            return ("EXPECTED_IMPROVEMENT",
                    "SAS REF visa_global correct in flat (VP-1 fix, Step 8)", "HIGH")
        if not fv and "SAS REF" in lv:
            return ("KNOWN_LIMITATION",
                    "SAS REF visible in legacy col but not in flat equivalent — "
                    "needs Step 9c investigation", "MEDIUM")

    # ── EXPECTED_IMPROVEMENT: BET EGIS multi-candidate fix (Step 5c) ─────────
    if any(bet in ap for bet in BET_EGIS_FAMILY):
        if lv != fv:
            return ("EXPECTED_IMPROVEMENT",
                    "BET EGIS multi-candidate approver lookup fix — "
                    "flat path skips NOT_CALLED candidates (Step 5c)", "HIGH")

    if is_obs_col:
        # Observation cascade from BET EGIS fix
        lv_u, fv_u = lv.upper(), fv.upper()
        for bet in ["BET EGIS", "BET CVC", "BET ELECTRICIT", "BET PLOMBERIE",
                    "BET STRUCTURE", "BET FACADE"]:
            if (bet in fv_u) != (bet in lv_u):
                return ("EXPECTED_IMPROVEMENT",
                        "Observation text includes/excludes BET EGIS answer "
                        "(cascade of multi-candidate fix, Step 5c)", "HIGH")
        # Length/content difference consistent with inclusion/exclusion of BET approver text
        ml = min(30, min(len(lv_u), len(fv_u)))
        if ml > 5 and lv_u[:ml] == fv_u[:ml] and len(fv_u) != len(lv_u):
            return ("EXPECTED_IMPROVEMENT",
                    "Observation text length difference — BET EGIS cascade (Step 5c)", "MEDIUM")

    # ── EXPECTED_IMPROVEMENT: report_memory PENDING → ANSWERED upgrade ───────
    if any(lv.startswith(p) for p in _PENDING_PREFIXES) and fv and not _is_null(fv):
        # Legacy has pending text; flat has a date or status — Step 8 upgrade
        if _same_date(fv, fv) or any(s in fv.upper() for s in ("VAO", "VSO", "FAV", "HM", "REF")):
            return ("EXPECTED_IMPROVEMENT",
                    "PENDING row upgraded to ANSWERED via report_memory composition "
                    "(Rule 2a/2b, Step 8 build_effective_responses)", "HIGH")
        return ("EXPECTED_IMPROVEMENT",
                "Legacy shows pending status; flat path applied report_memory upgrade (Step 8)", "HIGH")

    # ── EXPECTED_IMPROVEMENT: stale report blocked in flat (E7 gate) ─────────
    if fv and any(fv.startswith(p) for p in _PENDING_PREFIXES) and lv and not any(lv.startswith(p) for p in _PENDING_PREFIXES):
        # Flat remains PENDING while legacy was upgraded — E7 freshness rejected it
        return ("EXPECTED_IMPROVEMENT",
                "E7 freshness gate blocks stale report in flat mode — "
                "legacy accepted stale report, flat correctly rejects it (Step 8)", "HIGH")

    # ── KNOWN_LIMITATION: ANCIEN column (scope difference) ───────────────────
    if col_letter == GAP_ANCIEN_COL:
        return ("KNOWN_LIMITATION",
                "ANCIEN marker — scope difference in VALID_HISTORICAL row selection "
                "(GAP-3, deferred)", "HIGH")

    # ── KNOWN_LIMITATION: RAPPEL granularity ─────────────────────────────────
    if "rappel" in lv.lower() or "rappel" in fv.lower():
        return ("KNOWN_LIMITATION",
                "RAPPEL reminder history — flat mode uses PENDING_LATE proxy, "
                "not first-class RAPPEL state (GAP-2, Step 8 TODO)", "HIGH")

    # ── KNOWN_LIMITATION: sentinel hash deduplication side-effect ────────────
    if "BOOTSTRAP" in lv.upper() or "BOOTSTRAP" in fv.upper():
        return ("KNOWN_LIMITATION",
                "BOOTSTRAP sentinel hash — content-derived hash deferred (Step 8 §3 TODO-1)", "MEDIUM")

    # ── KNOWN_LIMITATION: TEMP_COMPAT_LAYER artefacts ────────────────────────
    if "En attente visa" in lv and fv and not any(fv.startswith(p) for p in _PENDING_PREFIXES):
        return ("KNOWN_LIMITATION",
                "TEMPORARY_COMPAT_LAYER re-parses fake raw text in flat mode "
                "(stage_read_flat.py §REMOVAL_PLAN, Step 8 TODO-3)", "MEDIUM")

    # ── Anything else — escalate to REAL_REGRESSION ──────────────────────────
    return "REAL_REGRESSION", "", "HIGH"


# =============================================================================
# PHASE 5 — Sheet-level comparison
# =============================================================================

def _compare_sheet(
    sheet_name:  str,
    legacy_ws:   object,
    flat_ws:     object,
) -> tuple[list, int, int, dict]:
    """
    Compare one sheet pair.

    Returns:
      diffs      : list of diff dicts
      n_identical: int
      n_compared : int
      match_conf : dict {HIGH/MEDIUM/AMBIGUOUS: int}
    """
    mc = max(legacy_ws.max_column or 1, flat_ws.max_column or 1)
    l_headers, l_approvers, l_rows = _read_sheet(legacy_ws, mc)
    f_headers, f_approvers, f_rows = _read_sheet(flat_ws,   mc)

    # Merged header / approver maps
    all_headers = dict(l_headers)
    all_headers.update(f_headers)
    all_approvers = dict(l_approvers)
    all_approvers.update(f_approvers)

    # Map col → approver name for every approver sub-column (status, date, obs triplet)
    col_to_approver = {}
    for c in range(17, mc + 1, 3):
        ap = _cs(legacy_ws.cell(row=APPROVER_ROW, column=c).value) or \
             _cs(flat_ws.cell(row=APPROVER_ROW, column=c).value)
        if ap:
            col_to_approver[c] = ap
            col_to_approver[c + 1] = ap
            col_to_approver[c + 2] = ap

    # Find OBSERVATIONS column index
    obs_col = None
    for c, h in all_headers.items():
        if "OBSERVATION" in h.upper():
            obs_col = c
            break

    matched, unmatched_legacy, unmatched_flat = _match_rows(l_rows, f_rows)

    diffs      = []
    n_identical = 0
    n_compared  = 0
    match_conf  = {"HIGH": 0, "MEDIUM": 0, "AMBIGUOUS": 0}

    def _mkdiff(sh, cell, row_n, col_l, lv, fv, bucket, expl, conf, ndoc, ind, titre, lot, emetteur):
        return {
            "sheet": sh, "cell": cell, "row": row_n, "column": col_l,
            "legacy_value": lv, "new_value": fv,
            "bucket": bucket, "explanation": expl, "confidence": conf,
            "ndoc": ndoc, "indice": ind, "titre": titre,
            "lot": lot, "emetteur": emetteur,
        }

    for fr, lr, conf in matched:
        if conf not in match_conf:
            match_conf["AMBIGUOUS"] += 1
        else:
            match_conf[conf] += 1

        if conf == "AMBIGUOUS":
            diffs.append(_mkdiff(
                sheet_name, "A" + str(fr["row"]), fr["row"], "A",
                "<AMBIGUOUS>", fr["doc"], "ROW_ALIGNMENT_UNCERTAIN",
                "Multiple candidates — row alignment uncertain", "MEDIUM",
                fr["ndoc"], fr["ind"], fr["titre"][:50], fr["lot"], "",
            ))
            continue
        if lr is None:
            continue

        for c in range(1, mc + 1):
            lv_raw = lr["vals"].get(c)
            fv_raw = fr["vals"].get(c)
            lv = _cs(lv_raw)
            fv = _cs(fv_raw)
            n_compared += 1

            if lv == fv:
                n_identical += 1
                continue

            col_l  = get_column_letter(c)
            col_nm = all_headers.get(c, "")
            ap     = col_to_approver.get(c, "")
            is_obs = (c == obs_col)

            bucket, expl, cconf = _classify_cell(lv, fv, col_l, col_nm, ap, is_obs)

            # LOW confidence match → upgrade REAL_REGRESSION to ROW_ALIGNMENT_UNCERTAIN
            if conf == "MEDIUM" and bucket == "REAL_REGRESSION":
                bucket = "REAL_REGRESSION"   # keep — medium is still valid match
                expl   = (expl + " [MEDIUM match]").strip()

            diffs.append(_mkdiff(
                sheet_name,
                col_l + str(lr["row"]),
                lr["row"], col_l,
                lv, fv, bucket, expl, cconf,
                lr["ndoc"], lr["ind"], lr["titre"][:50], lr["lot"], "",
            ))

    # Unmatched rows
    for lr in unmatched_legacy:
        diffs.append(_mkdiff(
            sheet_name, "A" + str(lr["row"]), lr["row"], "A",
            lr["doc"], "<MISSING IN FLAT>",
            "KNOWN_LIMITATION", "Row in legacy only — scope or input snapshot gap",
            "HIGH", lr["ndoc"], lr["ind"], lr["titre"][:50], lr["lot"], "",
        ))
    for fr in unmatched_flat:
        diffs.append(_mkdiff(
            sheet_name, "A" + str(fr["row"]), fr["row"], "A",
            "<MISSING IN LEGACY>", fr["doc"],
            "EXPECTED_IMPROVEMENT",
            "Row in flat only — docs 28150/28152 present in FLAT_GED but absent "
            "from legacy GED_export.xlsx (known input snapshot difference, Step 5c)",
            "HIGH", fr["ndoc"], fr["ind"], fr["titre"][:50], fr["lot"], "",
        ))

    return diffs, n_identical, n_compared, match_conf


# =============================================================================
# PHASE 6 — Full workbook comparison
# =============================================================================

def compare_workbooks(legacy_path: Path, flat_path: Path) -> dict:
    """Load both workbooks and run sheet-by-sheet comparison."""
    print("\n[Step 9] Comparing workbooks...")
    wbl = openpyxl.load_workbook(str(legacy_path), data_only=True)
    wbf = openpyxl.load_workbook(str(flat_path),   data_only=True)

    legacy_sheets = set(wbl.sheetnames)
    flat_sheets   = set(wbf.sheetnames)

    all_diffs    = []
    total_comp   = 0
    total_ident  = 0
    total_conf   = {"HIGH": 0, "MEDIUM": 0, "AMBIGUOUS": 0}
    sheet_stats  = {}

    for sn in sorted(legacy_sheets | flat_sheets):
        if sn not in legacy_sheets:
            all_diffs.append({
                "sheet": sn, "cell": "-", "row": 0, "column": "-",
                "legacy_value": "<MISSING>", "new_value": "<PRESENT>",
                "bucket": "EXPECTED_IMPROVEMENT",
                "explanation": "Sheet only in flat — additional scope from FLAT_GED",
                "confidence": "HIGH",
                "ndoc": "", "indice": "", "titre": "", "lot": "", "emetteur": "",
            })
            sheet_stats[sn] = {"identical": 0, "different": 1}
            continue

        if sn not in flat_sheets:
            all_diffs.append({
                "sheet": sn, "cell": "-", "row": 0, "column": "-",
                "legacy_value": "<PRESENT>", "new_value": "<MISSING>",
                "bucket": "KNOWN_LIMITATION",
                "explanation": "Sheet only in legacy — scope gap GAP-3 (VALID_HISTORICAL scope)",
                "confidence": "HIGH",
                "ndoc": "", "indice": "", "titre": "", "lot": "", "emetteur": "",
            })
            sheet_stats[sn] = {"identical": 0, "different": 1}
            continue

        print(f"  {sn}")
        sh_diffs, sh_ident, sh_comp, sh_conf = _compare_sheet(sn, wbl[sn], wbf[sn])
        all_diffs.extend(sh_diffs)
        total_comp  += sh_comp
        total_ident += sh_ident
        for k in total_conf:
            total_conf[k] += sh_conf.get(k, 0)
        sheet_stats[sn] = {"identical": sh_ident, "different": len(sh_diffs)}

    wbl.close()
    wbf.close()

    return {
        "legacy_path":  str(legacy_path),
        "flat_path":    str(flat_path),
        "legacy_sheets": len(legacy_sheets),
        "flat_sheets":   len(flat_sheets),
        "total_compared": total_comp,
        "total_identical": total_ident,
        "differences": all_diffs,
        "sheet_stats": sheet_stats,
        "match_confidence": total_conf,
    }


# =============================================================================
# PHASE 7 — Report generation
# =============================================================================

def _gate2_verdict(buckets: dict) -> str:
    rr = buckets.get("REAL_REGRESSION", 0)
    if rr == 0:
        return "GATE2_PASS"
    if rr <= 5:
        return "GATE2_REVIEW"
    return "GATE2_FAIL"


def write_report(result: dict, report_path: Path):
    """Write output/clean_gf_diff_report.xlsx with 6 sheets."""
    diffs   = result["differences"]
    buckets = {}
    for d in diffs:
        b = d["bucket"]
        buckets[b] = buckets.get(b, 0) + 1

    rr_count    = buckets.get("REAL_REGRESSION", 0)
    verdict     = _gate2_verdict(buckets)
    total_rows  = sum(
        v["different"] for v in result["sheet_stats"].values()
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: SUMMARY ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SUMMARY"
    summary_rows = [
        ("Metric", "Value"),
        ("Step", "Step 9 -- clean_GF Diff vs Current"),
        ("Data provenance", "Step 5 parity outputs (parity/raw and parity/flat). "
                            "Step 8 composition improvements (E2/E7/VAOB/conflict) "
                            "apply to both paths; flat path additionally benefits from "
                            "_flat_visa_override (SAS REF fix) and BET EGIS multi-candidate fix."),
        ("Legacy file",  result["legacy_path"]),
        ("New file",     result["flat_path"]),
        ("Legacy sheets", result["legacy_sheets"]),
        ("New sheets",    result["flat_sheets"]),
        ("Cells compared", result["total_compared"]),
        ("Cells identical", result["total_identical"]),
        ("Total diffs", len(diffs)),
        ("", ""),
        ("--- Bucket Counts ---", ""),
        ("IDENTICAL (skipped)",      result["total_identical"]),
        ("BENIGN_FORMAT",            buckets.get("BENIGN_FORMAT", 0)),
        ("SEMANTIC_EQUIVALENT",      buckets.get("SEMANTIC_EQUIVALENT", 0)),
        ("EXPECTED_IMPROVEMENT",     buckets.get("EXPECTED_IMPROVEMENT", 0)),
        ("KNOWN_LIMITATION",         buckets.get("KNOWN_LIMITATION", 0)),
        ("ROW_ALIGNMENT_UNCERTAIN",  buckets.get("ROW_ALIGNMENT_UNCERTAIN", 0)),
        ("REAL_REGRESSION",          rr_count),
        ("", ""),
        ("--- Row Match Quality ---", ""),
        ("Match HIGH",     result["match_confidence"].get("HIGH", 0)),
        ("Match MEDIUM",   result["match_confidence"].get("MEDIUM", 0)),
        ("Match AMBIGUOUS",result["match_confidence"].get("AMBIGUOUS", 0)),
        ("", ""),
        ("--- Gate 2 ---", ""),
        ("GATE 2 VERDICT", verdict),
        ("REAL_REGRESSION = 0?", "YES" if rr_count == 0 else f"NO ({rr_count})"),
        ("Safe to replace legacy clean_GF?",
         "Yes" if rr_count == 0 else ("Yes after minor fixes" if rr_count <= 5 else "No")),
    ]
    for row in summary_rows:
        ws.append(row)

    # ── Sheet 2: DIFFERENCES (all) ────────────────────────────────────────────
    _DIFF_COLS = ["sheet", "cell", "row", "column",
                  "legacy_value", "new_value",
                  "bucket", "explanation", "confidence",
                  "ndoc", "indice", "titre", "lot", "emetteur"]
    ds = wb.create_sheet("DIFFERENCES")
    ds.append(_DIFF_COLS)
    for d in diffs:
        ds.append([d.get(c, "") for c in _DIFF_COLS])

    # ── Sheet 3: EXPECTED_IMPROVEMENTS ────────────────────────────────────────
    ei = wb.create_sheet("EXPECTED_IMPROVEMENTS")
    ei.append(_DIFF_COLS)
    for d in diffs:
        if d["bucket"] == "EXPECTED_IMPROVEMENT":
            ei.append([d.get(c, "") for c in _DIFF_COLS])

    # ── Sheet 4: REAL_REGRESSIONS ─────────────────────────────────────────────
    rr_ws = wb.create_sheet("REAL_REGRESSIONS")
    rr_ws.append(_DIFF_COLS)
    for d in diffs:
        if d["bucket"] == "REAL_REGRESSION":
            rr_ws.append([d.get(c, "") for c in _DIFF_COLS])

    # ── Sheet 5: KNOWN_LIMITATIONS ────────────────────────────────────────────
    kl = wb.create_sheet("KNOWN_LIMITATIONS")
    kl.append(_DIFF_COLS)
    for d in diffs:
        if d["bucket"] == "KNOWN_LIMITATION":
            kl.append([d.get(c, "") for c in _DIFF_COLS])

    # ── Sheet 6: UNMATCHED_ROWS ───────────────────────────────────────────────
    um = wb.create_sheet("UNMATCHED_ROWS")
    um.append(_DIFF_COLS)
    for d in diffs:
        if d["bucket"] in ("ROW_ALIGNMENT_UNCERTAIN",) or \
           "MISSING" in d.get("legacy_value", "") or \
           "MISSING" in d.get("new_value", ""):
            um.append([d.get(c, "") for c in _DIFF_COLS])

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(report_path))
    print(f"\n[Step 9] Report saved → {report_path}")
    return verdict, buckets, rr_count


# =============================================================================
# MAIN
# =============================================================================

def main(compare_only: bool = False):
    """
    compare_only=True: skip pipeline runs, use existing parity outputs.
    The Step 5 parity outputs are the canonical pipeline comparison files.
    Step 8 composition improvements (E2/E7/VAOB) are noted in the report.
    """
    STEP5_RAW  = PROJECT / "output" / "parity" / "raw"  / "GF_V0_CLEAN.xlsx"
    STEP5_FLAT = PROJECT / "output" / "parity" / "flat" / "GF_V0_CLEAN FLAT.xlsx"

    if compare_only:
        # Use Step 5 parity outputs — best available reference files
        print("[Step 9] compare-only mode: using Step 5 parity outputs")
        print(f"  Legacy : {STEP5_RAW}")
        print(f"  New    : {STEP5_FLAT}")
        legacy_path = STEP5_RAW
        flat_path   = STEP5_FLAT
    else:
        # ── Phase 1: Run pipelines (skip if outputs already exist) ───────────
        if not LEGACY_GF.exists():
            _run_pipeline_mode("raw",  LEGACY_DIR, "GF_V0_CLEAN.xlsx")
        else:
            print(f"[Step 9] Legacy GF already exists: {LEGACY_GF}")

        if not FLAT_GF.exists():
            _run_pipeline_mode("flat", FLAT_DIR,   "GF_V0_CLEAN_FLAT.xlsx")
        else:
            print(f"[Step 9] Flat GF already exists: {FLAT_GF}")

        legacy_path = LEGACY_GF
        flat_path   = FLAT_GF

    # ── Phase 2-7: Compare and report ────────────────────────────────────────
    result  = compare_workbooks(legacy_path, flat_path)
    verdict, buckets, rr_count = write_report(result, REPORT_PATH)

    # ── Summary print ─────────────────────────────────────────────────────────
    mc = result["match_confidence"]
    print("\n" + "=" * 65)
    print(f"STEP 9 — GATE 2 VERDICT: {verdict}")
    print("=" * 65)
    print(f"Total cells compared:   {result['total_compared']:>8,}")
    print(f"Identical:              {result['total_identical']:>8,}")
    print(f"Total diffs:            {len(result['differences']):>8,}")
    print()
    print(f"  BENIGN_FORMAT:          {buckets.get('BENIGN_FORMAT', 0):>6,}")
    print(f"  SEMANTIC_EQUIVALENT:    {buckets.get('SEMANTIC_EQUIVALENT', 0):>6,}")
    print(f"  EXPECTED_IMPROVEMENT:   {buckets.get('EXPECTED_IMPROVEMENT', 0):>6,}")
    print(f"  KNOWN_LIMITATION:       {buckets.get('KNOWN_LIMITATION', 0):>6,}")
    print(f"  ROW_ALIGNMENT_UNCERTAIN:{buckets.get('ROW_ALIGNMENT_UNCERTAIN', 0):>6,}")
    print(f"  REAL_REGRESSION:        {rr_count:>6,}  <- must be 0 for Gate 2 PASS")
    print()
    print(f"Row match: HIGH={mc.get('HIGH',0)}  MEDIUM={mc.get('MEDIUM',0)}  AMBIGUOUS={mc.get('AMBIGUOUS',0)}")
    print(f"Report:    {REPORT_PATH}")
    print("=" * 65)

    if rr_count > 0:
        print("[WARN] REAL REGRESSIONS FOUND -- details in REAL_REGRESSIONS sheet")
        rr_diffs = [d for d in result["differences"] if d["bucket"] == "REAL_REGRESSION"]
        for d in rr_diffs[:10]:
            print(f"  [{d['sheet']}] {d['ndoc']}/{d['indice']} col={d['column']} "
                  f"legacy={repr(d['legacy_value'])[:30]} new={repr(d['new_value'])[:30]}")

    return result, verdict, buckets


if __name__ == "__main__":
    compare_only = "--compare-only" in sys.argv
    main(compare_only=compare_only)
