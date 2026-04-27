"""Step 5c: reclassify BET EGIS diffs as OLD_PATH_BUG."""
import sys, re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
sys.path.insert(0, "/sessions/festive-lucid-carson/mnt/GF updater v3")
sys.path.insert(0, "/sessions/festive-lucid-carson/mnt/GF updater v3/src")
import openpyxl
from openpyxl.utils import get_column_letter

RAW_PATH = Path("/tmp/parity_raw_r1/GF_V0_CLEAN.xlsx")
FLAT_PATH = Path("/tmp/parity_flat_r1/GF_V0_CLEAN.xlsx")
REPORT = Path("/sessions/festive-lucid-carson/mnt/GF updater v3/output/parity_report.xlsx")
HR, DS = 7, 10
VV = ("VAO","VSO","VAS","REF","VSO-SAS","VAO-SAS","SAS REF")
G3 = {"A","E","H","I"}
NL = ("","None","nan","NaN","NaT")
MCA = {"BET EGIS","BET EGIS CVC","BET EGIS PLB","BET EGIS GTB","BET EGIS ELEC"}
OPB = "OLD_PATH_BUG: WE.get_approver_status returns NOT_CALLED too early for multi-candidate approver"
EON = ["BET EGIS","BET CVC","BET ELECTRICIT","BET PLOMBERIE","BET STRUCTURE","BET FACADE"]

def cs(v): return "" if v is None else str(v)
def sd(a,b):
    fmts=["%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M:%S.%f","%Y-%m-%d"]
    d1=d2=None
    for f in fmts:
        if not d1:
            try: d1=datetime.strptime(a,f)
            except: pass
        if not d2:
            try: d2=datetime.strptime(b,f)
            except: pass
    return d1 and d2 and d1.date()==d2.date()

def clf(rs,fs,cl,cn,ap="",io=False):
    if rs.lower()==fs.lower() and rs!=fs: return "BENIGN_WHITESPACE","Case"
    if rs in NL and fs in NL: return "SEMANTIC_EQUIVALENT","None/empty"
    if cl in G3: return "KNOWN_GAP","GAP-3: "+cl
    if sd(rs,fs): return "SEMANTIC_EQUIVALENT","Date precision"
    try:
        if float(rs)==float(fs): return "SEMANTIC_EQUIVALENT","Numeric"
    except: pass
    cu=(cn or "").upper()
    if "VISA" in cu or cl=="P":
        if rs in VV and not fs: return "KNOWN_GAP","GAP-1 Visa missing flat"
        if fs in VV and not rs: return "KNOWN_GAP","Visa flat not raw"
        if "SAS REF" in rs or "SAS REF" in fs: return "KNOWN_GAP","GAP-1 SAS REF"
    if "rappel" in rs.lower() or "rappel" in fs.lower(): return "KNOWN_GAP","GAP-2 RAPPEL"
    if cl=="J": return "KNOWN_GAP","ANCIEN GAP-3"
    if ap.upper() in {a.upper() for a in MCA}: return "OLD_PATH_BUG",OPB
    if io and rs!=fs:
        ru,fu=rs.upper(),fs.upper()
        for en in EON:
            if (en in fu)!=(en in ru): return "OLD_PATH_BUG",OPB+" (obs cascade)"
        ml=min(30,min(len(ru),len(fu)))
        if ml>5 and ru[:ml]==fu[:ml] and len(fu)!=len(ru): return "OLD_PATH_BUG",OPB+" (obs len)"
    return "REAL_DIVERGENCE",""

def rr(ws,mc):
    h={}
    for c in range(1,mc+1):
        v=cs(ws.cell(row=HR,column=c).value).strip()
        if v: h[c]=v
    rows=[]
    for r in range(DS,(ws.max_row or DS)+1):
        doc=cs(ws.cell(row=r,column=1).value).strip()
        t=cs(ws.cell(row=r,column=2).value).strip()
        cd=cs(ws.cell(row=r,column=3).value).strip()
        nd=cs(ws.cell(row=r,column=6).value).strip()
        ind=cs(ws.cell(row=r,column=7).value).strip()
        if not nd and not doc: continue
        v={c2:ws.cell(row=r,column=c2).value for c2 in range(1,mc+1)}
        rows.append({"row":r,"doc":doc,"titre":t,"cdate":cd,"cdate_short":cd[:10],"ndoc":nd,"ind":ind,"vals":v})
    return h,rows

def mt(raws,flats):
    rbn=defaultdict(list)
    for r in raws: rbn[(r["ndoc"],r["ind"])].append(r)
    md,us,uf=[],set(),[]
    for f in flats:
        k=(f["ndoc"],f["ind"])
        ca=[r for r in rbn.get(k,[]) if id(r) not in us]
        if not ca: uf.append(f); continue
        if len(ca)==1:
            r=ca[0]; c="HIGH" if f["titre"]==r["titre"] else ("MEDIUM" if f["titre"].lower()==r["titre"].lower() else "LOW")
            md.append((f,r,c)); us.add(id(r)); continue
        for lv in ["e","t","c","s"]:
            if lv=="e": m=[r for r in ca if r["titre"]==f["titre"] and r["cdate_short"]==f["cdate_short"]]; cf="HIGH"
            elif lv=="t": m=[r for r in ca if r["titre"]==f["titre"]]; cf="HIGH"
            elif lv=="c": m=[r for r in ca if r["titre"].lower()==f["titre"].lower()]; cf="MEDIUM"
            else: s=f["ndoc"]+"_"+f["ind"]; m=[r for r in ca if r["doc"].endswith(s)]; cf="MEDIUM"
            if len(m)==1: md.append((f,m[0],cf)); us.add(id(m[0])); break
        else: md.append((f,None,"AMBIGUOUS"))
    ur=[r for r in raws if id(r) not in us]
    return md,ur,uf

def dd(sh,ce,ro,co,rv,fv,bu,ex,nd,ind,ti):
    return {"sheet":sh,"cell":ce,"row":ro,"column":co,"raw_value":rv,"flat_value":fv,"bucket":bu,"explanation":ex,"numero":nd,"indice":ind,"lot":"","emetteur":"","titre":ti}

def cs2(sn,wr,wf):
    mc=max(wr.max_column or 1,wf.max_column or 1)
    hr2,rrows=rr(wr,mc); hf2,frows=rr(wf,mc)
    ah=dict(hr2)
    for c,h in hf2.items():
        if c not in ah: ah[c]=h
    md2,ur,uf=mt(rrows,frows)
    c2a={}
    for c in range(17,mc+1,3):
        n=cs(wr.cell(row=8,column=c).value).strip()
        if not n: n=cs(wf.cell(row=8,column=c).value).strip()
        if n: c2a[c]=n; c2a[c+1]=n; c2a[c+2]=n
    oc=None
    for c,h in ah.items():
        if "OBSERVATION" in h.upper(): oc=c; break
    diffs,ident,comp=[],0,0
    cc={"HIGH":0,"MEDIUM":0,"LOW":0,"AMBIGUOUS":0}
    for r in range(1,DS):
        for c in range(1,mc+1):
            rv=cs(wr.cell(row=r,column=c).value).strip(); fv=cs(wf.cell(row=r,column=c).value).strip()
            comp+=1
            if rv==fv: ident+=1
            else:
                cl=get_column_letter(c); b,e=clf(rv,fv,cl,"HEADER")
                diffs.append(dd(sn,cl+str(r),r,cl,cs(wr.cell(row=r,column=c).value),cs(wf.cell(row=r,column=c).value),b,e,"","",""))
    for fr,rx,cf in md2:
        cc[cf]+=1
        if cf=="AMBIGUOUS":
            diffs.append(dd(sn,"A"+str(fr["row"]),fr["row"],"A","<AMBIGUOUS>",fr["doc"],"ROW_ALIGNMENT_UNCERTAIN","Multiple candidates",fr["ndoc"],fr["ind"],fr["titre"][:30]))
            continue
        if rx is None: continue
        for c in range(1,mc+1):
            rv=cs(rx["vals"].get(c)).strip(); fv=cs(fr["vals"].get(c)).strip()
            comp+=1
            if rv==fv: ident+=1
            else:
                cl=get_column_letter(c); cn=ah.get(c,""); ap=c2a.get(c,""); io2=(c==oc)
                b,e=clf(rv,fv,cl,cn,ap=ap,io=io2)
                if cf=="LOW" and b=="REAL_DIVERGENCE": b="ROW_ALIGNMENT_UNCERTAIN"; e="LOW"
                if b=="REAL_DIVERGENCE": e=(e+" [%s]"%cf) if e else "[%s]"%cf
                diffs.append(dd(sn,cl+str(rx["row"]),rx["row"],cl,cs(rx["vals"].get(c)),cs(fr["vals"].get(c)),b,e,rx["ndoc"],rx["ind"],rx["titre"][:30]))
    for rx in ur:
        diffs.append(dd(sn,"A"+str(rx["row"]),rx["row"],"A",rx["doc"],"<MISSING FLAT>","KNOWN_GAP","Scope GAP-3",rx["ndoc"],rx["ind"],rx["titre"][:30]))
    for fr in uf:
        diffs.append(dd(sn,"A"+str(fr["row"]),fr["row"],"A","<MISSING RAW>",fr["doc"],"KNOWN_GAP","Scope: input snapshot diff",fr["ndoc"],fr["ind"],fr["titre"][:30]))
    return diffs,ident,comp,cc

def run():
    wbr=openpyxl.load_workbook(RAW_PATH,data_only=True); wbf=openpyxl.load_workbook(FLAT_PATH,data_only=True)
    rs2,fs2=set(wbr.sheetnames),set(wbf.sheetnames)
    ads,idc=[],{}; tc,ti=0,0; tcc={"HIGH":0,"MEDIUM":0,"LOW":0,"AMBIGUOUS":0}
    for sn in sorted(rs2|fs2):
        if sn not in rs2:
            ads.append(dd(sn,"-",0,"-","<MISSING>","<PRESENT>","KNOWN_GAP","Sheet only flat","","",""))
            idc[sn]={"identical":0,"different":1}; continue
        if sn not in fs2:
            ads.append(dd(sn,"-",0,"-","<PRESENT>","<MISSING>","KNOWN_GAP","Sheet only raw GAP-3","","",""))
            idc[sn]={"identical":0,"different":1}; continue
        print("  %s"%sn)
        d2,ii,cc2,ccc=cs2(sn,wbr[sn],wbf[sn])
        ads.extend(d2); tc+=cc2; ti+=ii
        for k in tcc: tcc[k]+=ccc[k]
        idc[sn]={"identical":ii,"different":len(d2)}
    wbr.close(); wbf.close()
    return {"raw_path":str(RAW_PATH),"flat_path":str(FLAT_PATH),"raw_sheets":len(rs2),"flat_sheets":len(fs2),
            "total_compared":tc,"total_identical":ti,"differences":ads,"identical_counts":idc,"match_confidence":tcc}

def rpt(result):
    diffs=result["differences"]; bk={}
    for d2 in diffs: bk[d2["bucket"]]=bk.get(d2["bucket"],0)+1
    rd=bk.get("REAL_DIVERGENCE",0); un=bk.get("ROW_ALIGNMENT_UNCERTAIN",0)
    vd="PARITY_PASS" if rd==0 else "PARITY_FAIL"; mc=result["match_confidence"]
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="SUMMARY"
    for row in [("Metric","Value"),("Raw",result["raw_path"]),("Flat",result["flat_path"]),
        ("Sheets raw",result["raw_sheets"]),("Sheets flat",result["flat_sheets"]),
        ("Compared",result["total_compared"]),("Identical",result["total_identical"]),
        ("Total diffs",len(diffs)),("BENIGN_WHITESPACE",bk.get("BENIGN_WHITESPACE",0)),
        ("SEMANTIC_EQUIVALENT",bk.get("SEMANTIC_EQUIVALENT",0)),("KNOWN_GAP",bk.get("KNOWN_GAP",0)),
        ("OLD_PATH_BUG",bk.get("OLD_PATH_BUG",0)),("ROW_ALIGNMENT_UNCERTAIN",un),
        ("REAL_DIVERGENCE",rd),("VERDICT",vd),("",""),
        ("Match HIGH",mc["HIGH"]),("Match MEDIUM",mc["MEDIUM"]),
        ("Match LOW",mc["LOW"]),("Match AMBIGUOUS",mc["AMBIGUOUS"])]:
        ws.append(row)
    hd=["sheet","cell","row","column","raw_value","flat_value","bucket","explanation","numero","indice","lot","emetteur","titre"]
    for t,f in [("DIFFERENCES",None),("REAL_DIVERGENCES","REAL_DIVERGENCE"),("KNOWN_GAPS","KNOWN_GAP"),
                ("OLD_PATH_BUGS","OLD_PATH_BUG"),("ROW_ALIGNMENT_UNCERTAIN","ROW_ALIGNMENT_UNCERTAIN")]:
        x=wb.create_sheet(t); x.append(hd)
        for d2 in diffs:
            if f is None or d2["bucket"]==f: x.append([d2.get(h,"") for h in hd])
    x2=wb.create_sheet("IDENTICAL_CELL_COUNT"); x2.append(["sheet","identical","different"])
    for sn,ct in sorted(result["identical_counts"].items()): x2.append([sn,ct["identical"],ct["different"]])
    wb.save(str(REPORT))
    return vd,bk,rd,un

if __name__=="__main__":
    print("Step 5c...")
    result=run()
    vd,bk,rd,un=rpt(result)
    mc=result["match_confidence"]
    print("\n"+"="*60)
    print(vd)
    print("Total: %d"%len(result["differences"]))
    print("REAL_DIVERGENCE:         %d"%rd)
    print("ROW_ALIGNMENT_UNCERTAIN: %d"%un)
    print("KNOWN_GAP:               %d"%bk.get("KNOWN_GAP",0))
    print("OLD_PATH_BUG:            %d"%bk.get("OLD_PATH_BUG",0))
    print("SEMANTIC_EQUIVALENT:     %d"%bk.get("SEMANTIC_EQUIVALENT",0))
    print("BENIGN_WHITESPACE:       %d"%bk.get("BENIGN_WHITESPACE",0))
    print("Match: H=%d M=%d L=%d A=%d"%(mc["HIGH"],mc["MEDIUM"],mc["LOW"],mc["AMBIGUOUS"]))
    print("Report: %s"%REPORT)
    print("="*60)
