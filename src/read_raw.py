"""
read_raw.py
-----------
Reads the GED export Excel file (sheet: 'Doc. sous workflow, x versions').
Handles the 2-row merged header structure:
  Row 1: base fields + approver group names
  Row 2: sub-fields (Date réponse, Réponse, Commentaire, PJ)

Returns:
  - docs_df  : DataFrame of document base fields (one row per GED row)
  - responses_df : DataFrame of exploded approver responses (long format)
  - approver_names : list of canonical approver names discovered in this file
"""

import re
import uuid
from pathlib import Path

import pandas as pd


GED_SHEET = "Doc. sous workflow, x versions"

BASE_FIELDS = [
    "AFFAIRE", "PROJET", "BATIMENT", "PHASE", "EMETTEUR",
    "SPECIALITE", "LOT", "TYPE DE DOC", "ZONE", "NIVEAU",
    "NUMERO", "INDICE", "Libellé du document", "Créé le",
]

# Sub-headers expected for each approver group
APPROVER_SUB = ["Date réponse", "Réponse", "Commentaire", "PJ"]


def _parse_headers(row1: list, row2: list):
    """
    Parse the 2-row header into:
      - base_cols: {col_index: field_name}
      - approver_groups: [{name, date_col, response_col, comment_col, pj_col}]
    """
    base_cols = {}
    approver_groups = []

    current_approver = None
    current_start = None
    sub_idx = 0

    for i, (h1, h2) in enumerate(zip(row1, row2)):
        if h1 is not None and h1 != "":
            # Could be a base field or a new approver group
            if h1 in BASE_FIELDS:
                base_cols[i] = h1
                current_approver = None
            else:
                # New approver group
                if current_approver is not None and current_start is not None:
                    # close previous (shouldn't happen normally)
                    pass
                current_approver = str(h1).strip()
                current_start = i
                sub_idx = 0
                # First sub-col is "Date réponse" at this same column
                if current_approver not in ("", None):
                    approver_groups.append({
                        "name": current_approver,
                        "date_col": i,
                        "response_col": i + 1,
                        "comment_col": i + 2,
                        "pj_col": i + 3,
                    })

    return base_cols, approver_groups


def read_ged(filepath: str) -> tuple:
    """
    Main entry point. Returns (docs_df, responses_df, approver_names).
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"GED file not found: {filepath}")

    # Read raw with openpyxl to handle merged cells correctly
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if GED_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{GED_SHEET}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[GED_SHEET]
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 3:
        raise ValueError("GED sheet has fewer than 3 rows — no data.")

    row1 = list(all_rows[0])
    row2 = list(all_rows[1])
    data_rows = all_rows[2:]

    base_cols, approver_groups = _parse_headers(row1, row2)

    # Build docs records
    docs = []
    responses = []

    for raw_row in data_rows:
        # Skip fully empty rows
        vals = [c for c in raw_row if c is not None]
        if not vals:
            continue

        doc_id = str(uuid.uuid4())

        # Extract base fields
        record = {"doc_id": doc_id}
        for col_idx, field_name in base_cols.items():
            val = raw_row[col_idx] if col_idx < len(raw_row) else None
            # Normalise field name to snake_case internal key
            key = field_name.lower().replace(" ", "_").replace("é", "e").replace("è", "e").replace("ê", "e")
            key = re.sub(r"[^a-z0-9_]", "_", key).strip("_")
            record[key] = val

        docs.append(record)

        # Extract approver responses
        for ag in approver_groups:
            date_val = raw_row[ag["date_col"]] if ag["date_col"] < len(raw_row) else None
            resp_val = raw_row[ag["response_col"]] if ag["response_col"] < len(raw_row) else None
            comment_val = raw_row[ag["comment_col"]] if ag["comment_col"] < len(raw_row) else None
            pj_val = raw_row[ag["pj_col"]] if ag["pj_col"] < len(raw_row) else None

            responses.append({
                "doc_id": doc_id,
                "approver_raw": ag["name"],
                "response_date_raw": date_val,
                "response_status_raw": resp_val,
                "response_comment": comment_val,
                "pj_flag": 1 if pj_val not in (None, "", 0) else 0,
            })

    docs_df = pd.DataFrame(docs)
    responses_df = pd.DataFrame(responses)

    approver_names = [ag["name"] for ag in approver_groups]

    return docs_df, responses_df, approver_names
