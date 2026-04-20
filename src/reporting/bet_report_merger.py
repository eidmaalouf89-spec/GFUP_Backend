"""
bet_report_merger.py — Merge BET PDF report data into GED responses_df

Two merge tracks:
  Track A (status backfill): For docs where GED has no response (PENDING)
           but the PDF report has a status → inject a synthetic response row.
           Applies to: AVLS, SOCOTEC, Le Sommer. NOT Terrell (OBS-ONLY).

  Track B (observation enrichment): For ALL 4 BET consultants, even when GED
           already has a status, if the GED comment is empty or a placeholder
           → replace with the real observation from the PDF report.
"""

import logging
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# ── Canonical name → RAPPORT sheet name ──────────────────────────────────────
RAPPORT_SHEET_MAP = {
    "BET Acoustique":     "RAPPORT_AVLS",
    "Bureau de Contrôle": "RAPPORT_SOCOTEC",
    "AMO HQE":            "RAPPORT_LE_SOMMER",
    "BET Structure":      "RAPPORT_TERRELL",
}

# Consultants where PDF status can backfill GED (Track A).
# Terrell is OBS-ONLY: status comes from GED, only observations from PDF.
STATUS_BACKFILL_CONSULTANTS = {"BET Acoustique", "Bureau de Contrôle", "AMO HQE"}

# ── Placeholder detection ────────────────────────────────────────────────────
# GED comments that are empty or match these patterns should be replaced
# by the real observation from the PDF report.

_PLACEHOLDER_EXACT = {
    "", "nan", "none", "-", ".", "x", "ok", "ras", "nc", "n/c",
}

_PLACEHOLDER_CONTAINS = [
    "voir rapport",
    "cf rapport",
    "cf. rapport",
    "voir pj",
    "cf pj",
    "voir pièce",
    "voir piece",
    "rapport joint",
    "voir fiche visa",
    "voir fiche",
    "fiche visa en pj",
    "visa en pj",
    "en pj",
    "voir obs",
    "voir commentaire",
    "commentaires sur le document joint",
    "document joint",
    "se reporter",
    "voir annexe",
    "voir fichier",
    "pdf joint",
    "pièce jointe",
    "piece jointe",
    "voir le rapport",
    "voir rpt",
    "observations en pièce jointe",
    "observations par zones",
    "voir observations sur visa terrell",
    "voir observations sur le pdf",
    "voir nos observations",
    "voir visa avls",
    "voir fiche synthèse",
    "voir fiche synthese",
]


def is_placeholder_comment(comment: str) -> bool:
    """Return True if the GED comment is empty or a lazy placeholder."""
    if comment is None:
        return True
    s = str(comment).strip()
    if s.lower() in _PLACEHOLDER_EXACT:
        return True
    s_lower = s.lower()
    for pattern in _PLACEHOLDER_CONTAINS:
        if pattern in s_lower:
            return True
    return False


# ── Load RAPPORT sheets ──────────────────────────────────────────────────────

def _find_consultant_reports_xlsx(base_dir: Path) -> Optional[Path]:
    """Find consultant_reports.xlsx in known locations."""
    candidates = [
        base_dir / "output" / "repports output" / "consultant_reports.xlsx",
        base_dir / "output" / "consultant_reports.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _load_rapport_sheet(wb_path: Path, sheet_name: str) -> list[dict]:
    """Load one RAPPORT sheet as list of dicts."""
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(all_rows) < 2:
        return []
    header = [str(h).strip() if h else "" for h in all_rows[0]]
    data = []
    for row in all_rows[1:]:
        rec = {}
        for i, val in enumerate(row):
            if i < len(header):
                rec[header[i]] = str(val).strip() if val is not None else ""
        data.append(rec)
    return data


# ── NUMERO normalisation (must match GED's numero_normalized) ────────────────

def _normalize_numero(raw: str) -> str:
    """Strip leading zeros and whitespace to match GED numero_normalized."""
    s = raw.strip()
    try:
        return str(int(s))
    except (ValueError, TypeError):
        return s


# ── Main merge function ──────────────────────────────────────────────────────

def merge_bet_reports(
    docs_df: pd.DataFrame,
    responses_df: pd.DataFrame,
    base_dir: Path,
) -> tuple[pd.DataFrame, dict]:
    """
    Merge BET PDF report data into responses_df.

    Returns:
        (enriched_responses_df, merge_stats)

    The enriched DataFrame has two new columns:
        response_source: "GED" | "PDF_REPORT" | "GED+PDF_OBS"
        observation_pdf: str  (the PDF observation, for reference)
    """
    stats = {
        "rapport_file": None,
        "by_consultant": {},
        "total_status_backfilled": 0,
        "total_obs_enriched": 0,
    }

    # Find the consultant_reports.xlsx file
    rapport_path = _find_consultant_reports_xlsx(base_dir)
    if rapport_path is None:
        logger.warning("consultant_reports.xlsx not found — skipping BET merge")
        responses_df = responses_df.copy()
        responses_df["response_source"] = "GED"
        responses_df["observation_pdf"] = ""
        return responses_df, stats

    stats["rapport_file"] = str(rapport_path)
    logger.info("BET report merge: loading %s", rapport_path)

    # Initialise new columns
    responses_df = responses_df.copy()
    responses_df["response_source"] = "GED"
    responses_df["observation_pdf"] = ""

    # Build GED NUMERO → [(doc_id, indice), ...] index from all docs.
    # Stores the indice letter so we can match RAPPORT indice to the right version.
    numero_to_docs: dict[str, list] = {}
    for _, row in docs_df.iterrows():
        num = str(row.get("numero_normalized", "")).strip()
        if num and num != "nan":
            indice = str(row.get("indice", "")).strip().upper()
            numero_to_docs.setdefault(num, []).append((row["doc_id"], indice))

    # Pre-build a fast lookup: (doc_id, approver_canonical) → list of positional indices
    # This avoids a full O(n) scan of responses_df for every rapport row.
    logger.info("Building (doc_id, approver) index over %d response rows…", len(responses_df))
    _resp_doc_id = responses_df["doc_id"].values
    _resp_approver = responses_df["approver_canonical"].values
    pair_to_pos: dict[tuple, list[int]] = {}
    for pos, (did, app) in enumerate(zip(_resp_doc_id, _resp_approver)):
        pair_to_pos.setdefault((did, app), []).append(pos)
    logger.info("Index built: %d unique (doc_id, approver) pairs", len(pair_to_pos))

    # Work on numpy arrays for fast column writes
    col_status_clean    = list(responses_df.get("status_clean",    responses_df.index.map(lambda _: "")).values)
    col_date_status     = list(responses_df["date_status_type"].values)
    col_date_answered   = list(responses_df["date_answered"].values) if "date_answered" in responses_df.columns else [None] * len(responses_df)
    col_response_source = list(responses_df["response_source"].values)
    col_response_comment= list(responses_df["response_comment"].values) if "response_comment" in responses_df.columns else [""] * len(responses_df)
    col_obs_pdf         = list(responses_df["observation_pdf"].values)

    new_rows = []  # (kept for future use — Track A now mutates in-place)

    for canonical_name, sheet_name in RAPPORT_SHEET_MAP.items():
        rapport_rows = _load_rapport_sheet(rapport_path, sheet_name)
        if not rapport_rows:
            logger.info("  %s: no rapport rows found", canonical_name)
            continue

        cons_stats = {
            "rapport_rows": len(rapport_rows),
            "matched_to_ged": 0,
            "status_backfilled": 0,
            "obs_enriched": 0,
            "unmatched": 0,
        }

        can_backfill_status = canonical_name in STATUS_BACKFILL_CONSULTANTS

        for rr in rapport_rows:
            numero_raw = rr.get("NUMERO", "")
            numero_norm = _normalize_numero(numero_raw)
            if not numero_norm:
                cons_stats["unmatched"] += 1
                continue

            doc_entries = numero_to_docs.get(numero_norm, [])
            if not doc_entries:
                cons_stats["unmatched"] += 1
                continue

            cons_stats["matched_to_ged"] += 1

            rapport_status = rr.get("STATUT_NORM", "").strip()
            rapport_obs    = rr.get("COMMENTAIRE", "").strip()
            rapport_date   = rr.get("DATE_FICHE", "").strip()
            rapport_indice = rr.get("INDICE", "").strip().upper()

            # For each matching doc_id, find the GED response row.
            # If the RAPPORT has an INDICE, only match docs with that specific indice.
            # If no INDICE in RAPPORT, match all versions of this NUMERO.
            for doc_id, doc_indice in doc_entries:
                if rapport_indice and doc_indice and rapport_indice != doc_indice:
                    continue
                positions = pair_to_pos.get((doc_id, canonical_name), [])
                if not positions:
                    continue

                for pos in positions:
                    ged_status  = col_date_status[pos]
                    ged_comment = str(col_response_comment[pos] or "")

                    # ── Track A: Status backfill ──────────────────────────
                    if (can_backfill_status
                            and ged_status in ("PENDING_IN_DELAY", "PENDING_LATE")
                            and rapport_status):
                        col_status_clean[pos]    = rapport_status
                        col_date_status[pos]     = "ANSWERED"
                        parsed_date = _parse_date_fiche(rapport_date)
                        if parsed_date:
                            col_date_answered[pos] = parsed_date
                        col_response_source[pos] = "PDF_REPORT"
                        cons_stats["status_backfilled"] += 1
                        if rapport_obs:
                            col_response_comment[pos] = rapport_obs
                            col_obs_pdf[pos]          = rapport_obs
                        continue  # done with this row

                    # ── Track B: Observation enrichment ───────────────────
                    if rapport_obs and is_placeholder_comment(ged_comment):
                        col_response_comment[pos] = rapport_obs
                        col_obs_pdf[pos]          = rapport_obs
                        if col_response_source[pos] == "GED":
                            col_response_source[pos] = "GED+PDF_OBS"
                        cons_stats["obs_enriched"] += 1

        stats["by_consultant"][canonical_name] = cons_stats
        stats["total_status_backfilled"] += cons_stats["status_backfilled"]
        stats["total_obs_enriched"] += cons_stats["obs_enriched"]

        logger.info(
            "  %s: %d rapport rows, %d matched, %d status backfilled, %d obs enriched, %d unmatched",
            canonical_name,
            cons_stats["rapport_rows"],
            cons_stats["matched_to_ged"],
            cons_stats["status_backfilled"],
            cons_stats["obs_enriched"],
            cons_stats["unmatched"],
        )

    # Write mutated column arrays back to the DataFrame in one shot
    if "status_clean" in responses_df.columns:
        responses_df["status_clean"] = col_status_clean
    responses_df["date_status_type"]  = col_date_status
    if "date_answered" in responses_df.columns:
        responses_df["date_answered"] = col_date_answered
    responses_df["response_source"]   = col_response_source
    if "response_comment" in responses_df.columns:
        responses_df["response_comment"] = col_response_comment
    responses_df["observation_pdf"]   = col_obs_pdf

    logger.info(
        "BET merge complete: %d status backfilled, %d obs enriched",
        stats["total_status_backfilled"],
        stats["total_obs_enriched"],
    )

    return responses_df, stats


def _parse_date_fiche(raw: str) -> Optional[datetime]:
    """Parse DATE_FICHE string (dd/mm/yyyy or dd/mm/yy) to datetime."""
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw.strip(), fmt)
        except (ValueError, TypeError):
            continue
    return None
