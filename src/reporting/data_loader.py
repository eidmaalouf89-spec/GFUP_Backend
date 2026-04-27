"""
data_loader.py — Run-centric data access layer
Reads from run_memory.db artifact registry, NOT from input/ directly.
Verifies GED provenance before loading. Falls back to degraded mode if unverifiable.
"""
import hashlib
import json
import logging
import re
import sqlite3
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

# Import existing backend modules (src/ is already in sys.path from app.py)
from read_raw import read_ged
from normalize import load_mapping, normalize_docs, normalize_responses
from version_engine import VersionEngine
from workflow_engine import WorkflowEngine, compute_responsible_party, compute_moex_countdown
# bet_report_merger RETIRED (Step 8 / FLAT_GED_REPORT_COMPOSITION.md §8).
# UI now consumes effective_responses_df composed by the pipeline composition engine.
# from reporting.bet_report_merger import merge_bet_reports  # DO NOT RESTORE
from reporting.focus_ownership import compute_focus_ownership
from report_memory import load_persisted_report_responses
from effective_responses import build_effective_responses


@dataclass
class RunContext:
    run_number: int
    run_status: str
    run_date: str
    summary_json: dict
    gf_artifact_path: Optional[Path]
    ged_available: bool
    degraded_mode: bool
    docs_df: Optional[pd.DataFrame] = None
    responses_df: Optional[pd.DataFrame] = None
    approver_names: Optional[list] = None
    dernier_df: Optional[pd.DataFrame] = None          # dernier indice docs only
    workflow_engine: Optional[WorkflowEngine] = None
    responsible_parties: Optional[dict] = None          # {doc_id: responsible_party}
    gf_sheets: dict = field(default_factory=dict)       # {sheet_name: {lots, contractor}}
    artifact_paths: dict = field(default_factory=dict)  # {artifact_type: file_path}
    warnings: list = field(default_factory=list)
    data_date: Optional[date] = None                    # from GED Détails sheet
    ged_status_labels: dict = field(default_factory=dict)  # e.g. {"Terrell": {"s1": "VSO", "s2": "OBS", "s3": "REF"}}
    bet_merge_stats: dict = field(default_factory=dict)  # from BET report merger
    moex_countdown: dict = field(default_factory=dict)   # {doc_id: countdown_info}


# Module-level cache
_cached_context: Optional[RunContext] = None
_cached_run_number: Optional[int] = None
_SQLITE_READ_RETRY_DELAYS = (0.0, 0.2, 0.5)


def clear_cache():
    """Call this when a pipeline run completes or the user switches runs."""
    global _cached_context, _cached_run_number
    _cached_context = None
    _cached_run_number = None
    _resolve_artifact_file.cache_clear()


def _sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _is_sqlite_locked_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return "database is locked" in msg or "database table is locked" in msg


def _immutable_sqlite_uri(db_path: str) -> str:
    return "file:" + Path(db_path).resolve().as_posix() + "?mode=ro&immutable=1"


def _query_db(db_path, sql, params=()):
    if not Path(db_path).exists():
        return []
    last_exc = None
    for attempt, delay in enumerate(_SQLITE_READ_RETRY_DELAYS):
        try:
            conn = sqlite3.connect(db_path, timeout=5)
            conn.execute("PRAGMA busy_timeout=5000")
            conn.row_factory = sqlite3.Row
            try:
                return [dict(r) for r in conn.execute(sql, params).fetchall()]
            finally:
                conn.close()
        except sqlite3.OperationalError as exc:
            last_exc = exc
            if _is_sqlite_locked_error(exc) and attempt < len(_SQLITE_READ_RETRY_DELAYS) - 1:
                time.sleep(delay)
                continue
            break

    try:
        conn = sqlite3.connect(_immutable_sqlite_uri(db_path), uri=True, timeout=5)
        conn.row_factory = sqlite3.Row
        try:
            rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
        finally:
            conn.close()
        if last_exc is not None:
            logger.warning(
                "[SQLITE_IMMUTABLE_FALLBACK] Using read-only immutable mode for %s after: %s",
                db_path,
                last_exc,
            )
        return rows
    except sqlite3.OperationalError as exc:
        raise RuntimeError(
            f"Could not read SQLite database {db_path}: {exc}"
        ) from exc


def _resolve_latest_run(db_path: str) -> Optional[int]:
    """Find latest COMPLETED non-stale run number."""
    rows = _query_db(db_path,
        "SELECT run_number FROM runs "
        "WHERE status='COMPLETED' AND is_stale=0 "
        "ORDER BY run_number DESC LIMIT 1")
    return rows[0]["run_number"] if rows else None


@lru_cache(maxsize=1024)
def _resolve_artifact_file(db_path: str, file_path: str) -> Optional[str]:
    """Try stored path, then relocation-aware fallback relative to project root."""
    if not file_path:
        return None
    p = Path(file_path)
    if p.exists():
        return str(p)
    base_dir = Path(db_path).resolve().parent.parent
    path_str = str(p)
    for anchor in ("runs/", "runs\\", "output/", "output\\"):
        idx = path_str.find(anchor)
        if idx >= 0:
            candidate = base_dir / path_str[idx:]
            if candidate.exists():
                return str(candidate.resolve())
    # Parts-based fallback for Linux paths parsed on Windows (and vice versa)
    parts = p.parts
    for anchor in ("runs", "input", "output", "data", "debug"):
        for i, part in enumerate(parts):
            if part == anchor:
                candidate = base_dir.joinpath(*parts[i:])
                if candidate.exists():
                    return str(candidate.resolve())
                break

    # Cross-platform fallback: split stored path by backslash (Windows paths on Linux)
    win_parts = re.split(r'[/\\]', path_str)
    for anchor in ("runs", "input", "output", "data", "debug"):
        for i, part in enumerate(win_parts):
            if part == anchor:
                candidate = base_dir.joinpath(*win_parts[i:])
                if candidate.exists():
                    return str(candidate.resolve())
                break

    return None


def _get_artifact_path(db_path: str, run_number: int, artifact_type: str) -> Optional[str]:
    """Resolve an artifact file path, with relocation-aware fallback."""
    rows = _query_db(db_path,
        "SELECT file_path FROM run_artifacts "
        "WHERE run_number=? AND artifact_type=? "
        "ORDER BY created_at DESC LIMIT 1",
        (run_number, artifact_type))
    if not rows or not rows[0]["file_path"]:
        return None
    return _resolve_artifact_file(db_path, rows[0]["file_path"])


def _verify_ged_provenance(db_path: str, run_number: int) -> tuple:
    """
    Returns (ged_path, verified: bool).
    Checks if the GED file used for this run still exists and has the same hash.
    """
    rows = _query_db(db_path,
        "SELECT source_path, source_file_hash FROM run_inputs "
        "WHERE run_number=? AND input_type='GED'",
        (run_number,))
    if not rows:
        return None, False
    row = rows[0]
    source_path = row.get("source_path")
    stored_hash = row.get("source_file_hash")
    if not source_path:
        return source_path, False
    resolved_str = _resolve_artifact_file(db_path, source_path)
    if not resolved_str:
        return source_path, False
    if not stored_hash:
        return resolved_str, True  # No hash stored — trust it
    actual_hash = _sha256(resolved_str)
    return resolved_str, (actual_hash == stored_hash)




def _read_ged_data_date(ged_path: str) -> Optional[date]:
    """Extract DATA_DATE from the GED workbook's Détails sheet.

    The AxeoBIM export places the export timestamp at row ~15:
      Col B: "Date & heure de la demande"
      Col C: "∙"  (decorative dot)
      Col D: datetime value

    We scan for any cell containing 'date' AND one of: 'demande', 'extraction', 'export'.
    Then we look at column+2 first (skipping the dot separator), then column+1.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(ged_path, read_only=True, data_only=True)
        details_sheet = None
        for sn in wb.sheetnames:
            if sn.lower().replace("é", "e").replace("È", "e") in ("details", "détails"):
                details_sheet = wb[sn]
                break
        if details_sheet is None:
            wb.close()
            return None

        for row in details_sheet.iter_rows(min_row=1, max_row=30, max_col=10, values_only=False):
            for cell in row:
                val = str(cell.value or "").lower()
                if "date" not in val:
                    continue
                # Match any of the known label patterns
                if not any(kw in val for kw in ("demande", "extraction", "export")):
                    continue

                ws = cell.parent
                # Try col+2 first (AxeoBIM uses col B=label, col C=dot, col D=value)
                for offset in (2, 1, 3):
                    adj = ws.cell(row=cell.row, column=cell.column + offset).value
                    if adj is None:
                        continue
                    if isinstance(adj, datetime):
                        wb.close()
                        return adj.date()
                    if isinstance(adj, date):
                        wb.close()
                        return adj
                    try:
                        wb.close()
                        return datetime.strptime(str(adj).strip(), "%d/%m/%Y").date()
                    except Exception:
                        pass

                # Try cell below as last resort
                below = ws.cell(row=cell.row + 1, column=cell.column).value
                if below is not None:
                    if isinstance(below, datetime):
                        wb.close()
                        return below.date()
                    if isinstance(below, date):
                        wb.close()
                        return below
        wb.close()
    except Exception:
        pass
    return None


def _parse_gf_sheets(gf_path: str) -> dict:
    """
    Parse GF sheet names to extract contractor info.
    Returns: {sheet_name: {"lots": [str], "contractor_code": str, "contractor_name": str}}
    """
    import openpyxl
    wb = openpyxl.load_workbook(gf_path, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        if name.upper().startswith("OLD"):
            continue
        # Parse: "LOT 41-CVC-AXIMA" -> lots=["41"], code="AXIMA"
        parts = name.replace("LOT ", "").replace("Lot ", "").split("-")
        code = parts[-1].strip() if len(parts) >= 2 else name
        sheets[name] = {
            "lots": [],  # Will be enriched by routing if needed
            "contractor_code": code,
            "contractor_name": code,
        }
    wb.close()
    return sheets


def _load_from_flat_artifacts(base_dir: Path, db_path: str, run_number: int,
                              run_meta: dict, artifact_paths: dict,
                              warnings: list) -> Optional[RunContext]:
    """
    Artifact-first loader: builds RunContext from registered FLAT_GED artifact.
    Uses stage_read_flat() to parse FLAT_GED.xlsx, then normalize + WorkflowEngine.
    VersionEngine is skipped — flat GED contains only ACTIVE instances (all are dernier).
    Returns RunContext on success, None on failure (caller falls back to legacy).
    """
    flat_ged_path = _get_artifact_path(db_path, run_number, "FLAT_GED")
    if not flat_ged_path:
        logger.info("[FLAT_ARTIFACT] No FLAT_GED artifact for Run %d — will try fallback", run_number)
        return None

    try:
        from types import SimpleNamespace
        from pipeline.stages.stage_read_flat import stage_read_flat

        # Parse FLAT_GED.xlsx via stage_read_flat (reuses all flat adapter logic)
        mock_ctx = SimpleNamespace(FLAT_GED_FILE=flat_ged_path)
        stage_read_flat(mock_ctx, log=lambda msg: logger.debug(msg))

        docs_df = mock_ctx.docs_df
        responses_df = mock_ctx.responses_df
        approver_names = mock_ctx.ged_approver_names
        mapping = mock_ctx.mapping

        # Normalize (adds approver_canonical, date_status_type, lot_normalized, etc.)
        docs_df = normalize_docs(docs_df, mapping)
        responses_df = normalize_responses(responses_df, mapping)

        # Skip VersionEngine — flat GED only has ACTIVE instances.
        # All docs are dernier indice by definition (FLAT_GED_CONTRACT §Known Limitation).
        docs_df["is_dernier_indice"] = True
        dernier_df = docs_df.copy()

        # WorkflowEngine + responsible parties
        workflow_engine = WorkflowEngine(responses_df)
        dernier_ids = dernier_df["doc_id"].tolist()
        responsible_parties = compute_responsible_party(workflow_engine, dernier_ids)

        # Inject MOEX SAS synthetic consultant
        if not responses_df.empty:
            has_sas = (responses_df["approver_raw"] == "0-SAS").any()
            if has_sas and approver_names is not None and "MOEX SAS" not in approver_names:
                approver_names.append("MOEX SAS")

        # data_date from flat_ged_doc_meta (not from raw GED Détails sheet)
        data_date_val = None
        if mock_ctx.flat_ged_doc_meta:
            for meta in mock_ctx.flat_ged_doc_meta.values():
                dd_str = meta.get("data_date", "")
                if dd_str:
                    try:
                        data_date_val = datetime.strptime(dd_str[:10], "%Y-%m-%d").date()
                    except Exception:
                        pass
                    break

        # Report memory composition (NO bet_report_merger — already retired)
        try:
            report_memory_db = str(base_dir / "data" / "report_memory.db")
            persisted_df = load_persisted_report_responses(report_memory_db)
            effective_responses_df = build_effective_responses(responses_df, persisted_df)
            if "effective_source" in effective_responses_df.columns:
                effective_responses_df["response_source"] = effective_responses_df["effective_source"]
            if "observation_pdf" not in effective_responses_df.columns:
                effective_responses_df["observation_pdf"] = ""
            workflow_engine = WorkflowEngine(effective_responses_df)
            responsible_parties = compute_responsible_party(workflow_engine, dernier_ids)
            responses_df = effective_responses_df
            logger.info(
                "[FLAT_ARTIFACT] Report memory composition: %d effective rows, %d upgraded",
                len(responses_df),
                int(responses_df.get("report_memory_applied", pd.Series(dtype=bool)).sum())
                if "report_memory_applied" in responses_df.columns else 0,
            )
        except Exception as e:
            logger.warning("[FLAT_ARTIFACT] Report memory composition failed (non-fatal): %s", e)
            if "response_source" not in responses_df.columns:
                responses_df["response_source"] = "GED"
            if "observation_pdf" not in responses_df.columns:
                responses_df["observation_pdf"] = ""

        # MOEX countdown
        moex_countdown = compute_moex_countdown(workflow_engine, dernier_ids, data_date=data_date_val)

        # Focus columns + ownership
        try:
            _precompute_focus_columns(dernier_df, responses_df, workflow_engine, data_date_val)
            if data_date_val is not None:
                compute_focus_ownership(dernier_df, workflow_engine, data_date_val)
            logger.info("[FLAT_ARTIFACT] Focus columns pre-computed on dernier_df")
        except Exception as e:
            logger.warning("[FLAT_ARTIFACT] Focus column pre-computation failed (non-fatal): %s", e)

        # GF sheets
        gf_path = artifact_paths.get("FINAL_GF") or _get_artifact_path(db_path, run_number, "FINAL_GF")
        gf_sheets = {}
        if gf_path:
            try:
                gf_sheets = _parse_gf_sheets(gf_path)
            except Exception as e:
                warnings.append(f"Error parsing GF sheets: {e}")

        summary = {}
        if run_meta.get("summary_json"):
            try:
                summary = json.loads(run_meta["summary_json"])
            except Exception:
                pass

        logger.info(
            "[FLAT_ARTIFACT] RunContext loaded: %d docs, %d dernier, %d responses, data_date=%s",
            len(docs_df), len(dernier_df), len(responses_df), data_date_val,
        )

        return RunContext(
            run_number=run_number,
            run_status=run_meta["status"],
            run_date=run_meta.get("completed_at") or "",
            summary_json=summary,
            gf_artifact_path=Path(gf_path) if gf_path else None,
            ged_available=True,
            degraded_mode=False,
            docs_df=docs_df,
            responses_df=responses_df,
            approver_names=approver_names,
            dernier_df=dernier_df,
            workflow_engine=workflow_engine,
            responsible_parties=responsible_parties,
            gf_sheets=gf_sheets,
            artifact_paths=artifact_paths,
            warnings=warnings,
            data_date=data_date_val,
            ged_status_labels={},
            bet_merge_stats={},
            moex_countdown=moex_countdown,
        )

    except Exception as e:
        logger.warning("[FLAT_ARTIFACT] Failed: %s — will try legacy fallback", e)
        return None


def _precompute_focus_columns(dernier_df: pd.DataFrame,
                              responses_df: pd.DataFrame,
                              workflow_engine,
                              data_date_val) -> None:
    """Add pre-computed focus columns to dernier_df IN PLACE.

    Columns added:
        _visa_global       : str or None — MOEX visa status
        _visa_global_date  : date or None — date of MOEX visa
        _last_activity_date: date or None — max(submission, any response date)
        _days_since_last_activity: int or None — DATA_DATE - _last_activity_date
        _earliest_deadline : date or None — min(date_limite) among pending responses
        _days_to_deadline  : int or None — earliest_deadline - DATA_DATE (negative = overdue)
        _focus_priority    : int 1-5 — urgency tier from _days_to_deadline
    """
    if data_date_val is None:
        return

    dd = data_date_val.date() if hasattr(data_date_val, 'date') else data_date_val

    # ── 1. VISA GLOBAL per doc (O(N) using cached _doc_approvers) ───
    visa_globals = {}
    visa_dates = {}
    for doc_id in dernier_df["doc_id"]:
        v, vd = workflow_engine.compute_visa_global_with_date(doc_id)
        visa_globals[doc_id] = v
        if vd is not None:
            visa_dates[doc_id] = vd.date() if hasattr(vd, 'date') else vd
        else:
            visa_dates[doc_id] = None

    dernier_df["_visa_global"] = dernier_df["doc_id"].map(visa_globals)
    dernier_df["_visa_global_date"] = dernier_df["doc_id"].map(visa_dates)

    # ── 2. Last activity date (O(M) grouped aggregation) ────────────
    # Get max date_answered per doc from responses
    resp_answered = responses_df[responses_df["date_answered"].notna()].copy()
    if not resp_answered.empty:
        resp_answered["_da_date"] = pd.to_datetime(
            resp_answered["date_answered"], errors="coerce"
        ).dt.date
        last_resp = resp_answered.groupby("doc_id")["_da_date"].max().to_dict()
    else:
        last_resp = {}

    # Compute last activity = max(created_at, last_response_date)
    def _compute_last_activity(row):
        created = row.get("created_at")
        if created is not None and not pd.isna(created):
            created = created.date() if hasattr(created, 'date') else created
        else:
            created = None
        resp_date = last_resp.get(row["doc_id"])
        dates = [d for d in [created, resp_date] if d is not None]
        return max(dates) if dates else None

    dernier_df["_last_activity_date"] = dernier_df.apply(_compute_last_activity, axis=1)

    # Days since last activity
    def _days_since(last_act):
        if last_act is None:
            return None
        try:
            la = last_act.date() if hasattr(last_act, 'date') else last_act
            return (dd - la).days
        except Exception:
            return None

    dernier_df["_days_since_last_activity"] = dernier_df["_last_activity_date"].apply(_days_since)

    # ── 3. Earliest deadline among PENDING responses (O(M) grouped) ─
    pending_resp = responses_df[
        (responses_df["date_status_type"].isin(["PENDING_IN_DELAY", "PENDING_LATE"])) &
        (responses_df["date_limite"].notna())
    ].copy()

    if not pending_resp.empty:
        pending_resp["_dl_date"] = pd.to_datetime(
            pending_resp["date_limite"], errors="coerce"
        ).dt.date
        earliest_dl = pending_resp.groupby("doc_id")["_dl_date"].min().to_dict()
    else:
        earliest_dl = {}

    dernier_df["_earliest_deadline"] = dernier_df["doc_id"].map(earliest_dl)

    # Days to deadline (negative = overdue)
    def _days_to_dl(dl):
        if dl is None or pd.isna(dl):
            return None
        try:
            d = dl.date() if hasattr(dl, 'date') else dl
            return (d - dd).days
        except Exception:
            return None

    dernier_df["_days_to_deadline"] = dernier_df["_earliest_deadline"].apply(_days_to_dl)

    # ── 4. Focus priority tier ───────────────────────────��──────────
    def _priority(dtd):
        if dtd is None:
            return 5  # P5 — no deadline
        if dtd < 0:
            return 1  # P1 — overdue
        if dtd <= 5:
            return 2  # P2 — urgent
        if dtd <= 15:
            return 3  # P3 — soon
        return 4      # P4 — comfortable

    dernier_df["_focus_priority"] = dernier_df["_days_to_deadline"].apply(_priority)


def load_run_context(base_dir: Path, run_number: int = None) -> RunContext:
    """
    Load all data for a specific run. Uses cache if available.

    Default path (FLAT_ARTIFACT — CLEAN Step 11):
    1. Resolve run number (default: latest COMPLETED non-stale)
    2. Collect artifact paths from run_memory.db
    3. Try _load_from_flat_artifacts() — reads registered FLAT_GED artifact
    4. If successful, return artifact-based RunContext

    Legacy fallback (LEGACY_RAW_FALLBACK):
    5. Verify GED provenance
    6. If GED verified: read raw GED, normalize, VersionEngine, WorkflowEngine
    7. If not: degraded mode (GF-only data)
    8. Parse GF sheet structure, cache, return
    """
    global _cached_context, _cached_run_number

    db_path = str(base_dir / "data" / "run_memory.db")

    if run_number is None:
        run_number = _resolve_latest_run(db_path)
    if run_number is None:
        return RunContext(
            run_number=0, run_status="UNKNOWN", run_date="",
            summary_json={}, gf_artifact_path=None,
            ged_available=False, degraded_mode=True,
            warnings=["No completed run found"],
        )

    # Return cache if same run
    if _cached_context and _cached_run_number == run_number:
        return _cached_context

    warnings = []

    # Get run metadata
    run_rows = _query_db(db_path,
        "SELECT status, completed_at, summary_json FROM runs WHERE run_number=?",
        (run_number,))
    if not run_rows:
        return RunContext(
            run_number=run_number, run_status="NOT_FOUND", run_date="",
            summary_json={}, gf_artifact_path=None,
            ged_available=False, degraded_mode=True,
            warnings=[f"Run {run_number} not found in database"],
        )

    run_meta = run_rows[0]
    summary = {}
    if run_meta.get("summary_json"):
        try:
            summary = json.loads(run_meta["summary_json"])
        except Exception:
            pass

    # Resolve FINAL_GF artifact
    gf_path = _get_artifact_path(db_path, run_number, "FINAL_GF")
    if not gf_path:
        warnings.append(f"FINAL_GF artifact not found for Run {run_number}")

    # Collect all artifact paths
    art_rows = _query_db(db_path,
        "SELECT artifact_type, file_path FROM run_artifacts WHERE run_number=?",
        (run_number,))
    artifact_paths = {}
    for r in art_rows:
        resolved = _resolve_artifact_file(db_path, r["file_path"]) if r["file_path"] else None
        if resolved:
            artifact_paths[r["artifact_type"]] = resolved

    # ═══ ARTIFACT-FIRST PATH (CLEAN Step 11) ═══════════════════════
    # Try loading from registered FLAT_GED artifact. This is the default
    # product path. Raw GED rebuild is legacy fallback only.
    flat_ctx = _load_from_flat_artifacts(
        base_dir, db_path, run_number, run_meta, artifact_paths, list(warnings)
    )
    if flat_ctx is not None:
        _cached_context = flat_ctx
        _cached_run_number = run_number
        return flat_ctx

    # ═══ LEGACY FALLBACK ONLY — raw GED rebuild ═══════════════════
    # This path runs only when FLAT_GED artifact is missing or flat load failed.
    # It rebuilds everything from raw GED input. Will be deprecated.
    logger.warning("[LEGACY_RAW_FALLBACK] UI loader using legacy raw GED rebuild — "
                   "FLAT_GED artifact missing or load failed for Run %d", run_number)

    # Verify GED provenance
    ged_path, ged_verified = _verify_ged_provenance(db_path, run_number)
    ged_available = ged_verified and ged_path is not None
    degraded_mode = not ged_available

    if not ged_available:
        if ged_path:
            warnings.append(f"GED file hash mismatch for Run {run_number} — degraded mode")
        else:
            warnings.append(f"GED file not found for Run {run_number} — degraded mode")

    # Parse GF sheets
    gf_sheets = {}
    if gf_path:
        try:
            gf_sheets = _parse_gf_sheets(gf_path)
        except Exception as e:
            warnings.append(f"Error parsing GF sheets: {e}")

    # Extract DATA_DATE from GED Détails sheet
    data_date_val = None
    if ged_available and ged_path:
        try:
            data_date_val = _read_ged_data_date(ged_path)
            if data_date_val is None:
                warnings.append("DATA_DATE not found in GED Détails sheet")
        except Exception as e:
            warnings.append(f"Error reading DATA_DATE from GED: {e}")

    # Build DataFrames if GED is available
    docs_df = None
    responses_df = None
    approver_names = None
    dernier_df = None
    workflow_engine = None
    responsible_parties = None
    bet_merge_stats = {}
    moex_countdown = {}

    if ged_available:
        try:
            # Mapping is now hardcoded in normalize.py
            mapping = load_mapping()

            # Read and normalize GED
            docs_df, responses_df, approver_names = read_ged(ged_path)
            docs_df = normalize_docs(docs_df, mapping)
            responses_df = normalize_responses(responses_df, mapping)

            # Run version engine
            ve = VersionEngine(docs_df)
            versioned_df = ve.run()
            dernier_df = versioned_df[versioned_df["is_dernier_indice"] == True].copy()

            # Run workflow engine
            workflow_engine = WorkflowEngine(responses_df)

            # Compute responsible parties for dernier docs
            dernier_ids = dernier_df["doc_id"].tolist()
            responsible_parties = compute_responsible_party(workflow_engine, dernier_ids)

            # Inject MOEX SAS as a synthetic consultant if any 0-SAS rows exist
            if responses_df is not None and not responses_df.empty:
                has_sas = (responses_df["approver_raw"] == "0-SAS").any()
                if has_sas and approver_names is not None and "MOEX SAS" not in approver_names:
                    approver_names.append("MOEX SAS")

            # ── Report memory composition (replaces retired bet_report_merger) ──
            # Build effective_responses_df via pipeline composition engine.
            # This is the same path the main pipeline uses in stage_report_memory,
            # ensuring UI and GF output reflect identical state.
            try:
                report_memory_db = str(base_dir / "data" / "report_memory.db")
                persisted_df = load_persisted_report_responses(report_memory_db)
                effective_responses_df = build_effective_responses(responses_df, persisted_df)
                # Map effective_source → response_source for UI compatibility
                if "effective_source" in effective_responses_df.columns:
                    effective_responses_df["response_source"] = effective_responses_df["effective_source"]
                if "observation_pdf" not in effective_responses_df.columns:
                    effective_responses_df["observation_pdf"] = ""
                # Rebuild workflow engine from composed effective responses
                workflow_engine = WorkflowEngine(effective_responses_df)
                responsible_parties = compute_responsible_party(workflow_engine, dernier_ids)
                # Replace responses_df with effective version for downstream consumers
                responses_df = effective_responses_df
                logger.info(
                    "UI report composition: %d effective rows, %d upgraded via report memory",
                    len(responses_df),
                    int(responses_df.get("report_memory_applied", pd.Series(dtype=bool)).sum())
                    if "report_memory_applied" in responses_df.columns else 0,
                )
            except Exception as e:
                logger.warning("UI report memory composition failed (non-fatal): %s", e)
                # Ensure compatibility columns exist on failure
                if "response_source" not in responses_df.columns:
                    responses_df["response_source"] = "GED"
                if "observation_pdf" not in responses_df.columns:
                    responses_df["observation_pdf"] = ""

            # Compute MOEX 10-day countdown for all dernier docs
            moex_countdown = compute_moex_countdown(workflow_engine, dernier_ids, data_date=data_date_val)

            # ── Pre-compute focus columns on dernier_df ─────────────────
            # These columns enable Focus Mode to filter with DataFrame ops
            # instead of per-doc Python loops. Computed once, cached with ctx.
            try:
                _precompute_focus_columns(
                    dernier_df, responses_df, workflow_engine, data_date_val
                )
                # Ownership resolver — adds _focus_owner and _focus_owner_tier
                if data_date_val is not None:
                    dd_resolved = data_date_val.date() if hasattr(data_date_val, 'date') else data_date_val
                    compute_focus_ownership(dernier_df, workflow_engine, dd_resolved)
                logger.info("Focus columns + ownership pre-computed on dernier_df")
            except Exception as e:
                logger.warning("Focus column pre-computation failed (non-fatal): %s", e)

            logger.info("RunContext loaded: %d docs, %d dernier, %d responses",
                        len(docs_df), len(dernier_df), len(responses_df))

        except Exception as e:
            warnings.append(f"Error loading GED data: {e}")
            ged_available = False
            degraded_mode = True
            docs_df = None
            responses_df = None

    ctx = RunContext(
        run_number=run_number,
        run_status=run_meta["status"],
        run_date=run_meta.get("completed_at") or "",
        summary_json=summary,
        gf_artifact_path=Path(gf_path) if gf_path else None,
        ged_available=ged_available,
        degraded_mode=degraded_mode,
        docs_df=docs_df,
        responses_df=responses_df,
        approver_names=approver_names,
        dernier_df=dernier_df,
        workflow_engine=workflow_engine,
        responsible_parties=responsible_parties,
        gf_sheets=gf_sheets,
        artifact_paths=artifact_paths,
        warnings=warnings,
        data_date=data_date_val,
        ged_status_labels={},
        bet_merge_stats=bet_merge_stats,
        moex_countdown=moex_countdown,
    )

    # Cache it
    _cached_context = ctx
    _cached_run_number = run_number
    return ctx
