"""Step 5 — Per-bucket Excel export for the ACTION MOEX cockpit.

Reads the existing Phase 6A artifact (output/intermediate/COUNTER_ATTACK_ITEMS.csv)
filtered to one bucket. For each item, joins to ctx.dernier_df for the reception
date / titre and calls document_command_center._get_latest_responses_for_doc to
get the same composed-truth reviewer list the DCC already uses. No business
logic is recomputed.

Public entrypoint:
    build_action_moex_bucket_xlsx(ctx, bucket: str, dest_dir: Path) -> dict

Layout — Y (one row per item; reviewer detail in multi-line cells).
Columns (French):
    1. Numero
    2. Indice
    3. Emetteur (canonical via resolve_emetteur_name)
    4. Titre (libelle_du_document from dernier_df)
    5. Date de reception (created_at, ISO date)
    6. Date contractuelle de reponse (created_at + 30 days)
    7. Reviseurs (multi-line: "{name} ({tier})")
    8. Statuts (multi-line, same row order)
    9. Dates de reponse (multi-line, ISO date or empty)
   10. Commentaires (multi-line, raw comment text)
   11. Pourquoi il est ici (plain_reason from CSV)
   12. MOEX AVIS (always empty; for manual team input)

Empty bucket: header-only workbook, count=0, no exception.
Identity dtype: numero/indice/family_key/emetteur_code/item_id read as string.
Atomic write: temp file then rename.
"""
from __future__ import annotations

import os
import tempfile
from datetime import timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


_IDENTITY_DTYPES: Dict[str, str] = {
    "item_id": "string",
    "numero": "string",
    "indice": "string",
    "family_key": "string",
    "emetteur_code": "string",
}

# Tier ordering for multi-line reviewer cells
_TIER_ORDER = {"PRIMARY": 0, "SECONDARY": 1, "MOEX": 2, "MOEX_SAS": 3}


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, float) and pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value)


def _fmt_iso_date(value: Any) -> str:
    """Format a date / datetime / pandas Timestamp to ISO yyyy-mm-dd. Empty if unparseable."""
    if value is None:
        return ""
    try:
        if isinstance(value, float) and pd.isna(value):
            return ""
    except Exception:
        pass
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(value).strip()
    if not s:
        return ""
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m-%d")
    except Exception:
        return ""


def _add_30_days(value: Any) -> str:
    """Return ISO date for `value + 30 days`. Empty if value missing / unparseable."""
    if value is None:
        return ""
    try:
        if isinstance(value, float) and pd.isna(value):
            return ""
    except Exception:
        pass
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return ""
        return (ts + timedelta(days=30)).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _sort_responses(latest_responses: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort by tier (PRIMARY, SECONDARY, MOEX, MOEX_SAS, then UNKNOWN), then alphabetical reviewer name."""
    def key(r):
        tier = str(r.get("tier") or "").upper()
        tier_rank = _TIER_ORDER.get(tier, 99)
        name = str(r.get("reviewer") or "").lower()
        return (tier_rank, name)
    return sorted(latest_responses, key=key)


def _resolve_dernier_row(ctx, numero: str, indice: str) -> Optional[Dict[str, Any]]:
    """Fetch the matching latest-chain row by (numero, indice). Returns None if not found.

    Step 7: migrated to use latest_enriched_view(ctx) as the row source so a
    keyed (numero, indice) lookup cannot accidentally return a stale-indice
    row even if a future caller passes a non-latest indice.
    """
    if getattr(ctx, "dernier_df", None) is None:
        return None
    try:
        from reporting.latest_chain_view import latest_enriched_view
        d = latest_enriched_view(ctx)
    except Exception:
        return None
    if d is None or len(d) == 0:
        return None
    try:
        sub = d[(d["numero"].astype(str) == str(numero)) &
                (d["indice"].astype(str) == str(indice))]
    except Exception:
        return None
    if sub.empty:
        return None
    return sub.iloc[0].to_dict()


def _resolve_emetteur_display(emetteur_code: str, emetteur_name_csv: str) -> str:
    """Prefer canonical resolve_emetteur_name(code); fall back to CSV emetteur_name; then code."""
    code = _safe_str(emetteur_code).strip()
    if code:
        try:
            from reporting.contractor_fiche import resolve_emetteur_name
            resolved = resolve_emetteur_name(code)
            if resolved:
                return resolved
        except Exception:
            pass
    name = _safe_str(emetteur_name_csv).strip()
    if name:
        return name
    return code


def _build_row_layout_y(
    csv_row: pd.Series,
    dernier_row: Optional[Dict[str, Any]],
    latest_responses: List[Dict[str, Any]],
) -> Dict[str, str]:
    """Return the 12 column values for one item row."""
    numero = _safe_str(csv_row.get("numero"))
    indice = _safe_str(csv_row.get("indice"))
    emetteur_display = _resolve_emetteur_display(
        csv_row.get("emetteur_code"),
        csv_row.get("emetteur_name"),
    )

    # Titre — prefer dernier_df.libelle_du_document; fall back to subject_label sans "{em} — " prefix
    titre = ""
    if dernier_row is not None:
        titre = _safe_str(dernier_row.get("libelle_du_document")).strip()
    if not titre:
        subj = _safe_str(csv_row.get("subject_label")).strip()
        # subject_label format: "EMETTEUR — titre" — strip prefix once if present
        if " — " in subj:
            titre = subj.split(" — ", 1)[1].strip()
        else:
            titre = subj

    # Reception date — created_at on dernier_df row
    reception_iso = ""
    if dernier_row is not None:
        reception_iso = _fmt_iso_date(dernier_row.get("created_at"))

    # Contractual response date = reception + 30 days
    contractual_iso = ""
    if dernier_row is not None:
        contractual_iso = _add_30_days(dernier_row.get("created_at"))

    # Reviewer multi-line cells — sorted by tier, then name
    sorted_resps = _sort_responses(latest_responses or [])
    reviewers_lines: List[str] = []
    statuses_lines: List[str] = []
    dates_lines: List[str] = []
    comments_lines: List[str] = []
    for r in sorted_resps:
        name = _safe_str(r.get("reviewer")).strip()
        tier = _safe_str(r.get("tier")).strip()
        if name:
            reviewers_lines.append(f"{name} ({tier})" if tier else name)
        else:
            reviewers_lines.append("")
        status_clean = _safe_str(r.get("status")).strip()
        statuses_lines.append(status_clean if status_clean else "En attente")
        dates_lines.append(_fmt_iso_date(r.get("response_date")))
        comments_lines.append(_safe_str(r.get("comment")).strip())

    reviewers_cell = "\n".join(reviewers_lines)
    statuses_cell = "\n".join(statuses_lines)
    dates_cell = "\n".join(dates_lines)
    comments_cell = "\n".join(comments_lines)

    plain_reason = _safe_str(csv_row.get("plain_reason")).strip()

    return {
        "Numéro": numero,
        "Indice": indice,
        "Émetteur": emetteur_display,
        "Titre": titre,
        "Date de réception": reception_iso,
        "Date contractuelle de réponse": contractual_iso,
        "Réviseurs": reviewers_cell,
        "Statuts": statuses_cell,
        "Dates de réponse": dates_cell,
        "Commentaires": comments_cell,
        "Pourquoi il est ici": plain_reason,
        "MOEX AVIS": "",
    }


_COLUMNS_ORDER: List[str] = [
    "Numéro",
    "Indice",
    "Émetteur",
    "Titre",
    "Date de réception",
    "Date contractuelle de réponse",
    "Réviseurs",
    "Statuts",
    "Dates de réponse",
    "Commentaires",
    "Pourquoi il est ici",
    "MOEX AVIS",
]

# Per-column widths (Excel character units), keyed to _COLUMNS_ORDER
_COLUMN_WIDTHS: Dict[str, int] = {
    "Numéro": 12,
    "Indice": 7,
    "Émetteur": 18,
    "Titre": 50,
    "Date de réception": 14,
    "Date contractuelle de réponse": 16,
    "Réviseurs": 28,
    "Statuts": 16,
    "Dates de réponse": 14,
    "Commentaires": 50,
    "Pourquoi il est ici": 40,
    "MOEX AVIS": 22,
}

_MULTILINE_COLUMNS = {"Réviseurs", "Statuts", "Dates de réponse", "Commentaires"}


def _failure_payload(bucket: str, error_msg: str) -> Dict[str, Any]:
    return {
        "success": False,
        "path": None,
        "filename": None,
        "rows_exported": 0,
        "bucket": str(bucket or ""),
        "message": None,
        "error": error_msg,
    }


def _success_payload(bucket: str, path: Path, rows: int, message: Optional[str]) -> Dict[str, Any]:
    return {
        "success": True,
        "path": os.fspath(path.resolve()),
        "filename": path.name,
        "rows_exported": int(rows),
        "bucket": str(bucket or ""),
        "message": message,
        "error": None,
    }


def build_action_moex_bucket_xlsx(ctx, bucket: str, dest_dir: Path) -> Dict[str, Any]:
    """Build an Excel export for one Action MOEX bucket.

    Returns:
        {success, path, filename, rows_exported, bucket, message, error}
    """
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        bucket_str = str(bucket or "")

        # Validate bucket against the canonical list
        try:
            from reporting.counter_attack_query import (
                BUCKET_DISPLAY_ORDER,
                _artifact_path,
            )
        except Exception as e:
            return _failure_payload(bucket_str, f"counter_attack_query import failed: {e}")

        if bucket_str not in BUCKET_DISPLAY_ORDER:
            return _failure_payload(
                bucket_str, f"Unknown bucket: {bucket_str!r}"
            )

        # Locate / read the artifact
        artifact = _artifact_path()
        if not artifact.exists():
            return _failure_payload(
                bucket_str, f"COUNTER_ATTACK_ITEMS.csv not found at {artifact}"
            )

        try:
            df = pd.read_csv(artifact, dtype=_IDENTITY_DTYPES, keep_default_na=False)
        except Exception as e:
            return _failure_payload(bucket_str, f"CSV read failed: {type(e).__name__}: {e}")

        if "action_bucket" not in df.columns:
            return _failure_payload(bucket_str, "CSV missing action_bucket column")

        sub = df[df["action_bucket"].astype(str) == bucket_str]
        rows_count = int(len(sub))

        # Load the DCC reviewer-list helper. ctx may be None for the empty case.
        latest_responses_fn = None
        if ctx is not None and rows_count > 0:
            try:
                from reporting.document_command_center import _get_latest_responses_for_doc as _lr
                latest_responses_fn = _lr
            except Exception as e:
                return _failure_payload(
                    bucket_str,
                    f"document_command_center._get_latest_responses_for_doc unavailable: {e}",
                )

        # Build rows
        export_rows: List[Dict[str, str]] = []
        for _, csv_row in sub.iterrows():
            numero = _safe_str(csv_row.get("numero"))
            indice = _safe_str(csv_row.get("indice"))
            dernier_row = _resolve_dernier_row(ctx, numero, indice) if ctx is not None else None
            latest_responses: List[Dict[str, Any]] = []
            if latest_responses_fn is not None and dernier_row is not None:
                try:
                    latest_responses = latest_responses_fn(ctx, dernier_row) or []
                except Exception as e:
                    # If DCC helper fails for one doc, leave reviewer cells empty rather than abort.
                    print(f"[counter_attack_export] _get_latest_responses_for_doc failed for "
                          f"({numero}, {indice}): {type(e).__name__}: {e}")
                    latest_responses = []
            export_rows.append(_build_row_layout_y(csv_row, dernier_row, latest_responses))

        # Build the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"ACTION_MOEX_{bucket_str}"[:31]  # Excel sheet name max 31 chars

        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        body_font = Font(name="Calibri", size=10)
        wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        plain_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

        # Header row
        for col_idx, header in enumerate(_COLUMNS_ORDER, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Body rows
        for row_idx, item_row in enumerate(export_rows, start=2):
            for col_idx, header in enumerate(_COLUMNS_ORDER, 1):
                value = item_row.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.border = thin_border
                if header in _MULTILINE_COLUMNS or header in ("Titre", "Pourquoi il est ici"):
                    cell.alignment = wrap_align
                else:
                    cell.alignment = plain_align
                # Force string dtype for identity columns to preserve any leading zeros
                if header in ("Numéro", "Indice"):
                    cell.number_format = "@"

        # Column widths
        for col_idx, header in enumerate(_COLUMNS_ORDER, 1):
            try:
                from openpyxl.utils import get_column_letter
                letter = get_column_letter(col_idx)
            except Exception:
                letter = chr(64 + col_idx)
            ws.column_dimensions[letter].width = _COLUMN_WIDTHS.get(header, 18)

        ws.freeze_panes = "A2"
        last_col_letter = chr(64 + len(_COLUMNS_ORDER))
        ws.auto_filter.ref = f"A1:{last_col_letter}{max(len(export_rows) + 1, 1)}"

        # Atomic write — temp file in dest_dir, then rename
        try:
            dest_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return _failure_payload(bucket_str, f"dest_dir not writable: {type(e).__name__}: {e}")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ACTION_MOEX_{bucket_str}_{ts}.xlsx"
        final_path = dest_dir / filename

        with tempfile.NamedTemporaryFile(dir=str(dest_dir), suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        try:
            wb.save(str(tmp_path))
            if final_path.exists():
                final_path.unlink()
            tmp_path.rename(final_path)
        except Exception as e:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
            return _failure_payload(bucket_str, f"workbook save failed: {type(e).__name__}: {e}")

        message = None
        if rows_count == 0:
            message = "Bucket vide — fichier exporté avec en-tête uniquement."
        return _success_payload(bucket_str, final_path, len(export_rows), message)

    except Exception as e:
        return _failure_payload(str(bucket or ""), f"Unexpected: {type(e).__name__}: {e}")
