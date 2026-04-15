"""
reconciliation_engine.py
------------------------
GF UPDATER V3 — Post-processing Reconciliation Engine (Patch F)

Runs AFTER the primary comparison pass.
Processes unresolved MISSING_IN_GED_TRUE and MISSING_IN_GF_TRUE cases using
progressive fuzzy matching to detect:
  - number typos / transposed digits / prefix mismatches
  - wrong numero in GF vs GED
  - same logical document with broken identity
  - routing gaps (document exists in GED but unrouted to correct sheet)
  - title variants of the same document
"""

import re
import datetime as dt
from collections import defaultdict
from difflib import SequenceMatcher
from typing import Optional, Dict, List, Tuple

import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

SCORE_AUTO_RECONCILE   = 0.85
SCORE_PROBABLE_REVIEW  = 0.70

# Title-first pass thresholds (Patch G)
TITLE_SIM_GATE     = 0.55   # hard gate for title-first candidates
TITLE_SIM_STRONG   = 0.80   # → confirmed / reconciled
TITLE_SIM_PROBABLE = 0.65   # → ambiguous

MIG_TRUE_TYPES = {
    "MISSING_IN_GED_TRUE",
    "MISSING_IN_GED",   # legacy
}
MIF_TRUE_TYPES = {
    "MISSING_IN_GF_TRUE",
    "MISSING_IN_GF",    # legacy
}


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY: lot normalization (strip leading letters/zeros)
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_lot(lot_val) -> str:
    """
    Extract numeric core from lot string.
    'B006' → '6', 'I042' → '42', '42' → '42', 'AH07' → '7', '7' → '7'
    """
    s = str(lot_val or "").strip()
    # Remove leading non-numeric chars (letters, zeros combined)
    m = re.search(r'\d+', s)
    if m:
        try:
            return str(int(m.group()))
        except Exception:
            return m.group()
    return s


def _normalize_numero(val) -> str:
    """Strip leading zeros / .0 suffix."""
    if val is None:
        return ""
    try:
        return str(int(float(str(val).replace(",", "").strip())))
    except Exception:
        return str(val).strip()


def _normalize_title(val) -> str:
    """Lowercase, remove extensions, collapse whitespace, dedup tokens."""
    if not val:
        return ""
    s = str(val).strip()
    # Strip GED code prefix before ' - '
    m = re.match(r'^[A-Z0-9][A-Z0-9_\-]{10,}\s+-\s+(.+)$', s)
    if m:
        s = m.group(1).strip()
    s = s.lower()
    s = re.sub(r'\.(pdf|docx?|xlsx?|dwg|txt|pptx?|zip)\s*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[_\-]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip('.,;: ')
    tokens = s.split()
    deduped: list = []
    for t in tokens:
        if not deduped or t != deduped[-1]:
            deduped.append(t)
    return ' '.join(deduped)


def _title_similarity(a: str, b: str) -> float:
    """Token-based Jaccard + Containment similarity."""
    na = _normalize_title(a)
    nb = _normalize_title(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    ta = set(na.split())
    tb = set(nb.split())
    inter = len(ta & tb)
    union_len = len(ta | tb)
    jaccard = inter / union_len if union_len else 0.0
    containment = inter / min(len(ta), len(tb)) if min(len(ta), len(tb)) else 0.0
    return max(jaccard, containment)


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY: numero fuzzy score
# ─────────────────────────────────────────────────────────────────────────────

def _numero_fuzzy_score(num_a: str, num_b: str) -> float:
    """
    Fuzzy similarity between two numero strings.
    Returns 0.0–1.0.

    1.0   → exact match (already handled before calling)
    0.95  → strip-zero match
    0.70+ → one is a prefix/suffix of the other (e.g. 50939 vs 350939)
    0.50+ → high SequenceMatcher ratio (transposed digits, 1-char typo)
    """
    if not num_a or not num_b:
        return 0.0
    if num_a == num_b:
        return 1.0
    na = num_a.lstrip("0")
    nb = num_b.lstrip("0")
    if na == nb:
        return 0.95

    # One is a suffix of the other  (e.g. "50939" inside "350939")
    if na and nb:
        shorter = na if len(na) <= len(nb) else nb
        longer  = nb if len(na) <= len(nb) else na
        if longer.endswith(shorter) or longer.startswith(shorter):
            ratio = len(shorter) / len(longer) if longer else 0.0
            return min(0.90, 0.50 + ratio * 0.45)

    # SequenceMatcher for transpositions / single-digit typos
    ratio = SequenceMatcher(None, na, nb).ratio()
    return ratio * 0.70   # cap at 0.70 for fuzzy


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY: date proximity
# ─────────────────────────────────────────────────────────────────────────────

def _date_proximity_score(date_a, date_b, max_days: int = 180) -> float:
    """1.0 if same date, 0.0 if > max_days apart."""
    def _to_date(v):
        if v is None:
            return None
        if isinstance(v, (dt.date, dt.datetime)):
            return v.date() if isinstance(v, dt.datetime) else v
        try:
            s = str(v).strip()
            if not s or s.lower() in ("none", "n/a", "nan"):
                return None
            return pd.to_datetime(s).date()
        except Exception:
            return None

    d_a = _to_date(date_a)
    d_b = _to_date(date_b)
    if d_a is None or d_b is None:
        return 0.0
    diff = abs((d_a - d_b).days)
    return max(0.0, 1.0 - diff / max_days)


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY: title-first scoring (Patch G — numero is optional)
# ─────────────────────────────────────────────────────────────────────────────

def _score_title_first(
    cand_title: str, cand_date, cand_ind: str,
    cand_lot_num: str, cand_type_doc: str,
    tgt_title: str, tgt_date, tgt_ind: str,
    tgt_lot_num: str, tgt_type_doc: str,
) -> Tuple[float, dict]:
    """
    Title-first scoring for secondary reconciliation passes.
    Caller must pre-filter by emetteur (required gate).
    Numero is NOT a factor — it is deliberately ignored.

    Max score = 1.00:
      title_sim * 0.50   (primary; hard gate at TITLE_SIM_GATE)
      lot_match    0.20
      indice_match 0.15
      date_prox  * 0.10
      type_doc     0.05

    Returns (score, breakdown). -1.0 if hard title gate fails.
    """
    breakdown: dict = {}

    t_sim = _title_similarity(cand_title, tgt_title)
    if t_sim < TITLE_SIM_GATE:
        return -1.0, {"fail_title_sim": round(t_sim, 3)}

    score = t_sim * 0.50
    breakdown["title_sim"]   = round(t_sim, 3)
    breakdown["title_score"] = round(score, 3)

    # Lot match
    if cand_lot_num and tgt_lot_num and cand_lot_num == tgt_lot_num:
        score += 0.20
        breakdown["lot_match"] = 0.20

    # Indice match (exact or adjacent)
    if cand_ind and tgt_ind:
        ci = str(cand_ind).strip().upper()
        ti = str(tgt_ind).strip().upper()
        if ci == ti:
            score += 0.15
            breakdown["indice_match"] = 0.15
        elif abs(_INDICE_ORDER.get(ci, 999) - _INDICE_ORDER.get(ti, 999)) <= 1:
            score += 0.05
            breakdown["indice_adjacent"] = 0.05

    # Date proximity
    d_prox = _date_proximity_score(cand_date, tgt_date, max_days=180)
    d_score = d_prox * 0.10
    score += d_score
    if d_prox > 0:
        breakdown["date_prox"]  = round(d_prox, 3)
        breakdown["date_score"] = round(d_score, 3)

    # Type doc match
    ct = str(cand_type_doc or "").strip().upper()
    tt = str(tgt_type_doc  or "").strip().upper()
    if ct and tt and ct == tt:
        score += 0.05
        breakdown["type_doc_match"] = 0.05

    return round(score, 4), breakdown


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY: indice comparison
# ─────────────────────────────────────────────────────────────────────────────

_INDICE_ORDER = {c: i for i, c in enumerate(
    "0 AA A B C D E F G H I J K L M N O P Q R S T U V W X Y Z".split()
)}


def _indice_is_newer(ind_ged: str, ind_gf: str) -> bool:
    """Return True if ind_ged is strictly newer (higher) than ind_gf."""
    a = str(ind_ged or "").strip().upper()
    b = str(ind_gf  or "").strip().upper()
    if a == b:
        return False
    oa = _INDICE_ORDER.get(a, 999)
    ob = _INDICE_ORDER.get(b, 999)
    return oa > ob


# ─────────────────────────────────────────────────────────────────────────────
# BUILD INDEXES
# ─────────────────────────────────────────────────────────────────────────────

def _build_global_ged_index(dernier_df_all: pd.DataFrame) -> dict:
    """
    Build lookup tables over ALL GED dernier-indice rows (including unrouted).

    Returns:
      {
        "by_numero": {numero_str: [ged_entry, ...]},
        "by_emetteur_lot": {(emetteur, lot_num_str): [ged_entry, ...]},
      }

    Each ged_entry has:
      numero, indice, title, date, emetteur, lot_normalized,
      zone, niveau, type_doc, batiment, doc_id, gf_sheet_name
    """
    by_numero:       dict = defaultdict(list)
    by_emetteur_lot: dict = defaultdict(list)

    for _, row in dernier_df_all.iterrows():
        if not row.get("is_dernier_indice"):
            continue

        num      = _normalize_numero(row.get("numero_normalized") or row.get("numero"))
        emetteur = str(row.get("emetteur") or "").strip().upper()
        lot_raw  = str(row.get("lot_normalized") or row.get("lot") or "")
        lot_num  = _normalize_lot(lot_raw)

        entry = {
            "numero":       num,
            "indice":       str(row.get("indice") or "").strip().upper(),
            "title":        str(row.get("libelle_du_document") or ""),
            "date":         row.get("date_diffusion") or row.get("cree_le"),
            "emetteur":     emetteur,
            "lot_raw":      lot_raw,
            "lot_num":      lot_num,
            "zone":         str(row.get("zone")     or "").strip().upper(),
            "niveau":       str(row.get("niveau")   or "").strip().upper(),
            "type_doc":     str(row.get("type_de_doc") or "").strip().upper(),
            "batiment":     str(row.get("batiment") or "").strip().upper(),
            "doc_id":       str(row.get("doc_id")   or ""),
            "gf_sheet_name": row.get("gf_sheet_name"),
        }

        if num:
            by_numero[num].append(entry)
        if emetteur and lot_num:
            by_emetteur_lot[(emetteur, lot_num)].append(entry)

    return {
        "by_numero":       dict(by_numero),
        "by_emetteur_lot": dict(by_emetteur_lot),
    }


def _build_global_gf_index(gf_by_sheet: dict) -> dict:
    """
    Build lookup tables over ALL GF rows (all sheets).

    Returns:
      {
        "by_numero": {numero_str: [(sheet, ind, gf_row), ...]},
        "by_lot": {lot_num_str: [(sheet, ind, gf_row), ...]},
      }
    """
    by_numero: dict = defaultdict(list)
    by_lot:    dict = defaultdict(list)

    for sheet_name, gf_docs in gf_by_sheet.items():
        for (num_c, ind_c), gf_rows in gf_docs.items():
            for gf_row in gf_rows:
                lot_raw = str(gf_row.get("lot") or "")
                lot_num = _normalize_lot(lot_raw)
                entry = (sheet_name, ind_c, gf_row)
                if num_c:
                    by_numero[num_c].append(entry)
                if lot_num:
                    by_lot[lot_num].append(entry)

    return {
        "by_numero": dict(by_numero),
        "by_lot":    dict(by_lot),
    }


def _build_sheet_emetteur_map(dernier_df_all: pd.DataFrame) -> dict:
    """
    sheet_name → most common emetteur for that sheet
    (from GED documents that successfully routed to it).
    """
    sheet_to_emetteur: dict = defaultdict(lambda: defaultdict(int))
    for _, row in dernier_df_all.iterrows():
        sname = row.get("gf_sheet_name")
        emetteur = str(row.get("emetteur") or "").strip().upper()
        if sname and emetteur:
            sheet_to_emetteur[sname][emetteur] += 1

    result: dict = {}
    for sname, counts in sheet_to_emetteur.items():
        result[sname] = min(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0]
    return result


def _ged_tiebreak_key(entry: dict) -> tuple:
    """Stable ordering for GED candidates when scores tie."""
    return (
        str(entry.get("numero") or ""),
        str(entry.get("indice") or ""),
        str(entry.get("gf_sheet_name") or ""),
        str(entry.get("doc_id") or ""),
    )


def _gf_tiebreak_key(match_tuple: tuple) -> tuple:
    """Stable ordering for GF candidates when scores tie."""
    sheet_name, ind_c, gf_row = match_tuple
    return (
        str(gf_row.get("numero_normalized") or ""),
        str(ind_c or ""),
        str(sheet_name or ""),
        str(gf_row.get("document") or gf_row.get("titre") or ""),
    )


# ─────────────────────────────────────────────────────────────────────────────
# TITLE-FIRST INDEXES (Patch G)
# ─────────────────────────────────────────────────────────────────────────────

def _build_ged_by_emetteur(ged_index: dict) -> dict:
    """
    Flatten ged_index["by_emetteur_lot"] → {emetteur: [ged_entry, ...]}
    Used for title-first search when lot is unknown or when lot-filtered
    pool returns no candidates.
    """
    by_emetteur: dict = defaultdict(list)
    for (em, lot), entries in ged_index["by_emetteur_lot"].items():
        by_emetteur[em].extend(entries)
    return {k: list(v) for k, v in by_emetteur.items()}


def _build_gf_title_index(gf_by_sheet: dict, sheet_emetteur: dict) -> dict:
    """
    Returns {emetteur: [(sheet_name, ind, gf_row), ...]} for all GF rows.
    Used for title-first search in Direction 2 (MISSING_IN_GF → search GF).
    """
    by_emetteur: dict = defaultdict(list)
    for sheet_name, gf_docs in gf_by_sheet.items():
        emetteur = str(sheet_emetteur.get(sheet_name) or "").upper()
        if not emetteur:
            continue
        for (num_c, ind_c), rows in gf_docs.items():
            for gf_row in rows:
                by_emetteur[emetteur].append((sheet_name, ind_c, gf_row))
    return dict(by_emetteur)


# ─────────────────────────────────────────────────────────────────────────────
# SCORING
# ─────────────────────────────────────────────────────────────────────────────

def _score_ged_vs_gf(
    ged_entry:   dict,
    gf_rec:      dict,
    gf_emetteur: str,
    gf_lot_num:  str,
) -> Tuple[float, dict]:
    """
    Score a GED candidate against a GF unmatched record.

    gf_emetteur: derived from the sheet (not stored in gf_rec directly)
    gf_lot_num:  numerically-normalized GF lot

    Returns (total_score, breakdown_dict).
    -1.0 if required gates fail.
    """
    breakdown: dict = {}

    # ── Required gates ────────────────────────────────────────────────────────
    ged_emetteur = str(ged_entry.get("emetteur") or "").upper()
    if ged_emetteur != gf_emetteur.upper():
        return -1.0, {"fail": f"emetteur {ged_emetteur} != {gf_emetteur}"}

    ged_lot_num = str(ged_entry.get("lot_num") or "")
    if ged_lot_num != gf_lot_num:
        return -1.0, {"fail": f"lot {ged_lot_num} != {gf_lot_num}"}

    score = 0.0

    # ── Zone (+0.15 each if both present and match) ───────────────────────────
    ged_zone = str(ged_entry.get("zone") or "").upper()
    gf_zone  = str(gf_rec.get("zone") or "").upper()
    if ged_zone and gf_zone and ged_zone == gf_zone:
        score += 0.15
        breakdown["zone"] = 0.15

    # ── Niveau ───────────────────────────────────────────────────────────────
    ged_niv = str(ged_entry.get("niveau") or "").upper()
    gf_niv  = str(gf_rec.get("niveau") or "").upper()
    if ged_niv and gf_niv and ged_niv == gf_niv:
        score += 0.15
        breakdown["niveau"] = 0.15

    # ── Type doc ─────────────────────────────────────────────────────────────
    ged_type = str(ged_entry.get("type_doc") or "").upper()
    gf_type  = str(gf_rec.get("type_doc") or "").upper()
    if ged_type and gf_type and ged_type == gf_type:
        score += 0.10
        breakdown["type_doc"] = 0.10

    # ── Title similarity (+0.30 weighted) ────────────────────────────────────
    t_sim = _title_similarity(
        gf_rec.get("titre", ""),
        ged_entry.get("title", ""),
    )
    title_contrib = t_sim * 0.30
    score += title_contrib
    breakdown["title_sim"]    = round(t_sim, 3)
    breakdown["title_score"]  = round(title_contrib, 3)

    # ── Date proximity (+0.15 weighted) ──────────────────────────────────────
    d_prox = _date_proximity_score(
        gf_rec.get("date_diffusion"),
        ged_entry.get("date"),
    )
    date_contrib = d_prox * 0.15
    score += date_contrib
    breakdown["date_prox"]  = round(d_prox, 3)
    breakdown["date_score"] = round(date_contrib, 3)

    # ── Numero matching ───────────────────────────────────────────────────────
    gf_num  = str(gf_rec.get("numero_normalized") or gf_rec.get("numero") or "")
    ged_num = str(ged_entry.get("numero") or "")
    if gf_num and ged_num:
        if gf_num == ged_num:
            score += 0.15
            breakdown["numero_exact"] = 0.15
        else:
            n_fuz = _numero_fuzzy_score(gf_num, ged_num)
            if n_fuz > 0.30:
                num_contrib = n_fuz * 0.08
                score += num_contrib
                breakdown["numero_fuzzy"]  = round(n_fuz, 3)
                breakdown["numero_score"]  = round(num_contrib, 3)

    # ── Document code cross-reference (+0.15 bonus) ───────────────────────────
    # When GF's document code field literally contains the GED numero as a
    # token (e.g. GF doc_code = "P17_..._350939", GED numero = "350939"),
    # it's unambiguous evidence of a numero transcription error in GF.
    # The GED numero appears as an underscore-delimited token in doc codes.
    gf_doc_code = str(gf_rec.get("document") or "")
    if ged_num and gf_doc_code and not breakdown.get("numero_exact"):
        if (f"_{ged_num}_" in gf_doc_code or
                gf_doc_code.endswith(f"_{ged_num}") or
                gf_doc_code.endswith(f"_{ged_num}_A") or
                gf_doc_code.endswith(f"_{ged_num}_B")):
            score += 0.15
            breakdown["doc_code_xref"] = 0.15

    return round(score, 4), breakdown


# ─────────────────────────────────────────────────────────────────────────────
# RECONCILIATION RESULT SUBTYPES
# ─────────────────────────────────────────────────────────────────────────────

def _ged_result_subtype(score: float, has_numero_exact: bool, has_numero_fuzzy: bool) -> str:
    """Return the new MISSING_IN_GED_* subtype based on score and matching pattern."""
    if score >= SCORE_AUTO_RECONCILE:
        if has_numero_exact:
            return "MISSING_IN_GED_RECONCILED_BY_FUZZY"   # cross-sheet same-numero
        if has_numero_fuzzy:
            return "MISSING_IN_GED_POSSIBLE_NUMERO_ERROR"
        return "MISSING_IN_GED_RECONCILED_BY_FUZZY"
    elif score >= SCORE_PROBABLE_REVIEW:
        if has_numero_fuzzy:
            return "MISSING_IN_GED_POSSIBLE_NUMERO_ERROR"
        return "MISSING_IN_GED_POSSIBLE_TITLE_VARIANT"
    return "MISSING_IN_GED_AMBIGUOUS_RECONCILIATION"


def _gf_result_subtype(score: float, has_numero_exact: bool, has_numero_fuzzy: bool) -> str:
    """Return the new MISSING_IN_GF_* subtype based on score and matching pattern."""
    if score >= SCORE_AUTO_RECONCILE:
        if has_numero_exact:
            return "MISSING_IN_GF_RECONCILED_BY_FUZZY"
        if has_numero_fuzzy:
            return "MISSING_IN_GF_POSSIBLE_NUMERO_ERROR"
        return "MISSING_IN_GF_RECONCILED_BY_FUZZY"
    elif score >= SCORE_PROBABLE_REVIEW:
        if has_numero_fuzzy:
            return "MISSING_IN_GF_POSSIBLE_NUMERO_ERROR"
        return "MISSING_IN_GF_POSSIBLE_TITLE_VARIANT"
    return "MISSING_IN_GF_AMBIGUOUS_RECONCILIATION"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def run_reconciliation(
    discrepancies:    list,
    dernier_df_all:   "pd.DataFrame",
    gf_by_sheet:      dict,
    responses_df:     "pd.DataFrame",
    sheet_emetteur_override: dict = None,
) -> Tuple[list, list]:
    """
    Post-processing reconciliation pass.

    Args:
      discrepancies    : full list of discrepancy dicts from primary comparison
      dernier_df_all   : ALL GED dernier docs (BEFORE exclusions / routing filter)
      gf_by_sheet      : {sheet_name: {(num, ind): [gf_rows]}} from primary pass
      responses_df     : GED responses DataFrame (for future consultant-footprint)
      sheet_emetteur_override: manual {sheet_name: emetteur} overrides

    Returns:
      (updated_discrepancies, reconciliation_log)

    Updated discrepancies have some entries' flag_type changed to one of:
      MISSING_IN_GED_RECONCILED_BY_FUZZY
      MISSING_IN_GED_POSSIBLE_NUMERO_ERROR
      MISSING_IN_GED_POSSIBLE_TITLE_VARIANT
      MISSING_IN_GED_AMBIGUOUS_RECONCILIATION
      MISSING_IN_GED_HISTORICAL  (when global index confirms newer indice exists)

      MISSING_IN_GF_RECONCILED_BY_FUZZY
      MISSING_IN_GF_POSSIBLE_NUMERO_ERROR
      MISSING_IN_GF_POSSIBLE_TITLE_VARIANT
      MISSING_IN_GF_AMBIGUOUS_RECONCILIATION
    """
    print("  [Recon] Building global indexes...")
    ged_index  = _build_global_ged_index(dernier_df_all)
    gf_index   = _build_global_gf_index(gf_by_sheet)
    sheet_emetteur = _build_sheet_emetteur_map(dernier_df_all)
    if sheet_emetteur_override:
        sheet_emetteur.update(sheet_emetteur_override)

    ged_by_num = ged_index["by_numero"]
    ged_by_el  = ged_index["by_emetteur_lot"]
    gf_by_num  = gf_index["by_numero"]

    recon_log: list = []

    # ── Collect unresolved sets ───────────────────────────────────────────────
    # Indexes into discrepancies list for fast mutation
    mid_true_indices: list = []   # MISSING_IN_GED_TRUE
    mif_true_indices: list = []   # MISSING_IN_GF_TRUE

    for idx, d in enumerate(discrepancies):
        ft = d.get("flag_type", "")
        if ft in MIG_TRUE_TYPES:
            mid_true_indices.append(idx)
        elif ft in MIF_TRUE_TYPES:
            mif_true_indices.append(idx)

    print(f"  [Recon] Unresolved MIG_TRUE: {len(mid_true_indices)}, "
          f"MIF_TRUE: {len(mif_true_indices)}")

    # ── Track already-reconciled GED numeros (to avoid double-matching) ──────
    reconciled_ged_nums: set = set()
    reconciled_gf_pairs: set = set()  # (sheet, num, ind)

    # ══════════════════════════════════════════════════════════════════════════
    # DIRECTION 1: MISSING_IN_GED_TRUE → search global GED index
    # GF has a row with no match in primary pass → try to find it in ALL GED
    # ══════════════════════════════════════════════════════════════════════════
    mid_resolved = 0
    for idx in mid_true_indices:
        d = discrepancies[idx]
        sheet_name = d.get("sheet_name", "")
        gf_num     = d.get("numero", "")
        gf_ind     = d.get("indice", "")
        gf_title   = d.get("document_code", "")  # GF doc title stored here
        gf_lot_raw = ""  # will need from gf_by_sheet

        # Get the actual GF row data (for lot, zone, titre, etc.)
        gf_row = {}
        sheet_data = gf_by_sheet.get(sheet_name, {})
        for gf_rows in sheet_data.get((gf_num, gf_ind), []):
            gf_row = gf_rows
            break
        if not gf_row:
            # Try to find it in any matching key (in case indice stored differently)
            for (num_k, ind_k), rows in sheet_data.items():
                if num_k == gf_num and rows:
                    gf_row = rows[0]
                    break
        if not gf_row:
            gf_row = {"titre": gf_title, "numero_normalized": gf_num}

        gf_lot_raw = str(gf_row.get("lot") or "")
        gf_lot_num = _normalize_lot(gf_lot_raw) if gf_lot_raw else ""

        # Derive emetteur for this GF sheet
        gf_emetteur = sheet_emetteur.get(sheet_name, "")
        if not gf_emetteur:
            # Fallback: try from sheet name (last word after dash)
            parts = sheet_name.upper().replace(" ", "").split("-")
            gf_emetteur = parts[-1] if parts else ""

        # ── LEVEL 1: exact numero lookup in global GED ─────────────────────
        ged_matches_by_num = ged_by_num.get(gf_num, [])
        if ged_matches_by_num:
            # Found in GED! Check if it's historical (newer indice) or routing gap
            best_ged = max(
                ged_matches_by_num,
                key=lambda e: (
                    _INDICE_ORDER.get(e["indice"], 0),
                    _ged_tiebreak_key(e),
                ),
            )
            if _indice_is_newer(best_ged["indice"], gf_ind):
                # GED has a newer revision → the GF row is stale/historical
                new_ft = "MISSING_IN_GED_HISTORICAL"
                rationale = (f"Global GED has {gf_num}/{best_ged['indice']} "
                             f"which is newer than GF {gf_num}/{gf_ind}. "
                             f"Routing was {best_ged.get('gf_sheet_name') or 'UNMATCHED'}.")
                discrepancies[idx]["flag_type"] = new_ft
                discrepancies[idx]["reconciliation_note"] = rationale
                recon_log.append({
                    "direction":       "MISSING_IN_GED",
                    "sheet":           sheet_name,
                    "gf_numero":       gf_num,
                    "gf_indice":       gf_ind,
                    "ged_numero":      best_ged["numero"],
                    "ged_indice":      best_ged["indice"],
                    "ged_sheet":       best_ged.get("gf_sheet_name") or "UNROUTED",
                    "original_type":   "MISSING_IN_GED_TRUE",
                    "new_type":        new_ft,
                    "score":           1.0,
                    "level":           1,
                    "rationale":       rationale,
                })
                reconciled_ged_nums.add(gf_num)
                mid_resolved += 1
                continue

            elif best_ged["indice"] == gf_ind.upper():
                # Same indice — routing mismatch (GED exists but on different sheet)
                new_ft = "MISSING_IN_GED_RECONCILED_BY_FUZZY"
                rationale = (f"GED has {gf_num}/{best_ged['indice']} on sheet "
                             f"'{best_ged.get('gf_sheet_name') or 'UNROUTED'}' "
                             f"but GF recorded it on '{sheet_name}'. Routing mismatch.")
                discrepancies[idx]["flag_type"] = new_ft
                discrepancies[idx]["reconciliation_note"] = rationale
                recon_log.append({
                    "direction":     "MISSING_IN_GED",
                    "sheet":         sheet_name,
                    "gf_numero":     gf_num,
                    "gf_indice":     gf_ind,
                    "ged_numero":    best_ged["numero"],
                    "ged_indice":    best_ged["indice"],
                    "ged_sheet":     best_ged.get("gf_sheet_name") or "UNROUTED",
                    "original_type": "MISSING_IN_GED_TRUE",
                    "new_type":      new_ft,
                    "score":         1.0,
                    "level":         1,
                    "rationale":     rationale,
                })
                reconciled_ged_nums.add(gf_num)
                mid_resolved += 1
                continue

        # ── LEVEL 2–4: fuzzy by emetteur + lot ─────────────────────────────
        if not gf_emetteur or not gf_lot_num:
            continue   # can't do fuzzy without anchors

        candidates = ged_by_el.get((gf_emetteur, gf_lot_num), [])
        if not candidates:
            continue

        best_score  = 0.0
        best_cand   = None
        best_breakdown = {}

        for cand in candidates:
            if cand["numero"] == gf_num:
                continue   # already tried exact — this is a different-indice case
            sc, bd = _score_ged_vs_gf(cand, gf_row, gf_emetteur, gf_lot_num)
            if (
                sc > best_score or
                (
                    best_cand is not None and
                    sc == best_score and
                    _ged_tiebreak_key(cand) < _ged_tiebreak_key(best_cand)
                ) or
                (best_cand is None and sc == best_score and sc > 0)
            ):
                best_score     = sc
                best_cand      = cand
                best_breakdown = bd

        if best_cand is None or best_score < SCORE_PROBABLE_REVIEW:
            continue

        # Determine subtype
        has_num_exact = best_breakdown.get("numero_exact", 0) > 0
        has_num_fuzzy = best_breakdown.get("numero_fuzzy", 0) > 0
        new_ft = _ged_result_subtype(best_score, has_num_exact, has_num_fuzzy)

        discrepancies[idx]["flag_type"] = new_ft
        discrepancies[idx]["reconciliation_note"] = (
            f"Fuzzy match: GED {best_cand['numero']}/{best_cand['indice']} "
            f"(score={best_score}, title_sim={best_breakdown.get('title_sim',0)})"
        )
        recon_log.append({
            "direction":       "MISSING_IN_GED",
            "sheet":           sheet_name,
            "gf_numero":       gf_num,
            "gf_indice":       gf_ind,
            "ged_numero":      best_cand["numero"],
            "ged_indice":      best_cand["indice"],
            "ged_sheet":       best_cand.get("gf_sheet_name") or "UNROUTED",
            "original_type":   "MISSING_IN_GED_TRUE",
            "new_type":        new_ft,
            "score":           best_score,
            "level":           2,
            "title_sim":       best_breakdown.get("title_sim", 0),
            "date_prox":       best_breakdown.get("date_prox", 0),
            "numero_fuzzy":    best_breakdown.get("numero_fuzzy", 0),
            "rationale":       str(best_breakdown),
        })
        mid_resolved += 1

    # ══════════════════════════════════════════════════════════════════════════
    # DIRECTION 2: MISSING_IN_GF_TRUE → search global GF index
    # GED family has no GF match → try to find it on a different sheet or num
    # ══════════════════════════════════════════════════════════════════════════
    mif_resolved = 0
    for idx in mif_true_indices:
        d = discrepancies[idx]
        sheet_name  = d.get("sheet_name", "")
        ged_num     = d.get("numero", "")
        ged_ind     = d.get("indice", "")
        ged_title   = d.get("document_code", "")
        ged_date    = d.get("gfi_date", "")
        ged_lot_raw = d.get("gfi_lot", "")
        ged_lot_num = _normalize_lot(ged_lot_raw) if ged_lot_raw else ""
        ged_emetteur = sheet_emetteur.get(sheet_name, "")

        # ── LEVEL 1: exact numero on different GF sheet ─────────────────────
        gf_occ = gf_by_num.get(ged_num, [])
        other_sheet_occ = [(sn, ind, row) for sn, ind, row in gf_occ
                           if sn != sheet_name]

        if other_sheet_occ:
            best_other = sorted(other_sheet_occ, key=_gf_tiebreak_key)[0]
            new_ft = "MISSING_IN_GF_RECONCILED_BY_FUZZY"
            rationale = (f"GF has {ged_num}/{best_other[1]} on sheet "
                         f"'{best_other[0]}' but GED routed to '{sheet_name}'. "
                         f"Possible sheet/routing mismatch.")
            discrepancies[idx]["flag_type"] = new_ft
            discrepancies[idx]["reconciliation_note"] = rationale
            recon_log.append({
                "direction":     "MISSING_IN_GF",
                "sheet":         sheet_name,
                "ged_numero":    ged_num,
                "ged_indice":    ged_ind,
                "gf_numero":     ged_num,
                "gf_indice":     best_other[1],
                "gf_sheet":      best_other[0],
                "original_type": "MISSING_IN_GF_TRUE",
                "new_type":      new_ft,
                "score":         1.0,
                "level":         1,
                "rationale":     rationale,
            })
            mif_resolved += 1
            continue

        # ── LEVEL 2–4: fuzzy search in same emetteur/lot on all GF sheets ──
        if not ged_emetteur or not ged_lot_num:
            continue

        # Build GF candidate set for this emetteur/lot
        gf_candidates_for_lot = gf_index["by_lot"].get(ged_lot_num, [])
        if not gf_candidates_for_lot:
            continue

        best_score     = 0.0
        best_gf_match  = None
        best_bd        = {}

        # Build a fake gf_rec from ged info for scoring
        ged_as_gf_rec = {
            "titre":            ged_title,
            "numero_normalized": ged_num,
            "date_diffusion":   ged_date,
            "lot":              ged_lot_raw,
            "type_doc":         d.get("gfi_type_doc", ""),
            "zone":             "",
            "niveau":           "",
        }
        # GED entry dict
        ged_entry_for_score = {
            "emetteur":  ged_emetteur,
            "lot_num":   ged_lot_num,
            "numero":    ged_num,
            "indice":    ged_ind,
            "title":     ged_title,
            "date":      ged_date,
            "zone":      "",
            "niveau":    "",
            "type_doc":  d.get("gfi_type_doc", ""),
        }

        for gf_sheet, gf_ind_c, gf_row in gf_candidates_for_lot:
            if gf_sheet == sheet_name and gf_row.get("numero_normalized") == ged_num:
                continue  # already tried

            gf_row_num = str(gf_row.get("numero_normalized") or "")
            gf_lot_r   = str(gf_row.get("lot") or "")
            gf_lot_n   = _normalize_lot(gf_lot_r)

            sc, bd = _score_ged_vs_gf(
                ged_entry_for_score,
                gf_row,
                ged_emetteur,
                gf_lot_n,
            )
            if (
                sc > best_score or
                (
                    best_gf_match is not None and
                    sc == best_score and
                    _gf_tiebreak_key((gf_sheet, gf_ind_c, gf_row)) < _gf_tiebreak_key(best_gf_match)
                ) or
                (best_gf_match is None and sc == best_score and sc > 0)
            ):
                best_score    = sc
                best_gf_match = (gf_sheet, gf_ind_c, gf_row)
                best_bd       = bd

        if best_gf_match is None or best_score < SCORE_PROBABLE_REVIEW:
            continue

        gf_sheet_m, gf_ind_m, gf_row_m = best_gf_match
        has_num_exact = best_bd.get("numero_exact", 0) > 0
        has_num_fuzzy = best_bd.get("numero_fuzzy", 0) > 0
        new_ft = _gf_result_subtype(best_score, has_num_exact, has_num_fuzzy)

        discrepancies[idx]["flag_type"] = new_ft
        discrepancies[idx]["reconciliation_note"] = (
            f"Fuzzy match: GF {gf_row_m.get('numero_normalized')}/{gf_ind_m} "
            f"on sheet '{gf_sheet_m}' "
            f"(score={best_score}, title_sim={best_bd.get('title_sim', 0)})"
        )
        recon_log.append({
            "direction":     "MISSING_IN_GF",
            "sheet":         sheet_name,
            "ged_numero":    ged_num,
            "ged_indice":    ged_ind,
            "gf_numero":     gf_row_m.get("numero_normalized", ""),
            "gf_indice":     gf_ind_m,
            "gf_sheet":      gf_sheet_m,
            "original_type": "MISSING_IN_GF_TRUE",
            "new_type":      new_ft,
            "score":         best_score,
            "level":         2,
            "title_sim":     best_bd.get("title_sim", 0),
            "date_prox":     best_bd.get("date_prox", 0),
            "numero_fuzzy":  best_bd.get("numero_fuzzy", 0),
            "rationale":     str(best_bd),
        })
        mif_resolved += 1

    print(f"  [Recon L1-L2] Resolved: MIG_TRUE −{mid_resolved}, MIF_TRUE −{mif_resolved}")

    # ══════════════════════════════════════════════════════════════════════════
    # PASS 3a — Promote POSSIBLE_NUMERO_ERROR → GF_NUMERO_TYPO_CONFIRMED
    # If the fuzzy-numero match was already made AND title_sim >= 0.85,
    # the numero error is confirmed, not merely possible.
    # ══════════════════════════════════════════════════════════════════════════
    # Build lookup: (sheet, gf_num, gf_ind) → list of recon_log entries
    recon_by_key: dict = defaultdict(list)
    for le in recon_log:
        key = (le.get("sheet", ""), str(le.get("gf_numero", "")), str(le.get("gf_indice", "")))
        recon_by_key[key].append(le)

    pne_promoted = 0
    for idx, d in enumerate(discrepancies):
        if d.get("flag_type") != "MISSING_IN_GED_POSSIBLE_NUMERO_ERROR":
            continue
        key = (d.get("sheet_name", ""), str(d.get("numero", "")), str(d.get("indice", "")))
        for le in recon_by_key.get(key, []):
            title_sim_val = float(le.get("title_sim") or 0)
            if title_sim_val >= 0.85:
                d["flag_type"] = "MISSING_IN_GED_GF_NUMERO_TYPO_CONFIRMED"
                d["reconciliation_note"] = (
                    (d.get("reconciliation_note") or "") +
                    f" | Confirmed by title (sim={title_sim_val:.2f}) → GF_NUMERO_TYPO_CONFIRMED"
                )
                le["new_type"]    = "MISSING_IN_GED_GF_NUMERO_TYPO_CONFIRMED"
                le["promoted_by"] = "title_confirmation"
                pne_promoted += 1
                break  # first log entry is enough

    print(f"  [Recon Pass 3a] POSSIBLE_NUMERO_ERROR promoted: {pne_promoted}")

    # ══════════════════════════════════════════════════════════════════════════
    # PASS 3b — Title-first search for remaining MISSING_IN_GED_TRUE
    # Emetteur is locked; lot preferred but optional; numero ignored.
    # ══════════════════════════════════════════════════════════════════════════
    ged_by_emetteur = _build_ged_by_emetteur(ged_index)
    mid_title_resolved = 0

    for idx in mid_true_indices:
        d = discrepancies[idx]
        if d.get("flag_type") not in {"MISSING_IN_GED_TRUE", "MISSING_IN_GED"}:
            continue

        sheet_name  = d.get("sheet_name", "")
        gf_num      = d.get("numero", "")
        gf_ind      = d.get("indice", "")

        # Get the GF row (needed for titre, lot, type_doc, date)
        gf_row: dict = {}
        sheet_data = gf_by_sheet.get(sheet_name, {})
        for gf_rows_l in sheet_data.get((gf_num, gf_ind), []):
            gf_row = gf_rows_l
            break
        if not gf_row:
            for (num_k, _ind_k), rows in sheet_data.items():
                if num_k == gf_num and rows:
                    gf_row = rows[0]
                    break

        gf_titre    = str(gf_row.get("titre") or "")
        gf_lot_raw  = str(gf_row.get("lot") or "")
        gf_lot_num  = _normalize_lot(gf_lot_raw) if gf_lot_raw else ""
        gf_type_doc = str(gf_row.get("type_doc") or "")
        gf_date     = gf_row.get("date_diffusion")
        gf_emetteur = sheet_emetteur.get(sheet_name, "")

        if not gf_titre or not gf_emetteur:
            continue  # can't search without title or emetteur

        # Use lot-filtered candidates first; fall back to full emetteur pool
        if gf_lot_num:
            candidates = ged_by_el.get((gf_emetteur, gf_lot_num), [])
        else:
            candidates = ged_by_emetteur.get(gf_emetteur, [])
        if not candidates and gf_lot_num:
            candidates = ged_by_emetteur.get(gf_emetteur, [])

        if not candidates:
            continue

        best_score = 0.0
        best_cand  = None
        best_bd: dict = {}

        for cand in candidates:
            if cand["numero"] in reconciled_ged_nums:
                continue
            if cand["numero"] == gf_num:
                continue  # same numero already failed in Level 1+2

            sc, bd = _score_title_first(
                cand_title   = cand.get("title", ""),
                cand_date    = cand.get("date"),
                cand_ind     = cand.get("indice", ""),
                cand_lot_num = cand.get("lot_num", ""),
                cand_type_doc= cand.get("type_doc", ""),
                tgt_title    = gf_titre,
                tgt_date     = gf_date,
                tgt_ind      = gf_ind,
                tgt_lot_num  = gf_lot_num,
                tgt_type_doc = gf_type_doc,
            )
            if (
                sc > best_score or
                (
                    best_cand is not None and
                    sc == best_score and
                    _ged_tiebreak_key(cand) < _ged_tiebreak_key(best_cand)
                ) or
                (best_cand is None and sc == best_score and sc > 0)
            ):
                best_score = sc
                best_cand  = cand
                best_bd    = bd

        if best_cand is None or best_score < TITLE_SIM_PROBABLE:
            continue

        n_fuz = _numero_fuzzy_score(gf_num, best_cand["numero"])
        if best_score >= TITLE_SIM_STRONG:
            new_ft = ("MISSING_IN_GED_GF_NUMERO_TYPO_CONFIRMED"
                      if n_fuz >= 0.50
                      else "MISSING_IN_GED_RECONCILED_BY_TITLE")
        else:
            new_ft = "MISSING_IN_GED_AMBIGUOUS_TITLE_MATCH"

        d["flag_type"] = new_ft
        d["reconciliation_note"] = (
            f"Title-first L3: GED {best_cand['numero']}/{best_cand['indice']} "
            f"(title_sim={best_bd.get('title_sim', 0):.2f}, score={best_score:.3f})"
        )
        recon_log.append({
            "direction":     "MISSING_IN_GED",
            "sheet":         sheet_name,
            "gf_numero":     gf_num,
            "gf_indice":     gf_ind,
            "ged_numero":    best_cand["numero"],
            "ged_indice":    best_cand["indice"],
            "ged_sheet":     best_cand.get("gf_sheet_name") or "UNROUTED",
            "original_type": "MISSING_IN_GED_TRUE",
            "new_type":      new_ft,
            "score":         best_score,
            "level":         3,
            "title_sim":     best_bd.get("title_sim", 0),
            "date_prox":     best_bd.get("date_prox", 0),
            "numero_fuzzy":  n_fuz,
            "rationale":     str(best_bd),
        })
        reconciled_ged_nums.add(best_cand["numero"])
        mid_title_resolved += 1

    print(f"  [Recon Pass 3b] MIG title-first resolved: {mid_title_resolved}")

    # ══════════════════════════════════════════════════════════════════════════
    # PASS 4 — Title-first search for remaining MISSING_IN_GF_TRUE
    # Search GF rows by title for each unresolved GED family.
    # ══════════════════════════════════════════════════════════════════════════
    gf_title_idx = _build_gf_title_index(gf_by_sheet, sheet_emetteur)
    mif_title_resolved = 0

    for idx in mif_true_indices:
        d = discrepancies[idx]
        if d.get("flag_type") not in {"MISSING_IN_GF_TRUE", "MISSING_IN_GF"}:
            continue

        ged_num      = d.get("numero", "")
        ged_ind      = d.get("indice", "")
        sheet_name   = d.get("sheet_name", "")
        ged_title    = d.get("gfi_titre", "") or d.get("document_code", "")
        ged_date_raw = d.get("gfi_date", "")
        ged_lot_raw  = d.get("gfi_lot", "")
        ged_lot_num  = _normalize_lot(ged_lot_raw) if ged_lot_raw else ""
        ged_type_doc = d.get("gfi_type_doc", "")
        ged_emetteur = sheet_emetteur.get(sheet_name, "")

        if not ged_title or not ged_emetteur:
            continue

        candidates_gf = gf_title_idx.get(ged_emetteur, [])
        if not candidates_gf:
            continue

        best_score    = 0.0
        best_gf_match: Optional[tuple] = None
        best_bd: dict = {}

        for gf_sheet, gf_ind_c, gf_row in candidates_gf:
            gf_num_c = str(gf_row.get("numero_normalized") or "")
            if gf_num_c == ged_num:
                continue  # same numero — already tried
            if (gf_sheet, gf_num_c, gf_ind_c) in reconciled_gf_pairs:
                continue

            gf_lot_raw_c = str(gf_row.get("lot") or "")
            gf_lot_num_c = _normalize_lot(gf_lot_raw_c) if gf_lot_raw_c else ""

            sc, bd = _score_title_first(
                cand_title   = str(gf_row.get("titre") or ""),
                cand_date    = gf_row.get("date_diffusion"),
                cand_ind     = gf_ind_c,
                cand_lot_num = gf_lot_num_c,
                cand_type_doc= str(gf_row.get("type_doc") or ""),
                tgt_title    = ged_title,
                tgt_date     = ged_date_raw,
                tgt_ind      = ged_ind,
                tgt_lot_num  = ged_lot_num,
                tgt_type_doc = ged_type_doc,
            )
            if (
                sc > best_score or
                (
                    best_gf_match is not None and
                    sc == best_score and
                    _gf_tiebreak_key((gf_sheet, gf_ind_c, gf_row)) < _gf_tiebreak_key(best_gf_match)
                ) or
                (best_gf_match is None and sc == best_score and sc > 0)
            ):
                best_score    = sc
                best_gf_match = (gf_sheet, gf_ind_c, gf_row)
                best_bd       = bd

        if best_gf_match is None or best_score < TITLE_SIM_PROBABLE:
            continue

        gf_sheet_m, gf_ind_m, gf_row_m = best_gf_match
        gf_num_m = str(gf_row_m.get("numero_normalized") or "")
        n_fuz    = _numero_fuzzy_score(ged_num, gf_num_m)

        if best_score >= TITLE_SIM_STRONG:
            new_ft = ("MISSING_IN_GF_GF_NUMERO_TYPO"
                      if n_fuz >= 0.50
                      else "MISSING_IN_GF_RECONCILED_BY_TITLE")
        else:
            new_ft = "MISSING_IN_GF_AMBIGUOUS_TITLE_MATCH"

        d["flag_type"] = new_ft
        d["reconciliation_note"] = (
            f"Title-first L4: GF {gf_num_m}/{gf_ind_m} on '{gf_sheet_m}' "
            f"(title_sim={best_bd.get('title_sim', 0):.2f}, score={best_score:.3f})"
        )
        recon_log.append({
            "direction":     "MISSING_IN_GF",
            "sheet":         sheet_name,
            "ged_numero":    ged_num,
            "ged_indice":    ged_ind,
            "gf_numero":     gf_num_m,
            "gf_indice":     gf_ind_m,
            "gf_sheet":      gf_sheet_m,
            "original_type": "MISSING_IN_GF_TRUE",
            "new_type":      new_ft,
            "score":         best_score,
            "level":         3,
            "title_sim":     best_bd.get("title_sim", 0),
            "date_prox":     best_bd.get("date_prox", 0),
            "numero_fuzzy":  n_fuz,
            "rationale":     str(best_bd),
        })
        reconciled_gf_pairs.add((gf_sheet_m, gf_num_m, gf_ind_m))
        mif_title_resolved += 1

    print(f"  [Recon Pass 4]  MIF title-first resolved: {mif_title_resolved}")

    return discrepancies, recon_log


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT WRITER
# ─────────────────────────────────────────────────────────────────────────────

def write_reconciliation_outputs(
    recon_log:    list,
    log_path:     str,
    summary_path: str,
):
    """
    Write RECONCILIATION_LOG.xlsx and debug/reconciliation_summary.xlsx.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from pathlib import Path
    from collections import Counter

    Path(log_path).parent.mkdir(parents=True, exist_ok=True)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)

    # ── RECONCILIATION_LOG ────────────────────────────────────────────────────
    log_headers = [
        "Direction", "Sheet", "GF Numéro", "GF Indice",
        "GED Numéro", "GED Indice", "GED/GF Sheet",
        "Original Type", "New Type",
        "Score", "Level",
        "Title Sim", "Date Prox", "Numero Fuzzy",
        "Rationale",
    ]
    colour_map = {
        # Level 1-2 (numero-based)
        "MISSING_IN_GED_HISTORICAL":                   "E8F5E9",
        "MISSING_IN_GED_RECONCILED_BY_FUZZY":          "C8E6C9",
        "MISSING_IN_GED_POSSIBLE_NUMERO_ERROR":        "FFE0B2",
        "MISSING_IN_GED_POSSIBLE_TITLE_VARIANT":       "FFF9C4",
        "MISSING_IN_GED_AMBIGUOUS_RECONCILIATION":     "EEEEEE",
        "MISSING_IN_GF_RECONCILED_BY_FUZZY":           "BBDEFB",
        "MISSING_IN_GF_POSSIBLE_NUMERO_ERROR":         "FFE0B2",
        "MISSING_IN_GF_POSSIBLE_TITLE_VARIANT":        "FFF9C4",
        "MISSING_IN_GF_AMBIGUOUS_RECONCILIATION":      "EEEEEE",
        # Level 3 (title-first)
        "MISSING_IN_GED_GF_NUMERO_TYPO_CONFIRMED":     "FF8F00",  # amber — confirmed typo
        "MISSING_IN_GED_RECONCILED_BY_TITLE":          "A5D6A7",  # green — title match
        "MISSING_IN_GED_AMBIGUOUS_TITLE_MATCH":        "FFF176",  # yellow — ambiguous
        "MISSING_IN_GF_GF_NUMERO_TYPO":                "FF8F00",  # amber
        "MISSING_IN_GF_RECONCILED_BY_TITLE":           "90CAF9",  # blue-green
        "MISSING_IN_GF_AMBIGUOUS_TITLE_MATCH":         "FFF176",  # yellow
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation Log"

    for ci, h in enumerate(log_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A237E")

    for ri, entry in enumerate(recon_log, 2):
        new_type = entry.get("new_type", "")
        fill_hex = colour_map.get(new_type, "FFFFFF")
        row_data = [
            entry.get("direction", ""),
            entry.get("sheet", ""),
            entry.get("gf_numero", ""),
            entry.get("gf_indice", ""),
            entry.get("ged_numero", ""),
            entry.get("ged_indice", ""),
            entry.get("ged_sheet", entry.get("gf_sheet", "")),
            entry.get("original_type", ""),
            new_type,
            round(float(entry.get("score", 0)), 3),
            entry.get("level", ""),
            round(float(entry.get("title_sim", 0)), 3) if entry.get("title_sim") else "",
            round(float(entry.get("date_prox",  0)), 3) if entry.get("date_prox")  else "",
            round(float(entry.get("numero_fuzzy", 0)), 3) if entry.get("numero_fuzzy") else "",
            entry.get("rationale", "")[:200],
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_hex)

    widths = [14, 45, 12, 10, 12, 10, 45, 40, 45, 8, 6, 10, 10, 12, 120]
    for col_idx, width in zip(range(1, len(log_headers) + 1), widths):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(log_path)

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    counts = Counter(e.get("new_type") for e in recon_log)
    dir_counts = Counter(e.get("direction") for e in recon_log)
    level_counts = Counter(e.get("level") for e in recon_log)

    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.cell(row=1, column=1, value="Reconciliation Summary").font = Font(bold=True, size=12)

    ws_s.cell(row=3, column=1, value="By Direction").font = Font(bold=True)
    for ri, (k, v) in enumerate(dir_counts.items(), 4):
        ws_s.cell(row=ri, column=1, value=k)
        ws_s.cell(row=ri, column=2, value=v)

    ws_s.cell(row=8, column=1, value="By Level").font = Font(bold=True)
    for ri, (k, v) in enumerate(level_counts.items(), 9):
        ws_s.cell(row=ri, column=1, value=f"Level {k}")
        ws_s.cell(row=ri, column=2, value=v)

    ws_s.cell(row=13, column=1, value="By New Type").font = Font(bold=True)
    for ri, (k, v) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 14):
        ws_s.cell(row=ri, column=1, value=k)
        ws_s.cell(row=ri, column=2, value=v)

    ws_s.column_dimensions["A"].width = 50
    ws_s.column_dimensions["B"].width = 10
    wb_s.save(summary_path)

    print(f"    [Recon] Log written: {len(recon_log)} entries → {log_path}")
    print(f"    [Recon] Summary → {summary_path}")
    print(f"    [Recon] By new type: {dict(counts)}")
