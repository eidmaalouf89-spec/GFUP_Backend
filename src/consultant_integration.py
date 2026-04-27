"""
consultant_integration.py
JANSA VISASIST — Consultant Integration Orchestrator
Version 1.0 — April 2026

Integrates consultant PDF reports into the main GF Updater V3 pipeline.

Pipeline steps:
  A. Load consultant_reports.xlsx (or rebuild from PDFs)
  B. Load normalized GED document universe
  C. Match consultant rows to GED documents
  D. Generate output/consultant_match_report.xlsx  (REQUIRED before GF update)
  E. Enrich internal GED model with consultant data
  F. Stage 1 GF output: date + status updates only
  G. Stage 2 GF output: date + status + OBSERVATIONS updates

Usage:
    cd "GF updater v3"
    python src/consultant_integration.py

Or from Python:
    from src.consultant_integration import run_consultant_integration
    run_consultant_integration()
"""

import logging
import sys
from pathlib import Path

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Path setup (allows running as script or importing as module)
# ---------------------------------------------------------------------------

_HERE = Path(__file__).resolve().parent
_PROJECT_ROOT = _HERE.parent   # src → project root

if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))


# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------

import openpyxl
import pandas as pd

from consultant_ingest import build_consultant_reports
from consultant_matcher import (
    build_ged_index,
    match_all_consultants,
    build_enrichment_records,
)
from consultant_match_report import write_match_report
from consultant_gf_writer import write_gf_enriched


# ---------------------------------------------------------------------------
# Default paths (relative to project root)
# ---------------------------------------------------------------------------

def _proj(relative: str) -> Path:
    return _PROJECT_ROOT / relative

INPUT_CONSULTANT_ROOT = _proj("input/consultant_reports")
INPUT_GED             = _proj("input/GED_export.xlsx")
INPUT_GF              = _proj("output/GF_V0_CLEAN.xlsx")

OUTPUT_CONSULTANT_WB  = _proj("output/consultant_reports.xlsx")
OUTPUT_MATCH_REPORT   = _proj("output/consultant_match_report.xlsx")
OUTPUT_GF_STAGE1      = _proj("output/GF_consultant_enriched_stage1.xlsx")
OUTPUT_GF_STAGE2      = _proj("output/GF_consultant_enriched_stage2.xlsx")


# ---------------------------------------------------------------------------
# Step A: Load or rebuild consultant workbook
# ---------------------------------------------------------------------------

def _load_consultant_rows(
    wb_path: Path,
    rebuild: bool = False,
    input_root: Path | None = None,
) -> dict:
    """
    Load consultant rows from existing workbook (or rebuild if requested).

    Returns {source_key: [row_dict, ...], ...}
    """
    if rebuild and input_root:
        logger.info("Rebuilding consultant_reports.xlsx from PDFs...")
        result = build_consultant_reports(
            input_root=input_root,
            output_path=wb_path,
        )
        logger.info("Workbook rebuilt: %s", wb_path)
    else:
        if not wb_path.exists():
            raise FileNotFoundError(
                f"consultant_reports.xlsx not found: {wb_path}\n"
                "Run build_consultant_reports() first or pass rebuild=True"
            )
        logger.info("Loading existing consultant_reports.xlsx: %s", wb_path)

    # Load rows from workbook
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    rows_by_source = {}

    sheet_to_source = {
        "RAPPORT_LE_SOMMER": "LE_SOMMER",
        "RAPPORT_AVLS":      "AVLS",
        "RAPPORT_TERRELL":   "TERRELL",
        "RAPPORT_SOCOTEC":   "SOCOTEC",
    }

    for sheet_name, source_key in sheet_to_source.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook — skipping", sheet_name)
            rows_by_source[source_key] = []
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            rows_by_source[source_key] = []
            continue

        header = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        data   = []
        for row in all_rows[1:]:
            rec = {header[i]: (str(v).strip() if v is not None else "")
                   for i, v in enumerate(row)
                   if i < len(header)}
            data.append(rec)

        rows_by_source[source_key] = data
        logger.info("[%s] Loaded %d rows from '%s'", source_key, len(data), sheet_name)

    wb.close()
    return rows_by_source


# ---------------------------------------------------------------------------
# Step B: Load normalized GED universe
# ---------------------------------------------------------------------------

def _load_ged_docs(ged_path: Path) -> pd.DataFrame:
    """Load and normalize GED documents, including SAS response enrichment."""
    from read_raw import read_ged
    from normalize import load_mapping, normalize_docs, enrich_docs_with_sas

    logger.info("Loading GED export: %s", ged_path)
    docs_df, responses_df, _ = read_ged(str(ged_path))
    mapping = load_mapping()
    docs_norm = normalize_docs(docs_df, mapping)
    docs_norm = enrich_docs_with_sas(docs_norm, responses_df)
    logger.info("GED docs loaded: %d rows", len(docs_norm))
    return docs_norm


# ---------------------------------------------------------------------------
# Enrichment conflict detection
# ---------------------------------------------------------------------------

def _detect_conflicts(
    enrichments: list,
    docs_df:     pd.DataFrame,
) -> list:
    """
    Flag enrichments where consultant status disagrees with GED status.
    Both values are kept — conflict is flagged in match_flag field.
    """
    # Build GED doc lookup by doc_id for quick access
    ged_by_id = {r["doc_id"]: r for _, r in docs_df.iterrows()} if docs_df is not None else {}

    result = []
    for rec in enrichments:
        doc_id = rec.get("doc_id", "")
        ged_doc = ged_by_id.get(doc_id, {})

        # We don't have a "status" in GED docs directly (that lives in responses_df)
        # Mark conflict flag as False for now; can be enriched later
        rec_out = dict(rec)
        rec_out["has_conflict"] = False
        rec_out["conflict_note"] = ""
        result.append(rec_out)

    return result


# ---------------------------------------------------------------------------
# Main integration pipeline
# ---------------------------------------------------------------------------

def run_consultant_integration(
    rebuild_consultant_wb: bool = False,
    skip_gf_update:        bool = False,
) -> dict:
    """
    Full consultant integration pipeline.

    Parameters
    ----------
    rebuild_consultant_wb : if True, re-parse PDFs instead of using cached workbook
    skip_gf_update        : if True, skip Stage 1 and Stage 2 GF outputs

    Returns
    -------
    dict with all pipeline results
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    logger.info("=" * 72)
    logger.info("JANSA VISASIST — Consultant Integration Pipeline")
    logger.info("=" * 72)

    # ── A. Load consultant rows ────────────────────────────────────────────
    logger.info("\n── STEP A: Load consultant workbook ──")
    rows_by_source = _load_consultant_rows(
        wb_path    = OUTPUT_CONSULTANT_WB,
        rebuild    = rebuild_consultant_wb,
        input_root = INPUT_CONSULTANT_ROOT,
    )

    total_consultant_rows = sum(len(v) for v in rows_by_source.values())
    logger.info("Total consultant rows loaded: %d", total_consultant_rows)
    for src, rows in rows_by_source.items():
        logger.info("  %-14s: %d rows", src, len(rows))

    # ── B. Load GED documents ─────────────────────────────────────────────
    logger.info("\n── STEP B: Load GED document universe ──")
    docs_df = _load_ged_docs(INPUT_GED)

    # ── C. Build GED index ────────────────────────────────────────────────
    logger.info("\n── STEP C: Build GED lookup index ──")
    ged_index = build_ged_index(docs_df)

    # ── D. Match consultant rows to GED ───────────────────────────────────
    logger.info("\n── STEP D: Match consultant rows to GED documents ──")
    match_results = match_all_consultants(rows_by_source, ged_index)

    # ── E. Write match report ─────────────────────────────────────────────
    logger.info("\n── STEP E: Write consultant_match_report.xlsx ──")
    write_match_report(match_results, OUTPUT_MATCH_REPORT)
    logger.info("Match report written: %s", OUTPUT_MATCH_REPORT)

    # ── F. Build enrichment records ───────────────────────────────────────
    logger.info("\n── STEP F: Build enrichment records ──")
    enrichments = build_enrichment_records(match_results["matched"])
    enrichments = _detect_conflicts(enrichments, docs_df)
    logger.info("Enrichment records: %d", len(enrichments))

    # Confidence breakdown across ALL matched rows (including those excluded)
    from collections import Counter as _Counter
    all_matched  = match_results["matched"]
    conf_counter = _Counter(r.get("consultant_match_confidence", "") for r in all_matched)
    method_counter = _Counter(r.get("consultant_match_method", "") for r in all_matched)
    enrich_eligible = conf_counter.get("HIGH", 0) + conf_counter.get("MEDIUM", 0)

    logger.info("\n  ── Confidence breakdown (matched rows) ──")
    for conf, cnt in sorted(conf_counter.items()):
        eligible = conf in ("HIGH", "MEDIUM")
        logger.info("    %-26s: %4d  %s", conf, cnt,
                    "(will enrich GF)" if eligible else "(EXCLUDED from GF)")
    logger.info("  Enrich-eligible total : %d / %d matched",
                enrich_eligible, len(all_matched))

    logger.info("\n  ── Match method breakdown (matched rows) ──")
    for method, cnt in sorted(method_counter.items()):
        logger.info("    %-36s: %4d", method, cnt)

    gf_stage1_result = None
    gf_stage2_result = None

    # ── DEPRECATED: Direct GF write stages (Steps G and H) ───────────────────
    # As of Step 8 (FLAT_GED_REPORT_COMPOSITION.md §9), write_gf_enriched() is
    # DEPRECATED from the truth path.  GF_consultant_enriched_stage1.xlsx and
    # GF_consultant_enriched_stage2.xlsx are NOT authoritative GF outputs.
    # They are not registered in run_memory.db and bypass pipeline composition.
    #
    # The correct path:  consultant_match_report.xlsx → stage_report_memory →
    #                    effective_responses_df → stage_write_gf → GF_V0_CLEAN.xlsx
    #
    # consultant_gf_writer may be repurposed as a pipeline artifact writer
    # (reading from effective_responses_df) in a future step, but must not be
    # used to produce deliverable GF files until that repurposing is implemented.
    #
    # skip_gf_update is forced True here to prevent accidental direct-write runs.
    # Remove this guard only if explicitly reinstated as a pipeline artifact step.
    _DIRECT_GF_WRITE_DEPRECATED = True  # Step 8 deprecation guard
    if not skip_gf_update and not _DIRECT_GF_WRITE_DEPRECATED:
        # UNREACHABLE — guarded by _DIRECT_GF_WRITE_DEPRECATED
        if not INPUT_GF.exists():
            logger.warning("GF_V0_CLEAN.xlsx not found at %s — skipping GF update stages", INPUT_GF)
        else:
            logger.info("\n── STEP G: Stage 1 GF enrichment (date + status) ──")
            gf_stage1_result = write_gf_enriched(
                gf_source_path = INPUT_GF,
                enrichments    = enrichments,
                output_path    = OUTPUT_GF_STAGE1,
                stage          = 1,
            )
            logger.info("Stage 1 complete: %d rows enriched", gf_stage1_result["total_enriched"])

            logger.info("\n── STEP H: Stage 2 GF enrichment (+ observations) ──")
            gf_stage2_result = write_gf_enriched(
                gf_source_path = INPUT_GF,
                enrichments    = enrichments,
                output_path    = OUTPUT_GF_STAGE2,
                stage          = 2,
            )
            logger.info("Stage 2 complete: %d rows enriched", gf_stage2_result["total_enriched"])
    elif not skip_gf_update:
        logger.warning(
            "Direct GF write stages (Stage 1 / Stage 2) are DEPRECATED (Step 8). "
            "Skipping write_gf_enriched(). Run the main pipeline to produce GF_V0_CLEAN.xlsx."
        )

    # ── Summary ───────────────────────────────────────────────────────────
    logger.info("\n" + "=" * 72)
    logger.info("CONSULTANT INTEGRATION COMPLETE")
    logger.info("=" * 72)
    logger.info("Consultant rows:   %d", total_consultant_rows)
    logger.info("GED docs:          %d", len(docs_df))
    logger.info("Matched:           %d", len(match_results["matched"]))
    logger.info("Ambiguous:         %d", len(match_results["ambiguous"]))
    logger.info("Unmatched:         %d", len(match_results["unmatched"]))
    logger.info("Match report:      %s", OUTPUT_MATCH_REPORT)
    if gf_stage1_result:
        logger.info("Stage 1 output:    %s  (%d rows enriched)",
                    OUTPUT_GF_STAGE1, gf_stage1_result["total_enriched"])
    if gf_stage2_result:
        logger.info("Stage 2 output:    %s  (%d rows enriched)",
                    OUTPUT_GF_STAGE2, gf_stage2_result["total_enriched"])
    logger.info("=" * 72)

    return {
        "rows_by_source":   rows_by_source,
        "docs_df":          docs_df,
        "match_results":    match_results,
        "enrichments":      enrichments,
        "gf_stage1":        gf_stage1_result,
        "gf_stage2":        gf_stage2_result,
        "output_paths": {
            "consultant_workbook": OUTPUT_CONSULTANT_WB,
            "match_report":        OUTPUT_MATCH_REPORT,
            "gf_stage1":           OUTPUT_GF_STAGE1,
            "gf_stage2":           OUTPUT_GF_STAGE2,
        }
    }


# ---------------------------------------------------------------------------
# Script entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    sys.stderr.write(
        "Direct standalone execution of consultant_integration.py is blocked for production use.\n"
        "Run consultant integration through the main pipeline / orchestrated run path so "
        "artifacts are tracked in run memory.\n"
    )
    sys.exit(2)
