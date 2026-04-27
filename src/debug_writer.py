"""
debug_writer.py
---------------
Writes all debug artifacts to output/debug/.

Artifacts:
  routing_summary.xlsx    — per-sheet routing breakdown (written by routing.py)
  gf_sheet_schema.xlsx    — dynamic column parsing results (written by main.py)
  coarse_groups.xlsx      — version engine coarse group stats
  family_clusters.xlsx    — family cluster breakdown
  lifecycle_resolution.xlsx — all lifecycle decisions
  discrepancy_sample.xlsx — sample rows of each discrepancy flag type
"""

from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _header_row(ws, headers: list, fg_color: str = "1F4E79"):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fg_color)
        cell.alignment = Alignment(horizontal="center")


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────
# 1. COARSE GROUPS
# ─────────────────────────────────────────────────────────────

def write_coarse_groups(output_path: str, versioned_df: pd.DataFrame):
    """
    Write per-coarse-group stats: how many versions, families, dernier indices.
    """
    if "coarse_group_key" not in versioned_df.columns:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coarse Groups"

    headers = ["Coarse Group Key", "Total Versions", "Distinct Families",
               "Dernier Indices", "Has Anomaly", "Anomaly Flags"]
    _header_row(ws, headers, "2E75B6")
    _set_col_widths(ws, [55, 14, 15, 14, 12, 40])

    for r_idx, (key, grp) in enumerate(
        versioned_df.groupby("coarse_group_key", dropna=False), 2
    ):
        families = grp["family_cluster_id"].nunique() if "family_cluster_id" in grp.columns else 0
        dernier = grp["is_dernier_indice"].sum() if "is_dernier_indice" in grp.columns else 0
        all_flags = []
        for flags in grp.get("anomaly_flags", []):
            if isinstance(flags, list):
                all_flags.extend(flags)
        unique_flags = list(set(all_flags))

        ws.cell(row=r_idx, column=1, value=str(key))
        ws.cell(row=r_idx, column=2, value=len(grp))
        ws.cell(row=r_idx, column=3, value=families)
        ws.cell(row=r_idx, column=4, value=int(dernier))
        ws.cell(row=r_idx, column=5, value="Yes" if unique_flags else "")
        ws.cell(row=r_idx, column=6, value=", ".join(unique_flags))

        if unique_flags:
            for col in range(1, 7):
                ws.cell(row=r_idx, column=col).fill = PatternFill(
                    "solid", fgColor="FFD966"
                )

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────
# 2. FAMILY CLUSTERS
# ─────────────────────────────────────────────────────────────

def write_family_clusters(output_path: str, versioned_df: pd.DataFrame):
    """
    Write per-family-cluster breakdown: representative libellé, versions, routing.
    """
    if "family_cluster_id" not in versioned_df.columns:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Family Clusters"

    headers = [
        "Family Cluster ID (short)", "Coarse Group Key", "Lot",
        "Emetteur", "Versions", "Is Dernier", "Routing Status",
        "GF Sheet", "Representative Libellé", "Indice Range",
    ]
    _header_row(ws, headers, "375623")
    _set_col_widths(ws, [24, 45, 8, 14, 8, 10, 18, 40, 65, 12])

    for r_idx, (fid, grp) in enumerate(
        versioned_df.groupby("family_cluster_id", dropna=False), 2
    ):
        fid_short = str(fid)[:8] if fid else "?"
        coarse_key = grp["coarse_group_key"].iloc[0] if "coarse_group_key" in grp else ""
        lot = grp["lot_normalized"].iloc[0] if "lot_normalized" in grp else ""
        emetteur = grp["emetteur"].iloc[0] if "emetteur" in grp else ""
        is_dernier = grp["is_dernier_indice"].sum() if "is_dernier_indice" in grp else 0
        routing = grp["routing_status"].iloc[0] if "routing_status" in grp else ""
        sheet = grp["gf_sheet_name"].iloc[0] if "gf_sheet_name" in grp else ""
        rep_libelle = str(grp["libelle_du_document"].iloc[0])[:80] if "libelle_du_document" in grp else ""

        indices = sorted(grp["indice"].dropna().unique().tolist()) if "indice" in grp else []
        indice_range = f"{indices[0]}→{indices[-1]}" if len(indices) >= 2 else (indices[0] if indices else "")

        ws.cell(row=r_idx, column=1, value=fid_short)
        ws.cell(row=r_idx, column=2, value=str(coarse_key))
        ws.cell(row=r_idx, column=3, value=str(lot))
        ws.cell(row=r_idx, column=4, value=str(emetteur))
        ws.cell(row=r_idx, column=5, value=len(grp))
        ws.cell(row=r_idx, column=6, value=int(is_dernier))
        ws.cell(row=r_idx, column=7, value=str(routing))
        ws.cell(row=r_idx, column=8, value=str(sheet))
        ws.cell(row=r_idx, column=9, value=rep_libelle)
        ws.cell(row=r_idx, column=10, value=indice_range)

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────
# 3. LIFECYCLE RESOLUTION
# ─────────────────────────────────────────────────────────────

def write_lifecycle_resolution(output_path: str, versioned_df: pd.DataFrame):
    """
    Write all lifecycle decisions — one row per dernier_indice document,
    with its anomaly_flags and resolution_status.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lifecycle Resolution"

    headers = [
        "Doc ID", "EMETTEUR", "LOT", "NUMERO", "INDICE",
        "Lifecycle ID (short)", "Chain Position", "Is Dernier",
        "Is Excluded Lifecycle", "Resolution Status",
        "Anomaly Flags", "Confidence", "GF Sheet", "Libellé",
    ]
    _header_row(ws, headers, "7F0000")
    _set_col_widths(ws, [36, 14, 8, 10, 6, 16, 14, 10, 18, 18, 40, 10, 38, 60])

    STATUS_COLORS = {
        "OK":               "FFFFFF",
        "AUTO_RESOLVED":    "C6EFCE",
        "REVIEW_REQUIRED":  "FFC7CE",
        "IGNORED":          "FFD966",
    }

    for r_idx, (_, row) in enumerate(versioned_df.iterrows(), 2):
        flags = row.get("anomaly_flags", [])
        flags_str = ", ".join(flags) if isinstance(flags, list) else str(flags)
        resolution = row.get("resolution_status", "OK")
        lc_id = str(row.get("lifecycle_id", ""))[:8]

        ws.cell(row=r_idx, column=1, value=str(row.get("doc_id", "")))
        ws.cell(row=r_idx, column=2, value=str(row.get("emetteur", "")))
        ws.cell(row=r_idx, column=3, value=str(row.get("lot_normalized", "")))
        ws.cell(row=r_idx, column=4, value=str(row.get("numero_normalized", "")))
        ws.cell(row=r_idx, column=5, value=str(row.get("indice", "")))
        ws.cell(row=r_idx, column=6, value=lc_id)
        ws.cell(row=r_idx, column=7, value=row.get("chain_position", ""))
        ws.cell(row=r_idx, column=8, value=row.get("is_dernier_indice", ""))
        ws.cell(row=r_idx, column=9, value=row.get("is_excluded_lifecycle", ""))
        ws.cell(row=r_idx, column=10, value=resolution)
        ws.cell(row=r_idx, column=11, value=flags_str)
        ws.cell(row=r_idx, column=12, value=row.get("version_confidence", ""))
        ws.cell(row=r_idx, column=13, value=str(row.get("gf_sheet_name", "")))
        ws.cell(row=r_idx, column=14, value=str(row.get("libelle_du_document", ""))[:80])

        fill_color = STATUS_COLORS.get(resolution, "FFFFFF")
        if fill_color != "FFFFFF":
            for col in range(1, 15):
                ws.cell(row=r_idx, column=col).fill = PatternFill(
                    "solid", fgColor=fill_color
                )

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────
# 4. DISCREPANCY SAMPLE
# ─────────────────────────────────────────────────────────────

def write_discrepancy_sample(output_path: str, discrepancies: list, n_per_type: int = 20):
    """
    Write a sample (up to n_per_type rows) of each discrepancy flag type.
    Makes it easy to spot-check without loading 8,000 rows.
    """
    if not discrepancies:
        return

    import collections
    disc_df = pd.DataFrame(discrepancies)
    flag_types = disc_df["flag_type"].unique()

    FLAG_COLORS = {
        "MISSING_IN_GF":   "FFC7CE",
        "MISSING_IN_GED":  "FFEB9C",
        "INDICE_MISMATCH": "FCE4D6",
        "TITRE_MISMATCH":  "D9D9D9",
        "DATE_MISMATCH":   "D9EAD3",
        "SHEET_MISMATCH":  "CFE2F3",
    }

    wb = openpyxl.Workbook()

    for flag_type in sorted(flag_types):
        subset = disc_df[disc_df["flag_type"] == flag_type].head(n_per_type)
        sheet_name = flag_type[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = list(subset.columns)
        _header_row(ws, headers, "1F4E79")

        fill = PatternFill("solid", fgColor=FLAG_COLORS.get(flag_type, "FFFFFF"))
        for r_idx, (_, row) in enumerate(subset.iterrows(), 2):
            for c_idx, col in enumerate(headers, 1):
                ws.cell(row=r_idx, column=c_idx, value=str(row.get(col, ""))[:80])
                ws.cell(row=r_idx, column=c_idx).fill = fill

        for c_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 25

    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────
# WRITE ALL DEBUG ARTIFACTS
# ─────────────────────────────────────────────────────────────

def write_all_debug(
    debug_dir: str,
    versioned_df: pd.DataFrame,
    discrepancies: list,
):
    """
    Write all debug artifacts except routing_summary and gf_sheet_schema
    (those are written by routing.py and main.py respectively).
    """
    ddir = Path(debug_dir)
    ddir.mkdir(parents=True, exist_ok=True)

    write_coarse_groups(str(ddir / "coarse_groups.xlsx"), versioned_df)
    write_family_clusters(str(ddir / "family_clusters.xlsx"), versioned_df)
    write_lifecycle_resolution(str(ddir / "lifecycle_resolution.xlsx"), versioned_df)
    write_discrepancy_sample(str(ddir / "discrepancy_sample.xlsx"), discrepancies)
