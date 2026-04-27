"""
writer.py
---------
Generates the output Excel files:
  1. GF_V0_CLEAN.xlsx  — reconstructed GF per contractor sheet
  2. DISCREPANCY_REPORT.xlsx  — GED vs existing GF comparison
  3. ANOMALY_REPORT.xlsx  — per-contractor anomaly flags

Uses openpyxl for full formatting control.
"""

import os
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd


# ─────────────────────────────────────────────────────────────
# SAFE SAVE HELPER
# ─────────────────────────────────────────────────────────────

def _safe_save(wb: openpyxl.Workbook, target_path: str) -> str:
    """
    Save workbook to target_path.
    If target_path is locked (PermissionError), write to a temp file
    alongside it and print a warning. Returns the actual path saved to.
    """
    try:
        wb.save(target_path)
        return target_path
    except (PermissionError, FileNotFoundError):
        # Target is locked or unwritable (e.g. open in Excel on Windows mount).
        # Fall back to a _NEW suffix file.
        p = Path(target_path)
        alt_path = str(p.parent / f"{p.stem}_NEW{p.suffix}")
        wb.save(alt_path)
        print(f"  ⚠️  WARNING: '{p.name}' is locked / unwritable.")
        print(f"       Wrote to '{alt_path}' instead.")
        return alt_path


# ─────────────────────────────────────────────────────────────
# COLOURS & STYLES
# ─────────────────────────────────────────────────────────────

COLOUR_HEADER_BG = "1F4E79"      # Dark blue
COLOUR_APPROVER_BG = "2E75B6"    # Mid blue
COLOUR_SUBHEADER_BG = "9DC3E6"   # Light blue
COLOUR_VSO = "C6EFCE"            # Green
COLOUR_VAO = "FFEB9C"            # Yellow
COLOUR_REF = "FFC7CE"            # Red
COLOUR_HM = "D9D9D9"             # Grey
COLOUR_SUS = "FCE4D6"            # Orange
COLOUR_ANOMALY = "FFD966"        # Gold
COLOUR_EXCLUDED = "BFBFBF"       # Grey for old versions

STATUS_COLOURS = {
    "VSO": COLOUR_VSO,
    "VAO": COLOUR_VAO,
    "REF": COLOUR_REF,
    "SAS REF": COLOUR_REF,   # SAS-stage refusal — same red as REF
    "HM": COLOUR_HM,
    "SUS": COLOUR_SUS,
}

WHITE_FONT = Font(color="FFFFFF", bold=True, size=9)
BOLD_FONT = Font(bold=True, size=9)
NORMAL_FONT = Font(size=9)
SMALL_FONT = Font(size=8)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_cell(ws, row: int, col: int, value, font=None, fill=None,
                alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


# ─────────────────────────────────────────────────────────────
# OBSERVATION COLUMN HELPERS
# ─────────────────────────────────────────────────────────────

# Canonical approver sort order for the Observation column.
# MOEX GEMO first, then primary consultants, then secondary.
# Matched by substring of the upper-cased canonical approver name.
_OBSERVATION_ORDER_KEYWORDS = [
    # MOEX / GEMO — always first
    "MOEX", "GEMO",
    # Primary consultants (in specification order)
    "ARCHI MOX",
    "TERRELL",        # BET STR-TERRELL
    "EGIS CVC",
    "EGIS ELEC",
    "EGIS PLB",
    "EGIS GTB",
    "EGIS",           # catch-all for other EGIS variants
    "BET SPK",
    "BET ASC",
    "ELIOTH",         # BET ELIOTH / BET EV
    "BET EV",
    "FACADES",
    # Secondary consultants
    "AVLS",
    "AMO HQE",
    "SOCOTEC",
    "GEOLIA",
    "MUGO",
]


def _approver_obs_sort_key(name: str) -> int:
    """Return the sort priority of an approver name for the Observation column."""
    upper = name.upper()
    for i, kw in enumerate(_OBSERVATION_ORDER_KEYWORDS):
        if kw in upper:
            return i
    return len(_OBSERVATION_ORDER_KEYWORDS)  # unknowns go last


def _build_observation_text(approver_statuses: dict) -> str:
    """
    Build the Observation column text from approver_statuses.
    Only includes approvers that have a non-empty status or comment.
    Returns a newline-joined string in canonical priority order.
    Format per entry: "{Approver}: {STATUS} — {comment}"
    """
    parts = []
    ordered = sorted(approver_statuses.keys(), key=_approver_obs_sort_key)
    for approver in ordered:
        info = approver_statuses.get(approver, {})
        # Accept 'status' (writer dict key) or 'status_clean' (engine dict key)
        raw_status = info.get("status")
        if raw_status in (None, ""):
            raw_status = info.get("status_clean")

        raw_comment = info.get("comment", "")

        def _clean_obs_value(value) -> str:
            if value is None:
                return ""
            text = str(value).strip()
            return "" if text.lower() == "nan" else text

        status = _clean_obs_value(raw_status)
        comment = _clean_obs_value(raw_comment)
        if not status and not comment:
            continue
        line = approver
        if status:
            line += f": {status}"
        if comment:
            line += f" — {comment}"
        parts.append(line)
    return "\n".join(parts)


# ─────────────────────────────────────────────────────────────
# GF SHEET BUILDER
# ─────────────────────────────────────────────────────────────

class GFSheetWriter:
    """Writes one GF contractor sheet."""

    # Base column definitions
    BASE_COLS = [
        ("DOCUMENT", 28),
        ("TITRE", 45),
        ("Date diffusion", 14),
        ("LOT", 8),
        ("TYPE DOC", 10),
        ("N° Doc", 10),
        ("IND", 5),
        ("NIV", 6),
        ("Type", 6),
        ("ANCIEN", 7),
        ("N°BDX", 8),
        ("Date réception", 14),
        ("non reçu papier", 8),
        ("DATE\nCONTRACT.", 14),
        # PATCH 2: Real MOEX visa date column inserted between DATE CONTRACT. and VISA GLOBAL
        ("Date réel\nde visa", 14),
        ("VISA\nGLOBAL", 10),
    ]

    APPROVER_SUB_COLS = [("DATE", 12), ("N°", 5), ("STATUT", 8)]
    # PATCH 4: OBSERVATIONS first so it lands at AH; DISCREPANCES GED follows
    EXTRA_COLS = [("OBSERVATIONS", 30), ("DISCREPANCES GED", 30)]

    LEGEND_ROW = 1
    LEGEND2_ROW = 2
    STATUS_LEGEND_ROW_1 = 3
    STATUS_LEGEND_ROW_2 = 4
    STATUS_LEGEND_ROW_3 = 5
    BLANK_ROW = 6
    HEADER_ROW = 7
    APPROVER_ROW = 8
    SUBHEADER_ROW = 9
    DATA_START_ROW = 10

    def __init__(self, wb: openpyxl.Workbook, sheet_name: str,
                 approver_names: List[str]):
        """
        approver_names: list of GF approver names for this sheet
                        (e.g. ['MOEX GEMO', 'ARCHI MOX', 'BET EGIS', ...])
        """
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        self.ws = wb.create_sheet(sheet_name)
        self.sheet_name = sheet_name
        self.approver_names = approver_names

        # Column layout
        self.total_base_cols = len(self.BASE_COLS)
        self.approver_col_start = self.total_base_cols + 1  # 1-indexed
        self.approver_col_span = len(self.APPROVER_SUB_COLS)
        self.extra_col_start = (
            self.approver_col_start +
            len(approver_names) * self.approver_col_span
        )

        self._write_legend()
        self._write_headers()
        self._set_col_widths()

        self.current_data_row = self.DATA_START_ROW

    def _write_legend(self):
        ws = self.ws
        # Row 1: Title
        ws.merge_cells(f"C{self.LEGEND_ROW}:M{self.LEGEND_ROW}")
        _apply_cell(ws, self.LEGEND_ROW, 3, "S U I V I    D E S    V I S A",
                    font=Font(bold=True, size=14),
                    alignment=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[self.LEGEND_ROW].height = 20

        # Rows 3-5: Status legend
        legend_items = [
            (3, "VSO", "Visé sans observations", COLOUR_VSO,
             "REF", "Visé avec observation bloquant", COLOUR_REF),
            (4, "VAO", "Visé avec observations", COLOUR_VAO,
             "", "Retard sur visa", None),
            (5, "Annulé", "Voir visa indice supérieur", None,
             "HM", "Hors mission", COLOUR_HM),
        ]
        for row, s1, l1, c1, s2, l2, c2 in legend_items:
            _apply_cell(ws, row, 3, s1, font=BOLD_FONT,
                        fill=_fill(c1) if c1 else None)
            _apply_cell(ws, row, 5, l1, font=NORMAL_FONT)
            if s2:
                _apply_cell(ws, row, 10, s2, font=BOLD_FONT,
                            fill=_fill(c2) if c2 else None)
            if l2:
                _apply_cell(ws, row, 11, l2, font=NORMAL_FONT)

    def _write_headers(self):
        ws = self.ws
        border = _thin_border()
        header_fill = _fill(COLOUR_HEADER_BG)
        approver_fill = _fill(COLOUR_APPROVER_BG)
        sub_fill = _fill(COLOUR_SUBHEADER_BG)

        # Row 7: Base column headers
        for i, (col_name, _) in enumerate(self.BASE_COLS):
            col = i + 1
            _apply_cell(
                ws, self.HEADER_ROW, col, col_name,
                font=WHITE_FONT,
                fill=header_fill,
                alignment=Alignment(horizontal="center", vertical="center",
                                    wrap_text=True),
                border=border,
            )
            ws.merge_cells(
                start_row=self.HEADER_ROW, start_column=col,
                end_row=self.SUBHEADER_ROW, end_column=col,
            )

        # Row 7: "APPROBATEURS" spanning all approver cols
        if self.approver_names:
            app_start = self.approver_col_start
            app_end = app_start + len(self.approver_names) * self.approver_col_span - 1
            _apply_cell(
                ws, self.HEADER_ROW, app_start, "APPROBATEURS",
                font=WHITE_FONT, fill=header_fill,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=border,
            )
            ws.merge_cells(
                start_row=self.HEADER_ROW, start_column=app_start,
                end_row=self.HEADER_ROW, end_column=app_end,
            )

        # Extra cols header
        for i, (col_name, _) in enumerate(self.EXTRA_COLS):
            col = self.extra_col_start + i
            _apply_cell(ws, self.HEADER_ROW, col, col_name,
                        font=WHITE_FONT, fill=header_fill,
                        alignment=Alignment(horizontal="center"),
                        border=border)
            ws.merge_cells(
                start_row=self.HEADER_ROW, start_column=col,
                end_row=self.SUBHEADER_ROW, end_column=col,
            )

        # Row 8: Approver names
        for i, approver in enumerate(self.approver_names):
            col = self.approver_col_start + i * self.approver_col_span
            _apply_cell(
                ws, self.APPROVER_ROW, col, approver,
                font=WHITE_FONT, fill=approver_fill,
                alignment=Alignment(horizontal="center", wrap_text=True),
                border=border,
            )
            ws.merge_cells(
                start_row=self.APPROVER_ROW, start_column=col,
                end_row=self.APPROVER_ROW, end_column=col + self.approver_col_span - 1,
            )

        # Row 9: Sub-headers (DATE, N°, STATUT)
        for i in range(len(self.approver_names)):
            base_col = self.approver_col_start + i * self.approver_col_span
            for j, (sub_name, _) in enumerate(self.APPROVER_SUB_COLS):
                _apply_cell(
                    ws, self.SUBHEADER_ROW, base_col + j, sub_name,
                    font=BOLD_FONT, fill=sub_fill,
                    alignment=Alignment(horizontal="center"),
                    border=border,
                )

        ws.row_dimensions[self.HEADER_ROW].height = 30
        ws.row_dimensions[self.APPROVER_ROW].height = 25
        ws.row_dimensions[self.SUBHEADER_ROW].height = 16

    def _set_col_widths(self):
        ws = self.ws
        # Base cols
        for i, (_, width) in enumerate(self.BASE_COLS):
            ws.column_dimensions[get_column_letter(i + 1)].width = width
        # Approver sub-cols
        for i in range(len(self.approver_names)):
            base_col = self.approver_col_start + i * self.approver_col_span
            for j, (_, width) in enumerate(self.APPROVER_SUB_COLS):
                ws.column_dimensions[get_column_letter(base_col + j)].width = width
        # Extra cols
        for i, (_, width) in enumerate(self.EXTRA_COLS):
            ws.column_dimensions[get_column_letter(self.extra_col_start + i)].width = width

    def write_document_row(
        self,
        doc: dict,
        approver_statuses: dict,  # {approver_name: {date, status, comment}}
        visa_global: Optional[str] = None,
        ancien_value: str = "",   # "DI" | "1" | ""
        discrepancy_notes: str = "",
        anomaly_flags: list = None,
        moex_date=None,           # PATCH 2: real MOEX visa date
        sas_date=None,            # PATCH 2 (Round 2): SAS response date for DATE CONTRACT. Case B
    ):
        """
        Write one data row.
        doc: dict with keys matching GF base columns
        approver_statuses: {gf_approver_name: {'date': ..., 'status': ...}}
        """
        ws = self.ws
        row = self.current_data_row
        border = _thin_border()
        anomaly_flags = anomaly_flags or []

        # Row background
        row_fill = None
        if anomaly_flags:
            row_fill = _fill(COLOUR_ANOMALY)

        def _write(col, value, fmt=None, fill=None, font=None):
            _apply_cell(
                ws, row, col, value,
                font=font or NORMAL_FONT,
                fill=fill or row_fill,
                alignment=Alignment(vertical="center", wrap_text=True),
                border=border,
                number_format=fmt,
            )

        # Col 1: DOCUMENT code
        _write(1, doc.get("document_code", ""))
        # Col 2: TITRE
        _write(2, doc.get("titre", ""))
        # Col 3: Date diffusion
        date_val = doc.get("date_diffusion")
        _write(3, date_val, fmt="DD/MM/YYYY")
        # Col 4: LOT
        _write(4, doc.get("lot", ""))
        # Col 5: TYPE DOC
        _write(5, doc.get("type_doc", ""))
        # Col 6: N° Doc
        _write(6, doc.get("numero", ""))
        # Col 7: IND
        _write(7, doc.get("indice", ""))
        # Col 8: NIV
        _write(8, doc.get("niveau", ""))
        # Col 9: Type
        _write(9, doc.get("type", "pdf"))
        # Col 10: ANCIEN  ("DI" = latest version, "1" = older version, "" = n/a)
        _write(10, ancien_value)
        # Col 11: N°BDX (empty)
        _write(11, "")
        # Col 12: Date réception (formula referencing col 3 = Date diffusion)
        _write(12, f"=C{row}")
        # Col 13: non reçu papier
        _write(13, "X")
        # Col 14: DATE CONTRACTUELLE — Round 2 Patch 2: SAS-phase logic
        # Case B (SAS passed with VAO-SAS or VSO-SAS): base = SAS response date
        # Case A (SAS pending / not called):            base = Date réception (= date_diffusion)
        # DATE CONTRACT. = base + 15 calendar days
        _date_diffusion = doc.get("date_diffusion")
        _date_contr = None
        _base_date = sas_date if sas_date is not None else _date_diffusion
        if _base_date is not None:
            try:
                import datetime as _datetime
                import pandas as _pd_local
                _base_date_dt = _pd_local.to_datetime(_base_date)
                _date_contr = _base_date_dt + _datetime.timedelta(days=15)
            except Exception:
                _date_contr = None
        _write(14, _date_contr, fmt="DD/MM/YYYY")
        # Col 15: Date réel de visa — PATCH 2: real MOEX visa date
        _write(15, moex_date, fmt="DD/MM/YYYY")
        # Col 16: VISA GLOBAL — PATCH 3: equals MOEX status (SAS REF or REF or VSO etc.)
        if visa_global:
            status_fill = _fill(STATUS_COLOURS.get(visa_global, "FFFFFF"))
            _write(16, visa_global, fill=status_fill, font=BOLD_FONT)
        else:
            _write(16, "")

        # Approver columns
        for i, approver in enumerate(self.approver_names):
            base_col = self.approver_col_start + i * self.approver_col_span
            status_info = approver_statuses.get(approver, {})
            date_val = status_info.get("date")
            status_val = status_info.get("status")

            # DATE
            _write(base_col, date_val, fmt="DD/MM/YYYY")

            # N° (BDX number — empty for now)
            _write(base_col + 1, "")

            # STATUT
            if status_val:
                status_fill = _fill(STATUS_COLOURS.get(status_val, "FFFFFF"))
                _write(base_col + 2, status_val, fill=status_fill)
            else:
                _write(base_col + 2, "")

        # PATCH 4: OBSERVATIONS is now the first extra col (AH position)
        # DISCREPANCES GED follows it.
        _write(self.extra_col_start, _build_observation_text(approver_statuses))
        _write(self.extra_col_start + 1, discrepancy_notes)

        ws.row_dimensions[row].height = 14
        self.current_data_row += 1


# ─────────────────────────────────────────────────────────────
# MAIN WRITER ORCHESTRATOR
# ─────────────────────────────────────────────────────────────

class GFWriter:
    """Orchestrates writing of all GF sheets."""

    # Per-sheet approver configuration
    # These are the canonical GF approver names used per sheet type
    # Key: pattern in sheet name → list of approvers
    DEFAULT_APPROVERS_BY_LOT = {
        "41": ["MOEX GEMO", "ARCHI MOX", "BET EGIS CVC", "AMO HQE", "BC SOCOTEC"],
        "42": ["MOEX GEMO", "ARCHI MOX", "BET EGIS PLB", "AMO HQE", "BC SOCOTEC"],
        "35": ["MOEX GEMO", "ARCHI MOX", "BET EGIS GTB", "AMO HQE", "BC SOCOTEC"],
        "43": ["MOEX GEMO", "ARCHI MOX", "BET SPK", "AMO HQE", "BC SOCOTEC"],
        "51": ["MOEX GEMO", "ARCHI MOX", "BET ASC", "AMO HQE", "BC SOCOTEC"],
        "default": ["MOEX GEMO", "ARCHI MOX", "BET EGIS", "AMO HQE", "BC SOCOTEC"],
    }

    def __init__(self, output_path: str, existing_gf_path: Optional[str] = None):
        self.output_path = output_path
        self.existing_gf_path = existing_gf_path
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self._sheet_writers: Dict[str, GFSheetWriter] = {}

    def _get_approvers_for_sheet(self, sheet_name: str,
                                  existing_structure: Optional[dict] = None) -> List[str]:
        """Get approver list for a sheet, preferring existing structure."""
        if existing_structure and existing_structure.get("approvers"):
            return [a["name"] for a in existing_structure["approvers"]]

        # Fallback: find by lot number
        nums = re.findall(r'\d+', sheet_name)
        for num in nums:
            if num in self.DEFAULT_APPROVERS_BY_LOT:
                return self.DEFAULT_APPROVERS_BY_LOT[num]

        return self.DEFAULT_APPROVERS_BY_LOT["default"]

    def get_or_create_sheet_writer(
        self,
        sheet_name: str,
        existing_structure: Optional[dict] = None,
    ) -> GFSheetWriter:
        if sheet_name not in self._sheet_writers:
            approvers = self._get_approvers_for_sheet(sheet_name, existing_structure)
            writer = GFSheetWriter(self.wb, sheet_name, approvers)
            self._sheet_writers[sheet_name] = writer
        return self._sheet_writers[sheet_name]

    def write_all(
        self,
        docs_df: pd.DataFrame,
        responses_df: pd.DataFrame,
        workflow_engine,
        sheet_structures: Dict[str, dict],
        gf_to_ged_map: Optional[Dict] = None,
        ancien_df: Optional[pd.DataFrame] = None,
        sas_lookup: Optional[dict] = None,  # Round 2 Patch 2: SAS dates for DATE CONTRACT.
    ):
        """
        Main entry point.
        docs_df:    dernier-indice documents (ANCIEN = "DI")
        ancien_df:  older-indice documents belonging to the same families (ANCIEN = "1")
        responses_df: normalized responses
        gf_to_ged_map: {gf_display_name: [ged_canonical_names]}
        sas_lookup:  {doc_id: {sas_result, sas_date, ...}} for DATE CONTRACTUELLE Case B
        """
        _sas_lookup = sas_lookup or {}
        # Group by sheet
        for sheet_name, group in docs_df.groupby("gf_sheet_name", dropna=True):
            struct = sheet_structures.get(sheet_name)
            writer = self.get_or_create_sheet_writer(sheet_name, struct)

            # ── Write DI (dernier-indice) rows ──────────────────────────
            for _, doc_row in group.iterrows():
                doc_id = doc_row["doc_id"]

                # Build approver statuses — use GF→GED mapping
                # Include comment for Observation column
                approver_statuses = {}
                for approver in writer.approver_names:
                    status_info = workflow_engine.get_approver_status(
                        doc_id, approver, gf_name_to_ged=gf_to_ged_map
                    )
                    approver_statuses[approver] = {
                        "date": status_info.get("date_answered"),
                        "status": status_info.get("status_clean"),
                        "comment": status_info.get("comment", ""),
                    }

                # Derive VISA GLOBAL and Date réel de visa from the SAME MOEX entry
                # so they are always semantically consistent:
                #   SAS-only (VAO-SAS/VSO-SAS) → both empty
                #   SAS REF                    → ("SAS REF", sas_refusal_date)
                #   final MOEX visa            → (status, final_moex_date)
                visa_global, moex_date = workflow_engine.compute_visa_global_with_date(doc_id)

                # Round 2 Patch 2: SAS date for DATE CONTRACTUELLE Case B
                # Case B: SAS passed (VSO-SAS or VAO-SAS) → DATE CONTRACT = SAS date + 15
                # Case A: SAS pending / not called → DATE CONTRACT = date_diffusion + 15
                _sas_entry = _sas_lookup.get(doc_id, {})
                _sas_result = _sas_entry.get("sas_result", "") or ""
                _sas_passed = _sas_result in ("VSO-SAS", "VAO-SAS", "VSO", "VAO")
                sas_date_for_contract = _sas_entry.get("sas_date") if _sas_passed else None

                # Build doc dict for writer
                doc_dict = {
                    "document_code": _build_document_code(doc_row),
                    "titre": _truncate(str(doc_row.get("libelle_du_document", "")), 80),
                    "date_diffusion": doc_row.get("created_at"),
                    "lot": doc_row.get("lot", ""),
                    "type_doc": doc_row.get("type_de_doc", ""),
                    "numero": doc_row.get("numero_normalized", ""),
                    "indice": doc_row.get("indice", ""),
                    "niveau": doc_row.get("niveau", ""),
                    "type": "pdf",
                }

                anomaly_flags = doc_row.get("anomaly_flags", [])
                if isinstance(anomaly_flags, str):
                    anomaly_flags = []

                writer.write_document_row(
                    doc=doc_dict,
                    approver_statuses=approver_statuses,
                    visa_global=visa_global,
                    ancien_value="DI",
                    anomaly_flags=anomaly_flags,
                    moex_date=moex_date,
                    sas_date=sas_date_for_contract,
                )

            # ── Write VALID_HISTORICAL (older-indice) rows for this sheet ─────
            # These are genuine older versions of active document families.
            # They receive the same full approver-status treatment as DI rows
            # (each has its own doc_id with its own GED responses).
            # EXCEPTION rows (excluded lifecycles, BENTIN legacy docs marked
            # is_exception=True, SAS-old docs) are filtered OUT before this
            # loop in main.py and must NOT appear in ancien_df.
            if ancien_df is not None and len(ancien_df) > 0 and "gf_sheet_name" in ancien_df.columns:
                sheet_ancien = ancien_df[ancien_df["gf_sheet_name"] == sheet_name]
                for _, doc_row in sheet_ancien.iterrows():
                    doc_id = doc_row["doc_id"]

                    # ── Full approver status — same as DI rows ──────────────
                    approver_statuses = {}
                    for approver in writer.approver_names:
                        status_info = workflow_engine.get_approver_status(
                            doc_id, approver, gf_name_to_ged=gf_to_ged_map
                        )
                        approver_statuses[approver] = {
                            "date": status_info.get("date_answered"),
                            "status": status_info.get("status_clean"),
                            "comment": status_info.get("comment", ""),
                        }

                    # Same unified call for ANCIEN rows
                    visa_global, moex_date = workflow_engine.compute_visa_global_with_date(doc_id)

                    # Round 2 Patch 2: SAS date for ANCIEN rows
                    _sas_entry_a = _sas_lookup.get(doc_id, {})
                    _sas_result_a = _sas_entry_a.get("sas_result", "") or ""
                    _sas_passed_a = _sas_result_a in ("VSO-SAS", "VAO-SAS", "VSO", "VAO")
                    sas_date_for_contract_a = _sas_entry_a.get("sas_date") if _sas_passed_a else None

                    doc_dict = {
                        "document_code": _build_document_code(doc_row),
                        "titre": _truncate(str(doc_row.get("libelle_du_document", "")), 80),
                        "date_diffusion": doc_row.get("created_at"),
                        "lot": doc_row.get("lot", ""),
                        "type_doc": doc_row.get("type_de_doc", ""),
                        "numero": doc_row.get("numero_normalized", ""),
                        "indice": doc_row.get("indice", ""),
                        "niveau": doc_row.get("niveau", ""),
                        "type": "pdf",
                    }
                    anomaly_flags = doc_row.get("anomaly_flags", [])
                    if isinstance(anomaly_flags, str):
                        anomaly_flags = []
                    writer.write_document_row(
                        doc=doc_dict,
                        approver_statuses=approver_statuses,
                        visa_global=visa_global,
                        ancien_value="1",
                        anomaly_flags=anomaly_flags,
                        moex_date=moex_date,
                        sas_date=sas_date_for_contract_a,
                    )

    def save(self):
        """Save the workbook (uses safe save to handle locked files)."""
        return _safe_save(self.wb, self.output_path)


def _is_moex_gf_name(name: str) -> bool:
    """True if this GF display name refers to the MOEX/GEMO approver."""
    upper = name.upper()
    return "MOEX" in upper or "GEMO" in upper


def _get_moex_date(doc_id: str, approver_names: list,
                   workflow_engine, gf_to_ged_map):
    """
    PATCH 2 helper: extract the MOEX visa date for this document.
    Scans approver_names for the MOEX approver, then asks the engine.
    Returns a datetime object or None.
    """
    for approver in approver_names:
        if _is_moex_gf_name(approver):
            info = workflow_engine.get_approver_status(
                doc_id, approver, gf_name_to_ged=gf_to_ged_map
            )
            if info.get("date_answered"):
                return info["date_answered"]
    return None


def _build_document_code(doc_row) -> str:
    """Build the document code string from doc fields."""
    parts = [
        str(doc_row.get("affaire", "") or ""),
        str(doc_row.get("projet", "") or ""),
        str(doc_row.get("batiment", "") or ""),
        str(doc_row.get("phase", "") or ""),
        str(doc_row.get("emetteur", "") or ""),
        str(doc_row.get("specialite", "") or ""),
        str(doc_row.get("lot", "") or ""),
        str(doc_row.get("type_de_doc", "") or ""),
        str(doc_row.get("zone", "") or ""),
        str(doc_row.get("niveau", "") or ""),
        str(doc_row.get("numero_normalized", "") or ""),
        str(doc_row.get("indice", "") or ""),
    ]
    return "_".join(p for p in parts if p and p != "None")


def _truncate(s: str, max_len: int) -> str:
    if len(s) > max_len:
        return s[:max_len - 3] + "..."
    return s


# ─────────────────────────────────────────────────────────────
# DISCREPANCY REPORT WRITER
# ─────────────────────────────────────────────────────────────

def write_discrepancy_report(
    output_path: str,
    discrepancies: list,
    title_suffix: str = "",
):
    """
    Write a discrepancy report with flag-type breakdown summary.
    Patch D: adds SEVERITY column and severity breakdown tab.

    discrepancies: list of dicts with keys:
      sheet_name, document_code, numero, indice, field,
      ged_value, gf_value, flag_type, severity (optional),
      ged_value_normalized, gf_value_normalized (optional)
    title_suffix: appended to sheet tab titles for filtered views
    """
    import collections
    FLAG_COLORS = {
        "MISSING_IN_GF":            "FFC7CE",
        "MISSING_IN_GED":           "FFEB9C",
        "INDICE_MISMATCH":          "FCE4D6",
        "TITRE_MISMATCH":           "D9D9D9",
        "DATE_MISMATCH":            "D9EAD3",
        "SHEET_MISMATCH":           "CFE2F3",
        "BENTIN_LEGACY_EXCEPTION":  "D9D9D9",   # grey — excluded legacy
    }
    SEVERITY_COLORS = {
        "REVIEW_REQUIRED": "FF0000",  # red header
        "COSMETIC":        "92D050",  # green
        "EXCLUDED":        "BFBFBF",  # grey
        "INFO":            "9DC3E6",  # light blue
    }

    wb = openpyxl.Workbook()

    # ── Sheet 1: Full detail (with severity column) ──
    ws = wb.active
    ws.title = ("Discrepancies" + title_suffix)[:31]

    headers = [
        "Sheet", "Document Code", "Numéro", "Indice",
        "Field", "GED Value", "GF Value", "Flag Type",
        "Severity",                                    # Patch D
        "GED Value (normalized)", "GF Value (normalized)",  # Patch B debug
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for r, d in enumerate(discrepancies, 2):
        flag     = d.get("flag_type", "")
        severity = d.get("severity", "")
        row_fill = PatternFill("solid", fgColor=FLAG_COLORS.get(flag, "FFFFFF"))
        vals = [
            d.get("sheet_name", ""),
            d.get("document_code", ""),
            d.get("numero", ""),
            d.get("indice", ""),
            d.get("field", ""),
            str(d.get("ged_value", "")),
            str(d.get("gf_value", "")),
            flag,
            severity,
            str(d.get("ged_value_normalized", "")),
            str(d.get("gf_value_normalized", "")),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = row_fill
        # Override severity cell color for quick visual scan
        if severity:
            sev_fill_color = SEVERITY_COLORS.get(severity)
            if sev_fill_color:
                ws.cell(row=r, column=9).fill = PatternFill("solid",
                                                             fgColor=sev_fill_color)

    col_widths = [38, 42, 10, 8, 18, 35, 35, 18, 16, 35, 35]
    for col, width in zip(range(1, len(col_widths) + 1), col_widths):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Summary by flag type AND severity ──
    ws2 = wb.create_sheet(("Summary" + title_suffix)[:31])
    ws2.cell(row=1, column=1, value="Flag Type").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Severity").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Description").font = Font(bold=True)

    DESCRIPTIONS = {
        "MISSING_IN_GF":   "GED dernier indice absent du GF existant",
        "MISSING_IN_GED":  "GF a une ligne absente du GED dernier indice",
        "INDICE_MISMATCH": "Même numéro mais indice différent dans GF",
        "TITRE_MISMATCH":  "Même numéro+indice mais libellé différent",
        "DATE_MISMATCH":   "Même numéro+indice mais date diffusion différente",
        "SHEET_MISMATCH":  "Numéro présent dans GF sur une autre feuille",
    }

    # Group by (flag_type, severity)
    combo_counts = collections.Counter(
        (d.get("flag_type", ""), d.get("severity", "")) for d in discrepancies
    )
    for r, ((ftype, sev), cnt) in enumerate(sorted(combo_counts.items()), 2):
        c1 = ws2.cell(row=r, column=1, value=ftype)
        c2 = ws2.cell(row=r, column=2, value=cnt)
        c3 = ws2.cell(row=r, column=3, value=sev)
        c4 = ws2.cell(row=r, column=4, value=DESCRIPTIONS.get(ftype, ""))
        fgcolor = FLAG_COLORS.get(ftype, "FFFFFF")
        for cell in (c1, c2, c4):
            cell.fill = PatternFill("solid", fgColor=fgcolor)
        if sev:
            sev_color = SEVERITY_COLORS.get(sev)
            if sev_color:
                c3.fill = PatternFill("solid", fgColor=sev_color)

    total_row = len(combo_counts) + 3
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=len(discrepancies)).font = Font(bold=True)

    # Severity sub-totals
    sev_counts = collections.Counter(d.get("severity", "") for d in discrepancies)
    total_row += 2
    ws2.cell(row=total_row, column=1, value="By Severity:").font = Font(bold=True)
    for i, (sev, cnt) in enumerate(sorted(sev_counts.items()), 1):
        ws2.cell(row=total_row + i, column=1, value=sev)
        ws2.cell(row=total_row + i, column=2, value=cnt)
        sev_color = SEVERITY_COLORS.get(sev)
        if sev_color:
            ws2.cell(row=total_row + i, column=1).fill = PatternFill("solid",
                                                                       fgColor=sev_color)

    for col, w in zip(range(1, 5), [22, 8, 16, 50]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    return _safe_save(wb, output_path)


# ─────────────────────────────────────────────────────────────
# ANOMALY REPORT WRITER
# ─────────────────────────────────────────────────────────────

def write_anomaly_report(output_path: str, anomalies_df: pd.DataFrame):
    """
    Write anomaly report grouped by contractor (sheet/lot).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomaly Report"

    headers = [
        "Sheet", "Doc ID", "EMETTEUR", "LOT", "TYPE DOC",
        "NUMERO", "INDICE", "Libellé", "Anomaly Flags",
        "Version Confidence", "Created At",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C55A11")
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for _, doc in anomalies_df.iterrows():
        flags = doc.get("anomaly_flags", [])
        if isinstance(flags, list):
            flags_str = ", ".join(flags)
        else:
            flags_str = str(flags)

        ws.cell(row=row, column=1, value=doc.get("gf_sheet_name", ""))
        ws.cell(row=row, column=2, value=doc.get("doc_id", ""))
        ws.cell(row=row, column=3, value=doc.get("emetteur", ""))
        ws.cell(row=row, column=4, value=doc.get("lot", ""))
        ws.cell(row=row, column=5, value=doc.get("type_de_doc", ""))
        ws.cell(row=row, column=6, value=doc.get("numero_normalized", ""))
        ws.cell(row=row, column=7, value=doc.get("indice", ""))
        ws.cell(row=row, column=8, value=doc.get("libelle_du_document", ""))
        ws.cell(row=row, column=9, value=flags_str)
        ws.cell(row=row, column=10, value=doc.get("version_confidence", ""))
        ws.cell(row=row, column=11, value=doc.get("created_at", ""))

        if flags_str:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(
                    "solid", fgColor=COLOUR_ANOMALY
                )
        row += 1

    for col, width in zip(range(1, 12), [20, 38, 15, 10, 10, 12, 6, 50, 35, 12, 18]):
        ws.column_dimensions[get_column_letter(col)].width = width

    return _safe_save(wb, output_path)


def _write_resolution_log(output_path: str, df: pd.DataFrame,
                          title: str, header_color: str):
    """Generic writer for AUTO_RESOLUTION_LOG and IGNORED_ITEMS_LOG."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    headers = [
        "Resolution Status", "Doc ID", "EMETTEUR", "LOT", "TYPE DOC",
        "NUMERO", "INDICE", "Libellé", "Anomaly Flags",
        "Confidence", "Family Cluster", "Created At",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (_, doc) in enumerate(df.iterrows(), 2):
        flags = doc.get("anomaly_flags", [])
        flags_str = ", ".join(flags) if isinstance(flags, list) else str(flags)

        ws.cell(row=row_idx, column=1, value=doc.get("resolution_status", ""))
        ws.cell(row=row_idx, column=2, value=doc.get("doc_id", ""))
        ws.cell(row=row_idx, column=3, value=doc.get("emetteur", ""))
        ws.cell(row=row_idx, column=4, value=doc.get("lot", ""))
        ws.cell(row=row_idx, column=5, value=doc.get("type_de_doc", ""))
        ws.cell(row=row_idx, column=6, value=doc.get("numero_normalized", ""))
        ws.cell(row=row_idx, column=7, value=doc.get("indice", ""))
        ws.cell(row=row_idx, column=8, value=str(doc.get("libelle_du_document", ""))[:80])
        ws.cell(row=row_idx, column=9, value=flags_str)
        ws.cell(row=row_idx, column=10, value=doc.get("version_confidence", ""))
        ws.cell(row=row_idx, column=11, value=str(doc.get("family_cluster_id", ""))[:8])
        ws.cell(row=row_idx, column=12, value=doc.get("created_at", ""))

    for col, w in zip(range(1, 13), [18, 36, 15, 10, 10, 12, 6, 60, 35, 10, 10, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w

    return _safe_save(wb, output_path)


def write_auto_resolution_log(output_path: str, df: pd.DataFrame):
    """Write AUTO_RESOLVED anomaly items — automatically handled, no review needed."""
    return _write_resolution_log(
        output_path, df, "Auto Resolution Log", "2E75B6"
    )


def write_ignored_items_log(output_path: str, df: pd.DataFrame):
    """Write IGNORED items — excluded by config rules (LOT I01/I02, GOE filter, etc.)."""
    return _write_resolution_log(
        output_path, df, "Ignored Items Log", "7F6000"
    )


# ─────────────────────────────────────────────────────────────
# INSERT LOG WRITER
# ─────────────────────────────────────────────────────────────

def write_insert_log(output_path: str, insert_log: list) -> str:
    """
    Write INSERT_LOG.xlsx — explicit log of newly inserted rows in CLEAN GF.

    Each entry in insert_log is a dict with:
      Sheet, Emetteur, Lot, Numero, Indice, Titre, Type Doc,
      Date réception, Date contract, Date réel de visa, VISA Global,
      Reason, Confidence
    """
    COLS = [
        ("Sheet",             30),
        ("Emetteur",          10),
        ("Lot",               10),
        ("Numero",            14),
        ("Indice",             6),
        ("Titre",             45),
        ("Type Doc",          10),
        ("Date réception",    14),
        ("Date contract",     14),
        ("Date réel de visa", 14),
        ("VISA Global",       12),
        ("Reason",            22),
        ("Confidence",        12),
    ]

    HEADER_BG  = PatternFill("solid", fgColor="1F4E79")
    EVEN_BG    = PatternFill("solid", fgColor="EBF3FB")
    ODD_BG     = PatternFill("solid", fgColor="FFFFFF")
    VISA_FILLS = {
        "VSO":     PatternFill("solid", fgColor=COLOUR_VSO),
        "VAO":     PatternFill("solid", fgColor=COLOUR_VAO),
        "REF":     PatternFill("solid", fgColor=COLOUR_REF),
        "SAS REF": PatternFill("solid", fgColor=COLOUR_REF),
        "HM":      PatternFill("solid", fgColor=COLOUR_HM),
        "SUS":     PatternFill("solid", fgColor=COLOUR_SUS),
    }
    border = _thin_border()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INSERT LOG"
    ws.freeze_panes = "A2"

    # ── Header row ──────────────────────────────────────────────
    for col_idx, (col_name, col_w) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(color="FFFFFF", bold=True, size=9)
        cell.fill      = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[1].height = 20

    # ── Summary row at top (row 2) — injected after header ──────
    # (actual data starts at row 3 to make room; handled below by starting at row 2)

    # ── Data rows ───────────────────────────────────────────────
    DATE_COLS = {"Date réception", "Date contract", "Date réel de visa"}
    col_names  = [c for c, _ in COLS]

    for row_idx, entry in enumerate(insert_log, start=2):
        row_fill = EVEN_BG if (row_idx % 2 == 0) else ODD_BG
        for col_idx, col_name in enumerate(col_names, start=1):
            val = entry.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = border

            # Date formatting
            if col_name in DATE_COLS and val is not None:
                try:
                    cell.number_format = "DD/MM/YYYY"
                except Exception:
                    pass

            # VISA colour
            if col_name == "VISA Global":
                fill = VISA_FILLS.get(str(val or ""), row_fill)
                cell.fill = fill
                if val:
                    cell.font = BOLD_FONT
            else:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 14

    # ── Auto-filter on header ───────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    return _safe_save(wb, output_path)


# ─────────────────────────────────────────────────────────────
# NEW SUBMITTAL ANALYSIS WRITER
# ─────────────────────────────────────────────────────────────

# Status → fill colour (ARGB)
_NS_FILL = {
    "NEW_PENDING_SAS":        "FFFFF2CC",   # light yellow
    "NEW_RECENT_SAS_REF":     "FFFFC7CE",   # light red
    "NEW_RECENT_SAS_APPROVED":"FFC6EFCE",   # light green
    "NOT_NEW_BACKLOG":        "FFD9D9D9",   # light grey
    "EXCLUDED":               "FFFDE9D9",   # light orange
    "AMBIGUOUS":              "FFEDEDED",   # very light grey
    "ALREADY_IN_GF":          "FFDCE6F1",   # light blue
}

# Columns: (name, width)  — keys must match _build_new_submittal_analysis() output
_NS_COLUMNS = [
    ("new_submittal_status",  26),
    ("Sheet target",          36),
    ("Emetteur",              10),
    ("Lot",                   20),
    ("Numero",                20),
    ("Indice",                 8),
    ("Titre",                 46),
    ("SAS status type",       18),
    ("SAS result",            14),
    ("SAS date",              14),
    ("days_from_data_date",   18),
    ("exists_in_original_gf", 20),
    ("data_date",             14),
    ("rationale",             40),
]


def _write_summary_tab(ws, rows: list, group_key: str, title: str):
    """
    Write a summary table grouped by `group_key` value.
    Columns: group_key value | Count | % of Total
    """
    from collections import Counter
    counts = Counter(r.get(group_key, "") for r in rows)
    total  = sum(counts.values()) or 1

    HDR_BG   = PatternFill("solid", fgColor="FF1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFFFF", size=11)
    BOLD_F   = Font(bold=True, size=10)
    NORM_F   = Font(size=10)
    EVEN_BG  = PatternFill("solid", fgColor="FFDCE6F1")
    ODD_BG   = PatternFill("solid", fgColor="FFFFFFFF")
    border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # Title row
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells("A1:C1")

    # Header row
    for col_idx, hdr in enumerate([group_key, "Count", "% of Total"], 1):
        c = ws.cell(row=2, column=col_idx, value=hdr)
        c.font      = HDR_FONT
        c.fill      = HDR_BG
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    for r_idx, (key, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 3):
        bg = PatternFill("solid", fgColor=_NS_FILL.get(str(key), "FFFFFFFF"))
        row_fill = bg if key in _NS_FILL else (EVEN_BG if r_idx % 2 == 0 else ODD_BG)

        c1 = ws.cell(row=r_idx, column=1, value=str(key))
        c1.font = BOLD_F; c1.fill = row_fill; c1.border = border

        c2 = ws.cell(row=r_idx, column=2, value=cnt)
        c2.font = NORM_F; c2.fill = row_fill; c2.border = border
        c2.alignment = Alignment(horizontal="center")

        c3 = ws.cell(row=r_idx, column=3, value=round(cnt / total * 100, 1))
        c3.font = NORM_F; c3.fill = row_fill; c3.border = border
        c3.number_format = "0.0\"%\""
        c3.alignment = Alignment(horizontal="center")

    # Total row
    total_r = len(counts) + 3
    c = ws.cell(row=total_r, column=1, value="TOTAL")
    c.font = BOLD_F
    c2 = ws.cell(row=total_r, column=2, value=total)
    c2.font = BOLD_F
    c2.alignment = Alignment(horizontal="center")


def _write_pivot_tab(ws, rows: list, row_key: str, col_key: str, title: str):
    """
    Write a pivot-style cross-tab: rows = row_key unique values,
    columns = col_key unique values, cells = count.
    """
    from collections import defaultdict
    pivot: dict = defaultdict(lambda: defaultdict(int))
    for r in rows:
        rk = str(r.get(row_key, "") or "")
        ck = str(r.get(col_key, "") or "")
        pivot[rk][ck] += 1

    col_vals = sorted({str(r.get(col_key, "") or "") for r in rows})
    row_vals = sorted(pivot.keys())
    total    = len(rows)

    HDR_BG   = PatternFill("solid", fgColor="FF1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFFFF", size=10)
    BOLD_F   = Font(bold=True, size=10)
    NORM_F   = Font(size=10)
    border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    n_cols = len(col_vals) + 2
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Header row (row 2): row_key label, then each col_key value, then Total
    ws.cell(row=2, column=1, value=row_key).font = HDR_FONT
    ws.cell(row=2, column=1).fill   = HDR_BG
    ws.cell(row=2, column=1).border = border
    ws.column_dimensions["A"].width = 36

    for c_idx, cv in enumerate(col_vals, 2):
        cell = ws.cell(row=2, column=c_idx, value=cv)
        cell.font = HDR_FONT; cell.fill = HDR_BG; cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c_idx)].width = 20

    total_col = len(col_vals) + 2
    tc = ws.cell(row=2, column=total_col, value="Total")
    tc.font = HDR_FONT; tc.fill = HDR_BG; tc.border = border
    tc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(total_col)].width = 10

    EVEN_BG = PatternFill("solid", fgColor="FFDCE6F1")
    ODD_BG  = PatternFill("solid", fgColor="FFFFFFFF")

    for r_idx, rv in enumerate(row_vals, 3):
        row_fill = EVEN_BG if r_idx % 2 == 0 else ODD_BG
        rc = ws.cell(row=r_idx, column=1, value=rv)
        rc.font = BOLD_F; rc.fill = row_fill; rc.border = border

        row_total = 0
        for c_idx, cv in enumerate(col_vals, 2):
            cnt = pivot[rv][cv]
            row_total += cnt
            cell = ws.cell(row=r_idx, column=c_idx, value=cnt if cnt else "")
            cell.font = NORM_F; cell.fill = row_fill; cell.border = border
            cell.alignment = Alignment(horizontal="center")

        tot_cell = ws.cell(row=r_idx, column=total_col, value=row_total)
        tot_cell.font = BOLD_F; tot_cell.fill = row_fill; tot_cell.border = border
        tot_cell.alignment = Alignment(horizontal="center")

    # Grand-total footer
    footer_r = len(row_vals) + 3
    ws.cell(row=footer_r, column=1, value="TOTAL").font = BOLD_F
    col_totals = [sum(pivot[rv][cv] for rv in row_vals) for cv in col_vals]
    for c_idx, ct in enumerate(col_totals, 2):
        cell = ws.cell(row=footer_r, column=c_idx, value=ct)
        cell.font = BOLD_F
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=footer_r, column=total_col, value=total).font = BOLD_F


def write_new_submittal_analysis(
    analysis_path: str,
    summary_path:  str,
    rows: list,
) -> None:
    """
    Write two Excel files:
      1. analysis_path  — full row-level detail + per-status tabs
      2. summary_path   — summary + pivot tabs only

    `rows` is a list of dicts, one per analysed doc, with keys matching
    _NS_COLUMNS plus 'new_submittal_status', 'days_since_sas', 'rationale'.
    """
    Path(analysis_path).parent.mkdir(parents=True, exist_ok=True)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)

    HEADER_BG   = PatternFill("solid", fgColor="FF1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT   = Font(bold=True, size=10)
    EVEN_BG     = PatternFill("solid", fgColor="FFDCE6F1")
    ODD_BG      = PatternFill("solid", fgColor="FFFFFFFF")
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── 1. ANALYSIS workbook ─────────────────────────────────────
    wb_a = openpyxl.Workbook()

    # Tab: ALL DOCS (full detail)
    ws_all = wb_a.active
    ws_all.title = "ALL DOCS"
    col_names = [c for c, _ in _NS_COLUMNS]

    for col_idx, (col_name, col_w) in enumerate(_NS_COLUMNS, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = border
        ws_all.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws_all.row_dimensions[1].height = 22

    DATE_COLS_NS = {"SAS date", "data_date"}

    for row_idx, entry in enumerate(rows, start=2):
        status   = str(entry.get("new_submittal_status") or "")
        row_fill = PatternFill("solid", fgColor=_NS_FILL.get(status, "FFFFFFFF"))

        for col_idx, col_name in enumerate(col_names, start=1):
            val  = entry.get(col_name)
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = NORMAL_FONT
            cell.fill      = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = border

            if col_name in DATE_COLS_NS and val is not None:
                try:
                    cell.number_format = "DD/MM/YYYY"
                except Exception:
                    pass

        ws_all.row_dimensions[row_idx].height = 14

    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(_NS_COLUMNS))}1"
    ws_all.freeze_panes    = "A2"

    # Per-status detail tabs
    from collections import defaultdict
    by_status: dict = defaultdict(list)
    for r in rows:
        by_status[r.get("new_submittal_status", "UNKNOWN")].append(r)

    STATUS_ORDER = [
        "NEW_PENDING_SAS", "NEW_RECENT_SAS_REF", "NEW_RECENT_SAS_APPROVED",
        "NOT_NEW_BACKLOG", "EXCLUDED", "AMBIGUOUS", "ALREADY_IN_GF",
    ]
    for status in STATUS_ORDER:
        status_rows = by_status.get(status, [])
        if not status_rows:
            continue
        ws_s = wb_a.create_sheet(title=status[:31])  # sheet name ≤31 chars

        for col_idx, (col_name, col_w) in enumerate(_NS_COLUMNS, 1):
            cell = ws_s.cell(row=1, column=col_idx, value=col_name)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_BG
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = border
            ws_s.column_dimensions[get_column_letter(col_idx)].width = col_w
        ws_s.row_dimensions[1].height = 22

        sf = PatternFill("solid", fgColor=_NS_FILL.get(status, "FFFFFFFF"))
        for row_idx, entry in enumerate(status_rows, start=2):
            for col_idx, col_name in enumerate(col_names, start=1):
                val  = entry.get(col_name)
                cell = ws_s.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = NORMAL_FONT
                cell.fill      = sf
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border    = border
                if col_name in DATE_COLS_NS and val is not None:
                    try:
                        cell.number_format = "DD/MM/YYYY"
                    except Exception:
                        pass
            ws_s.row_dimensions[row_idx].height = 14

        ws_s.auto_filter.ref = f"A1:{get_column_letter(len(_NS_COLUMNS))}1"
        ws_s.freeze_panes    = "A2"

    # Summary tab in analysis workbook
    ws_sum_a = wb_a.create_sheet(title="SUMMARY")
    _write_summary_tab(ws_sum_a, rows, "new_submittal_status",
                       "New Submittal Classification — All Docs")

    _safe_save(wb_a, analysis_path)

    # ── 2. SUMMARY workbook ──────────────────────────────────────
    wb_s = openpyxl.Workbook()

    ws_by_status = wb_s.active
    ws_by_status.title = "By Status"
    _write_summary_tab(ws_by_status, rows, "new_submittal_status",
                       "New Submittal — by Status")

    ws_by_sheet = wb_s.create_sheet(title="By Sheet")
    # Only new submittals (not already in GF, not excluded)
    new_only = [r for r in rows if r.get("new_submittal_status") not in
                ("ALREADY_IN_GF", "EXCLUDED")]
    _write_summary_tab(ws_by_sheet, new_only, "Sheet target",
                       "New Submittals — by GF Sheet")

    ws_pivot = wb_s.create_sheet(title="Sheet × Status")
    _write_pivot_tab(ws_pivot, rows, "Sheet target", "new_submittal_status",
                     "New Submittals — Sheet × Status")

    ws_emit = wb_s.create_sheet(title="By Emetteur")
    _write_summary_tab(ws_emit, new_only, "Emetteur",
                       "New Submittals — by Emetteur")

    _safe_save(wb_s, summary_path)
    print(f"    → NEW_SUBMITTAL_ANALYSIS.xlsx written ({len(rows)} rows)")
    print(f"    → new_submittal_summary.xlsx written")
