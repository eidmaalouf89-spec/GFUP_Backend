"""
consultant_gf_writer.py
JANSA VISASIST — GF Consultant Enrichment Writer
Version 1.0 — April 2026

Loads GF_V0_CLEAN.xlsx and enriches it with consultant report data
in two stages:

Stage 1: Update consultant DATE + STATUT columns only.
         → output/GF_consultant_enriched_stage1.xlsx

Stage 2: Update consultant DATE + STATUT + OBSERVATIONS column.
         → output/GF_consultant_enriched_stage2.xlsx

Matching strategy:
  - For each GF row, extract NUMERO (col 6) and INDICE (col 7)
  - Look up enrichment records keyed by (numero, indice)
  - Update the matching consultant approver column(s)

GF approver name → consultant source mapping:
  "ACOUSTICIEN AVLS"   ← AVLS
  "AMO HQE LE SOMMER"  ← LE_SOMMER
  "BET STR-TERRELL"    ← TERRELL
  "SOCOTEC" or "BC SOCOTEC" ← SOCOTEC
"""

import copy
import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# GF approver canonical → consultant SOURCE mapping
# ---------------------------------------------------------------------------

GF_APPROVER_TO_CONSULTANT = {
    "ACOUSTICIEN AVLS":   "AVLS",
    "AMO HQE LE SOMMER":  "LE_SOMMER",
    "BET STR-TERRELL":    "TERRELL",
    "SOCOTEC":            "SOCOTEC",
    "BC SOCOTEC":         "SOCOTEC",
}

# ---------------------------------------------------------------------------
# GF structure constants
# ---------------------------------------------------------------------------

GF_HEADER_ROW      = 7   # Column names (DOCUMENT, TITRE, …)
GF_APPROVER_ROW    = 8   # Approver names
GF_SUBHEADER_ROW   = 9   # DATE / N° / STATUT
GF_DATA_START_ROW  = 10

GF_COL_NUMERO = 6   # "N° Doc"
GF_COL_INDICE = 7   # "IND"

STATUS_COLOURS = {
    "VSO":  "C6EFCE",
    "VAO":  "FFEB9C",
    "REF":  "FFC7CE",
    "HM":   "D9D9D9",
    "SUS":  "FCE4D6",
    "FAV":  "C6EFCE",   # Socotec FAVORABLE = same green as VSO
    "DEF":  "FFC7CE",   # Socotec DÉFAVORABLE = same red as REF
}

HIGHLIGHT_DATE_HEX  = "BDD7EE"   # light blue — consultant date filled
HIGHLIGHT_STAT_HEX  = "E2EFDA"   # light green — consultant status filled


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan") else s


def _normalize_numero(v) -> str:
    """Normalize to plain integer string."""
    s = _clean(v)
    digits = re.sub(r'\D', '', s)
    if not digits:
        return ""
    try:
        return str(int(digits))
    except ValueError:
        return digits


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


# ---------------------------------------------------------------------------
# Parse GF sheet column map
# ---------------------------------------------------------------------------

def _parse_gf_columns(ws) -> dict:
    """
    Read rows 7-9 of a GF sheet to build column map.

    Returns:
        {
          'numero_col': int,
          'indice_col': int,
          'observations_col': int,
          'approvers': {
              'ACOUSTICIEN AVLS': {'date': col, 'num': col, 'statut': col},
              ...
          }
        }
    """
    # Row 7: base column names
    row7 = list(ws.iter_rows(min_row=GF_HEADER_ROW, max_row=GF_HEADER_ROW, values_only=True))[0]
    # Row 8: approver names (merged cells show value only in first cell)
    row8 = list(ws.iter_rows(min_row=GF_APPROVER_ROW, max_row=GF_APPROVER_ROW, values_only=True))[0]
    # Row 9: sub-headers DATE/N°/STATUT
    row9 = list(ws.iter_rows(min_row=GF_SUBHEADER_ROW, max_row=GF_SUBHEADER_ROW, values_only=True))[0]

    result = {
        "numero_col": None,
        "indice_col": None,
        "observations_col": None,
        "approvers": {},
    }

    # Find base columns from row 7
    for i, val in enumerate(row7):
        if val is None:
            continue
        norm = str(val).strip().upper()
        if "N° DOC" in norm or norm == "N° DOC":
            result["numero_col"] = i + 1   # 1-indexed
        elif norm == "IND":
            result["indice_col"] = i + 1
        elif "OBSERVATIONS" in norm:
            result["observations_col"] = i + 1

    # Find approver columns from row 8
    current_approver = None
    current_col = None
    sub_count = 0
    for i, val in enumerate(row8):
        col = i + 1
        if val is not None and str(val).strip():
            current_approver = str(val).strip()
            current_col = col
            sub_count = 0

        if current_approver and sub_count < 3:
            if sub_count == 0:
                result["approvers"].setdefault(current_approver, {})["date"] = col
            elif sub_count == 1:
                result["approvers"][current_approver]["num"] = col
            elif sub_count == 2:
                result["approvers"][current_approver]["statut"] = col
            sub_count += 1

    return result


# ---------------------------------------------------------------------------
# Build enrichment lookup
# ---------------------------------------------------------------------------

def _build_enrichment_lookup(enrichments: list) -> dict:
    """
    Build a fast lookup: (numero_normalized, indice) → [enrichment_record, ...]

    Also build a numero-only lookup for when indice is absent.
    """
    by_num_ind = {}
    by_num = {}

    for rec in enrichments:
        num = _normalize_numero(rec.get("matched_numero", ""))
        ind = _clean(rec.get("matched_indice", "")).upper()
        if not num:
            continue

        key_full = (num, ind)
        by_num_ind.setdefault(key_full, []).append(rec)
        by_num.setdefault(num, []).append(rec)

    return {"by_num_ind": by_num_ind, "by_num": by_num}


# ---------------------------------------------------------------------------
# Observation text builder
# ---------------------------------------------------------------------------

_OBS_SORT_PRIORITY = [
    "MOEX", "ARCHI", "EGIS", "ELIOTH", "SPK", "ASC",
    "TERRELL", "GEOLIA", "POLLUTION",
    "AVLS", "ACOUSTICIEN",
    "AMO HQE", "LE SOMMER",
    "SOCOTEC",
]

def _obs_sort_key(name: str) -> int:
    upper = name.upper()
    for i, kw in enumerate(_OBS_SORT_PRIORITY):
        if kw in upper:
            return i
    return len(_OBS_SORT_PRIORITY)


def _build_updated_observation(
    existing_obs: str,
    new_entries: list,   # [{approver, status, comment}, ...]
) -> str:
    """
    Merge new consultant entries into existing observation text.

    Format per entry: "{Approver}: {STATUS} — {comment}"
    Existing approver entries are replaced; new ones are appended in order.
    """
    # Parse existing observation into per-approver blocks
    existing = {}
    if existing_obs:
        for line in existing_obs.split("\n"):
            line = line.strip()
            if not line:
                continue
            # Detect "APPROVER: STATUS — comment" format
            m = re.match(r'^([^:]+?):\s*([A-Z]+(?:\s[A-Z]+)?)\s*(?:—\s*(.*))?$', line)
            if m:
                existing[m.group(1).strip()] = line
            else:
                # Non-structured line — keep as-is under a blank key
                existing[f"_raw_{len(existing)}"] = line

    # Overwrite/add new consultant entries
    for entry in new_entries:
        approver = _clean(entry.get("approver", ""))
        status   = _clean(entry.get("status", ""))
        comment  = _clean(entry.get("comment", ""))
        if not approver:
            continue
        line = approver
        if status:
            line += f": {status}"
        if comment:
            line += f" — {comment}"
        # Remove only the previous structured entry for this approver.
        # Preserve any raw/unstructured observation lines.
        existing = {
            k: v for k, v in existing.items()
            if (
                (k.startswith("_raw_") and not str(v).upper().startswith(f"{approver.upper()}:"))
                or (not k.startswith("_raw_") and approver.upper() not in k.upper())
            )
        } | {approver: line}

    # Sort all lines by priority
    sorted_items = sorted(existing.items(), key=lambda x: _obs_sort_key(x[0]))
    return "\n".join(v for _, v in sorted_items)


def _resolve_row_enrichments(num: str, ind: str, by_num_ind: dict, by_num: dict) -> list:
    """
    Resolve consultant enrichments for one GF row safely.

    Prefer exact (NUMERO, INDICE) matches. Fall back to NUMERO-only only when
    the GF row itself has no INDICE and there is exactly one consultant
    enrichment candidate for that NUMERO.
    """
    exact = by_num_ind.get((num, ind), [])
    if exact:
        return exact

    if ind:
        return []

    num_only = by_num.get(num, [])
    return num_only if len(num_only) == 1 else []


# ---------------------------------------------------------------------------
# Core sheet enrichment function
# ---------------------------------------------------------------------------

def _enrich_sheet_inplace(
    wb:            openpyxl.Workbook,
    sheet_name:    str,
    enrich_lookup: dict,
    stage:         int,   # 1 = date+status only, 2 = also update observations
) -> dict:
    """
    Modify sheet_name in wb in-place.
    Called after the source workbook has been copied to the output path.

    Returns per-sheet stats dict.
    """
    ws_out = wb[sheet_name]

    # Parse column map from the sheet
    col_map = _parse_gf_columns(ws_out)

    numero_col       = col_map.get("numero_col", GF_COL_NUMERO)
    indice_col       = col_map.get("indice_col", GF_COL_INDICE)
    observations_col = col_map.get("observations_col")
    approver_cols    = col_map.get("approvers", {})

    stats = {"sheet": sheet_name, "rows_enriched": 0, "updates": []}

    by_num_ind = enrich_lookup["by_num_ind"]
    by_num     = enrich_lookup["by_num"]

    # Iterate data rows
    for row in ws_out.iter_rows(min_row=GF_DATA_START_ROW):
        row_idx = row[0].row

        # Extract NUMERO and INDICE from this row
        num_cell = ws_out.cell(row=row_idx, column=numero_col)
        ind_cell = ws_out.cell(row=row_idx, column=indice_col)
        num = _normalize_numero(num_cell.value)
        ind = _clean(ind_cell.value).upper()

        if not num:
            continue

        # Find matching enrichment records
        enrich_recs = _resolve_row_enrichments(num, ind, by_num_ind, by_num)
        if not enrich_recs:
            continue

        row_updated = False
        new_obs_entries = []

        for erec in enrich_recs:
            gf_canonical = erec.get("gf_approver_canonical", "")
            date_str     = erec.get("consultant_date_fiche", "")
            status       = erec.get("consultant_statut_norm", "")
            comment      = erec.get("consultant_commentaire", "")

            # Find approver column group
            # Try exact match first, then substring match
            app_cols = approver_cols.get(gf_canonical)
            if app_cols is None:
                for app_name, app_c in approver_cols.items():
                    if gf_canonical.upper() in app_name.upper() or app_name.upper() in gf_canonical.upper():
                        app_cols = app_c
                        break

            if app_cols is None:
                logger.debug(
                    "Approver '%s' not found in sheet '%s' columns %s",
                    gf_canonical, sheet_name, list(approver_cols.keys())
                )
                continue

            # Stage 1+2: update DATE and STATUT
            date_col   = app_cols.get("date")
            statut_col = app_cols.get("statut")

            if date_col and date_str:
                cell = ws_out.cell(row=row_idx, column=date_col)
                cell.value = date_str
                cell.fill  = _fill(HIGHLIGHT_DATE_HEX)

            if statut_col and status:
                cell = ws_out.cell(row=row_idx, column=statut_col)
                cell.value = status
                hex_color  = STATUS_COLOURS.get(status, HIGHLIGHT_STAT_HEX)
                cell.fill  = _fill(hex_color)

            if (date_str or status):
                row_updated = True
                stats["updates"].append({
                    "row": row_idx, "numero": num, "indice": ind,
                    "approver": gf_canonical,
                    "date": date_str, "status": status,
                    "comment": comment,
                })

            # Collect for observation update in stage 2
            if stage >= 2 and (status or comment):
                new_obs_entries.append({
                    "approver": gf_canonical,
                    "status":   status,
                    "comment":  comment,
                })

        # Stage 2: update OBSERVATIONS
        if stage >= 2 and observations_col and new_obs_entries:
            obs_cell    = ws_out.cell(row=row_idx, column=observations_col)
            existing    = _clean(obs_cell.value)
            updated_obs = _build_updated_observation(existing, new_obs_entries)
            obs_cell.value     = updated_obs
            obs_cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_updated = True

        if row_updated:
            stats["rows_enriched"] += 1

    return stats


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_gf_enriched(
    gf_source_path:  Path,
    enrichments:     list,
    output_path:     Path,
    stage:           int = 1,
) -> dict:
    """
    Load GF_V0_CLEAN.xlsx, apply consultant enrichment, save as output_path.

    Strategy: copy the source file to output_path first, then load and modify
    in-place (avoids cross-workbook copy issues with openpyxl).

    Parameters
    ----------
    gf_source_path : path to GF_V0_CLEAN.xlsx (read-only source)
    enrichments    : list of enrichment records from consultant_matcher
    output_path    : destination path for enriched GF
    stage          : 1 = date+status only, 2 = also update OBSERVATIONS

    Returns
    -------
    dict with stats per sheet and overall
    """
    import shutil

    logger.info(
        "Loading GF source: %s  (stage=%d, enrichments=%d)",
        gf_source_path, stage, len(enrichments)
    )

    # Step 1: Copy source to output path (preserves all formatting/styling)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(gf_source_path), str(output_path))
    logger.info("Copied GF source → %s", output_path)

    # Step 2: Load the copy and modify it in-place
    wb = openpyxl.load_workbook(str(output_path), data_only=False)

    enrich_lookup = _build_enrichment_lookup(enrichments)

    all_stats = []
    total_enriched = 0

    for sheet_name in wb.sheetnames:
        stats = _enrich_sheet_inplace(wb, sheet_name, enrich_lookup, stage)
        all_stats.append(stats)
        total_enriched += stats["rows_enriched"]
        logger.info(
            "  Sheet '%-35s': %d rows enriched",
            sheet_name, stats["rows_enriched"]
        )

    wb.save(str(output_path))
    wb.close()

    logger.info(
        "Stage %d GF output saved: %s  (total rows enriched: %d)",
        stage, output_path, total_enriched
    )

    return {
        "output_path":    output_path,
        "stage":          stage,
        "sheets":         all_stats,
        "total_enriched": total_enriched,
    }
