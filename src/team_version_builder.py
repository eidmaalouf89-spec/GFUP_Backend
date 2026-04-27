"""
GF TEAM VERSION BUILDER
Surgical patch of OGF using GF_V0_CLEAN as truth.
"""

import re
import shutil
import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy as copy_obj

# ─── PATHS ──────────────────────────────────────────────────────────────────
OGF_PATH   = "input/Grandfichier_v3.xlsx"
CLEAN_PATH = "output/GF_V0_CLEAN.xlsx"
OUT_PATH   = "output/GF_TEAM_VERSION.xlsx"

# ─── CONSTANTS ──────────────────────────────────────────────────────────────
# Column keys that hold date values (used for format normalisation)
DATE_COLUMN_KEYS = frozenset({'date_diff', 'reception', 'date_contract', 'date_reel'})
# Fallback format when no reference row is available
DATE_FORMAT_FALLBACK = 'DD/MM/YY'

STATUS_COLORS = {
    'VSO':    'FFA9D08E',
    'VAO':    'FFFFD965',
    'REF':    'FFF4B083',
    'Annulé': 'FFF2F2F2',
    'HM':     'FF9BC2E6',
}
GREY_COLOR = 'FFD9D9D9'
WHITE_COLOR = 'FFFFFFFF'

# ─── COLUMN DETECTION ───────────────────────────────────────────────────────

def detect_cols(ws):
    """Detect key column positions from header rows 7-9."""
    cols = {}
    # Row 7
    for c in range(1, ws.max_column + 2):
        val = ws.cell(7, c).value
        if val is None:
            continue
        v = str(val).strip()
        if 'DOCUMENT' == v:               cols['document'] = c
        elif 'TITRE' == v:                cols['titre'] = c
        elif 'Date diffusion' in v:       cols['date_diff'] = c
        elif v == 'LOT':                  cols['lot'] = c
        elif 'TYPE DOC' in v:             cols['type_doc'] = c
        elif 'N° Doc' in v or 'N°Doc' in v: cols['numero'] = c
        elif v == 'IND':                  cols['indice'] = c
        elif v.upper() == 'NIV':          cols['niv'] = c
        elif v == 'Type':                 cols['type'] = c
        elif v == 'ANCIEN':               cols['ancien'] = c
        elif 'N°BDX' in v:               cols['bdx'] = c
        elif 'réception' in v.lower():    cols['reception'] = c
        elif 'non reçu' in v.lower():     cols['non_recu'] = c
        elif 'CONTRAC' in v.upper():      cols['date_contract'] = c
        elif 'réel' in v.lower():         cols['date_reel'] = c
        elif 'VISA' in v and 'GLOBAL' in v: cols['visa_global'] = c
        elif 'DISCREPANCES' in v.upper(): cols['discrepances'] = c
        elif v.upper() == 'OBSERVATIONS': cols['obs'] = c  # Clean GF has obs in row 7

    # Rows 8-9: consultant names + OBSERVATIONS
    consultants = {}  # name -> {date_col, num_col, stat_col}
    obs_col = None
    current_consultant = None

    for c in range(1, ws.max_column + 2):
        r8 = ws.cell(8, c).value
        r9 = ws.cell(9, c).value
        if r8 is not None:
            r8s = str(r8).strip()
            if r8s == 'OBSERVATIONS':
                obs_col = c
                current_consultant = None
            elif r8s not in ('DATE', 'N°', 'N', 'STATUT', 'APPROBATEURS'):
                current_consultant = r8s
                consultants[current_consultant] = {}
        if r9 is not None and current_consultant:
            r9s = str(r9).strip()
            if r9s == 'DATE':
                consultants[current_consultant]['date_col'] = c
            elif r9s in ('N°', 'N'):
                consultants[current_consultant]['num_col'] = c
            elif r9s == 'STATUT':
                consultants[current_consultant]['stat_col'] = c

    # Only override row-7 obs detection if row-8 scan found it
    if obs_col is not None:
        cols['obs'] = obs_col
    # If obs not found in row 8 but found in row 7, keep the row-7 value
    cols['consultants'] = consultants
    return cols


def normalize_numero(val):
    """Normalize numero value for comparison."""
    if val is None:
        return None
    try:
        return str(int(float(str(val).strip())))
    except:
        return str(val).strip()


def normalize_indice(val):
    if val is None:
        return None
    return str(val).strip().upper()


def strip_ind_suffix(doc_id):
    """Strip trailing _INDICE suffix from clean GF document ID."""
    if doc_id is None:
        return None
    s = str(doc_id).strip()
    # Remove trailing _A, _B, _AB, etc. (1-2 uppercase letters)
    s2 = re.sub(r'_[A-Z]{1,2}$', '', s)
    return s2


def normalize_doc(doc_id):
    """Normalize document ID: strip IND suffix, remove leading zeros, remove underscores."""
    if doc_id is None:
        return None
    s = strip_ind_suffix(str(doc_id).strip())
    # Remove leading zeros from numeric segments: _028246 → _28246
    s = re.sub(r'_0+(\d)', r'_\1', s)
    return s.upper()


def normalize_doc_flat(doc_id):
    """Normalize document ID by removing all underscores and leading zeros."""
    if doc_id is None:
        return None
    s = strip_ind_suffix(str(doc_id).strip())
    s = re.sub(r'0+(\d)', r'\1', s)  # remove leading zeros in any number segment
    s = s.replace('_', '').upper()
    return s


# ─── OGF ROW INDEX ──────────────────────────────────────────────────────────

def build_ogf_index(ws_read, cols_ogf):
    """
    Build multi-key index from OGF sheet:
      - doc_idx:  {document_id: [row_numbers]}
      - num_idx:  {(numero, indice): [row_numbers]}
    """
    doc_idx = {}
    num_idx = {}
    numero_col = cols_ogf.get('numero')
    indice_col = cols_ogf.get('indice')
    doc_col    = cols_ogf.get('document', 1)

    for r in range(10, ws_read.max_row + 1):
        doc = ws_read.cell(r, doc_col).value
        num = normalize_numero(ws_read.cell(r, numero_col).value) if numero_col else None
        ind = normalize_indice(ws_read.cell(r, indice_col).value) if indice_col else None

        if doc:
            doc_str = str(doc).strip()
            doc_idx.setdefault(doc_str, []).append(r)
            # Normalized (remove leading zeros)
            doc_norm = normalize_doc(doc_str)
            if doc_norm and doc_norm != doc_str.upper():
                doc_idx.setdefault(doc_norm, []).append(r)
            # Flat (remove underscores + leading zeros) for old-style documents
            doc_flat = normalize_doc_flat(doc_str)
            if doc_flat and doc_flat not in (doc_str.upper().replace('_',''), doc_norm):
                doc_idx.setdefault(doc_flat, []).append(r)

        if num:
            key = (num, ind)
            num_idx.setdefault(key, []).append(r)

    return doc_idx, num_idx


def to_datetime(val):
    """Convert a cell value to a datetime object, or None if not parseable."""
    from datetime import datetime, date as date_type
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date_type):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def titre_similarity(s1, s2):
    """Return SequenceMatcher ratio between two title strings (0.0–1.0)."""
    import difflib
    if not s1 or not s2:
        return 0.0
    return difflib.SequenceMatcher(
        None,
        str(s1).strip().upper(),
        str(s2).strip().upper()
    ).ratio()


def closest_date_match(clean_date_val, candidate_rows, ws_read, date_col_ogf):
    """
    Among candidate_rows pick the one whose date_col_ogf value is closest
    to clean_date_val.  Returns (best_row, True) if a unique best exists,
    else (None, False).
    """
    if not clean_date_val or not date_col_ogf or not candidate_rows:
        return None, False
    clean_dt = to_datetime(clean_date_val)
    if clean_dt is None:
        return None, False

    best_rows   = []
    best_delta  = None
    for r in candidate_rows:
        ogf_dt = to_datetime(ws_read.cell(r, date_col_ogf).value)
        if ogf_dt is None:
            continue
        delta = abs((clean_dt - ogf_dt).total_seconds())
        if best_delta is None or delta < best_delta:
            best_delta = delta
            best_rows  = [r]
        elif delta == best_delta:
            best_rows.append(r)

    if len(best_rows) == 1:
        return best_rows[0], True
    return None, False


def find_ogf_row(ws_read, r_clean, ws_clean, cols_clean, cols_ogf, doc_idx, num_idx):
    """
    Try to find a matching OGF row for a clean GF row.
    Matching cascade:
      S1  DOCUMENT key  (exact / normalized / flat)
      S2  (numero, indice) + ANCIEN filter
      S3  Closest Date réception / Date diffusion
      S4  Closest TITRE similarity
      S5  All candidates share the same TITRE → OGF duplicates → update all

    Returns (row_or_list, tag) where tag is one of:
      'FOUND'          – row_or_list is a single int
      'DUPLICATE_ALL'  – row_or_list is a list[int] (true OGF duplicates)
      'NOT_FOUND_NEW'  – row_or_list is None
      'NOT_FOUND_OLD'  – row_or_list is None
      'SUSPICIOUS'     – row_or_list is None
    """
    doc_col_clean = cols_clean.get('document', 1)
    num_col_clean = cols_clean.get('numero')
    ind_col_clean = cols_clean.get('indice')
    anc_col_clean = cols_clean.get('ancien')
    anc_col_ogf   = cols_ogf.get('ancien')

    doc_clean = ws_clean.cell(r_clean, doc_col_clean).value
    num_clean = normalize_numero(ws_clean.cell(r_clean, num_col_clean).value) if num_col_clean else None
    ind_clean = normalize_indice(ws_clean.cell(r_clean, ind_col_clean).value) if ind_col_clean else None
    anc_clean = ws_clean.cell(r_clean, anc_col_clean).value if anc_col_clean else None
    anc_clean_str = str(anc_clean).strip() if anc_clean is not None else ''

    if not num_clean:
        return None, 'SKIP'

    # ── Strategy 1: DOCUMENT match (try exact, then normalized) ──
    if doc_clean:
        doc_stripped = strip_ind_suffix(str(doc_clean).strip())
        doc_norm     = normalize_doc(str(doc_clean).strip())

        doc_flat = normalize_doc_flat(str(doc_clean).strip())
        for key in [doc_stripped, doc_stripped.upper(), doc_norm, doc_flat]:
            if key is None:
                continue
            doc_matches = doc_idx.get(key, [])
            if len(doc_matches) == 1:
                return doc_matches[0], 'FOUND'
            if len(doc_matches) > 1:
                ind_col_ogf = cols_ogf.get('indice', 9)
                ind_filtered = [r for r in doc_matches
                                if normalize_indice(ws_read.cell(r, ind_col_ogf).value) == ind_clean]
                if len(ind_filtered) == 1:
                    return ind_filtered[0], 'FOUND'
                cur_filtered = [r for r in (ind_filtered or doc_matches)
                                if ws_read.cell(r, anc_col_ogf).value != 1 and
                                   ws_read.cell(r, anc_col_ogf).value != '1'] if anc_col_ogf else []
                if len(cur_filtered) == 1:
                    return cur_filtered[0], 'FOUND'

    # ── Strategy 2: (numero, indice) unique match ──
    key = (num_clean, ind_clean)
    num_matches = num_idx.get(key, [])

    if len(num_matches) == 1:
        return num_matches[0], 'FOUND'

    if len(num_matches) > 1:
        # Prefer ANCIEN != 1 (current)
        cur = num_matches
        if anc_col_ogf:
            cur = [r for r in num_matches
                   if ws_read.cell(r, anc_col_ogf).value != 1 and
                      ws_read.cell(r, anc_col_ogf).value != '1']
            if len(cur) == 1:
                return cur[0], 'FOUND'

        # ── Strategy 3: closest Date réception (or Date diffusion as fallback) ──
        # Work on the narrowed set (current rows); fall back to all matches if empty
        candidates = cur if cur else num_matches

        # Reference date: prefer clean GF reception → fall back to date_diff
        rec_col_clean  = cols_clean.get('reception')
        date_col_clean = cols_clean.get('date_diff')
        rec_col_ogf    = cols_ogf.get('reception')
        date_col_ogf   = cols_ogf.get('date_diff')

        clean_ref = None
        if rec_col_clean:
            clean_ref = ws_clean.cell(r_clean, rec_col_clean).value
        if (clean_ref is None or str(clean_ref).strip() in ('', 'None', 'nan')) and date_col_clean:
            clean_ref = ws_clean.cell(r_clean, date_col_clean).value

        # OGF comparison column: prefer reception → fall back to date_diff
        # Try reception first; if all candidates have None there, try date_diff
        for ogf_date_col in filter(None, [rec_col_ogf, date_col_ogf]):
            best_row, unique = closest_date_match(clean_ref, candidates, ws_read, ogf_date_col)
            if unique:
                return best_row, 'FOUND'

        # ── Strategy 4: closest TITRE (fuzzy similarity) ──
        titre_col_clean = cols_clean.get('titre')
        titre_col_ogf   = cols_ogf.get('titre')
        if titre_col_clean and titre_col_ogf:
            clean_titre = ws_clean.cell(r_clean, titre_col_clean).value
            if clean_titre and str(clean_titre).strip() not in ('', 'None', 'nan'):
                scores = [
                    (titre_similarity(clean_titre, ws_read.cell(r, titre_col_ogf).value), r)
                    for r in candidates
                ]
                max_score = max(s for s, _ in scores) if scores else 0.0
                best = [r for s, r in scores if s == max_score]

                if len(best) == 1 and max_score > 0:
                    # Unique best TITRE match
                    return best[0], 'FOUND'

                # ── Strategy 5: all candidates share the same TITRE → OGF duplicates ──
                if len(best) == len(candidates):
                    ogf_titres = set(
                        str(ws_read.cell(r, titre_col_ogf).value or '').strip().upper()
                        for r in candidates
                    )
                    if len(ogf_titres) == 1:
                        # True duplicates: identical num, ind, date, titre → update ALL
                        return list(candidates), 'DUPLICATE_ALL'

        return None, 'SUSPICIOUS'

    # ── Not found ──
    if anc_clean_str == '1':
        return None, 'NOT_FOUND_OLD'
    return None, 'NOT_FOUND_NEW'


# ─── OBSERVATION MERGING ────────────────────────────────────────────────────

CONSULTANT_PATTERNS = [
    r'^([A-Z][A-Z\s\-/&\.0-9]+?)[\s]*[:.]',
    r'\n([A-Z][A-Z\s\-/&\.0-9]+?)[\s]*[:.]\s',
]

def parse_consultants_in_obs(obs_text):
    """Return set of consultant name fragments found in observation text."""
    if not obs_text:
        return set()
    found = set()
    # Match lines starting with consultant name followed by : or .
    for line in obs_text.split('\n'):
        line = line.strip()
        m = re.match(r'^([A-Z][A-Z\s\-/&\.0-9]+?)\s*[:.]', line)
        if m:
            found.add(m.group(1).strip().upper())
    return found


def extract_consultant_blocks(obs_text):
    """Split observation text into per-consultant blocks."""
    if not obs_text:
        return []
    blocks = []
    current_key = None
    current_lines = []
    for line in obs_text.split('\n'):
        m = re.match(r'^([A-Z][A-Z\s\-/&\.0-9]+?)\s*[:.]\s*(.*)', line.strip())
        if m:
            if current_key:
                blocks.append((current_key, '\n'.join(current_lines)))
            current_key = m.group(1).strip()
            current_lines = [line.strip()]
        else:
            if line.strip():
                current_lines.append(line.strip())
    if current_key:
        blocks.append((current_key, '\n'.join(current_lines)))
    return blocks


def merge_observations(ogf_obs, clean_obs):
    """
    Merge observations: append consultant blocks from clean_obs
    that are not already present in ogf_obs.
    Returns merged text.
    """
    if not clean_obs or str(clean_obs).strip() in ('', 'nan', 'None'):
        return ogf_obs

    clean_obs_str = str(clean_obs).strip()
    if not ogf_obs or str(ogf_obs).strip() in ('', 'nan', 'None'):
        return clean_obs_str

    ogf_obs_str = str(ogf_obs).strip()
    ogf_upper = ogf_obs_str.upper()

    # Extract blocks from clean obs
    clean_blocks = extract_consultant_blocks(clean_obs_str)
    append_parts = []

    for consultant_name, block_text in clean_blocks:
        # Check if this consultant is already mentioned in ogf
        consultant_upper = consultant_name.upper()
        # Check various short forms
        already = False
        for fragment in [consultant_upper, consultant_upper.split()[0]]:
            if fragment in ogf_upper:
                already = True
                break
        if not already:
            append_parts.append(block_text)

    if append_parts:
        return ogf_obs_str + '\n' + '\n'.join(append_parts)
    return ogf_obs_str


# ─── ROW COPY UTILITIES ─────────────────────────────────────────────────────

def copy_cell_value_only(src_cell, dst_cell):
    """Copy only the value from src to dst."""
    dst_cell.value = src_cell.value


def copy_cell_format(src_cell, dst_cell):
    """Copy formatting from src cell to dst cell."""
    if src_cell.has_style:
        dst_cell.font = copy_obj(src_cell.font)
        dst_cell.fill = copy_obj(src_cell.fill)
        dst_cell.border = copy_obj(src_cell.border)
        dst_cell.alignment = copy_obj(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def copy_row_format_from_template(ws_write, template_row, target_row):
    """Copy row formatting from template_row to target_row."""
    for c in range(1, ws_write.max_column + 1):
        tmpl = ws_write.cell(template_row, c)
        tgt = ws_write.cell(target_row, c)
        if tmpl.has_style:
            tgt.font = copy_obj(tmpl.font)
            tgt.fill = copy_obj(tmpl.fill)
            tgt.border = copy_obj(tmpl.border)
            tgt.alignment = copy_obj(tmpl.alignment)
            tgt.number_format = tmpl.number_format


def apply_row_fill(ws_write, row_num, max_col, rgb_color, except_col=None):
    """Apply fill color to entire row."""
    fill = PatternFill('solid', fgColor=rgb_color)
    for c in range(1, max_col + 1):
        if except_col and c == except_col:
            continue
        ws_write.cell(row_num, c).fill = fill


def to_date_only(val):
    """
    Strip the time component from a datetime value.
    Returns a datetime at midnight (openpyxl requires datetime, not date).
    Non-datetime values are returned unchanged.
    """
    from datetime import datetime, date as date_type
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, date_type):       # date but not datetime
        return datetime(val.year, val.month, val.day)
    return val


def get_date_col_formats(ws, cols_ogf):
    """
    Read the number_format for each date column from the first populated
    data row (row 10+).  Returns {col_number: format_string}.
    """
    from datetime import datetime
    formats = {}
    for key in DATE_COLUMN_KEYS:
        col = cols_ogf.get(key)
        if not col:
            continue
        # Find first row in that column that has a value
        fmt = None
        for r in range(10, min(ws.max_row + 1, 20)):
            cell = ws.cell(r, col)
            if cell.number_format and cell.number_format != 'General':
                fmt = cell.number_format
                break
        formats[col] = fmt if fmt else DATE_FORMAT_FALLBACK
    return formats


def copy_row_borders(ws_write, source_row, target_row):
    """
    Copy border formatting from source_row to target_row within ws_write.
    Skips columns where the source cell has no border set.
    """
    max_col = ws_write.max_column
    for c in range(1, max_col + 1):
        src_border = ws_write.cell(source_row, c).border
        if src_border:
            ws_write.cell(target_row, c).border = copy_obj(src_border)


# ─── MAIN PROCESSING ────────────────────────────────────────────────────────

def process_sheet(ws_read, ws_clean, ws_write, sheet_name, report):
    """
    Process one sheet: match rows, update, insert.
    ws_read  : OGF data-only (for reading values)
    ws_clean : clean GF data-only
    ws_write : output workbook sheet (for writing)
    """
    cols_ogf   = detect_cols(ws_read)
    cols_clean = detect_cols(ws_clean)

    num_col_ogf  = cols_ogf.get('numero')
    ind_col_ogf  = cols_ogf.get('indice')
    anc_col_ogf  = cols_ogf.get('ancien')
    vis_col_ogf  = cols_ogf.get('visa_global')
    obs_col_ogf  = cols_ogf.get('obs')

    num_col_clean = cols_clean.get('numero')
    ind_col_clean = cols_clean.get('indice')
    anc_col_clean = cols_clean.get('ancien')
    obs_col_clean = cols_clean.get('obs')

    if not num_col_ogf or not ind_col_ogf:
        print(f"  [SKIP] {sheet_name}: cannot detect key columns in OGF")
        report['skipped_sheets'].append(sheet_name)
        return

    if not num_col_clean or not ind_col_clean:
        print(f"  [SKIP] {sheet_name}: cannot detect key columns in clean GF")
        report['skipped_sheets'].append(sheet_name)
        return

    # Build OGF row index
    doc_idx, num_idx = build_ogf_index(ws_read, cols_ogf)

    # Build column mapping for VALUE cells (excluding obs)
    # Map: for each column name in clean GF, find OGF column
    col_map = {}  # clean_col -> ogf_col
    # Fixed columns
    fixed_map = [
        ('document', 'document'), ('titre', 'titre'), ('date_diff', 'date_diff'),
        ('lot', 'lot'), ('type_doc', 'type_doc'), ('numero', 'numero'),
        ('indice', 'indice'), ('niv', 'niv'), ('type', 'type'),
        ('ancien', 'ancien'), ('bdx', 'bdx'), ('reception', 'reception'),
        ('non_recu', 'non_recu'), ('date_contract', 'date_contract'),
        ('visa_global', 'visa_global'),
    ]
    for clean_key, ogf_key in fixed_map:
        if clean_key in cols_clean and ogf_key in cols_ogf:
            col_map[cols_clean[clean_key]] = cols_ogf[ogf_key]

    # Consultant columns mapping by name
    for cons_name, clean_sub in cols_clean.get('consultants', {}).items():
        # Find matching consultant in OGF (by name similarity)
        ogf_sub = None
        for ogf_cons_name, sub in cols_ogf.get('consultants', {}).items():
            if cons_name.upper() == ogf_cons_name.upper():
                ogf_sub = sub
                break
            # fuzzy: one name contains the other
            if (cons_name.upper() in ogf_cons_name.upper() or
                    ogf_cons_name.upper() in cons_name.upper()):
                ogf_sub = sub
                break
        if ogf_sub:
            for sub_key in ('date_col', 'num_col', 'stat_col'):
                if sub_key in clean_sub and sub_key in ogf_sub:
                    col_map[clean_sub[sub_key]] = ogf_sub[sub_key]

    # ── PASS 1: Match rows from clean GF to OGF ──
    rows_to_insert = []
    rows_found     = 0
    rows_updated   = 0
    rows_suspicious = 0

    for r_clean in range(10, ws_clean.max_row + 1):
        num_clean = normalize_numero(ws_clean.cell(r_clean, num_col_clean).value)
        if not num_clean:
            continue

        ind_clean = normalize_indice(ws_clean.cell(r_clean, ind_col_clean).value)
        r_ogf, tag = find_ogf_row(ws_read, r_clean, ws_clean, cols_clean, cols_ogf, doc_idx, num_idx)

        if tag == 'FOUND':
            rows_found += 1
            updated = apply_cell_updates(
                ws_clean, r_clean, cols_clean,
                ws_read, ws_write, r_ogf, cols_ogf,
                col_map, obs_col_clean, obs_col_ogf
            )
            if updated:
                rows_updated += 1
                if len(report['examples_updated']) < 5:
                    report['examples_updated'].append({
                        'sheet': sheet_name, 'numero': num_clean,
                        'indice': ind_clean, 'ogf_row': r_ogf
                    })

        elif tag == 'DUPLICATE_ALL':
            # r_ogf is a list of true-duplicate OGF rows — update every one
            rows_found += 1
            any_updated = False
            for r_dup in r_ogf:
                updated = apply_cell_updates(
                    ws_clean, r_clean, cols_clean,
                    ws_read, ws_write, r_dup, cols_ogf,
                    col_map, obs_col_clean, obs_col_ogf
                )
                if updated:
                    any_updated = True
            if any_updated:
                rows_updated += 1
            report['duplicates_resolved'] = report.get('duplicates_resolved', 0) + 1

        elif tag == 'SUSPICIOUS':
            # New rule: if clean GF VISA GLOBAL = 'SAS REF' → promote to new insert
            vis_col_clean = cols_clean.get('visa_global')
            visa_val_clean = ws_clean.cell(r_clean, vis_col_clean).value if vis_col_clean else None
            if visa_val_clean and str(visa_val_clean).strip().upper() == 'SAS REF':
                rows_to_insert.append(r_clean)
                report['sas_ref_promoted'] = report.get('sas_ref_promoted', 0) + 1
            else:
                rows_suspicious += 1
                # Collect rich details for the report
                titre_col  = cols_clean.get('titre')
                date_col   = cols_clean.get('date_diff')
                doc_col_c  = cols_clean.get('document', 1)
                titre_val  = ws_clean.cell(r_clean, titre_col).value  if titre_col  else None
                date_val   = ws_clean.cell(r_clean, date_col).value   if date_col   else None
                doc_val    = ws_clean.cell(r_clean, doc_col_c).value

                # Determine OGF match details
                num_key    = (num_clean, ind_clean)
                ogf_match_rows = num_idx.get(num_key, [])
                nb_matchs  = len(ogf_match_rows)
                doc_col_ogf = cols_ogf.get('document', 1)
                detail_parts = []
                for mr in ogf_match_rows[:8]:  # cap at 8 for display
                    anc_v = ws_read.cell(mr, anc_col_ogf).value if anc_col_ogf else ''
                    d_v   = ws_read.cell(mr, doc_col_ogf).value
                    detail_parts.append(f"row{mr}|doc={d_v}|ANCIEN={anc_v}")
                detail_str = ' // '.join(detail_parts)

                report['suspicious'].append({
                    'sheet':   sheet_name,
                    'numero':  num_clean,
                    'indice':  ind_clean,
                    'titre':   str(titre_val).strip() if titre_val else '',
                    'date':    date_val,
                    'document': str(doc_val).strip() if doc_val else '',
                    'nb_matchs': nb_matchs,
                    'detail':  detail_str,
                })
        elif tag == 'NOT_FOUND_NEW':
            rows_to_insert.append(r_clean)
        elif tag == 'NOT_FOUND_OLD':
            report['skipped_old'] += 1

    # ── PASS 2: Insert new rows ──
    rows_inserted_a    = 0
    rows_inserted_chain = 0
    rows_inserted_fallback = 0

    # Pre-compute the last populated data row for border template use in
    # Case A end-of-sheet insertions (avoids scanning 935+ empty rows per insert)
    num_col_check = num_col_ogf or 1
    last_data_row = 10  # safe fallback
    for r in range(ws_write.max_row, 9, -1):
        if ws_write.cell(r, num_col_check).value not in (None, ''):
            last_data_row = r
            break

    # Process insertions (in order)
    for r_clean in rows_to_insert:
        num_clean = normalize_numero(ws_clean.cell(r_clean, num_col_clean).value)
        ind_clean = normalize_indice(ws_clean.cell(r_clean, ind_col_clean).value)

        if ind_clean == 'A':
            # CASE A: insert at end of sheet
            insert_row = ws_write.max_row + 1
            write_new_row(ws_clean, r_clean, cols_clean,
                          ws_write, insert_row, cols_ogf, col_map,
                          anc_override='DI', border_template_row=last_data_row)
            apply_new_row_formatting(ws_write, insert_row, cols_ogf)
            last_data_row = insert_row   # next Case A uses this row as template
            rows_inserted_a += 1
            report['examples_inserted_a'].append({
                'sheet': sheet_name, 'numero': num_clean, 'indice': ind_clean, 'at_row': insert_row
            })
        else:
            # CASE B: find document chain
            # Find all rows with same numero in write sheet
            chain_rows = []
            for r in range(10, ws_write.max_row + 1):
                n = normalize_numero(ws_write.cell(r, num_col_ogf).value)
                if n == num_clean:
                    chain_rows.append(r)

            if chain_rows:
                # Find latest row (ANCIEN empty or 'DI')
                latest_row = None
                for cr in reversed(chain_rows):
                    anc_val = ws_write.cell(cr, anc_col_ogf).value if anc_col_ogf else None
                    anc_s = str(anc_val).strip() if anc_val is not None else ''
                    if anc_s in ('', 'None', 'DI'):
                        latest_row = cr
                        break

                if latest_row:
                    # Insert after latest_row
                    insert_pos = latest_row + 1
                    ws_write.insert_rows(insert_pos)
                    # Set previous latest → ANCIEN = 1
                    if anc_col_ogf:
                        ws_write.cell(latest_row, anc_col_ogf).value = 1
                    # Write new row — border from latest_row (always a data row)
                    write_new_row(ws_clean, r_clean, cols_clean,
                                  ws_write, insert_pos, cols_ogf, col_map,
                                  anc_override='DI', border_template_row=latest_row)
                    apply_new_row_formatting(ws_write, insert_pos, cols_ogf)
                    rows_inserted_chain += 1
                    report['examples_inserted_chain'].append({
                        'sheet': sheet_name, 'numero': num_clean, 'indice': ind_clean,
                        'after_row': latest_row, 'insert_pos': insert_pos
                    })
                else:
                    # Chain found but no latest → fallback: end of sheet
                    insert_row = ws_write.max_row + 1
                    write_new_row(ws_clean, r_clean, cols_clean,
                                  ws_write, insert_row, cols_ogf, col_map,
                                  anc_override='DI', border_template_row=last_data_row)
                    apply_new_row_formatting(ws_write, insert_row, cols_ogf)
                    last_data_row = insert_row
                    rows_inserted_fallback += 1
                    report['examples_inserted_fallback'].append({
                        'sheet': sheet_name, 'numero': num_clean, 'indice': ind_clean, 'at_row': insert_row
                    })
            else:
                # No chain found → fallback: insert at end
                insert_row = ws_write.max_row + 1
                write_new_row(ws_clean, r_clean, cols_clean,
                              ws_write, insert_row, cols_ogf, col_map,
                              anc_override='DI', border_template_row=last_data_row)
                apply_new_row_formatting(ws_write, insert_row, cols_ogf)
                last_data_row = insert_row
                rows_inserted_fallback += 1
                report['examples_inserted_fallback'].append({
                    'sheet': sheet_name, 'numero': num_clean, 'indice': ind_clean, 'at_row': insert_row
                })

    # ── PASS 3: Remove empty rows ──
    removed = remove_empty_rows(ws_write)

    # ── PASS 4: Re-apply full formatting (overrides OGF colours) ──
    apply_full_sheet_formatting(ws_write, cols_ogf)

    # Summary
    total_ins = rows_inserted_a + rows_inserted_chain + rows_inserted_fallback
    sas_promoted = report.get('sas_ref_promoted', 0)
    print(f"  {sheet_name}: found={rows_found}, updated={rows_updated}, "
          f"inserted={total_ins}(A:{rows_inserted_a}/chain:{rows_inserted_chain}/fb:{rows_inserted_fallback}), "
          f"suspicious={rows_suspicious}, removed_empty={removed}, sas_ref_promoted(running)={sas_promoted}")

    report['total_matched'] += rows_found
    report['total_updated'] += rows_updated
    report['total_inserted'] += total_ins
    report['total_suspicious'] += rows_suspicious


def apply_cell_updates(ws_clean, r_clean, cols_clean,
                       ws_read, ws_write, r_ogf, cols_ogf,
                       col_map, obs_col_clean, obs_col_ogf):
    """Apply cell updates to a FOUND row. Returns True if any update was made."""
    updated = False

    # Build date-column map once for this call
    date_fmt_map = get_date_col_formats(ws_write, cols_ogf)  # {ogf_col: fmt_str}

    for clean_col, ogf_col in col_map.items():
        ogf_val = ws_read.cell(r_ogf, ogf_col).value
        clean_val = ws_clean.cell(r_clean, clean_col).value

        # Skip OBSERVATIONS (handled separately)
        if obs_col_ogf and ogf_col == obs_col_ogf:
            continue

        if (ogf_val is None or str(ogf_val).strip() in ('', 'nan', 'None')) \
                and clean_val is not None \
                and str(clean_val).strip() not in ('', 'nan', 'None'):
            # Strip time from date values; preserve OGF number format
            if ogf_col in date_fmt_map:
                clean_val = to_date_only(clean_val)
            ws_write.cell(r_ogf, ogf_col).value = clean_val
            if ogf_col in date_fmt_map:
                ws_write.cell(r_ogf, ogf_col).number_format = date_fmt_map[ogf_col]
            updated = True

    # Handle OBSERVATIONS separately
    if obs_col_clean and obs_col_ogf:
        ogf_obs = ws_read.cell(r_ogf, obs_col_ogf).value
        clean_obs = ws_clean.cell(r_clean, obs_col_clean).value
        merged = merge_observations(ogf_obs, clean_obs)
        if merged != ogf_obs:
            ws_write.cell(r_ogf, obs_col_ogf).value = merged
            updated = True

    return updated


def write_new_row(ws_clean, r_clean, cols_clean,
                  ws_write, target_row, cols_ogf, col_map,
                  anc_override=None, border_template_row=None):
    """Write a new row from clean GF to write sheet at target_row."""
    obs_col_clean = cols_clean.get('obs')
    obs_col_ogf   = cols_ogf.get('obs')
    anc_col_ogf   = cols_ogf.get('ancien')

    # ── Border: copy from border_template_row (caller supplies the nearest
    #    non-empty data row, avoiding the blank rows at the OGF sheet bottom) ──
    if border_template_row and border_template_row >= 10:
        copy_row_borders(ws_write, border_template_row, target_row)

    # ── Date column formats: read from reference row so we match OGF ──
    date_fmt_map = get_date_col_formats(ws_write, cols_ogf)  # {ogf_col: fmt_str}

    for clean_col, ogf_col in col_map.items():
        val = ws_clean.cell(r_clean, clean_col).value
        if val is not None and str(val).strip() not in ('', 'nan', 'None'):
            if ogf_col in date_fmt_map:
                val = to_date_only(val)
            ws_write.cell(target_row, ogf_col).value = val
            if ogf_col in date_fmt_map:
                ws_write.cell(target_row, ogf_col).number_format = date_fmt_map[ogf_col]

    # Write observations
    if obs_col_clean and obs_col_ogf:
        obs_val = ws_clean.cell(r_clean, obs_col_clean).value
        if obs_val and str(obs_val).strip() not in ('', 'nan', 'None'):
            ws_write.cell(target_row, obs_col_ogf).value = str(obs_val).strip()

    # Set ANCIEN override
    if anc_override and anc_col_ogf:
        ws_write.cell(target_row, anc_col_ogf).value = anc_override


def apply_new_row_formatting(ws_write, row_num, cols_ogf):
    """Apply formatting rules to a newly inserted row."""
    anc_col = cols_ogf.get('ancien')
    vis_col = cols_ogf.get('visa_global')
    max_col = ws_write.max_column

    anc_val = ws_write.cell(row_num, anc_col).value if anc_col else None
    vis_val = ws_write.cell(row_num, vis_col).value if vis_col else None

    anc_str = str(anc_val).strip() if anc_val is not None else ''
    vis_str = str(vis_val).strip() if vis_val is not None else ''

    if anc_str == '1':
        # RULE A: grey
        apply_row_fill(ws_write, row_num, max_col, GREY_COLOR, except_col=vis_col)
    elif vis_str in ('', 'None', 'nan'):
        # RULE B: white (no fill)
        pass
    else:
        # RULE C: color based on VISA GLOBAL
        color = STATUS_COLORS.get(vis_str, None)
        if color:
            apply_row_fill(ws_write, row_num, max_col, color, except_col=vis_col)
            # VISA GLOBAL cell gets its own color
            if vis_col:
                ws_write.cell(row_num, vis_col).fill = PatternFill('solid', fgColor=color)

    # Set alignment: wrap for obs, no-wrap for others
    obs_col = cols_ogf.get('obs')
    for c in range(1, max_col + 1):
        cell = ws_write.cell(row_num, c)
        if c == obs_col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        else:
            cell.alignment = Alignment(wrap_text=False, vertical='top')


def remove_empty_rows(ws_write):
    """
    Delete fully-empty data rows (row 10+) where every cell has no value.
    Scans in reverse to avoid index shifting.
    Returns number of rows removed.
    """
    max_col  = ws_write.max_column
    removed  = 0
    for r in range(ws_write.max_row, 9, -1):
        if all(ws_write.cell(r, c).value in (None, '') for c in range(1, max_col + 1)):
            ws_write.delete_rows(r)
            removed += 1
    return removed


def apply_full_sheet_formatting(ws_write, cols_ogf):
    """
    Apply our formatting rules to EVERY data row (10+), overriding whatever
    came from the OGF file.  Rules:
      • ANCIEN == 1  → grey row
      • VISA empty   → white row
      • VISA filled  → status colour row (matching STATUS_COLORS)
    Also ensures OBSERVATIONS column has wrap=True.
    """
    anc_col = cols_ogf.get('ancien')
    vis_col = cols_ogf.get('visa_global')
    obs_col = cols_ogf.get('obs')
    max_col = ws_write.max_column

    for row_num in range(10, ws_write.max_row + 1):
        # Skip completely empty rows (shouldn't exist after remove_empty_rows, but safe)
        if all(ws_write.cell(row_num, c).value in (None, '')
               for c in range(1, max_col + 1)):
            continue

        anc_val = ws_write.cell(row_num, anc_col).value if anc_col else None
        vis_val = ws_write.cell(row_num, vis_col).value if vis_col else None

        anc_str = str(anc_val).strip() if anc_val is not None else ''
        vis_str = str(vis_val).strip() if vis_val is not None else ''

        # Normalise ANCIEN: 1, 1.0, '1', '1.0' all mean "superseded"
        anc_is_old = anc_val in (1, '1') or (
            isinstance(anc_val, float) and anc_val == 1.0)

        if anc_is_old:
            apply_row_fill(ws_write, row_num, max_col, GREY_COLOR)
        elif vis_str in ('', 'None', 'nan', 'DI'):
            apply_row_fill(ws_write, row_num, max_col, WHITE_COLOR)
        else:
            color = STATUS_COLORS.get(vis_str)
            if color:
                apply_row_fill(ws_write, row_num, max_col, color)

        # OBSERVATIONS column: always wrap
        if obs_col:
            cell = ws_write.cell(row_num, obs_col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # ── Date format pass: strip time + apply per-column OGF format ──
    from datetime import datetime
    date_fmt_map = get_date_col_formats(ws_write, cols_ogf)  # {ogf_col: fmt_str}
    for row_num in range(10, ws_write.max_row + 1):
        for col, fmt in date_fmt_map.items():
            cell = ws_write.cell(row_num, col)
            if isinstance(cell.value, datetime):
                cell.value = to_date_only(cell.value)
                cell.number_format = fmt


def setup_page_layout(ws_write):
    """
    Configure print settings for a sheet:
      • Paper  : A3 landscape
      • Width  : fit to 1 page wide, unlimited height
      • Scale  : 65 % (above 55 % floor; acts as minimum when fitToPage=True)
      • Print area   : A1 → last populated row
      • Print titles : repeat rows 1:9
      • Auto-filter  : row 9, full column range
    """
    from openpyxl.utils import get_column_letter

    max_col = ws_write.max_column
    max_col_letter = get_column_letter(max_col)

    # Find last row that actually has data
    last_row = 9  # fallback: header only
    for r in range(ws_write.max_row, 9, -1):
        if any(ws_write.cell(r, c).value not in (None, '')
               for c in range(1, max_col + 1)):
            last_row = r
            break

    # ── Print area ──
    ws_write.print_area = f'A1:{max_col_letter}{last_row}'

    # ── Repeat header rows on every printed page ──
    ws_write.print_title_rows = '1:9'

    # ── Page setup: A3 landscape, fit to 1 page wide ──
    ws_write.sheet_properties.pageSetUpPr.fitToPage = True
    ws_write.page_setup.paperSize   = 8          # A3
    ws_write.page_setup.orientation = 'landscape'
    ws_write.page_setup.fitToWidth  = 1
    ws_write.page_setup.fitToHeight = 0          # unlimited pages tall
    ws_write.page_setup.scale       = 65         # floor: never print below 65 %

    # ── Auto-filter on row 9 ──
    ws_write.auto_filter.ref = f'A9:{max_col_letter}{last_row}'


# ─── SUSPICIOUS REPORT GENERATOR ───────────────────────────────────────────

SUSPICIOUS_REPORT_PATH = "output/SUSPICIOUS_ROWS_REPORT.xlsx"

def generate_suspicious_report(report):
    """
    Write output/SUSPICIOUS_ROWS_REPORT.xlsx from report['suspicious'].
    Columns: #, SHEET, N°DOC, IND, TITRE, DATE DIFFUSION, DOCUMENT (Clean GF),
             NB MATCHS OGF, DÉTAIL DES MATCHS OGF
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SUSPICIOUS"

    # ── Header ──
    headers = [
        '#', 'SHEET', 'N°DOC', 'IND', 'TITRE',
        'DATE DIFFUSION', 'DOCUMENT (Clean GF)',
        'NB MATCHS OGF', 'DÉTAIL DES MATCHS OGF'
    ]
    header_fill   = PatternFill('solid', fgColor='FF1F4E79')
    header_font   = Font(bold=True, color='FFFFFFFF', size=10)
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin          = Side(style='thin', color='FF000000')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # ── Column widths ──
    col_widths = [5, 30, 12, 6, 50, 16, 45, 10, 80]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Alternating row fills ──
    fill_even = PatternFill('solid', fgColor='FFDAE3F3')
    fill_odd  = PatternFill('solid', fgColor='FFFFFFFF')
    data_font = Font(size=9)
    data_align_wrap = Alignment(vertical='top', wrap_text=True)
    data_align      = Alignment(vertical='top', wrap_text=False)

    rows = report.get('suspicious', [])
    for i, entry in enumerate(rows, 1):
        row_num = i + 1
        fill = fill_even if i % 2 == 0 else fill_odd

        values = [
            i,
            entry.get('sheet', ''),
            entry.get('numero', ''),
            entry.get('indice', ''),
            entry.get('titre', ''),
            entry.get('date', ''),
            entry.get('document', ''),
            entry.get('nb_matchs', ''),
            entry.get('detail', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row_num, col_idx, val)
            cell.fill   = fill
            cell.font   = data_font
            cell.border = border
            # Wrap only TITRE and DÉTAIL columns
            cell.alignment = data_align_wrap if col_idx in (5, 9) else data_align

        # Auto-height hint for long detail rows
        ws.row_dimensions[row_num].height = 15

    # ── Summary row ──
    summary_row = len(rows) + 2
    ws.cell(summary_row, 1, f"Total: {len(rows)} suspicious rows")
    ws.cell(summary_row, 1).font = Font(bold=True, italic=True, size=9)

    # ── Freeze top row ──
    ws.freeze_panes = 'A2'

    wb.save(SUSPICIOUS_REPORT_PATH)
    print(f"  Suspicious report saved → {SUSPICIOUS_REPORT_PATH} ({len(rows)} rows)")


# ─── ENTRY POINT ────────────────────────────────────────────────────────────

def build_team_version(ogf_path, clean_path, out_path):
    """
    Pipeline-callable entry point.
    Same logic as main() but with explicit paths instead of module constants.
    Returns the report dict.
    """
    import os
    import sys
    import tempfile
    import warnings

    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    print("=== GF TEAM VERSION BUILDER ===")
    print(f"OGF:   {ogf_path}")
    print(f"Clean: {clean_path}")
    print(f"Out:   {out_path}")
    print()

    out_dir = os.path.dirname(os.path.abspath(out_path)) or "."
    with tempfile.NamedTemporaryFile(dir=out_dir, suffix=".xlsx", delete=False) as work_tmp:
        work_path = work_tmp.name
    shutil.copy2(ogf_path, work_path)
    print(f"Copied OGF to temporary workbook {work_path}")

    print("Loading workbooks...")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb_read = load_workbook(ogf_path, data_only=True)
        wb_clean = load_workbook(clean_path, data_only=True)
        wb_write = load_workbook(work_path)

    report = {
        'total_matched': 0,
        'total_updated': 0,
        'total_inserted': 0,
        'total_suspicious': 0,
        'skipped_old': 0,
        'skipped_sheets': [],
        'suspicious': [],
        'examples_updated': [],
        'examples_inserted_a': [],
        'examples_inserted_chain': [],
        'examples_inserted_fallback': [],
    }

    for sheet_name in wb_clean.sheetnames:
        if sheet_name not in wb_read.sheetnames:
            print(f"  [SKIP] {sheet_name}: not found in OGF")
            continue
        if sheet_name not in wb_write.sheetnames:
            print(f"  [SKIP] {sheet_name}: not found in output wb")
            continue
        ws_read = wb_read[sheet_name]
        ws_clean = wb_clean[sheet_name]
        ws_write = wb_write[sheet_name]
        process_sheet(ws_read, ws_clean, ws_write, sheet_name, report)

    print("\nSetting up page layout...")
    for sheet_name in wb_write.sheetnames:
        try:
            ws = wb_write[sheet_name]
            setup_page_layout(ws)
        except Exception as e:
            print(f"  [WARN] Page layout for {sheet_name}: {e}")

    print(f"\nSaving to {out_path}...")
    with tempfile.NamedTemporaryFile(dir=out_dir, suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        wb_write.save(tmp_path)
    finally:
        for wb in (wb_read, wb_clean, wb_write):
            try:
                wb.close()
            except Exception:
                pass
    try:
        if os.path.exists(out_path):
            os.chmod(out_path, 0o666)
    except Exception:
        pass
    os.replace(tmp_path, out_path)
    try:
        os.unlink(work_path)
    except Exception:
        pass
    print("Saved.")

    generate_suspicious_report(report)

    print(f"\nTeam version: matched={report['total_matched']}, "
          f"updated={report['total_updated']}, inserted={report['total_inserted']}, "
          f"suspicious={report['total_suspicious']}")
    return report


def main():
    print("=== GF TEAM VERSION BUILDER ===")
    print(f"OGF:   {OGF_PATH}")
    print(f"Clean: {CLEAN_PATH}")
    print(f"Out:   {OUT_PATH}")
    print()

    # Copy OGF → output
    shutil.copy(OGF_PATH, OUT_PATH)
    print(f"Copied OGF to {OUT_PATH}")

    # Load workbooks
    print("Loading workbooks...")
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb_read  = load_workbook(OGF_PATH, data_only=True)   # OGF values
        wb_clean = load_workbook(CLEAN_PATH, data_only=True)  # Clean GF values
        wb_write = load_workbook(OUT_PATH)                    # Output (preserves formulas)

    report = {
        'total_matched':   0,
        'total_updated':   0,
        'total_inserted':  0,
        'total_suspicious': 0,
        'skipped_old':     0,
        'skipped_sheets':  [],
        'suspicious':      [],
        'examples_updated':         [],
        'examples_inserted_a':      [],
        'examples_inserted_chain':  [],
        'examples_inserted_fallback': [],
    }

    # Process each sheet in clean GF
    for sheet_name in wb_clean.sheetnames:
        if sheet_name not in wb_read.sheetnames:
            print(f"  [SKIP] {sheet_name}: not found in OGF")
            continue
        if sheet_name not in wb_write.sheetnames:
            print(f"  [SKIP] {sheet_name}: not found in output wb")
            continue

        ws_read  = wb_read[sheet_name]
        ws_clean = wb_clean[sheet_name]
        ws_write = wb_write[sheet_name]

        process_sheet(ws_read, ws_clean, ws_write, sheet_name, report)

    # Setup page layout for all sheets (A3 landscape, fit-to-width, scale, autofilter)
    print("\nSetting up page layout...")
    for sheet_name in wb_write.sheetnames:
        try:
            ws = wb_write[sheet_name]
            setup_page_layout(ws)
            print(f"  {sheet_name}: print area A1:{ws.print_area.split(':')[1] if ws.print_area and ':' in ws.print_area else '?'}, "
                  f"filter={ws.auto_filter.ref}")
        except Exception as e:
            print(f"  [WARN] Page layout for {sheet_name}: {e}")

    # Save output
    print(f"\nSaving to {OUT_PATH}...")
    wb_write.save(OUT_PATH)
    print("Saved.")

    # ── GENERATE SUSPICIOUS REPORT ──
    print("\nGenerating suspicious rows report...")
    generate_suspicious_report(report)

    # ── VALIDATION REPORT ──
    print("\n" + "="*60)
    print("VALIDATION REPORT")
    print("="*60)
    print(f"Total rows matched (found in OGF):  {report['total_matched']}")
    print(f"  of which duplicates (S5):          {report.get('duplicates_resolved', 0)}")
    print(f"Total rows updated:                  {report['total_updated']}")
    print(f"Total rows inserted:                 {report['total_inserted']}")
    print(f"  of which SAS REF promoted:         {report.get('sas_ref_promoted', 0)}")
    print(f"Total suspicious (skipped):          {report['total_suspicious']}")
    print(f"Total old rows skipped:              {report['skipped_old']}")
    print(f"Sheets skipped:                      {report['skipped_sheets']}")
    print()

    print("── EXAMPLES: CELL UPDATED ──")
    for ex in report['examples_updated'][:3]:
        print(f"  Sheet={ex['sheet']}, N°Doc={ex['numero']}, IND={ex['indice']}, OGF row={ex['ogf_row']}")

    print("\n── EXAMPLES: ROW INSERTED (CASE A - indice A) ──")
    for ex in report['examples_inserted_a'][:3]:
        print(f"  Sheet={ex['sheet']}, N°Doc={ex['numero']}, IND={ex['indice']}, at row={ex['at_row']}")

    print("\n── EXAMPLES: ROW INSERTED (CASE B - chain) ──")
    for ex in report['examples_inserted_chain'][:3]:
        print(f"  Sheet={ex['sheet']}, N°Doc={ex['numero']}, IND={ex['indice']}, after={ex['after_row']}, pos={ex['insert_pos']}")

    print("\n── EXAMPLES: ROW INSERTED (FALLBACK) ──")
    for ex in report['examples_inserted_fallback'][:3]:
        print(f"  Sheet={ex['sheet']}, N°Doc={ex['numero']}, IND={ex['indice']}, at row={ex['at_row']}")

    print("\n── SUSPICIOUS MATCHES ──")
    for s in report['suspicious'][:5]:
        print(f"  Sheet={s['sheet']}, N°Doc={s['numero']}, IND={s['indice']}")

    return report


if __name__ == '__main__':
    main()
