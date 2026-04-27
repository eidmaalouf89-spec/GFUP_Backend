"""
src/pipeline/compute.py
-----------------------
Non-pure pipeline helpers extracted from main.py.
These functions perform file I/O (read workbooks, write xlsx, read filesystem).
"""

import re as _re
import datetime as _dt
from pathlib import Path

import pandas as pd

from pipeline.utils import _safe_console_print
from domain.normalization import (
    normalize_date_for_compare,
    normalize_title_for_compare,
    title_similarity,
    date_diff_days,
    normalize_numero_for_compare,
    normalize_indice_for_compare,
)
from domain.discrepancy import (
    classify_discrepancy,
    _is_excluded_sheet_for_discrepancy,
)
from domain.classification import (
    _classify_missing_in_gf,
    _classify_missing_in_ged,
)
from domain.sas_helpers import _build_sas_lookup
from domain.family_builder import _build_ged_families
from domain.gf_helpers import (
    _gf_row_stable_key,
    _sorted_family_doc_ids,
    _parse_gf_sheet_data,
)


def _determine_data_date(ged_file_path: str) -> _dt.date:
    """
    Returns the reference date for recency checks.
    Uses GED file mtime; falls back to today.
    """
    import os
    try:
        mtime = os.path.getmtime(str(ged_file_path))
        return _dt.datetime.fromtimestamp(mtime).date()
    except Exception:
        return _dt.date.today()


def _compute_discrepancies(
    dernier_df: pd.DataFrame,
    gf_filepath: str,
    debug_dir: str = None,
    excluded_sheets: set = None,
    sheet_year_filters: dict = None,
    responses_df: pd.DataFrame = None,   # Patch E: SAS data for GED docs
    data_date: _dt.date = None,          # Patch E: reference date for recency
) -> tuple:  # returns (discrepancies_list, gf_by_sheet_dict)
    """
    Compare GED dernier indices against existing GF data.

    Patch B: Uses normalize_date_for_compare / normalize_title_for_compare
             so cosmetic differences don't create false mismatches.
    Patch C: excluded_sheets and sheet_year_filters are passed in so GF rows
             from excluded populations are skipped entirely.

    Discrepancy types:
      MISSING_IN_GF      – GED dernier indice not present in GF at all
      MISSING_IN_GED     – GF has a row whose numero+indice is not in GED dernier
      INDICE_MISMATCH    – GF has this numero but with a different indice
      TITRE_MISMATCH     – numero+indice matched but title significantly differs
      DATE_MISMATCH      – numero+indice matched but date_diffusion differs
      SHEET_MISMATCH     – numero found in GF but on a different sheet than routed

    Returns list of discrepancy dicts (each has a 'severity' key after
    classify_discrepancy is applied in run_pipeline).
    """
    import openpyxl
    from routing import read_all_gf_sheet_structures

    excluded_sheets    = excluded_sheets    or set()
    sheet_year_filters = sheet_year_filters or {}

    # Patch E: Build SAS lookup once for all sheets
    sas_lookup: dict = {}
    if responses_df is not None:
        sas_lookup = _build_sas_lookup(responses_df)
        _safe_console_print(f"  [Patch E] SAS lookup built: {len(sas_lookup)} doc_ids with SAS data")
    if data_date is None:
        data_date = _dt.date.today()

    discrepancies = []

    try:
        wb = openpyxl.load_workbook(gf_filepath, read_only=True, data_only=True)
    except Exception as e:
        _safe_console_print(f"  WARNING: Could not read existing GF for discrepancy check: {e}")
        return discrepancies, {}

    sheet_names = [s for s in wb.sheetnames if not s.upper().startswith("OLD")]

    # Remove fully-excluded sheets (Patch C)
    sheet_names_for_compare = [
        s for s in sheet_names
        if not _is_excluded_sheet_for_discrepancy(s, excluded_sheets, sheet_year_filters)
    ]
    excluded_skipped = [s for s in sheet_names if s not in sheet_names_for_compare]
    if excluded_skipped:
        _safe_console_print(f"  [Patch C] Skipping {len(excluded_skipped)} excluded sheets: {excluded_skipped}")

    structures = read_all_gf_sheet_structures(gf_filepath, sheet_names_for_compare)

    if debug_dir:
        _write_gf_schema_debug(debug_dir, structures)

    # ── Build GF indexes ──────────────────────────────────────────────────────
    # gf_by_sheet[sheet] = {(num, ind): [row, row, ...]}   ← LIST per key (Part 1)
    # full_gf_index[num] = [(sheet, ind, row), ...]
    gf_by_sheet: dict = {}
    full_gf_index: dict = {}
    # For gf_duplicates debug: (sheet, num, ind) → count
    gf_dup_counts: dict = {}

    for sheet_name in sheet_names_for_compare:
        ws = wb[sheet_name]
        struct = structures.get(sheet_name, {})
        gf_docs = _parse_gf_sheet_data(ws, struct)  # → dict[(num,ind)] → [rows]
        gf_by_sheet[sheet_name] = gf_docs

        for (num_c, ind_c), rows in gf_docs.items():
            for row in rows:
                full_gf_index.setdefault(num_c, [])
                full_gf_index[num_c].append((sheet_name, ind_c, row))
            if len(rows) > 1:
                gf_dup_counts[(sheet_name, num_c, ind_c)] = len(rows)

    total_dup_keys = len(gf_dup_counts)
    total_dup_excess = sum(v - 1 for v in gf_dup_counts.values())
    if total_dup_keys:
        _safe_console_print(f"  [Part 1] GF duplicate keys on active sheets: "
              f"{total_dup_keys} keys, {total_dup_excess} excess rows")

    # Write gf_duplicates.xlsx debug artifact
    if debug_dir:
        _write_gf_duplicates_debug(debug_dir, gf_dup_counts)

    # ── Per-sheet comparison ─────────────────────────────────────────────────
    for sheet_name in sheet_names_for_compare:
        gf_docs = gf_by_sheet.get(sheet_name, {})
        ged_sheet_docs = dernier_df[dernier_df["gf_sheet_name"] == sheet_name]
        min_year = sheet_year_filters.get(sheet_name)

        # ── Part 2: Build GED families ──────────────────────────────────────
        # Group GED rows into logical document families so that multiple GED
        # submissions for the same document don't generate N discrepancies.
        # Pass GF numeros so Pass 2 merge skips numeros that have a GF match
        # (those rows must remain independent to match their own GF row).
        gf_numeros_in_sheet = {k[0] for k in gf_docs.keys()}
        ged_families = _build_ged_families(ged_sheet_docs, gf_numeros_in_sheet)

        # Collect RAW GED numeros directly from original rows (NOT from families)
        # This is used for the MISSING_IN_GED_HISTORICAL check:
        # "GED has this numero at some indice → old GF row is historical"
        # Using family["all_numeros"] would contaminate this with merged numeros
        # from different logical documents.
        all_ged_numeros_in_sheet: set = set()
        for _, raw_row in ged_sheet_docs.iterrows():
            n = normalize_numero_for_compare(
                raw_row.get("numero_normalized") or raw_row.get("numero")
            )
            if n:
                all_ged_numeros_in_sheet.add(n)

        # Track which GF keys have been matched (exact or fuzzy) — used to
        # suppress MISSING_IN_GED for GF rows that ARE covered by a GED family.
        matched_gf_keys: set = set()

        # Track GF keys where DUPLICATE_ACTIVE_IN_GF has already been emitted
        dup_flagged_gf_keys: set = set()

        for family in ged_families:
            numero    = family["numero"]
            indice    = family["indice"]
            ged_titre_raw = family["title"]
            ged_date_raw  = family["date"]
            doc_code  = family["doc_code"]
            all_fam_keys = sorted(family["all_keys"])  # all (num, ind) in this family

            # ── Step 1: Exact-key match: try every (num, ind) in the family ──
            exact_gf_key = None
            exact_cands  = None
            # CRITICAL: initialize to -1.0, not 0.0.
            # When GED libelle is a raw doc-code (e.g. "P17_T2_..._130027_A.pdf") and the
            # GF row has a meaningful title ("Platelage bois MOSO BAMBOO"), title_similarity
            # returns 0.0.  Using 0.0 as the floor means "0.0 > 0.0" is False and the exact
            # key match is never registered, causing a false MISSING_IN_GF_TRUE.
            # Initializing to -1.0 ensures ANY exact (numero, indice) match is accepted.
            exact_best_sim = -1.0

            for fam_key in all_fam_keys:
                if fam_key in gf_docs:
                    cands = gf_docs[fam_key]
                    for cand in cands:
                        sim = title_similarity(ged_titre_raw, cand.get("titre", ""))
                        if sim > exact_best_sim:
                            exact_best_sim = sim
                            exact_gf_key   = fam_key
                            exact_cands    = cands

            if exact_gf_key is not None:
                # ── Exact key match: full comparison ──────────────────────
                matched_gf_keys.add(exact_gf_key)
                scored = sorted(
                    [(title_similarity(ged_titre_raw, c.get("titre", "")), c)
                     for c in exact_cands],
                    key=lambda x: (-x[0], _gf_row_stable_key(x[1])),
                )
                best_sim, best_cand = scored[0]

                # Flag structural duplicates (once per GF key)
                if exact_gf_key not in dup_flagged_gf_keys and len(exact_cands) > 1:
                    uniq_titles = set(
                        normalize_title_for_compare(c.get("titre", ""))
                        for c in exact_cands
                    )
                    if len(uniq_titles) > 1:
                        dup_flagged_gf_keys.add(exact_gf_key)
                        discrepancies.append({
                            "sheet_name": sheet_name,
                            "document_code": doc_code,
                            "numero": numero,
                            "indice": indice,
                            "field": "DUPLICATE",
                            "ged_value": ged_titre_raw[:80],
                            "gf_value": f"{len(exact_cands)} GF rows with same (num,ind)",
                            "ged_value_normalized": normalize_title_for_compare(ged_titre_raw),
                            "gf_value_normalized": str(uniq_titles),
                            "flag_type": "DUPLICATE_ACTIVE_IN_GF",
                            "title_similarity": best_sim,
                            "date_diff_days": None,
                            "is_excluded_population": False,
                        })

                # Title comparison
                gf_titre_raw = str(best_cand.get("titre", "") or "")
                if ged_titre_raw and gf_titre_raw and best_sim < 0.85:
                    discrepancies.append({
                        "sheet_name": sheet_name,
                        "document_code": doc_code,
                        "numero": numero,
                        "indice": indice,
                        "field": "TITRE",
                        "ged_value": ged_titre_raw[:100],
                        "gf_value": gf_titre_raw[:100],
                        "ged_value_normalized": normalize_title_for_compare(ged_titre_raw),
                        "gf_value_normalized": normalize_title_for_compare(gf_titre_raw),
                        "flag_type": "TITRE_MISMATCH",
                        "title_similarity": round(best_sim, 3),
                        "date_diff_days": None,
                        "is_excluded_population": False,
                    })

                # Date comparison
                gf_date_raw = best_cand.get("date_diffusion")
                if ged_date_raw is not None and gf_date_raw is not None:
                    diff = date_diff_days(ged_date_raw, gf_date_raw)
                    if diff is not None and diff > 0:
                        discrepancies.append({
                            "sheet_name": sheet_name,
                            "document_code": doc_code,
                            "numero": numero,
                            "indice": indice,
                            "field": "DATE_DIFFUSION",
                            "ged_value": str(ged_date_raw),
                            "gf_value": str(gf_date_raw),
                            "ged_value_normalized": normalize_date_for_compare(ged_date_raw),
                            "gf_value_normalized": normalize_date_for_compare(gf_date_raw),
                            "flag_type": "DATE_MISMATCH",
                            "title_similarity": None,
                            "date_diff_days": diff,
                            "is_excluded_population": False,
                        })

            else:
                # ── No exact key match: try cross-key title fuzzy match ────
                # If a GF row exists with title_similarity >= 0.75 to this
                # family's representative title, suppress MISSING_IN_GF.
                # We do NOT emit TITRE/DATE discrepancies for fuzzy matches
                # (different numero = different document reference).
                fuzzy_gf_key = None
                fuzzy_best_key = None
                fuzzy_best_sim = 0.0
                for gf_key, gf_cands in gf_docs.items():
                    for gf_cand in gf_cands:
                        sim = title_similarity(ged_titre_raw, gf_cand.get("titre", ""))
                        cand_key = (gf_key, _gf_row_stable_key(gf_cand))
                        if (
                            sim > fuzzy_best_sim or
                            (
                                fuzzy_best_key is not None and
                                sim == fuzzy_best_sim and
                                cand_key < fuzzy_best_key
                            ) or
                            (fuzzy_best_key is None and sim == fuzzy_best_sim and sim >= 0.75)
                        ):
                            fuzzy_best_sim = sim
                            fuzzy_best_key = cand_key
                            if sim >= 0.75:
                                fuzzy_gf_key = gf_key

                if fuzzy_gf_key is not None:
                    # Fuzzy match found → suppress MISSING_IN_GF, mark GF key covered
                    matched_gf_keys.add(fuzzy_gf_key)
                    # No TITRE or DATE discrepancies for cross-key fuzzy matches
                else:
                    # No match at all — check INDICE/SHEET/MISSING_IN_GF
                    if numero in full_gf_index:
                        all_occ = full_gf_index[numero]
                        same_sheet_other_ind = [
                            (sn, ind2, row) for sn, ind2, row in all_occ
                            if sn == sheet_name and ind2 != indice
                        ]
                        other_sheet_same_ind = [
                            (sn, ind2, row) for sn, ind2, row in all_occ
                            if sn != sheet_name and ind2 == indice
                        ]

                        if same_sheet_other_ind:
                            same_sheet_other_ind = sorted(
                                same_sheet_other_ind,
                                key=lambda x: (str(x[1]), _gf_row_stable_key(x[2])),
                            )
                            gf_ind = same_sheet_other_ind[0][1]
                            # Patch D: include date/title info so the relaxation
                            # pass can accept INDICE_MISMATCH when dates match
                            _gf_row_ind  = same_sheet_other_ind[0][2]
                            _gf_date_ind = _gf_row_ind.get("date_diffusion")
                            _ind_date_diff = date_diff_days(ged_date_raw, _gf_date_ind)
                            _ind_title_sim = title_similarity(
                                ged_titre_raw,
                                str(_gf_row_ind.get("titre") or ""),
                            )
                            discrepancies.append({
                                "sheet_name": sheet_name,
                                "document_code": doc_code,
                                "numero": numero,
                                "indice": indice,
                                "field": "INDICE",
                                "ged_value": indice,
                                "gf_value": gf_ind,
                                "ged_value_normalized": indice,
                                "gf_value_normalized": gf_ind,
                                "flag_type": "INDICE_MISMATCH",
                                "title_similarity": round(_ind_title_sim, 3),
                                "date_diff_days": _ind_date_diff,
                                "is_excluded_population": False,
                                # For Patch D reconstruction
                                "ged_date_raw": str(ged_date_raw or ""),
                                "gf_date_raw":  str(_gf_date_ind or ""),
                            })
                        elif other_sheet_same_ind:
                            other_sheet_same_ind = sorted(
                                other_sheet_same_ind,
                                key=lambda x: (str(x[0]), str(x[1]), _gf_row_stable_key(x[2])),
                            )
                            gf_sheet = other_sheet_same_ind[0][0]
                            discrepancies.append({
                                "sheet_name": sheet_name,
                                "document_code": doc_code,
                                "numero": numero,
                                "indice": indice,
                                "field": "SHEET",
                                "ged_value": sheet_name,
                                "gf_value": gf_sheet,
                                "ged_value_normalized": sheet_name,
                                "gf_value_normalized": gf_sheet,
                                "flag_type": "SHEET_MISMATCH",
                                "title_similarity": None,
                                "date_diff_days": None,
                                "is_excluded_population": False,
                            })
                        else:
                            # Patch E: Classify MISSING_IN_GF using SAS timing
                            mig_subtype, _ = _classify_missing_in_gf(
                                family, sas_lookup, data_date
                            )
                            sas_info = {}
                            for did in _sorted_family_doc_ids(family):
                                if did in sas_lookup:
                                    sas_info = sas_lookup[did]
                                    break
                            discrepancies.append({
                                "sheet_name":    sheet_name,
                                "document_code": doc_code,
                                "numero":        numero,
                                "indice":        indice,
                                "field":         "DOCUMENT",
                                "ged_value":     f"{numero}/{indice}",
                                "gf_value":      "NOT FOUND IN GF (numero on other sheet)",
                                "ged_value_normalized": "",
                                "gf_value_normalized":  "",
                                "flag_type":     mig_subtype,
                                "title_similarity": None,
                                "date_diff_days":   None,
                                "is_excluded_population": False,
                                # Patch E extras
                                "sas_status_type":  sas_info.get("sas_status_type"),
                                "sas_result":       sas_info.get("sas_result"),
                                "sas_date":         str(sas_info.get("sas_date") or ""),
                                # GF_INSERTION_QUEUE fields (Part 9)
                                "gfi_numero":  numero,
                                "gfi_indice":  indice,
                                "gfi_titre":   ged_titre_raw,
                                "gfi_date":    str(ged_date_raw or ""),
                                "gfi_sheet":   sheet_name,
                                "gfi_lot":     family.get("lot", ""),
                                "gfi_type_doc": family.get("type_doc", ""),
                            })
                    else:
                        # Patch E: Classify MISSING_IN_GF using SAS timing
                        mig_subtype, _ = _classify_missing_in_gf(
                            family, sas_lookup, data_date
                        )
                        sas_info = {}
                        for did in _sorted_family_doc_ids(family):
                            if did in sas_lookup:
                                sas_info = sas_lookup[did]
                                break
                        discrepancies.append({
                            "sheet_name":    sheet_name,
                            "document_code": doc_code,
                            "numero":        numero,
                            "indice":        indice,
                            "field":         "DOCUMENT",
                            "ged_value":     f"{numero}/{indice}",
                            "gf_value":      "NOT FOUND IN GF",
                            "ged_value_normalized": "",
                            "gf_value_normalized":  "",
                            "flag_type":     mig_subtype,
                            "title_similarity": None,
                            "date_diff_days":   None,
                            "is_excluded_population": False,
                            # Patch E extras
                            "sas_status_type":  sas_info.get("sas_status_type"),
                            "sas_result":       sas_info.get("sas_result"),
                            "sas_date":         str(sas_info.get("sas_date") or ""),
                            # GF_INSERTION_QUEUE fields (Part 9)
                            "gfi_numero":   numero,
                            "gfi_indice":   indice,
                            "gfi_titre":    ged_titre_raw,
                            "gfi_date":     str(ged_date_raw or ""),
                            "gfi_sheet":    sheet_name,
                            "gfi_lot":      family.get("lot", ""),
                            "gfi_type_doc": family.get("type_doc", ""),
                        })

        # ── MISSING_IN_GED: GF rows not covered by any GED family ─────────
        # Part 4 / Patch E: classify into subtypes
        for (gf_num, gf_ind), gf_rows in gf_docs.items():
            if (gf_num, gf_ind) not in matched_gf_keys:
                gf_row    = sorted(gf_rows, key=_gf_row_stable_key)[0]
                doc_label = str(gf_row.get("document", "") or "")[:80]

                # Historical: GED has this numero but at a newer indice
                is_historical = gf_num in all_ged_numeros_in_sheet

                # Year-filter exclusion (Patch C)
                gf_date_raw = gf_row.get("date_diffusion")
                is_excluded_by_year = False
                if not is_historical and min_year and gf_date_raw:
                    try:
                        gf_year = pd.to_datetime(gf_date_raw).year
                        if gf_year < min_year:
                            is_excluded_by_year = True
                    except Exception:
                        pass

                # Patch E: refined subtype classification
                flag_type_val = _classify_missing_in_ged(
                    gf_num, gf_ind, gf_rows,
                    is_historical, is_excluded_by_year,
                    gf_dup_counts, sheet_name,
                )

                discrepancies.append({
                    "sheet_name":    sheet_name,
                    "document_code": doc_label,
                    "numero":        gf_num,
                    "indice":        gf_ind,
                    "field":         "DOCUMENT",
                    "ged_value":     "NOT IN GED DERNIER",
                    "gf_value":      f"{gf_num}/{gf_ind}",
                    "ged_value_normalized": "",
                    "gf_value_normalized":  "",
                    "flag_type":     flag_type_val,
                    "title_similarity": None,
                    "date_diff_days":   None,
                    "is_excluded_population": is_excluded_by_year,
                    # Patch E: GF SAS context
                    "gf_visa_global":  gf_row.get("gf_visa_global"),
                    "gf_has_sas_ref":  gf_row.get("gf_has_sas_ref", False),
                })

    return discrepancies, gf_by_sheet


def _write_missing_in_ged_diagnosis(
    diag_path: str,
    true_only_path: str,
    summary_path: str,
    discrepancies: list,
):
    """
    Part 7 / Patch E: Write MISSING_IN_GED diagnosis outputs.

    Columns written: sheet, numero, indice, document_code,
                     flag_type (subtype), gf_visa_global, gf_has_sas_ref, severity
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    mid_types = {
        "MISSING_IN_GED_TRUE", "MISSING_IN_GED_HISTORICAL",
        "MISSING_IN_GED_GF_SAS_REF", "MISSING_IN_GED_GF_DUPLICATE_ROW",
        "MISSING_IN_GED_EXCLUDED", "MISSING_IN_GED_AMBIGUOUS",
        "MISSING_IN_GED_FAMILY_MATCH_MISSED",
        # legacy
        "MISSING_IN_GED",
    }

    mid_recs = [d for d in discrepancies if d.get("flag_type") in mid_types]
    if not mid_recs:
        _safe_console_print("    [Patch E] No MISSING_IN_GED records found — skipping diagnosis files")
        return

    headers = [
        "Sheet", "Numéro", "Indice", "Document Code",
        "Subtype", "GF Visa Global", "GF Has SAS REF",
        "Severity",
    ]
    colour_map = {
        "MISSING_IN_GED_TRUE":            "FFCCCC",
        "MISSING_IN_GED_GF_SAS_REF":      "FFE8CC",
        "MISSING_IN_GED_HISTORICAL":      "E8F5E9",
        "MISSING_IN_GED_GF_DUPLICATE_ROW":"FFF9C4",
        "MISSING_IN_GED_EXCLUDED":        "E3F2FD",
        "MISSING_IN_GED_AMBIGUOUS":       "F3E5F5",
        "MISSING_IN_GED_FAMILY_MATCH_MISSED": "FFF3E0",
    }

    def _make_wb(recs):
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "MISSING_IN_GED"
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        for ri, rec in enumerate(recs, 2):
            ft = rec.get("flag_type", "")
            sev = classify_discrepancy(rec)
            row_data = [
                rec.get("sheet_name", ""),
                rec.get("numero", ""),
                rec.get("indice", ""),
                rec.get("document_code", "")[:80],
                ft,
                rec.get("gf_visa_global", ""),
                "YES" if rec.get("gf_has_sas_ref") else "",
                sev,
            ]
            fill_color = colour_map.get(ft, "FFFFFF")
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
        for col_idx, width in zip(range(1, len(headers) + 1),
                                   [45, 12, 8, 60, 38, 15, 15, 18]):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        return wb2

    Path(diag_path).parent.mkdir(parents=True, exist_ok=True)
    _make_wb(mid_recs).save(diag_path)

    true_recs = [d for d in mid_recs if d.get("flag_type") == "MISSING_IN_GED_TRUE"]
    _make_wb(true_recs).save(true_only_path)

    # Summary counts
    from collections import Counter
    counts = Counter(d.get("flag_type") for d in mid_recs)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.cell(row=1, column=1, value="Subtype").font = Font(bold=True)
    ws_s.cell(row=1, column=2, value="Count").font   = Font(bold=True)
    for ri, (ft, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        ws_s.cell(row=ri, column=1, value=ft)
        ws_s.cell(row=ri, column=2, value=cnt)
    ws_s.column_dimensions["A"].width = 42
    ws_s.column_dimensions["B"].width = 10
    wb_s.save(summary_path)

    _safe_console_print(f"    [Patch E] MISSING_IN_GED breakdown: {dict(counts)}")
    _safe_console_print(f"    [Patch E] TRUE only: {len(true_recs)}")


def _write_missing_in_gf_diagnosis(
    diag_path: str,
    true_only_path: str,
    summary_path: str,
    discrepancies: list,
):
    """
    Part 7 / Patch E: Write MISSING_IN_GF diagnosis outputs.

    Columns: sheet, numero, indice, document_code, subtype,
             sas_status_type, sas_result, sas_date, severity,
             GFI fields (for future auto-insertion queue, Part 9)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    mig_types = {
        "MISSING_IN_GF_TRUE", "MISSING_IN_GF_PENDING_SAS",
        "MISSING_IN_GF_RECENT_SAS_REMINDER", "MISSING_IN_GF_RECENT_REFUSAL",
        "MISSING_IN_GF_RECENT_ACCEPTED_SAS", "MISSING_IN_GF_SAME_KEY_COLLISION",
        "MISSING_IN_GF_AMBIGUOUS",
        # legacy
        "MISSING_IN_GF",
    }

    mig_recs = [d for d in discrepancies if d.get("flag_type") in mig_types]
    if not mig_recs:
        _safe_console_print("    [Patch E] No MISSING_IN_GF records found — skipping diagnosis files")
        return

    headers = [
        "Sheet", "Numéro", "Indice", "Document Code",
        "Subtype", "SAS Status Type", "SAS Result", "SAS Date",
        "Severity",
        # GF_INSERTION_QUEUE (Part 9)
        "GFI Titre", "GFI Date", "GFI Lot", "GFI Type Doc",
    ]
    colour_map = {
        "MISSING_IN_GF_TRUE":                 "FFCCCC",
        "MISSING_IN_GF_PENDING_SAS":          "E8F5E9",
        "MISSING_IN_GF_RECENT_SAS_REMINDER":  "FFF9C4",
        "MISSING_IN_GF_RECENT_REFUSAL":       "FFE8CC",
        "MISSING_IN_GF_RECENT_ACCEPTED_SAS":  "E3F2FD",
        "MISSING_IN_GF_SAME_KEY_COLLISION":   "F3E5F5",
        "MISSING_IN_GF_AMBIGUOUS":            "EEEEEE",
    }

    def _make_wb(recs):
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "MISSING_IN_GF"
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="7B2D00")
        for ri, rec in enumerate(recs, 2):
            ft  = rec.get("flag_type", "")
            sev = classify_discrepancy(rec)
            row_data = [
                rec.get("sheet_name", ""),
                rec.get("numero", ""),
                rec.get("indice", ""),
                rec.get("document_code", "")[:80],
                ft,
                rec.get("sas_status_type", ""),
                rec.get("sas_result", ""),
                rec.get("sas_date", ""),
                sev,
                # GFI queue
                rec.get("gfi_titre", "")[:80],
                rec.get("gfi_date", ""),
                rec.get("gfi_lot", ""),
                rec.get("gfi_type_doc", ""),
            ]
            fill_color = colour_map.get(ft, "FFFFFF")
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
        widths = [45, 12, 8, 60, 38, 18, 12, 12, 18, 70, 14, 12, 14]
        for col_idx, width in zip(range(1, len(headers) + 1), widths):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        return wb2

    Path(diag_path).parent.mkdir(parents=True, exist_ok=True)
    _make_wb(mig_recs).save(diag_path)

    true_recs = [d for d in mig_recs if d.get("flag_type") == "MISSING_IN_GF_TRUE"]
    _make_wb(true_recs).save(true_only_path)

    # Summary counts
    from collections import Counter
    counts = Counter(d.get("flag_type") for d in mig_recs)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.cell(row=1, column=1, value="Subtype").font = Font(bold=True)
    ws_s.cell(row=1, column=2, value="Count").font   = Font(bold=True)
    for ri, (ft, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        ws_s.cell(row=ri, column=1, value=ft)
        ws_s.cell(row=ri, column=2, value=cnt)
    ws_s.column_dimensions["A"].width = 42
    ws_s.column_dimensions["B"].width = 10
    wb_s.save(summary_path)

    _safe_console_print(f"    [Patch E] MISSING_IN_GF breakdown: {dict(counts)}")
    _safe_console_print(f"    [Patch E] TRUE only: {len(true_recs)}")


def _write_gf_schema_debug(debug_dir: str, structures: dict):
    """
    Write debug/gf_sheet_schema.xlsx from parsed GF sheet structures.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    Path(debug_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(debug_dir) / "gf_sheet_schema.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GF Sheet Schema"

    headers = [
        "sheet_name",
        "header_row",
        "data_start_row",
        "total_cols",
        "base_col_count",
        "approver_count",
        "col_map",
        "approvers",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row_idx, sheet_name in enumerate(sorted(structures.keys()), 2):
        struct = structures.get(sheet_name, {}) or {}
        approvers = struct.get("approvers", []) or []
        approver_names = ", ".join(
            str(a.get("name", "")).strip()
            for a in approvers
            if str(a.get("name", "")).strip()
        )
        col_map = struct.get("col_map", {}) or {}
        col_map_str = ", ".join(
            f"{key}:{value}" for key, value in sorted(col_map.items(), key=lambda kv: str(kv[0]))
        )
        row = [
            sheet_name,
            struct.get("header_row"),
            struct.get("data_start_row"),
            struct.get("total_cols"),
            struct.get("base_col_count"),
            len(approvers),
            col_map_str,
            approver_names,
        ]
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [42, 12, 14, 10, 14, 14, 60, 80]
    for col_idx, width in zip(range(1, len(headers) + 1), widths):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)


def _write_gf_duplicates_debug(debug_dir: str, gf_dup_counts: dict):
    """
    Part 1: Write debug/gf_duplicates.xlsx.

    Lists every (sheet_name, numero, indice) key that appeared more than once
    in the existing GF, along with the duplicate count.

    Columns: sheet_name, numero, indice, count_of_duplicates
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    Path(debug_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(debug_dir) / "gf_duplicates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GF Duplicate Keys"

    headers = ["sheet_name", "numero", "indice", "count_of_duplicates"]
    header_fill = PatternFill("solid", fgColor="C00000")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    # Sort by count desc, then sheet, num, ind for readability
    rows_sorted = sorted(
        gf_dup_counts.items(),
        key=lambda kv: (-kv[1], kv[0][0], kv[0][1], kv[0][2]),
    )

    row_fill = PatternFill("solid", fgColor="FFE0E0")
    for r_idx, ((sheet_name, num, ind), count) in enumerate(rows_sorted, 2):
        ws.cell(row=r_idx, column=1, value=sheet_name)
        ws.cell(row=r_idx, column=2, value=num)
        ws.cell(row=r_idx, column=3, value=ind)
        c = ws.cell(row=r_idx, column=4, value=count)
        if count >= 3:
            for col in range(1, 5):
                ws.cell(row=r_idx, column=col).fill = row_fill

    # Summary row
    total_keys = len(rows_sorted)
    total_excess = sum(v - 1 for v in gf_dup_counts.values())
    summary_row = total_keys + 2
    ws.cell(row=summary_row, column=1, value="TOTAL duplicate keys").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=total_keys).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="TOTAL excess rows").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=total_excess).font = Font(bold=True)

    for col_idx, width in zip(range(1, len(headers) + 1), [34, 18, 10, 20]):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)
