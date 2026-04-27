"""
config_loader.py
----------------
Exclusion rules and project-level configuration for GF Updater V3.

Rules:
  1. LOT I01-VDSTP : ignore all documents — sheet is infrastructure-only
  2. LOT I02-FKI   : ignore all documents — sub-lot split, handled separately
  3. LOT 03-GOE    : only include documents created in 2026 or later
  4. OLD-* sheets  : always skipped (routing.py already handles this)

Emetteur routing filter (Patch A2):
  Maps each GF sheet name → set of ALLOWED emetteur codes.
  Documents whose emetteur is NOT in the allowed set are marked
  ROUTING_EMETTEUR_MISMATCH and excluded from that sheet.
  If a sheet has no entry here → no emetteur filter applied (permissive).
"""

from typing import Callable, Dict, List, Optional, Set

import pandas as pd


# ─────────────────────────────────────────────────────────────
# EXCLUSION RULES
# ─────────────────────────────────────────────────────────────

# Sheets to completely exclude (all documents)
EXCLUDED_SHEETS: set = {
    "LOT I01-VDSTP",
    "LOT I02-FKI",
}

# Lot-normalized prefixes to completely exclude
EXCLUDED_LOT_CODES: set = set()  # Not used — handled via EXCLUDED_SHEETS + routing

# Year filters: {sheet_name: min_year} — exclude docs before min_year
SHEET_YEAR_FILTERS: Dict[str, int] = {
    "LOT 03-GOE-LGD": 2026,
    # BENTIN sheet: original GF has 135 rows, all dated 2026.
    # The GED contains 659 BEN DI rows (2024–2026) because OLD_BEN legacy docs
    # (originally in "OLD 31 à 34-IN-BX-CFO-BENTIN") are also stored in the GED.
    # Pre-2026 BEN docs = OLD_BEN legacy exceptions → excluded from clean GF.
    # Valid ANCIEN rows are kept via lifecycle_id membership in active 2026 families.
    "LOT 31 à 34-IN-BX-CFO-BENTIN": 2026,
}

# Lot prefix + year filter: {lot_prefix: min_year}
LOT_YEAR_FILTERS: Dict[str, int] = {}


# ─────────────────────────────────────────────────────────────
# EMETTEUR ROUTING FILTER (Patch A2)
# ─────────────────────────────────────────────────────────────
# Maps GF sheet name → frozenset of ALLOWED emetteur codes from GED.
#
# Derivation logic:
#   - Single-emetteur sheets: set contains exactly that emetteur.
#   - Multi-emetteur sheets where all emetteurs are legit: set contains all.
#   - Sheets not listed here: no emetteur filter applied (all emetteurs pass).
#
# This was derived from inspecting emetteur→lots in GED and the GF sheet names.
# Key contractor name ↔ GED emetteur code mapping:
#   AXIMA=AXI, UTB=UTB, SNIE/SNI=SNI, BENTIN=BEN, APILOG=API, SEPA+VRD=SPA+VRD,
#   SMAC=SMA, ICM=ICM, AMP=AMP, DUVAL/DUV=DUV, LGD=LGD, FER=FER, FKI=FKI,
#   SODIC=SOD, BANGUI=BAN, CMF=CMF, TLS=TLS, FRS=FRS, VALLEE=VAL, DBH=DBH,
#   JLE=JLE, HVA=HVA, LINDNER=LIN, LAC=LAC, IST=IST, CPL=CPL, FMC=FMC, etc.
#
SHEET_EMETTEUR_FILTER: Dict[str, frozenset] = {
    # ── 100% single-contractor sheets ──────────────────────────
    "LOT 03-GOE-LGD":                    frozenset({"LGD"}),
    "LOT 05-MEN EXT-ICM":                frozenset({"ICM"}),
    "LOT 04-06-ETANCH-COUV-SMAC":        frozenset({"SMA"}),
    "LOT AH06-RVT FAC - FER ":           frozenset({"FER"}),
    "LOT B06-AH07-BFUP-CAS BA-LGD":      frozenset({"LGD"}),
    "LOT 08-MUR-RID-DUV":                frozenset({"DUV"}),
    "LOT B09-Oc-SODIC":                  frozenset({"SOD"}),
    "LOT 11 - 16A -AMP-CLD - FPL":       frozenset({"AMP"}),
    "Lot 12A-LAC-MEN":                   frozenset({"LAC"}),
    # NOTE: old name was "Lot HA12B-IST-MEN" — renamed in v3 GF
    "Lot AH11, AH12B, AH016-IST":        frozenset({"IST"}),
    "LOT 12B-CPL-MEN PARQUET":           frozenset({"CPL"}),
    "LOT I13A-SERR-DUVAL":               frozenset({"DUV"}),
    "LOT I13B-RID CF-FER":               frozenset({"FMC"}),
    "LOT 13BAH-SERR-ATCH":               frozenset({"CHV"}),
    "LOT 14-PORTE-PK-FER":               frozenset({"FER"}),
    "LOT 16B-FP RAYON-LINDNER":          frozenset({"LIN"}),
    "LOT B17-BX-FPR-BANGUI":             frozenset({"BAN"}),
    "LOT A18-SOLS & MURS - CMF":         frozenset({"CMF"}),
    "LOT H18-SOLS & MURS TLS":           frozenset({"TLS"}),
    "LOT 018-019-BX-SOLS DS-FRS":        frozenset({"FRS"}),
    "LOT A19-SOLS DS-VALLEE":            frozenset({"VAL"}),
    "LOT 0I20-PEINT PK-DBH":             frozenset({"DBH"}),
    "LOT B020-PEINT - JLE":              frozenset({"JLE"}),
    "LOT A22-SDB PREFA-HVA":             frozenset({"HVA"}),
    "LOT 31 à 34-AU-HO-CFO-SNIE":        frozenset({"SNI"}),
    "LOT 31 à 34-IN-BX-CFO-BENTIN":      frozenset({"BEN"}),
    "LOT 35-GTB-APILOG":                 frozenset({"API"}),
    "LOT 41-CVC-AXIMA":                  frozenset({"AXI"}),
    "LOT 42-PLB-UTB":                    frozenset({"UTB"}),
    "LOT 42b-VSQ-CREA":                  frozenset({"CRE"}),
    "LOT 43-SPK-AAI":                    frozenset({"AAI"}),
    "LOT 51-ASC-SCHINDLER":              frozenset({"SCH"}),
    # ── Multi-contractor sheets (both emetteurs are legitimate) ─
    "LOT 6162-VRD-EV-SEPA":              frozenset({"VRD", "SPA"}),
    # ── Excluded sheets (already excluded globally) ─────────────
    "LOT I01-VDSTP":                     frozenset({"VDP", "VTP"}),
    "LOT I02-FKI":                       frozenset({"FKI", "LGD"}),
}


# ─────────────────────────────────────────────────────────────
# EXCLUSION ENGINE
# ─────────────────────────────────────────────────────────────

class ExclusionConfig:
    """
    Applies exclusion rules to a DataFrame of documents.

    Usage:
        config = ExclusionConfig()
        docs_df = config.apply(docs_df)
        # docs_df now has 'is_excluded_config' and 'exclusion_reason' columns
    """

    def __init__(
        self,
        excluded_sheets: Optional[set] = None,
        sheet_year_filters: Optional[Dict[str, int]] = None,
        lot_year_filters: Optional[Dict[str, int]] = None,
        sheet_emetteur_filter: Optional[Dict[str, frozenset]] = None,
    ):
        self.excluded_sheets       = excluded_sheets       or EXCLUDED_SHEETS
        self.sheet_year_filters    = sheet_year_filters    or SHEET_YEAR_FILTERS
        self.lot_year_filters      = lot_year_filters      or LOT_YEAR_FILTERS
        self.sheet_emetteur_filter = sheet_emetteur_filter or SHEET_EMETTEUR_FILTER

    def apply(self, docs_df: pd.DataFrame) -> pd.DataFrame:
        """
        Add 'is_excluded_config' and 'exclusion_reason' columns to docs_df.
        Modifies a copy; does not remove rows.

        Requires columns (all optional — graceful fallback to not-excluded):
          gf_sheet_name, created_at, lot_prefix, lot_normalized, emetteur
        """
        df = docs_df.copy()
        df["is_excluded_config"] = False
        df["exclusion_reason"]   = ""

        for idx, row in df.iterrows():
            excluded, reason = self._check_row(row)
            if excluded:
                df.at[idx, "is_excluded_config"] = True
                df.at[idx, "exclusion_reason"]   = reason

        return df

    def _check_row(self, row) -> tuple:
        """Return (is_excluded: bool, reason: str)."""

        sheet    = row.get("gf_sheet_name") or ""
        emetteur = str(row.get("emetteur") or "").strip()

        # Rule 1: Completely excluded sheets
        if sheet in self.excluded_sheets:
            return True, f"EXCLUDED_SHEET:{sheet}"

        # Rule 2: Sheet-specific year filter
        if sheet in self.sheet_year_filters:
            min_year = self.sheet_year_filters[sheet]
            created = row.get("created_at")
            if created is not None:
                try:
                    year = pd.to_datetime(created).year
                    if year < min_year:
                        return True, f"PRE_{min_year}:{sheet}"
                except Exception:
                    pass

        # Rule 3: Lot-prefix year filter
        prefix = row.get("lot_prefix") or ""
        if prefix in self.lot_year_filters:
            min_year = self.lot_year_filters[prefix]
            created = row.get("created_at")
            if created is not None:
                try:
                    year = pd.to_datetime(created).year
                    if year < min_year:
                        return True, f"PRE_{min_year}_PREFIX:{prefix}"
                except Exception:
                    pass

        # Rule 4: Emetteur routing mismatch (Patch A2)
        # If a routing_status of ROUTING_EMETTEUR_MISMATCH was set, exclude
        routing_status = row.get("routing_status") or ""
        if routing_status == "ROUTING_EMETTEUR_MISMATCH":
            return True, f"EMETTEUR_MISMATCH:sheet={sheet},emetteur={emetteur}"

        return False, ""

    def summary(self, docs_df: pd.DataFrame) -> dict:
        """Return exclusion counts after apply() has been called."""
        if "is_excluded_config" not in docs_df.columns:
            return {}
        excluded = docs_df[docs_df["is_excluded_config"] == True]
        reasons = excluded["exclusion_reason"].value_counts().to_dict()
        return {
            "total_excluded": len(excluded),
            "by_reason": reasons,
        }

    def write_exclusion_summary(
        self,
        docs_df: pd.DataFrame,
        output_path: str,
    ):
        """Write debug/exclusion_summary.xlsx (Patch C requirement)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        from collections import defaultdict

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if "is_excluded_config" not in docs_df.columns:
            return

        excluded = docs_df[docs_df["is_excluded_config"] == True].copy()

        # Aggregate by exclusion_reason
        agg: dict = defaultdict(lambda: {
            "row_count": 0,
            "emetteurs": set(),
            "lots": set(),
            "sheets": set(),
        })

        for _, row in excluded.iterrows():
            reason = row.get("exclusion_reason") or "UNKNOWN"
            agg[reason]["row_count"] += 1
            agg[reason]["emetteurs"].add(str(row.get("emetteur") or ""))
            agg[reason]["lots"].add(str(row.get("lot_normalized") or ""))
            agg[reason]["sheets"].add(str(row.get("gf_sheet_name") or ""))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exclusion Summary"

        headers = [
            "Exclusion Rule", "Row Count Excluded",
            "Affected Emetteurs", "Affected Lots (normalized)", "Affected Sheets",
        ]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="7F0000")

        for r_idx, (reason, data) in enumerate(sorted(agg.items()), 2):
            ws.cell(row=r_idx, column=1, value=reason)
            ws.cell(row=r_idx, column=2, value=data["row_count"])
            ws.cell(row=r_idx, column=3,
                    value=", ".join(sorted(data["emetteurs"] - {""})))
            ws.cell(row=r_idx, column=4,
                    value=", ".join(sorted(data["lots"] - {""})))
            ws.cell(row=r_idx, column=5,
                    value=", ".join(sorted(data["sheets"] - {""})))

        # Totals
        total_row = len(agg) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2,
                value=len(excluded)).font = Font(bold=True)

        for col, w in zip(range(1, 6), [40, 18, 30, 40, 60]):
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(output_path)
        print(f"    → exclusion_summary.xlsx written ({len(excluded)} excluded rows)")


# ─────────────────────────────────────────────────────────────
# CONVENIENCE FACTORY
# ─────────────────────────────────────────────────────────────

def load_exclusion_config() -> ExclusionConfig:
    """
    Return the default exclusion config for this project.
    To customise: edit EXCLUDED_SHEETS / SHEET_YEAR_FILTERS above,
    or pass overrides directly to ExclusionConfig().
    """
    return ExclusionConfig(
        excluded_sheets=EXCLUDED_SHEETS,
        sheet_year_filters=SHEET_YEAR_FILTERS,
        lot_year_filters=LOT_YEAR_FILTERS,
        sheet_emetteur_filter=SHEET_EMETTEUR_FILTER,
    )
