"""
routing.py  (v3 — Patch A)
-----------
Routes documents to GF contractor sheets using:
  - lot_normalized  (numeric lot id)
  - lot_prefix      (building code: A=AU, B=BX, H=HO, I=IN)
  - emetteur        (contractor code — new in Patch A)

Routing logic (in order):
  1. Find candidate sheets via lot_normalized + lot_prefix (existing logic)
  2. Validate emetteur against SHEET_EMETTEUR_FILTER:
       - If emetteur is in the allowed set → route OK
       - If emetteur NOT in allowed set but another sheet exists that
         (a) also matches the lot/prefix and (b) accepts this emetteur
         → route to that sheet instead
       - If no valid sheet found → ROUTING_EMETTEUR_MISMATCH

Routing status values:
  OK                        – matched by lot + prefix + emetteur
  ROUTING_AMBIGUOUS         – multiple sheets claim this combo
  ROUTING_UNMATCHED         – no sheet found for lot/prefix
  ROUTING_EMETTEUR_MISMATCH – lot/prefix matched but emetteur not expected

Debug output:
  - output/debug/routing_summary.xlsx
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
import pandas as pd


# ─────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────

# Map full building names (from sheet names) → single-letter lot prefix
BUILDING_NAME_TO_PREFIX: Dict[str, str] = {
    "AU": "A",
    "HO": "H",
    "BX": "B",
    "IN": "I",  # Infrastructure building
}

# Multi-letter building prefix in lot identifier → letter prefix
LOT_LETTER_TO_PREFIX: Dict[str, str] = {
    "A": "A",
    "B": "B",
    "H": "H",
    "I": "I",
    "G": "G",
    "0": "0",  # common/zero-prefix lots
}


# ─────────────────────────────────────────────────────────────
# 1. PARSE SHEET NAME → LOT IDS + BUILDING PREFIXES
# ─────────────────────────────────────────────────────────────

def extract_lot_numbers_from_sheet_name(sheet_name: str) -> List[str]:
    """
    Extract all numeric lot identifiers from a GF sheet name.
    Examples:
      'LOT 41-CVC-AXIMA'              → ['41']
      'LOT 31 à 34-AU-HO-CFO-SNIE'   → ['31', '32', '33', '34']
      'LOT 6162-VRD-EV-SEPA'         → ['61', '62']
      'Lot HA12B-IST-MEN'            → ['12B']
      'LOT I13A-SERR-DUVAL'          → ['13A']
      'LOT I13B-RID CF-FER'          → ['13B']
      'LOT 04-06-ETANCH-COUV-SMAC'  → ['4', '5', '6']
      'LOT 11 - 16A -AMP-CLD - FPL' → ['11', '16A']
      'LOT 16B-FP RAYON-LINDNER'     → ['16B']
    """
    name = re.sub(r'^LOT\s+', '', sheet_name, flags=re.IGNORECASE).strip()
    name = re.sub(r'^OLD\s+', '', name, flags=re.IGNORECASE).strip()

    # French range: 'NN à MM'
    range_match = re.match(r'^[A-Za-z]*(\d+)\s+[àa]\s+(\d+)', name)
    if range_match:
        return [str(i) for i in range(int(range_match.group(1)),
                                      int(range_match.group(2)) + 1)]

    # Multi-segment first part: 'NNA - MMB' or range '04-06'
    multi_match = re.match(r'^[A-Za-z]*(\d+[A-Za-z]?)\s*[-–]\s*(\d+[A-Za-z]?)', name)
    if multi_match:
        n1_raw, n2_raw = multi_match.group(1), multi_match.group(2)
        n1d = re.sub(r'[A-Za-z]', '', n1_raw)
        n2d = re.sub(r'[A-Za-z]', '', n2_raw)
        if n1d and n2d:
            n1i, n2i = int(n1d), int(n2d)
            has_suffix = bool(re.search(r'[A-Za-z]', n1_raw + n2_raw))
            if not has_suffix and n2i > n1i and (n2i - n1i) < 20:
                return [str(i) for i in range(n1i, n2i + 1)]
            else:
                results = []
                for raw in [n1_raw, n2_raw]:
                    dm = re.match(r'^(\d+)([A-Za-z]?)$', raw)
                    if dm:
                        n = str(int(dm.group(1)))
                        s = dm.group(2).upper()
                        results.append(n + s if s else n)
                if results:
                    return results

    # Concatenated '6162'
    first_seg = name.split('-')[0].strip().split()[0] if name else ""
    if re.match(r'^\d{4,}$', first_seg):
        raw = first_seg
        return [str(int(raw[i:i+2])) for i in range(0, len(raw), 2)]

    # General: letter* + digits + optional letter suffix
    m = re.match(r'^[A-Za-z]*(\d+)([A-Za-z]?)', first_seg)
    if m:
        num = str(int(m.group(1)))
        suffix = m.group(2).upper()
        return [num + suffix if suffix else num]

    return []


def extract_building_prefixes_from_sheet_name(sheet_name: str) -> Optional[List[str]]:
    """
    Extract which building-prefixes (A/B/H/I) a sheet covers.
    Returns None = wildcard (covers all buildings).

    Strategy:
    1. Scan for standalone building code words: AU→A, HO→H, BX→B
       ('IN' detected only when explicitly hyphen-delimited to avoid false matches
        in words like BENTIN, LINDNER, etc.)
    2. Check direct letter-prefix in the first lot identifier (B09→B, A18→A...)
    3. Handle compound prefixes: 'HA12B' → H + A, '0I20' → 0 + I
    """
    name_upper = sheet_name.upper()
    found: Set[str] = set()

    # --- Step 1: scan for standalone building code words ---
    # Split on spaces, dashes, underscores
    segments = re.split(r'[\s\-_–]+', name_upper)

    for seg in segments:
        # Clean trailing punctuation
        seg = seg.strip('.,;()')
        if seg in BUILDING_NAME_TO_PREFIX:
            found.add(BUILDING_NAME_TO_PREFIX[seg])

    if found:
        return sorted(found)

    # --- Step 2: check letter prefix in first lot identifier ---
    name_clean = re.sub(r'^LOT\s+', '', sheet_name, flags=re.IGNORECASE).strip()
    name_clean = re.sub(r'^OLD\s+', '', name_clean, flags=re.IGNORECASE).strip()
    first_seg = name_clean.split('-')[0].strip().split()[0] if name_clean else ""

    # Extract leading letters before digits (e.g. 'HA12B' → 'HA', 'B09' → 'B',
    # 'A18' → 'A', '0I20' → '0I', 'I01' → 'I')
    m = re.match(r'^([A-Za-z0-9]*[A-Za-z]+)(\d)', first_seg)
    if m:
        prefix_str = m.group(1).upper()
        # Filter to known single-letter prefixes
        detected = []
        for char in prefix_str:
            if char in ('A', 'B', 'H', 'I', 'G', '0'):
                detected.append(char)
        if detected:
            return sorted(set(detected))

    return None  # wildcard — matches all buildings


# ─────────────────────────────────────────────────────────────
# 2. BUILD ROUTING TABLE
# ─────────────────────────────────────────────────────────────

class RoutingTable:
    """
    Routing table mapping (lot_normalized, building_prefix) → sheet_name.

    Lookup priority:
    1. Exact: (lot_num, prefix) → sheet
    2. Wildcard: (lot_num, None) → sheet (sheet covers all buildings)
    3. No match → ROUTING_UNMATCHED
    """

    def __init__(self):
        # {(lot_num, prefix_or_None): [sheet_name, ...]}
        self._table: Dict[Tuple[str, Optional[str]], List[str]] = {}

    def add(self, lot_num: str, prefixes: Optional[List[str]], sheet_name: str):
        """
        Register a sheet for a lot + optional building prefixes.
        prefixes=None means wildcard (sheet covers all buildings for this lot).
        """
        if prefixes is None:
            key = (lot_num, None)
            self._table.setdefault(key, [])
            if sheet_name not in self._table[key]:
                self._table[key].append(sheet_name)
        else:
            for prefix in prefixes:
                key = (lot_num, prefix)
                self._table.setdefault(key, [])
                if sheet_name not in self._table[key]:
                    self._table[key].append(sheet_name)

    def lookup(self, lot_num: str, building_prefix: Optional[str]) -> Tuple[Optional[str], str]:
        """
        Return (sheet_name, routing_status).
        routing_status: 'OK' | 'ROUTING_AMBIGUOUS' | 'ROUTING_UNMATCHED'
        """
        if lot_num is None:
            return None, "ROUTING_UNMATCHED"

        # 1. Exact match
        exact_key = (str(lot_num), building_prefix)
        if exact_key in self._table:
            sheets = self._table[exact_key]
            if len(sheets) == 1:
                return sheets[0], "OK"
            else:
                return sheets[0], "ROUTING_AMBIGUOUS"

        # 2. Wildcard for this lot
        wild_key = (str(lot_num), None)
        if wild_key in self._table:
            sheets = self._table[wild_key]
            if len(sheets) == 1:
                return sheets[0], "OK"
            else:
                return sheets[0], "ROUTING_AMBIGUOUS"

        # 3. Fallback: try stripping A/B suffix from lot_num ('12A' → '12', '13B' → '13')
        m = re.match(r'^(\d+)[A-Z]$', str(lot_num).upper())
        if m:
            base = m.group(1)
            for key in [(base, building_prefix), (base, None)]:
                if key in self._table:
                    sheets = self._table[key]
                    return sheets[0], "OK"

        # 4. Try stripping all letters
        digits_only = re.sub(r'[A-Za-z]', '', str(lot_num))
        if digits_only and digits_only != str(lot_num):
            for key in [(digits_only, building_prefix), (digits_only, None)]:
                if key in self._table:
                    sheets = self._table[key]
                    return sheets[0], "OK"

        return None, "ROUTING_UNMATCHED"

    def lookup_all_candidates(
        self,
        lot_num: str,
        building_prefix: Optional[str],
    ) -> List[str]:
        """
        Return ALL candidate sheet names for a lot+prefix combination
        (including wildcard matches). Used for emetteur-based disambiguation.
        """
        if lot_num is None:
            return []

        candidates: List[str] = []

        for key in [
            (str(lot_num), building_prefix),
            (str(lot_num), None),
        ]:
            candidates.extend(self._table.get(key, []))

        # Fallback suffixes
        m = re.match(r'^(\d+)[A-Z]$', str(lot_num).upper())
        if m:
            base = m.group(1)
            for key in [(base, building_prefix), (base, None)]:
                candidates.extend(self._table.get(key, []))

        digits_only = re.sub(r'[A-Za-z]', '', str(lot_num))
        if digits_only and digits_only != str(lot_num):
            for key in [(digits_only, building_prefix), (digits_only, None)]:
                candidates.extend(self._table.get(key, []))

        # Deduplicate, preserving order
        seen: set = set()
        result: List[str] = []
        for s in candidates:
            if s not in seen:
                seen.add(s)
                result.append(s)
        return result

    def all_entries(self):
        """Iterate all (lot_num, prefix, [sheets]) for debugging."""
        for (lot_num, prefix), sheets in sorted(
            self._table.items(), key=lambda kv: (kv[0][0], kv[0][1] or "")
        ):
            yield lot_num, prefix, sheets


def build_routing_table(gf_filepath: str) -> "RoutingTable":
    """
    Build routing table from all sheets in existing GF file.
    Skips OLD sheets.

    Some sheet names use non-standard formats the parser cannot handle
    automatically (comma-separated lots, building-prefix letters embedded in lot
    code, combined digit+prefix codes).  These are listed in
    MANUAL_SHEET_LOT_OVERRIDES and take precedence over the auto-parsed values.
    """

    # Manual overrides: sheet_name → (list_of_lot_nums, list_of_building_prefixes or None)
    # Format: lot_nums are the same string keys used in GED lot_normalized
    # (numeric, no leading zeros, with letter suffix if applicable)
    MANUAL_SHEET_LOT_OVERRIDES: Dict[str, Tuple] = {
        # AH11, AH12B, AH016 → lots 11 / 12B / 16 in buildings A and H
        "Lot AH11, AH12B, AH016-IST":  (["11", "12B", "16"], ["A", "H"]),
        # 13BAH → lot 13 in buildings B, A, H
        "LOT 13BAH-SERR-ATCH":          (["13"],              ["A", "B", "H"]),
        # 0I20 → lot 20 in buildings I (and 0 if used)
        "LOT 0I20-PEINT PK-DBH":        (["20"],              ["I", "0"]),
    }

    table = RoutingTable()
    wb = openpyxl.load_workbook(gf_filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.upper().startswith("OLD"):
            continue

        if sheet_name in MANUAL_SHEET_LOT_OVERRIDES:
            lot_nums, bld_prefixes = MANUAL_SHEET_LOT_OVERRIDES[sheet_name]
            print(f"  [routing] Manual override for '{sheet_name}': "
                  f"lots={lot_nums}, prefixes={bld_prefixes}")
        else:
            lot_nums = extract_lot_numbers_from_sheet_name(sheet_name)
            bld_prefixes = extract_building_prefixes_from_sheet_name(sheet_name)

        for lot_num in lot_nums:
            table.add(lot_num, bld_prefixes, sheet_name)

    return table


def build_lot_to_sheet_map(gf_filepath: str) -> Dict[str, str]:
    """
    Legacy simple map: lot_number → first sheet found.
    Kept for backwards compatibility.
    """
    table = build_routing_table(gf_filepath)
    result = {}
    for lot_num, prefix, sheets in table.all_entries():
        key = lot_num if prefix is None else f"{lot_num}_{prefix}"
        if key not in result:
            result[key] = sheets[0]
    return result


# ─────────────────────────────────────────────────────────────
# 3. EMETTEUR-AWARE ROUTING (Patch A)
# ─────────────────────────────────────────────────────────────

def build_routing_key(document: dict) -> Tuple[Optional[str], Optional[str], str]:
    """
    Build the routing lookup key for a document.
    Returns (lot_normalized, lot_prefix, emetteur).
    """
    lot = document.get("lot_normalized")
    prefix = document.get("lot_prefix")
    if prefix and len(str(prefix)) > 1:
        prefix = str(prefix)[0]  # normalise to single letter
    emetteur = str(document.get("emetteur") or "").strip()
    return lot, prefix, emetteur


def match_route_with_emetteur(
    lot_normalized: Optional[str],
    lot_prefix: Optional[str],
    emetteur: str,
    routing_table: "RoutingTable",
    sheet_emetteur_filter: Dict[str, frozenset],
) -> Tuple[Optional[str], str]:
    """
    Route a document using lot + prefix + emetteur.

    Algorithm:
    1. Find all candidate sheets for (lot, prefix).
    2. Among candidates, find sheets whose emetteur filter includes this emetteur.
       - If exactly one match → OK
       - If multiple matches → ROUTING_AMBIGUOUS (pick first)
       - If zero matches from filtered set but candidates exist →
           ROUTING_EMETTEUR_MISMATCH (emetteur not expected on any matching sheet)
    3. If no candidates at all → ROUTING_UNMATCHED

    Returns (sheet_name_or_None, routing_status).
    """
    if lot_normalized is None:
        return None, "ROUTING_UNMATCHED"

    # Step 1: gather all candidate sheets for this lot+prefix
    candidates = routing_table.lookup_all_candidates(lot_normalized, lot_prefix)

    if not candidates:
        return None, "ROUTING_UNMATCHED"

    # Step 2: filter by emetteur
    emetteur_matches: List[str] = []
    for sheet in candidates:
        allowed = sheet_emetteur_filter.get(sheet)
        if allowed is None:
            # No filter for this sheet → emetteur passes
            emetteur_matches.append(sheet)
        elif emetteur in allowed:
            emetteur_matches.append(sheet)
        # else: this sheet does NOT accept this emetteur → skip it

    if len(emetteur_matches) == 1:
        return emetteur_matches[0], "OK"
    elif len(emetteur_matches) > 1:
        # Multiple sheets accept this emetteur — use first (already ranked by lookup priority)
        return emetteur_matches[0], "ROUTING_AMBIGUOUS"
    else:
        # All candidate sheets rejected this emetteur
        # Fall back to the first candidate but mark it mismatch
        return candidates[0], "ROUTING_EMETTEUR_MISMATCH"


# ─────────────────────────────────────────────────────────────
# 4. ROUTE DOCUMENTS
# ─────────────────────────────────────────────────────────────

def route_documents(
    docs_df: pd.DataFrame,
    routing_table: "RoutingTable",
    sheet_emetteur_filter: Optional[Dict[str, frozenset]] = None,
) -> pd.DataFrame:
    """
    Route documents using the RoutingTable + emetteur validation.
    Adds columns:
      - gf_sheet_name       : matched sheet, or None
      - routing_status      : 'OK' | 'ROUTING_AMBIGUOUS' |
                              'ROUTING_UNMATCHED' | 'ROUTING_EMETTEUR_MISMATCH'
      - routing_debug_notes : human-readable routing decision note

    If sheet_emetteur_filter is None, defaults to SHEET_EMETTEUR_FILTER
    from config_loader (imported lazily to avoid circular imports).
    """
    # Lazy import to avoid circular dependency
    if sheet_emetteur_filter is None:
        from config_loader import SHEET_EMETTEUR_FILTER
        sheet_emetteur_filter = SHEET_EMETTEUR_FILTER

    df = docs_df.copy()

    routed_sheets: list = []
    routed_statuses: list = []
    routed_notes: list = []

    for _, row in df.iterrows():
        lot = row.get("lot_normalized")
        raw_prefix = row.get("lot_prefix")
        emetteur = str(row.get("emetteur") or "").strip()

        # Normalise prefix to single letter
        prefix = raw_prefix
        if prefix and len(str(prefix)) > 1:
            # Multi-letter prefix (e.g., 'HA') → try emetteur-aware per letter
            best_sheet = None
            best_status = "ROUTING_UNMATCHED"
            for p in str(prefix):
                s, st = match_route_with_emetteur(
                    lot, p, emetteur, routing_table, sheet_emetteur_filter
                )
                if st == "OK":
                    best_sheet, best_status = s, st
                    break
                elif st in ("ROUTING_AMBIGUOUS",) and best_status == "ROUTING_UNMATCHED":
                    best_sheet, best_status = s, st
            if best_status == "ROUTING_UNMATCHED":
                # Final fallback using first letter
                prefix = str(prefix)[0]
                best_sheet, best_status = match_route_with_emetteur(
                    lot, prefix, emetteur, routing_table, sheet_emetteur_filter
                )
            sheet, status = best_sheet, best_status
        else:
            sheet, status = match_route_with_emetteur(
                lot, prefix, emetteur, routing_table, sheet_emetteur_filter
            )

        # Build debug note
        allowed_emetteurs = sheet_emetteur_filter.get(sheet or "", None) if sheet else None
        if status == "ROUTING_EMETTEUR_MISMATCH":
            note = (
                f"emetteur={emetteur} not in allowed set "
                f"{allowed_emetteurs} for sheet={sheet}"
            )
        elif status == "ROUTING_UNMATCHED":
            note = f"no sheet found for lot={lot} prefix={prefix}"
        elif status == "ROUTING_AMBIGUOUS":
            note = f"multiple candidate sheets for lot={lot} prefix={prefix} emetteur={emetteur}"
        else:
            note = ""

        routed_sheets.append(sheet)
        routed_statuses.append(status)
        routed_notes.append(note)

    df["gf_sheet_name"]       = routed_sheets
    df["routing_status"]      = routed_statuses
    df["routing_debug_notes"] = routed_notes

    return df


# ─────────────────────────────────────────────────────────────
# 5. GF SHEET STRUCTURE PARSING
# ─────────────────────────────────────────────────────────────

GF_DATA_START_ROW = 10
GF_HEADER_ROW = 7
GF_APPROVER_ROW = 8
GF_SUBHEADER_ROW = 9


def read_all_gf_sheet_structures(gf_filepath: str,
                                  sheet_names: List[str]) -> Dict[str, Dict]:
    """Batch-read all GF sheet structures in a single file open."""
    wb = openpyxl.load_workbook(gf_filepath, read_only=True)
    return {sn: _parse_sheet_structure_from_wb(wb, sn) for sn in sheet_names}


def read_gf_sheet_structure(gf_filepath: str, sheet_name: str) -> Dict:
    wb = openpyxl.load_workbook(gf_filepath, read_only=True)
    return _parse_sheet_structure_from_wb(wb, sheet_name)


def _parse_sheet_structure_from_wb(wb, sheet_name: str) -> Dict:
    if sheet_name not in wb.sheetnames:
        return {"approvers": [], "data_start_row": GF_DATA_START_ROW,
                "col_map": {}}
    return _parse_sheet_structure(wb[sheet_name])


def _parse_sheet_structure(ws) -> Dict:
    """
    Parse a GF sheet's header structure.
    Returns:
      approvers       : [{name, date_col, num_col, statut_col}]
      data_start_row  : first data row (1-indexed)
      col_map         : {canonical_col_name: 0-indexed column}
      header_row      : detected header row (0-indexed)
    """
    rows = list(ws.iter_rows(min_row=1,
                              max_row=min(GF_DATA_START_ROW + 2, ws.max_row or 12),
                              values_only=True))

    # --- Detect header row (row containing 'DOCUMENT' or 'TITRE') ---
    header_row_idx = GF_HEADER_ROW - 1  # default
    for r_idx, row in enumerate(rows):
        row_vals = [str(c).upper().strip() if c else "" for c in row]
        if "DOCUMENT" in row_vals or "TITRE" in row_vals:
            header_row_idx = r_idx
            break

    # --- Build col_map from header row ---
    COL_ALIASES = {
        "DOCUMENT": "document",
        "TITRE": "titre",
        "DATE DIFFUSION": "date_diffusion",
        "LOT": "lot",
        "TYPE DOC": "type_doc",
        "N° DOC": "numero",
        "NO DOC": "numero",
        "NUMERO": "numero",
        "N DOC": "numero",
        "IND": "indice",
        "INDICE": "indice",
        "NIV": "niveau",
        "NIVEAU": "niveau",
        "ZONE": "zone",
        "TYPE": "type_fichier",
        "ANCIEN": "ancien",
        "N°BDX": "bdx",
        "DATE RÉCEPTION": "date_reception",
        "VISA\nGLOBAL": "visa_global",
        "VISA GLOBAL": "visa_global",
        "APPROBATEURS": "approbateurs",
    }

    col_map: Dict[str, int] = {}
    if header_row_idx < len(rows):
        for col_idx, val in enumerate(rows[header_row_idx]):
            if val is None:
                continue
            key = str(val).upper().strip()
            if key in COL_ALIASES:
                col_map[COL_ALIASES[key]] = col_idx

    # --- Parse approver row (one row after header) ---
    approver_row_idx = header_row_idx + 1
    approvers = []
    if approver_row_idx < len(rows):
        approver_row = rows[approver_row_idx]
        for i, val in enumerate(approver_row):
            if val is not None and str(val).strip():
                name = str(val).strip()
                # Skip if it's actually a sub-header (DATE / N° / STATUT)
                # or a known GF column label that appears in the approver row
                if name.upper() in ("DATE", "N°", "STATUT", "N", "DATE RÉPONSE"):
                    continue
                # Skip known GF extra-column labels that are not approvers
                if name.upper() in ("OBSERVATIONS", "DISCREPANCES GED", "DISCREPANCES",
                                    "ANNEXES", "REMARQUES", "COMMENTAIRES"):
                    continue
                approvers.append({
                    "name": name,
                    "date_col": i,
                    "num_col": i + 1,
                    "statut_col": i + 2,
                })

    # --- Detect data start row ---
    data_start = GF_DATA_START_ROW
    for r_idx in range(header_row_idx + 3, len(rows)):
        row = rows[r_idx]
        if row and row[0] is not None and str(row[0]).strip() not in (
            "DOCUMENT", "", "None", "S U I V I    D E S    V I S A "
        ):
            data_start = r_idx + 1
            break

    return {
        "approvers": approvers,
        "data_start_row": data_start,
        "col_map": col_map,
        "header_row": header_row_idx,
        "total_cols": len(rows[header_row_idx]) if header_row_idx < len(rows) else 0,
        "base_col_count": col_map.get("approbateurs", 15),
    }


# ─────────────────────────────────────────────────────────────
# 6. GF APPROVER NAME MAPPING
# ─────────────────────────────────────────────────────────────

def build_approver_match_map(gf_approvers, ged_canonical_approvers):
    return {a: match_gf_approver_to_ged(a, ged_canonical_approvers)
            for a in gf_approvers}


def match_gf_approver_to_ged(gf_approver, ged_canonical_approvers):
    GF_TO_GED_KEYWORDS = {
        "MOEX":         ["Maître d'Oeuvre EXE"],
        "GEMO":         ["Maître d'Oeuvre EXE"],
        "ARCHI":        ["ARCHITECTE"],
        "BET EGIS":     ["BET CVC", "BET Electricité", "BET Plomberie", "BET Structure"],
        "ACOUSTICIEN":  ["BET Acoustique"],
        "AMO HQE":      ["AMO HQE"],
        "BC SOCOTEC":   ["Bureau de Contrôle"],
        "SOCOTEC":      ["Bureau de Contrôle"],
        "BET ASC":      ["BET Ascenseur"],
        "BET EV":       ["BET EV"],
        "BET SPK":      ["BET SPK"],
        "BET VRD":      ["BET VRD"],
    }
    upper = gf_approver.upper()
    for kw, candidates in GF_TO_GED_KEYWORDS.items():
        if kw.upper() in upper:
            for c in candidates:
                if c in ged_canonical_approvers:
                    return c
    for ga in ged_canonical_approvers:
        if ga.upper() in upper or upper in ga.upper():
            return ga
    return None


def build_gf_to_ged_map(ged_canonical_approvers: List[str]) -> Dict[str, List[str]]:
    GF_TO_GED = {
        "MOEX GEMO":          ["Maître d'Oeuvre EXE"],
        "GEMO":               ["Maître d'Oeuvre EXE"],
        "MOEX":               ["Maître d'Oeuvre EXE"],
        "MOE":                ["Maître d'Oeuvre EXE"],
        "ARCHI MOX":          ["ARCHITECTE"],
        "ARCHI":              ["ARCHITECTE"],
        "ARCHITECTE":         ["ARCHITECTE"],
        "BET EGIS":           ["BET CVC", "BET Electricité", "BET Plomberie",
                               "BET Structure", "BET Façade"],
        "BET EGIS CVC":       ["BET CVC"],
        "BET EGIS PLB":       ["BET Plomberie"],
        "BET EGIS GTB":       ["BET CVC"],
        "BET EGIS ELEC":      ["BET Electricité"],
        "ACOUSTICIEN AVLS":   ["BET Acoustique"],
        "ACOUSTICIEN":        ["BET Acoustique"],
        "AMO HQE LE SOMMER":  ["AMO HQE"],
        "AMO HQE":            ["AMO HQE"],
        "AMO":                ["AMO HQE"],
        "BC SOCOTEC":         ["Bureau de Contrôle"],
        "BC":                 ["Bureau de Contrôle"],
        "BET ASC":            ["BET Ascenseur"],
        "BET EV":             ["BET EV"],
        "BET SPK":            ["BET SPK"],
        "BET VRD":            ["BET VRD"],
        "BET POL":            ["BET POL"],
        "TERRELL":            ["TERRELL"],
    }
    result = {}
    for gf_name, candidates in GF_TO_GED.items():
        matched = [c for c in candidates if c in ged_canonical_approvers]
        result[gf_name] = matched if matched else candidates
    return result


# ─────────────────────────────────────────────────────────────
# 7. DEBUG: ROUTING SUMMARY
# ─────────────────────────────────────────────────────────────

def write_routing_summary(
    output_path: str,
    routed_df: pd.DataFrame,
    routing_table: "RoutingTable",
):
    """
    Write debug/routing_summary.xlsx showing:
      - Per-sheet: row_count, distinct_emetteurs, distinct_lots
        mapped_emetteur (from SHEET_EMETTEUR_FILTER), ambiguous/unmatched counts
      - Routing table entries
      - Ambiguous / unmatched / emetteur-mismatch docs
    """
    from config_loader import SHEET_EMETTEUR_FILTER
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Per-sheet summary ──
    ws1 = wb.active
    ws1.title = "Per-Sheet Summary"
    headers = [
        "Sheet Name", "Mapped Emetteur(s)", "Routed Row Count",
        "Distinct Emetteurs In Sheet", "Distinct Lots In Sheet",
        "Ambiguous Row Count", "Unmatched Row Count",
        "Emetteur Mismatch Count", "Sample Emetteurs Found",
    ]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")

    # Include all known sheets (even if zero rows)
    all_sheets = set(routing_table._table.values().__class__.__mro__)
    all_sheet_names: set = set()
    for sheets in routing_table._table.values():
        all_sheet_names.update(sheets)

    routed_only = routed_df[routed_df["gf_sheet_name"].notna()].copy()
    row = 2
    for sheet_name in sorted(all_sheet_names):
        grp = routed_only[routed_only["gf_sheet_name"] == sheet_name]
        emetteurs = sorted(grp["emetteur"].dropna().unique().tolist())
        lots = sorted(grp["lot_normalized"].dropna().unique().tolist())
        ambiguous_cnt = (grp["routing_status"] == "ROUTING_AMBIGUOUS").sum()
        mismatch_cnt = (grp["routing_status"] == "ROUTING_EMETTEUR_MISMATCH").sum()

        # Look up expected emetteurs from filter
        expected = SHEET_EMETTEUR_FILTER.get(sheet_name)
        expected_str = ", ".join(sorted(expected)) if expected else "(any)"

        ws1.cell(row=row, column=1, value=sheet_name)
        ws1.cell(row=row, column=2, value=expected_str)
        ws1.cell(row=row, column=3, value=len(grp))
        ws1.cell(row=row, column=4, value=len(emetteurs))
        ws1.cell(row=row, column=5, value=", ".join(str(l) for l in lots[:20]))
        ws1.cell(row=row, column=6, value=int(ambiguous_cnt))
        ws1.cell(row=row, column=7, value=0)  # unmatched in sheet — not applicable
        ws1.cell(row=row, column=8, value=int(mismatch_cnt))
        ws1.cell(row=row, column=9, value=", ".join(str(e) for e in emetteurs[:10]))

        # Highlight sheets with emetteur mismatches
        if mismatch_cnt > 0:
            for col in range(1, 10):
                ws1.cell(row=row, column=col).fill = PatternFill("solid",
                                                                  fgColor="FFD966")
        row += 1

    # Unmatched summary row
    unmatched_df = routed_df[routed_df["routing_status"] == "ROUTING_UNMATCHED"]
    ws1.cell(row=row + 1, column=1, value="[UNMATCHED]")
    ws1.cell(row=row + 1, column=3, value=len(unmatched_df))

    for col, width in zip(range(1, 10), [40, 20, 15, 18, 50, 15, 15, 18, 40]):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Routing table ──
    ws2 = wb.create_sheet("Routing Table")
    for i, h in enumerate(["Lot Num", "Building Prefix", "Sheet(s)"], 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
    row = 2
    for lot_num, prefix, sheets in routing_table.all_entries():
        ws2.cell(row=row, column=1, value=lot_num)
        ws2.cell(row=row, column=2, value=prefix or "(wildcard)")
        ws2.cell(row=row, column=3, value=", ".join(sheets))
        if len(sheets) > 1:
            for col in range(1, 4):
                ws2.cell(row=row, column=col).fill = PatternFill("solid",
                                                                  fgColor="FFD966")
        row += 1
    for col, w in zip(range(1, 4), [12, 16, 60]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Unmatched docs ──
    ws3 = wb.create_sheet("Unmatched Docs")
    unmatched = routed_df[routed_df["routing_status"] == "ROUTING_UNMATCHED"].copy()
    for i, h in enumerate(["routing_status", "emetteur", "lot", "lot_normalized",
                            "lot_prefix", "libelle_du_document"], 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C55A11")

    for r_idx, (_, row_data) in enumerate(unmatched.iterrows(), 2):
        ws3.cell(row=r_idx, column=1, value=row_data.get("routing_status", ""))
        ws3.cell(row=r_idx, column=2, value=row_data.get("emetteur", ""))
        ws3.cell(row=r_idx, column=3, value=str(row_data.get("lot", "")))
        ws3.cell(row=r_idx, column=4, value=str(row_data.get("lot_normalized", "")))
        ws3.cell(row=r_idx, column=5, value=str(row_data.get("lot_prefix", "")))
        ws3.cell(row=r_idx, column=6,
                 value=str(row_data.get("libelle_du_document", ""))[:60])
    for col, w in zip(range(1, 7), [20, 15, 12, 12, 12, 60]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Emetteur Mismatch docs ──
    ws4 = wb.create_sheet("Emetteur Mismatch")
    mismatch_df = routed_df[
        routed_df["routing_status"] == "ROUTING_EMETTEUR_MISMATCH"
    ].copy()
    mismatch_headers = [
        "gf_sheet_name", "emetteur", "lot", "lot_normalized",
        "lot_prefix", "routing_debug_notes",
    ]
    for i, h in enumerate(mismatch_headers, 1):
        c = ws4.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7F0000")
    for r_idx, (_, row_data) in enumerate(mismatch_df.iterrows(), 2):
        ws4.cell(row=r_idx, column=1, value=str(row_data.get("gf_sheet_name", "")))
        ws4.cell(row=r_idx, column=2, value=row_data.get("emetteur", ""))
        ws4.cell(row=r_idx, column=3, value=str(row_data.get("lot", "")))
        ws4.cell(row=r_idx, column=4, value=str(row_data.get("lot_normalized", "")))
        ws4.cell(row=r_idx, column=5, value=str(row_data.get("lot_prefix", "")))
        ws4.cell(row=r_idx, column=6,
                 value=str(row_data.get("routing_debug_notes", ""))[:120])
    for col, w in zip(range(1, 7), [40, 12, 12, 12, 12, 80]):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 5: Ambiguous docs ──
    ws5 = wb.create_sheet("Ambiguous Docs")
    ambiguous = routed_df[routed_df["routing_status"] == "ROUTING_AMBIGUOUS"].copy()
    for i, h in enumerate(["gf_sheet_name", "emetteur", "lot",
                            "lot_normalized", "lot_prefix"], 1):
        c = ws5.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7F0000")
    for r_idx, (_, row_data) in enumerate(ambiguous.iterrows(), 2):
        ws5.cell(row=r_idx, column=1, value=str(row_data.get("gf_sheet_name", "")))
        ws5.cell(row=r_idx, column=2, value=row_data.get("emetteur", ""))
        ws5.cell(row=r_idx, column=3, value=str(row_data.get("lot", "")))
        ws5.cell(row=r_idx, column=4, value=str(row_data.get("lot_normalized", "")))
        ws5.cell(row=r_idx, column=5, value=str(row_data.get("lot_prefix", "")))

    wb.save(output_path)
