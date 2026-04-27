"""
consultant_match_report.py
JANSA VISASIST — Consultant Match Report Writer
Version 2.0 — April 2026

Generates output/consultant_match_report.xlsx with 4 sheets:
  1. MATCHED      — high/medium/low confidence matches (colour-coded by confidence)
  2. UNMATCHED    — no GED link found
  3. AMBIGUOUS    — multiple GED candidates for same NUMERO, unresolved
  4. SUMMARY      — per-source counts, match rates, confidence breakdown

New in v2.0:
  - Trace columns: consultant_match_method, consultant_match_confidence,
    candidate_count, candidate_ids_considered, winning_candidate_id,
    match_rationale, date_distance_days, indice_fallback_used, deterministic_match_used
  - MATCHED sheet rows are colour-coded: HIGH=green, MEDIUM=light blue, LOW=amber
  - SUMMARY sheet includes confidence breakdown and enrichment eligibility counts
"""

import logging
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

DARK_TEAL     = "1F6B75"
DARK_GREEN    = "375623"
DARK_RED      = "9C0006"
DARK_AMBER    = "7E5109"
DARK_BLUE     = "1F4E79"
WHITE_HEX     = "FFFFFF"

# Row fill colours by confidence tier
FILL_HIGH      = "C6EFCE"   # light green  — HIGH confidence
FILL_MEDIUM    = "BDD7EE"   # light blue   — MEDIUM confidence
FILL_LOW       = "FFEB9C"   # light amber  — LOW confidence
FILL_AMBIGUOUS = "FFC7CE"   # light red    — AMBIGUOUS_UNRESOLVED
FILL_UNMATCHED = "FFC7CE"   # light red    — UNMATCHED
FILL_HEADER    = DARK_TEAL

CONFIDENCE_FILL = {
    "HIGH":                FILL_HIGH,
    "MEDIUM":              FILL_MEDIUM,
    "LOW":                 FILL_LOW,
    "AMBIGUOUS_UNRESOLVED": FILL_AMBIGUOUS,
}

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Consultant-side columns
_COL_CONSULTANT = [
    ("consultant_source",           16, "Consultant Source"),
    ("rapport_id",                  30, "Rapport ID"),
    ("c_ref_doc",                   52, "REF_DOC (Consultant)"),
    ("c_numero",                    12, "NUMERO"),
    ("c_indice",                     8, "INDICE"),
    ("c_statut_norm",               14, "STATUT_NORM"),
    ("c_date_fiche",                16, "DATE_FICHE"),
    ("c_commentaire",               45, "COMMENTAIRE"),
]

# GED matched-document columns
_COL_GED = [
    ("matched_doc_id",              38, "Matched GED doc_id"),
    ("matched_numero",              12, "Matched NUMERO"),
    ("matched_indice",               8, "Matched INDICE"),
    ("matched_lot",                 12, "Matched LOT"),
    ("matched_emetteur",            18, "Matched EMETTEUR"),
    ("matched_titre",               40, "Matched TITRE"),
    ("matched_type_doc",            12, "Type Doc"),
    ("matched_specialite",          14, "Specialite"),
]

# Match quality / trace columns (new in v2.0)
_COL_TRACE = [
    ("consultant_match_method",     30, "Match Method"),
    ("consultant_match_confidence", 16, "Confidence"),
    ("candidate_count",              8, "# Cands"),
    ("candidate_ids_considered",    55, "Candidates Considered"),
    ("winning_candidate_id",        40, "Winning Candidate ID"),
    ("match_rationale",             65, "Match Rationale"),
    ("date_distance_days",          14, "Date Gap (days)"),
    ("indice_fallback_used",        10, "Indice Fallback"),
    ("deterministic_match_used",    12, "Deterministic"),
]

REPORT_COLUMNS = _COL_CONSULTANT + _COL_GED + _COL_TRACE

SUMMARY_COLUMNS = [
    ("source",                      20, "Consultant Source"),
    ("total",                       10, "Total Rows"),
    ("matched",                     10, "Matched"),
    ("ambiguous",                   12, "Ambiguous"),
    ("unmatched",                   12, "Unmatched"),
    ("match_rate_pct",              14, "Match Rate %"),
    ("conf_high",                   10, "HIGH"),
    ("conf_medium",                 10, "MEDIUM"),
    ("conf_low",                    10, "LOW"),
    ("enrich_eligible",             14, "Enrich Eligible"),
    ("top_unmatched_reason",        55, "Top Unmatched Reason"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hfill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _write_cell(ws, row, col, value, font=None, fill=None, fmt=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    return cell


def _write_header_row(ws, columns, row=1):
    """Write header row with dark-teal background."""
    header_font = Font(bold=True, color=WHITE_HEX, name="Calibri", size=10)
    header_fill = _hfill(FILL_HEADER)
    for col_idx, (_, width, label) in enumerate(columns, 1):
        cell = _write_cell(ws, row, col_idx, label, font=header_font, fill=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[row].height = 16


def _set_col_widths(ws, columns):
    for col_idx, (_, width, _label) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_data_rows(
    ws,
    records,
    columns,
    row_fill=None,
    start_row=2,
    confidence_field: str = None,
):
    """
    Write data rows.

    If confidence_field is set, each row's fill is determined by the
    confidence value in that field (HIGH/MEDIUM/LOW/AMBIGUOUS_UNRESOLVED).
    Otherwise row_fill is applied uniformly (pass None for no fill).
    """
    data_font = Font(name="Calibri", size=9)
    for row_idx, rec in enumerate(records, start_row):
        # Determine fill for this row
        if confidence_field:
            conf = str(rec.get(confidence_field, "")).strip().upper()
            hex_color = CONFIDENCE_FILL.get(conf, FILL_MEDIUM)
            fill = _hfill(hex_color)
        else:
            fill = row_fill

        for col_idx, (field, _, _) in enumerate(columns, 1):
            val = rec.get(field, "")
            if val is None or str(val).lower() in ("none", "nan"):
                val = ""
            _write_cell(
                ws, row_idx, col_idx,
                str(val) if val != "" else "",
                font=data_font,
                fill=fill,
            )
        ws.row_dimensions[row_idx].height = 13


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def write_match_report(
    match_results: dict,
    output_path:   Path,
) -> None:
    """
    Write the 4-sheet match report workbook.

    Parameters
    ----------
    match_results : output of consultant_matcher.match_all_consultants()
    output_path   : destination .xlsx path
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    matched   = match_results["matched"]
    unmatched = match_results["unmatched"]
    ambiguous = match_results["ambiguous"]
    stats     = match_results["stats"]

    # ── Sheet 1: MATCHED ──────────────────────────────────────────────────
    # Rows are colour-coded by confidence (HIGH=green, MEDIUM=blue, LOW=amber)
    ws_matched = wb.create_sheet("MATCHED")
    _write_header_row(ws_matched, REPORT_COLUMNS)
    _set_col_widths(ws_matched, REPORT_COLUMNS)
    _write_data_rows(
        ws_matched, matched, REPORT_COLUMNS,
        confidence_field="consultant_match_confidence",
    )
    logger.info("MATCHED sheet: %d rows", len(matched))

    # ── Sheet 2: UNMATCHED ────────────────────────────────────────────────
    ws_unmatched = wb.create_sheet("UNMATCHED")
    _write_header_row(ws_unmatched, REPORT_COLUMNS)
    _set_col_widths(ws_unmatched, REPORT_COLUMNS)
    _write_data_rows(ws_unmatched, unmatched, REPORT_COLUMNS,
                     row_fill=_hfill(FILL_UNMATCHED))
    logger.info("UNMATCHED sheet: %d rows", len(unmatched))

    # ── Sheet 3: AMBIGUOUS ────────────────────────────────────────────────
    ws_ambiguous = wb.create_sheet("AMBIGUOUS")
    _write_header_row(ws_ambiguous, REPORT_COLUMNS)
    _set_col_widths(ws_ambiguous, REPORT_COLUMNS)
    _write_data_rows(ws_ambiguous, ambiguous, REPORT_COLUMNS,
                     row_fill=_hfill(FILL_AMBIGUOUS))
    logger.info("AMBIGUOUS sheet: %d rows", len(ambiguous))

    # ── Sheet 4: SUMMARY ──────────────────────────────────────────────────
    ws_summary = wb.create_sheet("SUMMARY")
    _write_header_row(ws_summary, SUMMARY_COLUMNS)
    _set_col_widths(ws_summary, SUMMARY_COLUMNS)

    # Pre-compute per-source confidence breakdown from matched rows
    from collections import defaultdict
    conf_by_source = defaultdict(Counter)
    for r in matched:
        src  = r.get("consultant_source", "")
        conf = r.get("consultant_match_confidence", "")
        conf_by_source[src][conf] += 1

    summary_rows = []
    for src, s in sorted(stats.items()):
        total       = s["total"]
        match_rate  = f"{s['matched']/total*100:.1f}%" if total else "n/a"
        conf_counts = conf_by_source.get(src, {})
        n_high      = conf_counts.get("HIGH", 0)
        n_medium    = conf_counts.get("MEDIUM", 0)
        n_low       = conf_counts.get("LOW", 0)
        n_enrich    = n_high + n_medium   # LOW excluded from enrichment

        # Top unmatched reason
        src_unmatched = [r for r in unmatched if r.get("consultant_source") == src]
        reasons   = [r.get("match_rationale", r.get("rationale", "")) for r in src_unmatched]
        top_reason = Counter(reasons).most_common(1)[0][0] if reasons else ""

        summary_rows.append({
            "source":               src,
            "total":                total,
            "matched":              s["matched"],
            "ambiguous":            s["ambiguous"],
            "unmatched":            s["unmatched"],
            "match_rate_pct":       match_rate,
            "conf_high":            n_high,
            "conf_medium":          n_medium,
            "conf_low":             n_low,
            "enrich_eligible":      n_enrich,
            "top_unmatched_reason": top_reason,
        })

    # Totals row
    total_all     = sum(s["total"]     for s in stats.values())
    total_matched = sum(s["matched"]   for s in stats.values())
    total_amb     = sum(s["ambiguous"] for s in stats.values())
    total_unm     = sum(s["unmatched"] for s in stats.values())
    all_conf      = Counter()
    for cc in conf_by_source.values():
        all_conf.update(cc)

    summary_rows.append({
        "source":               "TOTAL",
        "total":                total_all,
        "matched":              total_matched,
        "ambiguous":            total_amb,
        "unmatched":            total_unm,
        "match_rate_pct":       f"{total_matched/total_all*100:.1f}%" if total_all else "n/a",
        "conf_high":            all_conf.get("HIGH", 0),
        "conf_medium":          all_conf.get("MEDIUM", 0),
        "conf_low":             all_conf.get("LOW", 0),
        "enrich_eligible":      all_conf.get("HIGH", 0) + all_conf.get("MEDIUM", 0),
        "top_unmatched_reason": f"{total_unm} unmatched rows across all sources",
    })

    data_font  = Font(name="Calibri", size=9)
    total_font = Font(name="Calibri", size=9, bold=True)
    for row_idx, rec in enumerate(summary_rows, 2):
        is_total = rec.get("source") == "TOTAL"
        f    = total_font if is_total else data_font
        fill = _hfill(DARK_BLUE) if is_total else None
        for col_idx, (field, _, _) in enumerate(SUMMARY_COLUMNS, 1):
            val  = rec.get(field, "")
            cell = _write_cell(ws_summary, row_idx, col_idx,
                               str(val) if val != "" else "", font=f, fill=fill)
            if is_total:
                cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE_HEX)
        ws_summary.row_dimensions[row_idx].height = 14

    logger.info("SUMMARY sheet: %d source rows", len(summary_rows))

    # ── Save ──────────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Match report saved: %s", output_path)
