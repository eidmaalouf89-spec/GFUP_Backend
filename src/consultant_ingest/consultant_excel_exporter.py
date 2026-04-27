"""
consultant_excel_exporter.py
JANSA VISASIST — Consultant Report Builder
Version 1.0 — April 2026

Creates the final 4-sheet Excel workbook with controlled formatting.

Sheets:
  1. RAPPORT_LE_SOMMER
  2. RAPPORT_AVLS
  3. RAPPORT_TERRELL
  4. RAPPORT_SOCOTEC
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .consultant_transformers import (
    LS_COLUMNS,
    AVLS_COLUMNS,
    TERRELL_COLUMNS,
    SOCOTEC_COLUMNS,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants — dark teal header matching reference workbook
# ---------------------------------------------------------------------------

HEADER_FILL_HEX = "1F6B75"   # dark teal
HEADER_FONT_COLOR = "FFFFFF"  # white text

# Approximate column width heuristics (in Excel character units)
_COL_WIDTH_DEFAULTS = {
    "SOURCE":       12,
    "RAPPORT_ID":   28,
    "DATE_FICHE":   14,
    "NUMERO":       12,
    "INDICE":        8,
    "REF_DOC":      48,
    "STATUT_NORM":  14,
    "COMMENTAIRE":  50,
    "OBSERVATIONS": 50,
    "PDF_PAGE":     10,
    "UPSERT_KEY":   48,
    "LAST_UPDATED": 20,
    # Le Sommer extras
    "LOT_TYPE":     12,
    "SECTION":      12,
    "TABLE_TYPE":   14,
    # AVLS extras
    "LOT_LABEL":    18,
    "LOT_NUM":      12,
    "N_VISA":       10,
    "REVIEWER":     22,
    # Terrell extras
    "BAT":          10,
    "LOT":          12,
    "SPECIALITE":   14,
    "TYPE_DOC":     12,
    "NIVEAU":       10,
    "DATE_SOURCE":  14,
    "DESIGNATION":  40,
    # Socotec extras
    "CT_REF":       26,
    "OBS_NUM":      12,
}

SHEET_DEFS = [
    ("RAPPORT_LE_SOMMER", LS_COLUMNS),
    ("RAPPORT_AVLS",       AVLS_COLUMNS),
    ("RAPPORT_TERRELL",    TERRELL_COLUMNS),
    ("RAPPORT_SOCOTEC",    SOCOTEC_COLUMNS),
]


# ---------------------------------------------------------------------------
# Core export function
# ---------------------------------------------------------------------------

def export_workbook(
    ls_rows:      list[dict],
    avls_rows:    list[dict],
    terrell_rows: list[dict],
    socotec_rows: list[dict],
    output_path:  Path,
) -> None:
    """
    Write the final 4-sheet workbook to output_path.

    Parameters
    ----------
    ls_rows, avls_rows, terrell_rows, socotec_rows : transformed row lists
    output_path : destination .xlsx path
    """
    wb = openpyxl.Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    all_data = [ls_rows, avls_rows, terrell_rows, socotec_rows]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL_HEX,
    )
    header_font = Font(
        bold=True,
        color=HEADER_FONT_COLOR,
        name="Calibri",
        size=11,
    )
    data_font = Font(name="Calibri", size=10)

    for (sheet_name, columns), rows in zip(SHEET_DEFS, all_data):
        ws = wb.create_sheet(title=sheet_name)

        # ── Write header row ──────────────────────────────────────────────
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )

        # ── Write data rows ───────────────────────────────────────────────
        for row_idx, record in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = record.get(col_name, "")
                # Never write Python literal None/nan as string
                if value is None or str(value).lower() in ("none", "nan"):
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
                cell.font = data_font
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # ── Set column widths ─────────────────────────────────────────────
        for col_idx, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            width = _COL_WIDTH_DEFAULTS.get(col_name, 18)
            ws.column_dimensions[col_letter].width = width

        # ── Freeze pane at A2 ─────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Row height for header ─────────────────────────────────────────
        ws.row_dimensions[1].height = 20

        logger.info(
            "Sheet '%s': %d data rows, %d columns",
            sheet_name, len(rows), len(columns)
        )

    # ── Save workbook ─────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Workbook saved: %s", output_path)
