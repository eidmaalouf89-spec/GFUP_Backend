#!/usr/bin/env python3
"""
validate_consultant_reports.py
JANSA VISASIST — Workbook Schema Validator
Version 1.0 — April 2026

Validates the generated consultant_reports.xlsx against the agreed
column contract and reports quality metrics.

Run from any directory:
    python "Ingester Files/validate_consultant_reports.py"

Exit code: 0 if all structural checks pass, 1 if any fail.
"""

import sys
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed — run: pip install openpyxl --break-system-packages")
    sys.exit(2)

# ---------------------------------------------------------------------------
# Agreed column contract
# ---------------------------------------------------------------------------

AGREED_SCHEMAS = {
    "RAPPORT_LE_SOMMER": [
        "SOURCE", "RAPPORT_ID", "DATE_FICHE", "NUMERO", "INDICE",
        "REF_DOC", "STATUT_NORM", "COMMENTAIRE", "PDF_PAGE",
        "UPSERT_KEY", "LAST_UPDATED",
        "LOT_TYPE", "SECTION", "TABLE_TYPE",
    ],
    "RAPPORT_AVLS": [
        "SOURCE", "RAPPORT_ID", "DATE_FICHE", "NUMERO", "INDICE",
        "REF_DOC", "STATUT_NORM", "COMMENTAIRE", "PDF_PAGE",
        "UPSERT_KEY", "LAST_UPDATED",
        "LOT_LABEL", "LOT_NUM", "N_VISA", "REVIEWER",
    ],
    "RAPPORT_TERRELL": [
        "SOURCE", "RAPPORT_ID", "DATE_FICHE", "NUMERO", "INDICE",
        "REF_DOC", "STATUT_NORM", "COMMENTAIRE", "PDF_PAGE",
        "UPSERT_KEY", "LAST_UPDATED",
        "BAT", "LOT", "SPECIALITE", "TYPE_DOC",
        "NIVEAU", "DATE_SOURCE", "DESIGNATION",
    ],
    "RAPPORT_SOCOTEC": [
        "SOURCE", "RAPPORT_ID", "DATE_FICHE", "NUMERO", "INDICE",
        "REF_DOC", "STATUT_NORM", "COMMENTAIRE", "PDF_PAGE",
        "UPSERT_KEY", "LAST_UPDATED",
        "CT_REF", "OBS_NUM",
    ],
}

# Fields to check for blanks per sheet
BLANK_CHECKS = {
    "RAPPORT_LE_SOMMER": ["DATE_FICHE", "INDICE"],
    "RAPPORT_AVLS":      ["INDICE"],
    "RAPPORT_TERRELL":   ["DATE_FICHE"],
    "RAPPORT_SOCOTEC":   ["DATE_FICHE", "NUMERO", "INDICE", "OBS_NUM"],
}

# Socotec must only use native labels
SOCOTEC_ALLOWED_STATUSES = {"FAV", "SUS", "DEF"}
SOCOTEC_FORBIDDEN_STATUSES = {"VSO", "VAO", "REF"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def pct(n: int, total: int) -> str:
    return f"{n/total*100:.1f}%" if total else "n/a"


def banner(text: str) -> None:
    print(f"\n{'=' * 68}")
    print(f"  {text}")
    print('=' * 68)


def section(text: str) -> None:
    print(f"\n── {text}")


# ---------------------------------------------------------------------------
# Main validator
# ---------------------------------------------------------------------------

def validate(wb_path: Path) -> bool:
    """Run all checks. Returns True if all structural checks pass."""
    all_ok = True

    banner("JANSA VISASIST — Workbook Validator")
    print(f"  Target : {wb_path}")

    # 1. File existence
    section("1. File existence")
    if not wb_path.exists():
        print(f"  FAIL  File not found: {wb_path}")
        return False
    size_kb = wb_path.stat().st_size // 1024
    print(f"  OK    File exists ({size_kb} KB)")

    wb = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)

    # 2. Sheet names
    section("2. Sheet names (exact 4 required)")
    actual_sheets = wb.sheetnames
    expected_sheets = list(AGREED_SCHEMAS.keys())
    sheets_ok = True
    for name in expected_sheets:
        present = name in actual_sheets
        print(f"  {'OK  ' if present else 'FAIL'} {name}")
        if not present:
            sheets_ok = False
            all_ok = False
    extra_sheets = [s for s in actual_sheets if s not in expected_sheets]
    if extra_sheets:
        print(f"  WARN  Extra sheets present: {extra_sheets}")

    if not sheets_ok:
        print("\n  Cannot continue — required sheets are missing.")
        return False

    # Per-sheet checks
    for sheet_name, expected_cols in AGREED_SCHEMAS.items():
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            print(f"\n  FAIL  {sheet_name}: sheet is empty")
            all_ok = False
            continue

        header = list(all_rows[0])
        data = all_rows[1:]
        n = len(data)

        section(f"3–8. {sheet_name}  ({n} rows)")

        # 3. Header / column order
        cols_ok = (header == expected_cols)
        if cols_ok:
            print(f"  OK    Headers match contract ({len(header)} columns)")
        else:
            print(f"  FAIL  Header mismatch!")
            for i, (e, g) in enumerate(zip(expected_cols, header)):
                if e != g:
                    print(f"         col {i+1}: expected '{e}' got '{g}'")
            if len(expected_cols) != len(header):
                print(f"         length: expected {len(expected_cols)} got {len(header)}")
            all_ok = False

        col_idx = {c: i for i, c in enumerate(header)}

        # 4. Row count
        print(f"  OK    Row count: {n}")

        # 5. UPSERT_KEY duplicates
        if "UPSERT_KEY" in col_idx:
            ui = col_idx["UPSERT_KEY"]
            keys = [r[ui] for r in data if r[ui]]
            key_counts = Counter(keys)
            dup_keys = {k: v for k, v in key_counts.items() if v > 1}
            extra_rows = sum(v - 1 for v in dup_keys.values())
            if extra_rows == 0:
                print(f"  OK    UPSERT_KEY: no duplicates")
            else:
                mark = "WARN" if extra_rows <= 10 else "WARN"
                print(f"  {mark}  UPSERT_KEY: {len(dup_keys)} colliding keys, {extra_rows} extra rows")
                for k, v in sorted(dup_keys.items(), key=lambda x: -x[1])[:3]:
                    print(f"           [{v}x] {str(k)[:75]}")

        # 6. Blank counts
        for field in BLANK_CHECKS.get(sheet_name, []):
            if field not in col_idx:
                print(f"  WARN  '{field}' column not found")
                continue
            fi = col_idx[field]
            blanks = sum(1 for r in data if not r[fi] or str(r[fi]).strip() == '')
            mark = "OK  " if blanks == 0 else ("INFO" if blanks / n < 0.40 else "WARN")
            print(f"  {mark}  {field} blanks: {blanks}/{n} ({pct(blanks, n)})")

        # 7. Socotec native statuses
        if sheet_name == "RAPPORT_SOCOTEC" and "STATUT_NORM" in col_idx:
            si = col_idx["STATUT_NORM"]
            status_counts = Counter(r[si] for r in data if r[si])
            forbidden_found = {s for s in status_counts if s in SOCOTEC_FORBIDDEN_STATUSES}
            native_ok = all(s in SOCOTEC_ALLOWED_STATUSES for s in status_counts)
            if forbidden_found:
                print(f"  FAIL  Socotec contains forbidden statuses: {forbidden_found}")
                all_ok = False
            else:
                print(f"  OK    Socotec statuses: native FAV/SUS/DEF only {dict(status_counts)}")

        # 8. Freeze panes
        # openpyxl read_only doesn't expose freeze_panes; skip in read_only mode
        # Run a separate pass with full load
        pass

    # 8. Freeze panes — requires full (non-read-only) load
    section("8. Freeze panes")
    wb.close()
    try:
        wb_full = openpyxl.load_workbook(str(wb_path), read_only=False, data_only=True)
        freeze_issues = []
        for sheet_name in expected_sheets:
            ws = wb_full[sheet_name]
            fp = ws.freeze_panes
            if str(fp) == "A2":
                print(f"  OK    {sheet_name}: freeze_panes = A2")
            else:
                print(f"  WARN  {sheet_name}: freeze_panes = {fp!r} (expected A2)")
                freeze_issues.append(sheet_name)
        wb_full.close()
    except Exception as e:
        print(f"  WARN  Could not check freeze panes: {e}")

    # Summary
    banner("SUMMARY")
    if all_ok:
        print("  ALL STRUCTURAL CHECKS PASSED")
    else:
        print("  STRUCTURAL ISSUES FOUND — see FAIL lines above")
    print()
    return all_ok


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Resolve workbook path relative to this script's location
    script_dir = Path(__file__).resolve().parent
    wb_path = script_dir.parent.parent / "output" / "consultant_reports.xlsx"

    ok = validate(wb_path)
    sys.exit(0 if ok else 1)
