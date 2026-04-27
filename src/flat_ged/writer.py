"""
writer.py — Output file generation.

Two output paths:

  Single mode (write_flat_ged):
    output/FLAT_GED.xlsx  — 3-sheet styled workbook (GED_RAW_FLAT, GED_OPERATIONS, DEBUG_TRACE)
    output/run_report.json

  Batch mode (write_flat_ged_batch + write_debug_trace_csv):
    output/FLAT_GED.xlsx  — 2-sheet workbook (GED_RAW_FLAT, GED_OPERATIONS)
                            uses write_only mode; data rows written as plain values
                            (headers still styled)
    output/DEBUG_TRACE.csv — full debug audit trail as CSV (400k rows/0.5s vs 47s in Excel)
    output/run_report.json

All per-cell colours on single-mode output are preserved from prototype_v4.py.
"""

import csv
import json
import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment


# ── Style helpers ─────────────────────────────────────────────────────────────

def _HF(c="FFFFFF"):
    return Font(name="Arial", bold=True, color=c, size=10)

def _DF(bold=False, c="000000"):
    return Font(name="Arial", bold=bold, color=c, size=10)

def _IT():
    return Font(name="Arial", italic=True, size=9, color="595959")

_AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_AC = Alignment(horizontal="center", vertical="center")
_AR = Alignment(horizontal="right",  vertical="center")

_P = {
    "hdr_raw":   PatternFill("solid", start_color="1F4E79"),
    "hdr_ops":   PatternFill("solid", start_color="203864"),
    "hdr_dbg":   PatternFill("solid", start_color="4B0082"),
    "OPEN_DOC":  PatternFill("solid", start_color="BDD7EE"),
    "SAS":       PatternFill("solid", start_color="FFE699"),
    "CONSULTANT":PatternFill("solid", start_color="E2EFDA"),
    "MOEX":      PatternFill("solid", start_color="FCE4D6"),
    "PEND_LATE": PatternFill("solid", start_color="C00000"),
    "PEND_DLY":  PatternFill("solid", start_color="FFC000"),
    "ANSWERED":  PatternFill("solid", start_color="70AD47"),
    "NOT_CALLED":PatternFill("solid", start_color="D9D9D9"),
    "RETARD":    PatternFill("solid", start_color="FFCCCC"),
    "AVANCE":    PatternFill("solid", start_color="CCFFCC"),
    "BLOCKING":  PatternFill("solid", start_color="C00000"),
    "DONE":      PatternFill("solid", start_color="70AD47"),
    "UNKNOWN":   PatternFill("solid", start_color="FF6600"),
    "COMPUTED":  PatternFill("solid", start_color="FFF2CC"),
}

_DS_FILL = {
    "PENDING_LATE":     _P["PEND_LATE"],
    "PENDING_IN_DELAY": _P["PEND_DLY"],
    "ANSWERED":         _P["ANSWERED"],
    "NOT_CALLED":       _P["NOT_CALLED"],
}

_ST_FILL = {
    "OPEN_DOC":   _P["OPEN_DOC"],
    "SAS":        _P["SAS"],
    "CONSULTANT": _P["CONSULTANT"],
    "MOEX":       _P["MOEX"],
}


def _write_hdr(ws, cols, fill):
    ws.append(cols)
    for cell in ws[1]:
        cell.font      = _HF()
        cell.fill      = fill
        cell.alignment = _AC
    ws.row_dimensions[1].height = 20


def _row_style(ws, rn, row_fill=None):
    for cell in ws[rn]:
        cell.font      = _DF()
        cell.alignment = _AL
        if row_fill:
            cell.fill = row_fill


def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w


# ── Sheet 1: GED_RAW_FLAT ─────────────────────────────────────────────────────

R_COLS = [
    "numero", "indice", "lot", "emetteur", "titre",
    "approver_raw", "approver_canonical", "actor_type",
    "response_status_raw", "response_status_clean",
    "response_status_code", "response_status_scope",
    "response_date_raw", "response_date",
    "date_status_type", "deadline_raw", "deadline",
    "commentaire", "pj_flag", "is_sas", "raw_trace_key",
]


def _write_raw_flat(ws, raw_flat_rows: list[dict]) -> None:
    ws.freeze_panes = "A2"
    _write_hdr(ws, R_COLS, _P["hdr_raw"])

    ds_col = R_COLS.index("date_status_type")    + 1
    cn_col = R_COLS.index("approver_canonical")  + 1

    for r in raw_flat_rows:
        ws.append([r.get(c) for c in R_COLS])
        rn = ws.max_row
        _row_style(ws, rn)
        ds = r.get("date_status_type", "")
        ws.cell(rn, ds_col).fill = _DS_FILL.get(ds, _P["NOT_CALLED"])
        if ds in ("PENDING_LATE", "PENDING_IN_DELAY"):
            ws.cell(rn, ds_col).font = _DF(bold=True, c="FFFFFF")
        if r.get("approver_canonical") == "UNKNOWN":
            ws.cell(rn, cn_col).fill = _P["UNKNOWN"]

    _widths(ws, {
        "A": 10, "B": 7,  "C": 9,  "D": 10, "E": 46,
        "F": 36, "G": 28, "H": 12,
        "I": 22, "J": 22, "K": 14, "L": 14,
        "M": 36, "N": 14,
        "O": 20, "P": 42, "Q": 14,
        "R": 55, "S": 8,  "T": 8,  "U": 36,
    })


# ── Sheet 2: GED_OPERATIONS ───────────────────────────────────────────────────

O_COLS = [
    "numero", "indice", "lot", "emetteur", "titre",
    "step_order", "step_type", "actor_type",
    "actor_raw", "actor_clean",
    "status_raw", "status_clean", "status_code", "status_scope", "status_family",
    "is_completed", "is_blocking", "requires_new_cycle",
    "submittal_date", "sas_response_date", "response_date", "data_date",
    "global_deadline", "phase_deadline", "deadline_source",
    "retard_avance_days", "retard_avance_status",
    "step_delay_days", "delay_contribution_days", "cumulative_delay_days", "delay_actor",
    "chrono_source",
    "observation", "pj_flag",
    "source_trace", "source_rows", "operation_rule_used",
]


def _write_operations(ws, ops_rows: list[dict]) -> None:
    ws.freeze_panes = "A2"
    _write_hdr(ws, O_COLS, _P["hdr_ops"])

    st_c  = O_COLS.index("step_type")            + 1
    sf_c  = O_COLS.index("status_family")        + 1
    bl_c  = O_COLS.index("is_blocking")          + 1
    dn_c  = O_COLS.index("is_completed")         + 1
    rc_c  = O_COLS.index("requires_new_cycle")   + 1
    rv_c  = O_COLS.index("retard_avance_days")   + 1
    rs_c  = O_COLS.index("retard_avance_status") + 1
    sd_c  = O_COLS.index("step_delay_days")      + 1
    dc_c  = O_COLS.index("delay_contribution_days") + 1
    cu_c  = O_COLS.index("cumulative_delay_days")   + 1
    da_c  = O_COLS.index("delay_actor")          + 1
    pd_c  = O_COLS.index("phase_deadline")       + 1
    dl_c  = O_COLS.index("deadline_source")      + 1
    cs_c  = O_COLS.index("chrono_source")        + 1

    for step in ops_rows:
        ws.append([step.get(c) for c in O_COLS])
        rn    = ws.max_row
        stype = step.get("step_type", "")
        _row_style(ws, rn, _ST_FILL.get(stype))

        ws.cell(rn, st_c).font = _DF(bold=True)

        if step.get("phase_deadline"):
            ws.cell(rn, pd_c).fill = _P["COMPUTED"]
        ws.cell(rn, dl_c).font = _IT()

        sf = step.get("status_family", "")
        if sf == "PENDING":
            ws.cell(rn, sf_c).fill = _P["PEND_LATE"]
            ws.cell(rn, sf_c).font = _DF(bold=True, c="FFFFFF")
        elif sf in ("APPROVED", "OPENED"):
            ws.cell(rn, sf_c).fill = _P["ANSWERED"]
        elif sf == "APPROVED_WITH_REMARKS":
            ws.cell(rn, sf_c).fill = _P["PEND_DLY"]
        elif sf == "REFUSED":
            ws.cell(rn, sf_c).fill = _P["PEND_LATE"]
            ws.cell(rn, sf_c).font = _DF(c="FFFFFF")

        if step.get("is_blocking"):
            ws.cell(rn, bl_c).fill = _P["BLOCKING"]
            ws.cell(rn, bl_c).font = _DF(bold=True, c="FFFFFF")
        if step.get("is_completed"):
            ws.cell(rn, dn_c).fill = _P["DONE"]
            ws.cell(rn, dn_c).font = _DF(bold=True, c="FFFFFF")
        if step.get("requires_new_cycle"):
            ws.cell(rn, rc_c).fill = _P["BLOCKING"]
            ws.cell(rn, rc_c).font = _DF(bold=True, c="FFFFFF")

        rv = step.get("retard_avance_days")
        if isinstance(rv, int):
            ws.cell(rn, rv_c).alignment = _AR
            if rv < 0:
                ws.cell(rn, rv_c).fill = _P["RETARD"]
                ws.cell(rn, rv_c).font = _DF(bold=True, c="C00000")
                ws.cell(rn, rs_c).fill = _P["RETARD"]
                ws.cell(rn, rs_c).font = _DF(bold=True, c="C00000")
            elif rv > 0:
                ws.cell(rn, rv_c).fill = _P["AVANCE"]
                ws.cell(rn, rv_c).font = _DF(bold=True, c="375623")
                ws.cell(rn, rs_c).fill = _P["AVANCE"]
                ws.cell(rn, rs_c).font = _DF(bold=True, c="375623")

        sd2 = step.get("step_delay_days")
        if isinstance(sd2, int):
            ws.cell(rn, sd_c).alignment = _AR
            if sd2 > 0:
                ws.cell(rn, sd_c).fill = _P["RETARD"]
                ws.cell(rn, sd_c).font = _DF(bold=True, c="C00000")

        dc = step.get("delay_contribution_days")
        if isinstance(dc, int):
            ws.cell(rn, dc_c).alignment = _AR
            if dc > 0:
                ws.cell(rn, dc_c).fill = _P["RETARD"]
                ws.cell(rn, dc_c).font = _DF(bold=True, c="C00000")

        cu = step.get("cumulative_delay_days")
        if isinstance(cu, int):
            ws.cell(rn, cu_c).alignment = _AR
            if cu > 0:
                ws.cell(rn, cu_c).fill = _P["RETARD"]
                ws.cell(rn, cu_c).font = _DF(bold=True, c="C00000")

        if step.get("delay_actor") not in (None, "NONE"):
            ws.cell(rn, da_c).font = _DF(bold=True)

        ws.cell(rn, cs_c).font = _IT()

    _widths(ws, {
        "A": 10, "B": 7,  "C": 9,  "D": 10, "E": 46,
        "F": 10, "G": 13, "H": 12,
        "I": 36, "J": 32,
        "K": 16, "L": 16, "M": 14, "N": 14, "O": 24,
        "P": 13, "Q": 12, "R": 16,
        "S": 14, "T": 14, "U": 14, "V": 14,
        "W": 14, "X": 14, "Y": 32,
        "Z": 18, "AA": 14,
        "AB": 16, "AC": 16, "AD": 55, "AE": 8,
        "AF": 36, "AG": 12, "AH": 55,
    })


# ── Sheet 3: DEBUG_TRACE ──────────────────────────────────────────────────────

D_COLS = [
    "numero", "approver_raw", "actor_type",
    "ged_ref_date", "ged_ref_resp", "ged_ref_cmt",
    "raw_date", "raw_status", "raw_comment",
    "mapped_canonical", "mapped_status_clean", "status_code", "status_scope",
    "ged_extracted_deadline",
    "date_status_type",
    "computed_sas_deadline",
    "computed_phase_deadline",
    "deadline_source",
    "data_date_used",
    "doc_code",
    "submission_instance_id",
    "instance_role",
    "instance_resolution_reason",
]


def _write_debug_trace(ws, debug_rows: list[dict]) -> None:
    ws.freeze_panes = "A2"
    _write_hdr(ws, D_COLS, _P["hdr_dbg"])

    dt_d = D_COLS.index("date_status_type")        + 1
    cn_d = D_COLS.index("mapped_canonical")         + 1
    sd_d = D_COLS.index("computed_sas_deadline")    + 1
    pd_d = D_COLS.index("computed_phase_deadline")  + 1
    dl_d = D_COLS.index("deadline_source")          + 1

    for r in debug_rows:
        ws.append([r.get(c) for c in D_COLS])
        rn  = ws.max_row
        _row_style(ws, rn)
        mds = r.get("date_status_type", "")
        ws.cell(rn, dt_d).fill = _DS_FILL.get(mds, _P["NOT_CALLED"])
        if mds in ("PENDING_LATE", "PENDING_IN_DELAY"):
            ws.cell(rn, dt_d).font = _DF(bold=True, c="FFFFFF")
        if r.get("mapped_canonical") == "UNKNOWN":
            ws.cell(rn, cn_d).fill = _P["UNKNOWN"]
        if r.get("computed_sas_deadline"):
            ws.cell(rn, sd_d).fill = _P["COMPUTED"]
        if r.get("computed_phase_deadline"):
            ws.cell(rn, pd_d).fill = _P["COMPUTED"]
        ws.cell(rn, dl_d).font = _IT()

    _widths(ws, {
        "A": 10, "B": 38, "C": 12,
        "D": 14, "E": 14, "F": 14,
        "G": 38, "H": 18, "I": 55,
        "J": 30, "K": 20, "L": 14, "M": 14,
        "N": 14, "O": 20,
        "P": 14, "Q": 14, "R": 32, "S": 14,
        "T": 22, "U": 36, "V": 20, "W": 36,
    })


# ── Batch write helpers (write_only mode) ─────────────────────────────────────

def _styled_header_row(ws, cols: list, fill: PatternFill) -> list:
    """Build a styled header row for write_only worksheets."""
    row = []
    for col_name in cols:
        c = WriteOnlyCell(ws, value=col_name)
        c.font      = _HF()
        c.fill      = fill
        c.alignment = _AC
        row.append(c)
    return row


def _write_raw_flat_batch(ws, raw_flat_rows: list[dict]) -> None:
    """Write GED_RAW_FLAT in write_only mode: styled header, plain data values."""
    ws.freeze_panes = "A2"
    ws.append(_styled_header_row(ws, R_COLS, _P["hdr_raw"]))
    _widths(ws, {
        "A": 10, "B": 7,  "C": 9,  "D": 10, "E": 46,
        "F": 36, "G": 28, "H": 12,
        "I": 22, "J": 22, "K": 14, "L": 14,
        "M": 36, "N": 14,
        "O": 20, "P": 42, "Q": 14,
        "R": 55, "S": 8,  "T": 8,  "U": 36,
    })
    for r in raw_flat_rows:
        ws.append([r.get(c) for c in R_COLS])


def _write_operations_batch(ws, ops_rows: list[dict]) -> None:
    """Write GED_OPERATIONS in write_only mode: styled header, plain data values."""
    ws.freeze_panes = "A2"
    ws.append(_styled_header_row(ws, O_COLS, _P["hdr_ops"]))
    _widths(ws, {
        "A": 10, "B": 7,  "C": 9,  "D": 10, "E": 46,
        "F": 10, "G": 13, "H": 12,
        "I": 36, "J": 32,
        "K": 16, "L": 16, "M": 14, "N": 14, "O": 24,
        "P": 13, "Q": 12, "R": 16,
        "S": 14, "T": 14, "U": 14, "V": 14,
        "W": 14, "X": 14, "Y": 32,
        "Z": 18, "AA": 14,
        "AB": 16, "AC": 16, "AD": 55, "AE": 8,
        "AF": 36, "AG": 12, "AH": 55,
    })
    for step in ops_rows:
        ws.append([step.get(c) for c in O_COLS])


# ── Public API ────────────────────────────────────────────────────────────────

def write_flat_ged(
    output_dir:    Path,
    all_raw_flat:  list[dict],
    all_ops:       list[dict],
    all_debug:     list[dict],
) -> Path:
    """Write FLAT_GED.xlsx with three sheets. Returns the output path."""
    out_path = output_dir / "FLAT_GED.xlsx"
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "GED_RAW_FLAT"
    _write_raw_flat(ws1, all_raw_flat)

    ws2 = wb.create_sheet("GED_OPERATIONS")
    _write_operations(ws2, all_ops)

    ws3 = wb.create_sheet("DEBUG_TRACE")
    _write_debug_trace(ws3, all_debug)

    wb.save(str(out_path))
    return out_path


def write_flat_ged_batch(
    output_dir:   Path,
    all_raw_flat: list[dict],
    all_ops:      list[dict],
) -> Path:
    """Write FLAT_GED.xlsx for batch mode using write_only streaming.

    Contains GED_RAW_FLAT and GED_OPERATIONS only.
    DEBUG_TRACE is written separately as CSV by write_debug_trace_csv().

    Uses write_only=True for minimal memory footprint and faster streaming:
      - Headers are styled (same colours as single-mode output)
      - Data rows are plain values (no per-cell fills — not practical at 50k+ rows)
    """
    out_path = output_dir / "FLAT_GED.xlsx"
    wb = Workbook(write_only=True)

    ws1 = wb.create_sheet("GED_RAW_FLAT")
    _write_raw_flat_batch(ws1, all_raw_flat)

    ws2 = wb.create_sheet("GED_OPERATIONS")
    _write_operations_batch(ws2, all_ops)

    wb.save(str(out_path))
    return out_path


def write_debug_trace_csv(output_dir: Path, all_debug: list[dict]) -> Path:
    """Write DEBUG_TRACE as a CSV file for batch mode.

    Writing 400k+ rows as Excel (even write_only) takes ~47s due to openpyxl's
    Python overhead. CSV is ~90× faster (0.5s) and preserves all column values.
    Open with Excel or any CSV tool for inspection.
    """
    out_path = output_dir / "DEBUG_TRACE.csv"
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        # utf-8-sig adds BOM so Excel auto-detects UTF-8 (important for French chars)
        writer = csv.writer(f)
        writer.writerow(D_COLS)           # header
        for r in all_debug:
            writer.writerow([r.get(c, "") for c in D_COLS])
    return out_path


def write_run_report(output_dir: Path, stats: dict) -> Path:
    """Write run_report.json. Returns the output path."""
    out_path = output_dir / "run_report.json"

    def _serial(obj):
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()
        return str(obj)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2, default=_serial, ensure_ascii=False)

    return out_path
