"""
reader.py — GED Excel file I/O.

Opens the workbook, reads DATA_DATE from Détails!D15,
and parses the GED sheet header into base_cols + approver_groups.

Two load paths:
  open_workbook()      — read_only=False  (single/debug mode, full cell access)
  open_workbook_fast() — read_only=True   (batch mode, SAX streaming, ~40× faster open)

No business logic here — parsing and structure only.
"""

import datetime
import types

import openpyxl

from config import GED_SHEET_NAME, BASE_FIELD_NAMES


class GEDParseError(Exception):
    """Raised when the GED file is structurally invalid."""


# ── Workbook open ─────────────────────────────────────────────────────────────

def open_workbook(path: str) -> openpyxl.Workbook:
    """Open workbook in full (read-write) mode. Used by single/debug mode."""
    try:
        return openpyxl.load_workbook(path, read_only=False, data_only=True)
    except FileNotFoundError:
        raise GEDParseError(f"[FAIL] Input file not found: {path}")
    except Exception as e:
        raise GEDParseError(f"[FAIL] Cannot open workbook {path!r}: {e}")


def open_workbook_fast(path: str) -> openpyxl.Workbook:
    """Open workbook in read-only streaming mode. Used by batch mode.

    read_only=True uses SAX-based XML parsing — ~40× faster open than
    read_only=False, and does not load the entire XML tree into memory.
    """
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise GEDParseError(f"[FAIL] Input file not found: {path}")
    except Exception as e:
        raise GEDParseError(f"[FAIL] Cannot open workbook {path!r}: {e}")


# ── DATA_DATE ─────────────────────────────────────────────────────────────────

def read_data_date(wb: openpyxl.Workbook) -> datetime.date:
    """Read DATA_DATE from Détails!D15 via direct cell access (non-read_only wb)."""
    if "Détails" not in wb.sheetnames:
        raise GEDParseError("[FAIL] Sheet 'Détails' not found.")
    raw = wb["Détails"]["D15"].value
    if raw is None:
        raise GEDParseError("[FAIL] Détails!D15 is empty — DATA_DATE unavailable.")
    if not isinstance(raw, (datetime.datetime, datetime.date)):
        raise GEDParseError(f"[FAIL] Détails!D15 is not a date: {raw!r}")
    return raw.date() if isinstance(raw, datetime.datetime) else raw


def read_data_date_fast(wb: openpyxl.Workbook) -> datetime.date:
    """Read DATA_DATE from Détails!D15 via row iteration (read_only-safe).

    Direct cell access is not available in read_only mode, so we iterate
    to row 15 and read column D (index 3, 0-based).
    """
    if "Détails" not in wb.sheetnames:
        raise GEDParseError("[FAIL] Sheet 'Détails' not found.")
    ws = wb["Détails"]
    raw = None
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
        if i == 14:          # row 15 is 0-indexed as 14
            raw = row[3]     # column D = index 3
            break
    if raw is None:
        raise GEDParseError("[FAIL] Détails sheet has fewer than 15 rows — DATA_DATE unavailable.")
    if not isinstance(raw, (datetime.datetime, datetime.date)):
        raise GEDParseError(f"[FAIL] Détails!D15 is not a date: {raw!r}")
    return raw.date() if isinstance(raw, datetime.datetime) else raw


# ── GED sheet ─────────────────────────────────────────────────────────────────

def read_ged_sheet(wb: openpyxl.Workbook):
    """Return the GED workflow sheet. Raises GEDParseError if not found."""
    if GED_SHEET_NAME not in wb.sheetnames:
        raise GEDParseError(f"[FAIL] Sheet '{GED_SHEET_NAME}' not found in workbook.")
    return wb[GED_SHEET_NAME]


# ── Header parsing (shared logic) ─────────────────────────────────────────────

def _build_header_maps(row1: list) -> tuple[dict, list]:
    """Build base_cols and approver_groups from the first header row."""
    base_cols = {i: h for i, h in enumerate(row1) if h in BASE_FIELD_NAMES}

    approver_groups = []
    for i, h in enumerate(row1):
        if h and str(h).strip() and h not in BASE_FIELD_NAMES:
            approver_groups.append({
                "name":      str(h).strip(),
                "ged_order": len(approver_groups),
                "col_date":  i,
                "col_resp":  i + 1,
                "col_cmt":   i + 2,
                "col_pj":    i + 3,
            })

    if not base_cols:
        raise GEDParseError("[FAIL] No CORE_COLUMNS found in GED sheet row 1 — wrong sheet?")

    return base_cols, approver_groups


def parse_ged_header(ws) -> tuple[dict, list, list]:
    """Parse GED header — materialises all rows. Used by single/debug mode.

    Returns (base_cols, approver_groups, all_rows).
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise GEDParseError("[FAIL] GED sheet is empty.")

    base_cols, approver_groups = _build_header_maps(list(all_rows[0]))
    return base_cols, approver_groups, all_rows


def parse_ged_header_batch(ws) -> tuple[dict, list, types.GeneratorType]:
    """Parse GED header in streaming mode — does NOT materialise the full sheet.

    Consumes only the first two rows of the sheet iterator, then returns
    the still-open iterator positioned at row 3 (data start).

    Returns (base_cols, approver_groups, data_rows_iterator).
    The caller must exhaust data_rows_iterator in a single forward pass.
    """
    rows_iter = ws.iter_rows(values_only=True)
    try:
        row1_tuple = next(rows_iter)   # header row 1
        next(rows_iter)                # header row 2 — skip (sub-field labels)
    except StopIteration:
        raise GEDParseError("[FAIL] GED sheet is empty or has fewer than 2 rows.")

    base_cols, approver_groups = _build_header_maps(list(row1_tuple))
    return base_cols, approver_groups, rows_iter
