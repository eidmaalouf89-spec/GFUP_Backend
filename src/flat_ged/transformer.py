"""
transformer.py — Document processing pipeline.

Given a resolved candidate, builds:
  - GED_RAW_FLAT rows   (factual GED layer)
  - GED_OPERATIONS rows (operational reconstruction with computed deadlines)
  - DEBUG_TRACE rows    (full audit: raw GED + computed values side by side)

All business rules are preserved exactly from prototype_v4.py.
"""

import sys
import datetime

import openpyxl.utils

from config import (
    RAW_TO_CANONICAL,
    MOEX_CANONICAL,
    SAS_CANONICAL,
    DISPLAY,
    ALL_COMPLETED,
    SAS_WINDOW_DAYS,
    GLOBAL_WINDOW_DAYS,
)
from utils import (
    parse_status,
    interpret_date,
    actor_type_for,
    calc_status_family,
    calc_retard_avance,
)


class GEDValidationError(Exception):
    """Raised when a hard invariant or business rule is violated."""


# ── Step 3: Build GED_RAW_FLAT + DEBUG_TRACE skeleton ────────────────────────

def build_raw_flat(
    candidate:       dict,
    base_cols:       dict,
    approver_groups: list,
) -> tuple[list[dict], list[dict]]:
    """Build GED_RAW_FLAT rows and the DEBUG_TRACE skeleton for one document.

    raw_flat rows:   one entry per called, non-exception approver
    debug_rows:      one entry per approver column (all, including skipped ones)

    The computed deadline fields in debug_rows are filled later by
    fill_debug_deadlines().
    """
    target_row  = candidate["row_data"]
    ged_row_idx = candidate["ged_row_index"]
    doc_code    = candidate["doc_code"]
    sub_id      = candidate["submission_instance_id"]
    inst_role   = candidate["instance_role"]
    inst_reason = candidate["instance_resolution_reason"]

    doc = {field: target_row[idx] for idx, field in base_cols.items()}
    numero        = doc["NUMERO"]
    indice        = doc["INDICE"]
    lot           = doc["LOT"]
    emetteur      = doc["EMETTEUR"]
    titre         = doc["Libellé du document"]
    cree_le_raw   = doc["Créé le"]
    submittal_date = (cree_le_raw.date() if isinstance(cree_le_raw, datetime.datetime)
                      else cree_le_raw)

    raw_flat   = []
    debug_rows = []

    for ag in approver_groups:
        aname    = ag["name"]
        date_raw = target_row[ag["col_date"]] if ag["col_date"] < len(target_row) else None
        resp_raw = target_row[ag["col_resp"]] if ag["col_resp"] < len(target_row) else None
        cmt_raw  = target_row[ag["col_cmt"]]  if ag["col_cmt"]  < len(target_row) else None
        pj_raw   = target_row[ag["col_pj"]]   if ag["col_pj"]   < len(target_row) else None

        canonical                          = RAW_TO_CANONICAL.get(aname, "UNKNOWN")
        ds, response_date, dl_raw, ged_dl  = interpret_date(date_raw)
        status_code, status_scope          = parse_status(resp_raw)
        pj_flag                            = 1 if pj_raw not in (None, "", 0) else 0
        atype                              = actor_type_for(canonical)

        # Cell references for traceability
        col_d = openpyxl.utils.get_column_letter(ag["col_date"] + 1)
        col_r = openpyxl.utils.get_column_letter(ag["col_resp"] + 1)
        col_c = openpyxl.utils.get_column_letter(ag["col_cmt"]  + 1)

        debug_rows.append({
            "numero":              numero,
            "approver_raw":        aname,
            "actor_type":          atype if canonical not in ("Exception List", "UNKNOWN") else "—",
            "ged_ref_date":        f"{col_d}{ged_row_idx}",
            "ged_ref_resp":        f"{col_r}{ged_row_idx}",
            "ged_ref_cmt":         f"{col_c}{ged_row_idx}",
            "raw_date":            str(date_raw)   if date_raw is not None else "",
            "raw_status":          str(resp_raw)   if resp_raw is not None else "",
            "raw_comment":         str(cmt_raw)    if cmt_raw  is not None else "",
            "mapped_canonical":    canonical,
            "mapped_status_clean": status_code or "",
            "status_code":         status_code or "",
            "status_scope":        status_scope or "",
            "ged_extracted_deadline": str(ged_dl) if ged_dl else "",
            "date_status_type":    ds,
            # Filled later by fill_debug_deadlines()
            "computed_sas_deadline":   "",
            "computed_phase_deadline": "",
            "deadline_source":         "",
            "data_date_used":          "",
            # Instance resolution fields
            "doc_code":                   doc_code,
            "submission_instance_id":     sub_id,
            "instance_role":              inst_role,
            "instance_resolution_reason": inst_reason,
        })

        # GED_RAW_FLAT: skip Exception List and NOT_CALLED rows
        if canonical == "Exception List" or ds == "NOT_CALLED":
            continue

        raw_flat.append({
            "numero": numero, "indice": indice, "lot": lot,
            "emetteur": emetteur, "titre": titre,
            "approver_raw":          aname,
            "approver_canonical":    canonical,
            "actor_type":            atype,
            "response_status_raw":   resp_raw,
            "response_status_clean": status_code,
            "response_status_code":  status_code,
            "response_status_scope": status_scope,
            "response_date_raw":     date_raw,
            "response_date":         str(response_date) if response_date else "",
            "date_status_type":      ds,
            "deadline_raw":          dl_raw or "",
            "deadline":              str(ged_dl) if ged_dl else "",
            "commentaire":           cmt_raw,
            "pj_flag":               pj_flag,
            "is_sas":                (aname == SAS_CANONICAL),
            "raw_trace_key":         f"{numero}|{aname}",
            # Internal fields (used by build_operations and sorting)
            "_canonical":      canonical,
            "_response_date":  response_date,
            "_ged_dl":         ged_dl,
            "_ged_order":      ag["ged_order"],
            "_date_status":    ds,
            "_status_clean":   status_code,
            "_status_code":    status_code,
            "_status_scope":   status_scope,
            # Convenience
            "_submittal_date": submittal_date,
            "_numero":         numero,
            "_indice":         indice,
            "_lot":            lot,
            "_emetteur":       emetteur,
            "_titre":          titre,
        })

    return raw_flat, debug_rows, submittal_date, numero, indice, lot, emetteur, titre


# ── Step 4: SAS state detection ───────────────────────────────────────────────

def detect_sas_state(raw_flat: list[dict]) -> tuple[str, dict | None]:
    """Determine SAS state from raw_flat rows.

    Returns (sas_state, sas_entry):
      sas_state: "ABSENT" | "ANSWERED" | "PENDING"
      sas_entry: the SAS row dict, or None if absent
    """
    sas_entry = next((r for r in raw_flat if r["_canonical"] == SAS_CANONICAL), None)
    if sas_entry is None:
        return "ABSENT", None
    if sas_entry["_response_date"] is not None:
        return "ANSWERED", sas_entry
    if sas_entry["_date_status"] in ("PENDING_IN_DELAY", "PENDING_LATE"):
        return "PENDING", sas_entry
    raise GEDValidationError(
        f"[FAIL] SAS row exists but has no response date and unrecognised "
        f"date_status_type: {sas_entry['_date_status']!r}"
    )


def add_synthetic_sas(
    raw_flat:      list[dict],
    debug_rows:    list[dict],
    submittal_date: datetime.date,
    numero,
    indice:        str,
    lot,
    emetteur,
    titre,
    data_date:     datetime.date,
    doc_code:      str,
    sub_id:        str,
    inst_role:     str,
    inst_reason:   str,
) -> dict:
    """Insert a synthetic SAS row when SAS was never called in the GED.

    The synthetic SAS models the implicit approval: SAS was not explicitly
    solicited, so it is treated as auto-cleared on the submittal date.

    Returns the synthetic_sas dict (already appended to raw_flat and debug_rows).
    """
    synthetic_sas = {
        "numero": numero, "indice": indice, "lot": lot,
        "emetteur": emetteur, "titre": titre,
        "approver_raw":          SAS_CANONICAL,
        "approver_canonical":    SAS_CANONICAL,
        "actor_type":            "SAS",
        "response_status_raw":   "SYNTHETIC_VSO-SAS",
        "response_status_clean": "VSO",
        "response_status_code":  "VSO",
        "response_status_scope": "SAS",
        "response_date_raw":     submittal_date,
        "response_date":         str(submittal_date),
        "date_status_type":      "ANSWERED",
        "deadline_raw":          "",
        "deadline":              "",
        "commentaire":           "Synthetic SAS (auto-cleared — not called in GED)",
        "pj_flag":               0,
        "is_sas":                True,
        "raw_trace_key":         f"{numero}|SYNTHETIC_SAS",
        # Internal fields
        "_canonical":      SAS_CANONICAL,
        "_response_date":  submittal_date,
        "_ged_dl":         None,
        "_ged_order":      -1,
        "_date_status":    "ANSWERED",
        "_status_clean":   "VSO",
        "_status_code":    "VSO",
        "_status_scope":   "SAS",
        "_submittal_date": submittal_date,
        "_numero":         numero,
        "_indice":         indice,
        "_lot":            lot,
        "_emetteur":       emetteur,
        "_titre":          titre,
    }
    raw_flat.append(synthetic_sas)

    debug_rows.append({
        "numero":              numero,
        "approver_raw":        f"{SAS_CANONICAL} [SYNTHETIC]",
        "actor_type":          "SAS",
        "ged_ref_date":        "SYNTHETIC",
        "ged_ref_resp":        "SYNTHETIC",
        "ged_ref_cmt":         "SYNTHETIC",
        "raw_date":            "",
        "raw_status":          "",
        "raw_comment":         "SAS not called — synthetic VSO/SAS inserted",
        "mapped_canonical":    SAS_CANONICAL,
        "mapped_status_clean": "VSO",
        "status_code":         "VSO",
        "status_scope":        "SAS",
        "ged_extracted_deadline": "",
        "date_status_type":    "ANSWERED",
        # Filled by fill_debug_deadlines()
        "computed_sas_deadline":   "",
        "computed_phase_deadline": "",
        "deadline_source":         "",
        "data_date_used":          str(data_date),
        # Instance resolution fields
        "doc_code":                   doc_code,
        "submission_instance_id":     sub_id,
        "instance_role":              inst_role,
        "instance_resolution_reason": inst_reason,
    })

    return synthetic_sas


# ── Step 4 continued: compute phase deadlines ─────────────────────────────────

def compute_phase_deadlines(
    sas_entry:      dict | None,
    submittal_date: datetime.date,
) -> tuple[datetime.date, datetime.date | None, str]:
    """Compute SAS and consultant/MOEX phase deadlines.

    Returns (sas_phase_deadline, cm_deadline, cm_dl_source):
      sas_phase_deadline: submittal + SAS_WINDOW_DAYS  (always computable)
      cm_deadline:        None if SAS has not responded yet
      cm_dl_source:       string describing which rule produced cm_deadline
    """
    import datetime as dt
    sas_phase_deadline = submittal_date + dt.timedelta(days=SAS_WINDOW_DAYS)
    global_deadline    = submittal_date + dt.timedelta(days=GLOBAL_WINDOW_DAYS)

    sas_response_date = sas_entry["_response_date"] if sas_entry else None

    if sas_response_date is None:
        cm_deadline  = None
        cm_dl_source = "WAITING_FOR_SAS"
    else:
        sas_consumed = (sas_response_date - submittal_date).days
        if sas_consumed <= SAS_WINDOW_DAYS:
            cm_deadline  = global_deadline
            cm_dl_source = "GLOBAL_30D_AFTER_ON_TIME_SAS"
        else:
            cm_deadline  = sas_response_date + dt.timedelta(days=SAS_WINDOW_DAYS)
            cm_dl_source = "COMPUTED_15D_AFTER_LATE_SAS"

    return sas_phase_deadline, cm_deadline, cm_dl_source


def fill_debug_deadlines(
    debug_rows:         list[dict],
    sas_phase_deadline: datetime.date,
    cm_deadline:        datetime.date | None,
    cm_dl_source:       str,
    data_date:          datetime.date,
) -> None:
    """Back-fill computed deadline fields into the debug_rows (in-place)."""
    for dr in debug_rows:
        dr["computed_sas_deadline"] = str(sas_phase_deadline)
        dr["data_date_used"]        = str(data_date)
        canonical = dr["mapped_canonical"]
        if canonical == SAS_CANONICAL:
            dr["computed_phase_deadline"] = str(sas_phase_deadline)
            dr["deadline_source"]         = "COMPUTED_SAS_15D"
        elif canonical in ("Exception List", "UNKNOWN"):
            dr["computed_phase_deadline"] = ""
            dr["deadline_source"]         = "N/A"
        else:
            dr["computed_phase_deadline"] = str(cm_deadline) if cm_deadline else ""
            dr["deadline_source"]         = cm_dl_source


# ── Step 5: Build GED_OPERATIONS ─────────────────────────────────────────────

def _make_step(
    step_type:      str,
    actor_type:     str,
    actor_raw:      str,
    actor_clean:    str,
    status_raw,
    sc:             str | None,
    status_scope:   str | None,
    ds:             str,
    response_date:  datetime.date | None,
    phase_deadline: datetime.date | None,
    dl_source:      str,
    chrono_source:  str,
    op_rule:        str,
    observation,
    pj_flag,
    source_trace:   str,
    source_rows:    int,
    submittal_date: datetime.date,
    sas_response_date: datetime.date | None,
    global_deadline:   datetime.date,
    data_date:         datetime.date,
    numero,
    indice,
    lot,
    emetteur,
    titre,
) -> dict:
    sf        = calc_status_family(ds, sc) if step_type != "OPEN_DOC" else "OPENED"
    completed = (sc in ALL_COMPLETED) if sc else (ds == "ANSWERED")
    blocking  = ds in ("PENDING_IN_DELAY", "PENDING_LATE")
    new_cycle = (sc or "") in ("REF", "DEF")
    rv, rvs   = calc_retard_avance(phase_deadline, response_date, ds, data_date)

    return {
        "numero":   numero,
        "indice":   indice,
        "lot":      lot,
        "emetteur": emetteur,
        "titre":    titre,
        "step_type":           step_type,
        "actor_type":          actor_type,
        "actor_raw":           actor_raw,
        "actor_clean":         actor_clean,
        "status_raw":          status_raw or "",
        "status_clean":        sc or "",
        "status_code":         sc or "",
        "status_scope":        status_scope or "",
        "status_family":       sf,
        "is_completed":        completed,
        "is_blocking":         blocking,
        "requires_new_cycle":  new_cycle,
        "submittal_date":      str(submittal_date),
        "sas_response_date":   str(sas_response_date) if sas_response_date else "",
        "response_date":       str(response_date) if response_date else "",
        "data_date":           str(data_date),
        "global_deadline":     str(global_deadline),
        "phase_deadline":      str(phase_deadline) if phase_deadline else "",
        "deadline_source":     dl_source,
        "retard_avance_days":  rv if rv is not None else "",
        "retard_avance_status": rvs or "",
        "chrono_source":       chrono_source,
        "observation":         observation or "",
        "pj_flag":             pj_flag,
        "source_trace":        source_trace,
        "source_rows":         source_rows,
        "operation_rule_used": op_rule,
    }


def _sort_key(r: dict):
    rd = r["_response_date"]
    return (0, rd, r["_ged_order"]) if rd else (1, datetime.date.max, r["_ged_order"])


def build_operations(
    raw_flat:           list[dict],
    sas_entry:          dict | None,
    sas_state:          str,
    sas_not_called:     bool,
    sas_phase_deadline: datetime.date,
    cm_deadline:        datetime.date | None,
    cm_dl_source:       str,
    submittal_date:     datetime.date,
    sas_response_date:  datetime.date | None,
    global_deadline:    datetime.date,
    data_date:          datetime.date,
    numero,
    indice,
    lot,
    emetteur,
    titre,
) -> list[dict]:
    """Build the GED_OPERATIONS step list for one document.

    Step order:
      1. OPEN_DOC   (synthetic, always first)
      2. SAS        (real or synthetic)
      3. CONSULTANTs (sorted: response_date ASC → ged_order, skipped if SAS pending)
      4. MOEX       (forced last, skipped if SAS pending)
    """
    _step_kwargs = dict(
        submittal_date=submittal_date,
        sas_response_date=sas_response_date,
        global_deadline=global_deadline,
        data_date=data_date,
        numero=numero, indice=indice, lot=lot, emetteur=emetteur, titre=titre,
    )

    ops = []

    # ── OPEN_DOC (synthetic) ──────────────────────────────────────────────────
    open_step = _make_step(
        "OPEN_DOC", "EMETTEUR", emetteur, emetteur,
        "", "OPENED", "STANDARD", "ANSWERED",
        response_date=None, phase_deadline=None,
        dl_source="NONE",
        chrono_source="SYNTHETIC", op_rule="OPEN_DOC_INIT",
        observation="", pj_flag="",
        source_trace=f"{numero}|OPEN_DOC", source_rows=0,
        **_step_kwargs,
    )
    open_step.update({
        "is_completed": True, "is_blocking": False,
        "requires_new_cycle": False,
        "retard_avance_days": "", "retard_avance_status": "",
    })
    ops.append(open_step)

    # ── SAS ───────────────────────────────────────────────────────────────────
    if sas_entry:
        if sas_phase_deadline is None:
            raise GEDValidationError("[FAIL] SAS exists but sas_phase_deadline could not be computed.")
        r = sas_entry
        ops.append(_make_step(
            "SAS", "SAS", r["approver_raw"],
            DISPLAY.get(SAS_CANONICAL, "SAS (GEMO)"),
            r["response_status_raw"], r["_status_code"], r["_status_scope"],
            r["_date_status"], r["_response_date"],
            phase_deadline=sas_phase_deadline,
            dl_source="COMPUTED_SAS_15D",
            chrono_source="response_date" if r["_response_date"] else "ged_order",
            op_rule="SYNTHETIC_SAS_IF_NOT_CALLED" if sas_not_called else "SAS_FROM_0-SAS",
            observation=r["commentaire"], pj_flag=r["pj_flag"],
            source_trace=r["raw_trace_key"], source_rows=1,
            **_step_kwargs,
        ))

    # ── CONSULTANTs (sorted: response_date ASC → ged_order) ──────────────────
    if sas_state != "PENDING":
        c_rows = [r for r in raw_flat
                  if r["_canonical"] not in (SAS_CANONICAL, MOEX_CANONICAL,
                                             "Exception List", "UNKNOWN")]
        c_rows.sort(key=_sort_key)
        for r in c_rows:
            if cm_deadline is None:
                raise GEDValidationError(
                    "[FAIL] Consultant phase deadline not available — SAS has not responded."
                )
            ops.append(_make_step(
                "CONSULTANT", "CONSULTANT", r["approver_raw"],
                DISPLAY.get(r["_canonical"], r["_canonical"]),
                r["response_status_raw"], r["_status_code"], r["_status_scope"],
                r["_date_status"], r["_response_date"],
                phase_deadline=cm_deadline,
                dl_source=cm_dl_source,
                chrono_source="response_date" if r["_response_date"] else "ged_order",
                op_rule="CONSULTANT_FROM_RAW",
                observation=r["commentaire"], pj_flag=r["pj_flag"],
                source_trace=r["raw_trace_key"], source_rows=1,
                **_step_kwargs,
            ))

    # ── MOEX (forced last) ────────────────────────────────────────────────────
    moex_entry = next((r for r in raw_flat if r["_canonical"] == MOEX_CANONICAL), None)
    if moex_entry and sas_state != "PENDING":
        if cm_deadline is None:
            raise GEDValidationError(
                "[FAIL] MOEX phase deadline not available — SAS has not responded."
            )
        r = moex_entry
        ops.append(_make_step(
            "MOEX", "MOEX", r["approver_raw"],
            DISPLAY.get(MOEX_CANONICAL, MOEX_CANONICAL),
            r["response_status_raw"], r["_status_code"], r["_status_scope"],
            r["_date_status"], r["_response_date"],
            phase_deadline=cm_deadline,
            dl_source=cm_dl_source,
            chrono_source="FORCED_LAST",
            op_rule="MOEX_FINAL_STEP",
            observation=r["commentaire"], pj_flag=r["pj_flag"],
            source_trace=r["raw_trace_key"], source_rows=1,
            **_step_kwargs,
        ))

    # Assign step_order
    for i, s in enumerate(ops, 1):
        s["step_order"] = i

    return ops


# ── Delay contribution (cumulative, no double-counting) ───────────────────────

def compute_delay_contribution(ops: list[dict], data_date: datetime.date) -> None:
    """Compute step_delay_days, delay_contribution_days, cumulative_delay_days (in-place).

    For each step in order:
      effective_date = response_date (closed) or data_date (open)
      step_delay     = max(0, effective_date - phase_deadline)
      contribution   = max(0, step_delay - cumulative_so_far)
      cumulative    += contribution
    """
    cumulative = 0
    for s in ops:
        pd = s["phase_deadline"]
        if not pd:
            # OPEN_DOC has no phase_deadline — no delay contribution
            s["step_delay_days"]         = 0
            s["delay_contribution_days"] = 0
            s["cumulative_delay_days"]   = cumulative
            s["delay_actor"]             = "NONE"
            continue

        pd_date = datetime.date.fromisoformat(pd) if isinstance(pd, str) else pd
        rd      = s["response_date"]

        if rd:
            rd_date  = datetime.date.fromisoformat(rd) if isinstance(rd, str) else rd
            eff_date = rd_date
        else:
            eff_date = data_date

        step_lateness = (pd_date - eff_date).days
        step_delay    = max(0, -step_lateness)
        contribution  = max(0, step_delay - cumulative)
        cumulative   += contribution

        s["step_delay_days"]         = step_delay
        s["delay_contribution_days"] = contribution
        s["cumulative_delay_days"]   = cumulative
        s["delay_actor"]             = s["actor_clean"] if contribution > 0 else "NONE"


# ── Cycle closure determination ───────────────────────────────────────────────

def determine_cycle_closure(
    ops:       list[dict],
    data_date: datetime.date,
) -> tuple[str, str, datetime.date]:
    """Determine cycle_state, closure_mode, and effective_cycle_end_date.

    Rule 1 — MOEX_VISA:
      MOEX step exists, is completed, and has a response date.
      Pending consultant rows remain factual but are no longer blocking.

    Rule 2 — ALL_RESPONDED_NO_MOEX:
      No MOEX, but all SAS + CONSULTANT rows have responded.

    Rule 3 — WAITING_RESPONSES:
      Neither of the above — cycle is still open.
    """
    moex_step = next((s for s in ops if s["step_type"] == "MOEX"), None)

    if moex_step is not None and moex_step["is_completed"] and moex_step["response_date"]:
        # Rule 1: MOEX visa closes the cycle
        for s in ops:
            if s["is_blocking"]:
                s["is_blocking"] = False
        return (
            "CLOSED",
            "MOEX_VISA",
            datetime.date.fromisoformat(moex_step["response_date"]),
        )

    required_rows = [s for s in ops if s["step_type"] in ("SAS", "CONSULTANT")]
    if required_rows and all(s["is_completed"] for s in required_rows):
        # Rule 2: All non-MOEX rows responded
        return (
            "CLOSED",
            "ALL_RESPONDED_NO_MOEX",
            max(datetime.date.fromisoformat(s["response_date"]) for s in required_rows),
        )

    return "OPEN", "WAITING_RESPONSES", data_date


# ── Full document pipeline ────────────────────────────────────────────────────

def process_document(
    candidate:       dict,
    base_cols:       dict,
    approver_groups: list,
    data_date:       datetime.date,
    quiet:           bool = False,
) -> tuple[list[dict], list[dict], list[dict], dict]:
    """Run the full pipeline for one resolved candidate.

    Returns (raw_flat, ops, debug_rows, doc_stats).
    Raises GEDValidationError on hard failure.
    doc_stats contains per-document summary counts for the run report.
    """
    # Build raw_flat + debug skeleton
    raw_flat, debug_rows, submittal_date, numero, indice, lot, emetteur, titre = \
        build_raw_flat(candidate, base_cols, approver_groups)

    # Candidate instance metadata (for pass-through to add_synthetic_sas)
    doc_code    = candidate["doc_code"]
    sub_id      = candidate["submission_instance_id"]
    inst_role   = candidate["instance_role"]
    inst_reason = candidate["instance_resolution_reason"]

    global_deadline = submittal_date + datetime.timedelta(days=GLOBAL_WINDOW_DAYS)

    if not quiet:
        print(f"[OK] Document:   NUMERO={numero} | INDICE={indice} | "
              f"LOT={lot} | EMETTEUR={emetteur}")
        print(f"     GED row:    {candidate['ged_row_index']}")
        print(f"     submittal:  {submittal_date}")
        print(f"     global_dl:  {global_deadline}  (submittal + {GLOBAL_WINDOW_DAYS}d)")

    # SAS state detection (must run before any synthetic insertion)
    sas_state, sas_entry = detect_sas_state(raw_flat)
    sas_not_called = (sas_entry is None)

    # Synthetic SAS (only when SAS was never called in GED)
    if sas_not_called:
        if not quiet:
            print(f"[INFO] SAS not called in GED — inserting synthetic SAS "
                  f"(status=VSO/SAS, response_date=submittal_date)")
        sas_entry = add_synthetic_sas(
            raw_flat, debug_rows,
            submittal_date, numero, indice, lot, emetteur, titre,
            data_date, doc_code, sub_id, inst_role, inst_reason,
        )
        sas_state = "ANSWERED"

    sas_response_date = sas_entry["_response_date"] if sas_entry else None

    # Phase deadlines
    sas_phase_deadline, cm_deadline, cm_dl_source = compute_phase_deadlines(
        sas_entry, submittal_date
    )
    fill_debug_deadlines(debug_rows, sas_phase_deadline, cm_deadline, cm_dl_source, data_date)

    if not quiet:
        print(f"[OK] sas_state          = {sas_state}")
        print(f"[OK] sas_response_date: {sas_response_date}")
        if sas_response_date:
            print(f"     sas_consumed_days:  {(sas_response_date - submittal_date).days}")
        print(f"[OK] cm_deadline:       {cm_deadline}  ({cm_dl_source})")
        if sas_state == "PENDING":
            print("[INFO] sas_state=PENDING — consultant/MOEX phase skipped")

    # Build operations
    ops = build_operations(
        raw_flat=raw_flat,
        sas_entry=sas_entry,
        sas_state=sas_state,
        sas_not_called=sas_not_called,
        sas_phase_deadline=sas_phase_deadline,
        cm_deadline=cm_deadline,
        cm_dl_source=cm_dl_source,
        submittal_date=submittal_date,
        sas_response_date=sas_response_date,
        global_deadline=global_deadline,
        data_date=data_date,
        numero=numero, indice=indice, lot=lot, emetteur=emetteur, titre=titre,
    )

    # Delay contribution
    compute_delay_contribution(ops, data_date)

    # Cycle closure
    cycle_state, closure_mode, effective_cycle_end_date = determine_cycle_closure(ops, data_date)

    if not quiet:
        print(f"\n── Cycle Closure ──────────────────────────────────────")
        print(f"  cycle_state:              {cycle_state}")
        print(f"  closure_mode:             {closure_mode}")
        print(f"  effective_cycle_end_date: {effective_cycle_end_date}")
        print(f"──────────────────────────────────────────────────────\n")

    doc_stats = {
        "sas_state":           sas_state,
        "sas_not_called":      sas_not_called,
        "closure_mode":        closure_mode,
        "num_candidates":      1,  # will be updated by caller from resolver
        "raw_flat_count":      len(raw_flat),
        "ops_count":           len(ops),
        "debug_rows_count":    len(debug_rows),
        "final_cumulative_delay": ops[-1]["cumulative_delay_days"] if ops else 0,
    }

    return raw_flat, ops, debug_rows, doc_stats, cycle_state, closure_mode, effective_cycle_end_date, sas_state, cm_dl_source, global_deadline
