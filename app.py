"""
JANSA VISASIST — Desktop application launcher
PyWebView + React, single-user, single-project
"""
import sys
import os
import json
import math
import threading
import subprocess
import sqlite3
import time
import webbrowser
from pathlib import Path

import traceback

import webview

# ── Path resolution ──────────────────────────────────────────
def _resolve_base_dir():
    """Resolve base directory. Handles both dev mode and PyInstaller frozen."""
    if getattr(sys, 'frozen', False):
        # PyInstaller: exe is in dist/, data is alongside it
        return Path(sys.executable).parent
    return Path(__file__).parent

BASE_DIR = _resolve_base_dir()
DATA_DIR = BASE_DIR / "data"
RUN_MEMORY_DB = str(DATA_DIR / "run_memory.db")
REPORT_MEMORY_DB = str(DATA_DIR / "report_memory.db")
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
RUNS_DIR = BASE_DIR / "runs"

# ── Add src/ to path so existing modules can find each other ──
sys.path.insert(0, str(BASE_DIR / "src"))

# ── UI resolution ────────────────────────────────────────────
def _resolve_ui():
    """Return the single production UI entrypoint for PyWebView."""
    jansa = BASE_DIR / "ui" / "jansa-connected.html"
    if jansa.exists():
        return str(jansa)
    raise FileNotFoundError(
        "JANSA production UI not found: ui/jansa-connected.html"
    )


def _ui_target_for_browser(ui_url: str) -> str:
    """Return a browser-openable target for the JANSA HTML entrypoint."""
    if ui_url.startswith("http://") or ui_url.startswith("https://"):
        return ui_url
    return Path(ui_url).resolve().as_uri()

# ── Database helpers ─────────────────────────────────────────
_SQLITE_READ_RETRY_DELAYS = (0.0, 0.2, 0.5)


def _is_sqlite_locked_error(exc):
    msg = str(exc).lower()
    return "database is locked" in msg or "database table is locked" in msg


def _immutable_sqlite_uri(db_path):
    return "file:" + Path(db_path).resolve().as_posix() + "?mode=ro&immutable=1"


def _query_db(db_path, sql, params=()):
    """Execute a read query and return results as list of dicts."""
    if not Path(db_path).exists():
        return []
    last_exc = None
    for attempt, delay in enumerate(_SQLITE_READ_RETRY_DELAYS):
        try:
            conn = sqlite3.connect(db_path, timeout=5)
            conn.execute("PRAGMA busy_timeout=5000")
            conn.row_factory = sqlite3.Row
            try:
                rows = conn.execute(sql, params).fetchall()
                return [dict(r) for r in rows]
            finally:
                conn.close()
        except sqlite3.OperationalError as exc:
            last_exc = exc
            if _is_sqlite_locked_error(exc) and attempt < len(_SQLITE_READ_RETRY_DELAYS) - 1:
                time.sleep(delay)
                continue
            break

    conn = sqlite3.connect(_immutable_sqlite_uri(db_path), uri=True, timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(sql, params).fetchall()
        if last_exc is not None:
            print(f"[app][SQLITE_IMMUTABLE_FALLBACK] {db_path}: {last_exc}")
        return [dict(r) for r in rows]
    finally:
        conn.close()

# ── JSON sanitizer for PyWebView bridge ──────────────────────
def _sanitize_for_json(obj):
    """
    Recursively replace NaN, Infinity, -Infinity with None
    so the result is valid JSON for PyWebView's JS bridge.
    Also converts pandas Timestamps and other non-serializable types.
    """
    if obj is None:
        return None
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_for_json(item) for item in obj]
    if isinstance(obj, (int, bool, str)):
        return obj
    # Handle pandas Timestamp, numpy types, etc.
    try:
        import pandas as pd
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat() if not pd.isna(obj) else None
        if pd.isna(obj):
            return None
    except Exception:
        pass
    try:
        # numpy int64, float64, etc.
        import numpy as np
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return None if np.isnan(obj) else float(obj)
        if isinstance(obj, np.bool_):
            return bool(obj)
    except Exception:
        pass
    # Fallback: convert to string
    return str(obj)

# ── Chain data freshness helpers ──────────────────────────────

CHAIN_ONION_REFRESH_TIMEOUT_S = 300


def _chain_onion_outputs_stale(base_dir: Path) -> tuple:
    """Decide whether chain_onion outputs need to be regenerated.

    Returns (is_stale, reason). True if any required file is missing,
    OR if FLAT_GED.xlsx is newer than CHAIN_REGISTER.csv (chain_onion's
    inputs changed since last run).
    """
    co = base_dir / "output" / "chain_onion"
    required = ["CHAIN_REGISTER.csv", "CHAIN_EVENTS.csv", "CHAIN_VERSIONS.csv"]
    for name in required:
        p = co / name
        if not p.exists():
            return True, f"missing: {name}"
    flat_ged = base_dir / "output" / "intermediate" / "FLAT_GED.xlsx"
    register = co / "CHAIN_REGISTER.csv"
    if flat_ged.exists() and flat_ged.stat().st_mtime > register.stat().st_mtime:
        return True, "FLAT_GED.xlsx newer than CHAIN_REGISTER.csv"
    return False, "fresh"


def _run_chain_onion_subprocess(base_dir: Path) -> tuple:
    """Invoke run_chain_onion.py as a subprocess. Returns (success, message).

    Captures stdout/stderr; logs them via print. Times out at
    CHAIN_ONION_REFRESH_TIMEOUT_S. NEVER raises — failures return (False, msg).
    """
    script = base_dir / "run_chain_onion.py"
    if not script.exists():
        return False, f"run_chain_onion.py not found at {script}"
    try:
        result = subprocess.run(
            [sys.executable, str(script)],
            cwd=str(base_dir),
            capture_output=True,
            text=True,
            timeout=CHAIN_ONION_REFRESH_TIMEOUT_S,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            co = base_dir / "output" / "chain_onion"
            required = ["CHAIN_REGISTER.csv", "CHAIN_EVENTS.csv", "CHAIN_VERSIONS.csv"]
            if all((co / name).exists() for name in required):
                print(f"[app][CHAIN_ONION] exit {result.returncode} but artifacts present — partial success")
                return True, f"partial success (exit {result.returncode})"
            print(f"[app][CHAIN_ONION] subprocess exited {result.returncode}")
            stderr_safe = (result.stderr or "")[-500:].encode("utf-8", errors="replace").decode("utf-8")
            print(f"[app][CHAIN_ONION] stderr tail: {stderr_safe}")
            return False, f"exit code {result.returncode}"
        print(f"[app][CHAIN_ONION] subprocess OK")
        return True, "ok"
    except subprocess.TimeoutExpired:
        return False, f"timeout after {CHAIN_ONION_REFRESH_TIMEOUT_S}s"
    except Exception as exc:
        return False, f"exception: {exc}"


def _refresh_chain_timeline(ctx, base_dir: Path) -> tuple:
    """Recompute CHAIN_TIMELINE_ATTRIBUTION.{json,csv} from current chain_onion
    outputs + ctx. Returns (success, message). NEVER raises.
    """
    try:
        from reporting.chain_timeline_attribution import (
            _load_chain_data,
            compute_all_chain_timelines,
            write_chain_timeline_artifact,
        )
        co_dir = base_dir / "output" / "chain_onion"
        intermediate = base_dir / "output" / "intermediate"
        events_df, register_df, versions_df = _load_chain_data(co_dir)
        timelines = compute_all_chain_timelines(ctx, events_df, register_df, versions_df)
        json_path, csv_path = write_chain_timeline_artifact(timelines, intermediate)
        print(f"[app][CHAIN_TIMELINE] wrote {json_path.name} ({len(timelines)} chains)")
        return True, f"{len(timelines)} chains"
    except Exception as exc:
        print(f"[app][CHAIN_TIMELINE] refresh failed: {exc}")
        traceback.print_exc()
        return False, str(exc)


def _ensure_chain_data_fresh(ctx, base_dir: Path) -> None:
    """Top-level orchestrator: refresh chain_onion if stale, then chain_timeline.

    Called from _prewarm_cache AFTER _cache_ready is set, so UI is not blocked.
    NEVER raises.
    """
    try:
        is_stale, reason = _chain_onion_outputs_stale(base_dir)
        if is_stale:
            print(f"[app][CHAIN_DATA] chain_onion stale ({reason}) — regenerating")
            ok, msg = _run_chain_onion_subprocess(base_dir)
            if not ok:
                print(f"[app][CHAIN_DATA] chain_onion regeneration failed: {msg}")
                return
        else:
            print(f"[app][CHAIN_DATA] chain_onion fresh — skipping subprocess")
        _refresh_chain_timeline(ctx, base_dir)
    except Exception as exc:
        print(f"[app][CHAIN_DATA] _ensure_chain_data_fresh outer exception: {exc}")
        traceback.print_exc()

# ── API class exposed to JavaScript ──────────────────────────
class Api:
    """
    Every public method here is callable from React via:
        window.pywebview.api.method_name(args)

    Returns are auto-serialized to JSON by PyWebView.
    """

    def __init__(self):
        self._pipeline_lock = threading.Lock()
        self._pipeline_status = {
            "running": False,
            "message": "",
            "error": None,
            "completed_run": None,
            "warnings": [],
        }
        self._cache_ready = threading.Event()
        self._chain_data_ready = threading.Event()
        threading.Thread(target=self._prewarm_cache, daemon=True).start()

    def _prewarm_cache(self):
        ctx = None
        try:
            from reporting.data_loader import load_run_context
            ctx = load_run_context(BASE_DIR)
            print("[app] Cache pre-warm complete.")
        except Exception as exc:
            print(f"[app] Cache pre-warm failed: {exc}")
        finally:
            self._cache_ready.set()

        # Phase 3: chain_onion + chain_timeline auto-refresh (background, after cache_ready)
        if ctx is None or getattr(ctx, "data_date", None) is None or getattr(ctx, "dernier_df", None) is None:
            # No usable ctx → skip refresh, but unblock chain_data_ready so callers don't deadlock
            print("[app][CHAIN_DATA] skipping refresh: ctx is missing data_date or dernier_df")
            self._chain_data_ready.set()
            return
        try:
            _ensure_chain_data_fresh(ctx, BASE_DIR)
        finally:
            self._chain_data_ready.set()

    def get_app_state(self):
        """Called by React on mount. Returns system state."""
        has_baseline = False
        current_run = None
        current_run_date = None
        total_runs = 0
        warnings = []

        if Path(RUN_MEMORY_DB).exists():
            runs = _query_db(RUN_MEMORY_DB,
                "SELECT run_number, status, is_stale, is_current, completed_at "
                "FROM runs ORDER BY run_number DESC")
            total_runs = len(runs)
            has_baseline = any(r["run_number"] == 0 for r in runs)

            # Find latest COMPLETED non-stale run
            for r in runs:
                if r["status"] == "COMPLETED" and not r["is_stale"]:
                    current_run = r["run_number"]
                    current_run_date = r["completed_at"]
                    break

            if not has_baseline:
                warnings.append("Run 0 baseline not found. Run bootstrap_run_zero.py first.")
        else:
            warnings.append("run_memory.db not found. Initialize the database first.")

        # Detect input files
        ged = None
        for f in INPUT_DIR.glob("*.xlsx"):
            if "GED" in f.name.upper() or "17CO" in f.name.upper():
                ged = str(f)
                break
        gf = None
        for f in INPUT_DIR.glob("*.xlsx"):
            if "grandfichier" in f.name.lower():
                gf = str(f)
                break

        return _sanitize_for_json({
            "has_baseline": has_baseline,
            "current_run": current_run,
            "current_run_date": current_run_date,
            "total_runs": total_runs,
            "ged_file_detected": ged,
            "gf_file_detected": gf,
            "data_dir": str(DATA_DIR),
            "pipeline_running": False,
            "app_version": "1.0.0",
            "warnings": warnings,
        })

    # ── Run explorer — delegate to existing service layer ──────

    def get_all_runs(self):
        """Delegate to src.run_explorer.get_all_runs."""
        try:
            from run_explorer import get_all_runs
            return _sanitize_for_json(get_all_runs(RUN_MEMORY_DB))
        except Exception as exc:
            return {"error": str(exc)}

    def get_run_summary(self, run_number):
        """Delegate to src.run_explorer.get_run_summary."""
        try:
            from run_explorer import get_run_summary
            return _sanitize_for_json(get_run_summary(RUN_MEMORY_DB, run_number))
        except Exception as exc:
            return {"error": str(exc)}

    def compare_runs(self, run_a, run_b):
        """Delegate to src.run_explorer.compare_runs."""
        try:
            from run_explorer import compare_runs
            return _sanitize_for_json(compare_runs(RUN_MEMORY_DB, run_a, run_b))
        except Exception as exc:
            return {"error": str(exc)}

    def export_run_bundle(self, run_number):
        """
        Export run artifacts as ZIP. Returns file path to the ZIP.
        Saves to output/exports/run_N_bundle.zip
        """
        try:
            export_dir = OUTPUT_DIR / "exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            zip_path = str(export_dir / f"run_{run_number}_bundle.zip")

            from run_explorer import export_run_bundle
            result_path = export_run_bundle(RUN_MEMORY_DB, run_number, zip_path)
            return _sanitize_for_json({"success": True, "path": result_path})
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    # ── Pipeline execution (background thread) ───────────────

    def run_pipeline_async(self, run_mode, ged_path=None, gf_path=None,
                           reports_dir=None):
        """
        Start pipeline in a background thread.
        Uses src.run_orchestrator.run_pipeline_controlled().
        Returns: {"started": True} or {"started": False, "errors": [...]}
        """
        with self._pipeline_lock:
            if self._pipeline_status["running"]:
                return _sanitize_for_json({"started": False, "errors": ["Pipeline is already running"]})

        # Use defaults from input/ if paths not provided
        if not ged_path:
            ged_path = self._detect_file("GED")
        if not gf_path:
            gf_path = self._detect_file("GF")

        # Pre-validate via orchestrator
        from run_orchestrator import validate_run_inputs
        validation = validate_run_inputs(run_mode, {
            "ged_path": ged_path,
            "gf_path": gf_path,
            "reports_dir": reports_dir,
        })
        if not validation["valid"]:
            return _sanitize_for_json({"started": False, "errors": validation["errors"],
                    "warnings": validation.get("warnings", [])})

        # Reset status and start worker thread
        with self._pipeline_lock:
            self._pipeline_status = {
                "running": True,
                "message": "Starting pipeline...",
                "error": None,
                "completed_run": None,
                "warnings": validation.get("warnings", []),
            }

        def worker():
            try:
                from run_orchestrator import run_pipeline_controlled
                self._update_status("Executing run_pipeline_controlled...")
                result = run_pipeline_controlled(
                    run_mode=run_mode,
                    ged_path=ged_path,
                    gf_path=gf_path,
                    reports_dir=reports_dir,
                )
                with self._pipeline_lock:
                    if result.get("success"):
                        # Invalidate reporting cache so dashboard reloads fresh data
                        try:
                            from reporting.data_loader import clear_cache
                            clear_cache()
                        except Exception:
                            pass
                        self._pipeline_status["completed_run"] = result.get("run_number")
                        self._pipeline_status["message"] = (
                            f"Run {result.get('run_number')} completed — "
                            f"{result.get('artifact_count', 0)} artifacts"
                        )
                        self._pipeline_status["warnings"] = result.get("warnings", [])
                    else:
                        self._pipeline_status["error"] = "; ".join(
                            result.get("errors", ["Unknown error"])
                        )
                        self._pipeline_status["message"] = "Pipeline failed"
            except Exception as exc:
                with self._pipeline_lock:
                    self._pipeline_status["error"] = f"{type(exc).__name__}: {exc}"
                    self._pipeline_status["message"] = "Pipeline failed with exception"
            finally:
                with self._pipeline_lock:
                    self._pipeline_status["running"] = False

        t = threading.Thread(target=worker, daemon=True)
        t.start()
        return _sanitize_for_json({"started": True, "warnings": validation.get("warnings", [])})

    def get_pipeline_status(self):
        """Polled by React every 500ms during execution."""
        with self._pipeline_lock:
            return _sanitize_for_json(dict(self._pipeline_status))

    def _update_status(self, msg):
        with self._pipeline_lock:
            self._pipeline_status["message"] = msg

    def _build_live_operational_numeros(self):
        try:
            from chain_onion.query_hooks import (
                QueryContext,
                get_live_operational,
                get_legacy_backlog,
            )
            co_dir = BASE_DIR / "output" / "chain_onion"
            if not co_dir.exists():
                return None, 0
            ctx = QueryContext(output_dir=co_dir)
            scores = ctx.scores()
            if scores.empty:
                return None, 0
            live_df = get_live_operational(ctx)
            legacy_df = get_legacy_backlog(ctx)
            if live_df.empty:
                return None, 0
            live_numeros = set(
                live_df["family_key"].dropna().astype(str).tolist()
            )
            legacy_count = len(legacy_df) if legacy_df is not None else 0
            return live_numeros, legacy_count
        except Exception as exc:
            print(exc)
            return None, 0

    @staticmethod
    def _apply_live_narrowing(focus_result, live_numeros, legacy_count):
        if live_numeros is None:
            return
        fdf = focus_result.focused_df
        if fdf is None:
            return
        if "numero_normalized" not in fdf.columns:
            return
        mask = fdf["numero_normalized"].astype(str).isin(live_numeros)
        focus_result.focused_df = fdf[mask].copy()
        focus_result.focused_doc_ids = set(
            focus_result.focused_df["doc_id"].tolist()
        )
        surviving = focus_result.focused_doc_ids
        focus_result.priority_queue = [
            x for x in focus_result.priority_queue
            if x.get("doc_id") in surviving
        ]
        focus_result.stats["focused_count"] = len(surviving)
        focus_result.stats["legacy_backlog_count"] = legacy_count

    # ── Dashboard / reporting data ─────────────────────────────

    def get_dashboard_data(self, focus=False, stale_days=90):
        """Full dashboard payload: KPIs + monthly chart + consultant/contractor summaries."""
        try:
            from reporting.data_loader import load_run_context
            from reporting.aggregator import (
                compute_project_kpis,
                compute_monthly_timeseries,
                compute_weekly_timeseries,
                compute_consultant_summary,
                compute_contractor_summary,
            )
            from reporting.focus_filter import apply_focus_filter, FocusConfig

            ctx = load_run_context(BASE_DIR)
            focus_config = FocusConfig(enabled=bool(focus), stale_threshold_days=int(stale_days))
            focus_result = apply_focus_filter(ctx, focus_config)
            _legacy_count = 0
            _live_numeros, _legacy_count = self._build_live_operational_numeros()
            self._apply_live_narrowing(focus_result, _live_numeros, _legacy_count)

            kpis = compute_project_kpis(ctx, focus_result=focus_result)
            consultants = compute_consultant_summary(ctx, focus_result=focus_result)
            contractors = compute_contractor_summary(ctx, focus_result=focus_result)

            if focus:
                timeseries = compute_weekly_timeseries(ctx, focus_result=focus_result)
            else:
                timeseries = compute_monthly_timeseries(ctx)

            payload = {
                "kpis": kpis,
                "monthly": timeseries,
                "consultants": consultants,
                "contractors": contractors,
                "focus": focus_result.stats,
            }
            if focus:
                payload["priority_queue"] = focus_result.priority_queue[:50]
                payload["legacy_backlog_count"] = _legacy_count

            return _sanitize_for_json(payload)
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return {"error": str(exc), "kpis": {}, "monthly": [], "consultants": [], "contractors": []}

    def get_consultant_list(self, focus=False, stale_days=90):
        """Consultant summary list for the Consultants page."""
        try:
            from reporting.data_loader import load_run_context
            from reporting.aggregator import compute_consultant_summary
            from reporting.focus_filter import apply_focus_filter, FocusConfig
            ctx = load_run_context(BASE_DIR)
            focus_config = FocusConfig(enabled=bool(focus), stale_threshold_days=int(stale_days))
            focus_result = apply_focus_filter(ctx, focus_config)
            _legacy_count = 0
            _live_numeros, _legacy_count = self._build_live_operational_numeros()
            self._apply_live_narrowing(focus_result, _live_numeros, _legacy_count)
            return _sanitize_for_json(compute_consultant_summary(ctx, focus_result=focus_result))
        except Exception as exc:
            return {"error": str(exc)}

    def get_contractor_list(self, focus=False, stale_days=90):
        """Contractor summary list for the Contractors page."""
        try:
            from reporting.data_loader import load_run_context
            from reporting.aggregator import compute_contractor_summary
            from reporting.focus_filter import apply_focus_filter, FocusConfig
            ctx = load_run_context(BASE_DIR)
            focus_config = FocusConfig(enabled=bool(focus), stale_threshold_days=int(stale_days))
            focus_result = apply_focus_filter(ctx, focus_config)
            _legacy_count = 0
            _live_numeros, _legacy_count = self._build_live_operational_numeros()
            self._apply_live_narrowing(focus_result, _live_numeros, _legacy_count)
            return _sanitize_for_json(compute_contractor_summary(ctx, focus_result=focus_result))
        except Exception as exc:
            return {"error": str(exc)}

    def get_consultant_fiche(self, consultant_name, focus=False, stale_days=90):
        """Full fiche data for one consultant."""
        import traceback
        try:
            from reporting.data_loader import load_run_context
            from reporting.consultant_fiche import build_consultant_fiche, resolve_consultant_name
            from reporting.focus_filter import apply_focus_filter, FocusConfig
            ctx = load_run_context(BASE_DIR)
            canonical = resolve_consultant_name(consultant_name)
            focus_config = FocusConfig(enabled=bool(focus), stale_threshold_days=int(stale_days))
            focus_result = apply_focus_filter(ctx, focus_config)
            result = build_consultant_fiche(ctx, canonical, focus_result=focus_result)
            return _sanitize_for_json(result)
        except Exception as exc:
            traceback.print_exc()
            return _sanitize_for_json({"error": str(exc), "consultant_name": consultant_name})

    def get_contractor_fiche(self, contractor_code, focus=False, stale_days=90):
        """Full fiche data for one contractor."""
        import traceback
        try:
            from reporting.data_loader import load_run_context
            from reporting.contractor_fiche import build_contractor_fiche
            from reporting.focus_filter import apply_focus_filter, FocusConfig
            ctx = load_run_context(BASE_DIR)
            focus_config = FocusConfig(enabled=bool(focus), stale_threshold_days=int(stale_days))
            focus_result = apply_focus_filter(ctx, focus_config)
            result = build_contractor_fiche(ctx, contractor_code, focus_result=focus_result)
            return _sanitize_for_json(result)
        except Exception as exc:
            traceback.print_exc()
            return _sanitize_for_json({"error": str(exc), "contractor_code": contractor_code})

    # ── Input validation (pre-flight check) ──────────────────

    def export_team_version(self):
        """Export the GF TEAM_VERSION artifact with a readable name."""
        import shutil
        import tempfile
        try:
            from reporting.data_loader import load_run_context
            ctx = load_run_context(BASE_DIR)
            team_path = ctx.artifact_paths.get("GF_TEAM_VERSION")
            if not team_path or not Path(team_path).exists():
                from reporting.data_loader import _get_artifact_path
                team_path = _get_artifact_path(str(DATA_DIR / "run_memory.db"), ctx.run_number, "GF_TEAM_VERSION")
            if not team_path or not Path(team_path).exists():
                return _sanitize_for_json({
                    "success": False,
                    "path": None,
                    "error": f"GF_TEAM_VERSION artifact not found for Run {ctx.run_number}. Run the pipeline first.",
                })
            date_str = ctx.data_date.strftime("%d_%m_%Y") if ctx.data_date else __import__("datetime").date.today().strftime("%d_%m_%Y")
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            dest = OUTPUT_DIR / f"Tableau de suivi de visa {date_str}.xlsx"
            with tempfile.NamedTemporaryFile(dir=str(OUTPUT_DIR), suffix=".xlsx", delete=False) as tmp:
                tmp_path = Path(tmp.name)
            shutil.copy2(str(team_path), str(tmp_path))
            if dest.exists():
                dest.unlink()
            tmp_path.rename(dest)
            return _sanitize_for_json({"success": True, "path": str(dest), "error": None})
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return _sanitize_for_json({"success": False, "path": None, "error": str(exc)})

    def get_doc_details(self, consultant_name, filter_key, lot_name=None, focus=False, stale_days=90):
        """Return document-level detail for a specific consultant fiche KPI cell."""
        import traceback
        try:
            from reporting.data_loader import load_run_context
            from reporting.consultant_fiche import (
                resolve_consultant_name, _filter_for_consultant,
                _attach_derived, _resolve_status_labels, _resolve_data_date,
            )
            ctx = load_run_context(BASE_DIR)
            if ctx.degraded_mode or ctx.docs_df is None:
                return _sanitize_for_json({"docs": [], "count": 0, "error": "Degraded mode - no GED data"})
            canonical = resolve_consultant_name(consultant_name)
            data_date = _resolve_data_date(ctx)
            s1, s2, s3 = _resolve_status_labels(ctx, canonical)
            docs = _filter_for_consultant(ctx, canonical)
            if docs.empty:
                return _sanitize_for_json({"docs": [], "count": 0})
            docs = _attach_derived(docs, data_date, s1=s1, s2=s2, s3=s3, ctx=ctx)

            if focus:
                from reporting.focus_filter import apply_focus_filter, FocusConfig
                focus_config = FocusConfig(enabled=True, stale_threshold_days=int(stale_days))
                focus_result = apply_focus_filter(ctx, focus_config)
                _legacy_count = 0
                _live_numeros, _legacy_count = self._build_live_operational_numeros()
                self._apply_live_narrowing(focus_result, _live_numeros, _legacy_count)
                id_col = "doc_id_resp" if "doc_id_resp" in docs.columns else "doc_id"
                focused_df = getattr(focus_result, "focused_df", None)
                if focused_df is not None and "_focus_owner" in focused_df.columns:
                    owned_ids = set()
                    for _, frow in focused_df.iterrows():
                        owners = frow.get("_focus_owner", [])
                        if isinstance(owners, list) and canonical in owners:
                            owned_ids.add(frow["doc_id"])
                    if canonical == "Maître d'Oeuvre EXE":
                        for _, frow in focused_df.iterrows():
                            if frow.get("_focus_owner_tier") == "MOEX":
                                owned_ids.add(frow["doc_id"])
                else:
                    owned_ids = focus_result.focused_doc_ids
                docs = docs[docs[id_col].isin(owned_ids)]
                if docs.empty:
                    return _sanitize_for_json({"docs": [], "count": 0})

            actual_filter = filter_key
            actual_lot = lot_name
            if filter_key.startswith("lot:"):
                parts = filter_key.split(":", 2)
                if len(parts) == 3:
                    actual_lot = parts[1]
                    actual_filter = parts[2]
                elif len(parts) == 2:
                    actual_lot = parts[1]
                    actual_filter = "total"
            if actual_lot:
                docs = docs[docs["_gf_sheet"] == actual_lot]

            if actual_filter == "total":
                mask = docs.index == docs.index
            elif actual_filter == "answered":
                mask = docs["_status_for_consultant"].isin({s1, s2, s3, "HM"})
            elif actual_filter in ("s1", s1):
                mask = docs["_status_for_consultant"] == s1
            elif actual_filter in ("s2", s2):
                mask = docs["_status_for_consultant"] == s2
            elif actual_filter in ("s3", s3):
                mask = docs["_status_for_consultant"] == s3
            elif actual_filter in ("hm", "HM"):
                mask = docs["_status_for_consultant"] == "HM"
            elif actual_filter == "open_count":
                mask = docs["_is_open"]
            elif actual_filter == "open_ok":
                mask = docs["_is_open"] & docs["_on_time"]
            elif actual_filter == "open_late":
                mask = docs["_is_open"] & ~docs["_on_time"]
            elif actual_filter == "open_blocking":
                mask = docs["_is_blocking"]
            elif actual_filter == "open_blocking_ok":
                mask = docs["_is_blocking"] & docs["_on_time"]
            elif actual_filter == "open_blocking_late":
                mask = docs["_is_blocking"] & ~docs["_on_time"]
            elif actual_filter == "open_non_blocking":
                mask = docs["_is_open"] & ~docs["_is_blocking"]
            else:
                mask = docs["_status_for_consultant"] == actual_filter
            filtered = docs[mask]

            import math as _math
            import pandas as _pd

            def _safe(v):
                if v is None or (isinstance(v, float) and _math.isnan(v)):
                    return None
                return v

            def _fmt_date(v):
                if v is None or _pd.isna(v):
                    return None
                if hasattr(v, "strftime"):
                    return v.strftime("%d/%m/%Y")
                return str(v)[:10]

            def _remaining_days(row):
                dl_col = "date_limite_resp" if "date_limite_resp" in row.index else "date_limite"
                dl = row.get(dl_col)
                if dl is None or (isinstance(dl, float) and _math.isnan(dl)) or _pd.isna(dl):
                    return None
                if hasattr(dl, "date"):
                    dl = dl.date()
                elif isinstance(dl, str):
                    try:
                        from datetime import datetime as _dt
                        dl = _dt.strptime(dl[:10], "%Y-%m-%d").date()
                    except Exception:
                        return None
                return (dl - data_date).days

            result_docs = []
            for _, row in filtered.iterrows():
                dl_col = "date_limite_resp" if "date_limite_resp" in row.index else "date_limite"
                result_docs.append({
                    "numero": _safe(row.get("numero")),
                    "indice": _safe(row.get("indice")),
                    "emetteur": _safe(row.get("emetteur")),
                    "titre": _safe(row.get("libelle_du_document", "")),
                    "date_soumission": _fmt_date(row.get("created_at")),
                    "date_limite": _fmt_date(row.get(dl_col)),
                    "remaining_days": _remaining_days(row),
                    "status": _safe(row.get("_status_for_consultant", "")),
                    "lot": _safe(row.get("_gf_sheet", "")),
                })
            result_docs.sort(key=lambda d: (
                0 if d["remaining_days"] is not None and d["remaining_days"] < 0 else 1,
                d["remaining_days"] if d["remaining_days"] is not None else 9999,
                str(d["numero"] or ""),
            ))
            return _sanitize_for_json({
                "docs": result_docs,
                "count": len(result_docs),
                "filter_key": filter_key,
                "consultant": consultant_name,
            })
        except Exception as exc:
            traceback.print_exc()
            return _sanitize_for_json({"docs": [], "count": 0, "error": str(exc)})

    def search_documents(self, query, focus=False, stale_days=30, limit=50):
        """Substring search across dernier_df. Returns ranked list of document summaries."""
        self._cache_ready.wait()
        try:
            from reporting.data_loader import load_run_context
            from reporting.document_command_center import search_documents
            ctx = load_run_context(BASE_DIR)
            results = search_documents(ctx, str(query or ""),
                                       bool(focus), int(stale_days), int(limit))
            return _sanitize_for_json(results)
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return _sanitize_for_json({"error": str(exc), "query": query})

    def get_document_command_center(self, numero, indice=None, focus=False, stale_days=30):
        """Full Document Command Center payload for one document."""
        self._cache_ready.wait()
        try:
            from reporting.data_loader import load_run_context
            from reporting.document_command_center import build_document_command_center
            ctx = load_run_context(BASE_DIR)
            payload = build_document_command_center(
                ctx, str(numero),
                None if indice is None else str(indice),
                bool(focus), int(stale_days),
            )
            return _sanitize_for_json(payload)
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return _sanitize_for_json({"error": str(exc), "numero": numero})

    def export_drilldown_xlsx(self, consultant_name, filter_key, lot_name=None, focus=False, stale_days=90):
        """Export the currently filtered drilldown documents to an Excel file."""
        import inspect
        import tempfile
        import traceback
        try:
            sig = inspect.signature(self.get_doc_details)
            if "focus" in sig.parameters:
                result = self.get_doc_details(consultant_name, filter_key, lot_name, focus, stale_days)
            else:
                result = self.get_doc_details(consultant_name, filter_key, lot_name)

            docs = result.get("docs", []) if isinstance(result, dict) else []
            if not docs:
                return _sanitize_for_json({"success": False, "error": "No documents to export"})

            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Drilldown"

            headers = [
                "Numero", "Indice", "Emetteur", "Titre", "Lot",
                "Date Soumission", "Date Echeance", "Jours Restants", "Statut",
            ]
            keys = [
                "numero", "indice", "emetteur", "titre", "lot",
                "date_soumission", "date_limite", "remaining_days", "status",
            ]

            header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            late_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
            for row_idx, doc in enumerate(docs, 2):
                is_late = doc.get("remaining_days") is not None and doc["remaining_days"] < 0
                for col_idx, key in enumerate(keys, 1):
                    value = doc.get(key)
                    if value is None:
                        value = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=10)
                    if is_late:
                        cell.fill = late_fill

            widths = [16, 6, 14, 50, 16, 14, 14, 12, 10]
            for col_idx, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(docs) + 1}"

            ws_meta = wb.create_sheet("Info")
            ws_meta["A1"] = "Consultant"
            ws_meta["B1"] = consultant_name
            ws_meta["A2"] = "Filtre"
            ws_meta["B2"] = filter_key
            ws_meta["A3"] = "Documents"
            ws_meta["B3"] = len(docs)
            for row_idx in range(1, 4):
                ws_meta.cell(row=row_idx, column=1).font = Font(bold=True)

            def _safe_filename(value):
                return "".join(c if c.isalnum() or c in " _-" else "_" for c in str(value)).strip()

            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            safe_consultant = _safe_filename(consultant_name)
            safe_filter = _safe_filename(filter_key)
            date_str = __import__("datetime").date.today().strftime("%d%m%Y")
            dest = OUTPUT_DIR / f"Drilldown_{safe_consultant}_{safe_filter}_{date_str}.xlsx"

            with tempfile.NamedTemporaryFile(dir=str(OUTPUT_DIR), suffix=".xlsx", delete=False) as tmp:
                tmp_path = Path(tmp.name)
            wb.save(str(tmp_path))
            if dest.exists():
                dest.unlink()
            tmp_path.rename(dest)

            return _sanitize_for_json({"success": True, "path": str(dest), "count": len(docs)})
        except Exception as exc:
            traceback.print_exc()
            return _sanitize_for_json({"success": False, "error": str(exc)})

    def get_documents_drilldown(self, kind, params=None, focus=False, stale_days=90):
        """Return drilldown rows for an Overview interaction (Phase 3 wiring)."""
        import traceback
        try:
            from reporting.data_loader import load_run_context
            from reporting.drilldown_builder import build_drilldown
            ctx = load_run_context(BASE_DIR)
            focus_result = None
            if focus:
                from reporting.focus_filter import apply_focus_filter, FocusConfig
                focus_config = FocusConfig(enabled=True, stale_threshold_days=int(stale_days))
                focus_result = apply_focus_filter(ctx, focus_config)
                _live_numeros, _legacy_count = self._build_live_operational_numeros()
                self._apply_live_narrowing(focus_result, _live_numeros, _legacy_count)
            payload = build_drilldown(
                ctx,
                str(kind or ""),
                params or {},
                focus_result=focus_result,
            )
            return _sanitize_for_json(payload)
        except Exception as exc:
            traceback.print_exc()
            return _sanitize_for_json({
                "error": str(exc),
                "rows": [],
                "total_count": 0,
                "truncated": False,
            })

    def validate_inputs(self, run_mode, ged_path=None, gf_path=None,
                        reports_dir=None):
        """
        Pre-validates without running. Delegates to run_orchestrator.validate_run_inputs().
        Called by UI on mode/file change for inline validation.
        """
        from run_orchestrator import validate_run_inputs
        return _sanitize_for_json(validate_run_inputs(run_mode, {
            "ged_path": ged_path or self._detect_file("GED"),
            "gf_path": gf_path or self._detect_file("GF"),
            "reports_dir": reports_dir,
        }))

    # ── File selection (native OS dialog) ────────────────────

    def select_file(self, file_type):
        """
        Opens native file dialog. Returns selected path or None.
        file_type: "ged" | "gf" | "mapping" | "report_dir"
        """
        windows = webview.windows
        if not windows:
            return None
        win = windows[0]

        if file_type == "report_dir":
            result = win.create_file_dialog(
                webview.FOLDER_DIALOG,
                directory=str(INPUT_DIR),
            )
            return result[0] if result else None
        else:
            result = win.create_file_dialog(
                webview.OPEN_DIALOG,
                directory=str(INPUT_DIR),
                file_types=("Excel files (*.xlsx)",),
            )
            return result[0] if result else None

    # ── Open file in Windows Explorer ────────────────────────

    def open_file_in_explorer(self, file_path):
        """Open the containing folder in Windows Explorer and select the file."""
        path = Path(file_path)
        if path.exists():
            subprocess.Popen(f'explorer /select,"{path}"')
            return _sanitize_for_json({"success": True})
        elif path.parent.exists():
            subprocess.Popen(f'explorer "{path.parent}"')
            return _sanitize_for_json({"success": True})
        return _sanitize_for_json({"success": False, "error": "Path not found"})

    # ── Helpers ──────────────────────────────────────────────

    def _detect_file(self, file_type):
        """Auto-detect input files from input/ directory."""
        if file_type == "GED":
            for f in INPUT_DIR.glob("*.xlsx"):
                if "GED" in f.name.upper() or "17CO" in f.name.upper():
                    return str(f)
        elif file_type == "GF":
            for f in INPUT_DIR.glob("*.xlsx"):
                if "grandfichier" in f.name.lower():
                    return str(f)
        return None

    # ── JANSA Dashboard UI adapter methods ──────────────────────

    def get_overview_for_ui(self, focus=False, stale_days=90):
        """Return OVERVIEW payload shaped for JANSA standalone dashboard."""
        self._cache_ready.wait()
        try:
            dashboard = self.get_dashboard_data(focus, stale_days)
            if isinstance(dashboard, dict) and "error" in dashboard:
                return _sanitize_for_json(dashboard)
            app_state = self.get_app_state()
            from reporting.ui_adapter import adapt_overview
            return _sanitize_for_json(adapt_overview(dashboard, app_state))
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return {"error": str(exc)}

    def get_consultants_for_ui(self, focus=False, stale_days=90):
        """Return CONSULTANTS list shaped for JANSA standalone dashboard."""
        self._cache_ready.wait()
        try:
            raw = self.get_consultant_list(focus, stale_days)
            if isinstance(raw, dict) and "error" in raw:
                return _sanitize_for_json(raw)
            from reporting.ui_adapter import adapt_consultants
            return _sanitize_for_json(adapt_consultants(raw))
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return {"error": str(exc)}

    def get_contractors_for_ui(self, focus=False, stale_days=90):
        """Return CONTRACTORS + CONTRACTORS_LIST for JANSA standalone dashboard."""
        self._cache_ready.wait()
        try:
            raw = self.get_contractor_list(focus, stale_days)
            if isinstance(raw, dict) and "error" in raw:
                return _sanitize_for_json(raw)
            from reporting.ui_adapter import adapt_contractors_lookup, adapt_contractors_list
            return _sanitize_for_json({
                "lookup": adapt_contractors_lookup(raw),
                "list":   adapt_contractors_list(raw, focus=focus),
            })
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return {"error": str(exc)}

    def get_fiche_for_ui(self, consultant_name, focus=False, stale_days=90):
        """Return FICHE_DATA for one consultant — already shaped correctly."""
        return self.get_consultant_fiche(consultant_name, focus, stale_days)

    def get_contractor_fiche_for_ui(self, contractor_code, focus=False, stale_days=90):
        """Return contractor fiche payload for the UI.

        Thin wrapper around get_contractor_fiche. Adds canonical-name
        enrichment so the UI header shows 'Bentin' rather than 'BEN'.
        Also merges the V1 quality payload (5 KPIs, polar histogram,
        dormant lists, open/finished, long-chains) under payload['quality'].
        A quality-build failure degrades to payload['quality'] = {'error': str}
        rather than failing the whole fiche.
        """
        payload = self.get_contractor_fiche(contractor_code, focus=focus, stale_days=stale_days)
        if isinstance(payload, dict) and "error" not in payload:
            try:
                from reporting.contractor_fiche import resolve_emetteur_name
                raw_code = payload.get("contractor_code") or contractor_code
                payload["contractor_name"] = resolve_emetteur_name(raw_code)
            except Exception:
                pass

            # Merge quality payload (best-effort — degrade gracefully)
            try:
                from reporting.data_loader import load_run_context
                from reporting.contractor_quality import (
                    build_contractor_quality,
                    build_contractor_quality_peer_stats,
                )
                ctx = load_run_context(BASE_DIR)
                peer = build_contractor_quality_peer_stats(ctx)
                payload["quality"] = build_contractor_quality(
                    ctx, contractor_code, peer_stats=peer
                )
            except Exception as exc:
                import traceback
                traceback.print_exc()
                payload["quality"] = {"error": str(exc)}

        return payload

    def get_chain_onion_intel(self, limit=20):
        """Return top_issues + dashboard_summary from chain_onion output for the UI panel."""
        try:
            co_dir = OUTPUT_DIR / "chain_onion"
            top_path = co_dir / "top_issues.json"
            dash_path = co_dir / "dashboard_summary.json"

            top_issues = []
            if top_path.exists():
                data = json.loads(top_path.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    top_issues = data[:int(limit)]

            summary = {}
            if dash_path.exists():
                summary = json.loads(dash_path.read_text(encoding="utf-8"))

            from reporting.narrative_translation import translate_top_issue
            top_issues = [translate_top_issue(i) for i in top_issues]
            return _sanitize_for_json({"top_issues": top_issues, "summary": summary})
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return {"top_issues": [], "summary": {}, "error": str(exc)}

    def get_chain_timeline(self, numero):
        """Return the chain timeline payload for one document by numero.

        Phase 4 (Document Command Center) consumes this for the
        "Chronologie de la chaîne" panel section.

        Returns the per-chain dict (header + indices[] + totals + tag flags +
        attribution_breakdown), or {"error": ...} if the artifact is unavailable
        or the numero has no chain.
        """
        self._chain_data_ready.wait()
        try:
            from reporting.chain_timeline_attribution import load_chain_timeline_artifact
            intermediate = BASE_DIR / "output" / "intermediate"
            timelines = load_chain_timeline_artifact(intermediate)
            # Normalize numero to chain_onion's zero-padded 6-digit family_key form
            s = str(numero).strip()
            if s.isdigit():
                family_key = s.zfill(6)
            else:
                family_key = s
            payload = timelines.get(family_key)
            if payload is None:
                return _sanitize_for_json({"error": f"No chain timeline for numero {numero!r}", "numero": numero})
            return _sanitize_for_json(payload)
        except FileNotFoundError as exc:
            return _sanitize_for_json({"error": str(exc), "numero": numero})
        except Exception as exc:
            import traceback
            traceback.print_exc()
            return _sanitize_for_json({"error": str(exc), "numero": numero})

# ── Main ─────────────────────────────────────────────────────
def main():
    ui_url = _resolve_ui()
    browser_mode = "--browser" in sys.argv
    if browser_mode:
        target = _ui_target_for_browser(ui_url)
        print(f"[app] Opening browser mode: {target}")
        webbrowser.open(target)
        return
    print("[app] Creating API...")
    api = Api()
    print("[app] Creating WebView window...")
    webview.create_window(
        title="JANSA VISASIST \u2014 P17&CO T2",
        url=ui_url,
        js_api=api,
        width=1400,
        height=900,
        min_size=(1024, 700),
        text_select=False,
    )
    print("[app] Starting WebView...")
    try:
        webview.start(
            debug="--debug" in sys.argv,
        )
    except Exception as exc:
        target = _ui_target_for_browser(ui_url)
        print(f"[app] Embedded WebView startup failed: {exc}")
        print(f"[app] Falling back to browser mode: {target}")
        webbrowser.open(target)
        return


if __name__ == "__main__":
    main()
