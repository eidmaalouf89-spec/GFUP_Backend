"""
scripts/build_shadow_flat_ged.py — Phase 8B.9 shadow corrected FLAT model.

Starts from raw_ged_trace.csv and applies, in order, the documented §5
projection rules (the same buckets used in audit sheets 02–09):

    1. Drop  MALFORMED_RESPONSE       — _MISSING_ numero rows
    2. Drop  ACTIVE_VERSION_PROJECTION — RAW carries an old indice
    3. Apply SAS_CYCLE_COLLAPSED       — collapse C1+C2 (no-op in run 0)
    4. Drop  DUPLICATE_FAVORABLE_KEPT  — REF dropped when same-actor VAO/VSO present in FLAT
    5. Drop  DUPLICATE_MERGED          — multiple RAW rows for same (num,ind,actor) → keep first
    6. Drop  EXCEPTION_COLUMN          — non-actor columns
    7. Drop  NON_OPERATIONAL_RESPONSE  — status not in operational set
    8. Drop  UNKNOWN_ACTOR             — Exception List actors

The set of applied rules MUST match the §5 taxonomy exactly.  We do NOT
invent new rules here — every drop is justified by an entry that already
appears in audit sheets 02..09.

The classification index is materialised from the existing audit workbook
(raw_flat_reconcile.xlsx, sheets 03/06/07) — reusing the classifier outputs
from 8B.4/8B.5/8B.6 keeps the shadow reproducible and traceable.

Inputs (read only):
    output/debug/raw_ged_trace.csv
    output/debug/flat_ged_trace.csv
    output/debug/raw_flat_reconcile.xlsx  (sheets 01..10)
    output/debug/report_to_flat_trace.json (from 8B.8)

Outputs:
    output/debug/SHADOW_FLAT_GED_OPERATIONS.csv
    output/debug/SHADOW_FLAT_GED_TRACE.xlsx
    output/debug/raw_flat_reconcile.xlsx — appends sheet 11_SHADOW_DIFFS

Stdout summary line — exact format (per §16.4):
    SHADOW_FLAT: shadow_rows=<n> current_flat_rows=<n> shadow_minus_flat=<n> unexplained=<n>
"""
from __future__ import annotations

import csv
import json
import subprocess
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_CSV   = REPO_ROOT / "output" / "debug" / "raw_ged_trace.csv"
FLAT_CSV  = REPO_ROOT / "output" / "debug" / "flat_ged_trace.csv"
AUDIT_XLSX = REPO_ROOT / "output" / "debug" / "raw_flat_reconcile.xlsx"
REPORT_JSON = REPO_ROOT / "output" / "debug" / "report_to_flat_trace.json"

OUT_OPS_CSV = REPO_ROOT / "output" / "debug" / "SHADOW_FLAT_GED_OPERATIONS.csv"
OUT_TRACE_XLSX = REPO_ROOT / "output" / "debug" / "SHADOW_FLAT_GED_TRACE.xlsx"

# Operational status set from 8B.5
_OPERATIONAL_STATUSES = {"VAO", "VSO", "VSO-SAS", "REF", "HM", "FAV", "SUS", "DEF"}

# Shadow operations columns — match production GED_OPERATIONS shape
_SHADOW_OPS_COLS = [
    "numero", "indice", "lot", "emetteur", "titre",
    "step_order", "step_type",
    "actor_canonical", "cycle_id",
    "status_raw", "status_clean",
    "submission_date", "response_date", "deadline_date",
    "date_status_type",
    "comment_raw",
    "pj_flag",
    "shadow_source",     # which RAW row(s) produced this shadow row
    "shadow_rule_kept",  # the §5 classification this row evaded
]

_SHEET11_COLS = [
    "numero", "indice",
    "current_flat_state", "shadow_flat_state",
    "delta_explanation",
]


# ── Loaders ───────────────────────────────────────────────────────────────────

def _read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def _load_audit_classifications(xlsx_path: Path) -> dict:
    """Build per-row classification indices from the audit workbook.

    Returns dict with keys:
        sheet03  — list of {numero, indice, actor_canonical, side, classification}
        sheet06  — list of {numero, indice, actor_canonical, side, classification, status_raw, status_clean}
        sheet07  — list of {numero, indice, classification, raw_cycle, ...}
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    def _read(sheet_name):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = rows[0]
        return [dict(zip(header, r)) for r in rows[1:]]

    out = {
        "sheet03": _read("03_ACTOR_CALL_DIFFS"),
        "sheet06": _read("06_RESPONSE_DIFFS"),
        "sheet07": _read("07_SAS_REF_TRACE"),
    }
    wb.close()
    return out


def _load_flat_summary(flat_csv: Path) -> dict:
    """Per-(numero, indice) summary of the current FLAT side.

    Returns dict with keys:
        flat_pairs                  — set of (numero, indice)
        flat_actor_count_per_pair   — {(numero, indice): #actor_canonical observations}
    """
    flat_pairs: set = set()
    counts: Counter = Counter()
    with flat_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            num = r.get("numero", "").strip()
            ind = r.get("indice", "").strip()
            if not num:
                continue
            flat_pairs.add((num, ind))
            counts[(num, ind)] += 1
    return {"flat_pairs": flat_pairs, "flat_actor_counts": dict(counts)}


# ── Per-row classifier (mirrors §5 taxonomy rules) ───────────────────────────

def _build_dropset_from_audit(audit: dict) -> dict:
    """For each (numero, indice, actor_canonical, side='RAW_ONLY') tuple in the
    audit, record the §5 classification.  This is the authoritative dropset
    used to project RAW → shadow without re-running the classifiers.
    """
    drop_index: dict[tuple, str] = {}
    sheet07_index: dict[tuple, dict] = {}
    unexplained_07: list[dict] = []

    for r in audit["sheet03"]:
        if r.get("side") != "RAW_ONLY":
            continue
        cls = r.get("classification") or "UNEXPLAINED"
        key = (str(r["numero"]).strip(),
               str(r["indice"]).strip(),
               (r.get("actor_canonical") or "").strip())
        drop_index.setdefault(key, cls)

    for r in audit["sheet06"]:
        if r.get("side") != "RAW_ONLY":
            continue
        cls = r.get("classification") or "UNEXPLAINED"
        key = (str(r["numero"]).strip(),
               str(r["indice"]).strip(),
               (r.get("actor_canonical") or "").strip())
        # don't overwrite a stronger sheet 03 verdict
        drop_index.setdefault(key, cls)

    for r in audit["sheet07"]:
        # Sheet 07 carries classification=None for the 283 matched rows
        # (i.e. RAW SAS REF that has a clean FLAT counterpart — no projection
        # needed).  Only treat explicit string verdicts as drop-buckets; None
        # means "matched cleanly, keep as-is".
        cls_raw = r.get("classification")
        cls = cls_raw if cls_raw else "MATCHED"
        key = (str(r["numero"]).strip(),
               str(r["indice"]).strip(),
               "0-SAS")
        sheet07_index[key] = {"classification": cls, "raw": r}
        if cls == "UNEXPLAINED":
            unexplained_07.append(r)

    return {
        "drop_index": drop_index,
        "sheet07_index": sheet07_index,
        "unexplained_07": unexplained_07,
    }


def _classify_raw_event_for_shadow(
    row: dict,
    drop_index: dict,
    sheet07_index: dict,
    raw_row_count_per_key: dict,
) -> tuple[str, str]:
    """Return (verdict, rule_applied) for a single RAW event.

    verdict ∈ {"KEEP", "DROP"}
    rule_applied: §5 bucket name, "OPERATIONAL_KEEP", or "UNEXPLAINED_KEEP"
    """
    num   = (row.get("numero") or "").strip()
    ind   = (row.get("indice") or "").strip()
    actor = (row.get("actor_canonical") or "").strip()
    status_clean = (row.get("status_clean") or "").strip()
    et    = (row.get("event_type") or "").strip()
    cycle = (row.get("cycle_id") or "").strip()

    # 1 — MALFORMED_RESPONSE: _MISSING_ numeros
    if num.startswith("_MISSING_"):
        return "DROP", "MALFORMED_RESPONSE"

    # 8 — UNKNOWN_ACTOR: Exception List
    if actor == "Exception List" or actor == "":
        return "DROP", "UNKNOWN_ACTOR"

    # 6 — EXCEPTION_COLUMN: collapsed under UNKNOWN_ACTOR for current data
    # (Exception List columns are the only EXCEPTION_COLUMN producers.)

    # 7 — NON_OPERATIONAL_RESPONSE: only applies to RESPONSE rows with a status
    if et == "RESPONSE" and status_clean and status_clean.upper() not in _OPERATIONAL_STATUSES:
        return "DROP", "NON_OPERATIONAL_RESPONSE"

    # 2 — ACTIVE_VERSION_PROJECTION: from sheet 03/06 (audit-supplied verdicts)
    key = (num, ind, actor)
    audit_verdict = drop_index.get(key)
    if audit_verdict in ("ACTIVE_VERSION_PROJECTION",
                         "DUPLICATE_MERGED"):
        return "DROP", audit_verdict

    # 4 — DUPLICATE_FAVORABLE_KEPT: SAS REF rows so classified by sheet 07
    if actor == "0-SAS" and (status_clean or "").upper() == "REF":
        s07 = sheet07_index.get(key)
        if s07:
            cls = s07["classification"]
            if cls == "DUPLICATE_FAVORABLE_KEPT":
                return "DROP", "DUPLICATE_FAVORABLE_KEPT"
            if cls == "DUPLICATE_MERGED":
                return "DROP", "DUPLICATE_MERGED"
            if cls == "ACTIVE_VERSION_PROJECTION":
                return "DROP", "ACTIVE_VERSION_PROJECTION"
            if cls == "MALFORMED_RESPONSE":
                return "DROP", "MALFORMED_RESPONSE"
            if cls == "UNEXPLAINED":
                return "KEEP", "UNEXPLAINED_KEEP"
            # cls == "MATCHED" → fall through to OPERATIONAL_KEEP

    # 5 — DUPLICATE_MERGED at the (num, ind, actor) tuple level (≥2 RAW rows
    # collapse to one).  The classifier kept this for non-SAS already; here
    # we additionally collapse same-key duplicates that didn't reach sheet 03.
    if raw_row_count_per_key.get(key, 0) > 1:
        # The first occurrence is the canonical row; subsequent are merged.
        if not row.get("__is_canonical_for_key__", False):
            return "DROP", "DUPLICATE_MERGED"

    return "KEEP", "OPERATIONAL_KEEP"


# ── Core projection ───────────────────────────────────────────────────────────

def project_shadow(raw_rows: list[dict], audit_index: dict) -> dict:
    """Apply §5 rules to RAW rows in order, returning shadow rows + per-rule deltas."""
    drop_index = audit_index["drop_index"]
    sheet07_index = audit_index["sheet07_index"]

    # First pass: count RAW rows per (num, ind, actor) for DUPLICATE_MERGED
    counts_per_key: Counter = Counter()
    for r in raw_rows:
        num   = (r.get("numero") or "").strip()
        ind   = (r.get("indice") or "").strip()
        actor = (r.get("actor_canonical") or "").strip()
        if num and not num.startswith("_MISSING_") and actor:
            counts_per_key[(num, ind, actor)] += 1

    seen_canonical: set = set()
    rule_buckets: dict[str, list] = defaultdict(list)
    shadow_rows: list[dict] = []
    unexplained_residual: list[dict] = []

    for raw_row in raw_rows:
        # mark canonical-for-key on first appearance of duplicates
        num   = (raw_row.get("numero") or "").strip()
        ind   = (raw_row.get("indice") or "").strip()
        actor = (raw_row.get("actor_canonical") or "").strip()
        key = (num, ind, actor)
        is_canonical = (counts_per_key.get(key, 0) > 0
                         and key not in seen_canonical)
        if is_canonical:
            seen_canonical.add(key)
        raw_row["__is_canonical_for_key__"] = is_canonical

        verdict, rule = _classify_raw_event_for_shadow(
            raw_row, drop_index, sheet07_index, counts_per_key
        )

        rule_buckets[rule].append(raw_row)
        if verdict == "KEEP":
            shadow_rows.append(_to_shadow_op_row(raw_row, rule))
            if rule == "UNEXPLAINED_KEEP":
                unexplained_residual.append(raw_row)

    return {
        "shadow_rows":          shadow_rows,
        "rule_buckets":         dict(rule_buckets),
        "unexplained_residual": unexplained_residual,
        "counts_per_key":       counts_per_key,
    }


def _to_shadow_op_row(raw: dict, rule: str) -> dict:
    """Convert a RAW event row into a shadow GED_OPERATIONS-shaped row."""
    return {
        "numero":            raw.get("numero", ""),
        "indice":            raw.get("indice", ""),
        "lot":               raw.get("lot", ""),
        "emetteur":          raw.get("emetteur", ""),
        "titre":             raw.get("titre", ""),
        "step_order":        "",  # filled later
        "step_type":         _step_type_from_actor(raw.get("actor_canonical", "")),
        "actor_canonical":   raw.get("actor_canonical", ""),
        "cycle_id":          raw.get("cycle_id", ""),
        "status_raw":        raw.get("status_raw", ""),
        "status_clean":      raw.get("status_clean", ""),
        "submission_date":   raw.get("submission_date", ""),
        "response_date":     raw.get("response_date", ""),
        "deadline_date":     raw.get("deadline_date", ""),
        "date_status_type":  raw.get("date_status_type", ""),
        "comment_raw":       raw.get("comment_raw", ""),
        "pj_flag":           raw.get("pj_flag", ""),
        "shadow_source":     f"raw_excel_row={raw.get('source_excel_row', '')}",
        "shadow_rule_kept":  rule,
    }


def _step_type_from_actor(actor: str) -> str:
    if actor == "0-SAS":
        return "SAS"
    if actor in ("Maître d'Oeuvre EXE", "Maître d'Oeuvre EXE", "MOEX"):
        return "MOEX"
    if actor == "":
        return ""
    return "CONSULTANT"


def assign_step_order(shadow_rows: list[dict]) -> None:
    """Fill step_order in-place per (numero, indice), starting at 1."""
    grouped: dict[tuple, list[dict]] = defaultdict(list)
    for r in shadow_rows:
        grouped[(r["numero"], r["indice"])].append(r)
    for rows in grouped.values():
        for i, r in enumerate(rows, 1):
            r["step_order"] = i


# ── Sheet 11 deltas ───────────────────────────────────────────────────────────

def compute_sheet11_diffs(
    shadow_rows: list[dict],
    flat_csv: Path,
) -> list[dict]:
    """One row per (numero, indice) where current FLAT differs from shadow.

    Two simple deltas:
      * pair-existence: shadow has the pair but FLAT does not, or vice-versa
      * actor-count:    shadow has K rows for the pair, FLAT has J ≠ K
    """
    # FLAT counts per pair
    flat_counts: Counter = Counter()
    flat_pairs: set = set()
    with flat_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            num = (r.get("numero") or "").strip()
            ind = (r.get("indice") or "").strip()
            if not num or num.startswith("_MISSING_"):
                continue
            flat_pairs.add((num, ind))
            flat_counts[(num, ind)] += 1

    # Shadow counts per pair
    shadow_counts: Counter = Counter()
    shadow_pairs: set = set()
    for r in shadow_rows:
        num = r.get("numero", "").strip()
        ind = r.get("indice", "").strip()
        if not num:
            continue
        shadow_pairs.add((num, ind))
        shadow_counts[(num, ind)] += 1

    diffs: list[dict] = []
    all_pairs = flat_pairs | shadow_pairs
    for (num, ind) in sorted(all_pairs):
        fc = flat_counts.get((num, ind), 0)
        sc = shadow_counts.get((num, ind), 0)
        if fc == sc:
            continue
        if fc == 0:
            current_state, shadow_state = "ABSENT", f"{sc} rows"
            explanation = "shadow includes pair, current FLAT excludes (UNEXPLAINED_KEEP)"
        elif sc == 0:
            current_state, shadow_state = f"{fc} rows", "ABSENT"
            explanation = "current FLAT includes pair, shadow drops it (all rows classified into §5 buckets)"
        else:
            current_state, shadow_state = f"{fc} rows", f"{sc} rows"
            explanation = (
                "row count mismatch — typically RAW had additional rows "
                "the §5 taxonomy keeps that current FLAT had already merged "
                "or vice-versa"
            )
        diffs.append({
            "numero": num, "indice": ind,
            "current_flat_state": current_state,
            "shadow_flat_state":  shadow_state,
            "delta_explanation":  explanation,
        })

    return diffs


# ── Writers ───────────────────────────────────────────────────────────────────

def write_shadow_csv(shadow_rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=_SHADOW_OPS_COLS)
        writer.writeheader()
        for r in shadow_rows:
            writer.writerow({c: r.get(c, "") for c in _SHADOW_OPS_COLS})


def write_shadow_trace_xlsx(
    rule_buckets: dict[str, list],
    unexplained_residual: list[dict],
    path: Path,
) -> None:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Summary sheet
    ws = wb.create_sheet("00_SUMMARY")
    ws.append(["rule", "rows_affected", "verdict"])
    summary_order = [
        ("MALFORMED_RESPONSE",        "DROP"),
        ("ACTIVE_VERSION_PROJECTION", "DROP"),
        ("DUPLICATE_FAVORABLE_KEPT",  "DROP"),
        ("DUPLICATE_MERGED",          "DROP"),
        ("EXCEPTION_COLUMN",          "DROP"),
        ("NON_OPERATIONAL_RESPONSE",  "DROP"),
        ("UNKNOWN_ACTOR",             "DROP"),
        ("OPERATIONAL_KEEP",          "KEEP"),
        ("UNEXPLAINED_KEEP",          "KEEP"),
    ]
    seen_rules = set(rule_buckets.keys())
    for rule, verdict in summary_order:
        ws.append([rule, len(rule_buckets.get(rule, [])), verdict])
        seen_rules.discard(rule)
    for rule in sorted(seen_rules):
        ws.append([rule, len(rule_buckets[rule]), "OTHER"])

    # Per-rule sheets — first 200 rows of each (large rules truncated for size)
    for rule, rows in sorted(rule_buckets.items()):
        sheet_name = ("RULE_" + rule)[:31]  # Excel limit
        ws_r = wb.create_sheet(sheet_name)
        cols = ["numero", "indice", "actor_canonical", "cycle_id",
                "event_type", "status_raw", "status_clean",
                "source_excel_row"]
        ws_r.append(cols)
        for r in rows[:200]:
            ws_r.append([r.get(c, "") for c in cols])

    # Residual UNEXPLAINED kept-rows sheet for §17 gate input
    ws_u = wb.create_sheet("UNEXPLAINED_KEPT")
    cols = ["numero", "indice", "actor_canonical", "cycle_id",
            "event_type", "status_raw", "status_clean",
            "source_excel_row"]
    ws_u.append(cols)
    for r in unexplained_residual[:500]:
        ws_u.append([r.get(c, "") for c in cols])

    wb.save(path)


def append_sheet11_to_audit(diffs: list[dict], xlsx_path: Path) -> None:
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    if "11_SHADOW_DIFFS" in wb.sheetnames:
        del wb["11_SHADOW_DIFFS"]
    ws = wb.create_sheet("11_SHADOW_DIFFS")
    ws.append(_SHEET11_COLS)
    for d in diffs:
        ws.append([d.get(c, "") for c in _SHEET11_COLS])
    wb.save(xlsx_path)
    del wb


def trust_but_verify(
    ops_csv: Path,
    trace_xlsx: Path,
    audit_xlsx: Path,
    expected_shadow_rows: int,
    expected_sheet11_rows: int,
) -> None:
    code = f"""
import csv, os, zipfile
from openpyxl import load_workbook
from collections import Counter

# CSV
csv_path = r{str(ops_csv)!r}
csv_size = os.path.getsize(csv_path)
with open(csv_path, "r", encoding="utf-8", newline="") as f:
    rows = list(csv.DictReader(f))
print(f"shadow_csv: size={{csv_size}} rows={{len(rows)}}")
assert len(rows) == {expected_shadow_rows}, f"shadow csv row mismatch: {{len(rows)}} != {expected_shadow_rows}"

# trace XLSX
tx_path = r{str(trace_xlsx)!r}
tx_size = os.path.getsize(tx_path)
with zipfile.ZipFile(tx_path) as z:
    assert z.testzip() is None, "trace ZIP corrupt"
print(f"shadow_trace_xlsx: size={{tx_size}}")
wb = load_workbook(tx_path, read_only=True, data_only=True)
print(f"  sheets={{wb.sheetnames}}")
wb.close()

# audit XLSX (sheet 11)
ax_path = r{str(audit_xlsx)!r}
ax_size = os.path.getsize(ax_path)
with zipfile.ZipFile(ax_path) as z:
    assert z.testzip() is None, "audit ZIP corrupt"
wb = load_workbook(ax_path, read_only=True, data_only=True)
sheet11_rows = sum(1 for _ in wb["11_SHADOW_DIFFS"].iter_rows())
all_sheets = wb.sheetnames
wb.close()
print(f"audit_xlsx: size={{ax_size}} sheets={{all_sheets}}")
print(f"  11_data_rows={{sheet11_rows-1}}")
assert sheet11_rows - 1 == {expected_sheet11_rows}, f"sheet 11 row mismatch: {{sheet11_rows-1}} != {expected_sheet11_rows}"
print("ZIP_integrity=OK")
"""
    proc = subprocess.run(["python", "-c", code], capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(
            f"Subprocess self-check FAILED:\nSTDOUT: {proc.stdout}\nSTDERR: {proc.stderr}"
        )
    print(proc.stdout.strip())


# ── Driver ────────────────────────────────────────────────────────────────────

def compute_shadow() -> dict:
    raw_rows = _read_csv(RAW_CSV)
    audit = _load_audit_classifications(AUDIT_XLSX)
    audit_index = _build_dropset_from_audit(audit)

    project = project_shadow(raw_rows, audit_index)
    assign_step_order(project["shadow_rows"])

    shadow_rows = project["shadow_rows"]
    rule_buckets = project["rule_buckets"]
    unexplained = project["unexplained_residual"]

    # FLAT current count: post-_MISSING_ filter, all rows
    flat_summary = _load_flat_summary(FLAT_CSV)
    current_flat_rows = sum(flat_summary["flat_actor_counts"].values())

    diffs = compute_sheet11_diffs(shadow_rows, FLAT_CSV)

    summary_line = (
        f"SHADOW_FLAT: shadow_rows={len(shadow_rows)} "
        f"current_flat_rows={current_flat_rows} "
        f"shadow_minus_flat={len(shadow_rows) - current_flat_rows} "
        f"unexplained={len(unexplained)}"
    )

    return {
        "shadow_rows":          shadow_rows,
        "rule_buckets":         rule_buckets,
        "unexplained":          unexplained,
        "current_flat_rows":    current_flat_rows,
        "diffs":                diffs,
        "summary_line":         summary_line,
        "rule_counts":          {k: len(v) for k, v in rule_buckets.items()},
    }


def main() -> None:
    print(f"Loading RAW trace:    {RAW_CSV}")
    print(f"Loading FLAT trace:   {FLAT_CSV}")
    print(f"Loading audit XLSX:   {AUDIT_XLSX}")
    result = compute_shadow()

    print(f"Shadow rows:          {len(result['shadow_rows'])}")
    print(f"Current FLAT rows:    {result['current_flat_rows']}")
    print(f"Sheet 11 diff rows:   {len(result['diffs'])}")
    print(f"Per-rule counts:      {result['rule_counts']}")
    print(f"UNEXPLAINED residual: {len(result['unexplained'])}")

    write_shadow_csv(result["shadow_rows"], OUT_OPS_CSV)
    write_shadow_trace_xlsx(result["rule_buckets"], result["unexplained"], OUT_TRACE_XLSX)
    append_sheet11_to_audit(result["diffs"], AUDIT_XLSX)

    trust_but_verify(
        OUT_OPS_CSV, OUT_TRACE_XLSX, AUDIT_XLSX,
        expected_shadow_rows=len(result["shadow_rows"]),
        expected_sheet11_rows=len(result["diffs"]),
    )

    print(result["summary_line"])


if __name__ == "__main__":
    main()
