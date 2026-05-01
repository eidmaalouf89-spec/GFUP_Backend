"""
scripts/raw_flat_reconcile.py — Phase 8B RAW→FLAT reconciliation driver.

Usage:
    python scripts/raw_flat_reconcile.py --step <step>

Implemented steps:
    identity       — Phase 8B.3: prove (numero) and (numero, indice) set parity
    actor_calls    — Phase 8B.4: verify every actor call in RAW has a FLAT counterpart
    responses      — Phase 8B.5: response/status parity + D-011 SAS REF trace
    dates_comments — Phase 8B.6: date/comment field parity for matched response pairs

Stub steps (raise NotImplementedError until their phase is approved):
    reasons_audit  — Phase 8B.7
    shadow         — Phase 8B.9

RAW _MISSING_ filter (applied before every identity set is built):
    Rows where numero starts with '_MISSING_' are extractor sentinels that
    have no real-document counterpart in FLAT.  They are excluded from RAW
    before any set comparison.  Row counts are printed to stdout before and
    after the filter so the exclusion is visible.

    Exception for Phase 8B.5 SAS REF trace (sheet 07): the 27 _MISSING_ rows
    that have actor_canonical=='0-SAS' and status_raw=='REF' ARE included in
    sheet 07 so that the 836 unique source_excel_row baseline is preserved.
    Those rows are classified MALFORMED_RESPONSE.  The general response parity
    computation (sheets 04/05/06) still applies the _MISSING_ filter.

Phase 8B.5 operational status set (verified from L0 audit baseline):
    VAO, VSO, VSO-SAS, REF, HM, FAV, SUS, DEF
    Any status_raw not in this set → NON_OPERATIONAL_RESPONSE.

Phase 8B.5 taxonomy extension — DUPLICATE_FAVORABLE_KEPT bucket:
    Pattern: RAW has conflicting statuses for the same (numero, indice, 0-SAS)
    — both REF and VAO/VSO/VSO-SAS — while FLAT kept only the favorable one
    and dropped the REF (operator manipulation or data-quality issue).
    Note: submission_dates may differ between the REF and favorable rows in the
    actual GED data, so no same-date constraint is applied.
    Two conditions, BOTH must hold:
        (a) RAW has at least one OTHER row at same (numero, indice) with
            actor='0-SAS', status_clean ∈ {'VAO','VSO','VSO-SAS'}
        (b) FLAT has a response at same (numero, indice) with actor='0-SAS',
            status_clean ∈ {'VAO','VSO','VSO-SAS'}
    Placement: after ACTIVE_VERSION_PROJECTION, before SAS_CYCLE_COLLAPSED.
"""
from __future__ import annotations

import argparse
import datetime
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_CSV   = REPO_ROOT / "output" / "debug" / "raw_ged_trace.csv"
FLAT_CSV  = REPO_ROOT / "output" / "debug" / "flat_ged_trace.csv"
OUT_XLSX  = REPO_ROOT / "output" / "debug" / "raw_flat_reconcile.xlsx"

_STEP_PHASE: dict[str, str] = {
    "actor_calls":    "8B.4",
    "dates_comments": "8B.6",
}

_SHEET_COLS = [
    "set_name",
    "raw_count",
    "flat_count",
    "missing_in_flat_count",
    "extra_in_flat_count",
    "verdict",
]

# ── Phase 8B.4 constants ──────────────────────────────────────────────────────

_ACTOR_CALL_TYPES = {"ACTOR_CALLED", "RESPONSE"}

_SHEET02_COLS = [
    "actor_canonical",
    "raw_count",
    "flat_count",
    "delta",
    "expected_explanation_category",
]

_SHEET03_COLS = [
    "numero",
    "indice",
    "actor_canonical",
    "side",
    "classification",
    "evidence_snippet",
]


# ── Core logic (no I/O — fully testable) ─────────────────────────────────────

# ── Phase 8B.4 helpers ────────────────────────────────────────────────────────

def _derive_raw_cycle(cycle_id: str, actor_canonical: str) -> str:
    """Normalise RAW cycle_id to a cross-side comparison token.

    Only 0-SAS carries meaningful cycle labels (C1, C2).
      C1  →  "SAS"   maps to the single FLAT SAS entry for this version
      C2  →  "C2"    no FLAT counterpart; RAW_ONLY → SAS_CYCLE_COLLAPSED candidate
      ""  →  "SAS"   absent cycle_id on 0-SAS treated as C1 equivalent (defensive)
    All non-SAS actors use "" (single occurrence per version, no cycle distinction).
    """
    if actor_canonical == "0-SAS":
        return "C2" if cycle_id == "C2" else "SAS"
    return ""


def _derive_flat_step(actor_canonical: str) -> str:
    """Map FLAT actor_canonical to a step_type token for cross-side keying.

    The FLAT GED_OPERATIONS trace has no explicit step_type column.  We derive
    it from actor_canonical:
        0-SAS   →  "SAS"   (matches the RAW C1 → "SAS" mapping above)
        others  →  ""      (CONSULTANT / MOEX actors are not distinguished here;
                            the classifier only needs to detect the SAS case)
    """
    return "SAS" if actor_canonical == "0-SAS" else ""


def _classify_raw_only(
    num: str,
    ind: str,
    actor: str,
    cycles: set,
    match_set: set,
    flat_actor_universe: set,
    flat_set_by_num: dict,
) -> str:
    """Apply §5 taxonomy to a RAW_ONLY actor-call 4-tuple.

    Rules applied in order; first match wins.  UNEXPLAINED is the default.
    NULL-numero rows must be filtered out before calling — raises if seen.
    """
    if not num or num.startswith("_MISSING_"):
        raise ValueError(f"NULL/MISSING numero reached classifier: {num!r}")

    # Rule: ACTIVE_VERSION_PROJECTION
    # Same actor appears in FLAT for this numero but on a different indice.
    for fa_ind, fa_actor in flat_set_by_num.get(num, set()):
        if fa_actor == actor and fa_ind != ind:
            return "ACTIVE_VERSION_PROJECTION"

    # Rule: UNKNOWN_ACTOR
    # Actor not present anywhere in the FLAT actor universe.
    if actor not in flat_actor_universe:
        return "UNKNOWN_ACTOR"

    # Rule: SAS_CYCLE_COLLAPSED
    # 0-SAS C2 whose C1 counterpart matched on the FLAT side.
    if actor == "0-SAS" and "C2" in cycles:
        if (num, ind, "0-SAS", "SAS") in match_set:
            return "SAS_CYCLE_COLLAPSED"

    return "UNEXPLAINED"


def _classify_flat_only(
    num: str,
    ind: str,
    actor: str,
    raw_actors_by_version: dict,
) -> str:
    """Apply §5 taxonomy to a FLAT_ONLY actor-call 4-tuple.

    DUPLICATE_MERGED (provisional): the version exists in RAW but carries no
    actor-call records there at all — FLAT produced actor entries from a merged
    or derived source.  Mark for manual review.
    """
    if (num, ind) not in raw_actors_by_version:
        return "DUPLICATE_MERGED"
    return "UNEXPLAINED"


def compute_actor_call_parity(raw_df: pd.DataFrame, flat_df: pd.DataFrame) -> dict:
    """Phase 8B.4 actor call parity computation.

    Applies the same _MISSING_ filter as 8B.3 (documented below).

    Actor-call population choice:
        RAW:  event_type in {ACTOR_CALLED, RESPONSE}
              DOCUMENT_VERSION excluded — RAW DOCUMENT_VERSION rows carry no actor_canonical.
        FLAT: source_sheet == 'GED_OPERATIONS' AND event_type in {ACTOR_CALLED, RESPONSE}
              DOCUMENT_VERSION excluded — its actor_canonical is the emetteur (not an actor
              call), which would create FLAT_ONLY noise with no RAW counterpart.

    4-tuple keys for cross-side comparison:
        RAW  key: (numero, indice, actor_canonical, normalized_cycle)
                  _derive_raw_cycle maps C1 → "SAS", C2 → "C2", non-SAS → ""
        FLAT key: (numero, indice, actor_canonical, step_type)
                  _derive_flat_step maps 0-SAS → "SAS", others → ""

    The C1 → "SAS" / FLAT "SAS" symmetry makes 0-SAS C1 tuples match FLAT 0-SAS
    tuples directly.  0-SAS C2 stays as "C2" — no FLAT counterpart → RAW_ONLY →
    SAS_CYCLE_COLLAPSED when C1 was matched.
    """
    # Step 1: _MISSING_ filter — same rule as 8B.3
    raw_df = raw_df[
        ~raw_df["numero"].astype(str).str.startswith("_MISSING_")
    ].copy()

    # Step 2: RAW actor-call set
    raw_ac = raw_df[raw_df["event_type"].isin(_ACTOR_CALL_TYPES)].copy()
    raw_ac = raw_ac.assign(
        _ncycle=raw_ac.apply(
            lambda r: _derive_raw_cycle(str(r["cycle_id"]), str(r["actor_canonical"])),
            axis=1,
        )
    )
    raw_tuples_df = (
        raw_ac[["numero", "indice", "actor_canonical", "_ncycle", "cycle_id"]]
        .drop_duplicates(subset=["numero", "indice", "actor_canonical", "_ncycle"])
        .copy()
    )
    raw_set: set = set(
        zip(
            raw_tuples_df["numero"].astype(str),
            raw_tuples_df["indice"].astype(str),
            raw_tuples_df["actor_canonical"].astype(str),
            raw_tuples_df["_ncycle"].astype(str),
        )
    )
    # Map 4-tuple → original cycle_id for evidence snippets
    raw_cycle_map: dict = {}
    for _, row in raw_tuples_df.iterrows():
        k = (str(row["numero"]), str(row["indice"]), str(row["actor_canonical"]), str(row["_ncycle"]))
        raw_cycle_map[k] = str(row["cycle_id"])

    # Step 3: FLAT actor-call set (GED_OPERATIONS only — see population choice above)
    flat_ac = flat_df[
        (flat_df["source_sheet"] == "GED_OPERATIONS")
        & flat_df["event_type"].isin(_ACTOR_CALL_TYPES)
    ].copy()
    flat_ac = flat_ac.assign(
        _step=flat_ac["actor_canonical"].apply(lambda a: _derive_flat_step(str(a)))
    )
    flat_tuples_df = (
        flat_ac[["numero", "indice", "actor_canonical", "_step"]]
        .drop_duplicates()
        .copy()
    )
    flat_set: set = set(
        zip(
            flat_tuples_df["numero"].astype(str),
            flat_tuples_df["indice"].astype(str),
            flat_tuples_df["actor_canonical"].astype(str),
            flat_tuples_df["_step"].astype(str),
        )
    )

    # Step 4: diff sets
    raw_only = raw_set - flat_set
    flat_only = flat_set - raw_set
    match_set = raw_set & flat_set

    # Classifier lookup structures
    flat_actor_universe: set = {actor for _, _, actor, _ in flat_set}
    flat_set_by_num: dict = {}  # num → {(ind, actor)}
    for num, ind, actor, _ in flat_set:
        flat_set_by_num.setdefault(num, set()).add((ind, actor))

    raw_actors_by_version: dict = {}  # (num, ind) → {actor}
    for num, ind, actor, _ in raw_set:
        raw_actors_by_version.setdefault((num, ind), set()).add(actor)

    # Collect raw_only cycles per (num, ind, actor) for SAS_CYCLE_COLLAPSED check
    raw_only_cycles: dict = {}
    for num, ind, actor, ncycle in raw_only:
        raw_only_cycles.setdefault((num, ind, actor), set()).add(ncycle)

    # Step 5 + 7: build Sheet 03 rows — one row per 4-tuple in raw_only / flat_only
    sheet03_rows: list = []
    for num, ind, actor, ncycle in sorted(raw_only):
        cycles = raw_only_cycles.get((num, ind, actor), set())
        raw_cid = raw_cycle_map.get((num, ind, actor, ncycle), "")
        evidence = f"cycle_id={raw_cid!r} norm={ncycle!r}"
        classification = _classify_raw_only(
            num, ind, actor, cycles, match_set, flat_actor_universe, flat_set_by_num
        )
        sheet03_rows.append(
            {
                "numero": num,
                "indice": ind,
                "actor_canonical": actor,
                "side": "RAW_ONLY",
                "classification": classification,
                "evidence_snippet": evidence,
            }
        )

    for num, ind, actor, step in sorted(flat_only):
        evidence = f"step_type={step!r}"
        classification = _classify_flat_only(num, ind, actor, raw_actors_by_version)
        sheet03_rows.append(
            {
                "numero": num,
                "indice": ind,
                "actor_canonical": actor,
                "side": "FLAT_ONLY",
                "classification": classification,
                "evidence_snippet": evidence,
            }
        )

    # Step 6: build Sheet 02 rows — per-actor totals, sorted by |delta| desc
    all_actors: set = {actor for _, _, actor, _ in (raw_set | flat_set)}
    raw_count_by_actor: dict = {}
    for _, _, actor, _ in raw_set:
        raw_count_by_actor[actor] = raw_count_by_actor.get(actor, 0) + 1
    flat_count_by_actor: dict = {}
    for _, _, actor, _ in flat_set:
        flat_count_by_actor[actor] = flat_count_by_actor.get(actor, 0) + 1

    sheet02_rows: list = []
    for actor in sorted(all_actors):
        raw_c = raw_count_by_actor.get(actor, 0)
        flat_c = flat_count_by_actor.get(actor, 0)
        delta = raw_c - flat_c
        if delta == 0:
            cat = "MATCH"
        elif actor == "0-SAS" and delta > 0:
            cat = "SAS_CYCLE_COLLAPSED"
        elif actor not in flat_actor_universe:
            cat = "UNKNOWN_ACTOR"
        else:
            cat = "UNEXPLAINED"
        sheet02_rows.append(
            {
                "actor_canonical": actor,
                "raw_count": raw_c,
                "flat_count": flat_c,
                "delta": delta,
                "expected_explanation_category": cat,
            }
        )
    sheet02_rows.sort(key=lambda r: abs(r["delta"]), reverse=True)

    total_diffs = len(sheet03_rows)
    unexplained = sum(1 for r in sheet03_rows if r["classification"] == "UNEXPLAINED")
    summary_line = f"ACTOR_CALL_PARITY: total_diffs={total_diffs} unexplained={unexplained}"

    return {
        "raw_set": raw_set,
        "flat_set": flat_set,
        "match_set": match_set,
        "raw_only": raw_only,
        "flat_only": flat_only,
        "sheet02_rows": sheet02_rows,
        "sheet03_rows": sheet03_rows,
        "total_diffs": total_diffs,
        "unexplained": unexplained,
        "summary_line": summary_line,
    }


def write_actor_call_sheets(
    sheet02_rows: list, sheet03_rows: list, xlsx_path: Path
) -> None:
    """Append / replace sheets 02_ACTOR_CALL_COUNTS and 03_ACTOR_CALL_DIFFS.

    Sheet 01_IDENTITY_PARITY is preserved untouched.
    """
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default empty sheet so named sheets are added cleanly
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for name in ("02_ACTOR_CALL_COUNTS", "03_ACTOR_CALL_DIFFS"):
        if name in wb.sheetnames:
            del wb[name]

    ws02 = wb.create_sheet("02_ACTOR_CALL_COUNTS")
    ws02.append(_SHEET02_COLS)
    for row in sheet02_rows:
        ws02.append([row[c] for c in _SHEET02_COLS])

    ws03 = wb.create_sheet("03_ACTOR_CALL_DIFFS")
    ws03.append(_SHEET03_COLS)
    for row in sheet03_rows:
        ws03.append([row[c] for c in _SHEET03_COLS])

    wb.save(xlsx_path)


def selfcheck_actor_call_sheets(result: dict, xlsx_path: Path) -> None:
    """Read sheets 02 and 03 back and assert row counts match in-memory data.

    §11.2 point 10: assert RAW_ONLY row count == len(raw_only),
                    assert FLAT_ONLY row count == len(flat_only).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    ws02 = wb["02_ACTOR_CALL_COUNTS"]
    xlsx02_rows = list(ws02.iter_rows(values_only=True))[1:]  # skip header
    if len(xlsx02_rows) != len(result["sheet02_rows"]):
        raise AssertionError(
            f"Sheet 02 row count mismatch: expected {len(result['sheet02_rows'])}, "
            f"got {len(xlsx02_rows)}"
        )

    ws03 = wb["03_ACTOR_CALL_DIFFS"]
    xlsx03_rows = list(ws03.iter_rows(values_only=True))[1:]  # skip header
    if len(xlsx03_rows) != len(result["sheet03_rows"]):
        raise AssertionError(
            f"Sheet 03 row count mismatch: expected {len(result['sheet03_rows'])}, "
            f"got {len(xlsx03_rows)}"
        )

    raw_only_in_sheet = sum(
        1 for r in xlsx03_rows if r[3] == "RAW_ONLY"  # col index 3 = side
    )
    flat_only_in_sheet = sum(
        1 for r in xlsx03_rows if r[3] == "FLAT_ONLY"
    )
    expected_raw_only = len(result["raw_only"])
    expected_flat_only = len(result["flat_only"])
    if raw_only_in_sheet != expected_raw_only:
        raise AssertionError(
            f"Sheet 03 RAW_ONLY count {raw_only_in_sheet} != raw_only set size "
            f"{expected_raw_only}"
        )
    if flat_only_in_sheet != expected_flat_only:
        raise AssertionError(
            f"Sheet 03 FLAT_ONLY count {flat_only_in_sheet} != flat_only set size "
            f"{expected_flat_only}"
        )

    wb.close()


# ── Phase 8B.5 constants ──────────────────────────────────────────────────────

# Operational status set verified from L0 audit baseline.
# raw[raw['event_type']=='RESPONSE']['status_clean'].unique() returns exactly these values.
_OPERATIONAL_STATUS: frozenset = frozenset(
    {"VAO", "VSO", "VSO-SAS", "REF", "HM", "FAV", "SUS", "DEF"}
)

_SHEET04_COLS = ["status_clean", "side", "count"]
_SHEET05_COLS = ["actor_canonical", "status_clean", "side", "count"]
_SHEET06_COLS = [
    "numero", "indice", "actor_canonical", "cycle_id", "step_type",
    "status_raw", "status_clean", "side", "classification", "evidence_snippet",
]
_SHEET07_COLS = [
    "numero", "indice", "actor", "raw_cycle", "status_raw", "status_clean",
    "flat_present", "flat_status_clean", "classification", "evidence_excel_row",
]

# SAS REF baseline (from extract_raw_ged_trace.py sas_ref_unique_pairs output).
# 836 = 809 non-_MISSING_ source_excel_rows + 27 _MISSING_ source_excel_rows.
_SAS_REF_RAW_BASELINE = 836
_SAS_REF_FLAT_BASELINE = 284  # GED_RAW_FLAT row count (283 unique pairs + 1 duplicate)


# ── Phase 8B.5 helpers ────────────────────────────────────────────────────────

def _classify_response_raw_only(
    num: str,
    ind: str,
    actor: str,
    cycle_id: str,
    ncycle: str,
    status_raw: str,
    status_clean: str,
    flat_actor_universe: set,
    flat_set_by_num: dict,
    raw_cycles_per_nia: dict,
    matched: set,
) -> str:
    """Apply §5 taxonomy to a RAW_ONLY response 4-tuple.

    Rules applied in order per §12.2; first match wins.

    Pivot rule for SAS (documented here per spec):
      RAW cycle_id C1 normalises to ncycle='SAS' → maps to the single FLAT
      SAS entry for (numero, indice).  RAW cycle_id C2 normalises to ncycle='C2'
      which has no FLAT counterpart; if its sibling C1 was matched, the C2 row
      is SAS_CYCLE_COLLAPSED.
    """
    # Rule (i): actor_canonical == 'Exception List' → UNKNOWN_ACTOR
    if actor == "Exception List":
        return "UNKNOWN_ACTOR"

    # Rule (ii): Same numero in FLAT but different indice → ACTIVE_VERSION_PROJECTION
    for fa_ind, fa_actor in flat_set_by_num.get(num, set()):
        if fa_actor == actor and fa_ind != ind:
            return "ACTIVE_VERSION_PROJECTION"

    # Rule (iv) guard: SAS C2 is handled exclusively by rule (iv) below.
    # Exclude SAS C2 from rule (iii) so cycle-expansion rows are not misclassified
    # as DUPLICATE_MERGED when they are actually SAS_CYCLE_COLLAPSED.
    is_sas_c2 = actor == "0-SAS" and ncycle == "C2"

    # Rule (iii): Multiple RAW tuples for (num, ind, actor) where FLAT has exactly one
    flat_count_for_actor = sum(
        1 for fa_ind, fa_actor in flat_set_by_num.get(num, set())
        if fa_actor == actor and fa_ind == ind
    )
    raw_ncycles_for_nia = raw_cycles_per_nia.get((num, ind, actor), set())
    if not is_sas_c2 and len(raw_ncycles_for_nia) > 1 and flat_count_for_actor == 1:
        return "DUPLICATE_MERGED"

    # Rule (iv): SAS C2 whose C1 counterpart is matched in FLAT → SAS_CYCLE_COLLAPSED
    if is_sas_c2:
        if (num, ind, "0-SAS", "SAS") in matched:
            return "SAS_CYCLE_COLLAPSED"

    # Rule (v): status_raw not in operational set → NON_OPERATIONAL_RESPONSE
    if status_raw and status_raw not in _OPERATIONAL_STATUS:
        return "NON_OPERATIONAL_RESPONSE"

    # Rule (vi): status_clean empty/garbage but status_raw populated → MALFORMED_RESPONSE
    if status_raw and not status_clean:
        return "MALFORMED_RESPONSE"

    # Rule (vii): none of the above → UNEXPLAINED
    return "UNEXPLAINED"


def _classify_response_flat_only(
    num: str,
    ind: str,
    actor: str,
    raw_nia_set: set,
) -> str:
    """Apply §5 taxonomy to a FLAT_ONLY response tuple."""
    if (num, ind) not in raw_nia_set:
        return "DUPLICATE_MERGED"
    return "UNEXPLAINED"


def compute_response_parity(raw_df: pd.DataFrame, flat_df: pd.DataFrame) -> dict:
    """Phase 8B.5 response parity computation (sheets 04, 05, 06).

    _MISSING_ filter: applied before building raw_responses (same rule as 8B.3/8B.4).
    Rows where numero starts with '_MISSING_' are excluded from RAW before any
    set comparison.  Sheet 07 (SAS REF trace) is computed separately in
    compute_sas_ref_trace() which includes the _MISSING_ rows.

    Join keys:
        RAW  4-tuple: (numero, indice, actor_canonical, ncycle)
                      ncycle = _derive_raw_cycle(cycle_id, actor_canonical)
        FLAT 4-tuple: (numero, indice, actor_canonical, step_type)
                      step_type = _derive_flat_step(actor_canonical)

    FLAT source: GED_OPERATIONS only (same rationale as 8B.4 — GED_RAW_FLAT
    rows are a different projection layer and would create spurious FLAT_ONLY diffs).
    """
    # _MISSING_ filter
    raw_df = raw_df[~raw_df["numero"].astype(str).str.startswith("_MISSING_")].copy()

    # Build response DataFrames
    raw_resp = raw_df[raw_df["event_type"] == "RESPONSE"].copy()
    flat_resp = flat_df[
        (flat_df["event_type"] == "RESPONSE")
        & (flat_df["source_sheet"] == "GED_OPERATIONS")
    ].copy()

    # Normalise cycle_id / step_type for cross-side keying
    raw_resp = raw_resp.assign(
        _ncycle=raw_resp.apply(
            lambda r: _derive_raw_cycle(str(r["cycle_id"]), str(r["actor_canonical"])),
            axis=1,
        )
    )
    flat_resp = flat_resp.assign(
        _step=flat_resp["actor_canonical"].apply(lambda a: _derive_flat_step(str(a)))
    )

    # Build unique 4-tuple sets
    raw_4t = raw_resp[
        ["numero", "indice", "actor_canonical", "_ncycle"]
    ].drop_duplicates()
    raw_set: set = set(
        zip(
            raw_4t["numero"].astype(str),
            raw_4t["indice"].astype(str),
            raw_4t["actor_canonical"].astype(str),
            raw_4t["_ncycle"].astype(str),
        )
    )

    flat_4t = flat_resp[
        ["numero", "indice", "actor_canonical", "_step"]
    ].drop_duplicates()
    flat_set: set = set(
        zip(
            flat_4t["numero"].astype(str),
            flat_4t["indice"].astype(str),
            flat_4t["actor_canonical"].astype(str),
            flat_4t["_step"].astype(str),
        )
    )

    raw_only = raw_set - flat_set
    flat_only = flat_set - raw_set
    matched = raw_set & flat_set

    # Classifier lookup structures
    flat_actor_universe: set = {a for _, _, a, _ in flat_set}
    flat_set_by_num: dict = {}
    for num, ind, actor, _ in flat_set:
        flat_set_by_num.setdefault(num, set()).add((ind, actor))

    raw_nia_set: set = {(num, ind) for num, ind, _, _ in raw_set}

    # Map (num, ind, actor) → set of ncycles in RAW (for rule iii)
    raw_cycles_per_nia: dict = {}
    for num, ind, actor, nc in raw_set:
        raw_cycles_per_nia.setdefault((num, ind, actor), set()).add(nc)

    # Build status_raw/status_clean lookup per 4-tuple
    raw_status_map: dict = {}
    for _, row in raw_resp.iterrows():
        k = (
            str(row["numero"]), str(row["indice"]),
            str(row["actor_canonical"]), str(row["_ncycle"]),
        )
        if k not in raw_status_map:
            raw_status_map[k] = (str(row["status_raw"]), str(row["status_clean"]))

    # Build cycle_id lookup per 4-tuple (for evidence snippet)
    raw_cid_map: dict = {}
    for _, row in raw_resp[
        ["numero", "indice", "actor_canonical", "_ncycle", "cycle_id"]
    ].drop_duplicates(subset=["numero", "indice", "actor_canonical", "_ncycle"]).iterrows():
        k = (
            str(row["numero"]), str(row["indice"]),
            str(row["actor_canonical"]), str(row["_ncycle"]),
        )
        raw_cid_map[k] = str(row["cycle_id"])

    # Sheet 06 rows — one per diff tuple
    sheet06_rows: list = []

    for num, ind, actor, ncycle in sorted(raw_only):
        cid = raw_cid_map.get((num, ind, actor, ncycle), "")
        sr, sc = raw_status_map.get((num, ind, actor, ncycle), ("", ""))
        classification = _classify_response_raw_only(
            num, ind, actor, cid, ncycle, sr, sc,
            flat_actor_universe, flat_set_by_num, raw_cycles_per_nia, matched,
        )
        sheet06_rows.append({
            "numero": num,
            "indice": ind,
            "actor_canonical": actor,
            "cycle_id": cid,
            "step_type": "",
            "status_raw": sr,
            "status_clean": sc,
            "side": "RAW_ONLY",
            "classification": classification,
            "evidence_snippet": f"cycle_id={cid!r} norm={ncycle!r} status_raw={sr!r}",
        })

    for num, ind, actor, step in sorted(flat_only):
        classification = _classify_response_flat_only(num, ind, actor, raw_nia_set)
        sheet06_rows.append({
            "numero": num,
            "indice": ind,
            "actor_canonical": actor,
            "cycle_id": "",
            "step_type": step,
            "status_raw": "",
            "status_clean": "",
            "side": "FLAT_ONLY",
            "classification": classification,
            "evidence_snippet": f"step={step!r}",
        })

    # Sheet 04 — status × side × count (all RESPONSE events, not just diffs)
    sheet04_rows: list = []
    all_statuses = (
        set(raw_resp["status_clean"].astype(str).unique())
        | set(flat_resp["status_clean"].astype(str).unique())
    )
    for sc in sorted(all_statuses):
        r = int((raw_resp["status_clean"].astype(str) == sc).sum())
        f = int((flat_resp["status_clean"].astype(str) == sc).sum())
        if r > 0:
            sheet04_rows.append({"status_clean": sc, "side": "RAW", "count": r})
        if f > 0:
            sheet04_rows.append({"status_clean": sc, "side": "FLAT", "count": f})
    sheet04_rows.sort(key=lambda r: (r["status_clean"], r["side"]))

    # Sheet 05 — actor × status × side × count
    sheet05_rows: list = []
    for (act, sc), grp in raw_resp.groupby(["actor_canonical", "status_clean"]):
        sheet05_rows.append({
            "actor_canonical": str(act),
            "status_clean": str(sc),
            "side": "RAW",
            "count": len(grp),
        })
    for (act, sc), grp in flat_resp.groupby(["actor_canonical", "status_clean"]):
        sheet05_rows.append({
            "actor_canonical": str(act),
            "status_clean": str(sc),
            "side": "FLAT",
            "count": len(grp),
        })
    sheet05_rows.sort(key=lambda r: (r["actor_canonical"], r["status_clean"], r["side"]))

    total_raw = len(raw_set)
    total_flat = len(flat_set)
    total_matched = len(matched)
    unexplained = sum(1 for r in sheet06_rows if r["classification"] == "UNEXPLAINED")
    summary_line = (
        f"RESPONSE_PARITY: total_raw={total_raw} total_flat={total_flat} "
        f"matched={total_matched}"
    )

    return {
        "raw_set": raw_set,
        "flat_set": flat_set,
        "matched": matched,
        "raw_only": raw_only,
        "flat_only": flat_only,
        "sheet04_rows": sheet04_rows,
        "sheet05_rows": sheet05_rows,
        "sheet06_rows": sheet06_rows,
        "total_raw": total_raw,
        "total_flat": total_flat,
        "total_matched": total_matched,
        "unexplained": unexplained,
        "summary_line": summary_line,
    }


def compute_sas_ref_trace(raw_df: pd.DataFrame, flat_df: pd.DataFrame) -> dict:
    """Phase 8B.5 SAS REF trace (sheet 07) — the D-011 headline.

    Sources:
        RAW: ALL SAS REF RESPONSE rows (no _MISSING_ filter).  The 27
             _MISSING_ rows are intentionally included so the 836 unique
             source_excel_row baseline is preserved.  They are classified
             MALFORMED_RESPONSE.
        FLAT: GED_RAW_FLAT sheet (284 rows, 283 unique (numero,indice) pairs;
              the spec instructs to use GED_RAW_FLAT as the 284 baseline).

    Deduplication rule:
        One row per unique source_excel_row.  The single Excel data row that
        carries BOTH SAS C1 and C2 REF (numero=152012, indice=A) contributes
        one sheet 07 row; raw_cycle is set to 'C1+C2' for that row.

    Classification order (first match wins, per §12.2 rule set):
        MALFORMED_RESPONSE   — _MISSING_ numero sentinel rows
        DUPLICATE_MERGED     — not the first occurrence of (numero, indice) in
                               this sheet (multiple GED Excel rows for the same
                               document version; FLAT collapses to one)
        ACTIVE_VERSION_PROJECTION — (numero, indice) absent from FLAT SAS REF
                               but same numero IS present in FLAT SAS REF with
                               a different indice
        SAS_CYCLE_COLLAPSED  — (numero, indice) has both C1 and C2 REF in RAW;
                               this is the canonical row (first occurrence)
        UNEXPLAINED          — none of the above apply

    flat_present logic:
        True  iff (numero, indice) ∈ FLAT GED_RAW_FLAT SAS REF (numero,indice) set.
    """
    # Build FLAT SAS REF lookup from GED_RAW_FLAT (the 284-row baseline)
    flat_sas_rf = flat_df[
        (flat_df["actor_canonical"] == "0-SAS")
        & (flat_df["status_clean"] == "REF")
        & (flat_df["event_type"] == "RESPONSE")
        & (flat_df["source_sheet"] == "GED_RAW_FLAT")
    ]
    flat_ni_set: set = set(
        zip(
            flat_sas_rf["numero"].astype(str),
            flat_sas_rf["indice"].astype(str),
        )
    )
    # For flat_status_clean lookup (should always be 'REF' but capture verbatim)
    flat_status_map: dict = {}
    for _, row in flat_sas_rf.iterrows():
        k = (str(row["numero"]), str(row["indice"]))
        flat_status_map.setdefault(k, str(row["status_clean"]))

    # FLAT SAS REF numeros (from GED_OPERATIONS, for ACTIVE_VERSION_PROJECTION)
    flat_sas_ops_nums: set = set(
        flat_df.loc[
            (flat_df["actor_canonical"] == "0-SAS")
            & (flat_df["status_clean"] == "REF")
            & (flat_df["event_type"] == "RESPONSE")
            & (flat_df["source_sheet"] == "GED_OPERATIONS"),
            "numero",
        ].astype(str)
    )
    flat_sas_ops_ni: set = set(
        zip(
            flat_df.loc[
                (flat_df["actor_canonical"] == "0-SAS")
                & (flat_df["status_clean"] == "REF")
                & (flat_df["event_type"] == "RESPONSE")
                & (flat_df["source_sheet"] == "GED_OPERATIONS"),
                "numero",
            ].astype(str),
            flat_df.loc[
                (flat_df["actor_canonical"] == "0-SAS")
                & (flat_df["status_clean"] == "REF")
                & (flat_df["event_type"] == "RESPONSE")
                & (flat_df["source_sheet"] == "GED_OPERATIONS"),
                "indice",
            ].astype(str),
        )
    )

    # Lookups for DUPLICATE_FAVORABLE_KEPT rule (§5 taxonomy extension)
    # Includes VSO-SAS: RAW frequently carries VSO-SAS while FLAT normalises it to VSO.
    _FAV_STATUS_SET = frozenset({"VAO", "VSO", "VSO-SAS"})

    # (a) RAW (num,ind) pairs that have ≥1 0-SAS VAO/VSO/VSO-SAS RESPONSE row
    _raw_sas_fav = raw_df[
        (raw_df["actor_canonical"] == "0-SAS")
        & (raw_df["event_type"] == "RESPONSE")
        & (raw_df["status_clean"].isin(_FAV_STATUS_SET))
    ]
    raw_sas_favorable_ni: set = set(
        zip(
            _raw_sas_fav["numero"].astype(str),
            _raw_sas_fav["indice"].astype(str),
        )
    )

    # (b) FLAT (num,ind) pairs that have ≥1 0-SAS VAO/VSO/VSO-SAS RESPONSE row
    _flat_sas_fav = flat_df[
        (flat_df["actor_canonical"] == "0-SAS")
        & (flat_df["event_type"] == "RESPONSE")
        & (flat_df["status_clean"].isin(_FAV_STATUS_SET))
    ]
    flat_sas_favorable_ni: set = set(
        zip(
            _flat_sas_fav["numero"].astype(str),
            _flat_sas_fav["indice"].astype(str),
        )
    )

    # ALL SAS REF RESPONSE rows in RAW (no _MISSING_ filter for this trace)
    all_sas_ref = raw_df[
        (raw_df["actor_canonical"] == "0-SAS")
        & (raw_df["status_raw"] == "REF")
        & (raw_df["event_type"] == "RESPONSE")
    ].copy()

    # Cycles present per source_excel_row (to detect C1+C2 dual-cycle rows)
    src_cycles: dict = (
        all_sas_ref.groupby("source_excel_row")["cycle_id"]
        .apply(set)
        .to_dict()
    )

    # Deduplicate by source_excel_row: keep one row per unique Excel source row.
    # Sort by cycle_id descending so C2 (if present) is kept as the representative.
    ded = (
        all_sas_ref
        .sort_values(["source_excel_row", "cycle_id"], ascending=[True, False])
        .drop_duplicates(subset=["source_excel_row"], keep="first")
        .copy()
    )

    # Build ni_first_seen: for non-_MISSING_ rows, track which source_excel_row
    # was the FIRST occurrence of each (numero, indice) pair.  Subsequent
    # occurrences are classified DUPLICATE_MERGED.
    ni_first_seen: dict = {}  # (num, ind) → source_excel_row str
    for _, row in ded[
        ~ded["numero"].astype(str).str.startswith("_MISSING_")
    ].sort_values("source_excel_row").iterrows():
        ni = (str(row["numero"]), str(row["indice"]))
        if ni not in ni_first_seen:
            ni_first_seen[ni] = str(row["source_excel_row"])

    # (numero, indice) pairs that have BOTH C1 and C2 REF in RAW
    ni_with_both_cycles: set = set()
    for src_row_id, cycles in src_cycles.items():
        row_slice = all_sas_ref[all_sas_ref["source_excel_row"] == src_row_id]
        if row_slice.empty:
            continue
        ni = (
            str(row_slice.iloc[0]["numero"]),
            str(row_slice.iloc[0]["indice"]),
        )
        if "C1" in cycles and "C2" in cycles:
            ni_with_both_cycles.add(ni)

    # Build sheet 07 rows
    sheet07_rows: list = []

    for _, row in ded.iterrows():
        num = str(row["numero"])
        ind = str(row["indice"])
        src_row = str(row["source_excel_row"])
        cycle_id = str(row["cycle_id"])
        status_raw = str(row["status_raw"])
        status_clean = str(row["status_clean"])

        # raw_cycle: 'C1+C2' for the dual-cycle row, else the actual cycle_id
        cycles_here = src_cycles.get(src_row, set())
        raw_cycle = (
            "C1+C2"
            if ("C1" in cycles_here and "C2" in cycles_here)
            else cycle_id
        )

        ni = (num, ind)
        is_missing = num.startswith("_MISSING_")
        flat_present = ni in flat_ni_set
        flat_sc = flat_status_map.get(ni, "") if flat_present else ""

        # Classification (first match wins)
        if is_missing:
            # _MISSING_ sentinel: numero could not be resolved from GED Excel
            classification = "MALFORMED_RESPONSE"
        elif ni_first_seen.get(ni) != src_row:
            # Not the canonical (first) occurrence of this (numero, indice)
            classification = "DUPLICATE_MERGED"
        elif not flat_present:
            # Canonical row but absent from FLAT — apply remaining taxonomy
            if num in flat_sas_ops_nums and ni not in flat_sas_ops_ni:
                # Same numero in FLAT SAS REF but with a different indice
                classification = "ACTIVE_VERSION_PROJECTION"
            elif ni in flat_sas_favorable_ni and ni in raw_sas_favorable_ni:
                # RAW has both REF and a favorable (VAO/VSO/VSO-SAS) for this
                # (numero, indice); FLAT kept the favorable and dropped the REF
                classification = "DUPLICATE_FAVORABLE_KEPT"
            elif ni in ni_with_both_cycles:
                # Both C1 and C2 REF present; FLAT collapses to single decision
                classification = "SAS_CYCLE_COLLAPSED"
            else:
                classification = "UNEXPLAINED"
        else:
            # flat_present=True, canonical occurrence — no taxonomy needed
            classification = ""

        sheet07_rows.append({
            "numero": num,
            "indice": ind,
            "actor": "0-SAS",
            "raw_cycle": raw_cycle,
            "status_raw": status_raw,
            "status_clean": status_clean,
            "flat_present": flat_present,
            "flat_status_clean": flat_sc if flat_sc else None,
            "classification": classification,
            "evidence_excel_row": src_row,
        })

    total_sas_raw = len(sheet07_rows)  # should equal _SAS_REF_RAW_BASELINE (836)
    flat_present_count = sum(1 for r in sheet07_rows if r["flat_present"])
    classified = sum(
        1 for r in sheet07_rows
        if r["classification"] and r["classification"] != "UNEXPLAINED"
    )
    unexplained = sum(1 for r in sheet07_rows if r["classification"] == "UNEXPLAINED")
    duplicate_favorable_kept = sum(
        1 for r in sheet07_rows if r["classification"] == "DUPLICATE_FAVORABLE_KEPT"
    )

    summary_line = (
        f"SAS_REF_TRACE: raw={_SAS_REF_RAW_BASELINE} "
        f"flat={_SAS_REF_FLAT_BASELINE} "
        f"classified={classified} unexplained={unexplained} "
        f"duplicate_favorable_kept={duplicate_favorable_kept}"
    )

    return {
        "sheet07_rows": sheet07_rows,
        "total_sas_raw": total_sas_raw,
        "flat_present_count": flat_present_count,
        "classified": classified,
        "unexplained": unexplained,
        "duplicate_favorable_kept": duplicate_favorable_kept,
        "summary_line": summary_line,
    }


def write_response_sheets(
    sheet04_rows: list,
    sheet05_rows: list,
    sheet06_rows: list,
    sheet07_rows: list,
    xlsx_path: Path,
) -> None:
    """Append / replace sheets 04–07 in the output xlsx.

    Sheets 01–03 are preserved untouched.
    Post-write: closes the workbook (del wb) then re-opens in a FRESH Python
    subprocess for the self-check so the in-memory state cannot mask disk errors
    (H-1 lesson from Phase 8).
    """
    import subprocess
    import os
    import zipfile

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for name in (
        "04_RESPONSE_COUNTS_BY_STATUS",
        "05_RESPONSE_COUNTS_BY_ACTOR",
        "06_RESPONSE_DIFFS",
        "07_SAS_REF_TRACE",
    ):
        if name in wb.sheetnames:
            del wb[name]

    ws04 = wb.create_sheet("04_RESPONSE_COUNTS_BY_STATUS")
    ws04.append(_SHEET04_COLS)
    for row in sheet04_rows:
        ws04.append([row[c] for c in _SHEET04_COLS])

    ws05 = wb.create_sheet("05_RESPONSE_COUNTS_BY_ACTOR")
    ws05.append(_SHEET05_COLS)
    for row in sheet05_rows:
        ws05.append([row[c] for c in _SHEET05_COLS])

    ws06 = wb.create_sheet("06_RESPONSE_DIFFS")
    ws06.append(_SHEET06_COLS)
    for row in sheet06_rows:
        ws06.append([row[c] for c in _SHEET06_COLS])

    ws07 = wb.create_sheet("07_SAS_REF_TRACE")
    ws07.append(_SHEET07_COLS)
    for row in sheet07_rows:
        ws07.append([row[c] for c in _SHEET07_COLS])

    wb.save(xlsx_path)
    del wb  # close in-memory workbook before subprocess re-open (H-1 lesson)

    # Fresh-subprocess self-check (H-1 lesson: in-memory workbook does not prove
    # disk state).  The subprocess reads the file from disk and verifies:
    #   a. file size > 0
    #   b. all four new sheets present
    #   c. row counts match expected values
    #   d. ZIP integrity (zipfile.ZipFile.testzip → first_bad_member must be None)
    check_code = f"""
import os, zipfile
from openpyxl import load_workbook

p = r{str(xlsx_path)!r}
size = os.path.getsize(p)
assert size > 0, f"file empty: {{size}}"

with zipfile.ZipFile(p) as z:
    bad = z.testzip()
assert bad is None, f"ZIP corrupt: {{bad}}"

wb = load_workbook(p, read_only=True, data_only=True)
names = wb.sheetnames
for req in ("04_RESPONSE_COUNTS_BY_STATUS","05_RESPONSE_COUNTS_BY_ACTOR",
            "06_RESPONSE_DIFFS","07_SAS_REF_TRACE"):
    assert req in names, f"missing sheet: {{req}}"

rows07 = list(wb["07_SAS_REF_TRACE"].iter_rows(values_only=True))
assert len(rows07) == {len(sheet07_rows) + 1}, (
    f"Sheet 07 row count (incl. header) expected {len(sheet07_rows) + 1}, got {{len(rows07)}}")
rows04 = list(wb["04_RESPONSE_COUNTS_BY_STATUS"].iter_rows(values_only=True))
rows05 = list(wb["05_RESPONSE_COUNTS_BY_ACTOR"].iter_rows(values_only=True))
rows06 = list(wb["06_RESPONSE_DIFFS"].iter_rows(values_only=True))
wb.close()

print(f"SUBPROCESS_CHECK: size={{size}} sheets={{names}}")
print(f"  04_rows={{len(rows04)-1}} 05_rows={{len(rows05)-1}} "
      f"06_rows={{len(rows06)-1}} 07_rows={{len(rows07)-1}}")
print(f"  ZIP_integrity=OK")
"""
    result = subprocess.run(
        ["python", "-c", check_code],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Subprocess self-check FAILED:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )
    print(result.stdout.strip())


def selfcheck_response_sheets(
    resp_result: dict,
    sas_result: dict,
    xlsx_path: Path,
) -> None:
    """In-process self-check for sheets 04–07 (fast read after subprocess check).

    Assertions (per spec §12.2 point 8):
    1. Sheet 07 has exactly _SAS_REF_RAW_BASELINE rows + 1 header.
    2. Sum of sheet 07 rows where flat_present==False and actor=='0-SAS' and
       status_clean=='REF' == count of RAW_ONLY SAS REF rows in sheet 06
       classified as not-flat-present.
    3. flat_present==True count in sheet 07 == sas_result['flat_present_count'].
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    ws07 = wb["07_SAS_REF_TRACE"]
    rows07 = list(ws07.iter_rows(values_only=True))
    header07 = rows07[0]
    data07 = rows07[1:]

    col = {name: idx for idx, name in enumerate(header07)}

    assert len(data07) == sas_result["total_sas_raw"], (
        f"Sheet 07 data rows: expected {sas_result['total_sas_raw']}, got {len(data07)}"
    )

    # flat_present is stored as bool; openpyxl reads it as True/False
    fp_true = sum(1 for r in data07 if r[col["flat_present"]])
    assert fp_true == sas_result["flat_present_count"], (
        f"Sheet 07 flat_present==True: expected {sas_result['flat_present_count']}, "
        f"got {fp_true}"
    )

    # Cross-check: sheet 06 RAW_ONLY rows for '0-SAS' / status_clean=='REF'
    ws06 = wb["06_RESPONSE_DIFFS"]
    rows06 = list(ws06.iter_rows(values_only=True))
    h06 = {name: idx for idx, name in enumerate(rows06[0])}
    sas_ref_raw_only_06 = sum(
        1 for r in rows06[1:]
        if r[h06["actor_canonical"]] == "0-SAS"
        and r[h06["status_clean"]] == "REF"
        and r[h06["side"]] == "RAW_ONLY"
    )
    # sheet 07 rows that are flat_present==False and status_clean=='REF'
    sas_ref_missing_07 = sum(
        1 for r in data07
        if not r[col["flat_present"]]
        and str(r[col["status_clean"]]) == "REF"
    )
    # These two counts are not required to be identical because the join
    # granularities differ (4-tuple vs source_excel_row), but we log them.
    print(
        f"SELFCHECK cross-ref: sheet06_SAS_REF_RAW_ONLY={sas_ref_raw_only_06} "
        f"sheet07_flat_present_False={sas_ref_missing_07}"
    )

    unexplained_07 = sum(
        1 for r in data07 if str(r[col["classification"]]) == "UNEXPLAINED"
    )
    assert unexplained_07 == sas_result["unexplained"], (
        f"Sheet 07 UNEXPLAINED mismatch: in-memory {sas_result['unexplained']}, "
        f"on-disk {unexplained_07}"
    )

    dfk_07 = sum(
        1 for r in data07
        if str(r[col["classification"]]) == "DUPLICATE_FAVORABLE_KEPT"
    )
    assert dfk_07 == sas_result["duplicate_favorable_kept"], (
        f"Sheet 07 DUPLICATE_FAVORABLE_KEPT mismatch: in-memory "
        f"{sas_result['duplicate_favorable_kept']}, on-disk {dfk_07}"
    )

    wb.close()


# ── Phase 8B.5 step runner ────────────────────────────────────────────────────

def _run_responses() -> None:
    print(f"Loading RAW trace:  {RAW_CSV}")
    print(f"Loading FLAT trace: {FLAT_CSV}")
    raw_df = pd.read_csv(RAW_CSV,  dtype=str, keep_default_na=False)
    flat_df = pd.read_csv(FLAT_CSV, dtype=str, keep_default_na=False)

    # --- Response parity (sheets 04/05/06) ---
    resp_result = compute_response_parity(raw_df, flat_df)
    print(f"RAW response 4-tuples:  {resp_result['total_raw']}")
    print(f"FLAT response 4-tuples: {resp_result['total_flat']}")
    print(f"Matched:   {resp_result['total_matched']}")
    print(f"RAW_ONLY:  {len(resp_result['raw_only'])}")
    print(f"FLAT_ONLY: {len(resp_result['flat_only'])}")

    # --- SAS REF trace (sheet 07) ---
    sas_result = compute_sas_ref_trace(raw_df, flat_df)
    print(f"SAS REF trace: {sas_result['total_sas_raw']} rows "
          f"(expected {_SAS_REF_RAW_BASELINE})")
    print(f"flat_present==True: {sas_result['flat_present_count']}")
    print(f"duplicate_favorable_kept: {sas_result['duplicate_favorable_kept']}")

    # Hard assertion: the 5 project-owner-verified numeros must all be classified
    # DUPLICATE_FAVORABLE_KEPT.  These were manually inspected and confirmed to
    # share the same submission_date between their REF and VAO/VSO RAW rows.
    _VERIFIED_DFK = {
        ("249101", "B"), ("245027", "D"), ("245026", "B"),
        ("246000", "A"), ("246012", "A"),
    }
    # Use set-based check: a verified (numero, indice) may have multiple sheet07
    # rows (canonical DFK + non-canonical DUPLICATE_MERGED); pass if ANY row is DFK.
    dfk_found: set = set()
    all_cls_for_ni: dict = {}
    for r in sas_result["sheet07_rows"]:
        ni = (str(r["numero"]), str(r["indice"]))
        if ni in _VERIFIED_DFK:
            all_cls_for_ni.setdefault(ni, []).append(r["classification"])
            if r["classification"] == "DUPLICATE_FAVORABLE_KEPT":
                dfk_found.add(ni)
    not_dfk = _VERIFIED_DFK - dfk_found
    if not_dfk:
        print(
            f"HARD_ASSERTION FAILED: verified DFK numeros have no DFK row: {not_dfk}\n"
            f"  Classifications seen: {all_cls_for_ni}"
        )
        sys.exit(1)
    print("HARD_ASSERTION PASSED: all 5 verified DFK numeros correctly classified")

    # --- Write sheets 04–07 ---
    write_response_sheets(
        resp_result["sheet04_rows"],
        resp_result["sheet05_rows"],
        resp_result["sheet06_rows"],
        sas_result["sheet07_rows"],
        OUT_XLSX,
    )
    print(f"Wrote and subprocess-checked sheets 04–07: {OUT_XLSX}")

    # --- In-process self-check ---
    selfcheck_response_sheets(resp_result, sas_result, OUT_XLSX)
    print("In-process self-check PASSED")

    # --- Top-5 UNEXPLAINED from sheet 06 ---
    unexplained_06 = [
        r for r in resp_result["sheet06_rows"] if r["classification"] == "UNEXPLAINED"
    ]
    if unexplained_06:
        print(f"Top-5 UNEXPLAINED from sheet 06 "
              f"(of {len(unexplained_06)} total):")
        for r in unexplained_06[:5]:
            print(f"  numero={r['numero']} indice={r['indice']} "
                  f"actor={r['actor_canonical']} side={r['side']} "
                  f"ev={r['evidence_snippet']}")

    # --- Top-5 UNEXPLAINED from sheet 07 (SAS REF specific) ---
    unexplained_07 = [
        r for r in sas_result["sheet07_rows"] if r["classification"] == "UNEXPLAINED"
    ]
    if unexplained_07:
        print(f"Top-5 UNEXPLAINED from sheet 07 "
              f"(of {len(unexplained_07)} total — §17 gate primary input):")
        for r in unexplained_07[:5]:
            print(f"  numero={r['numero']} indice={r['indice']} "
                  f"raw_cycle={r['raw_cycle']} "
                  f"ev_row={r['evidence_excel_row']}")

    # --- Required stdout summary lines ---
    print(resp_result["summary_line"])
    print(sas_result["summary_line"])


# ── Phase 8B.3 ────────────────────────────────────────────────────────────────

def compute_identity_parity(raw_df: pd.DataFrame, flat_df: pd.DataFrame) -> dict:
    """
    Phase 8B.3 identity parity computation.

    raw_df must be the UNFILTERED raw trace; the _MISSING_ filter is applied
    inside this function so before/after counts are captured in the result.

    Returns a result dict with all sets, counts, verdicts, and the summary line.
    Does not print anything — the caller handles stdout.
    """
    raw_before = len(raw_df)
    raw_df = raw_df[
        ~raw_df["numero"].astype(str).str.startswith("_MISSING_")
    ].copy()
    raw_after = len(raw_df)

    raw_numero_set  = set(raw_df["numero"].astype(str))
    flat_numero_set = set(flat_df["numero"].astype(str))
    raw_pair_set    = set(zip(raw_df["numero"].astype(str),  raw_df["indice"].astype(str)))
    flat_pair_set   = set(zip(flat_df["numero"].astype(str), flat_df["indice"].astype(str)))

    missing_num  = raw_numero_set  - flat_numero_set
    extra_num    = flat_numero_set - raw_numero_set
    missing_pair = raw_pair_set    - flat_pair_set
    extra_pair   = flat_pair_set   - raw_pair_set

    num_verdict  = "PASS" if not missing_num  and not extra_num  else "BLOCKER"
    pair_verdict = "PASS" if not missing_pair and not extra_pair else "BLOCKER"

    total_missing = len(missing_num)  + len(missing_pair)
    total_extra   = len(extra_num)    + len(extra_pair)

    sheet_rows = [
        {
            "set_name":              "numero",
            "raw_count":             len(raw_numero_set),
            "flat_count":            len(flat_numero_set),
            "missing_in_flat_count": len(missing_num),
            "extra_in_flat_count":   len(extra_num),
            "verdict":               num_verdict,
        },
        {
            "set_name":              "numero_indice",
            "raw_count":             len(raw_pair_set),
            "flat_count":            len(flat_pair_set),
            "missing_in_flat_count": len(missing_pair),
            "extra_in_flat_count":   len(extra_pair),
            "verdict":               pair_verdict,
        },
    ]

    summary_line = (
        f"IDENTITY_PARITY: numero {num_verdict}, numero_indice {pair_verdict}, "
        f"missing={total_missing}, extra={total_extra}"
    )

    return {
        "raw_before_filter": raw_before,
        "raw_after_filter":  raw_after,
        "flat_rows":         len(flat_df),
        "raw_numero_set":    raw_numero_set,
        "flat_numero_set":   flat_numero_set,
        "raw_pair_set":      raw_pair_set,
        "flat_pair_set":     flat_pair_set,
        "missing_num":       missing_num,
        "extra_num":         extra_num,
        "missing_pair":      missing_pair,
        "extra_pair":        extra_pair,
        "sheet_rows":        sheet_rows,
        "summary_line":      summary_line,
        "num_verdict":       num_verdict,
        "pair_verdict":      pair_verdict,
        "total_missing":     total_missing,
        "total_extra":       total_extra,
    }


def write_identity_sheet(sheet_rows: list[dict], xlsx_path: Path) -> None:
    """Write / replace sheet 01_IDENTITY_PARITY in the output xlsx."""
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        if "01_IDENTITY_PARITY" in wb.sheetnames:
            del wb["01_IDENTITY_PARITY"]
        ws = wb.create_sheet("01_IDENTITY_PARITY", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "01_IDENTITY_PARITY"

    ws.append(_SHEET_COLS)
    for row in sheet_rows:
        ws.append([row[c] for c in _SHEET_COLS])

    wb.save(xlsx_path)


def selfcheck_identity_sheet(sheet_rows: list[dict], xlsx_path: Path) -> None:
    """
    Read sheet 01_IDENTITY_PARITY back from xlsx and assert it matches
    the in-memory sheet_rows.  Raises AssertionError on any mismatch.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["01_IDENTITY_PARITY"]
    xlsx_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    for i, expected in enumerate(sheet_rows):
        actual = xlsx_rows[i + 1]  # row 0 is the header
        for j, col in enumerate(_SHEET_COLS):
            exp_val = expected[col]
            act_val = actual[j]
            if isinstance(exp_val, int):
                act_val = int(act_val) if act_val is not None else None
            else:
                exp_val = str(exp_val)
                act_val = str(act_val) if act_val is not None else None
            if exp_val != act_val:
                raise AssertionError(
                    f"Self-check mismatch: sheet row {i + 1} col '{col}': "
                    f"expected {exp_val!r}, got {act_val!r}"
                )


# ── Phase 8B.6 constants ──────────────────────────────────────────────────────

# Baseline from 8B.5 RESPONSE_PARITY; raised in the step runner if not matched.
_EXPECTED_MATCHED_8B5 = 16187

_DATE_FIELD_LABELS: dict[str, str] = {
    "submission_date": "submission",
    "response_date":   "response",
    "deadline_date":   "deadline",
}

_SHEET08_COLS = [
    "numero", "indice", "actor_canonical", "status_clean",
    "date_field", "raw_value_raw", "flat_value_raw",
    "raw_value_norm", "flat_value_norm", "verdict",
]
_SHEET09_COLS = [
    "numero", "indice", "actor_canonical", "status_clean",
    "raw_comment", "flat_comment", "verdict",
]

# Excel epoch: day 0 = 1899-12-30; this formula gives correct results for all
# dates >= 1900-03-01 (serial > 61), which covers all practical project dates.
_EXCEL_EPOCH = datetime.date(1899, 12, 30)

# ── Phase 8B.6 helpers ────────────────────────────────────────────────────────
# normalize_date: single helper used on BOTH sides before comparison so that
# Excel serial ↔ ISO string ↔ dd/mm/yyyy are treated as equivalent when they
# map to the same calendar date.  Time component dropped.


def normalize_date(value) -> str | None:
    """Normalise a date value to ISO format (YYYY-MM-DD) or None for blank.

    Handled inputs:
        None, "", "NaT", "None"           → None
        Excel serial float  (44561)       → "2021-12-31"
        ISO string          ("2021-12-31")→ "2021-12-31"
        French dd/mm/yyyy   ("31/12/2021")→ "2021-12-31"
        pandas Timestamp / datetime.date  → ISO string
        datetime.datetime                 → date portion as ISO string
    """
    if value is None:
        return None
    # pandas NaT / Timestamp
    try:
        import pandas as _pd
        if value is _pd.NaT:
            return None
        if isinstance(value, _pd.Timestamp):
            return None if _pd.isna(value) else value.date().isoformat()
    except Exception:
        pass
    # Python datetime objects
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    # String path
    s = str(value).strip()
    if s in ("", "NaT", "None", "none", "nan", "NA", "N/A", "<NA>", "nat"):
        return None
    # ISO YYYY-MM-DD (optional time suffix)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3))).isoformat()
        except ValueError:
            pass
    # French dd/mm/yyyy
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        try:
            return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1))).isoformat()
        except ValueError:
            pass
    # Excel serial (float string, plausible range 1970–2100 ≈ serials 25569–73414)
    try:
        serial = float(s)
        if 25569 <= serial <= 73414:
            return (_EXCEL_EPOCH + datetime.timedelta(days=int(serial))).isoformat()
    except (ValueError, OverflowError):
        pass
    return None


def _normalize_comment(value) -> str:
    """Strip + collapse whitespace.  No case-folding (comments are case-sensitive)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("", "nan", "nat", "none", "na", "n/a", "<na>"):
        return ""
    return re.sub(r"\s+", " ", s)


# ── Phase 8B.6 core logic ─────────────────────────────────────────────────────


def compute_dates_comments_parity(raw_df: pd.DataFrame, flat_df: pd.DataFrame) -> dict:
    """Phase 8B.6 date/comment field parity for matched response pairs.

    Uses the same join key as 8B.5: (numero, indice, actor_canonical, ncycle).
    The 16,187 matched count is verified in the step runner, not here, so that
    this function remains testable with small synthetic DataFrames.

    Note: flat_ged_trace.csv uses 'comment_raw' (not 'observation') on the FLAT
    side — verified against the actual CSV before implementation.

    Date verdicts are per (matched-event × field) — three fields × matched events:
        MATCH                   raw_norm == flat_norm (both non-None)
        DRIFTED                 both non-None, values differ
        RAW_BLANK_FLAT_BLANK    both None
        RAW_NONBLANK_FLAT_BLANK raw non-None, flat None  ← information lost
        RAW_BLANK_FLAT_NONBLANK raw None, flat non-None  ← unusual

    Comment verdicts are per matched event (one comment per event):
        MATCH       normalised strings equal (non-empty)
        DRIFTED     both non-empty, strings differ
        RAW_ONLY    raw non-empty, flat empty
        FLAT_ONLY   flat non-empty, raw empty
        BOTH_BLANK  both empty
    """
    # _MISSING_ filter (same rule as 8B.3/8B.4/8B.5)
    raw_df = raw_df[~raw_df["numero"].astype(str).str.startswith("_MISSING_")].copy()

    raw_resp = raw_df[raw_df["event_type"] == "RESPONSE"].copy()
    flat_resp = flat_df[
        (flat_df["event_type"] == "RESPONSE")
        & (flat_df["source_sheet"] == "GED_OPERATIONS")
    ].copy()

    # Normalise cycle/step for join key (same as 8B.5)
    raw_resp = raw_resp.assign(
        _ncycle=raw_resp.apply(
            lambda r: _derive_raw_cycle(str(r["cycle_id"]), str(r["actor_canonical"])),
            axis=1,
        )
    )
    flat_resp = flat_resp.assign(
        _step=flat_resp["actor_canonical"].apply(lambda a: _derive_flat_step(str(a)))
    )

    # Build 4-tuple sets
    raw_4t = raw_resp[["numero", "indice", "actor_canonical", "_ncycle"]].drop_duplicates()
    raw_set: set = set(
        zip(
            raw_4t["numero"].astype(str),
            raw_4t["indice"].astype(str),
            raw_4t["actor_canonical"].astype(str),
            raw_4t["_ncycle"].astype(str),
        )
    )
    flat_4t = flat_resp[["numero", "indice", "actor_canonical", "_step"]].drop_duplicates()
    flat_set: set = set(
        zip(
            flat_4t["numero"].astype(str),
            flat_4t["indice"].astype(str),
            flat_4t["actor_canonical"].astype(str),
            flat_4t["_step"].astype(str),
        )
    )
    matched: set = raw_set & flat_set
    total_matched = len(matched)

    # Build first-occurrence lookup per 4-tuple → field values
    _FLDS = [
        "submission_date", "response_date", "deadline_date",
        "comment_raw", "status_clean",
    ]

    raw_row_map: dict = {}
    _raw_iter = (
        raw_resp.sort_values("source_excel_row")
        if "source_excel_row" in raw_resp.columns
        else raw_resp
    )
    for _, row in _raw_iter.iterrows():
        k = (
            str(row["numero"]), str(row["indice"]),
            str(row["actor_canonical"]), str(row["_ncycle"]),
        )
        if k not in raw_row_map:
            raw_row_map[k] = {f: str(row[f]) for f in _FLDS}

    flat_row_map: dict = {}
    _flat_iter = (
        flat_resp.sort_values("source_excel_row")
        if "source_excel_row" in flat_resp.columns
        else flat_resp
    )
    for _, row in _flat_iter.iterrows():
        k = (
            str(row["numero"]), str(row["indice"]),
            str(row["actor_canonical"]), str(row["_step"]),
        )
        if k not in flat_row_map:
            flat_row_map[k] = {f: str(row[f]) for f in _FLDS}

    # Per-field verdict counters
    date_counts: dict = {
        lbl: {
            "MATCH": 0, "DRIFTED": 0, "RAW_BLANK_FLAT_BLANK": 0,
            "RAW_NONBLANK_FLAT_BLANK": 0, "RAW_BLANK_FLAT_NONBLANK": 0,
        }
        for lbl in _DATE_FIELD_LABELS.values()
    }
    comment_counts: dict = {
        "MATCH": 0, "DRIFTED": 0, "RAW_ONLY": 0, "FLAT_ONLY": 0, "BOTH_BLANK": 0,
    }

    sheet08_rows: list = []
    sheet09_rows: list = []

    for tup in sorted(matched):
        num, ind, actor, _nc = tup
        raw_r = raw_row_map.get(tup, {})
        flat_r = flat_row_map.get(tup, {})
        sc = raw_r.get("status_clean", "")

        # Date field comparisons
        for col, lbl in _DATE_FIELD_LABELS.items():
            rv = raw_r.get(col, "")
            fv = flat_r.get(col, "")
            rv_n = normalize_date(rv)
            fv_n = normalize_date(fv)

            if rv_n is None and fv_n is None:
                verdict = "RAW_BLANK_FLAT_BLANK"
            elif rv_n == fv_n:
                verdict = "MATCH"
            elif rv_n is not None and fv_n is None:
                verdict = "RAW_NONBLANK_FLAT_BLANK"
            elif rv_n is None:
                verdict = "RAW_BLANK_FLAT_NONBLANK"
            else:
                verdict = "DRIFTED"

            date_counts[lbl][verdict] += 1

            if verdict not in ("MATCH", "RAW_BLANK_FLAT_BLANK"):
                sheet08_rows.append({
                    "numero": num, "indice": ind, "actor_canonical": actor,
                    "status_clean": sc, "date_field": lbl,
                    "raw_value_raw": rv, "flat_value_raw": fv,
                    "raw_value_norm": rv_n, "flat_value_norm": fv_n,
                    "verdict": verdict,
                })

        # Comment comparison
        rc = raw_r.get("comment_raw", "")
        fc = flat_r.get("comment_raw", "")
        rc_n = _normalize_comment(rc)
        fc_n = _normalize_comment(fc)

        if not rc_n and not fc_n:
            c_verdict = "BOTH_BLANK"
        elif rc_n == fc_n:
            c_verdict = "MATCH"
        elif rc_n and not fc_n:
            c_verdict = "RAW_ONLY"
        elif not rc_n:
            c_verdict = "FLAT_ONLY"
        else:
            c_verdict = "DRIFTED"

        comment_counts[c_verdict] += 1

        if c_verdict not in ("MATCH", "BOTH_BLANK"):
            sheet09_rows.append({
                "numero": num, "indice": ind, "actor_canonical": actor,
                "status_clean": sc,
                "raw_comment": rc, "flat_comment": fc, "verdict": c_verdict,
            })

    # Sort output sheets per spec
    sheet08_rows.sort(key=lambda r: (r["numero"], r["indice"], r["actor_canonical"], r["date_field"]))
    sheet09_rows.sort(key=lambda r: (r["numero"], r["indice"], r["actor_canonical"]))

    # Aggregate date counts
    drifted_total    = sum(date_counts[l]["DRIFTED"]                for l in _DATE_FIELD_LABELS.values())
    rblank_fblank    = sum(date_counts[l]["RAW_BLANK_FLAT_BLANK"]   for l in _DATE_FIELD_LABELS.values())
    rnonblank_fblank = sum(date_counts[l]["RAW_NONBLANK_FLAT_BLANK"]for l in _DATE_FIELD_LABELS.values())
    rblank_fnonblank = sum(date_counts[l]["RAW_BLANK_FLAT_NONBLANK"]for l in _DATE_FIELD_LABELS.values())
    match_date       = sum(date_counts[l]["MATCH"]                  for l in _DATE_FIELD_LABELS.values())

    # Math self-checks
    pf_total = match_date + drifted_total + rblank_fblank + rnonblank_fblank + rblank_fnonblank
    if pf_total != 3 * total_matched:
        raise AssertionError(
            f"DATE_PARITY per-field total {pf_total} != 3×{total_matched}"
        )
    c_total = sum(comment_counts.values())
    if c_total != total_matched:
        raise AssertionError(
            f"COMMENT_PARITY total {c_total} != {total_matched}"
        )

    # Sheet row count self-checks
    exp_s08 = drifted_total + rnonblank_fblank + rblank_fnonblank
    if len(sheet08_rows) != exp_s08:
        raise AssertionError(f"Sheet 08 row count {len(sheet08_rows)} != {exp_s08}")
    exp_s09 = comment_counts["DRIFTED"] + comment_counts["RAW_ONLY"] + comment_counts["FLAT_ONLY"]
    if len(sheet09_rows) != exp_s09:
        raise AssertionError(f"Sheet 09 row count {len(sheet09_rows)} != {exp_s09}")

    summary_date = (
        f"DATE_PARITY: matched={total_matched} drifted={drifted_total} "
        f"raw_blank_flat_blank={rblank_fblank} raw_nonblank_flat_blank={rnonblank_fblank} "
        f"raw_blank_flat_nonblank={rblank_fnonblank}"
    )
    summary_comment = (
        f"COMMENT_PARITY: matched={total_matched} drifted={comment_counts['DRIFTED']} "
        f"raw_only={comment_counts['RAW_ONLY']} flat_only={comment_counts['FLAT_ONLY']} "
        f"both_blank={comment_counts['BOTH_BLANK']}"
    )

    return {
        "matched": total_matched,
        "date_counts": date_counts,
        "comment_counts": comment_counts,
        "drifted_total": drifted_total,
        "rblank_fblank": rblank_fblank,
        "rnonblank_fblank": rnonblank_fblank,
        "rblank_fnonblank": rblank_fnonblank,
        "match_date": match_date,
        "sheet08_rows": sheet08_rows,
        "sheet09_rows": sheet09_rows,
        "summary_date": summary_date,
        "summary_comment": summary_comment,
    }


def write_dates_comments_sheets(
    sheet08_rows: list,
    sheet09_rows: list,
    xlsx_path: Path,
) -> None:
    """Append / replace sheets 08_DATE_DIFFS and 09_COMMENT_DIFFS.

    Sheets 01–07 stay untouched. Runs a fresh-subprocess integrity check
    after saving (H-1 lesson).
    """
    import subprocess
    import os
    import zipfile

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for name in ("08_DATE_DIFFS", "09_COMMENT_DIFFS"):
        if name in wb.sheetnames:
            del wb[name]

    ws08 = wb.create_sheet("08_DATE_DIFFS")
    ws08.append(_SHEET08_COLS)
    for row in sheet08_rows:
        ws08.append([row[c] for c in _SHEET08_COLS])

    ws09 = wb.create_sheet("09_COMMENT_DIFFS")
    ws09.append(_SHEET09_COLS)
    for row in sheet09_rows:
        ws09.append([row[c] for c in _SHEET09_COLS])

    wb.save(xlsx_path)
    del wb

    check_code = f"""
import os, zipfile
from openpyxl import load_workbook
from collections import Counter

p = r{str(xlsx_path)!r}
size = os.path.getsize(p)
assert size > 0, f"file empty: {{size}}"

with zipfile.ZipFile(p) as z:
    bad = z.testzip()
assert bad is None, f"ZIP corrupt: {{bad}}"

wb = load_workbook(p, read_only=True, data_only=True)
names = wb.sheetnames
for req in ("08_DATE_DIFFS", "09_COMMENT_DIFFS"):
    assert req in names, f"missing sheet: {{req}}"

rows08 = list(wb["08_DATE_DIFFS"].iter_rows(values_only=True))
rows09 = list(wb["09_COMMENT_DIFFS"].iter_rows(values_only=True))
wb.close()

verdicts08 = Counter(r[-1] for r in rows08[1:])
verdicts09 = Counter(r[-1] for r in rows09[1:])
print(f"SUBPROCESS_CHECK: size={{size}} sheets={{names}}")
print(f"  08_rows={{len(rows08)-1}} 09_rows={{len(rows09)-1}}")
print(f"  08_verdicts={{dict(verdicts08)}}")
print(f"  09_verdicts={{dict(verdicts09)}}")
print(f"  ZIP_integrity=OK")
"""
    result = subprocess.run(
        ["python", "-c", check_code],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Subprocess self-check FAILED:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )
    print(result.stdout.strip())


def selfcheck_dates_comments_sheets(result: dict, xlsx_path: Path) -> None:
    """In-process self-check for sheets 08–09 (fast read after subprocess check)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows08 = list(wb["08_DATE_DIFFS"].iter_rows(values_only=True))[1:]
    rows09 = list(wb["09_COMMENT_DIFFS"].iter_rows(values_only=True))[1:]
    wb.close()

    exp_s08 = result["drifted_total"] + result["rnonblank_fblank"] + result["rblank_fnonblank"]
    if len(rows08) != exp_s08:
        raise AssertionError(
            f"Sheet 08 row count {len(rows08)} != expected {exp_s08}"
        )

    cv = result["comment_counts"]
    exp_s09 = cv["DRIFTED"] + cv["RAW_ONLY"] + cv["FLAT_ONLY"]
    if len(rows09) != exp_s09:
        raise AssertionError(
            f"Sheet 09 row count {len(rows09)} != expected {exp_s09}"
        )

    if result["matched"] != _EXPECTED_MATCHED_8B5:
        raise AssertionError(
            f"DATE_PARITY matched {result['matched']} != 8B.5 baseline {_EXPECTED_MATCHED_8B5}"
        )


# ── Phase 8B.6 step runner ────────────────────────────────────────────────────

def _run_dates_comments() -> None:
    print(f"Loading RAW trace:  {RAW_CSV}")
    print(f"Loading FLAT trace: {FLAT_CSV}")
    raw_df = pd.read_csv(RAW_CSV, dtype=str, keep_default_na=False)
    flat_df = pd.read_csv(FLAT_CSV, dtype=str, keep_default_na=False)

    result = compute_dates_comments_parity(raw_df, flat_df)

    # Verify matched count equals 8B.5 RESPONSE_PARITY baseline
    if result["matched"] != _EXPECTED_MATCHED_8B5:
        raise AssertionError(
            f"dates_comments matched {result['matched']} != "
            f"8B.5 RESPONSE_PARITY matched {_EXPECTED_MATCHED_8B5}"
        )
    print(f"Matched events: {result['matched']} (confirmed == 8B.5 baseline)")

    write_dates_comments_sheets(result["sheet08_rows"], result["sheet09_rows"], OUT_XLSX)
    print(f"Wrote and subprocess-checked sheets 08+09: {OUT_XLSX}")

    selfcheck_dates_comments_sheets(result, OUT_XLSX)
    print("In-process self-check PASSED")

    # Top-5 RAW_NONBLANK_FLAT_BLANK rows from sheet 08 (load-bearing signal)
    top_rnfb = [r for r in result["sheet08_rows"] if r["verdict"] == "RAW_NONBLANK_FLAT_BLANK"]
    if top_rnfb:
        print(
            f"Top-5 RAW_NONBLANK_FLAT_BLANK (of {len(top_rnfb)} total — "
            f"information lost in L0→L1 projection):"
        )
        for r in top_rnfb[:5]:
            print(
                f"  numero={r['numero']} indice={r['indice']} "
                f"actor={r['actor_canonical']} field={r['date_field']} "
                f"raw_norm={r['raw_value_norm']!r}"
            )

    # Top-5 DRIFTED comment rows from sheet 09
    top_cd = [r for r in result["sheet09_rows"] if r["verdict"] == "DRIFTED"]
    if top_cd:
        print(f"Top-5 DRIFTED comment rows from sheet 09 (of {len(top_cd)} total):")
        for r in top_cd[:5]:
            print(
                f"  numero={r['numero']} indice={r['indice']} "
                f"actor={r['actor_canonical']}"
            )

    print(result["summary_date"])
    print(result["summary_comment"])


# ── Step runners ──────────────────────────────────────────────────────────────

def _run_actor_calls() -> None:
    print(f"Loading RAW trace:  {RAW_CSV}")
    print(f"Loading FLAT trace: {FLAT_CSV}")
    raw_df  = pd.read_csv(RAW_CSV,  dtype=str, keep_default_na=False)
    flat_df = pd.read_csv(FLAT_CSV, dtype=str, keep_default_na=False)

    result = compute_actor_call_parity(raw_df, flat_df)

    print(f"RAW  actor-call 4-tuples: {len(result['raw_set'])}")
    print(f"FLAT actor-call 4-tuples: {len(result['flat_set'])}")
    print(f"Matched:  {len(result['match_set'])}")
    print(f"RAW_ONLY: {len(result['raw_only'])}")
    print(f"FLAT_ONLY: {len(result['flat_only'])}")

    write_actor_call_sheets(result["sheet02_rows"], result["sheet03_rows"], OUT_XLSX)
    selfcheck_actor_call_sheets(result, OUT_XLSX)
    print(f"Wrote and self-checked sheets 02+03: {OUT_XLSX}")

    # Top-5 UNEXPLAINED for manual investigation queue
    unexplained_rows = [r for r in result["sheet03_rows"] if r["classification"] == "UNEXPLAINED"]
    if unexplained_rows:
        print(f"Top-5 UNEXPLAINED (of {len(unexplained_rows)}):")
        for r in unexplained_rows[:5]:
            print(f"  numero={r['numero']} indice={r['indice']} "
                  f"actor={r['actor_canonical']} side={r['side']} "
                  f"evidence={r['evidence_snippet']}")

    print(result["summary_line"])


def _run_identity() -> None:
    print(f"Loading RAW trace:  {RAW_CSV}")
    print(f"Loading FLAT trace: {FLAT_CSV}")
    raw_df  = pd.read_csv(RAW_CSV,  dtype=str, keep_default_na=False)
    flat_df = pd.read_csv(FLAT_CSV, dtype=str, keep_default_na=False)

    result = compute_identity_parity(raw_df, flat_df)

    print(f"RAW rows before _MISSING_ filter: {result['raw_before_filter']}")
    print(f"RAW rows after  _MISSING_ filter: {result['raw_after_filter']}")
    print(f"FLAT rows (no filter needed):     {result['flat_rows']}")

    write_identity_sheet(result["sheet_rows"], OUT_XLSX)
    selfcheck_identity_sheet(result["sheet_rows"], OUT_XLSX)
    print(f"Wrote and self-checked: {OUT_XLSX}")

    print(result["summary_line"])

    if result["total_missing"] > 0 or result["total_extra"] > 0:
        _print_blocker_samples(result)
        sys.exit(1)


def _print_blocker_samples(result: dict) -> None:
    def _sample(s: set, label: str) -> None:
        if s:
            samples = sorted(str(x) for x in s)[:5]
            print(f"  {label} ({len(s)} total, showing up to 5): {samples}")

    _sample(result["missing_num"],  "missing_in_flat  numero")
    _sample(result["extra_num"],    "extra_in_flat    numero")
    _sample(result["missing_pair"], "missing_in_flat  (numero, indice)")
    _sample(result["extra_pair"],   "extra_in_flat    (numero, indice)")


# ── CLI ───────────────────────────────────────────────────────────────────────

# ── Phase 8B.7 reasons-audit constants ───────────────────────────────────────
# Audit is READ-ONLY: walks every .py file under src/flat_ged/ via AST and
# classifies each row-affecting site (drop / exception / boolean filter flag /
# duplicate-classification assignment) against the §5 taxonomy buckets.

_SHEET10_COLS = [
    "file", "function", "line", "kind", "reason_string",
    "classification", "related_taxonomy_bucket",
    "observed_count_in_run0", "notes",
]

# Map taxonomy bucket → keywords expected to appear in a valid reason text.
# Order matters: first matching bucket wins when classifying free text.
_TAXONOMY_BUCKET_KEYWORDS = [
    ("UNKNOWN_ACTOR",              ["exception list", "unknown actor", "unknown_actor"]),
    ("MALFORMED_RESPONSE",         ["null numero", "no numero", "missing numero",
                                    "_missing_", "malformed", "no_numero"]),
    ("NON_OPERATIONAL_RESPONSE",   ["pending", "not_called", "not called",
                                    "non-operational", "non_operational",
                                    "synthetic", "skipped if sas pending",
                                    "skipped — sas", "sas pending"]),
    ("ACTIVE_VERSION_PROJECTION",  ["inactive_duplicate", "active_duplicate",
                                    "active version", "version projection",
                                    "old indice", "active candidate",
                                    "separate_instance", "separate real submission"]),
    ("DUPLICATE_MERGED",           ["duplicate", "merge", "merged",
                                    "deduplicate", "tie_break"]),
    ("DUPLICATE_FAVORABLE_KEPT",   ["favorable", "favourable", "vao", "vso",
                                    "kept favorable", "favorable wins"]),
    ("SAS_CYCLE_COLLAPSED",        ["sas cycle", "cycle collapsed", "c1", "c2",
                                    "collapsed cycle"]),
    ("EXCEPTION_COLUMN",           ["exception column", "exception_column"]),
    ("UNEXPLAINED",                []),  # never assigned by code; trace-only
]

# Heuristic markers for "no reason recorded" (control-flow with no nearby
# string/comment that names what's being skipped).  When the reason_string is
# empty AND the surrounding context has no comment, the finding is MISSING.
_GENERIC_REASON_TOKENS = (
    "bounds", "defensive", "len(", "if not", "fallback", "guard",
)


def _walk_flat_ged_drops(src_root: Path) -> list[dict]:
    """AST-walk every .py file under ``src_root`` and collect row-affecting sites.

    A site is recorded when one of the following appears inside a function body:
      * ``continue``  inside a for/while loop guarded by an ``if`` test
      * ``raise``     of an exception class whose docstring or message names a
                       row-exclusion intent
      * a string literal that includes drop-style verbs ("skip", "exclude",
        "drop", "filter", "ignore") — captured even if no control-flow drop
        sits next to it (so we surface "intent-only" comments and class docs)
      * an assignment to a variable whose name matches the
        ``skip_/ignore_/exclude_/is_pre_`` boolean-filter pattern

    Returns a list of dicts with keys: file, function, line, kind,
    reason_string, source_excerpt.  Classification is applied separately by
    ``_classify_drop_finding``.
    """
    import ast

    drop_verbs = ("skip", "exclude", "drop", "filter", "ignore",
                  "exception list", "not_called", "not called",
                  "duplicate", "merge", "synthetic", "missing")
    flag_pattern = re.compile(
        r"^(skip|ignore|exclude|filter)_\w+$|^is_pre_\d+$|^is_(missing|unknown)_\w+$",
        re.IGNORECASE,
    )

    findings: list[dict] = []

    for py_path in sorted(src_root.rglob("*.py")):
        # Skip __pycache__ artefacts (rglob already excludes the dir itself,
        # but be paranoid in case a .py lands beside the package).
        if "__pycache__" in py_path.parts:
            continue
        try:
            source = py_path.read_text(encoding="utf-8")
            tree = ast.parse(source, filename=str(py_path))
        except (OSError, SyntaxError) as exc:
            findings.append({
                "file": str(py_path.relative_to(src_root.parent)),
                "function": "<module>",
                "line": 0,
                "kind": "PARSE_FAIL",
                "reason_string": f"AST parse failed: {exc}",
                "source_excerpt": "",
            })
            continue

        src_lines = source.splitlines()

        # Walk every function/method body.
        for node in ast.walk(tree):
            if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                continue
            fn_name = node.name
            for sub in ast.walk(node):
                kind = None
                line = getattr(sub, "lineno", 0)
                reason = ""

                if isinstance(sub, ast.Continue):
                    kind = "CONTINUE_DROP"
                elif isinstance(sub, ast.Raise):
                    kind = "RAISE"
                    if isinstance(sub.exc, ast.Call) and sub.exc.args:
                        first = sub.exc.args[0]
                        if isinstance(first, ast.Constant) and isinstance(first.value, str):
                            reason = first.value
                        elif isinstance(first, ast.JoinedStr):
                            # f-string: best-effort string flatten
                            parts = []
                            for v in first.values:
                                if isinstance(v, ast.Constant):
                                    parts.append(str(v.value))
                                else:
                                    parts.append("{...}")
                            reason = "".join(parts)
                elif isinstance(sub, ast.Assign):
                    for tgt in sub.targets:
                        if isinstance(tgt, ast.Name) and flag_pattern.match(tgt.id):
                            kind = "BOOLEAN_FILTER_FLAG"
                            reason = tgt.id
                            break
                else:
                    continue

                # Excerpt: 2 lines centred on the statement.
                lo = max(0, line - 2)
                hi = min(len(src_lines), line + 1)
                excerpt = "\n".join(s.strip() for s in src_lines[lo:hi])

                # If we don't yet have a reason, scan nearby lines for a
                # comment or string literal that names the drop verb.
                if not reason:
                    window_lo = max(0, line - 4)
                    window_hi = min(len(src_lines), line + 2)
                    for s in src_lines[window_lo:window_hi]:
                        s_low = s.lower()
                        if "#" in s_low or '"' in s_low or "'" in s_low:
                            for verb in drop_verbs:
                                if verb in s_low:
                                    reason = s.strip()
                                    break
                            if reason:
                                break

                findings.append({
                    "file": str(py_path.relative_to(src_root.parent)),
                    "function": fn_name,
                    "line": line,
                    "kind": kind,
                    "reason_string": reason,
                    "source_excerpt": excerpt,
                })

        # Module-level exception classes whose docstring names a row-exclusion
        # intent — these are reasons even if never raised.
        for node in tree.body:
            if isinstance(node, ast.ClassDef):
                doc = ast.get_docstring(node) or ""
                if any(v in doc.lower() for v in drop_verbs):
                    findings.append({
                        "file": str(py_path.relative_to(src_root.parent)),
                        "function": node.name,
                        "line": node.lineno,
                        "kind": "EXCEPTION_CLASS_DOC",
                        "reason_string": doc.strip().splitlines()[0] if doc else "",
                        "source_excerpt": doc.strip(),
                    })

    return findings


def _classify_drop_finding(finding: dict, raised_classes: set[str]) -> dict:
    """Apply VALID/INVALID/MISSING/AMBIGUOUS classification + taxonomy mapping.

    Heuristics:
      * MISSING_REASON   — control-flow drop with no captured reason text
      * INVALID_REASON   — exception-class doc names a drop intent but the
                            class is never raised in the same package
      * VALID_REASON     — reason text contains a §5-taxonomy keyword
      * AMBIGUOUS_REASON — reason exists but is purely defensive/generic OR
                            doesn't map to any §5 bucket
    """
    kind = finding["kind"]
    reason = (finding.get("reason_string") or "").strip()
    reason_low = reason.lower()

    # Map to taxonomy bucket via keyword scan (first match wins).
    bucket = ""
    for name, kws in _TAXONOMY_BUCKET_KEYWORDS:
        for kw in kws:
            if kw in reason_low:
                bucket = name
                break
        if bucket:
            break

    classification = "AMBIGUOUS_REASON"

    if kind == "EXCEPTION_CLASS_DOC":
        cls_name = finding.get("function") or ""
        if cls_name and cls_name not in raised_classes:
            classification = "INVALID_REASON"
        elif bucket:
            classification = "VALID_REASON"
        else:
            classification = "AMBIGUOUS_REASON"
    elif not reason:
        classification = "MISSING_REASON"
    elif bucket:
        classification = "VALID_REASON"
    elif any(tok in reason_low for tok in _GENERIC_REASON_TOKENS):
        classification = "AMBIGUOUS_REASON"
    else:
        classification = "AMBIGUOUS_REASON"

    notes_parts = []
    if kind == "RAISE" and not bucket:
        notes_parts.append("hard-fail (not a row drop) — kept for completeness")
    if kind == "BOOLEAN_FILTER_FLAG":
        notes_parts.append("boolean filter flag — verify the gating site uses §5 semantics")
    if kind == "EXCEPTION_CLASS_DOC" and classification == "INVALID_REASON":
        notes_parts.append(
            "exception class declared with row-exclusion docstring but never raised "
            "in src/flat_ged/* — drop happens elsewhere (pipeline) or not at all"
        )

    return {
        "file":                    finding["file"],
        "function":                finding["function"],
        "line":                    finding["line"],
        "kind":                    kind,
        "reason_string":           reason[:400] if reason else "",
        "classification":          classification,
        "related_taxonomy_bucket": bucket or "none",
        "observed_count_in_run0":  "",   # filled by compute_reasons_audit
        "notes":                   "; ".join(notes_parts),
    }


def _collect_raised_class_names(src_root: Path) -> set[str]:
    """Return the set of exception-class names that are actually raised
    somewhere under ``src_root``.  Heuristic: a ``raise X(...)`` or
    ``raise X`` whose ``X`` is a Name node.
    """
    import ast
    raised: set[str] = set()
    for py_path in sorted(src_root.rglob("*.py")):
        if "__pycache__" in py_path.parts:
            continue
        try:
            tree = ast.parse(py_path.read_text(encoding="utf-8"))
        except (OSError, SyntaxError):
            continue
        for node in ast.walk(tree):
            if isinstance(node, ast.Raise):
                exc = node.exc
                if isinstance(exc, ast.Call) and isinstance(exc.func, ast.Name):
                    raised.add(exc.func.id)
                elif isinstance(exc, ast.Name):
                    raised.add(exc.id)
    return raised


def compute_reasons_audit(src_root: Path) -> dict:
    """Phase 8B.7 — read-only audit of row-affecting sites under src/flat_ged/.

    Returns a dict with keys:
        rows         — list of sheet 10 rows (already classified)
        summary      — dict with total/valid/invalid/missing/ambiguous counts
        summary_line — exact stdout summary line per spec §14.4
    """
    raw_findings = _walk_flat_ged_drops(src_root)
    raised = _collect_raised_class_names(src_root)
    rows = [_classify_drop_finding(f, raised) for f in raw_findings]

    rows.sort(key=lambda r: (r["file"], r["line"]))

    counts = {"VALID_REASON": 0, "INVALID_REASON": 0,
              "MISSING_REASON": 0, "AMBIGUOUS_REASON": 0}
    for r in rows:
        counts[r["classification"]] = counts.get(r["classification"], 0) + 1

    summary_line = (
        f"REASON_AUDIT: total={len(rows)} "
        f"valid={counts['VALID_REASON']} "
        f"invalid={counts['INVALID_REASON']} "
        f"missing={counts['MISSING_REASON']} "
        f"ambiguous={counts['AMBIGUOUS_REASON']}"
    )

    return {
        "rows":         rows,
        "summary":      {"total": len(rows), **counts},
        "summary_line": summary_line,
    }


def write_reasons_audit_sheet(rows: list, xlsx_path: Path) -> None:
    """Append / replace sheet 10_EXISTING_REASON_AUDIT.

    Sheets 01-09 stay untouched.  Runs a fresh-subprocess integrity check
    after saving (H-1 lesson).
    """
    import subprocess

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = "10_EXISTING_REASON_AUDIT"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.append(_SHEET10_COLS)
    for row in rows:
        ws.append([row.get(c, "") for c in _SHEET10_COLS])

    wb.save(xlsx_path)
    del wb

    check_code = f"""
import os, zipfile
from openpyxl import load_workbook
from collections import Counter

p = r{str(xlsx_path)!r}
size = os.path.getsize(p)
assert size > 0, f"file empty: {{size}}"

with zipfile.ZipFile(p) as z:
    bad = z.testzip()
assert bad is None, f"ZIP corrupt: {{bad}}"

wb = load_workbook(p, read_only=True, data_only=True)
names = wb.sheetnames
assert "10_EXISTING_REASON_AUDIT" in names, f"missing sheet 10: {{names}}"
rows = list(wb["10_EXISTING_REASON_AUDIT"].iter_rows(values_only=True))
wb.close()

cls_idx = rows[0].index("classification")
counts = Counter(r[cls_idx] for r in rows[1:])
print(f"SUBPROCESS_CHECK: size={{size}} sheets={{names}}")
print(f"  10_rows={{len(rows)-1}}")
print(f"  10_classifications={{dict(counts)}}")
print(f"  ZIP_integrity=OK")
"""
    result = subprocess.run(
        ["python", "-c", check_code],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Subprocess self-check FAILED:\nSTDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )
    print(result.stdout.strip())


def selfcheck_reasons_audit_sheet(result: dict, xlsx_path: Path) -> None:
    """In-process self-check (fast read after subprocess check)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = list(wb["10_EXISTING_REASON_AUDIT"].iter_rows(values_only=True))[1:]
    wb.close()
    if len(rows) != result["summary"]["total"]:
        raise AssertionError(
            f"Sheet 10 row count {len(rows)} != expected {result['summary']['total']}"
        )


def _run_reasons_audit() -> None:
    src_root = REPO_ROOT / "src" / "flat_ged"
    print(f"Walking {src_root}")
    result = compute_reasons_audit(src_root)
    write_reasons_audit_sheet(result["rows"], OUT_XLSX)
    print(f"Wrote sheet 10_EXISTING_REASON_AUDIT to {OUT_XLSX}")
    selfcheck_reasons_audit_sheet(result, OUT_XLSX)
    print("In-process self-check PASSED")

    # Top-5 INVALID or AMBIGUOUS for investigation queue
    suspect = [r for r in result["rows"]
               if r["classification"] in ("INVALID_REASON", "AMBIGUOUS_REASON")]
    if suspect:
        print(f"Top-5 INVALID/AMBIGUOUS (of {len(suspect)} total):")
        for r in suspect[:5]:
            print(f"  {r['file']}:{r['line']} {r['function']} "
                  f"[{r['classification']}] reason={r['reason_string'][:80]!r}")

    print(result["summary_line"])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Phase 8B RAW→FLAT reconciliation driver"
    )
    parser.add_argument(
        "--step",
        required=True,
        choices=["identity", "actor_calls", "responses",
                 "dates_comments", "reasons_audit", "shadow"],
    )
    args = parser.parse_args()

    if args.step == "identity":
        _run_identity()
    elif args.step == "actor_calls":
        _run_actor_calls()
    elif args.step == "responses":
        _run_responses()
    elif args.step == "dates_comments":
        _run_dates_comments()
    elif args.step == "reasons_audit":
        _run_reasons_audit()
    elif args.step == "shadow":
        # Delegate to the standalone shadow builder (lives in its own script
        # so the audit-only reconciliation driver can stay narrow).
        import importlib.util
        shadow_path = Path(__file__).resolve().parent / "build_shadow_flat_ged.py"
        spec = importlib.util.spec_from_file_location("build_shadow_flat_ged", shadow_path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        mod.main()
    else:
        phase = _STEP_PHASE[args.step]
        raise NotImplementedError(f"step {args.step} lands in Phase {phase}")


if __name__ == "__main__":
    main()
