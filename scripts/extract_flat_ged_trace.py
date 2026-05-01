#!/usr/bin/env python3
"""
scripts/extract_flat_ged_trace.py  —  Phase 8B, Step 8B.2
----------------------------------------------------------
FLAT GED Event Extraction.

Reads output/intermediate/FLAT_GED.xlsx (sheets GED_OPERATIONS and
GED_RAW_FLAT) and emits one Event row per source row into the same
column shape as scripts/extract_raw_ged_trace.py.

Output:
  output/debug/flat_ged_trace.csv
  output/debug/flat_ged_trace.xlsx

═══════════════════════════════════════════════════════════════════════════════
RESOLVED OPEN QUESTIONS (§23 of PHASE_8B_RAW_TO_FLAT_RECONCILIATION.md)
═══════════════════════════════════════════════════════════════════════════════

OQ-1  Multi-cycle SAS in RAW:
  "0-SAS" appears twice in RAW GED (C1 + C2). FLAT collapses both cycles
  into a single step_type="SAS" row per (numero, indice). There is no cycle_id
  column in FLAT. Consequence: every RAW multi-cycle SAS pair is a
  SAS_CYCLE_COLLAPSED candidate in the §5 explanation taxonomy.

OQ-2  FLAT cycle index:
  GED_OPERATIONS.step_order carries the sequential step position within
  a (numero, indice) workflow — NOT a cycle label. It is stored in the
  cycle_id column of this trace (per §5: "step_order for FLAT"). RAW
  trace stores "C1"/"C2" in cycle_id; FLAT trace stores step_order
  numerics. The fields are structurally compatible but semantically
  different — 8B.3+ reconciliation must not equate them.

OQ-3  Actor canonical map source:
  Re-implemented INLINE below — identical to extract_raw_ged_trace.py.
  src/flat_ged/* is NEVER imported from this script, keeping the trace
  independent of production code.
  Mapping: RAW_TO_CANONICAL.get(actor_raw, actor_raw).

OQ-4  status_clean normalisation:
  Same inline implementation as extract_raw_ged_trace.py:
    strip() → if empty/none/nan → None; lstrip('.') → strip() → upper()
  Applied to status_raw values from both sheets.

OQ-5  _MISSING_ synthetic keys:
  746 RAW rows have NUMERO=None and carry synthetic _MISSING_{row} keys
  in raw_ged_trace.csv. These rows are NOT present in FLAT_GED.xlsx —
  the FLAT builder drops docs without a standard NUMERO. This script
  never emits _MISSING_ rows because FLAT simply doesn't carry them.
  Consequence for 8B.3: apply the documented _MISSING_ filter on the
  RAW side before building identity sets.

═══════════════════════════════════════════════════════════════════════════════
FLAT SHEET STRUCTURE (verified 2026-04-30)
═══════════════════════════════════════════════════════════════════════════════

GED_OPERATIONS (32099 data rows):
  step_type distribution: OPEN_DOC=4848, SAS=4848, CONSULTANT=18911, MOEX=3492
  MOEX actor_raw values: 0-/A-/B-/H-Maître d'Oeuvre EXE (all → canonical)
  SAS  actor_raw value:  "0-SAS" (all SAS rows, no cycle label)

GED_RAW_FLAT (27261 data rows):
  One row per (doc, approver) combination.

═══════════════════════════════════════════════════════════════════════════════
STEP COUNT DERIVATION FROM CSV (post-write self-check contract)
═══════════════════════════════════════════════════════════════════════════════

All step_* counters in the stdout summary are re-derived from the written CSV
after each run. If any counter disagrees, the script raises immediately.

  step_OPEN_DOC    = rows WHERE source_sheet="GED_OPERATIONS"
                      AND event_type="DOCUMENT_VERSION"
  step_SAS         = rows WHERE source_sheet="GED_OPERATIONS"
                      AND actor_raw="0-SAS"
  step_MOEX        = rows WHERE source_sheet="GED_OPERATIONS"
                      AND actor_canonical="Maître d'Oeuvre EXE"
  step_CONSULTANT  = (total GED_OPERATIONS rows)
                      - step_OPEN_DOC - step_SAS - step_MOEX
  sas_ref_rows     = rows WHERE source_sheet="GED_RAW_FLAT"
                      AND actor_raw="0-SAS"
                      AND status_clean="REF"
  NOTE: GED_OPERATIONS has 283 SAS REF rows (step_type=SAS, status_clean=REF).
  GED_RAW_FLAT has 284 (is_sas=True, response_status_clean=REF). The spec §7
  baseline of 284 uses the GED_RAW_FLAT definition. The 1-row gap is a
  GED_OPERATIONS projection artefact that 8B.5 will investigate.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
_ROOT     = Path(__file__).resolve().parent.parent
_FLAT     = _ROOT / "output" / "intermediate" / "FLAT_GED.xlsx"
_OUT_DIR  = _ROOT / "output" / "debug"
_OUT_CSV  = _OUT_DIR / "flat_ged_trace.csv"
_OUT_XLSX = _OUT_DIR / "flat_ged_trace.xlsx"

# ── Inline actor canonical map ─────────────────────────────────────────────────
# IDENTICAL to extract_raw_ged_trace.py. Re-implemented here so both traces
# use the same map — RAW and FLAT actor_canonical values are directly comparable.
# src/flat_ged/* is NEVER imported from this script.

_CONSULTANT_ROLES: list[str] = [
    "AMO HQE", "ARCHITECTE", "BET Acoustique", "BET Ascenseur",
    "BET CVC", "BET Electricité", "BET EV", "BET Façade",
    "BET Plomberie", "BET POL", "BET SPK", "BET Structure",
    "BET VRD", "Bureau de Contrôle", "Maître d'Oeuvre EXE",
]

RAW_TO_CANONICAL: dict[str, str] = {}
for _pfx in ("0-", "A-", "B-", "H-"):
    for _role in _CONSULTANT_ROLES:
        RAW_TO_CANONICAL[f"{_pfx}{_role}"] = _role

RAW_TO_CANONICAL["0-SAS"] = "0-SAS"

_EXCEPTION_COLS: frozenset[str] = frozenset({
    "0-BET Géotech", "0-BET Synthèse", "0-BIM Manager", "0-CSPS",
    "A05-MNS EXT", "A06-REVET FAC", "A07-CSQ PREFA", "A08-MR",
    "A22-SDB Préfa", "A31-33-34-ELEC", "A41-CVC", "A42 PLB",
    "B05-MNS EXT", "B06 - REVÊTEMENT EXT", "B13 - METALLERIE SERRURERIE",
    "B31-33-34-CFO-CFA", "B35-GTB", "B41-CVC", "B42 PLB",
    "H05-MNS EXT", "H06-REVET FAC", "H07-CSQ PREFA", "H08-MUR RIDEAUX",
    "H31-33-34-CFO-CFA", "H35-GTB", "H41-CVC", "H42 PLB", "H51-ASC",
    "00-TCE", "01-TERRASSEMENTS", "02-FONDATIONS SPECIALES", "03-GOE",
    "08-MURS RIDEAUX", "35-GTB", "41-CVC", "42-PLB",
    "Sollicitation supplémentaire",
})
for _ec in _EXCEPTION_COLS:
    RAW_TO_CANONICAL[_ec] = "Exception List"

del _pfx, _role, _ec


# ── Inline status_clean (from normalize.clean_status) ─────────────────────────
# IDENTICAL to extract_raw_ged_trace.py.

def _clean_status(raw) -> str | None:
    """Strip leading dots and normalise; returns None for blank/none/nan."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s.lower() in ("", "none", "nan"):
        return None
    s = s.lstrip(".")
    s = s.strip()
    return s.upper() if s else None


# ── CSV columns — IDENTICAL to extract_raw_ged_trace._COLS ────────────────────
# Do NOT reorder or rename. 8B.3+ joins on these columns.

_COLS: list[str] = [
    "source_file", "source_sheet", "source_excel_row",
    "numero", "indice", "lot", "emetteur", "titre",
    "cycle_id", "actor_raw", "actor_canonical",
    "event_type",
    "status_raw", "status_clean",
    "response_date", "submission_date", "deadline_date", "date_status_type",
    "comment_raw", "pj_flag",
]

_MOEX_CANONICAL = "Maître d'Oeuvre EXE"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _str_numero(v) -> str | None:
    """Convert a numero cell to string, matching raw trace logic."""
    if v is None:
        return None
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v).strip() or None


def _to_str(v) -> str | None:
    """Coerce cell value to stripped string; None for blank/none/nan."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("none", "nan") else None


def _hmap(header_row) -> dict[str, int]:
    return {str(h): i for i, h in enumerate(header_row) if h is not None}


def _get(row_vals: tuple, hmap: dict, col: str):
    idx = hmap.get(col)
    return row_vals[idx] if idx is not None and idx < len(row_vals) else None


# ── Main extraction ────────────────────────────────────────────────────────────

def extract(
    flat_path: Path = _FLAT,
    out_csv: Path = _OUT_CSV,
    out_xlsx: Path = _OUT_XLSX,
) -> dict:
    """
    Extract FLAT GED events from GED_OPERATIONS and GED_RAW_FLAT.

    Returns summary dict:
      total_rows, unique_numero, unique_numero_indice,
      step_OPEN_DOC, step_SAS, step_CONSULTANT, step_MOEX, sas_ref_rows
    """
    _OUT_DIR.mkdir(parents=True, exist_ok=True)

    # read_only=True is safe here: FLAT_GED.xlsx is a data-only file (no formulas).
    wb = openpyxl.load_workbook(str(flat_path), data_only=True, read_only=True)

    total_rows = 0
    all_numeros: set[str] = set()
    all_pairs:   set[tuple] = set()
    step_counts  = {"OPEN_DOC": 0, "SAS": 0, "CONSULTANT": 0, "MOEX": 0}
    sas_ref_rows = 0

    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_COLS)
        writer.writeheader()

        # ── GED_OPERATIONS ────────────────────────────────────────────────────
        ws_ops   = wb["GED_OPERATIONS"]
        ops_iter = ws_ops.iter_rows(values_only=True)
        ops_h    = _hmap(next(ops_iter))  # consume header row

        for excel_row, row_vals in enumerate(ops_iter, 2):  # data rows start at Excel row 2
            numero = _str_numero(_get(row_vals, ops_h, "numero"))
            if numero is None:
                continue

            indice      = _to_str(_get(row_vals, ops_h, "indice")) or ""
            lot         = _to_str(_get(row_vals, ops_h, "lot")) or ""
            emetteur    = _to_str(_get(row_vals, ops_h, "emetteur")) or ""
            titre       = _to_str(_get(row_vals, ops_h, "titre")) or ""
            step_type   = _to_str(_get(row_vals, ops_h, "step_type")) or ""
            step_order  = _get(row_vals, ops_h, "step_order")
            actor_raw   = _to_str(_get(row_vals, ops_h, "actor_raw"))
            actor_clean = _to_str(_get(row_vals, ops_h, "actor_clean"))
            actor_canonical = RAW_TO_CANONICAL.get(
                actor_raw or "",
                actor_clean or (actor_raw or ""),
            )
            status_raw_v   = _to_str(_get(row_vals, ops_h, "status_raw"))
            status_clean_v = _clean_status(_get(row_vals, ops_h, "status_clean"))
            response_date_v  = _to_str(_get(row_vals, ops_h, "response_date"))
            submittal_date_v = _to_str(_get(row_vals, ops_h, "submittal_date"))
            global_dl        = _to_str(_get(row_vals, ops_h, "global_deadline"))
            phase_dl         = _to_str(_get(row_vals, ops_h, "phase_deadline"))
            deadline_v       = global_dl or phase_dl
            observation_v    = _to_str(_get(row_vals, ops_h, "observation"))
            pj_flag_v        = _get(row_vals, ops_h, "pj_flag")

            # OPEN_DOC rows → DOCUMENT_VERSION regardless of status_clean
            if step_type == "OPEN_DOC":
                event_type = "DOCUMENT_VERSION"
            elif status_clean_v is not None:
                event_type = "RESPONSE"
            else:
                event_type = "ACTOR_CALLED"

            all_numeros.add(numero)
            all_pairs.add((numero, indice))

            if step_type in step_counts:
                step_counts[step_type] += 1

            writer.writerow({
                "source_file":      "FLAT_GED.xlsx",
                "source_sheet":     "GED_OPERATIONS",
                "source_excel_row": excel_row,
                "numero":      numero,
                "indice":      indice,
                "lot":         lot,
                "emetteur":    emetteur,
                "titre":       titre,
                "cycle_id":    str(step_order) if step_order is not None else None,
                "actor_raw":        actor_raw,
                "actor_canonical":  actor_canonical,
                "event_type":       event_type,
                "status_raw":       status_raw_v,
                "status_clean":     status_clean_v,
                "response_date":    response_date_v,
                "submission_date":  submittal_date_v,
                "deadline_date":    deadline_v,
                "date_status_type": None,   # field not present in GED_OPERATIONS
                "comment_raw":      observation_v,
                "pj_flag":          1 if pj_flag_v else 0,
            })
            total_rows += 1

        # ── GED_RAW_FLAT ──────────────────────────────────────────────────────
        ws_flat   = wb["GED_RAW_FLAT"]
        flat_iter = ws_flat.iter_rows(values_only=True)
        flat_h    = _hmap(next(flat_iter))

        for excel_row, row_vals in enumerate(flat_iter, 2):
            numero = _str_numero(_get(row_vals, flat_h, "numero"))
            if numero is None:
                continue

            indice      = _to_str(_get(row_vals, flat_h, "indice")) or ""
            lot         = _to_str(_get(row_vals, flat_h, "lot")) or ""
            emetteur    = _to_str(_get(row_vals, flat_h, "emetteur")) or ""
            titre       = _to_str(_get(row_vals, flat_h, "titre")) or ""
            approver_raw       = _to_str(_get(row_vals, flat_h, "approver_raw"))
            approver_canonical = _to_str(_get(row_vals, flat_h, "approver_canonical"))
            actor_canonical    = RAW_TO_CANONICAL.get(
                approver_raw or "",
                approver_canonical or (approver_raw or ""),
            )
            status_raw_v   = _to_str(_get(row_vals, flat_h, "response_status_raw"))
            status_clean_v = _clean_status(_get(row_vals, flat_h, "response_status_clean"))
            response_date_v      = _to_str(_get(row_vals, flat_h, "response_date"))
            deadline_v           = _to_str(_get(row_vals, flat_h, "deadline"))
            date_status_type_v   = _to_str(_get(row_vals, flat_h, "date_status_type"))
            commentaire_v        = _to_str(_get(row_vals, flat_h, "commentaire"))
            pj_flag_v            = _get(row_vals, flat_h, "pj_flag")

            event_type = "RESPONSE" if status_clean_v is not None else "ACTOR_CALLED"

            all_numeros.add(numero)
            all_pairs.add((numero, indice))

            # sas_ref baseline (284) counts from GED_RAW_FLAT: is_sas=True AND REF
            if approver_raw == "0-SAS" and status_clean_v == "REF":
                sas_ref_rows += 1

            writer.writerow({
                "source_file":      "FLAT_GED.xlsx",
                "source_sheet":     "GED_RAW_FLAT",
                "source_excel_row": excel_row,
                "numero":      numero,
                "indice":      indice,
                "lot":         lot,
                "emetteur":    emetteur,
                "titre":       titre,
                "cycle_id":    None,         # no cycle info in GED_RAW_FLAT
                "actor_raw":        approver_raw,
                "actor_canonical":  actor_canonical,
                "event_type":       event_type,
                "status_raw":       status_raw_v,
                "status_clean":     status_clean_v,
                "response_date":    response_date_v,
                "submission_date":  None,    # not in GED_RAW_FLAT
                "deadline_date":    deadline_v,
                "date_status_type": date_status_type_v,
                "comment_raw":      commentaire_v,
                "pj_flag":          1 if pj_flag_v else 0,
            })
            total_rows += 1

    wb.close()

    # ── Post-write self-check ──────────────────────────────────────────────────
    # Re-read the CSV and recompute every summary counter from it.
    # Raises AssertionError immediately if any counter disagrees.
    _verify_csv(out_csv, total_rows, all_numeros, all_pairs, step_counts, sas_ref_rows)

    # ── Write XLSX ─────────────────────────────────────────────────────────────
    _write_xlsx(out_csv, out_xlsx)

    return {
        "total_rows":           total_rows,
        "unique_numero":        len(all_numeros),
        "unique_numero_indice": len(all_pairs),
        "step_OPEN_DOC":        step_counts["OPEN_DOC"],
        "step_SAS":             step_counts["SAS"],
        "step_CONSULTANT":      step_counts["CONSULTANT"],
        "step_MOEX":            step_counts["MOEX"],
        "sas_ref_rows":         sas_ref_rows,
    }


def _verify_csv(
    csv_path: Path,
    expected_total: int,
    expected_numeros: set,
    expected_pairs: set,
    expected_steps: dict,
    expected_sas_ref: int,
) -> None:
    """Re-read the written CSV and assert every counter matches in-memory values."""
    rows = list(csv.DictReader(csv_path.open(encoding="utf-8")))

    if len(rows) != expected_total:
        raise AssertionError(
            f"CSV row count mismatch: wrote {expected_total}, read back {len(rows)}"
        )

    csv_numeros = {r["numero"] for r in rows if r.get("numero")}
    if csv_numeros != expected_numeros:
        raise AssertionError(
            f"unique_numero mismatch: in-memory={len(expected_numeros)} "
            f"csv={len(csv_numeros)}"
        )

    csv_pairs = {(r["numero"], r["indice"]) for r in rows if r.get("numero")}
    if csv_pairs != expected_pairs:
        raise AssertionError(
            f"unique_numero_indice mismatch: in-memory={len(expected_pairs)} "
            f"csv={len(csv_pairs)}"
        )

    ops = [r for r in rows if r["source_sheet"] == "GED_OPERATIONS"]

    csv_open_doc = sum(1 for r in ops if r["event_type"] == "DOCUMENT_VERSION")
    if csv_open_doc != expected_steps["OPEN_DOC"]:
        raise AssertionError(
            f"step_OPEN_DOC mismatch: in-memory={expected_steps['OPEN_DOC']} "
            f"csv={csv_open_doc}"
        )

    csv_sas = sum(1 for r in ops if r["actor_raw"] == "0-SAS")
    if csv_sas != expected_steps["SAS"]:
        raise AssertionError(
            f"step_SAS mismatch: in-memory={expected_steps['SAS']} "
            f"csv={csv_sas}"
        )

    csv_moex = sum(1 for r in ops if r["actor_canonical"] == _MOEX_CANONICAL)
    if csv_moex != expected_steps["MOEX"]:
        raise AssertionError(
            f"step_MOEX mismatch: in-memory={expected_steps['MOEX']} "
            f"csv={csv_moex}"
        )

    csv_consultant = len(ops) - csv_open_doc - csv_sas - csv_moex
    if csv_consultant != expected_steps["CONSULTANT"]:
        raise AssertionError(
            f"step_CONSULTANT mismatch: in-memory={expected_steps['CONSULTANT']} "
            f"csv={csv_consultant}"
        )

    # sas_ref baseline comes from GED_RAW_FLAT (284), not GED_OPERATIONS (283).
    flat_rows = [r for r in rows if r["source_sheet"] == "GED_RAW_FLAT"]
    csv_sas_ref = sum(
        1 for r in flat_rows
        if r["actor_raw"] == "0-SAS"
        and r["status_clean"] == "REF"
    )
    if csv_sas_ref != expected_sas_ref:
        raise AssertionError(
            f"sas_ref_rows mismatch: in-memory={expected_sas_ref} "
            f"csv={csv_sas_ref}"
        )


def _write_xlsx(csv_path: Path, out_xlsx: Path) -> None:
    """Write trace rows to xlsx. Reads the already-validated CSV."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "flat_ged_trace"
    ws.append(_COLS)
    with csv_path.open(encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            ws.append([row.get(c, "") for c in _COLS])
    wb.save(str(out_xlsx))


# ── Entry point ────────────────────────────────────────────────────────────────

_BASELINES = {
    "unique_numero":        2819,
    "unique_numero_indice": 4848,
    "step_OPEN_DOC":        4848,
    "step_SAS":             4848,
    "step_CONSULTANT":      18911,
    "step_MOEX":            3492,
    "sas_ref_rows":         284,
}

if __name__ == "__main__":
    summary = extract()
    print(
        f"FLAT_TRACE: rows={summary['total_rows']} "
        f"unique_numero={summary['unique_numero']} "
        f"unique_numero_indice={summary['unique_numero_indice']} "
        f"step_OPEN_DOC={summary['step_OPEN_DOC']} "
        f"step_SAS={summary['step_SAS']} "
        f"step_CONSULTANT={summary['step_CONSULTANT']} "
        f"step_MOEX={summary['step_MOEX']} "
        f"sas_ref_rows={summary['sas_ref_rows']}"
    )

    drifted = [
        f"{k}: expected={_BASELINES[k]} actual={summary[k]}"
        for k in _BASELINES
        if summary[k] != _BASELINES[k]
    ]
    if drifted:
        raise SystemExit(
            "BASELINE DRIFT — STOP and report; do not adjust baselines:\n"
            + "\n".join(f"  {d}" for d in drifted)
        )
