"""Find the single INVALID_REASON entry in sheet 10 for the receipt."""
from openpyxl import load_workbook

wb = load_workbook(r"output\debug\raw_flat_reconcile.xlsx", read_only=True)
ws = wb["10_EXISTING_REASON_AUDIT"]
header = next(ws.iter_rows(values_only=True))
col = {h: i for i, h in enumerate(header)}
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[col["classification"]] == "INVALID_REASON":
        print("INVALID_REASON entry:")
        for c in header:
            print(f"  {c}: {r[col[c]]}")
wb.close()
