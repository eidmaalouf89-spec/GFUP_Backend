"""Step 0.1 verification — sheets 08, 09 row counts and verdict breakdown."""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook(r"output\debug\raw_flat_reconcile.xlsx", read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    n = sum(1 for _ in ws.iter_rows())
    print(f"{name}: {n} rows")

ws = wb["08_DATE_DIFFS"]
header = next(ws.iter_rows(values_only=True))
col = {h: i for i, h in enumerate(header)}
rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"08_verdicts={dict(Counter(r[col['verdict']] for r in rows))}")

ws = wb["09_COMMENT_DIFFS"]
header = next(ws.iter_rows(values_only=True))
col = {h: i for i, h in enumerate(header)}
rows = list(ws.iter_rows(min_row=2, values_only=True))
print(f"09_verdicts={dict(Counter(r[col['verdict']] for r in rows))}")

wb.close()
