"""Inspect approver_canonical in GED_RAW_FLAT for the four target consultants."""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook(r"output\intermediate\FLAT_GED.xlsx", read_only=True)
ws = wb["GED_RAW_FLAT"]
header = next(ws.iter_rows(values_only=True))
i = header.index("approver_canonical")
counts = Counter(r[i] for r in ws.iter_rows(min_row=2, values_only=True))
wb.close()
print("GED_RAW_FLAT.approver_canonical top-25:")
for a, n in sorted(counts.items(), key=lambda x: -x[1])[:25]:
    print(f"  {n:5d}  {a}")
