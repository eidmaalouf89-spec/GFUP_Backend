"""Step 2 trust-but-verify — fresh Windows process check on report_to_flat_trace artifacts."""
import json, os, zipfile
from openpyxl import load_workbook

px = r"output\debug\report_to_flat_trace.xlsx"
pj = r"output\debug\report_to_flat_trace.json"

print(f"xlsx_size={os.path.getsize(px)}")
print(f"json_size={os.path.getsize(pj)}")

with zipfile.ZipFile(px) as z:
    bad = z.testzip()
print(f"xlsx_zip_integrity={'OK' if bad is None else f'CORRUPT: {bad}'}")

wb = load_workbook(px, read_only=True, data_only=True)
print(f"xlsx_sheets={wb.sheetnames}")
for n in wb.sheetnames:
    ws = wb[n]
    rc = sum(1 for _ in ws.iter_rows())
    print(f"  {n}: {rc} rows")
wb.close()

j = json.loads(open(pj, "r", encoding="utf-8").read())
print(f"\njson.report_count={j['report_count']}")
print(f"json.counts_match_type={j['counts_match_type']}")
print(f"json.counts_applied={j['counts_applied']}")
print(f"json.counts_confidence={j['counts_confidence']}")
print(f"json.counts_eff_source={j['counts_eff_source']}")
expected_keys = {"report_count","counts_match_type","counts_applied",
                 "counts_confidence","counts_eff_source","rows"}
missing = expected_keys - set(j.keys())
print(f"json.key_set_match={'OK' if not missing else f'MISSING: {missing}'}")
