"""Peek consultant_match_report.xlsx for sample rows."""
from openpyxl import load_workbook
wb = load_workbook(r"runs\run_0000\consultant_match_report.xlsx", read_only=True)
ws = wb["MATCHED"]
header = next(ws.iter_rows(values_only=True))
keep = ["Matched GED doc_id", "Matched NUMERO", "Matched INDICE",
        "Consultant Source", "Match Method", "Confidence", "Indice Fallback"]
ix = {h: header.index(h) for h in keep if h in header}
print("First 5 MATCHED rows:")
for r in list(ws.iter_rows(min_row=2, values_only=True))[:5]:
    print({k: r[i] for k, i in ix.items()})

# Distinct consultants in match report
ws2 = wb["MATCHED"]
hdr = next(ws2.iter_rows(values_only=True))
i_cons = hdr.index("Consultant Source")
distinct = sorted({r[i_cons] for r in ws2.iter_rows(min_row=2, values_only=True) if r[i_cons]})
print(f"\nMATCHED distinct Consultant Source: {distinct}")

i_method = hdr.index("Match Method")
i_conf = hdr.index("Confidence")
from collections import Counter
methods = Counter(r[i_method] for r in ws2.iter_rows(min_row=2, values_only=True))
confs = Counter(r[i_conf] for r in ws2.iter_rows(min_row=2, values_only=True))
print(f"\nMATCHED Match Method distribution: {dict(methods)}")
print(f"MATCHED Confidence distribution: {dict(confs)}")
wb.close()
