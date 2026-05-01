"""Inspect actor_clean values used in FLAT GED_OPERATIONS for join with persisted reports."""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook(r"output\intermediate\FLAT_GED.xlsx", read_only=True)
ws = wb["GED_OPERATIONS"]
header = next(ws.iter_rows(values_only=True))
i_actor = header.index("actor_clean")
i_step = header.index("step_type")
actors = Counter()
for r in ws.iter_rows(min_row=2, values_only=True):
    actors[r[i_actor]] += 1
wb.close()

# Top 20 actor_clean values
print("FLAT_GED.GED_OPERATIONS actor_clean distribution (top 25):")
for a, n in sorted(actors.items(), key=lambda x: -x[1])[:25]:
    print(f"  {n:5d}  {a}")
