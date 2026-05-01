"""Peek FLAT_GED.xlsx column names — does it carry doc_id?"""
from openpyxl import load_workbook
wb = load_workbook(r"output\intermediate\FLAT_GED.xlsx", read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    header = next(ws.iter_rows(values_only=True))
    print(f"{name} ({sum(1 for _ in ws.iter_rows())} rows):")
    for i, h in enumerate(header):
        print(f"  [{i}] {h}")
    print()
wb.close()
