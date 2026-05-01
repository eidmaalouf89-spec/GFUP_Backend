"""Step 1 trust-but-verify — sheet 10 row count + classification distribution + top-5 INVALID/AMBIGUOUS."""
import os, zipfile
from collections import Counter
from openpyxl import load_workbook

p = r"output\debug\raw_flat_reconcile.xlsx"
print(f"file_size_bytes={os.path.getsize(p)}")
with zipfile.ZipFile(p) as z:
    bad = z.testzip()
print(f"zip_integrity={'OK' if bad is None else f'CORRUPT: {bad}'}")

wb = load_workbook(p, read_only=True, data_only=True)
print(f"sheets={wb.sheetnames}")
for n in wb.sheetnames:
    ws = wb[n]
    rc = sum(1 for _ in ws.iter_rows())
    print(f"  {n}: {rc} rows")

ws = wb["10_EXISTING_REASON_AUDIT"]
header = next(ws.iter_rows(values_only=True))
col = {h: i for i, h in enumerate(header)}
rows = list(ws.iter_rows(min_row=2, values_only=True))
counts = Counter(r[col["classification"]] for r in rows)
print(f"\n10_classifications={dict(counts)}")

print("\nTop 5 INVALID + AMBIGUOUS findings (file:line function classification reason):")
shown = 0
for r in rows:
    if r[col["classification"]] in ("INVALID_REASON", "AMBIGUOUS_REASON"):
        print(f"  {r[col['file']]}:{r[col['line']]} {r[col['function']]} "
              f"[{r[col['classification']]}] {r[col['reason_string']][:80]!r}")
        shown += 1
        if shown >= 5:
            break

wb.close()
