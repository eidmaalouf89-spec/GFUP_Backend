"""Final trust-but-verify pass for all artifacts produced in Phase 8B finalization.

Runs from a fresh Windows-shell Python subprocess (H-1 mitigation).
"""
import csv, json, os, zipfile
from openpyxl import load_workbook
from collections import Counter

ARTIFACTS = [
    ("output/debug/raw_flat_reconcile.xlsx", "xlsx"),
    ("output/debug/report_to_flat_trace.xlsx", "xlsx"),
    ("output/debug/report_to_flat_trace.json", "json"),
    ("output/debug/SHADOW_FLAT_GED_TRACE.xlsx", "xlsx"),
    ("output/debug/SHADOW_FLAT_GED_OPERATIONS.csv", "csv"),
]

for path, kind in ARTIFACTS:
    print(f"\n=== {path} ===")
    size = os.path.getsize(path)
    print(f"size: {size}")
    if kind == "xlsx":
        with zipfile.ZipFile(path) as z:
            bad = z.testzip()
        print(f"zip_integrity: {'OK' if bad is None else f'CORRUPT: {bad}'}")
        wb = load_workbook(path, read_only=True, data_only=True)
        print(f"sheets: {wb.sheetnames}")
        for n in wb.sheetnames:
            rc = sum(1 for _ in wb[n].iter_rows())
            print(f"  {n}: {rc} rows")
        wb.close()
    elif kind == "json":
        j = json.loads(open(path, "r", encoding="utf-8").read())
        keys = sorted(j.keys()) if isinstance(j, dict) else "<not a dict>"
        print(f"top-level keys: {keys}")
        if isinstance(j, dict) and "report_count" in j:
            print(f"  report_count: {j['report_count']}")
            print(f"  counts_match_type: {j['counts_match_type']}")
            print(f"  counts_applied: {j['counts_applied']}")
    elif kind == "csv":
        with open(path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        print(f"csv_rows: {len(rows)}")
        print(f"csv_columns: {reader.fieldnames}")
        if rows:
            shadow_rules = Counter(r.get("shadow_rule_kept", "") for r in rows)
            print(f"  shadow_rule_kept distribution: {dict(shadow_rules)}")
