"""
scripts/report_to_flat_trace.py — Phase 8B.8 report integration trace.

For every persisted report response in data/report_memory.db, compute the
fields specified in PHASE_8B_RAW_TO_FLAT_RECONCILIATION.md §15.3:

    report_doc_key       (numero, indice) — resolved via consultant_match_report
    report_actor         the persisted DB consultant string
    report_status        the persisted DB report_status (VAO, REF, ...)
    report_comment       the persisted DB report_comment
    report_confidence    HIGH / MEDIUM / LOW / UNKNOWN  (E2 thresholds)
    report_match_type    EXACT / FAMILY / FUZZY / NO_MATCH
                          (mapped from the persisted match_method)
    report_applied       APPLIED_AS_PRIMARY    — GED was PENDING, eligible report would upgrade
                          APPLIED_AS_ENRICHMENT — GED was ANSWERED, report adds comment
                          BLOCKED_BY_GED        — GED was ANSWERED with conflicting status
                          BLOCKED_BY_CONFIDENCE — confidence is LOW or UNKNOWN
                          NOT_APPLIED           — no GED row exists at (numero, indice, actor)
    effective_status     post-merge predicted status
    effective_source     GED / REPORT / GED+REPORT (per spec)

The trace is OFFLINE: it does not call build_effective_responses, instead it
re-implements the decision rules at the trace level so the result is
auditable line-by-line and independent of the production code under audit.

Inputs (read only):
    data/report_memory.db
    runs/run_0000/consultant_match_report.xlsx   (doc_id ↔ numero/indice bridge)
    output/intermediate/FLAT_GED.xlsx            (GED_OPERATIONS for join)

Outputs:
    output/debug/report_to_flat_trace.xlsx
    output/debug/report_to_flat_trace.json

Stdout summary line — exact format (per §15.4):
    REPORT_TO_FLAT: report_count=<n> exact=<n> family=<n> fuzzy=<n> \\
                    no_match=<n> applied_primary=<n> applied_enrichment=<n> \\
                    blocked_ged=<n> blocked_confidence=<n> not_applied=<n>
"""
from __future__ import annotations

import json
import sqlite3
import subprocess
import zipfile
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
REPORT_DB = REPO_ROOT / "data" / "report_memory.db"
MATCH_XLSX = REPO_ROOT / "runs" / "run_0000" / "consultant_match_report.xlsx"
FLAT_GED_XLSX = REPO_ROOT / "output" / "intermediate" / "FLAT_GED.xlsx"
OUT_XLSX = REPO_ROOT / "output" / "debug" / "report_to_flat_trace.xlsx"
OUT_JSON = REPO_ROOT / "output" / "debug" / "report_to_flat_trace.json"


# ── Mappings ──────────────────────────────────────────────────────────────────

# Consultant Source label (consultant_match_report) → GED_RAW_FLAT.approver_canonical
# (the join target; GED_OPERATIONS.actor_clean uses different display names but
#  the same underlying canonical actor — we key on approver_canonical because
#  GED_RAW_FLAT carries date_status_type directly).
#
# Mapping derived by inspection of approver_canonical distribution + the
# four ingested consultant families:
#   AVLS       (acousticien)         → "BET Acoustique"
#   LE_SOMMER  (AMO HQE Le Sommer)   → "AMO HQE"
#   SOCOTEC    (bureau de contrôle)  → "Bureau de Contrôle"
#   TERRELL    (BET structure)       → "BET Structure"
_CONSULTANT_SOURCE_TO_ACTOR = {
    "AVLS":      "BET Acoustique",
    "LE_SOMMER": "AMO HQE",
    "SOCOTEC":   "Bureau de Contrôle",
    "TERRELL":   "BET Structure",
}

# Match method → match_type (per spec §15.3)
_MATCH_METHOD_TO_TYPE = {
    "MATCH_BY_NUMERO_INDICE":          "EXACT",
    "MATCH_BY_RECENT_INDICE_FALLBACK": "FAMILY",
    "MATCH_BY_DATE_PROXIMITY":         "FUZZY",
    "MATCH_BY_MIXED_HEURISTIC":        "FUZZY",
}

# Operational status families (used to predict conflict between GED and report)
_FAVORABLE   = {"VAO", "VSO", "FAV", "VAOB"}
_REFUSED     = {"REF"}
_OTHER_OK    = {"SUS", "DEF", "HM"}

# date_status_type values used in GED_OPERATIONS
_PENDING_DST = {"PENDING_IN_DELAY", "PENDING_LATE"}
_ANSWERED_DST = {"ANSWERED"}


def _confidence_bucket(value) -> str:
    """Match the E2 threshold used by effective_responses._confidence_eligible.

    HIGH   ≥ 0.95
    MEDIUM 0.75 – 0.95   (E2 passes)
    LOW    < 0.75        (E2 fails)
    UNKNOWN value is None / non-numeric  (E2 fails)
    """
    if value is None:
        return "UNKNOWN"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "UNKNOWN"
    if v >= 0.95:
        return "HIGH"
    if v >= 0.75:
        return "MEDIUM"
    return "LOW"


def _is_conflict(ged_status: str, report_status: str) -> bool:
    """Approximation of effective_responses._is_conflict — enough to predict
    BLOCKED_BY_GED at the trace level.  REF vs favorable is the canonical
    conflict; mixed favorable variants are not conflicts.
    """
    g = (ged_status or "").upper()
    r = (report_status or "").upper()
    if not g or not r:
        return False
    if g in _REFUSED and r in (_FAVORABLE | _OTHER_OK):
        return True
    if r in _REFUSED and g in (_FAVORABLE | _OTHER_OK):
        return True
    return False


# ── Loaders ───────────────────────────────────────────────────────────────────

def _load_persisted_responses(db_path: Path) -> list[dict]:
    """Read every active row from persisted_report_responses (read-only)."""
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    cur = con.execute("""
        SELECT id, consultant, doc_id, report_status, report_response_date,
               report_comment, source_filename, source_file_hash, ingested_at,
               match_confidence, match_method
          FROM persisted_report_responses
         WHERE is_active = 1
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    con.close()
    return rows


def _load_doc_id_bridge(match_xlsx: Path) -> dict[str, dict]:
    """Build doc_id → {numero, indice, consultant_source, match_method, confidence}.

    Reads MATCHED + AMBIGUOUS sheets from consultant_match_report.xlsx.
    UNMATCHED has no doc_id, so it does not contribute to this map.
    """
    wb = load_workbook(match_xlsx, read_only=True, data_only=True)
    bridge: dict[str, dict] = {}
    for sheet_name in ("MATCHED", "AMBIGUOUS"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = next(ws.iter_rows(values_only=True))
        ix = {h: i for i, h in enumerate(header)}
        i_doc  = ix.get("Matched GED doc_id")
        i_num  = ix.get("Matched NUMERO")
        i_ind  = ix.get("Matched INDICE")
        i_src  = ix.get("Consultant Source")
        i_meth = ix.get("Match Method")
        i_conf = ix.get("Confidence")
        if i_doc is None:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            doc_id = r[i_doc]
            if not doc_id:
                continue
            bridge[str(doc_id)] = {
                "numero":            r[i_num] if i_num is not None else "",
                "indice":            r[i_ind] if i_ind is not None else "",
                "consultant_source": r[i_src] if i_src is not None else "",
                "match_method_xlsx": r[i_meth] if i_meth is not None else "",
                "confidence_xlsx":   r[i_conf] if i_conf is not None else "",
            }
    wb.close()
    return bridge


def _load_ged_operations_lookup(flat_ged_xlsx: Path) -> dict[tuple, dict]:
    """Build (numero, indice, approver_canonical) → {date_status_type, status_clean}.

    Reads GED_RAW_FLAT (which has date_status_type and approver_canonical
    directly) — GED_OPERATIONS uses different actor naming and lacks the
    date_status_type field.

    When multiple cycles exist for the same key, keep the most
    operationally-significant: ANSWERED > PENDING > NOT_CALLED.
    """
    wb = load_workbook(flat_ged_xlsx, read_only=True, data_only=True)
    ws = wb["GED_RAW_FLAT"]
    header = next(ws.iter_rows(values_only=True))
    ix = {h: i for i, h in enumerate(header)}

    priority = {"ANSWERED": 3, "PENDING_LATE": 2, "PENDING_IN_DELAY": 2, "NOT_CALLED": 1}
    lookup: dict[tuple, dict] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        num   = str(r[ix["numero"]]).strip()
        ind   = str(r[ix["indice"]]).strip()
        actor = (r[ix["approver_canonical"]] or "").strip()
        dst   = (r[ix["date_status_type"]] or "").strip()
        sc    = (r[ix["response_status_clean"]] or "").strip()
        if not num or not actor:
            continue
        key = (num, ind, actor)
        existing = lookup.get(key)
        new_priority = priority.get(dst, 0)
        if existing is None or new_priority > priority.get(existing["date_status_type"], 0):
            lookup[key] = {"date_status_type": dst, "status_clean": sc}
    wb.close()
    return lookup


# ── Per-row classifier ────────────────────────────────────────────────────────

def _classify_report_row(
    row: dict,
    bridge: dict,
    ged_lookup: dict,
) -> dict:
    """Return a single trace dict for one persisted report row."""
    doc_id = str(row["doc_id"])
    confidence = _confidence_bucket(row["match_confidence"])
    match_type = _MATCH_METHOD_TO_TYPE.get(row.get("match_method") or "", "NO_MATCH")

    info = bridge.get(doc_id, {})
    numero = info.get("numero") or ""
    indice = info.get("indice") or ""
    consultant_src = info.get("consultant_source") or ""
    actor_clean = _CONSULTANT_SOURCE_TO_ACTOR.get(consultant_src, "")

    # Default: GED unaffected
    effective_status = ""
    effective_source = "GED"

    if confidence in ("LOW", "UNKNOWN"):
        applied = "BLOCKED_BY_CONFIDENCE"
    elif not actor_clean or not numero:
        # Confidence is HIGH/MEDIUM but the doc_id can't be mapped to a current
        # FLAT GED key — the ingestion-time match doesn't survive a fresh
        # pipeline run.  Treat as NOT_APPLIED.
        applied = "NOT_APPLIED"
    else:
        ged = ged_lookup.get((str(numero).strip(), str(indice).strip(), actor_clean))
        if ged is None:
            applied = "NOT_APPLIED"
        else:
            ged_dst = ged["date_status_type"]
            ged_sc  = ged["status_clean"]
            if ged_dst in _PENDING_DST:
                applied = "APPLIED_AS_PRIMARY"
                effective_status = (row["report_status"] or "").strip() or ged_sc
                effective_source = "GED+REPORT_STATUS"
            elif ged_dst in _ANSWERED_DST:
                if _is_conflict(ged_sc, row["report_status"]):
                    applied = "BLOCKED_BY_GED"
                    effective_status = ged_sc
                    effective_source = "GED_CONFLICT_REPORT"
                else:
                    applied = "APPLIED_AS_ENRICHMENT"
                    effective_status = ged_sc
                    effective_source = "GED+REPORT_COMMENT"
            else:
                # NOT_CALLED — Rule 4: reports cannot create new rows
                applied = "NOT_APPLIED"

    return {
        "report_id":            row["id"],
        "report_doc_id":        doc_id,
        "report_numero":        numero,
        "report_indice":        indice,
        "report_actor_db":      row["consultant"],
        "report_actor_resolved": actor_clean,
        "report_status":        row["report_status"] or "",
        "report_response_date": row["report_response_date"] or "",
        "report_comment":       (row["report_comment"] or "")[:300],
        "report_confidence":    confidence,
        "report_match_type":    match_type,
        "report_match_method":  row.get("match_method") or "",
        "report_applied":       applied,
        "effective_status":     effective_status,
        "effective_source":     effective_source,
        "source_filename":      row["source_filename"],
    }


# ── Driver ────────────────────────────────────────────────────────────────────

def compute_report_to_flat_trace() -> dict:
    persisted = _load_persisted_responses(REPORT_DB)
    bridge    = _load_doc_id_bridge(MATCH_XLSX)
    ged_lookup = _load_ged_operations_lookup(FLAT_GED_XLSX)

    rows = [_classify_report_row(r, bridge, ged_lookup) for r in persisted]

    counts_match_type = Counter(r["report_match_type"] for r in rows)
    counts_applied    = Counter(r["report_applied"]    for r in rows)
    counts_confidence = Counter(r["report_confidence"] for r in rows)
    counts_eff_source = Counter(r["effective_source"]  for r in rows)

    summary_line = (
        f"REPORT_TO_FLAT: report_count={len(rows)} "
        f"exact={counts_match_type.get('EXACT', 0)} "
        f"family={counts_match_type.get('FAMILY', 0)} "
        f"fuzzy={counts_match_type.get('FUZZY', 0)} "
        f"no_match={counts_match_type.get('NO_MATCH', 0)} "
        f"applied_primary={counts_applied.get('APPLIED_AS_PRIMARY', 0)} "
        f"applied_enrichment={counts_applied.get('APPLIED_AS_ENRICHMENT', 0)} "
        f"blocked_ged={counts_applied.get('BLOCKED_BY_GED', 0)} "
        f"blocked_confidence={counts_applied.get('BLOCKED_BY_CONFIDENCE', 0)} "
        f"not_applied={counts_applied.get('NOT_APPLIED', 0)}"
    )

    return {
        "rows":              rows,
        "counts_match_type": dict(counts_match_type),
        "counts_applied":    dict(counts_applied),
        "counts_confidence": dict(counts_confidence),
        "counts_eff_source": dict(counts_eff_source),
        "summary_line":      summary_line,
    }


# ── Writers ───────────────────────────────────────────────────────────────────

_SHEET_COLS = [
    "report_id", "report_doc_id", "report_numero", "report_indice",
    "report_actor_db", "report_actor_resolved",
    "report_status", "report_response_date", "report_comment",
    "report_confidence", "report_match_type", "report_match_method",
    "report_applied", "effective_status", "effective_source",
    "source_filename",
]


def write_outputs(result: dict, xlsx_path: Path, json_path: Path) -> None:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("01_REPORT_ROWS")
    ws.append(_SHEET_COLS)
    for r in result["rows"]:
        ws.append([r.get(c, "") for c in _SHEET_COLS])

    ws2 = wb.create_sheet("02_DISTRIBUTIONS")
    ws2.append(["category", "key", "count"])
    for cat, dist in (
        ("match_type",     result["counts_match_type"]),
        ("applied",        result["counts_applied"]),
        ("confidence",     result["counts_confidence"]),
        ("effective_source", result["counts_eff_source"]),
    ):
        for k, v in sorted(dist.items()):
            ws2.append([cat, k, v])

    wb.save(xlsx_path)
    del wb

    json_payload = {
        "report_count":       len(result["rows"]),
        "counts_match_type":  result["counts_match_type"],
        "counts_applied":     result["counts_applied"],
        "counts_confidence":  result["counts_confidence"],
        "counts_eff_source":  result["counts_eff_source"],
        "rows":               result["rows"],
    }
    json_path.write_text(json.dumps(json_payload, indent=2, default=str), encoding="utf-8")

    # Trust-but-verify in a fresh subprocess (H-1 mitigation).
    check_code = f"""
import os, zipfile, json
from openpyxl import load_workbook
from collections import Counter

px = r{str(xlsx_path)!r}
pj = r{str(json_path)!r}
sx = os.path.getsize(px)
sj = os.path.getsize(pj)
assert sx > 0 and sj > 0, f"empty: xlsx={{sx}} json={{sj}}"
with zipfile.ZipFile(px) as z:
    assert z.testzip() is None, "xlsx ZIP corrupt"

wb = load_workbook(px, read_only=True, data_only=True)
names = wb.sheetnames
assert "01_REPORT_ROWS" in names and "02_DISTRIBUTIONS" in names, names
rows = list(wb["01_REPORT_ROWS"].iter_rows(values_only=True))
dist = list(wb["02_DISTRIBUTIONS"].iter_rows(values_only=True))
wb.close()

j = json.loads(open(pj, "r", encoding="utf-8").read())

print(f"SUBPROCESS_CHECK: xlsx_size={{sx}} json_size={{sj}} sheets={{names}}")
print(f"  01_data_rows={{len(rows)-1}} 02_data_rows={{len(dist)-1}}")
print(f"  json.report_count={{j['report_count']}}")
print(f"  json.counts_applied={{j['counts_applied']}}")
print(f"  json.counts_match_type={{j['counts_match_type']}}")
print(f"  ZIP_integrity=OK")
"""
    result_proc = subprocess.run(
        ["python", "-c", check_code], capture_output=True, text=True,
    )
    if result_proc.returncode != 0:
        raise RuntimeError(
            f"Subprocess self-check FAILED:\nSTDOUT: {result_proc.stdout}\nSTDERR: {result_proc.stderr}"
        )
    print(result_proc.stdout.strip())


def main() -> None:
    print(f"Loading persisted responses: {REPORT_DB}")
    print(f"Loading doc_id bridge:       {MATCH_XLSX}")
    print(f"Loading FLAT GED operations: {FLAT_GED_XLSX}")
    result = compute_report_to_flat_trace()
    print(f"Report rows traced: {len(result['rows'])}")
    write_outputs(result, OUT_XLSX, OUT_JSON)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_JSON}")
    print(result["summary_line"])


if __name__ == "__main__":
    main()
