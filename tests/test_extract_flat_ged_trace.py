"""
tests/test_extract_flat_ged_trace.py  —  Phase 8B, Step 8B.2
-------------------------------------------------------------
Unit tests for scripts/extract_flat_ged_trace.py.

All tests in the default run are mock-only and do NOT read
output/intermediate/FLAT_GED.xlsx. The live xlsx test is marked
@pytest.mark.slow so it can be skipped in sandbox (H-4/H-5) and
confirmed on Windows shell.
"""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

import sys
_SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(_SCRIPTS))

from extract_flat_ged_trace import (
    RAW_TO_CANONICAL,
    _COLS,
    _MOEX_CANONICAL,
    _clean_status,
    _str_numero,
    _to_str,
    extract,
)


# ── Column parity with raw trace ───────────────────────────────────────────────

class TestColumnParity:
    """FLAT trace columns must be identical (same names, same order) to RAW trace."""

    def test_cols_match_raw_trace_exactly(self):
        from extract_raw_ged_trace import _COLS as RAW_COLS
        assert _COLS == RAW_COLS, (
            "Column set mismatch with raw_ged_trace.csv.\n"
            f"flat: {_COLS}\nraw:  {RAW_COLS}"
        )

    def test_cols_count(self):
        assert len(_COLS) == 20

    def test_required_columns_present(self):
        required = {
            "source_file", "source_sheet", "source_excel_row",
            "numero", "indice", "cycle_id",
            "actor_raw", "actor_canonical", "event_type",
            "status_raw", "status_clean",
            "response_date", "submission_date", "deadline_date",
            "comment_raw", "pj_flag",
        }
        assert required.issubset(set(_COLS))


# ── _clean_status ──────────────────────────────────────────────────────────────

class TestCleanStatus:
    def test_none_returns_none(self):
        assert _clean_status(None) is None

    def test_empty_string_returns_none(self):
        assert _clean_status("") is None
        assert _clean_status("   ") is None

    def test_nan_returns_none(self):
        assert _clean_status("nan") is None
        assert _clean_status("NaN") is None
        assert _clean_status("none") is None

    def test_strips_leading_dot(self):
        assert _clean_status(".REF") == "REF"
        assert _clean_status("..VSO") == "VSO"

    def test_uppercases(self):
        assert _clean_status("ref") == "REF"
        assert _clean_status("Vso") == "VSO"

    def test_vso_sas(self):
        assert _clean_status("VSO-SAS") == "VSO-SAS"

    def test_opened(self):
        assert _clean_status("OPENED") == "OPENED"
        assert _clean_status("opened") == "OPENED"


# ── _str_numero ────────────────────────────────────────────────────────────────

class TestStrNumero:
    def test_integer(self):
        assert _str_numero(248000) == "248000"

    def test_float(self):
        assert _str_numero(248000.0) == "248000"

    def test_string_int(self):
        assert _str_numero("248000") == "248000"

    def test_none_returns_none(self):
        assert _str_numero(None) is None


# ── RAW_TO_CANONICAL ──────────────────────────────────────────────────────────

class TestRawToCanonical:
    def test_sas_stays_sas(self):
        assert RAW_TO_CANONICAL["0-SAS"] == "0-SAS"

    def test_moex_all_prefixes(self):
        for pfx in ("0-", "A-", "B-", "H-"):
            key = f"{pfx}Maître d'Oeuvre EXE"
            assert RAW_TO_CANONICAL[key] == _MOEX_CANONICAL, f"Failed for {key}"

    def test_consultant_all_prefixes(self):
        for pfx in ("0-", "A-", "B-", "H-"):
            assert RAW_TO_CANONICAL[f"{pfx}ARCHITECTE"] == "ARCHITECTE"

    def test_exception_col_maps_to_exception_list(self):
        assert RAW_TO_CANONICAL["0-BET Géotech"] == "Exception List"
        assert RAW_TO_CANONICAL["Sollicitation supplémentaire"] == "Exception List"

    def test_moex_canonical_constant(self):
        assert _MOEX_CANONICAL == "Maître d'Oeuvre EXE"


# ── Mock xlsx builder ──────────────────────────────────────────────────────────

def _build_mock_flat_xlsx(tmp_path: Path) -> Path:
    """
    Build a minimal FLAT_GED.xlsx with both GED_OPERATIONS and GED_RAW_FLAT sheets.

    GED_OPERATIONS rows:
      1 OPEN_DOC  (numero=100001, indice=A, actor_raw=API)
      1 SAS       (numero=100001, indice=A, actor_raw=0-SAS, status_clean=REF)
      1 SAS       (numero=100002, indice=A, actor_raw=0-SAS, status_clean=VSO)
      1 CONSULTANT (numero=100001, indice=A, actor_raw=B-ARCHITECTE, status_clean=VAO)
      1 MOEX      (numero=100001, indice=A, actor_raw=B-Maître d'Oeuvre EXE, no status)
      1 OPEN_DOC  (numero=100002, indice=A, actor_raw=API)

    GED_RAW_FLAT rows:
      2 approver rows for (100001, A)
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # ── GED_OPERATIONS ────────────────────────────────────────────────────────
    ws_ops = wb.active
    ws_ops.title = "GED_OPERATIONS"

    ops_headers = [
        "numero", "indice", "lot", "emetteur", "titre",
        "step_order", "step_type", "actor_type", "actor_raw", "actor_clean",
        "status_raw", "status_clean", "status_code", "status_scope",
        "status_family", "is_completed", "is_blocking", "requires_new_cycle",
        "submittal_date", "sas_response_date", "response_date", "data_date",
        "global_deadline", "phase_deadline", "deadline_source",
        "retard_avance_days", "retard_avance_status",
        "step_delay_days", "delay_contribution_days", "cumulative_delay_days",
        "delay_actor", "chrono_source", "observation", "pj_flag",
        "source_trace", "source_rows", "operation_rule_used",
    ]
    ws_ops.append(ops_headers)

    def ops_row(num, ind, step_order, step_type, actor_raw, actor_clean,
                status_raw=None, status_clean=None, response_date=None,
                submittal_date="2025-01-01", observation=None, pj_flag=0):
        r = [None] * len(ops_headers)
        h = {h: i for i, h in enumerate(ops_headers)}
        r[h["numero"]] = num
        r[h["indice"]] = ind
        r[h["lot"]] = "L01"
        r[h["emetteur"]] = "API"
        r[h["titre"]] = f"doc_{num}.pdf"
        r[h["step_order"]] = step_order
        r[h["step_type"]] = step_type
        r[h["actor_raw"]] = actor_raw
        r[h["actor_clean"]] = actor_clean
        r[h["status_raw"]] = status_raw
        r[h["status_clean"]] = status_clean
        r[h["response_date"]] = response_date
        r[h["submittal_date"]] = submittal_date
        r[h["observation"]] = observation
        r[h["pj_flag"]] = pj_flag
        return r

    ws_ops.append(ops_row(100001, "A", 1, "OPEN_DOC",  "API",  "API",
                          status_raw="OPENED", status_clean="OPENED"))
    ws_ops.append(ops_row(100001, "A", 2, "SAS",       "0-SAS", "SAS (GEMO)",
                          status_raw="REF", status_clean="REF",
                          response_date="2025-02-01"))
    ws_ops.append(ops_row(100002, "A", 1, "OPEN_DOC",  "API",  "API",
                          status_raw="OPENED", status_clean="OPENED"))
    ws_ops.append(ops_row(100002, "A", 2, "SAS",       "0-SAS", "SAS (GEMO)",
                          status_raw="VSO-SAS", status_clean="VSO",
                          response_date="2025-02-02"))
    ws_ops.append(ops_row(100001, "A", 3, "CONSULTANT", "B-ARCHITECTE", "Hardel + Le Bihan",
                          status_raw="VAO", status_clean="VAO",
                          response_date="2025-03-01"))
    ws_ops.append(ops_row(100001, "A", 4, "MOEX",
                          "B-Maître d'Oeuvre EXE", "GEMO"))

    # ── GED_RAW_FLAT ──────────────────────────────────────────────────────────
    ws_flat = wb.create_sheet("GED_RAW_FLAT")
    flat_headers = [
        "numero", "indice", "lot", "emetteur", "titre",
        "approver_raw", "approver_canonical", "actor_type",
        "response_status_raw", "response_status_clean",
        "response_status_code", "response_status_scope",
        "response_date_raw", "response_date", "date_status_type",
        "deadline_raw", "deadline", "commentaire", "pj_flag",
        "is_sas", "raw_trace_key",
    ]
    ws_flat.append(flat_headers)

    def flat_row(num, ind, approver_raw, approver_canonical, actor_type,
                 status_raw=None, status_clean=None, response_date=None,
                 deadline=None, date_status_type=None, commentaire=None,
                 is_sas=False, pj_flag=0):
        r = [None] * len(flat_headers)
        h = {h: i for i, h in enumerate(flat_headers)}
        r[h["numero"]] = num
        r[h["indice"]] = ind
        r[h["lot"]] = "L01"
        r[h["emetteur"]] = "API"
        r[h["titre"]] = f"doc_{num}.pdf"
        r[h["approver_raw"]] = approver_raw
        r[h["approver_canonical"]] = approver_canonical
        r[h["actor_type"]] = actor_type
        r[h["response_status_raw"]] = status_raw
        r[h["response_status_clean"]] = status_clean
        r[h["response_date"]] = response_date
        r[h["deadline"]] = deadline
        r[h["date_status_type"]] = date_status_type
        r[h["commentaire"]] = commentaire
        r[h["is_sas"]] = is_sas
        r[h["pj_flag"]] = pj_flag
        r[h["raw_trace_key"]] = f"{num}|{approver_raw}"
        return r

    ws_flat.append(flat_row(100001, "A", "0-SAS", "0-SAS", "SAS",
                            status_raw="REF", status_clean="REF",
                            response_date="2025-02-01", is_sas=True))
    ws_flat.append(flat_row(100001, "A", "B-ARCHITECTE", "ARCHITECTE", "CONSULTANT",
                            status_raw="VAO", status_clean="VAO",
                            response_date="2025-03-01"))

    out = tmp_path / "mock_flat.xlsx"
    wb.save(str(out))
    return out


# ── extract (mock-based) ───────────────────────────────────────────────────────

class TestExtractFunction:
    def test_extract_returns_dict(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        summary = extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        assert isinstance(summary, dict)
        for k in ("total_rows", "unique_numero", "unique_numero_indice",
                  "step_OPEN_DOC", "step_SAS", "step_CONSULTANT",
                  "step_MOEX", "sas_ref_rows"):
            assert k in summary, f"Missing key: {k}"

    def test_step_counts(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        s = extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        # mock has 2 OPEN_DOC, 2 SAS, 1 CONSULTANT, 1 MOEX
        assert s["step_OPEN_DOC"] == 2, s
        assert s["step_SAS"]      == 2, s
        assert s["step_CONSULTANT"] == 1, s
        assert s["step_MOEX"]     == 1, s

    def test_unique_counts(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        s = extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        # numeros: 100001, 100002 → 2
        assert s["unique_numero"] == 2, s
        # pairs: (100001,A), (100002,A) → 2
        assert s["unique_numero_indice"] == 2, s

    def test_sas_ref_counted(self, tmp_path):
        # sas_ref baseline comes from GED_RAW_FLAT (is_sas=True AND REF), not
        # GED_OPERATIONS (which may have 1 fewer due to projection). Mock
        # GED_RAW_FLAT has one 0-SAS/REF row for (100001, A).
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        s = extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        assert s["sas_ref_rows"] == 1, s

    def test_open_doc_emits_document_version(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        ops_dv = [r for r in rows
                  if r["source_sheet"] == "GED_OPERATIONS"
                  and r["event_type"] == "DOCUMENT_VERSION"]
        assert len(ops_dv) == 2  # 2 OPEN_DOC rows

    def test_sas_actor_raw_preserved(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        sas_rows = [r for r in rows
                    if r["source_sheet"] == "GED_OPERATIONS"
                    and r["actor_raw"] == "0-SAS"]
        assert len(sas_rows) == 2  # 2 SAS steps in mock

    def test_moex_actor_canonical(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        moex_rows = [r for r in rows
                     if r["source_sheet"] == "GED_OPERATIONS"
                     and r["actor_canonical"] == _MOEX_CANONICAL]
        assert len(moex_rows) == 1  # 1 MOEX step in mock

    def test_csv_column_header_matches_raw_trace(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        with out_csv.open(encoding="utf-8") as fh:
            header = next(csv.reader(fh))
        from extract_raw_ged_trace import _COLS as RAW_COLS
        assert header == RAW_COLS, (
            f"CSV header mismatch.\ngot:      {header}\nexpected: {RAW_COLS}"
        )

    def test_both_sheets_in_csv(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        sheets = {r["source_sheet"] for r in rows}
        assert "GED_OPERATIONS" in sheets
        assert "GED_RAW_FLAT" in sheets

    def test_xlsx_written(self, tmp_path):
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        assert out_xlsx.exists()
        wb = openpyxl.load_workbook(str(out_xlsx))
        assert "flat_ged_trace" in wb.sheetnames
        wb.close()

    def test_gед_raw_flat_date_status_type_preserved(self, tmp_path):
        """date_status_type from GED_RAW_FLAT must be preserved in CSV."""
        import openpyxl as _xl
        mock = _build_mock_flat_xlsx(tmp_path)
        # Add a GED_RAW_FLAT row with a non-null date_status_type
        wb = _xl.load_workbook(str(mock))
        ws = wb["GED_RAW_FLAT"]
        # peek at headers
        hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
        h = {v: i for i, v in enumerate(hdrs)}
        # Append a row with date_status_type = PENDING_LATE
        row = [None] * len(hdrs)
        row[h["numero"]] = 100003
        row[h["indice"]] = "A"
        row[h["approver_raw"]] = "0-SAS"
        row[h["approver_canonical"]] = "0-SAS"
        row[h["actor_type"]] = "SAS"
        row[h["date_status_type"]] = "PENDING_LATE"
        row[h["is_sas"]] = True
        ws.append(row)
        # Also add matching GED_OPERATIONS row so numero is not orphaned
        wb.save(str(mock))

        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        pending_rows = [r for r in rows
                        if r["date_status_type"] == "PENDING_LATE"]
        assert len(pending_rows) >= 1

    def test_ops_date_status_type_is_blank(self, tmp_path):
        """GED_OPERATIONS rows have no date_status_type — must be blank in CSV."""
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        ops_rows = [r for r in rows if r["source_sheet"] == "GED_OPERATIONS"]
        for r in ops_rows:
            assert r["date_status_type"] in ("", "None", None, ""), (
                f"Expected blank date_status_type for GED_OPERATIONS row, got {r['date_status_type']!r}"
            )

    def test_cycle_id_is_step_order(self, tmp_path):
        """GED_OPERATIONS cycle_id stores step_order value."""
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        ops_rows = [r for r in rows if r["source_sheet"] == "GED_OPERATIONS"]
        cycle_ids = {r["cycle_id"] for r in ops_rows if r["cycle_id"]}
        # mock has step_order values 1..4
        assert "1" in cycle_ids or "2" in cycle_ids

    def test_grd_raw_flat_cycle_id_is_blank(self, tmp_path):
        """GED_RAW_FLAT rows have no cycle info — cycle_id must be blank."""
        mock = _build_mock_flat_xlsx(tmp_path)
        out_csv  = tmp_path / "flat_trace.csv"
        out_xlsx = tmp_path / "flat_trace.xlsx"
        extract(flat_path=mock, out_csv=out_csv, out_xlsx=out_xlsx)
        rows = list(csv.DictReader(out_csv.open(encoding="utf-8")))
        flat_rows = [r for r in rows if r["source_sheet"] == "GED_RAW_FLAT"]
        for r in flat_rows:
            assert r["cycle_id"] in ("", "None", None, "")


import openpyxl  # needed in test body above


# ── Live xlsx smoke test (skip in sandbox due to H-5) ─────────────────────────

@pytest.mark.slow
def test_live_extraction_baselines(tmp_path):
    """Full extraction against real FLAT_GED.xlsx — run on Windows shell."""
    root     = Path(__file__).resolve().parent.parent
    flat     = root / "output" / "intermediate" / "FLAT_GED.xlsx"
    out_csv  = tmp_path / "flat_trace_live.csv"
    out_xlsx = tmp_path / "flat_trace_live.xlsx"

    if not flat.exists():
        pytest.skip("FLAT_GED.xlsx not found")

    summary = extract(flat_path=flat, out_csv=out_csv, out_xlsx=out_xlsx)
    assert summary["unique_numero"]        == 2819,  summary
    assert summary["unique_numero_indice"] == 4848,  summary
    assert summary["step_OPEN_DOC"]        == 4848,  summary
    assert summary["step_SAS"]             == 4848,  summary
    assert summary["step_CONSULTANT"]      == 18911, summary
    assert summary["step_MOEX"]            == 3492,  summary
    assert summary["sas_ref_rows"]         == 284,   summary

    # Verify CSV header is byte-for-byte identical to raw_ged_trace.csv header
    root_debug = root / "output" / "debug" / "raw_ged_trace.csv"
    if root_debug.exists():
        with root_debug.open(encoding="utf-8") as fh:
            raw_header = next(csv.reader(fh))
        with out_csv.open(encoding="utf-8") as fh:
            flat_header = next(csv.reader(fh))
        assert flat_header == raw_header, (
            f"CSV header mismatch (byte-level).\n"
            f"flat: {flat_header}\nraw:  {raw_header}"
        )
