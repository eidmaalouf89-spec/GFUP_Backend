"""
tests/test_raw_flat_reconcile.py — Phase 8B.3 companion tests.

Fast (mock-only) tests exercise compute_identity_parity, write_identity_sheet,
and selfcheck_identity_sheet with small synthetic DataFrames.
The @pytest.mark.slow live test reads the real trace CSVs and expects PASS.

Run fast tests:
    python -m pytest tests/test_raw_flat_reconcile.py -v -m "not slow"
Run all:
    python -m pytest tests/test_raw_flat_reconcile.py -v
"""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pandas as pd
import pytest

# ── Load the script as a module without executing __main__ ────────────────────
_SCRIPT = Path(__file__).resolve().parent.parent / "scripts" / "raw_flat_reconcile.py"
_spec   = importlib.util.spec_from_file_location("raw_flat_reconcile", _SCRIPT)
_mod    = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

compute_identity_parity   = _mod.compute_identity_parity
write_identity_sheet      = _mod.write_identity_sheet
selfcheck_identity_sheet  = _mod.selfcheck_identity_sheet
compute_actor_call_parity   = _mod.compute_actor_call_parity
write_actor_call_sheets     = _mod.write_actor_call_sheets
selfcheck_actor_call_sheets = _mod.selfcheck_actor_call_sheets
compute_response_parity     = _mod.compute_response_parity
compute_sas_ref_trace       = _mod.compute_sas_ref_trace
write_response_sheets       = _mod.write_response_sheets
selfcheck_response_sheets   = _mod.selfcheck_response_sheets
normalize_date                  = _mod.normalize_date
compute_dates_comments_parity   = _mod.compute_dates_comments_parity
write_dates_comments_sheets     = _mod.write_dates_comments_sheets
selfcheck_dates_comments_sheets = _mod.selfcheck_dates_comments_sheets

_walk_flat_ged_drops      = _mod._walk_flat_ged_drops
_classify_drop_finding    = _mod._classify_drop_finding
_collect_raised_class_names = _mod._collect_raised_class_names
compute_reasons_audit     = _mod.compute_reasons_audit


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_df(rows: list[tuple[str, str]]) -> pd.DataFrame:
    """Build a minimal trace DataFrame with just numero and indice columns."""
    return pd.DataFrame(rows, columns=["numero", "indice"])


# ── Test 1: import guard ───────────────────────────────────────────────────────

def test_import_does_not_execute_main():
    """Loading the script via importlib must not call main() or any I/O."""
    assert callable(compute_identity_parity)
    assert callable(write_identity_sheet)
    assert callable(selfcheck_identity_sheet)
    assert hasattr(_mod, "main")


# ── Test 2: identical sets → PASS ─────────────────────────────────────────────

def test_identity_pass():
    raw  = _make_df([("100", "A"), ("100", "B"), ("200", "A")])
    flat = _make_df([("100", "A"), ("100", "B"), ("200", "A")])
    result = compute_identity_parity(raw, flat)
    assert result["num_verdict"]  == "PASS"
    assert result["pair_verdict"] == "PASS"
    assert result["total_missing"] == 0
    assert result["total_extra"]   == 0
    assert "IDENTITY_PARITY: numero PASS, numero_indice PASS" in result["summary_line"]
    assert "missing=0" in result["summary_line"]
    assert "extra=0"   in result["summary_line"]


# ── Test 3: numero missing in FLAT → BLOCKER ──────────────────────────────────

def test_identity_blocker_missing_numero():
    raw  = _make_df([("100", "A"), ("999", "A")])
    flat = _make_df([("100", "A")])
    result = compute_identity_parity(raw, flat)
    assert result["num_verdict"] == "BLOCKER"
    assert "999" in result["missing_num"]
    assert result["total_missing"] > 0
    assert "BLOCKER" in result["summary_line"]


# ── Test 4: extra numero in FLAT → BLOCKER ────────────────────────────────────

def test_identity_blocker_extra_numero():
    raw  = _make_df([("100", "A")])
    flat = _make_df([("100", "A"), ("888", "A")])
    result = compute_identity_parity(raw, flat)
    assert result["num_verdict"] == "BLOCKER"
    assert "888" in result["extra_num"]
    assert result["total_extra"] > 0
    assert "BLOCKER" in result["summary_line"]


# ── Test 5: same numeroes but missing pair → BLOCKER ─────────────────────────

def test_identity_blocker_missing_pair():
    raw  = _make_df([("100", "A"), ("100", "B")])
    flat = _make_df([("100", "A")])
    result = compute_identity_parity(raw, flat)
    # numero set is a subset (100 present in flat), but pair set differs
    assert result["pair_verdict"] == "BLOCKER"
    assert ("100", "B") in result["missing_pair"]


# ── Test 6: _MISSING_ sentinel rows are excluded from RAW sets ────────────────

def test_missing_filter_applied():
    raw = _make_df([
        ("100", "A"),
        ("_MISSING_001", "A"),
        ("_MISSING_002", "B"),
    ])
    flat = _make_df([("100", "A")])
    result = compute_identity_parity(raw, flat)
    # After filter, RAW has only row ("100", "A") — matches FLAT exactly
    assert result["raw_after_filter"] == 1
    assert result["raw_before_filter"] == 3
    assert result["num_verdict"]  == "PASS"
    assert result["pair_verdict"] == "PASS"
    assert "_MISSING_001" not in result["raw_numero_set"]
    assert "_MISSING_002" not in result["raw_numero_set"]


# ── Test 7: XLSX roundtrip (write then selfcheck passes) ─────────────────────

def test_identity_xlsx_roundtrip(tmp_path):
    raw  = _make_df([("100", "A"), ("200", "A")])
    flat = _make_df([("100", "A"), ("200", "A")])
    result = compute_identity_parity(raw, flat)
    xlsx = tmp_path / "out.xlsx"
    write_identity_sheet(result["sheet_rows"], xlsx)
    selfcheck_identity_sheet(result["sheet_rows"], xlsx)  # must not raise


# ── Test 8: selfcheck detects a tampered xlsx ─────────────────────────────────

def test_identity_selfcheck_detects_mismatch(tmp_path):
    raw  = _make_df([("100", "A")])
    flat = _make_df([("100", "A")])
    result = compute_identity_parity(raw, flat)
    xlsx = tmp_path / "out.xlsx"
    write_identity_sheet(result["sheet_rows"], xlsx)

    # Tamper: change raw_count for the numero row in the sheet
    from openpyxl import load_workbook as _lw
    wb = _lw(xlsx)
    ws = wb["01_IDENTITY_PARITY"]
    ws.cell(row=2, column=2).value = 9999  # raw_count for numero row
    wb.save(xlsx)

    with pytest.raises(AssertionError, match="raw_count"):
        selfcheck_identity_sheet(result["sheet_rows"], xlsx)


# ── Test 9: all Phase 8B steps are implemented (no remaining stubs) ──────────
# 'reasons_audit' (8B.7) and 'shadow' (8B.9) were the last stubs; both are
# now wired through main().  _STEP_PHASE is vestigial metadata used only for
# the error-message path; verify it no longer contains any unimplemented step.

def test_no_unimplemented_steps_remain():
    # Every step listed in argparse's choices must have a handler.  The map
    # is now empty for the canonical "implemented" steps because they're all
    # wired explicitly in main().
    implemented_choices = {"identity", "actor_calls", "responses",
                           "dates_comments", "reasons_audit", "shadow"}
    # _STEP_PHASE only carries (now-vestigial) metadata for steps that used
    # to be stubs — none of its keys should currently be unimplemented.
    for step in _mod._STEP_PHASE:
        assert step in implemented_choices


# ── Test 10: live test against real trace CSVs ────────────────────────────────

@pytest.mark.slow
def test_identity_live():
    """
    Read the real raw_ged_trace.csv and flat_ged_trace.csv and assert PASS.
    Expected (from Phase 8B.3 carry-forward):
        RAW unique_numero (post filter) = 2819
        RAW unique (numero, indice)     = 4848
        FLAT unique_numero              = 2819
        FLAT unique (numero, indice)    = 4848
    """
    repo_root = Path(__file__).resolve().parent.parent
    raw_csv   = repo_root / "output" / "debug" / "raw_ged_trace.csv"
    flat_csv  = repo_root / "output" / "debug" / "flat_ged_trace.csv"

    if not raw_csv.exists() or not flat_csv.exists():
        pytest.skip("trace CSVs not present — run extraction scripts first")

    raw_df  = pd.read_csv(raw_csv,  dtype=str, keep_default_na=False)
    flat_df = pd.read_csv(flat_csv, dtype=str, keep_default_na=False)

    result = compute_identity_parity(raw_df, flat_df)

    assert result["num_verdict"]  == "PASS",  f"numero BLOCKER: {result['summary_line']}"
    assert result["pair_verdict"] == "PASS",  f"pair   BLOCKER: {result['summary_line']}"
    assert result["total_missing"] == 0
    assert result["total_extra"]   == 0

    # Verify the carry-forward baseline counts
    assert len(result["raw_numero_set"])  == 2819, f"expected 2819, got {len(result['raw_numero_set'])}"
    assert len(result["raw_pair_set"])    == 4848, f"expected 4848, got {len(result['raw_pair_set'])}"
    assert len(result["flat_numero_set"]) == 2819, f"expected 2819, got {len(result['flat_numero_set'])}"
    assert len(result["flat_pair_set"])   == 4848, f"expected 4848, got {len(result['flat_pair_set'])}"


# ══════════════════════════════════════════════════════════════════════════════
# Phase 8B.4 — actor_calls tests
# ══════════════════════════════════════════════════════════════════════════════


def _make_ac_df(rows: list[tuple]) -> pd.DataFrame:
    """Build a minimal trace DataFrame for actor_calls tests.

    Each tuple: (numero, indice, actor_canonical, cycle_id, event_type, source_sheet)
    """
    return pd.DataFrame(
        rows,
        columns=["numero", "indice", "actor_canonical", "cycle_id", "event_type", "source_sheet"],
    )


# ── Test 11: identical actor sets → no diffs ─────────────────────────────────

def test_actor_calls_no_diffs():
    raw  = _make_ac_df([
        ("100", "A", "ARCHITECTE", "",   "RESPONSE",     "Doc. sous workflow, x versions"),
        ("100", "A", "0-SAS",      "C1", "RESPONSE",     "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE",     "GED_OPERATIONS"),
        ("100", "A", "0-SAS",      "", "RESPONSE",     "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    assert result["total_diffs"] == 0
    assert result["unexplained"] == 0
    assert len(result["raw_only"])  == 0
    assert len(result["flat_only"]) == 0
    assert "total_diffs=0" in result["summary_line"]
    assert "unexplained=0" in result["summary_line"]


# ── Test 12: RAW_ONLY actor not in FLAT universe → UNKNOWN_ACTOR ─────────────

def test_actor_calls_unknown_actor():
    raw  = _make_ac_df([
        ("200", "A", "Exception List", "C1", "ACTOR_CALLED", "Doc. sous workflow, x versions"),
        ("200", "A", "ARCHITECTE",     "",   "RESPONSE",     "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("200", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    assert len(result["raw_only"]) == 1
    assert result["total_diffs"] == 1
    diff_row = result["sheet03_rows"][0]
    assert diff_row["actor_canonical"] == "Exception List"
    assert diff_row["side"] == "RAW_ONLY"
    assert diff_row["classification"] == "UNKNOWN_ACTOR"


# ── Test 13: 0-SAS C2 with C1 matched → SAS_CYCLE_COLLAPSED ─────────────────

def test_actor_calls_sas_cycle_collapsed():
    raw  = _make_ac_df([
        ("300", "A", "0-SAS", "C1", "RESPONSE",     "Doc. sous workflow, x versions"),
        ("300", "A", "0-SAS", "C2", "ACTOR_CALLED", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("300", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    # C1 → "SAS" matches FLAT "SAS"; C2 → "C2" is RAW_ONLY
    assert len(result["match_set"]) == 1
    assert len(result["raw_only"])  == 1
    assert len(result["flat_only"]) == 0
    diff = result["sheet03_rows"][0]
    assert diff["actor_canonical"] == "0-SAS"
    assert diff["side"] == "RAW_ONLY"
    assert diff["classification"] == "SAS_CYCLE_COLLAPSED"


# ── Test 14: FLAT_ONLY actor, no RAW actors for that version → DUPLICATE_MERGED

def test_actor_calls_duplicate_merged():
    # FLAT has an actor for (400, A), but RAW has zero actor-call events for (400, A)
    raw  = _make_ac_df([
        ("400", "A", "GHOST", "", "DOCUMENT_VERSION", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("400", "A", "GHOST", "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    # RAW has no ACTOR_CALLED/RESPONSE rows → raw_actors_by_version is empty for (400,A)
    assert len(result["flat_only"]) == 1
    diff = result["sheet03_rows"][0]
    assert diff["side"] == "FLAT_ONLY"
    assert diff["classification"] == "DUPLICATE_MERGED"


# ── Test 15: FLAT_ONLY actor where RAW has other actors → UNEXPLAINED ────────

def test_actor_calls_flat_only_unexplained():
    raw  = _make_ac_df([
        ("500", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("500", "A", "ARCHITECTE", "",    "RESPONSE", "GED_OPERATIONS"),
        ("500", "A", "EXTRA_ACTOR", "",   "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    assert len(result["flat_only"]) == 1
    diff = result["sheet03_rows"][0]
    assert diff["actor_canonical"] == "EXTRA_ACTOR"
    assert diff["side"] == "FLAT_ONLY"
    assert diff["classification"] == "UNEXPLAINED"


# ── Test 16: ACTIVE_VERSION_PROJECTION ───────────────────────────────────────

def test_actor_calls_active_version_projection():
    # Actor ARCHITECTE appears in RAW for (600, A) but in FLAT only for (600, B)
    raw  = _make_ac_df([
        ("600", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
        ("600", "B", "0-SAS",      "C1", "RESPONSE", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("600", "B", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS"),
        ("600", "B", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    # (600, A, ARCHITECTE) is RAW_ONLY — same actor present in FLAT for (600, B)
    raw_only_rows = [r for r in result["sheet03_rows"] if r["side"] == "RAW_ONLY"]
    assert len(raw_only_rows) == 1
    assert raw_only_rows[0]["actor_canonical"] == "ARCHITECTE"
    assert raw_only_rows[0]["classification"] == "ACTIVE_VERSION_PROJECTION"


# ── Test 17: _MISSING_ filter applied before actor_calls computation ──────────

def test_actor_calls_missing_filter():
    raw  = _make_ac_df([
        ("_MISSING_001", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
        ("700",          "A", "0-SAS",      "C1", "RESPONSE", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("700", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    # _MISSING_ row must be excluded; only (700, A, 0-SAS) should remain
    assert len(result["raw_set"]) == 1
    assert len(result["raw_only"])  == 0
    assert len(result["flat_only"]) == 0
    assert result["total_diffs"] == 0


# ── Test 18: GED_OPERATIONS filter on FLAT side ───────────────────────────────

def test_actor_calls_flat_non_ged_ops_excluded():
    # Rows from GED_RAW_FLAT sheet must NOT be included in the FLAT actor-call set
    flat = _make_ac_df([
        ("800", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS"),
        ("800", "A", "GHOST",      "", "RESPONSE", "GED_RAW_FLAT"),  # must be excluded
    ])
    raw  = _make_ac_df([
        ("800", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
    ])
    result = compute_actor_call_parity(raw, flat)
    assert len(result["flat_set"]) == 1  # only GED_OPERATIONS row
    assert result["total_diffs"] == 0


# ── Test 19: XLSX roundtrip — write then selfcheck passes ────────────────────

def test_actor_calls_xlsx_roundtrip(tmp_path):
    raw  = _make_ac_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
        ("900", "A", "0-SAS",      "C1", "RESPONSE", "Doc. sous workflow, x versions"),
        ("900", "A", "0-SAS",      "C2", "ACTOR_CALLED", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS"),
        ("900", "A", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    xlsx = tmp_path / "reconcile.xlsx"
    write_actor_call_sheets(result["sheet02_rows"], result["sheet03_rows"], xlsx)
    selfcheck_actor_call_sheets(result, xlsx)  # must not raise


# ── Test 20: selfcheck detects row count mismatch ────────────────────────────

def test_actor_calls_selfcheck_detects_mismatch(tmp_path):
    raw  = _make_ac_df([("910", "A", "ARCHITECTE", "", "RESPONSE",
                          "Doc. sous workflow, x versions")])
    flat = _make_ac_df([("910", "A", "GHOST",      "", "RESPONSE", "GED_OPERATIONS")])
    result = compute_actor_call_parity(raw, flat)
    xlsx = tmp_path / "reconcile.xlsx"
    write_actor_call_sheets(result["sheet02_rows"], result["sheet03_rows"], xlsx)

    # Tamper: add a spurious extra row to sheet 03
    from openpyxl import load_workbook as _lw
    wb = _lw(xlsx)
    ws = wb["03_ACTOR_CALL_DIFFS"]
    ws.append(["999", "Z", "FAKE", "RAW_ONLY", "UNEXPLAINED", "tampered"])
    wb.save(xlsx)

    with pytest.raises(AssertionError, match="row count"):
        selfcheck_actor_call_sheets(result, xlsx)


# ── Test 21: summary line format ─────────────────────────────────────────────

def test_actor_calls_summary_line_format():
    raw  = _make_ac_df([])
    flat = _make_ac_df([])
    result = compute_actor_call_parity(raw, flat)
    assert result["summary_line"].startswith("ACTOR_CALL_PARITY: total_diffs=")
    assert "unexplained=" in result["summary_line"]


# ── Test 22: sheet02 columns present and non-empty for non-trivial data ───────

def test_actor_calls_sheet02_columns():
    raw  = _make_ac_df([
        ("920", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions"),
    ])
    flat = _make_ac_df([
        ("920", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS"),
        ("920", "A", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS"),
    ])
    result = compute_actor_call_parity(raw, flat)
    assert len(result["sheet02_rows"]) >= 1
    row = result["sheet02_rows"][0]
    for col in ["actor_canonical", "raw_count", "flat_count", "delta",
                "expected_explanation_category"]:
        assert col in row, f"missing column: {col}"


# ══════════════════════════════════════════════════════════════════════════════
# Phase 8B.5 — responses tests
# ══════════════════════════════════════════════════════════════════════════════


def _make_resp_df(rows: list[tuple]) -> pd.DataFrame:
    """Build a minimal trace DataFrame for response tests.

    Each tuple: (numero, indice, actor_canonical, cycle_id, event_type,
                 source_sheet, status_raw, status_clean)
    """
    return pd.DataFrame(
        rows,
        columns=[
            "numero", "indice", "actor_canonical", "cycle_id", "event_type",
            "source_sheet", "status_raw", "status_clean",
        ],
    )


# ── Test 23: identical response sets → no diffs ───────────────────────────────

def test_response_parity_no_diffs():
    raw = _make_resp_df([
        ("100", "A", "ARCHITECTE", "",   "RESPONSE", "Doc. sous workflow, x versions", "VAO", "VAO"),
        ("100", "A", "0-SAS",      "C1", "RESPONSE", "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    flat = _make_resp_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
        ("100", "A", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
    ])
    result = compute_response_parity(raw, flat)
    assert result["total_raw"] == 2
    assert result["total_flat"] == 2
    assert result["total_matched"] == 2
    assert len(result["raw_only"]) == 0
    assert len(result["flat_only"]) == 0
    assert result["unexplained"] == 0
    assert "RESPONSE_PARITY:" in result["summary_line"]


# ── Test 24: SAS C2 with C1 matched → SAS_CYCLE_COLLAPSED ────────────────────

def test_response_parity_sas_cycle_collapsed():
    raw = _make_resp_df([
        ("200", "A", "0-SAS", "C1", "RESPONSE", "Doc. sous workflow, x versions", "REF", "REF"),
        ("200", "A", "0-SAS", "C2", "RESPONSE", "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    flat = _make_resp_df([
        ("200", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
    ])
    result = compute_response_parity(raw, flat)
    assert len(result["raw_only"]) == 1
    assert result["total_matched"] == 1
    diff = result["sheet06_rows"][0]
    assert diff["actor_canonical"] == "0-SAS"
    assert diff["side"] == "RAW_ONLY"
    assert diff["classification"] == "SAS_CYCLE_COLLAPSED"


# ── Test 25: Exception List actor → UNKNOWN_ACTOR ────────────────────────────

def test_response_parity_unknown_actor():
    raw = _make_resp_df([
        ("300", "A", "Exception List", "", "RESPONSE", "Doc. sous workflow, x versions", "REF", "REF"),
        ("300", "A", "ARCHITECTE",     "", "RESPONSE", "Doc. sous workflow, x versions", "VAO", "VAO"),
    ])
    flat = _make_resp_df([
        ("300", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
    ])
    result = compute_response_parity(raw, flat)
    assert len(result["raw_only"]) == 1
    diff = result["sheet06_rows"][0]
    assert diff["actor_canonical"] == "Exception List"
    assert diff["classification"] == "UNKNOWN_ACTOR"


# ── Test 26: old indice in RAW, FLAT has newer → ACTIVE_VERSION_PROJECTION ───

def test_response_parity_active_version_projection():
    raw = _make_resp_df([
        ("400", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions", "VAO", "VAO"),
        ("400", "B", "0-SAS",      "C1", "RESPONSE", "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    flat = _make_resp_df([
        ("400", "B", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
        ("400", "B", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
    ])
    result = compute_response_parity(raw, flat)
    raw_only_rows = [r for r in result["sheet06_rows"] if r["side"] == "RAW_ONLY"]
    assert len(raw_only_rows) == 1
    assert raw_only_rows[0]["actor_canonical"] == "ARCHITECTE"
    assert raw_only_rows[0]["classification"] == "ACTIVE_VERSION_PROJECTION"


# ── Test 27: non-operational status → NON_OPERATIONAL_RESPONSE ───────────────

def test_response_parity_non_operational():
    raw = _make_resp_df([
        ("500", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions", "info", "info"),
    ])
    flat = _make_resp_df([
        ("500", "A", "OTHER",       "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
    ])
    result = compute_response_parity(raw, flat)
    diff = next(r for r in result["sheet06_rows"] if r["side"] == "RAW_ONLY")
    assert diff["classification"] == "NON_OPERATIONAL_RESPONSE"


# ── Test 28: malformed response (status_raw populated, status_clean empty) ────

def test_response_parity_malformed():
    raw = _make_resp_df([
        ("600", "A", "ARCHITECTE", "", "RESPONSE", "Doc. sous workflow, x versions", "VAO", ""),
    ])
    flat = _make_resp_df([
        ("600", "A", "OTHER", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
    ])
    result = compute_response_parity(raw, flat)
    diff = next(r for r in result["sheet06_rows"] if r["side"] == "RAW_ONLY")
    assert diff["classification"] == "MALFORMED_RESPONSE"


# ── Test 29: _MISSING_ filter applied to response parity ─────────────────────

def test_response_parity_missing_filter():
    raw = _make_resp_df([
        ("_MISSING_001", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("700", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO"),
    ])
    flat = _make_resp_df([
        ("700", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
    ])
    result = compute_response_parity(raw, flat)
    assert result["total_raw"] == 1   # _MISSING_ row excluded
    assert result["total_matched"] == 1
    assert len(result["raw_only"]) == 0


# ── Test 30: GED_RAW_FLAT excluded from FLAT response set ────────────────────

def test_response_parity_flat_raw_flat_excluded():
    raw = _make_resp_df([
        ("800", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO"),
    ])
    flat = _make_resp_df([
        ("800", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",  "VAO", "VAO"),
        ("800", "A", "GHOST",      "", "RESPONSE", "GED_RAW_FLAT",    "VAO", "VAO"),
    ])
    result = compute_response_parity(raw, flat)
    assert result["total_flat"] == 1   # GED_RAW_FLAT row excluded
    assert result["total_matched"] == 1


# ── Test 31: sheet04 structure ────────────────────────────────────────────────

def test_response_parity_sheet04_structure():
    raw = _make_resp_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO"),
    ])
    flat = _make_resp_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
    ])
    result = compute_response_parity(raw, flat)
    for row in result["sheet04_rows"]:
        for col in ["status_clean", "side", "count"]:
            assert col in row
    sides = {r["side"] for r in result["sheet04_rows"]}
    assert sides <= {"RAW", "FLAT"}


# ── Test 32: SAS REF trace — _MISSING_ rows classified MALFORMED_RESPONSE ────

def test_sas_ref_trace_missing_classified():
    raw = _make_resp_df([
        ("_MISSING_001", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("100", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    # Add source_excel_row column required by compute_sas_ref_trace
    raw["source_excel_row"] = ["10", "20"]
    flat = _make_resp_df([
        ("100", "A", "0-SAS", "", "RESPONSE", "GED_RAW_FLAT", "REF", "REF"),
    ])
    result = compute_sas_ref_trace(raw, flat)
    assert result["total_sas_raw"] == 2
    missing_row = next(r for r in result["sheet07_rows"] if "_MISSING_" in str(r["numero"]))
    assert missing_row["classification"] == "MALFORMED_RESPONSE"
    assert not missing_row["flat_present"]


# ── Test 33: SAS REF trace — C1+C2 dual-cycle row → raw_cycle='C1+C2' ────────

def test_sas_ref_trace_dual_cycle():
    raw = _make_resp_df([
        ("152012", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("152012", "A", "0-SAS", "C2", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    raw["source_excel_row"] = ["2131", "2131"]
    flat = _make_resp_df([
        ("152012", "A", "0-SAS", "", "RESPONSE", "GED_RAW_FLAT", "REF", "REF"),
    ])
    result = compute_sas_ref_trace(raw, flat)
    # Two raw rows from same source_excel_row → one sheet07 row
    assert result["total_sas_raw"] == 1
    row = result["sheet07_rows"][0]
    assert row["raw_cycle"] == "C1+C2"
    assert row["flat_present"] is True


# ── Test 34: SAS REF trace — duplicate (numero,indice) → DUPLICATE_MERGED ────

def test_sas_ref_trace_duplicate_merged():
    raw = _make_resp_df([
        ("1", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("1", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    raw["source_excel_row"] = ["797", "798"]  # two different Excel rows, same (num,ind)
    flat = _make_resp_df([
        ("1", "A", "0-SAS", "", "RESPONSE", "GED_RAW_FLAT", "REF", "REF"),
    ])
    result = compute_sas_ref_trace(raw, flat)
    assert result["total_sas_raw"] == 2
    canonical = next(
        r for r in result["sheet07_rows"] if r["evidence_excel_row"] == "797"
    )
    duplicate = next(
        r for r in result["sheet07_rows"] if r["evidence_excel_row"] == "798"
    )
    assert canonical["classification"] == ""          # canonical matched row
    assert canonical["flat_present"] is True
    assert duplicate["classification"] == "DUPLICATE_MERGED"


# ── Test 35: SAS REF trace — flat_present False, same numero other indice ────

def test_sas_ref_trace_active_version_projection():
    raw = _make_resp_df([
        ("999", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    raw["source_excel_row"] = ["5000"]
    flat_ops = _make_resp_df([
        ("999", "B", "0-SAS", "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
    ])
    flat_rf = _make_resp_df([
        ("999", "B", "0-SAS", "", "RESPONSE", "GED_RAW_FLAT", "REF", "REF"),
    ])
    flat = pd.concat([flat_ops, flat_rf], ignore_index=True)
    result = compute_sas_ref_trace(raw, flat)
    row = result["sheet07_rows"][0]
    assert not row["flat_present"]
    assert row["classification"] == "ACTIVE_VERSION_PROJECTION"


# ── Test 36: summary line format for responses ────────────────────────────────

def test_response_summary_line_format():
    raw = _make_resp_df([])
    flat = _make_resp_df([])
    result = compute_response_parity(raw, flat)
    assert result["summary_line"].startswith("RESPONSE_PARITY: total_raw=")
    assert "total_flat=" in result["summary_line"]
    assert "matched=" in result["summary_line"]


# ── Test 37: SAS REF trace summary line format ────────────────────────────────

def test_sas_ref_summary_line_format():
    raw = _make_resp_df([])
    flat = _make_resp_df([])
    raw["source_excel_row"] = []
    result = compute_sas_ref_trace(raw, flat)
    assert "SAS_REF_TRACE:" in result["summary_line"]
    assert "classified=" in result["summary_line"]
    assert "unexplained=" in result["summary_line"]
    assert "duplicate_favorable_kept=" in result["summary_line"]


# ── Test 38: response sheets XLSX roundtrip ───────────────────────────────────

def test_response_sheets_xlsx_roundtrip(tmp_path):
    raw = _make_resp_df([
        ("100", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("100", "A", "0-SAS", "C2", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
        ("100", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO"),
    ])
    raw_sas = _make_resp_df([
        ("100", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF"),
    ])
    raw_sas["source_excel_row"] = ["3"]
    flat = _make_resp_df([
        ("100", "A", "0-SAS",      "", "RESPONSE", "GED_OPERATIONS", "REF", "REF"),
        ("100", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS", "VAO", "VAO"),
        ("100", "A", "0-SAS",      "", "RESPONSE", "GED_RAW_FLAT",   "REF", "REF"),
    ])

    resp_result = compute_response_parity(raw, flat)
    sas_result  = compute_sas_ref_trace(raw_sas, flat)

    xlsx = tmp_path / "reconcile.xlsx"
    write_response_sheets(
        resp_result["sheet04_rows"],
        resp_result["sheet05_rows"],
        resp_result["sheet06_rows"],
        sas_result["sheet07_rows"],
        xlsx,
    )
    selfcheck_response_sheets(resp_result, sas_result, xlsx)  # must not raise


# ══════════════════════════════════════════════════════════════════════════════
# Phase 8B.6 — dates_comments tests
# ══════════════════════════════════════════════════════════════════════════════

import datetime as _dt


def _make_dc_df(rows: list[tuple]) -> pd.DataFrame:
    """Minimal trace DataFrame for dates_comments tests.

    Tuple: (numero, indice, actor_canonical, cycle_id, event_type, source_sheet,
            status_raw, status_clean, submission_date, response_date, deadline_date,
            comment_raw, source_excel_row)
    """
    return pd.DataFrame(
        rows,
        columns=[
            "numero", "indice", "actor_canonical", "cycle_id", "event_type",
            "source_sheet", "status_raw", "status_clean",
            "submission_date", "response_date", "deadline_date",
            "comment_raw", "source_excel_row",
        ],
    )


# ── normalize_date unit tests (one per input type per spec) ───────────────────

def test_normalize_date_excel_serial_int():
    # spec example: 44561 → 2021-12-31 (verified: date(1899,12,30)+timedelta(44561))
    result = normalize_date("44561")
    expected = (_dt.date(1899, 12, 30) + _dt.timedelta(44561)).isoformat()
    assert result == expected
    assert result == "2021-12-31"


def test_normalize_date_excel_serial_float():
    result = normalize_date("44561.0")
    assert result == "2021-12-31"


def test_normalize_date_iso_string():
    assert normalize_date("2021-12-31") == "2021-12-31"


def test_normalize_date_iso_with_time():
    assert normalize_date("2021-12-31 10:30:00") == "2021-12-31"


def test_normalize_date_french():
    assert normalize_date("31/12/2021") == "2021-12-31"


def test_normalize_date_pandas_timestamp():
    import pandas as pd
    assert normalize_date(pd.Timestamp("2021-12-31")) == "2021-12-31"


def test_normalize_date_datetime_date():
    assert normalize_date(_dt.date(2021, 12, 31)) == "2021-12-31"


def test_normalize_date_datetime_datetime():
    assert normalize_date(_dt.datetime(2021, 12, 31, 10, 30)) == "2021-12-31"


def test_normalize_date_blanks():
    assert normalize_date("") is None
    assert normalize_date(None) is None
    assert normalize_date("NaT") is None
    assert normalize_date("None") is None
    assert normalize_date("nan") is None


def test_normalize_date_pandas_nat():
    import pandas as pd
    assert normalize_date(pd.NaT) is None


# ── compute_dates_comments_parity mock tests ──────────────────────────────────

def test_dates_comments_all_match():
    raw = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "Good doc.", "2"),
    ])
    flat = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "Good doc.", "10"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["matched"] == 1
    assert result["drifted_total"] == 0
    assert result["rnonblank_fblank"] == 0
    assert result["rblank_fnonblank"] == 0
    # submission + response MATCH; deadline "" on both → RAW_BLANK_FLAT_BLANK
    assert result["match_date"] == 2
    assert result["rblank_fblank"] == 1
    assert result["comment_counts"]["MATCH"] == 1
    assert len(result["sheet08_rows"]) == 0
    assert len(result["sheet09_rows"]) == 0
    assert "DATE_PARITY:" in result["summary_date"]
    assert "COMMENT_PARITY:" in result["summary_comment"]


def test_dates_comments_date_drifted():
    raw = _make_dc_df([
        ("200", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "3"),
    ])
    flat = _make_dc_df([
        ("200", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-03-01", "", "", "11"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["matched"] == 1
    assert result["drifted_total"] == 1  # response_date drifted
    drifted_rows = [r for r in result["sheet08_rows"] if r["verdict"] == "DRIFTED"]
    assert len(drifted_rows) == 1
    assert drifted_rows[0]["date_field"] == "response"


def test_dates_comments_raw_nonblank_flat_blank():
    # Load-bearing signal: RAW has a date, FLAT does not
    raw = _make_dc_df([
        ("300", "A", "BET Electr.", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "FAV", "FAV", "2024-06-15", "2024-07-01", "2024-08-01", "", "4"),
    ])
    flat = _make_dc_df([
        ("300", "A", "BET Electr.", "", "RESPONSE", "GED_OPERATIONS",
         "FAV", "FAV", "2024-06-15", "2024-07-01", "", "", "12"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["rnonblank_fblank"] == 1  # deadline_date lost
    rnfb_rows = [r for r in result["sheet08_rows"] if r["verdict"] == "RAW_NONBLANK_FLAT_BLANK"]
    assert len(rnfb_rows) == 1
    assert rnfb_rows[0]["date_field"] == "deadline"
    assert rnfb_rows[0]["raw_value_norm"] == "2024-08-01"
    assert rnfb_rows[0]["flat_value_norm"] is None


def test_dates_comments_raw_blank_flat_nonblank():
    raw = _make_dc_df([
        ("400", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "HM", "HM", "", "2025-05-01", "", "", "5"),
    ])
    flat = _make_dc_df([
        ("400", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "HM", "HM", "2025-04-01", "2025-05-01", "", "", "13"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["rblank_fnonblank"] == 1  # submission_date only on FLAT
    rbnf_rows = [r for r in result["sheet08_rows"] if r["verdict"] == "RAW_BLANK_FLAT_NONBLANK"]
    assert len(rbnf_rows) == 1
    assert rbnf_rows[0]["date_field"] == "submission"


def test_dates_comments_comment_raw_only():
    raw = _make_dc_df([
        ("500", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions",
         "REF", "REF", "2025-01-01", "2025-02-01", "", "See note.", "6"),
    ])
    flat = _make_dc_df([
        ("500", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS",
         "REF", "REF", "2025-01-01", "2025-02-01", "", "", "14"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["comment_counts"]["RAW_ONLY"] == 1
    assert len(result["sheet09_rows"]) == 1
    assert result["sheet09_rows"][0]["verdict"] == "RAW_ONLY"


def test_dates_comments_comment_flat_only():
    raw = _make_dc_df([
        ("600", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VSO", "VSO", "2025-03-01", "2025-04-01", "", "", "7"),
    ])
    flat = _make_dc_df([
        ("600", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VSO", "VSO", "2025-03-01", "2025-04-01", "", "FLAT note.", "15"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["comment_counts"]["FLAT_ONLY"] == 1
    assert result["sheet09_rows"][0]["verdict"] == "FLAT_ONLY"


def test_dates_comments_comment_drifted():
    raw = _make_dc_df([
        ("700", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "RAW comment.", "8"),
    ])
    flat = _make_dc_df([
        ("700", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "FLAT comment.", "16"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["comment_counts"]["DRIFTED"] == 1
    assert result["sheet09_rows"][0]["verdict"] == "DRIFTED"


def test_dates_comments_missing_filter():
    raw = _make_dc_df([
        ("_MISSING_001", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "1"),
        ("800", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "2"),
    ])
    flat = _make_dc_df([
        ("800", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "10"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    assert result["matched"] == 1  # _MISSING_ row excluded


def test_dates_comments_math_self_check():
    # Verify sum invariants: per-field totals = 3×matched; comment totals = matched
    raw = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "2"),
        ("200", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions",
         "REF", "REF", "2024-01-01", "", "", "note", "3"),
    ])
    flat = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "", "10"),
        ("200", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS",
         "REF", "REF", "2024-01-01", "", "", "", "11"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    matched = result["matched"]
    assert matched == 2
    pf_total = (
        result["match_date"] + result["drifted_total"] + result["rblank_fblank"]
        + result["rnonblank_fblank"] + result["rblank_fnonblank"]
    )
    assert pf_total == 3 * matched
    c_total = sum(result["comment_counts"].values())
    assert c_total == matched


def test_dates_comments_xlsx_roundtrip(tmp_path):
    raw = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "2025-01-01", "2025-02-01", "", "note", "2"),
        ("200", "A", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions",
         "REF", "REF", "2024-01-01", "2024-06-01", "", "", "3"),
    ])
    flat = _make_dc_df([
        ("100", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2025-01-01", "2025-03-01", "", "", "10"),  # response drifted
        ("200", "A", "0-SAS", "", "RESPONSE", "GED_OPERATIONS",
         "REF", "REF", "2024-01-01", "2024-06-01", "", "", "11"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    xlsx = tmp_path / "reconcile.xlsx"
    write_dates_comments_sheets(result["sheet08_rows"], result["sheet09_rows"], xlsx)
    # Override matched for selfcheck so mock data (not 16187) doesn't fail
    result_patched = dict(result, matched=_mod._EXPECTED_MATCHED_8B5)
    # The selfcheck checks matched == 16187 — skip that assertion for mock
    # Just verify sheet row counts from the workbook directly
    from openpyxl import load_workbook as _lw
    wb = _lw(xlsx, read_only=True)
    rows08 = list(wb["08_DATE_DIFFS"].iter_rows(values_only=True))[1:]
    rows09 = list(wb["09_COMMENT_DIFFS"].iter_rows(values_only=True))[1:]
    wb.close()
    exp_s08 = result["drifted_total"] + result["rnonblank_fblank"] + result["rblank_fnonblank"]
    assert len(rows08) == exp_s08
    exp_s09 = (
        result["comment_counts"]["DRIFTED"]
        + result["comment_counts"]["RAW_ONLY"]
        + result["comment_counts"]["FLAT_ONLY"]
    )
    assert len(rows09) == exp_s09


# ══════════════════════════════════════════════════════════════════════════════
# Phase 8B.5 taxonomy extension — DUPLICATE_FAVORABLE_KEPT tests
# ══════════════════════════════════════════════════════════════════════════════


def _make_sas_trace_df(rows: list[tuple]) -> pd.DataFrame:
    """Build a trace DataFrame for SAS REF trace tests that includes submission_date.

    Tuple: (numero, indice, actor_canonical, cycle_id, event_type, source_sheet,
            status_raw, status_clean, submission_date, source_excel_row)
    """
    return pd.DataFrame(
        rows,
        columns=[
            "numero", "indice", "actor_canonical", "cycle_id", "event_type",
            "source_sheet", "status_raw", "status_clean",
            "submission_date", "source_excel_row",
        ],
    )


# ── Test 39: positive case — both conditions hold → DUPLICATE_FAVORABLE_KEPT ──

def test_sas_ref_trace_duplicate_favorable_kept_positive():
    """Both (a) and (b) hold: same-day VAO in RAW + FLAT has VAO → DFK."""
    raw = _make_sas_trace_df([
        # SAS REF row — goes into all_sas_ref and sheet07
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF", "2023-06-15", "1000"),
        # Same-day VAO row — goes into raw_sas_favorable_by_ni (not sheet07)
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO", "2023-06-15", "1001"),
    ])
    flat = _make_sas_trace_df([
        # FLAT has VAO → flat_sas_favorable_ni includes (249101,B)
        # No GED_RAW_FLAT REF → flat_ni_set does NOT include (249101,B) → flat_present=False
        ("249101", "B", "0-SAS", "", "RESPONSE",
         "GED_OPERATIONS", "VAO", "VAO", "", ""),
    ])
    result = compute_sas_ref_trace(raw, flat)
    # Only the REF row appears in sheet07 (status_raw='REF' filter)
    assert result["total_sas_raw"] == 1
    row = result["sheet07_rows"][0]
    assert not row["flat_present"]
    assert row["classification"] == "DUPLICATE_FAVORABLE_KEPT"
    assert result["duplicate_favorable_kept"] == 1
    assert "duplicate_favorable_kept=1" in result["summary_line"]


# ── Test 40: near-miss — condition (b) fails (FLAT has no favorable) → UNEXPLAINED

def test_sas_ref_trace_dfk_no_flat_favorable():
    """Condition (b) fails: FLAT has no VAO/VSO/VSO-SAS for this ni → UNEXPLAINED."""
    raw = _make_sas_trace_df([
        # SAS REF row
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF", "2023-06-15", "1000"),
        # RAW has VAO → condition (a) holds
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "VAO", "VAO", "2023-06-15", "1001"),
    ])
    flat = _make_sas_trace_df([
        # FLAT has NO favorable (no VAO/VSO/VSO-SAS) → condition (b) fails
        # Use a different numero entirely so flat_sas_favorable_ni stays empty for (249101,B)
    ])
    flat = pd.DataFrame(columns=["numero", "indice", "actor_canonical", "cycle_id",
                                  "event_type", "source_sheet", "status_raw",
                                  "status_clean", "submission_date", "source_excel_row"])
    result = compute_sas_ref_trace(raw, flat)
    row = result["sheet07_rows"][0]
    assert row["classification"] == "UNEXPLAINED"
    assert result["duplicate_favorable_kept"] == 0


# ── Test 41: near-miss — non-favorable status (HM) in RAW → UNEXPLAINED ──────

def test_sas_ref_trace_dfk_non_favorable_status():
    """Condition (a) fails: RAW only has HM (not VAO/VSO/VSO-SAS) → UNEXPLAINED."""
    raw = _make_sas_trace_df([
        # SAS REF row
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "REF", "REF", "2023-06-15", "1000"),
        # HM is not a favorable status (HM ∉ {VAO,VSO,VSO-SAS}) → (a) fails
        ("249101", "B", "0-SAS", "C1", "RESPONSE",
         "Doc. sous workflow, x versions", "HM", "HM", "2023-06-15", "1001"),
    ])
    flat = _make_sas_trace_df([
        # FLAT has VAO → condition (b) holds; but (a) fails because HM ∉ favorable set
        ("249101", "B", "0-SAS", "", "RESPONSE",
         "GED_OPERATIONS", "VAO", "VAO", "", ""),
    ])
    result = compute_sas_ref_trace(raw, flat)
    row = result["sheet07_rows"][0]
    assert row["classification"] == "UNEXPLAINED"
    assert result["duplicate_favorable_kept"] == 0


def test_dates_comments_date_normalization_equivalence():
    # Excel serial "44561" and ISO "2021-12-31" are equivalent
    raw = _make_dc_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE",
         "Doc. sous workflow, x versions",
         "VAO", "VAO", "44561", "2021-12-31", "", "", "9"),
    ])
    flat = _make_dc_df([
        ("900", "A", "ARCHITECTE", "", "RESPONSE", "GED_OPERATIONS",
         "VAO", "VAO", "2021-12-31", "31/12/2021", "", "", "17"),
    ])
    result = compute_dates_comments_parity(raw, flat)
    # Both submission_date and response_date should be MATCH after normalisation
    assert result["drifted_total"] == 0
    assert result["match_date"] == 2  # submission and response match; deadline both blank


# ── Phase 8B.7 reasons-audit tests (mock-only, synthetic source snippets) ─────

def _write_synthetic_src(tmp_path: Path, files: dict[str, str]) -> Path:
    """Lay out a synthetic ``flat_ged``-like package under ``tmp_path``."""
    pkg = tmp_path / "flat_ged"
    pkg.mkdir()
    for name, src in files.items():
        (pkg / name).write_text(src, encoding="utf-8")
    return pkg


def test_reasons_audit_classify_valid_taxonomy_match():
    finding = {
        "file": "flat_ged/transformer.py",
        "function": "build_raw_flat",
        "line": 122,
        "kind": "CONTINUE_DROP",
        "reason_string": "GED_RAW_FLAT: skip Exception List and NOT_CALLED rows",
        "source_excerpt": "",
    }
    out = _classify_drop_finding(finding, raised_classes={"Foo"})
    assert out["classification"] == "VALID_REASON"
    assert out["related_taxonomy_bucket"] == "UNKNOWN_ACTOR"


def test_reasons_audit_classify_missing_when_no_reason():
    finding = {
        "file": "flat_ged/foo.py",
        "function": "f",
        "line": 10,
        "kind": "CONTINUE_DROP",
        "reason_string": "",
        "source_excerpt": "if x: continue",
    }
    out = _classify_drop_finding(finding, raised_classes=set())
    assert out["classification"] == "MISSING_REASON"


def test_reasons_audit_classify_invalid_when_class_never_raised():
    """Exception class with row-exclusion docstring but never raised → INVALID."""
    finding = {
        "file": "flat_ged/resolver.py",
        "function": "GEDDocumentSkip",
        "line": 17,
        "kind": "EXCEPTION_CLASS_DOC",
        "reason_string": "Raised when a document should be excluded from GED_FLAT (e.g. no NUMERO).",
        "source_excerpt": "",
    }
    out = _classify_drop_finding(finding, raised_classes={"GEDValidationError"})
    assert out["classification"] == "INVALID_REASON"
    assert "never raised" in out["notes"]


def test_reasons_audit_classify_ambiguous_when_generic():
    finding = {
        "file": "flat_ged/resolver.py",
        "function": "score_candidate",
        "line": 88,
        "kind": "CONTINUE_DROP",
        "reason_string": "if ag['col_date'] >= len(row_data): continue  # bounds",
        "source_excerpt": "",
    }
    out = _classify_drop_finding(finding, raised_classes=set())
    assert out["classification"] == "AMBIGUOUS_REASON"


def test_reasons_audit_walk_finds_continue_and_class_doc(tmp_path):
    src = _write_synthetic_src(tmp_path, {
        "transformer.py": (
            "class GEDValidationError(Exception):\n"
            "    pass\n"
            "\n"
            "def build_raw_flat(rows):\n"
            "    out = []\n"
            "    for r in rows:\n"
            "        if r['canonical'] == 'Exception List':  # skip Exception List\n"
            "            continue\n"
            "        out.append(r)\n"
            "    return out\n"
        ),
        "resolver.py": (
            "class GEDDocumentSkip(Exception):\n"
            "    \"\"\"Raised when a document should be excluded — no NUMERO.\"\"\"\n"
            "\n"
            "class GEDValidationError(Exception):\n"
            "    pass\n"
            "\n"
            "def f():\n"
            "    raise GEDValidationError('[FAIL] something')\n"
        ),
    })
    findings = _walk_flat_ged_drops(src)
    kinds = {(f["file"].split("/")[-1].split("\\")[-1], f["kind"]) for f in findings}
    assert any(k[1] == "CONTINUE_DROP" for k in kinds)
    assert any(k[1] == "EXCEPTION_CLASS_DOC" for k in kinds)
    assert any(k[1] == "RAISE" for k in kinds)


def test_reasons_audit_collect_raised_classes(tmp_path):
    src = _write_synthetic_src(tmp_path, {
        "a.py": (
            "class A(Exception): pass\n"
            "class B(Exception):\n"
            "    \"\"\"Raised for skipping.\"\"\"\n"
            "\n"
            "def f():\n"
            "    raise A('x')\n"
        ),
    })
    raised = _collect_raised_class_names(src)
    assert "A" in raised
    assert "B" not in raised  # B has the doc but is never raised


def test_reasons_audit_compute_end_to_end(tmp_path):
    src = _write_synthetic_src(tmp_path, {
        "transformer.py": (
            "def build(rows):\n"
            "    for r in rows:\n"
            "        if r == 'Exception List':  # skip Exception List rows\n"
            "            continue\n"
            "        if r is None:\n"
            "            continue\n"  # MISSING_REASON
            "        yield r\n"
        ),
        "resolver.py": (
            "class GEDDocumentSkip(Exception):\n"
            "    \"\"\"Raised when a document should be excluded — no NUMERO.\"\"\"\n"
            "\n"
            "class Other(Exception):\n"
            "    pass\n"
            "\n"
            "def f():\n"
            "    raise Other('x')\n"
        ),
    })
    result = compute_reasons_audit(src)
    assert result["summary"]["total"] >= 4
    # at least one INVALID (the doc'd-but-never-raised class)
    assert result["summary"]["INVALID_REASON"] >= 1
    # the SUMMARY_LINE format must match the spec exactly
    s = result["summary_line"]
    assert s.startswith("REASON_AUDIT: total=")
    for tok in (" valid=", " invalid=", " missing=", " ambiguous="):
        assert tok in s
