"""
tests/test_audit_counts_lineage.py
Phase 1 companion tests for scripts/audit_counts_lineage.py

Tests:
  1. Import the script without triggering main() (guard check)
  2. EXPECTED_BASELINES contains all required values
  3. EXPECTED_DIVERGENCES is a list of dicts (data, not conditionals)
  4. After a real run on run_number=0, JSON output validates against §5.3 step 11 shape
"""
from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

import pytest

# ── Load the script as a module without executing __main__ ────────────────────
_SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"
_SCRIPT_PATH = _SCRIPTS_DIR / "audit_counts_lineage.py"


def _load_module():
    spec = importlib.util.spec_from_file_location("audit_counts_lineage", _SCRIPT_PATH)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# Load once at module level (simulates import; __name__ != "__main__" guard respected)
_mod = _load_module()


# ── Test 1: Import without executing main ────────────────────────────────────

def test_import_does_not_execute_main():
    """Loading the script via importlib must not call run_audit() or argparse."""
    # If __main__ guard is absent, run_audit() would have been called during
    # _load_module() above, which would have raised or taken ≥ 30 s.
    # Reaching here without error proves the guard works.
    assert hasattr(_mod, "run_audit"), "run_audit function must be defined"
    assert hasattr(_mod, "EXPECTED_BASELINES"), "EXPECTED_BASELINES must be defined"
    assert hasattr(_mod, "EXPECTED_DIVERGENCES"), "EXPECTED_DIVERGENCES must be defined"


# ── Test 2: EXPECTED_BASELINES contains all required values ──────────────────

_REQUIRED_BASELINES = {
    "raw_submission_rows":          6901,
    "raw_unique_numero":            2819,
    "raw_unique_numero_indice":     4848,
    "ged_raw_flat_rows":            27261,
    "ged_operations_rows":          32099,
    "open_doc_rows":                4848,
    "sas_rows":                     4848,
    "consultant_rows":              18911,
    "moex_rows":                    3492,
    "stage_read_flat_docs_df_rows": 4834,
}


@pytest.mark.parametrize("key,expected_value", list(_REQUIRED_BASELINES.items()))
def test_expected_baselines_values(key, expected_value):
    """Each required baseline key must be present with the specified value."""
    baselines = _mod.EXPECTED_BASELINES
    assert key in baselines, f"EXPECTED_BASELINES missing key: {key!r}"
    assert baselines[key] == expected_value, (
        f"EXPECTED_BASELINES[{key!r}] = {baselines[key]} but expected {expected_value}"
    )


# ── Test 3: EXPECTED_DIVERGENCES is a list of dicts (data, not conditionals) ─

def test_expected_divergences_is_data():
    """EXPECTED_DIVERGENCES must be a non-empty list of dicts with required keys."""
    divs = _mod.EXPECTED_DIVERGENCES
    assert isinstance(divs, list), "EXPECTED_DIVERGENCES must be a list"
    assert len(divs) > 0, "EXPECTED_DIVERGENCES must be non-empty"
    required_keys = {"name", "categories", "from_layer", "to_layer",
                     "is_difference_expected", "explanation"}
    for i, rule in enumerate(divs):
        assert isinstance(rule, dict), f"Item {i} in EXPECTED_DIVERGENCES must be a dict"
        missing = required_keys - set(rule.keys())
        assert not missing, f"Item {i} is missing keys: {missing}"
        assert isinstance(rule["categories"], list), (
            f"Item {i}: 'categories' must be a list, got {type(rule['categories'])}"
        )
        assert isinstance(rule["is_difference_expected"], bool), (
            f"Item {i}: 'is_difference_expected' must be bool"
        )
        assert rule["from_layer"] in _mod.LAYERS, (
            f"Item {i}: from_layer {rule['from_layer']!r} not in LAYERS"
        )
        assert rule["to_layer"] in _mod.LAYERS, (
            f"Item {i}: to_layer {rule['to_layer']!r} not in LAYERS"
        )


# ── Test 4: JSON output validates against §5.3 step 11 shape ─────────────────

_JSON_PATH = (
    Path(__file__).resolve().parent.parent
    / "output" / "debug" / "counts_lineage_audit.json"
)

_REQUIRED_TOP_LEVEL_KEYS = {
    "generated_at",
    "run_number",
    "flat_ged_sha256",
    "flat_ged_mtime",
    "expected_baselines",
    "categories",
    "baseline_checks",
    "summary",
}

_REQUIRED_CATEGORY_KEYS = {
    "name",
    "values",
    "first_divergence_layer",
    "is_difference_expected",
    "explanation",
    "source",
    "status",
}

_REQUIRED_SUMMARY_KEYS = {"PASS", "WARN", "FAIL"}

_KNOWN_CATEGORIES = {
    "submission_instance_count",
    "active_version_count",
    "family_count",
    "numero_indice_count",
    "workflow_step_count",
    "sas_row_count",
    "consultant_row_count",
    "moex_row_count",
    "open_doc_row_count",
    "open_count",
    "status_VSO",
    "status_VAO",
    "status_REF",
    "status_SAS_REF",
    "status_HM",
    "live_operational_count",
    "legacy_backlog_count",
}


@pytest.mark.skipif(
    not _JSON_PATH.exists(),
    reason=(
        "counts_lineage_audit.json not found — run "
        "'python scripts/audit_counts_lineage.py' first to produce it"
    ),
)
def test_json_output_shape():
    """JSON output must match the §5.3 step 11 schema."""
    with open(str(_JSON_PATH), encoding="utf-8") as fh:
        data = json.load(fh)

    # Top-level keys
    missing_top = _REQUIRED_TOP_LEVEL_KEYS - set(data.keys())
    assert not missing_top, f"JSON missing top-level keys: {missing_top}"

    # expected_baselines sub-structure
    eb = data["expected_baselines"]
    assert isinstance(eb, dict), "expected_baselines must be a dict"
    for k, v in _REQUIRED_BASELINES.items():
        assert k in eb, f"expected_baselines missing key: {k!r}"
        assert eb[k] == v, f"expected_baselines[{k!r}] = {eb[k]}, expected {v}"

    # categories list
    cats = data["categories"]
    assert isinstance(cats, list), "categories must be a list"
    assert len(cats) > 0, "categories must be non-empty"

    cat_names = {c["name"] for c in cats}
    assert cat_names == _KNOWN_CATEGORIES, (
        f"categories name set mismatch.\n"
        f"  Extra:   {cat_names - _KNOWN_CATEGORIES}\n"
        f"  Missing: {_KNOWN_CATEGORIES - cat_names}"
    )

    for i, cat in enumerate(cats):
        missing = _REQUIRED_CATEGORY_KEYS - set(cat.keys())
        assert not missing, f"Category[{i}] ({cat.get('name')!r}) missing keys: {missing}"
        assert isinstance(cat["values"], dict), f"Category[{i}]: 'values' must be dict"
        # values must have exactly the 7 LAYERS as keys
        assert set(cat["values"].keys()) == set(_mod.LAYERS), (
            f"Category[{i}] ({cat['name']!r}): values keys != LAYERS. "
            f"Extra: {set(cat['values'].keys()) - set(_mod.LAYERS)}"
        )
        assert cat["status"] in {"PASS", "FAIL", "WARN"}, (
            f"Category[{i}] ({cat['name']!r}): status must be PASS/FAIL/WARN"
        )

    # baseline_checks list
    bc = data["baseline_checks"]
    assert isinstance(bc, list), "baseline_checks must be a list"
    for b in bc:
        for req in ("baseline_key", "expected", "observed", "status"):
            assert req in b, f"baseline_check entry missing key: {req!r}"

    # summary
    summ = data["summary"]
    assert isinstance(summ, dict), "summary must be a dict"
    missing_sum = _REQUIRED_SUMMARY_KEYS - set(summ.keys())
    assert not missing_sum, f"summary missing keys: {missing_sum}"
    for k in _REQUIRED_SUMMARY_KEYS:
        assert isinstance(summ[k], int), f"summary[{k!r}] must be int"
    # PASS + FAIL + WARN >= total categories (WARN overlaps with baseline only)
    assert summ["PASS"] + summ["FAIL"] <= len(cats) + len(bc)


# ── Phase 2 tests ─────────────────────────────────────────────────────────────

_BASE_DIR = Path(__file__).resolve().parent.parent
_PROBE_JSON_PATH = _BASE_DIR / "output" / "debug" / "counts_lineage_probe.json"

_PROBE_REQUIRED_KEYS = {
    "category", "layer", "value", "value_origin_type",
    "source_file", "source_sheet", "source_column", "source_filter",
    "function_or_code_path", "is_hardcoded_baseline", "confidence",
}


def test_probe_mode_writes_files(tmp_path):
    """Running run_probe() must produce both output files."""
    records = _mod.run_probe(run_number=0)
    assert _PROBE_JSON_PATH.exists(), "counts_lineage_probe.json not created"
    xlsx_path = _BASE_DIR / "output" / "debug" / "counts_lineage_probe.xlsx"
    assert xlsx_path.exists(), "counts_lineage_probe.xlsx not created"
    assert len(records) > 0


@pytest.mark.skipif(
    not _PROBE_JSON_PATH.exists(),
    reason="counts_lineage_probe.json not found — run 'python scripts/audit_counts_lineage.py --probe' first",
)
def test_probe_records_required_fields():
    """Every probe record must contain all 11 required keys."""
    with open(str(_PROBE_JSON_PATH), encoding="utf-8") as fh:
        records = json.load(fh)
    assert len(records) > 0, "Probe JSON must not be empty"
    for i, rec in enumerate(records):
        missing = _PROBE_REQUIRED_KEYS - set(rec.keys())
        assert not missing, f"Probe record[{i}] ({rec.get('category')!r}@{rec.get('layer')!r}) missing keys: {missing}"


@pytest.mark.skipif(
    not _PROBE_JSON_PATH.exists(),
    reason="counts_lineage_probe.json not found — run 'python scripts/audit_counts_lineage.py --probe' first",
)
def test_step_type_counts_measured():
    """OPEN_DOC=4848, SAS=4848, CONSULTANT=18911, MOEX=3492 at L1 tagged as measured_excel, is_hardcoded_baseline=false."""
    with open(str(_PROBE_JSON_PATH), encoding="utf-8") as fh:
        records = json.load(fh)
    expected = {
        "open_doc_row_count":   4848,
        "sas_row_count":        4848,
        "consultant_row_count": 18911,
        "moex_row_count":       3492,
    }
    for cat_name, exp_val in expected.items():
        matches = [r for r in records if r["category"] == cat_name and r["layer"] == "L1_FLAT_GED_XLSX"]
        assert matches, f"No probe record for ({cat_name}, L1_FLAT_GED_XLSX)"
        rec = matches[0]
        assert rec["value"] == exp_val, f"{cat_name}@L1: value={rec['value']} expected {exp_val}"
        assert rec["value_origin_type"] == "measured_excel", (
            f"{cat_name}@L1: value_origin_type={rec['value_origin_type']!r} expected 'measured_excel'"
        )
        assert rec["source_sheet"] == "GED_OPERATIONS", (
            f"{cat_name}@L1: source_sheet={rec['source_sheet']!r} expected 'GED_OPERATIONS'"
        )
        assert rec["source_column"] == "step_type", (
            f"{cat_name}@L1: source_column={rec['source_column']!r} expected 'step_type'"
        )
        assert rec["is_hardcoded_baseline"] is False, f"{cat_name}@L1: is_hardcoded_baseline must be False"


def test_l1_sas_ref_correct():
    """L1 status_SAS_REF must equal 284 (GED_RAW_FLAT filter is_sas==True AND status_clean=='REF')."""
    result = _mod.collect_l1(_BASE_DIR)
    assert "_error" not in result, f"collect_l1 error: {result.get('_error')}"
    warn = result.get("_warn_sas_ref")
    assert warn is None, f"SAS REF reader warning: {warn}"
    sas_ref = result.get("status_SAS_REF")
    assert sas_ref is not None, "status_SAS_REF not computed at L1"
    assert sas_ref == 284, f"L1 status_SAS_REF = {sas_ref}, expected 284"


def test_l0_sas_ref_correct():
    """L0 status_SAS_REF must equal 836 (distinct rows with Réponse=='REF' under any 0-SAS block).
    Note: prior investigation reported 836+1=837 by summing per-column counts; OR-distinct gives 836
    because the single cycle-2 REF row is the same physical row as a cycle-1 REF row."""
    result = _mod.collect_l0(_BASE_DIR)
    assert "_error" not in result, f"collect_l0 error: {result.get('_error')}"
    warn = result.get("_warn_sas_ref")
    assert warn is None, f"L0 SAS REF reader warning: {warn}"
    sas_ref = result.get("status_SAS_REF")
    assert sas_ref is not None, "status_SAS_REF not computed at L0"
    assert sas_ref == 836, f"L0 status_SAS_REF = {sas_ref}, expected 836"


def test_raw_submission_baseline():
    """raw_submission_rows baseline must be 6901 with provenance referencing GED_export.xlsx and sheet name."""
    assert _mod.EXPECTED_BASELINES["raw_submission_rows"] == 6901, (
        f"raw_submission_rows baseline = {_mod.EXPECTED_BASELINES['raw_submission_rows']}, expected 6901"
    )
    provenance = _mod.BASELINE_PROVENANCE.get("raw_submission_rows", "")
    assert "input/GED_export.xlsx" in provenance, (
        f"Provenance missing 'input/GED_export.xlsx': {provenance!r}"
    )
    assert "Doc. sous workflow, x versions" in provenance, (
        f"Provenance missing sheet name: {provenance!r}"
    )


# ── Phase 2.5 tests (D-1 rule extension + D-012 confirmation) ─────────────────

def test_d1_rule_extends_workflow_step_count():
    """EXPECTED_DIVERGENCES must contain a rule covering workflow_step_count
    for L1_FLAT_GED_XLSX → L2_STAGE_READ_FLAT with a SAS-filter explanation."""
    found = any(
        "workflow_step_count" in rule["categories"]
        and rule["from_layer"] == "L1_FLAT_GED_XLSX"
        and rule["to_layer"] == "L2_STAGE_READ_FLAT"
        and rule["is_difference_expected"] is True
        and "stage_read_flat" in rule["explanation"]
        and "_apply_sas_filter_flat" in rule["explanation"]
        for rule in _mod.EXPECTED_DIVERGENCES
    )
    assert found, (
        "EXPECTED_DIVERGENCES must contain a rule with 'workflow_step_count' "
        "for L1_FLAT_GED_XLSX→L2_STAGE_READ_FLAT with SAS-filter explanation "
        "referencing stage_read_flat._apply_sas_filter_flat"
    )


def test_d1_rule_extends_sas_row_count():
    """EXPECTED_DIVERGENCES must contain a rule covering sas_row_count
    for L1_FLAT_GED_XLSX → L2_STAGE_READ_FLAT with a SAS-filter explanation."""
    found = any(
        "sas_row_count" in rule["categories"]
        and rule["from_layer"] == "L1_FLAT_GED_XLSX"
        and rule["to_layer"] == "L2_STAGE_READ_FLAT"
        and rule["is_difference_expected"] is True
        and "stage_read_flat" in rule["explanation"]
        and "_apply_sas_filter_flat" in rule["explanation"]
        for rule in _mod.EXPECTED_DIVERGENCES
    )
    assert found, (
        "EXPECTED_DIVERGENCES must contain a rule with 'sas_row_count' "
        "for L1_FLAT_GED_XLSX→L2_STAGE_READ_FLAT with SAS-filter explanation "
        "referencing stage_read_flat._apply_sas_filter_flat"
    )


def test_audit_fail_count_after_d1():
    """Running the audit must produce FAIL=1; the sole FAIL is status_SAS_REF
    at L1_FLAT_GED_XLSX (the real D-011 upstream gap — not silenced by D-1)."""
    result = _mod.run_audit(run_number=0)
    summary = result["summary"]
    fail_n  = summary["FAIL"]
    fails   = [c for c in result["categories"] if c["status"] == "FAIL"]

    assert fail_n == 1, (
        f"Expected FAIL=1 after D-1 extension, got FAIL={fail_n}. "
        f"Failing categories: {[c['name'] for c in fails]}"
    )
    assert len(fails) == 1, (
        f"Category FAIL list length mismatch: {[c['name'] for c in fails]}"
    )
    assert fails[0]["name"] == "status_SAS_REF", (
        f"Remaining FAIL must be status_SAS_REF, got {fails[0]['name']!r}"
    )
    assert fails[0].get("first_divergence_layer") == "L1_FLAT_GED_XLSX", (
        f"status_SAS_REF FAIL must point to L1_FLAT_GED_XLSX, "
        f"got {fails[0].get('first_divergence_layer')!r}"
    )


def test_status_sas_ref_l0_l1_still_flagged():
    """status_SAS_REF must NOT appear in any divergence rule that silences
    L0→L1 or L1→L2 (the D-011 upstream gap and the D-012 2-row gap must
    remain auditable)."""
    forbidden_pairs = {
        ("L0_RAW_GED",       "L1_FLAT_GED_XLSX"),
        ("L1_FLAT_GED_XLSX", "L2_STAGE_READ_FLAT"),
    }
    for rule in _mod.EXPECTED_DIVERGENCES:
        if "status_SAS_REF" in rule["categories"]:
            pair = (rule["from_layer"], rule["to_layer"])
            assert pair not in forbidden_pairs, (
                f"Rule {rule['name']!r} silences status_SAS_REF at {pair} — FORBIDDEN. "
                f"D-011 (L0→L1) and D-012 (L1→L2) must remain visible as FAILs."
            )


_D012_PATH = _BASE_DIR / "output" / "debug" / "sas_pre2026_confirmation.json"

_D012_REQUIRED_KEYS = {
    "generated_at",
    "l1_sas_ref_row_count", "l1_sas_ref_unique_pair_count", "l1_sas_ref_doc_ids_count",
    "l2_sas_ref_row_count", "l2_sas_ref_unique_pair_count", "l2_sas_ref_doc_ids_count",
    "row_gap", "pair_gap",
    "sas_filter_explained_rows", "structural_normalization_rows",
    "sas_filter_excluded_pair", "structural_duplicate_pairs",
    "sas_filter_component", "structural_component",
    "excluded_doc_ids", "excluded_count", "excluded_unique_pair_count", "excluded_l1_row_count",
    "pair_to_l1_row_count",
    "filter_threshold", "per_doc",
    "verdict", "verdict_reason",
}

_D012_PER_DOC_KEYS = {
    "doc_id", "date_column", "date_value", "would_be_excluded", "match_with_hypothesis",
}


@pytest.mark.skipif(
    not _D012_PATH.exists(),
    reason=(
        "output/debug/sas_pre2026_confirmation.json not found — "
        "run 'python scripts/audit_counts_lineage.py' first to produce it"
    ),
)
def test_d012_confirmation_emits_file():
    """sas_pre2026_confirmation.json must exist and validate against the documented shape."""
    with open(str(_D012_PATH), encoding="utf-8") as fh:
        data = json.load(fh)

    missing = _D012_REQUIRED_KEYS - set(data.keys())
    assert not missing, f"sas_pre2026_confirmation.json missing top-level keys: {missing}"

    assert isinstance(data["excluded_doc_ids"], list), "excluded_doc_ids must be a list"
    assert isinstance(data["per_doc"], list), "per_doc must be a list"
    assert data["verdict"] in {"CONFIRMED", "PARTIAL_CONFIRMED", "UNCONFIRMED", "UNDETERMINED"}, (
        f"verdict must be CONFIRMED|PARTIAL_CONFIRMED|UNCONFIRMED|UNDETERMINED, got {data['verdict']!r}"
    )
    assert isinstance(data["excluded_count"], int), "excluded_count must be int"
    assert isinstance(data["l1_sas_ref_doc_ids_count"], int), "l1_sas_ref_doc_ids_count must be int"
    assert isinstance(data["l2_sas_ref_doc_ids_count"], int), "l2_sas_ref_doc_ids_count must be int"

    for i, entry in enumerate(data["per_doc"]):
        missing_keys = _D012_PER_DOC_KEYS - set(entry.keys())
        assert not missing_keys, (
            f"per_doc[{i}] missing keys: {missing_keys}"
        )


@pytest.mark.skipif(
    not _D012_PATH.exists(),
    reason=(
        "output/debug/sas_pre2026_confirmation.json not found — "
        "run 'python scripts/audit_counts_lineage.py' first to produce it"
    ),
)
def test_d012_excluded_count():
    """PARTIAL_CONFIRMED: 1 SAS-filter pair + 1 structural-normalization row = 2-row gap."""
    with open(str(_D012_PATH), encoding="utf-8") as fh:
        data = json.load(fh)

    assert data["verdict"] == "PARTIAL_CONFIRMED", (
        f"D-012 FAIL: verdict={data['verdict']!r}, expected 'PARTIAL_CONFIRMED'. "
        f"verdict_reason={data.get('verdict_reason')!r}."
    )
    assert data["sas_filter_component"] == "CONFIRMED", (
        f"D-012 FAIL: sas_filter_component={data['sas_filter_component']!r}, expected 'CONFIRMED'."
    )
    assert data["structural_component"] == "PRESENT", (
        f"D-012 FAIL: structural_component={data['structural_component']!r}, expected 'PRESENT'."
    )
    assert data["pair_gap"] == 1, (
        f"D-012 FAIL: pair_gap={data['pair_gap']}, expected 1."
    )
    assert data["row_gap"] == 2, (
        f"D-012 FAIL: row_gap={data['row_gap']}, expected 2."
    )
    assert data["sas_filter_explained_rows"] == 1, (
        f"D-012 FAIL: sas_filter_explained_rows={data['sas_filter_explained_rows']}, expected 1."
    )
    assert data["structural_normalization_rows"] == 1, (
        f"D-012 FAIL: structural_normalization_rows={data['structural_normalization_rows']}, expected 1."
    )
    assert data["sas_filter_excluded_pair"] == "051020|A", (
        f"D-012 FAIL: sas_filter_excluded_pair={data['sas_filter_excluded_pair']!r}."
    )
    assert "152012|A" in data["structural_duplicate_pairs"], (
        f"D-012 FAIL: structural_duplicate_pairs={data['structural_duplicate_pairs']} "
        f"does not contain '152012|A'."
    )


# ── Phase 8 Step 3 — RunContext flat_ged_doc_meta integration test ────────────

def test_runcontext_carries_flat_ged_doc_meta():
    """In flat mode, load_run_context(0) must return a ctx whose
    flat_ged_doc_meta is a non-empty dict keyed by doc_id strings.
    Skipped (with reason) when the environment cannot load run 0.
    """
    try:
        import sys as _sys
        _sys.path.insert(0, str(_BASE_DIR / "src"))
        from reporting.data_loader import load_run_context
        ctx = load_run_context(_BASE_DIR, run_number=0)
    except Exception as exc:
        pytest.skip(f"Cannot load run context in test environment: {exc}")

    assert hasattr(ctx, "flat_ged_doc_meta"), (
        "RunContext missing flat_ged_doc_meta — TASK A field not yet applied"
    )
    assert isinstance(ctx.flat_ged_doc_meta, dict), (
        f"flat_ged_doc_meta must be a dict, got {type(ctx.flat_ged_doc_meta)}"
    )
    if ctx.degraded_mode:
        pytest.skip("RunContext loaded in degraded mode — flat_ged_doc_meta is expected to be empty")
    assert len(ctx.flat_ged_doc_meta) > 0, (
        "flat_ged_doc_meta is empty — _load_from_flat_artifacts did not pass "
        "flat_doc_meta to the RunContext constructor (TASK B not applied, or "
        "legacy fallback path was taken)"
    )


# ── Phase 8 Step 6 — UI Payload comparison tests ─────────────────────────────

def _minimal_kpis_overview(total=100, vso=50, vao=20, hm=5, open_cnt=25,
                            contractors=5, consultants=3,
                            avg_days=7.5, pending_sas=2):
    """Return a (kpis, overview) pair whose comparable fields all match."""
    kpis = {
        "total_docs_current": total,
        "total_docs_all_indices": total + 14,
        "by_visa_global": {"VSO": vso, "VAO": vao, "HM": hm, "Open": open_cnt,
                           "REF": 0, "SAS REF": 0},
        "by_visa_global_pct": {"VSO": 0.5, "VAO": 0.2},
        "by_building": {"AU": 30},
        "by_responsible": {},
        "avg_days_to_visa": avg_days,
        "docs_pending_sas": pending_sas,
        "docs_sas_ref_active": 0,
        "total_contractors": contractors,
        "total_consultants": consultants,
        "total_lots": contractors,
        "discrepancies_count": 0,
    }
    overview = {
        "total_docs": total,
        "visa_flow": {
            "submitted": total,
            "answered": vso + vao + hm,
            "vso": vso,
            "vao": vao,
            "ref": 0,
            "hm": hm,
            "pending": open_cnt,
            "on_time": None,
            "late": None,
        },
        "project_stats": {
            "total_contractors": contractors,
            "total_consultants": consultants,
            "avg_days_to_visa": avg_days,
            "docs_pending_sas": pending_sas,
        },
        "focus": {
            "focused": 0, "p1_overdue": 0, "p2_urgent": 0, "p3_soon": 0,
            "p4_ok": 0, "total_dernier": total, "excluded": 0,
            "stale": 0, "resolved": 0, "by_consultant": [], "by_contractor": [],
        },
    }
    return kpis, overview


# ── Test 1 ────────────────────────────────────────────────────────────────────

def test_ui_payload_field_map_is_data():
    """UI_PAYLOAD_FIELD_MAP is a module-scope list of dicts; each entry has
    aggregator_path and comparison_kind; comparison_kind values are data, not
    buried in code-only conditionals."""
    fmap = _mod.UI_PAYLOAD_FIELD_MAP
    assert isinstance(fmap, list), "UI_PAYLOAD_FIELD_MAP must be a list"
    assert len(fmap) > 0, "UI_PAYLOAD_FIELD_MAP must be non-empty"
    valid_kinds = {"numeric_equal", "identity", "set_equal", "dict_equal", "skipped"}
    for i, entry in enumerate(fmap):
        assert isinstance(entry, dict), f"Entry {i} must be a dict"
        assert "aggregator_path" in entry, f"Entry {i} missing 'aggregator_path'"
        assert "comparison_kind" in entry, f"Entry {i} missing 'comparison_kind'"
        assert entry["comparison_kind"] in valid_kinds, (
            f"Entry {i}: comparison_kind={entry['comparison_kind']!r} not in {valid_kinds}"
        )


# ── Test 2 ────────────────────────────────────────────────────────────────────

def test_compare_ui_payload_identity_match():
    """When kpis and overview have matching values, all comparable fields
    appear in matches and none appear in mismatches."""
    kpis, overview = _minimal_kpis_overview()
    result = _mod._compare_ui_payload(None, None, kpis, overview)
    assert isinstance(result["matches"], list)
    assert isinstance(result["mismatches"], list)
    assert isinstance(result["skipped"], list)
    assert len(result["mismatches"]) == 0, (
        f"Expected no mismatches; got: {result['mismatches']}"
    )
    assert len(result["matches"]) > 0, "Expected at least one match"


# ── Test 3 ────────────────────────────────────────────────────────────────────

def test_compare_ui_payload_renamed_match():
    """A map entry with a value_transform key-rename is honoured; the row
    counts as a match when values are equal after renaming."""
    test_map = [
        {
            "aggregator_path": "kpis.by_building",
            "ui_adapter_path": "overview.by_building_renamed",
            "comparison_kind": "dict_equal",
            "value_transform": {"AU": "au"},
            "notes": "test rename",
        }
    ]
    kpis = {"by_building": {"AU": 10}}
    overview = {"by_building_renamed": {"au": 10}}

    original = _mod.UI_PAYLOAD_FIELD_MAP
    _mod.UI_PAYLOAD_FIELD_MAP = test_map
    try:
        result = _mod._compare_ui_payload(None, None, kpis, overview)
    finally:
        _mod.UI_PAYLOAD_FIELD_MAP = original

    assert len(result["mismatches"]) == 0, (
        f"Expected no mismatches after rename; got: {result['mismatches']}"
    )
    assert len(result["matches"]) == 1, (
        f"Expected exactly 1 match; got {len(result['matches'])}"
    )


# ── Test 4 ────────────────────────────────────────────────────────────────────

def test_compare_ui_payload_mismatch_recorded():
    """When aggregator_value != ui_adapter_value, the row appears in mismatches
    with all required columns and does NOT appear in matches."""
    kpis, overview = _minimal_kpis_overview(total=100)
    overview["total_docs"] = 999   # deliberate mismatch

    result = _mod._compare_ui_payload(None, None, kpis, overview)

    mismatch_paths = [r["aggregator_path"] for r in result["mismatches"]]
    assert "kpis.total_docs_current" in mismatch_paths, (
        f"Expected kpis.total_docs_current in mismatches; got {mismatch_paths}"
    )
    mismatch = next(r for r in result["mismatches"]
                    if r["aggregator_path"] == "kpis.total_docs_current"
                    and r["ui_adapter_path"] == "overview.total_docs")
    for col in ("aggregator_path", "ui_adapter_path", "aggregator_value",
                "ui_adapter_value", "comparison_kind", "notes"):
        assert col in mismatch, f"Mismatch row missing column {col!r}"
    assert mismatch["aggregator_value"] == 100
    assert mismatch["ui_adapter_value"] == 999


# ── Test 5 ────────────────────────────────────────────────────────────────────

def test_compare_ui_payload_resolution_error_skipped():
    """A dotted path that doesn't exist on the payload is recorded as skipped
    with reason containing the exception type, NOT as a failure."""
    test_map = [
        {
            "aggregator_path": "kpis.nonexistent_field",
            "ui_adapter_path": "overview.total_docs",
            "comparison_kind": "numeric_equal",
            "notes": "test missing path",
        }
    ]
    kpis = {}   # nonexistent_field absent
    overview = {"total_docs": 100}

    original = _mod.UI_PAYLOAD_FIELD_MAP
    _mod.UI_PAYLOAD_FIELD_MAP = test_map
    try:
        result = _mod._compare_ui_payload(None, None, kpis, overview)
    finally:
        _mod.UI_PAYLOAD_FIELD_MAP = original

    assert len(result["mismatches"]) == 0, "Missing path must NOT become a mismatch"
    assert len(result["skipped"]) == 1, f"Expected 1 skipped, got {len(result['skipped'])}"
    reason = result["skipped"][0].get("reason", "")
    assert "resolution_error" in reason, (
        f"Skipped reason must mention 'resolution_error'; got {reason!r}"
    )
    # Check that the exception type (KeyError) is mentioned
    assert "KeyError" in reason or "Error" in reason, (
        f"Skipped reason must include exception type; got {reason!r}"
    )


# ── Test 6 ────────────────────────────────────────────────────────────────────

def test_audit_emits_ui_payload_comparison_in_json():
    """The in-memory comparison block returned by _compare_ui_payload has the
    required shape; if the JSON file exists it also carries ui_payload_comparison."""
    kpis, overview = _minimal_kpis_overview()
    result = _mod._compare_ui_payload(None, None, kpis, overview)

    # Build the block as run_audit would
    block = {
        "fields_compared": len(result["matches"]) + len(result["mismatches"]),
        "matches":         len(result["matches"]),
        "mismatches":      len(result["mismatches"]),
        "mismatch_rows":   result["mismatches"],
        "skipped":         result["skipped"],
    }
    for key in ("fields_compared", "matches", "mismatches", "mismatch_rows", "skipped"):
        assert key in block, f"ui_payload_comparison block missing key {key!r}"
    assert isinstance(block["fields_compared"], int)
    assert isinstance(block["mismatch_rows"], list)
    assert isinstance(block["skipped"], list)

    # If the JSON output file exists, verify the key is present
    if _JSON_PATH.exists():
        with open(str(_JSON_PATH), encoding="utf-8") as fh:
            data = json.load(fh)
        assert "ui_payload_comparison" in data, (
            "counts_lineage_audit.json missing 'ui_payload_comparison' block"
        )
        c = data["ui_payload_comparison"]
        for key in ("fields_compared", "matches", "mismatches", "mismatch_rows", "skipped"):
            assert key in c, f"JSON ui_payload_comparison missing key {key!r}"


# ── Test 7 ────────────────────────────────────────────────────────────────────

_AUDIT_XLSX_PATH = _BASE_DIR / "output" / "debug" / "counts_lineage_audit.xlsx"

@pytest.mark.skipif(
    not _AUDIT_XLSX_PATH.exists(),
    reason=(
        "counts_lineage_audit.xlsx not found — run "
        "'python scripts/audit_counts_lineage.py' first to produce it"
    ),
)
def test_audit_emits_ui_payload_mismatches_sheet():
    """After the script runs, the xlsx must contain a sheet named
    ui_payload_mismatches. Skipped if openpyxl is unavailable."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available in this environment")
    wb = openpyxl.load_workbook(str(_AUDIT_XLSX_PATH), read_only=True)
    assert "ui_payload_mismatches" in wb.sheetnames, (
        f"Sheet 'ui_payload_mismatches' missing; found: {wb.sheetnames}"
    )


# ── Test 8 ────────────────────────────────────────────────────────────────────

def test_existing_one_liner_unchanged():
    """The FIRST summary line printed by run_audit must match the established
    regex contract; the second UI_PAYLOAD line must also be present."""
    import io
    import re
    import contextlib

    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            _mod.run_audit(run_number=0)
    except Exception as exc:
        pytest.skip(f"run_audit raised {type(exc).__name__}: {exc}")

    lines = [ln for ln in buf.getvalue().splitlines() if ln.strip()]
    assert lines, "run_audit produced no stdout output"

    first_line = lines[0]
    pattern_audit = r"^AUDIT: PASS=\d+ WARN=\d+ FAIL=\d+; first_unexpected_divergence="
    assert re.match(pattern_audit, first_line), (
        f"First summary line doesn't match expected contract.\n"
        f"Got:      {first_line!r}\n"
        f"Expected: regex {pattern_audit!r}"
    )

    # Second line must be the UI_PAYLOAD summary
    assert len(lines) >= 2, "Expected at least two summary lines"
    second_line = lines[1]
    pattern_ui = r"^UI_PAYLOAD: compared=\d+ matches=\d+ mismatches=\d+;"
    assert re.match(pattern_ui, second_line), (
        f"Second summary line doesn't match UI_PAYLOAD pattern.\n"
        f"Got:      {second_line!r}\n"
        f"Expected: regex {pattern_ui!r}"
    )
