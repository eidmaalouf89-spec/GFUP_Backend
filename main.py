"""
main.py
-------
GF UPDATER V3 — Main Pipeline Orchestrator (Patch A+B+C+D+E)
Part E: MISSING_IN_GED/GF subtypes, GF SAS REF rule, smart MISSING_IN_GF classification

Usage:
  python main.py

Inputs (in ./input/):
  - GED_export.xlsx     : GED raw export
  - Grandfichier_v2.xlsx : Existing GF (for structure reference + discrepancy check)
  - Approver mapping    : Hardcoded in src/normalize.py

Outputs (in ./output/):
  - GF_V0_CLEAN.xlsx                : Rebuilt GF
  - DISCREPANCY_REPORT.xlsx         : GED vs existing GF (all, with severity column)
  - DISCREPANCY_REVIEW_REQUIRED.xlsx: Only REVIEW_REQUIRED discrepancies
  - ANOMALY_REPORT.xlsx             : Per-contractor anomaly flags
  - AUTO_RESOLUTION_LOG.xlsx
  - IGNORED_ITEMS_LOG.xlsx

Debug (in ./output/debug/):
  - routing_summary.xlsx
  - exclusion_summary.xlsx
  - gf_sheet_schema.xlsx
  - family_clusters.xlsx
  - lifecycle_resolution.xlsx
  - discrepancy_sample.xlsx
  - coarse_groups.xlsx
"""

import sys
import traceback
import json
import hashlib
from pathlib import Path

import pandas as pd

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

# Report memory imports — loaded here so they are available inside run_pipeline()
from report_memory import (
    init_report_memory_db,
    sha256_file,
    is_report_already_ingested,
    register_ingested_report,
    upsert_report_responses,
    load_persisted_report_responses,
)
from effective_responses import build_effective_responses
from run_memory import (
    init_run_memory_db as init_run_memory_db_fn,
    baseline_run_exists,
    get_next_run_number,
    create_run,
    update_run_metadata,
    finalize_run_success,
    finalize_run_failure,
    mark_run_current,
    register_run_input,
    register_run_artifact,
    copy_artifact_to_run_dir,
    get_run_dir,
    get_current_run,
)

for _stream_name in ("stdout", "stderr"):
    _stream = getattr(sys, _stream_name, None)
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

from read_raw import read_ged
from normalize import (
    load_mapping,
    normalize_docs,
    normalize_responses,
)
from version_engine import VersionEngine
from workflow_engine import WorkflowEngine
from config_loader import load_exclusion_config
from routing import (
    build_routing_table,
    read_gf_sheet_structure,
    read_all_gf_sheet_structures,
    route_documents,
    build_approver_match_map,
    build_gf_to_ged_map,
    write_routing_summary,
)
from writer import (
    GFWriter,
    write_discrepancy_report,
    write_anomaly_report,
    write_auto_resolution_log,
    write_ignored_items_log,
    write_insert_log,
    write_new_submittal_analysis,
)
from debug_writer import write_all_debug
from reconciliation_engine import run_reconciliation, write_reconciliation_outputs


def _safe_console_print(*args, **kwargs):
    """
    Print safely on Windows consoles using legacy encodings.
    """
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        sep = kwargs.get("sep", " ")
        end = kwargs.get("end", "\n")
        text = sep.join(str(arg) for arg in args)
        sanitized = (
            text.replace("—", "-")
            .replace("–", "-")
            .replace("â€”", "-")
            .replace("→", "->")
            .replace("â†’", "->")
            .replace("✅", "[OK]")
            .replace("âœ…", "[OK]")
            .replace("❌", "[ERROR]")
            .replace("âŒ", "[ERROR]")
            .replace("⚠️", "[WARN]")
            .replace("⚠", "[WARN]")
        )
        sys.stdout.write(sanitized + end)


def _build_input_signature(input_entries: list[dict]) -> str:
    """Build a stable hash of the concrete inputs that affected a run."""
    payload = json.dumps(input_entries, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


# ─────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

# Persistent report memory DB — created automatically on first run
REPORT_MEMORY_DB = str(BASE_DIR / "data" / "report_memory.db")

# Run-history database — tracks every pipeline execution and its artifacts
RUN_MEMORY_DB = str(BASE_DIR / "data" / "run_memory.db")

# Location of consultant source PDF files (used for SHA-256 deduplication)
CONSULTANT_REPORTS_ROOT = INPUT_DIR / "consultant_reports"

# Output of consultant_integration.py — used as the source for report responses
# when new ingestion is triggered inside main.py
CONSULTANT_MATCH_REPORT = OUTPUT_DIR / "consultant_match_report.xlsx"

GED_FILE = INPUT_DIR / "GED_export.xlsx"
GF_FILE = INPUT_DIR / "Grandfichier_v3.xlsx"  # NEW BASELINE (P17-T2-VISA-Tableau de suivi (2).xlsx)

OUTPUT_GF = OUTPUT_DIR / "GF_V0_CLEAN.xlsx"
OUTPUT_DISCREPANCY = OUTPUT_DIR / "DISCREPANCY_REPORT.xlsx"
OUTPUT_DISCREPANCY_REVIEW = OUTPUT_DIR / "DISCREPANCY_REVIEW_REQUIRED.xlsx"
OUTPUT_ANOMALY = OUTPUT_DIR / "ANOMALY_REPORT.xlsx"
OUTPUT_AUTO_RESOLUTION = OUTPUT_DIR / "AUTO_RESOLUTION_LOG.xlsx"
OUTPUT_IGNORED = OUTPUT_DIR / "IGNORED_ITEMS_LOG.xlsx"

DEBUG_DIR = OUTPUT_DIR / "debug"

OUTPUT_DIR.mkdir(exist_ok=True)
DEBUG_DIR.mkdir(exist_ok=True)

# ── New diagnosis output paths (Part E) ──────────────────────
OUTPUT_MISSING_GED_DIAGNOSIS  = OUTPUT_DIR / "MISSING_IN_GED_DIAGNOSIS.xlsx"
OUTPUT_MISSING_GED_TRUE       = OUTPUT_DIR / "MISSING_IN_GED_TRUE_ONLY.xlsx"
OUTPUT_MISSING_GF_DIAGNOSIS   = OUTPUT_DIR / "MISSING_IN_GF_DIAGNOSIS.xlsx"
OUTPUT_MISSING_GF_TRUE        = OUTPUT_DIR / "MISSING_IN_GF_TRUE_ONLY.xlsx"

# ── Reconciliation engine output paths (Patch F) ──────────────
OUTPUT_RECONCILIATION_LOG     = OUTPUT_DIR / "RECONCILIATION_LOG.xlsx"
OUTPUT_RECONCILIATION_SUMMARY = DEBUG_DIR / "reconciliation_summary.xlsx"

# ── INSERT LOG: newly inserted rows in CLEAN GF ───────────────
OUTPUT_INSERT_LOG = OUTPUT_DIR / "INSERT_LOG.xlsx"

# ── NEW SUBMITTAL ANALYSIS ────────────────────────────────────
OUTPUT_NEW_SUBMITTAL_ANALYSIS = OUTPUT_DIR / "NEW_SUBMITTAL_ANALYSIS.xlsx"
OUTPUT_NEW_SUBMITTAL_SUMMARY  = DEBUG_DIR  / "new_submittal_summary.xlsx"
OUTPUT_CONSULTANT_REPORTS_WB  = OUTPUT_DIR / "consultant_reports.xlsx"
OUTPUT_GF_STAGE1             = OUTPUT_DIR / "GF_consultant_enriched_stage1.xlsx"
OUTPUT_GF_STAGE2             = OUTPUT_DIR / "GF_consultant_enriched_stage2.xlsx"
OUTPUT_GF_TEAM_VERSION       = OUTPUT_DIR / "GF_TEAM_VERSION.xlsx"
OUTPUT_SUSPICIOUS_ROWS       = OUTPUT_DIR / "SUSPICIOUS_ROWS_REPORT.xlsx"
RUN_MEMORY_CORE_VERSION     = "P1"

_ACTIVE_RUN_NUMBER = None
_ACTIVE_RUN_FINALIZED = False
_RUN_CONTROL_CONTEXT = None


# ─────────────────────────────────────────────────────────────
# PATCH B — NORMALIZATION HELPERS
# ─────────────────────────────────────────────────────────────

import re as _re
import datetime as _dt
from typing import Optional as _Optional
from collections import defaultdict as _defaultdict


def normalize_date_for_compare(value) -> str:
    """
    Patch B1: Normalize a date value to YYYY-MM-DD string for comparison.
    Strips time-of-day component — only calendar date matters here.

    Returns '' (empty string) for None/invalid values.

    Examples:
      '2026-03-13 08:23:24' → '2026-03-13'
      '2026-03-13 00:00:00' → '2026-03-13'
      datetime(2026, 3, 13)  → '2026-03-13'
      None / '' / 'N/A'     → ''
    """
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return str(value.date() if isinstance(value, _dt.datetime) else value)
    s = str(value).strip()
    if not s or s.lower() in ("none", "n/a", ""):
        return ""
    try:
        return str(pd.to_datetime(s).date())
    except Exception:
        return ""


def normalize_title_for_compare(value) -> str:
    """
    Part 3 (v2): Normalize a document title for similarity comparison.

    1. Strip GED code-path prefix before ' - '
       'P17_T2_AU_EXE_LGD_GOE_A003_ARM_AZ_R7_228193_A - Armatures Poteaux'
       → 'Armatures Poteaux'
    2. Lowercase
    3. Remove file extensions (.pdf .docx .xlsx .dwg .txt .pptx .zip)
    4. Normalize separators: underscores/dashes → space
    5. Collapse whitespace
    6. Strip trailing/leading punctuation
    7. Remove consecutive duplicate tokens
    """
    if value is None:
        return ""
    s = str(value).strip()

    prefix_match = _re.match(r'^[A-Z0-9][A-Z0-9_\-]{10,}\s+-\s+(.+)$', s)
    if prefix_match:
        s = prefix_match.group(1).strip()

    s = s.lower()
    s = _re.sub(r'\.(pdf|docx?|xlsx?|dwg|txt|pptx?|zip)\s*$', '', s, flags=_re.IGNORECASE)
    s = _re.sub(r'[_\-]+', ' ', s)
    s = _re.sub(r'\s+', ' ', s)
    s = s.strip('.,;: ')

    tokens = s.split()
    deduped: list = []
    for t in tokens:
        if not deduped or t != deduped[-1]:
            deduped.append(t)
    return ' '.join(deduped)


def title_similarity(a: str, b: str) -> float:
    """
    Token-based similarity: max(Jaccard, Containment).

    Containment = |A∩B| / min(|A|, |B|) — gives 1.0 when one set is a
    subset of the other, so a short GED title that is entirely contained
    in a longer GF title (e.g. GF appends building/level info) scores 1.0.

    Thresholds (used in classify_discrepancy):
      >= 0.85 → suppress (COSMETIC or no discrepancy)
      0.65–0.84 → COSMETIC
      < 0.65 → REVIEW_REQUIRED
    """
    na = normalize_title_for_compare(a)
    nb = normalize_title_for_compare(b)
    if not na and not nb:
        return 1.0
    if not na or not nb:
        return 0.0
    ta = set(na.split())
    tb = set(nb.split())
    if not ta and not tb:
        return 1.0
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union_len = len(ta | tb)
    jaccard = inter / union_len if union_len else 0.0
    containment = inter / min(len(ta), len(tb)) if min(len(ta), len(tb)) else 0.0
    return max(jaccard, containment)


def date_diff_days(ged_date, gf_date) -> _Optional[int]:
    """
    Absolute calendar-day difference between two date values.
    Returns None if either is missing or unparseable.
    """
    gd = normalize_date_for_compare(ged_date)
    gf = normalize_date_for_compare(gf_date)
    if not gd or not gf:
        return None
    try:
        d1 = _dt.date.fromisoformat(gd)
        d2 = _dt.date.fromisoformat(gf)
        return abs((d1 - d2).days)
    except Exception:
        return None


def normalize_status_for_compare(value) -> str:
    """
    Patch B3: Normalize a visa/workflow status for comparison.
    Removes leading dots, uppercases, trims spaces.
    """
    if value is None:
        return ""
    s = str(value).strip().lstrip(".").strip().upper()
    return s


def normalize_numero_for_compare(value) -> str:
    """
    Normalize document numero: strip leading zeros and .0 suffix, cast to string.

    Handles:
      int     49202   → '49202'
      float   49202.0 → '49202'   (new GF stores as float)
      str    '49202'  → '49202'
      str    '49202.0'→ '49202'   (via float→int)
    """
    if value is None:
        return ""
    try:
        # float() handles both int-strings and float-strings ('49202', '49202.0')
        return str(int(float(str(value).replace(",", "").strip())))
    except (ValueError, TypeError):
        return str(value).strip()


def normalize_indice_for_compare(value) -> str:
    """Normalize indice: strip whitespace, uppercase."""
    if value is None:
        return ""
    return str(value).strip().upper()


# ─────────────────────────────────────────────────────────────
# PATCH D — DISCREPANCY SEVERITY CLASSIFIER
# ─────────────────────────────────────────────────────────────

def classify_discrepancy(record: dict) -> str:
    """
    Part 6 / Patch E: Unified severity classification.

    REVIEW_REQUIRED — operational truth affected; human must act
    COSMETIC        — values differ but meaning is equivalent
    INFO            — non-blocking observation (historical rows, minor diffs)
    EXCLUDED        — excluded population (traceability only)
    """
    flag_type = record.get("flag_type", "")
    is_excluded = record.get("is_excluded_population", False)

    if is_excluded:
        return "EXCLUDED"

    # ── Final pass: BENTIN legacy exceptions ─────────────────
    if flag_type == "BENTIN_LEGACY_EXCEPTION":
        return "EXCLUDED"

    # ── Hard REVIEW_REQUIRED ──────────────────────────────────
    if flag_type == "SHEET_MISMATCH":
        return "REVIEW_REQUIRED"

    # ── Part 1: INDICE_MISMATCH is no longer blocking ────────
    # Identity is confirmed by emetteur/lot match; indice differences are cosmetic.
    if flag_type == "INDICE_MISMATCH":
        return "INFO"

    # Legacy flat types (backward compat — should not appear after Patch E)
    if flag_type in ("MISSING_IN_GF", "MISSING_IN_GED"):
        return "REVIEW_REQUIRED"

    # ── MISSING_IN_GF subtypes (Part 5 / Patch E) ────────────
    if flag_type == "MISSING_IN_GF_TRUE":
        return "REVIEW_REQUIRED"
    if flag_type == "MISSING_IN_GF_SAME_KEY_COLLISION":
        return "REVIEW_REQUIRED"
    if flag_type in (
        "MISSING_IN_GF_PENDING_SAS",
        "MISSING_IN_GF_RECENT_SAS_REMINDER",
        "MISSING_IN_GF_RECENT_REFUSAL",
        "MISSING_IN_GF_RECENT_ACCEPTED_SAS",
        "MISSING_IN_GF_AMBIGUOUS",
    ):
        return "INFO"

    # ── MISSING_IN_GED subtypes (Part 4 / Patch E) ────────────
    if flag_type == "MISSING_IN_GED_TRUE":
        return "REVIEW_REQUIRED"
    if flag_type == "MISSING_IN_GED_FAMILY_MATCH_MISSED":
        return "REVIEW_REQUIRED"
    if flag_type in (
        "MISSING_IN_GED_HISTORICAL",
        "MISSING_IN_GED_GF_SAS_REF",
        "MISSING_IN_GED_GF_DUPLICATE_ROW",
        "MISSING_IN_GED_EXCLUDED",
        "MISSING_IN_GED_AMBIGUOUS",
    ):
        return "INFO"

    # ── Reconciliation subtypes (Patch F — numero-based) ─────
    if flag_type == "MISSING_IN_GED_HISTORICAL":
        return "INFO"
    if flag_type in (
        "MISSING_IN_GED_RECONCILED_BY_FUZZY",
        "MISSING_IN_GF_RECONCILED_BY_FUZZY",
    ):
        return "INFO"
    if flag_type in (
        "MISSING_IN_GED_POSSIBLE_NUMERO_ERROR",
        "MISSING_IN_GF_POSSIBLE_NUMERO_ERROR",
    ):
        return "REVIEW_REQUIRED"
    if flag_type in (
        "MISSING_IN_GED_POSSIBLE_TITLE_VARIANT",
        "MISSING_IN_GF_POSSIBLE_TITLE_VARIANT",
    ):
        return "REVIEW_REQUIRED"
    if flag_type in (
        "MISSING_IN_GED_AMBIGUOUS_RECONCILIATION",
        "MISSING_IN_GF_AMBIGUOUS_RECONCILIATION",
    ):
        return "REVIEW_REQUIRED"

    # ── Reconciliation subtypes (Patch G — title-first) ──────
    # Confirmed GF numero typo: same document, emetteur, title, indice — only numero differs
    if flag_type in (
        "MISSING_IN_GED_GF_NUMERO_TYPO_CONFIRMED",
        "MISSING_IN_GF_GF_NUMERO_TYPO",
    ):
        return "INFO"   # GED is truth; GF had a typo — not a real missing
    # Reconciled by title (routing mismatch or naming variant)
    if flag_type in (
        "MISSING_IN_GED_RECONCILED_BY_TITLE",
        "MISSING_IN_GF_RECONCILED_BY_TITLE",
    ):
        return "INFO"
    # Ambiguous title match — needs human check but not REVIEW_REQUIRED urgency
    if flag_type in (
        "MISSING_IN_GED_AMBIGUOUS_TITLE_MATCH",
        "MISSING_IN_GF_AMBIGUOUS_TITLE_MATCH",
    ):
        return "REVIEW_REQUIRED"

    # ── Patch C: title variant accepted (TITRE_MISMATCH relaxed) ──
    if flag_type == "TITRE_VARIANT_ACCEPTED":
        return "INFO"

    # ── Patch D: indice variant accepted for reconstruction ──────
    if flag_type == "INDICE_VARIANT_ACCEPTED_BY_GED":
        return "COSMETIC"

    # ── Other INFO types ──────────────────────────────────────
    if flag_type == "DUPLICATE_ACTIVE_IN_GF":
        return "INFO"

    # ── Part 1: TITRE_MISMATCH is no longer blocking ────────────
    # Title is not a reliable identifier; emetteur+lot+numero confirm identity.
    # Kept in logs for traceability but downgraded from REVIEW_REQUIRED.
    if flag_type == "TITRE_MISMATCH":
        return "INFO"

    # ── Part 4: DATE_MISMATCH is no longer blocking ─────────────
    # Date cannot be trusted as a matching field; title determines identity.
    # Remove from REVIEW_REQUIRED; keep as INFO for traceability.
    if flag_type == "DATE_MISMATCH":
        return "INFO"

    return "INFO"


# ─────────────────────────────────────────────────────────────
# PART 1 — SAS FILTER: RAPPEL_EN_ATTENTE + PRE-2026
# ─────────────────────────────────────────────────────────────

def _apply_sas_filter(
    docs_df: pd.DataFrame,
    responses_df: pd.DataFrame,
) -> tuple:
    """
    Business rule: Documents that are simultaneously:
      (a) still pending SAS visa with a reminder ("Rappel : En attente visa ..."),
      (b) created BEFORE 2026

    are considered INVALID (old unresolved SAS items) and must be removed
    entirely from processing — before version engine, routing, discrepancy.

    Detection:
      - The '0-SAS' approver column in responses_df is the SAS visa track
      - If response_date_raw for 0-SAS contains "rappel" (case-insensitive)
        → the SAS visa is still pending with a reminder = stale unresolved
      - Cross-reference doc created_at year < 2026

    Returns:
      (docs_df_clean, responses_df_clean, sas_filtered_df)
    """
    SAS_APPROVER = "0-SAS"
    RAPPEL_TOKEN = "rappel"

    # Step 1: Find doc_ids where 0-SAS has "Rappel" in date_raw
    sas_rows = responses_df[responses_df["approver_raw"] == SAS_APPROVER].copy()
    sas_rappel = sas_rows[
        sas_rows["response_date_raw"].apply(
            lambda x: isinstance(x, str) and RAPPEL_TOKEN in str(x).lower()
        )
    ]
    sas_rappel_doc_ids: set = set(sas_rappel["doc_id"].unique())

    if not sas_rappel_doc_ids:
        return docs_df, responses_df, pd.DataFrame()

    # Step 2: Cross with docs created before 2026
    sas_candidates = docs_df[docs_df["doc_id"].isin(sas_rappel_doc_ids)].copy()
    sas_candidates["_year"] = pd.to_datetime(
        sas_candidates["created_at"], errors="coerce"
    ).dt.year
    invalid_mask = sas_candidates["_year"] < 2026
    invalid_doc_ids: set = set(sas_candidates.loc[invalid_mask, "doc_id"].tolist())

    if not invalid_doc_ids:
        return docs_df, responses_df, pd.DataFrame()

    # Step 3: Extract filtered rows for logging
    sas_filtered_df = docs_df[docs_df["doc_id"].isin(invalid_doc_ids)].copy()
    sas_filtered_df["sas_status"] = "RAPPEL_EN_ATTENTE"

    # Step 4: Remove from both DataFrames
    docs_clean = docs_df[~docs_df["doc_id"].isin(invalid_doc_ids)].copy()
    responses_clean = responses_df[~responses_df["doc_id"].isin(invalid_doc_ids)].copy()

    return docs_clean, responses_clean, sas_filtered_df


# ─────────────────────────────────────────────────────────────
# PART E — SAS LOOKUP HELPERS (for MISSING subtypes)
# ─────────────────────────────────────────────────────────────

def _determine_data_date(ged_file_path: str) -> _dt.date:
    """
    Returns the reference date for recency checks.
    Uses GED file mtime; falls back to today.
    """
    import os
    try:
        mtime = os.path.getmtime(str(ged_file_path))
        return _dt.datetime.fromtimestamp(mtime).date()
    except Exception:
        return _dt.date.today()


def _build_sas_lookup(responses_df: pd.DataFrame) -> dict:
    """
    Build per-doc SAS status lookup from GED responses.

    Returns: {doc_id: {sas_status_type, sas_result, sas_date, has_rappel}}
      sas_status_type: ANSWERED | PENDING_IN_DELAY | NOT_CALLED | RAPPEL
      sas_result:      VSO-SAS | REF | VSO | HM | RAPPEL | None
      sas_date:        date object (response date) or None
      has_rappel:      bool
    """
    SAS_APPROVER = "0-SAS"
    RAPPEL_TOKEN = "rappel"
    DATE_RE      = _re.compile(r'\((\d{4}/\d{2}/\d{2})\)')

    sas = responses_df[responses_df["approver_raw"] == SAS_APPROVER].copy()
    if sas.empty:
        return {}

    lookup: dict = {}
    for doc_id, grp in sas.groupby("doc_id"):
        # Check for rappel in any response_date_raw
        has_rappel = grp["response_date_raw"].apply(
            lambda x: isinstance(x, str) and RAPPEL_TOKEN in str(x).lower()
        ).any()

        answered = grp[grp["date_status_type"] == "ANSWERED"]
        pending  = grp[grp["date_status_type"] == "PENDING_IN_DELAY"]

        if has_rappel:
            sas_status_type = "RAPPEL"
            sas_result      = "RAPPEL"
            sas_date        = None
            # Extract date from "Rappel : En attente visa (YYYY/MM/DD)"
            for _, r in grp[grp["response_date_raw"].apply(
                lambda x: isinstance(x, str) and RAPPEL_TOKEN in str(x).lower()
            )].iterrows():
                m = DATE_RE.search(str(r["response_date_raw"]))
                if m:
                    try:
                        sas_date = _dt.date.fromisoformat(m.group(1).replace("/", "-"))
                    except Exception:
                        pass
                    break
        elif not answered.empty:
            sas_status_type = "ANSWERED"
            best = answered.sort_values("date_answered", ascending=False, na_position="last").iloc[0]
            sas_result = str(best.get("status_clean") or best.get("response_status_raw") or "")
            raw_date   = best.get("date_answered")
            try:
                sas_date = pd.to_datetime(raw_date).date()
            except Exception:
                sas_date = None
        elif not pending.empty:
            sas_status_type = "PENDING_IN_DELAY"
            sas_result      = "PENDING"
            sas_date        = None
        else:
            sas_status_type = "NOT_CALLED"
            sas_result      = None
            sas_date        = None

        lookup[doc_id] = {
            "sas_status_type": sas_status_type,
            "sas_result":      sas_result,
            "sas_date":        sas_date,
            "has_rappel":      bool(has_rappel),
        }
    return lookup


def _classify_missing_in_gf(
    family:     dict,
    sas_lookup: dict,
    data_date:  _dt.date,
) -> tuple:
    """
    Part 5: Classify MISSING_IN_GF using SAS timing.

    Returns (subtype_str, severity_str).

    Priority (highest to lowest):
      PENDING_SAS          — SAS not yet reached (NOT_CALLED or PENDING_IN_DELAY)
      RECENT_SAS_REMINDER  — doc has an active rappel reminder
      RECENT_REFUSAL       — REF within 30 days
      RECENT_ACCEPTED_SAS  — VSO-SAS / VSO within 14 days
      TRUE                 — everything else (genuinely missing)
      AMBIGUOUS            — no SAS data at all
    """
    all_doc_ids = family.get("all_doc_ids", set())
    if not all_doc_ids:
        return "MISSING_IN_GF_AMBIGUOUS", "INFO"

    # Aggregate across all doc_ids (multi-submission families can have multiple)
    best: dict = {}
    for did in sorted(str(did) for did in all_doc_ids if did is not None):
        info = sas_lookup.get(did)
        if info:
            best = info
            # Prefer answered over not-called; rappel takes priority
            if info.get("has_rappel"):
                break
            if info.get("sas_status_type") == "ANSWERED":
                break

    if not best:
        return "MISSING_IN_GF_AMBIGUOUS", "INFO"

    stype   = best["sas_status_type"]
    sresult = (best.get("sas_result") or "").upper()
    sdate   = best.get("sas_date")

    # Rule 1: SAS not yet reached
    if stype in ("NOT_CALLED", "PENDING_IN_DELAY"):
        return "MISSING_IN_GF_PENDING_SAS", "INFO"

    # Rule 2: Active rappel reminder
    # Note: pre-2026 rappels already removed by _apply_sas_filter → remaining are recent
    if best.get("has_rappel"):
        if sdate and data_date:
            try:
                diff = abs((data_date - sdate).days)
                if diff <= 30:
                    return "MISSING_IN_GF_RECENT_SAS_REMINDER", "INFO"
            except Exception:
                pass
        # Still treat as INFO even if recency unclear (post-2026 rappel)
        return "MISSING_IN_GF_RECENT_SAS_REMINDER", "INFO"

    # Rule 3: Refusal (REF)
    if "REF" in sresult:
        if sdate and data_date:
            try:
                diff = abs((data_date - sdate).days)
                if diff <= 30:
                    return "MISSING_IN_GF_RECENT_REFUSAL", "INFO"
            except Exception:
                pass
        # Old refusal: still not a true missing from GF perspective
        return "MISSING_IN_GF_RECENT_REFUSAL", "INFO"

    # Rule 4: Recently accepted
    if sresult in ("VSO-SAS", "VSO"):
        if sdate and data_date:
            try:
                diff = abs((data_date - sdate).days)
                if diff <= 14:
                    return "MISSING_IN_GF_RECENT_ACCEPTED_SAS", "INFO"
            except Exception:
                pass

    # Rule 5: True missing (accepted SAS but not in GF, not recent)
    return "MISSING_IN_GF_TRUE", "REVIEW_REQUIRED"


# ─────────────────────────────────────────────────────────────
# NEW SUBMITTAL CLASSIFICATION
# ─────────────────────────────────────────────────────────────

NEW_SUBMITTAL_WINDOW_DAYS = 30


def _classify_new_submittal_status(
    doc_id:      str,
    is_absent:   bool,   # True if NOT found in original GF
    sas_lookup:  dict,
    data_date:   _dt.date,
    window_days: int = NEW_SUBMITTAL_WINDOW_DAYS,
) -> tuple:
    """
    Classify a single GED current doc as a new submittal or not.

    Returns: (new_submittal_status, days_from_data_date, rationale)

    new_submittal_status values:
      ALREADY_IN_GF        — doc matches a row in the original GF
      NEW_PENDING_SAS      — absent + SAS not yet answered (Case A)
      NEW_RECENT_SAS_REF   — absent + SAS REF within window (Case B)
      NEW_RECENT_SAS_APPROVED — absent + SAS VAO/VSO within window (Case C)
      NOT_NEW_BACKLOG      — absent + SAS older / doesn't fit new-window (Case D)
      AMBIGUOUS            — absent + no SAS data
    """
    if not is_absent:
        return "ALREADY_IN_GF", None, "Present in original GF"

    sas_info = sas_lookup.get(doc_id, {})
    if not sas_info:
        return "AMBIGUOUS", None, "No SAS data found for this doc"

    sas_type   = sas_info.get("sas_status_type", "NOT_CALLED")
    sas_result = (sas_info.get("sas_result") or "").upper()
    sas_date   = sas_info.get("sas_date")

    # Days from data_date (positive = sas_date is in the past relative to data_date)
    days_diff = None
    if sas_date and data_date:
        try:
            days_diff = (data_date - sas_date).days
        except Exception:
            pass

    # Case A: SAS not yet answered (pending / en attente / rappel)
    if sas_type in ("NOT_CALLED", "PENDING_IN_DELAY", "RAPPEL"):
        return "NEW_PENDING_SAS", days_diff, f"SAS pending (type={sas_type})"

    # Case B: SAS REF within window
    if "REF" in sas_result:
        if days_diff is not None and 0 <= days_diff <= window_days:
            return "NEW_RECENT_SAS_REF", days_diff, \
                f"SAS REF within {window_days} days (diff={days_diff})"
        return "NOT_NEW_BACKLOG", days_diff, \
            f"SAS REF but outside window (diff={days_diff})"

    # Case C: SAS approved (any accepted SAS status) within window
    if sas_result in ("VSO-SAS", "VAO-SAS", "VSO", "VAO"):
        if days_diff is not None and 0 <= days_diff <= window_days:
            return "NEW_RECENT_SAS_APPROVED", days_diff, \
                f"SAS {sas_result} within {window_days} days (diff={days_diff})"
        return "NOT_NEW_BACKLOG", days_diff, \
            f"SAS {sas_result} outside window (diff={days_diff})"

    # Case D: anything else (HM, SUS, unknown) not within window → backlog
    return "NOT_NEW_BACKLOG", days_diff, \
        f"SAS result={sas_result}, type={sas_type}, diff={days_diff}"


def _build_new_submittal_analysis(
    dernier_df_for_gf:    pd.DataFrame,
    dernier_df_excluded:  pd.DataFrame,
    discrepancies:        list,
    sas_lookup:           dict,
    data_date:            _dt.date,
) -> list:
    """
    Build the new-submittal classification for every current GED family.

    Returns a list of analysis record dicts (one per doc).

    Algorithm:
      1. Build `absent_keys` = set of (sheet, numero, indice) for all
         MISSING_IN_GF_* discrepancies.  Every doc NOT in this set is
         ALREADY_IN_GF.
      2. For each valid current GED doc (dernier_df_for_gf):
         - check absence → classify via _classify_new_submittal_status
      3. For each config-excluded doc (dernier_df_excluded):
         - mark as EXCLUDED
    """
    # ── Step 1: build absent set from post-reconciliation discrepancies ───────
    # Only MISSING_IN_GF_* records signal genuine absence from original GF.
    # SHEET_MISMATCH, TITRE_MISMATCH etc. mean the doc IS in GF (just mismatched).
    absent_keys: set = set()
    for d in discrepancies:
        if str(d.get("flag_type", "")).startswith("MISSING_IN_GF"):
            absent_keys.add((
                str(d.get("sheet_name", "")),
                str(d.get("numero",     "")),
                str(d.get("indice",     "")),
            ))

    rows = []

    # ── Step 2: valid current GED docs ───────────────────────────────────────
    for _, doc in dernier_df_for_gf.iterrows():
        doc_id   = str(doc.get("doc_id", "") or "")
        sheet    = str(doc.get("gf_sheet_name", "") or "")
        num      = str(doc.get("numero_normalized", "") or "")
        ind      = str(doc.get("indice", "") or "")
        emetteur = str(doc.get("emetteur", "") or "")
        lot      = str(doc.get("lot_normalized", "") or "")
        titre    = str(doc.get("libelle_du_document", "") or "")[:80]
        type_doc = str(doc.get("type_de_doc", "") or "")

        is_absent = (sheet, num, ind) in absent_keys

        status, days_diff, rationale = _classify_new_submittal_status(
            doc_id, is_absent, sas_lookup, data_date
        )

        sas_info = sas_lookup.get(doc_id, {})
        rows.append({
            "Sheet target":          sheet,
            "Emetteur":              emetteur,
            "Lot":                   lot,
            "Numero":                num,
            "Indice":                ind,
            "Titre":                 titre,
            "SAS status type":       sas_info.get("sas_status_type", "") if sas_info else "",
            "SAS result":            sas_info.get("sas_result", "") if sas_info else "",
            "SAS date":              str(sas_info.get("sas_date", "") or "") if sas_info else "",
            "data_date":             str(data_date or ""),
            "days_from_data_date":   days_diff,
            "exists_in_original_gf": "no" if is_absent else "yes",
            "new_submittal_status":  status,
            "rationale":             rationale,
        })

    # ── Step 3: excluded docs → EXCLUDED ─────────────────────────────────────
    for _, doc in dernier_df_excluded.iterrows():
        doc_id   = str(doc.get("doc_id", "") or "")
        sheet    = str(doc.get("gf_sheet_name", "") or doc.get("exclusion_reason", ""))
        num      = str(doc.get("numero_normalized", "") or "")
        ind      = str(doc.get("indice", "") or "")
        emetteur = str(doc.get("emetteur", "") or "")
        lot      = str(doc.get("lot_normalized", "") or "")
        titre    = str(doc.get("libelle_du_document", "") or "")[:80]
        type_doc = str(doc.get("type_de_doc", "") or "")
        reason   = str(doc.get("exclusion_reason", "config exclusion") or "config exclusion")

        sas_info = sas_lookup.get(doc_id, {})
        sas_date  = sas_info.get("sas_date")  if sas_info else None
        days_diff = None
        if sas_date and data_date:
            try:
                days_diff = (data_date - sas_date).days
            except Exception:
                pass

        rows.append({
            "Sheet target":          sheet,
            "Emetteur":              emetteur,
            "Lot":                   lot,
            "Numero":                num,
            "Indice":                ind,
            "Titre":                 titre,
            "SAS status type":       sas_info.get("sas_status_type", "") if sas_info else "",
            "SAS result":            sas_info.get("sas_result", "") if sas_info else "",
            "SAS date":              str(sas_info.get("sas_date", "") or "") if sas_info else "",
            "data_date":             str(data_date or ""),
            "days_from_data_date":   days_diff,
            "exists_in_original_gf": "?",
            "new_submittal_status":  "EXCLUDED",
            "rationale":             f"Config-excluded: {reason}",
        })

    return rows


def _classify_missing_in_ged(
    gf_num:           str,
    gf_ind:           str,
    gf_rows:          list,
    is_historical:    bool,
    is_excluded:      bool,
    gf_dup_counts:    dict,
    sheet_name:       str,
) -> str:
    """
    Part 4: Classify MISSING_IN_GED into subtypes.

    MISSING_IN_GED_TRUE             — genuinely absent from GED, actionable
    MISSING_IN_GED_HISTORICAL       — GED has this numero at a newer indice
    MISSING_IN_GED_GF_SAS_REF       — GF row was refused at SAS; may have been retracted
    MISSING_IN_GED_GF_DUPLICATE_ROW — GF has >1 row for same (num, ind)
    MISSING_IN_GED_EXCLUDED         — filtered by year threshold
    """
    if is_excluded:
        return "MISSING_IN_GED_EXCLUDED"

    if is_historical:
        return "MISSING_IN_GED_HISTORICAL"

    # GF row has SAS REF — may have been retracted from GED after refusal
    if any((row or {}).get("gf_has_sas_ref") for row in gf_rows):
        return "MISSING_IN_GED_GF_SAS_REF"

    # GF key is duplicated (>1 rows for same num/ind on same sheet)
    if gf_dup_counts.get((sheet_name, gf_num, gf_ind), 1) > 1:
        return "MISSING_IN_GED_GF_DUPLICATE_ROW"

    return "MISSING_IN_GED_TRUE"


def _gf_row_stable_key(gf_row: dict) -> tuple:
    """Stable ordering for duplicate GF rows representing the same key."""
    return (
        normalize_title_for_compare(gf_row.get("titre", "")),
        normalize_date_for_compare(gf_row.get("date_diffusion")),
        normalize_status_for_compare(gf_row.get("gf_visa_global")),
        str(gf_row.get("document", "") or ""),
        str(gf_row.get("numero_normalized", "") or ""),
        str(gf_row.get("indice_normalized", "") or ""),
    )


def _sorted_family_doc_ids(family: dict) -> list[str]:
    """Stable ordering for family doc ids when the caller stops on the first eligible SAS hit."""
    return sorted(str(did) for did in family.get("all_doc_ids", set()) if did is not None)


# ─────────────────────────────────────────────────────────────
# PIPELINE
# ─────────────────────────────────────────────────────────────

def run_pipeline(verbose: bool = True):
    global _ACTIVE_RUN_NUMBER, _ACTIVE_RUN_FINALIZED, _RUN_CONTROL_CONTEXT

    def log(msg: str):
        if verbose:
            _safe_console_print(f"  {msg}")

    _safe_console_print("=" * 60)
    _safe_console_print("GF UPDATER V3 — RUN 0 (CLEAN REBUILD) — Patch A+B+C+D")
    _safe_console_print("=" * 60)

    # ── RUN HISTORY: initialise and determine run number ──────
    _run_number = None
    _run_dir    = None
    _run_input_entries = []
    _arts_registered = 0
    _ACTIVE_RUN_NUMBER = None
    _ACTIVE_RUN_FINALIZED = False
    try:
        init_run_memory_db_fn(RUN_MEMORY_DB)
        if not baseline_run_exists(RUN_MEMORY_DB):
            raise RuntimeError(
                "run_memory baseline missing: bootstrap Run 0 first with "
                "scripts/bootstrap_run_zero.py"
            )
        _run_number = get_next_run_number(RUN_MEMORY_DB)
        _run_type   = "INCREMENTAL"

        # Determine lineage
        _current_run_df   = get_current_run(RUN_MEMORY_DB)
        _parent_run_number = (
            int(_current_run_df.iloc[0]["run_number"])
            if not _current_run_df.empty else None
        )
        # root is always 0 once it exists; before Run 0 exists it self-references
        _root_run_number = 0

        _run_dir = get_run_dir(str(BASE_DIR), _run_number)
        import os as _os
        _os.makedirs(_run_dir, exist_ok=True)
        _os.makedirs(_run_dir + "/debug", exist_ok=True)

        _run_notes = "Auto-created by run_pipeline()"
        if _RUN_CONTROL_CONTEXT and _RUN_CONTROL_CONTEXT.get("run_mode"):
            _run_notes = f"{_run_notes} | run_mode={_RUN_CONTROL_CONTEXT['run_mode']}"

        create_run(
            db_path              = RUN_MEMORY_DB,
            run_number           = _run_number,
            run_type             = _run_type,
            parent_run_number    = _parent_run_number,
            root_run_number      = _root_run_number,
            based_on_run_number  = _parent_run_number,
            is_baseline          = False,
            run_label            = f"Run {_run_number}",
            notes                = _run_notes,
            core_version         = RUN_MEMORY_CORE_VERSION,
        )
        log(f"Run history: run {_run_number} ({_run_type}) — folder: runs/run_{_run_number:04d}/")
        _ACTIVE_RUN_NUMBER = _run_number
    except Exception as _rm_init_err:
        _safe_console_print(f"  [WARN] Run history init error (non-fatal): {_rm_init_err}")
        _run_number = None
        _run_dir    = None

    # ── STEP 1: Read GED ──────────────────────────────────────
    _safe_console_print("\n[1/7] Reading GED export...")
    docs_df, responses_df, ged_approver_names = read_ged(str(GED_FILE))
    log(f"Documents: {len(docs_df)} rows")
    log(f"Responses: {len(responses_df)} rows")
    log(f"Approvers discovered: {len(ged_approver_names)}")

    # ── STEP 2: Load Mapping (hardcoded) ─────────────────────
    _safe_console_print("\n[2/7] Loading Mapping...")
    mapping = load_mapping()
    log(f"Mapping entries: {len(mapping)} (hardcoded)")

    exception_count = sum(1 for v in mapping.values() if v == "Exception List")
    log(f"Exception List entries: {exception_count}")

    # ── RUN HISTORY: register inputs ─────────────────────────
    if _run_number is not None:
        try:
            def _register_input_entry(
                input_type: str,
                source_path: Path,
                source_filename: str,
                metadata: dict | None = None,
            ) -> None:
                _inp_hash = None
                if source_path.exists() and source_path.is_file():
                    try:
                        from run_memory import sha256_file as _sha256
                        _inp_hash = _sha256(str(source_path))
                    except Exception:
                        pass
                metadata_json = json.dumps(metadata, sort_keys=True) if metadata else None
                register_run_input(
                    db_path          = RUN_MEMORY_DB,
                    run_number       = _run_number,
                    input_type       = input_type,
                    source_filename  = source_filename,
                    source_file_hash = _inp_hash,
                    source_path      = str(source_path) if source_path.exists() else None,
                    metadata_json    = metadata_json,
                )
                _run_input_entries.append({
                    "input_type": input_type,
                    "source_filename": source_filename,
                    "source_file_hash": _inp_hash,
                    "source_path": str(source_path) if source_path.exists() else None,
                    "metadata": metadata or {},
                })

            for _inp_path, _inp_type, _inp_name in [
                (GED_FILE,     "GED",     "GED_export.xlsx"),
                (GF_FILE,      "GF",      GF_FILE.name),
            ]:
                _register_input_entry(_inp_type, Path(_inp_path), _inp_name)

            _register_input_entry(
                "REPORT_MEMORY",
                Path(REPORT_MEMORY_DB),
                Path(REPORT_MEMORY_DB).name,
            )

            if CONSULTANT_MATCH_REPORT.exists():
                _register_input_entry(
                    "CONSULTANT_MATCH_REPORT",
                    CONSULTANT_MATCH_REPORT,
                    CONSULTANT_MATCH_REPORT.name,
                )

            if CONSULTANT_REPORTS_ROOT.exists():
                _pdf_files = sorted(CONSULTANT_REPORTS_ROOT.rglob("*.pdf"))
                _register_input_entry(
                    "REPORT",
                    CONSULTANT_REPORTS_ROOT,
                    "consultant_reports/",
                    metadata={
                        "pdf_count": len(_pdf_files),
                        "sample_files": [p.name for p in _pdf_files[:10]],
                    },
                )
                for _pdf_path in _pdf_files:
                    _register_input_entry(
                        "REPORT_FILE",
                        _pdf_path,
                        _pdf_path.name,
                    )

            update_run_metadata(
                RUN_MEMORY_DB,
                _run_number,
                input_signature=_build_input_signature(_run_input_entries),
            )
        except Exception as _ri_err:
            log(f"[WARN] Run history input registration error (non-fatal): {_ri_err}")

    # ── STEP 3: Normalize ─────────────────────────────────────
    _safe_console_print("\n[3/7] Normalizing data...")
    docs_df = normalize_docs(docs_df, mapping)
    responses_df = normalize_responses(responses_df, mapping)

    # Filter out exception approvers
    non_exception_responses = responses_df[~responses_df["is_exception_approver"]]
    log(f"Non-exception responses: {len(non_exception_responses)}")

    # ── STEP 3b: SAS FILTER — RAPPEL_EN_ATTENTE + pre-2026 ───
    # Business rule: If 0-SAS approver has an unresolved reminder ("Rappel")
    # AND the document was created before 2026 → INVALID, remove entirely.
    # These are old unresolved SAS visa requests that are stale/superseded.
    _safe_console_print("\n  [3b] Applying SAS filter (RAPPEL_EN_ATTENTE + pre-2026)...")
    docs_df, responses_df, sas_filtered_df = _apply_sas_filter(docs_df, responses_df)
    sas_count = len(sas_filtered_df)
    if sas_count:
        log(f"SAS filter removed: {sas_count} docs (SAS_OLD_UNRESOLVED)")
        emetteur_counts = sas_filtered_df["emetteur"].value_counts()
        log(f"  By emetteur: {emetteur_counts.to_dict()}")
    else:
        log("SAS filter: no docs removed")

    # ── STEP 4: Version Engine ────────────────────────────────
    _safe_console_print("\n[4/7] Running Version Engine...")
    engine = VersionEngine(docs_df)
    versioned_df = engine.run()

    total = len(versioned_df)
    dernier_count = versioned_df["is_dernier_indice"].sum()
    anomaly_count = versioned_df["anomaly_flags"].apply(lambda x: len(x) > 0).sum()
    excluded_count = versioned_df["is_excluded_lifecycle"].sum()

    log(f"Total document versions: {total}")
    log(f"Dernier indices (latest versions): {dernier_count}")
    log(f"Documents with anomaly flags: {anomaly_count}")
    log(f"Excluded (old lifecycle): {excluded_count}")

    # ── STEP 5: Route to Contractor Sheets ───────────────────
    _safe_console_print("\n[5/7] Routing to contractor sheets...")
    routing_table = build_routing_table(str(GF_FILE))
    routing_entries = sum(1 for _ in routing_table.all_entries())
    log(f"Routing table entries: {routing_entries}")

    # Route all versions (not just dernier_indice) — uses emetteur now (Patch A)
    versioned_df = route_documents(versioned_df, routing_table)

    # Write routing summary debug artifact
    routing_summary_path = str(DEBUG_DIR / "routing_summary.xlsx")
    write_routing_summary(routing_summary_path, versioned_df, routing_table)
    log(f"  → debug/routing_summary.xlsx written")

    # Only keep dernier_indice docs for output
    dernier_df = versioned_df[versioned_df["is_dernier_indice"] == True].copy()

    # Count routing outcomes
    ok_count          = (dernier_df["routing_status"] == "OK").sum()
    ambiguous_count   = (dernier_df["routing_status"] == "ROUTING_AMBIGUOUS").sum()
    unmatched_count   = (dernier_df["routing_status"] == "ROUTING_UNMATCHED").sum()
    mismatch_count    = (dernier_df["routing_status"] == "ROUTING_EMETTEUR_MISMATCH").sum()
    routed = ok_count + ambiguous_count
    unrouted = unmatched_count

    log(f"Documents routed OK: {ok_count}")
    log(f"Documents routed (ambiguous): {ambiguous_count}")
    log(f"Documents unmatched (no sheet): {unmatched_count}")
    log(f"Documents with emetteur mismatch (wrong contractor): {mismatch_count}")

    if unrouted > 0:
        missing_lots = dernier_df[
            dernier_df["routing_status"] == "ROUTING_UNMATCHED"
        ]["lot_normalized"].value_counts()
        log(f"  Lots unmatched: {missing_lots.index.tolist()[:10]}")

    if mismatch_count > 0:
        mismatch_emetteurs = dernier_df[
            dernier_df["routing_status"] == "ROUTING_EMETTEUR_MISMATCH"
        ]["emetteur"].value_counts()
        log(f"  Emetteur mismatches: {mismatch_emetteurs.to_dict()}")

    # ── STEP 5b: Apply exclusion config (Patch C — BEFORE discrepancy) ──
    _safe_console_print("\n  [5b] Applying exclusion config (Patch C: before discrepancy generation)...")
    exclusion_config = load_exclusion_config()
    dernier_df = exclusion_config.apply(dernier_df)
    exclusion_summary = exclusion_config.summary(dernier_df)
    if exclusion_summary.get("total_excluded", 0) > 0:
        log(f"Excluded by config rules: {exclusion_summary['total_excluded']}")
        for reason, cnt in exclusion_summary["by_reason"].items():
            log(f"    {reason}: {cnt}")
        # Mark excluded docs in versioned_df for reporting
        excluded_doc_ids = set(dernier_df[dernier_df["is_excluded_config"]]["doc_id"].tolist())
        versioned_df.loc[
            versioned_df["doc_id"].isin(excluded_doc_ids), "resolution_status"
        ] = "IGNORED"

    # Write exclusion summary debug artifact (Patch C requirement)
    exclusion_config.write_exclusion_summary(
        dernier_df,
        str(DEBUG_DIR / "exclusion_summary.xlsx"),
    )

    # Only keep non-excluded dernier_indice docs for GF output AND discrepancy
    # Patch C: exclusions applied HERE before discrepancy in step 7
    dernier_df_for_gf = dernier_df[~dernier_df["is_excluded_config"]].copy()
    log(f"Dernier docs for GF (after exclusions): {len(dernier_df_for_gf)}")
    log(f"  Emetteur mismatch excluded: {mismatch_count}")

    # ── STEP 6: Read GF Sheet Structures ─────────────────────
    _safe_console_print("\n[6/7] Reading existing GF sheet structures...")
    sheet_names_to_read = list(dernier_df_for_gf["gf_sheet_name"].dropna().unique())
    sheet_structures = read_all_gf_sheet_structures(str(GF_FILE), sheet_names_to_read)
    for sheet_name, struct in sheet_structures.items():
        log(f"  {sheet_name}: {len(struct['approvers'])} approvers")

    # ── STEP 7: Write Outputs ─────────────────────────────────
    _safe_console_print("\n[7/7] Writing outputs...")

    # ── REPORT MEMORY INTEGRATION ─────────────────────────────────────────────
    # Merge persisted consultant report answers with GED-normalized responses so
    # that WorkflowEngine uses the EFFECTIVE state for every (doc_id, approver)
    # pair, not just what GED currently shows.
    #
    # Flow:
    #  1. Init DB (idempotent — safe to call every run).
    #  2. Load previously persisted responses from DB.
    #  3. If consultant_match_report.xlsx is present, check for new rapport_ids
    #     not yet in the DB → persist them + register their source files.
    #  4. Reload from DB after any new ingestion.
    #  5. Build effective_responses_df = merge(GED, report_memory).
    #  6. Feed effective_responses_df to WorkflowEngine.

    _safe_console_print("\n[7a/7] Report memory integration...")

    # ── Step 1: Init DB ───────────────────────────────────────────────────────
    init_report_memory_db(REPORT_MEMORY_DB)
    log(f"Report memory DB: {REPORT_MEMORY_DB}")

    # ── Step 2: Load previously persisted responses ───────────────────────────
    persisted_df = load_persisted_report_responses(REPORT_MEMORY_DB)
    log(f"Persisted report responses loaded: {len(persisted_df)}")

    # ── Step 3: Check for new reports in consultant_match_report.xlsx ─────────
    # This file is produced by consultant_integration.py.  If it exists and
    # contains rapport_ids not yet in the DB, we persist them now so the
    # current run already benefits from the latest consultant data.
    if CONSULTANT_MATCH_REPORT.exists():
        log(f"Checking {CONSULTANT_MATCH_REPORT.name} for new consultant responses...")
        try:
            match_df = pd.read_excel(str(CONSULTANT_MATCH_REPORT))

            # Keep only rows with a valid matched doc_id
            match_df = match_df[
                match_df["Matched GED doc_id"].notna() &
                (match_df["Matched GED doc_id"].astype(str).str.strip() != "")
            ].copy()

            # Canonical source → GED approver name mapping
            _src_canonical = {
                "LE_SOMMER":  "AMO HQE LE SOMMER",
                "LESOMMER":   "AMO HQE LE SOMMER",
                "AVLS":       "ACOUSTICIEN AVLS",
                "TERRELL":    "BET STR-TERRELL",
                "SOCOTEC":    "SOCOTEC",
                "BC_SOCOTEC": "SOCOTEC",
            }
            def _to_canonical(src: str) -> str:
                upper = src.upper().replace(" ", "_").replace("-", "_")
                for k, v in _src_canonical.items():
                    if k in upper:
                        return v
                return src

            # Group by rapport_id to process file-by-file
            new_rapport_ids_found = 0
            new_responses_persisted = 0

            for rapport_id, group in match_df.groupby("Rapport ID"):
                rapport_id_str = str(rapport_id).strip()

                # Compute hash: try real PDF first, fall back to sentinel
                file_hash = None
                if CONSULTANT_REPORTS_ROOT.exists():
                    for pdf_path in CONSULTANT_REPORTS_ROOT.rglob("*.pdf"):
                        if rapport_id_str in pdf_path.stem or rapport_id_str in pdf_path.name:
                            try:
                                file_hash = sha256_file(str(pdf_path))
                            except OSError:
                                pass
                            break
                if file_hash is None:
                    file_hash = f"BOOTSTRAP::{rapport_id_str}"

                # Skip if already ingested under this hash
                if is_report_already_ingested(REPORT_MEMORY_DB, file_hash):
                    continue

                new_rapport_ids_found += 1
                log(f"  New rapport: {rapport_id_str}")

                # Build persistence-ready DataFrame for this batch
                records = []
                for _, row in group.iterrows():
                    src = str(row.get("Consultant Source", "")).strip()
                    records.append({
                        "consultant":           _to_canonical(src),
                        "doc_id":               str(row.get("Matched GED doc_id", "")).strip(),
                        "report_status":        str(row.get("STATUT_NORM",  "") or "").strip() or None,
                        "report_response_date": str(row.get("DATE_FICHE",   "") or "").strip() or None,
                        "report_comment":       str(row.get("COMMENTAIRE",  "") or "").strip() or None,
                        "source_filename":      rapport_id_str,
                        "source_file_hash":     file_hash,
                        "match_confidence":     str(row.get("Confidence",   "") or "").strip() or None,
                        "match_method":         str(row.get("Match Method", "") or "").strip() or None,
                    })

                batch_df = pd.DataFrame(records)
                written  = upsert_report_responses(REPORT_MEMORY_DB, batch_df)
                new_responses_persisted += written

                # Register the source file
                source_type = str(group.iloc[0].get("Consultant Source", "UNKNOWN")).strip()
                is_sentinel = file_hash.startswith("BOOTSTRAP::")
                register_ingested_report(
                    db_path          = REPORT_MEMORY_DB,
                    report_type      = source_type,
                    source_filename  = rapport_id_str,
                    source_file_hash = file_hash,
                    row_count        = len(records),
                    status           = "INGESTED_BOOTSTRAP" if is_sentinel else "INGESTED",
                )

            log(
                f"  New rapport_ids ingested: {new_rapport_ids_found} "
                f"({new_responses_persisted} rows persisted)"
            )

            # ── Step 4: Reload from DB after any new ingestion ─────────────────
            if new_rapport_ids_found > 0:
                persisted_df = load_persisted_report_responses(REPORT_MEMORY_DB)
                log(f"  Persisted responses after new ingestion: {len(persisted_df)}")

        except Exception as _rm_err:
            # Report memory is an enrichment layer — never block the main pipeline
            _safe_console_print(
                f"  [WARN] Report memory ingestion error (non-fatal): {_rm_err}"
            )
            import traceback as _tb
            _tb.print_exc()
    else:
        log(f"  {CONSULTANT_MATCH_REPORT.name} not found — skipping new report ingestion")

    # ── Step 5: Build effective responses ─────────────────────────────────────
    # Merge GED responses (source of truth for structure) with report memory
    # (fills in answers that GED still shows as pending).
    if _run_number is not None:
        try:
            _report_memory_hash = None
            _report_memory_path = Path(REPORT_MEMORY_DB)
            if _report_memory_path.exists():
                _report_memory_hash = sha256_file(str(_report_memory_path))
            update_run_metadata(
                RUN_MEMORY_DB,
                _run_number,
                report_memory_snapshot_hash=_report_memory_hash,
            )
        except Exception as _rm_meta_err:
            log(f"[WARN] Run history report-memory snapshot error (non-fatal): {_rm_meta_err}")

    effective_responses_df = build_effective_responses(responses_df, persisted_df)

    upgraded_count = int(effective_responses_df.get("report_memory_applied", pd.Series(dtype=bool)).sum()) \
        if "report_memory_applied" in effective_responses_df.columns else 0
    log(
        f"Effective responses built: {len(effective_responses_df)} rows "
        f"({upgraded_count} upgraded from PENDING→ANSWERED via report memory)"
    )

    # ── Step 6: Build workflow engine using EFFECTIVE responses ───────────────
    wf_engine = WorkflowEngine(effective_responses_df)

    # Build GF→GED approver name map (use effective_responses_df for full canonical list)
    ged_canonical_approvers = effective_responses_df["approver_canonical"].unique().tolist()
    gf_to_ged_map = build_gf_to_ged_map(ged_canonical_approvers)
    log(f"GF→GED approver mappings: {len(gf_to_ged_map)} entries")

    # ── Build ancien_df: VALID_HISTORICAL rows only ──────────────────────────
    # Distinguish two conceptually different row types:
    #
    #   VALID_HISTORICAL — older indices of an active document family.
    #     • is_dernier_indice == False
    #     • is_excluded_lifecycle == False  (no superseded/reused-numero lifecycle)
    #     • lifecycle_id belongs to a DI row that made it into dernier_df_for_gf
    #       (i.e. the family is not config-excluded / emetteur-mismatch excluded)
    #     • has a valid routing destination (gf_sheet_name not null)
    #     → Written to GF_V0_CLEAN with ANCIEN="1" and FULL approver_statuses.
    #
    #   EXCEPTION — must NOT appear in GF_V0_CLEAN:
    #     • SAS_OLD_UNRESOLVED docs → already removed before VersionEngine runs
    #     • is_excluded_lifecycle == True → superseded lifecycles (reused numero)
    #     • Config-excluded families → filtered by lifecycle_id not in active set
    #     → Written only to IGNORED / DEBUG outputs.
    #
    active_lifecycle_ids = set(dernier_df_for_gf["lifecycle_id"].dropna().unique())
    ancien_df = versioned_df[
        (versioned_df["is_dernier_indice"] == False) &
        (versioned_df["is_excluded_lifecycle"] == False) &   # VALID_HISTORICAL only
        (versioned_df["lifecycle_id"].isin(active_lifecycle_ids)) &  # active families
        (versioned_df["gf_sheet_name"].notna())              # routed to a GF sheet
    ].copy()
    log(f"Ancien VALID_HISTORICAL rows for GF: {len(ancien_df)}")

    # ── Build SAS lookup for DATE CONTRACTUELLE Case B (Round 2 Patch 2) ──
    # sas_lookup[doc_id] = {sas_result, sas_date, ...}
    # Case B: SAS passed (VSO-SAS / VAO-SAS) → DATE CONTRACT = SAS date + 15
    # Case A: SAS not yet passed → DATE CONTRACT = date_diffusion + 15
    log("Building SAS lookup for DATE CONTRACTUELLE...")
    gf_sas_lookup = _build_sas_lookup(responses_df)
    sas_case_b_count = sum(
        1 for v in gf_sas_lookup.values()
        if v.get("sas_result", "") in ("VSO-SAS", "VAO-SAS", "VSO", "VAO")
        and v.get("sas_date") is not None
    )
    log(f"  SAS lookup: {len(gf_sas_lookup)} docs with SAS data, {sas_case_b_count} Case B (SAS passed + date)")

    # Write GF
    log("Writing GF_V0_CLEAN.xlsx...")
    gf_writer = GFWriter(str(OUTPUT_GF), str(GF_FILE))
    gf_writer.write_all(
        docs_df=dernier_df_for_gf,
        responses_df=responses_df,
        workflow_engine=wf_engine,
        sheet_structures=sheet_structures,
        gf_to_ged_map=gf_to_ged_map,
        ancien_df=ancien_df,
        sas_lookup=gf_sas_lookup,
    )
    gf_writer.save()
    log(f"  → {OUTPUT_GF}")

    # Split anomalies by resolution_status
    # Include excluded docs from config in IGNORED bucket
    excluded_flags_df = dernier_df[dernier_df["is_excluded_config"] == True].copy()
    excluded_flags_df["anomaly_flags"] = excluded_flags_df.apply(
        lambda r: (r.get("anomaly_flags") or []) + [f"EXCLUDED:{r.get('exclusion_reason','')}"],
        axis=1
    )

    all_anomaly_df = pd.concat([
        versioned_df[versioned_df["anomaly_flags"].apply(lambda x: len(x) > 0)],
        excluded_flags_df[~excluded_flags_df["doc_id"].isin(
            versioned_df[versioned_df["anomaly_flags"].apply(lambda x: len(x) > 0)]["doc_id"]
        )]
    ], ignore_index=True).copy()

    review_df   = all_anomaly_df[all_anomaly_df["resolution_status"] == "REVIEW_REQUIRED"]
    auto_df     = all_anomaly_df[all_anomaly_df["resolution_status"] == "AUTO_RESOLVED"]
    ignored_df  = all_anomaly_df[all_anomaly_df["resolution_status"] == "IGNORED"]

    log(f"Anomalies — REVIEW_REQUIRED: {len(review_df)}, AUTO_RESOLVED: {len(auto_df)}, IGNORED: {len(ignored_df)}")

    log("Writing ANOMALY_REPORT.xlsx (REVIEW_REQUIRED only)...")
    write_anomaly_report(str(OUTPUT_ANOMALY), review_df)
    log(f"  → {OUTPUT_ANOMALY}")

    log("Writing AUTO_RESOLUTION_LOG.xlsx...")
    write_auto_resolution_log(str(OUTPUT_AUTO_RESOLUTION), auto_df)
    log(f"  → {OUTPUT_AUTO_RESOLUTION}")

    log("Writing IGNORED_ITEMS_LOG.xlsx...")
    # Merge workflow-level ignored_df with SAS-filtered docs
    if len(sas_filtered_df) > 0:
        sas_log_df = sas_filtered_df.copy()
        sas_log_df["resolution_status"] = "IGNORED"
        sas_log_df["anomaly_flags"] = sas_log_df.apply(lambda _: ["SAS_OLD_UNRESOLVED"], axis=1)
        sas_log_df["exclusion_reason"] = "SAS_OLD_UNRESOLVED"
        combined_ignored = pd.concat([ignored_df, sas_log_df], ignore_index=True)
    else:
        combined_ignored = ignored_df
    write_ignored_items_log(str(OUTPUT_IGNORED), combined_ignored)
    log(f"  → {OUTPUT_IGNORED} ({len(combined_ignored)} rows total, {len(sas_filtered_df)} SAS_OLD_UNRESOLVED)")

    # ── Discrepancy computation (Patch B+C+D+E) ───────────────
    # Patch C: only dernier_df_for_gf is passed — exclusions already applied
    # Patch B: normalization happens inside _compute_discrepancies
    # Patch E: responses_df and data_date for SAS-based MISSING subtypes
    data_date = _determine_data_date(str(GED_FILE))
    log(f"Data reference date (from GED file mtime): {data_date}")
    log("Computing discrepancies (Patch B+C+E: normalized + SAS subtypes)...")
    discrepancies, gf_by_sheet = _compute_discrepancies(
        dernier_df_for_gf,
        str(GF_FILE),
        debug_dir=str(DEBUG_DIR),
        excluded_sheets=exclusion_config.excluded_sheets,
        sheet_year_filters=exclusion_config.sheet_year_filters,
        responses_df=responses_df,
        data_date=data_date,
    )

    # ── Patch F: Reconciliation engine — runs before severity classification ──
    # Build dernier_df_all: ALL GED dernier docs (before exclusions/routing filter)
    # This lets the reconciliation engine find docs that were unrouted or excluded
    dernier_df_all = versioned_df[versioned_df["is_dernier_indice"] == True].copy()

    log("Running reconciliation engine (Patch F)...")
    mid_before = sum(1 for d in discrepancies if d.get("flag_type") in (
        "MISSING_IN_GED_TRUE", "MISSING_IN_GED"))
    mif_before = sum(1 for d in discrepancies if d.get("flag_type") in (
        "MISSING_IN_GF_TRUE", "MISSING_IN_GF"))
    log(f"  Before reconciliation: MIG_TRUE={mid_before}, MIF_TRUE={mif_before}")

    discrepancies, recon_log = run_reconciliation(
        discrepancies=discrepancies,
        dernier_df_all=dernier_df_all,
        gf_by_sheet=gf_by_sheet,
        responses_df=responses_df,
    )

    mid_after = sum(1 for d in discrepancies if d.get("flag_type") in (
        "MISSING_IN_GED_TRUE", "MISSING_IN_GED"))
    mif_after = sum(1 for d in discrepancies if d.get("flag_type") in (
        "MISSING_IN_GF_TRUE", "MISSING_IN_GF"))
    log(f"  After  reconciliation: MIG_TRUE={mid_after} (−{mid_before - mid_after}), "
        f"MIF_TRUE={mif_after} (−{mif_before - mif_after})")
    log(f"  Reconciliation events logged: {len(recon_log)}")

    log("Writing RECONCILIATION_LOG.xlsx...")
    write_reconciliation_outputs(
        recon_log,
        str(OUTPUT_RECONCILIATION_LOG),
        str(OUTPUT_RECONCILIATION_SUMMARY),
    )
    log(f"  → {OUTPUT_RECONCILIATION_LOG}")
    log(f"  → debug/reconciliation_summary.xlsx")

    # ── Patch G-C: TITRE_MISMATCH relaxation ─────────────────────────────────
    # Identity is already confirmed (same sheet/numero/indice) — title alone
    # is not a blocking discrepancy.  Keep only truly-foreign-document cases.
    titre_relaxed = 0
    for rec in discrepancies:
        if rec.get("flag_type") != "TITRE_MISMATCH":
            continue
        sim = float(rec.get("title_similarity") or 0)
        # Keep REVIEW_REQUIRED only when titles suggest completely different docs
        # (sim < 0.30 — very low overlap, possibly wrong document filed under same numero)
        if sim >= 0.30:
            rec["flag_type"] = "TITRE_VARIANT_ACCEPTED"
            rec["reconciliation_note"] = (
                f"Title variant accepted (sim={sim:.2f}): "
                "numero+indice+emetteur all aligned; title difference is cosmetic"
            )
            titre_relaxed += 1
    log(f"  Patch G-C: TITRE_MISMATCH relaxed → TITRE_VARIANT_ACCEPTED: {titre_relaxed}")

    # ── Patch G-D: INDICE_MISMATCH relaxation ────────────────────────────────
    # If the GED and GF dates are within 7 days, the document identity is clear.
    # Use GED indice for reconstruction; downgrade to COSMETIC.
    #
    # NOTE: title_sim is NOT required here. INDICE_MISMATCH already guarantees
    # an exact numero match; GF rows for different indices often have empty titles
    # (titre="" → title_sim=0.0). Since numero is exact, date proximity alone is
    # sufficient to confirm identity and accept the indice variant.
    indice_accepted = 0
    for rec in discrepancies:
        if rec.get("flag_type") != "INDICE_MISMATCH":
            continue
        diff = rec.get("date_diff_days")
        title_sim_ind = float(rec.get("title_similarity") or 0)
        # Accept if: date within 7 days (numero already matched exactly)
        if diff is not None and diff <= 7:
            rec["flag_type"] = "INDICE_VARIANT_ACCEPTED_BY_GED"
            rec["reconciliation_note"] = (
                f"Indice variant accepted: date_diff={diff}d"
                + (f", title_sim={title_sim_ind:.2f}" if title_sim_ind > 0 else ", title_sim=n/a (GF row has no titre)")
                + f". Exact numero match; GED indice '{rec.get('ged_value')}' wins over "
                f"GF indice '{rec.get('gf_value')}'."
            )
            indice_accepted += 1
    log(f"  Patch G-D: INDICE_MISMATCH accepted → INDICE_VARIANT_ACCEPTED_BY_GED: {indice_accepted}")

    # ── Part H-1: BENTIN legacy exception pass ───────────────────────────────
    # The BENTIN sheet (LOT 31 à 34-IN-BX-CFO-BENTIN) contains legacy
    # inconsistencies from an old contractor batch that predates the current
    # GF structure.  All MISSING_IN_GF discrepancies on this sheet are excluded.
    BENTIN_SHEET = "LOT 31 à 34-IN-BX-CFO-BENTIN"
    BENTIN_TARGET_TYPES = {
        "MISSING_IN_GF_TRUE",
        "MISSING_IN_GF_AMBIGUOUS_TITLE_MATCH",
        "INDICE_MISMATCH",          # now INFO globally, but explicit BENTIN trace
        "DATE_MISMATCH",            # now INFO globally, but explicit BENTIN trace
    }
    bentin_count = 0
    for rec in discrepancies:
        if (rec.get("sheet_name") == BENTIN_SHEET
                and rec.get("flag_type") in BENTIN_TARGET_TYPES):
            rec["flag_type"] = "BENTIN_LEGACY_EXCEPTION"
            rec["reconciliation_note"] = (
                "BENTIN legacy: excluded from review queue. "
                "This sheet contains pre-2026 contractor inconsistencies."
            )
            bentin_count += 1
    log(f"  Part H-1: BENTIN legacy exceptions flagged: {bentin_count}")

    # ── Part H-2: Global title reconciliation — promote AMBIGUOUS to INFO ────
    # MISSING_IN_GED_AMBIGUOUS_TITLE_MATCH and MISSING_IN_GF_AMBIGUOUS_TITLE_MATCH
    # were already matched by the reconciliation engine at title_sim >= TITLE_SIM_PROBABLE
    # (0.65).  The "ambiguous" label means the score was [0.65, 0.80).
    # These are NOT unresolved — they have a probable match.
    # Per the final pass rules, they become REVIEW_REQUIRED but with low urgency.
    # No reclassification here; they remain REVIEW_REQUIRED as the spec requires.
    # (Kept as a documented no-op for clarity.)

    # ── Severity classification (must come after all flag_type mutations) ──
    for rec in discrepancies:
        rec["severity"] = classify_discrepancy(rec)

    disc_review = [d for d in discrepancies if d["severity"] == "REVIEW_REQUIRED"]
    disc_cosmetic = [d for d in discrepancies if d["severity"] == "COSMETIC"]
    disc_excluded = [d for d in discrepancies if d["severity"] == "EXCLUDED"]
    disc_info = [d for d in discrepancies if d["severity"] == "INFO"]

    log(f"Discrepancies total: {len(discrepancies)}")
    log(f"  REVIEW_REQUIRED:  {len(disc_review)}")
    log(f"  COSMETIC:         {len(disc_cosmetic)}")
    log(f"  EXCLUDED:         {len(disc_excluded)}")
    log(f"  INFO:             {len(disc_info)}")
    # Final counts for Patch G validation
    mid_final = sum(1 for d in discrepancies if d.get("flag_type") in ("MISSING_IN_GED_TRUE","MISSING_IN_GED"))
    mif_final = sum(1 for d in discrepancies if d.get("flag_type") in ("MISSING_IN_GF_TRUE","MISSING_IN_GF"))
    log(f"  MISSING_IN_GED_TRUE remaining: {mid_final}")
    log(f"  MISSING_IN_GF_TRUE remaining:  {mif_final}")
    titre_remaining = sum(1 for d in discrepancies if d.get("flag_type") == "TITRE_MISMATCH")
    indice_remaining = sum(1 for d in discrepancies if d.get("flag_type") == "INDICE_MISMATCH")
    log(f"  TITRE_MISMATCH remaining (blocked):  {titre_remaining}")
    log(f"  INDICE_MISMATCH remaining (blocked): {indice_remaining}")

    if discrepancies:
        disc_df = pd.DataFrame(discrepancies)
        counts = disc_df["flag_type"].value_counts()
        log("Breakdown by flag_type:")
        for ftype, cnt in counts.items():
            log(f"    {ftype}: {cnt}")

    log("Writing DISCREPANCY_REPORT.xlsx (all, with severity)...")
    write_discrepancy_report(str(OUTPUT_DISCREPANCY), discrepancies)
    log(f"  → {OUTPUT_DISCREPANCY}")

    log("Writing DISCREPANCY_REVIEW_REQUIRED.xlsx...")
    write_discrepancy_report(str(OUTPUT_DISCREPANCY_REVIEW), disc_review,
                             title_suffix=" — REVIEW REQUIRED ONLY")
    log(f"  → {OUTPUT_DISCREPANCY_REVIEW}")

    # ── Part H-1: Append BENTIN_LEGACY_EXCEPTION to IGNORED_ITEMS_LOG ────────
    bentin_excluded = [d for d in discrepancies if d.get("flag_type") == "BENTIN_LEGACY_EXCEPTION"]
    if bentin_excluded:
        bentin_df = pd.DataFrame([{
            "numero":           d.get("numero"),
            "indice":           d.get("indice"),
            "sheet":            d.get("sheet_name"),
            "document_code":    d.get("document_code"),
            "exclusion_reason": "BENTIN_LEGACY_EXCEPTION",
            "reconciliation_note": d.get("reconciliation_note", ""),
            "anomaly_flags":    ["BENTIN_LEGACY"],
        } for d in bentin_excluded])
        # Reload existing ignored log and append
        try:
            existing_ignored = pd.read_excel(str(OUTPUT_IGNORED))
            combined_with_bentin = pd.concat([existing_ignored, bentin_df], ignore_index=True)
        except Exception:
            combined_with_bentin = bentin_df
        combined_with_bentin.to_excel(str(OUTPUT_IGNORED), index=False)
        log(f"  Part H-1: {len(bentin_excluded)} BENTIN exceptions appended to IGNORED_ITEMS_LOG")

    # ── Patch E: Diagnosis outputs ────────────────────────────
    log("Writing MISSING_IN_GED_DIAGNOSIS.xlsx and MISSING_IN_GED_TRUE_ONLY.xlsx...")
    _write_missing_in_ged_diagnosis(
        str(OUTPUT_MISSING_GED_DIAGNOSIS),
        str(OUTPUT_MISSING_GED_TRUE),
        str(DEBUG_DIR / "missing_in_ged_summary.xlsx"),
        discrepancies,
    )
    log(f"  → {OUTPUT_MISSING_GED_DIAGNOSIS}")
    log(f"  → {OUTPUT_MISSING_GED_TRUE}")

    log("Writing MISSING_IN_GF_DIAGNOSIS.xlsx and MISSING_IN_GF_TRUE_ONLY.xlsx...")
    _write_missing_in_gf_diagnosis(
        str(OUTPUT_MISSING_GF_DIAGNOSIS),
        str(OUTPUT_MISSING_GF_TRUE),
        str(DEBUG_DIR / "missing_in_gf_summary.xlsx"),
        discrepancies,
    )
    log(f"  → {OUTPUT_MISSING_GF_DIAGNOSIS}")
    log(f"  → {OUTPUT_MISSING_GF_TRUE}")

    # ── INSERT LOG: newly inserted rows in CLEAN GF ───────────────────────────
    # "Inserted" = MISSING_IN_GF_TRUE after reconciliation.
    # These are GED docs confirmed absent from the original GF and now written
    # into CLEAN GF as current DI rows.  All other MISSING subtypes
    # (PENDING_SAS, RECENT_REFUSAL, etc.) are explicitly excluded.
    #
    # Variables in scope: discrepancies, dernier_df_for_gf, wf_engine, gf_sas_lookup
    log("Building INSERT LOG...")

    # Index dernier_df_for_gf by (sheet, numero_normalized, indice) for O(1) join
    _gf_key_index: dict = {}
    for _, _dr in dernier_df_for_gf.iterrows():
        _k = (
            str(_dr.get("gf_sheet_name", "") or ""),
            str(_dr.get("numero_normalized", "") or ""),
            str(_dr.get("indice", "") or ""),
        )
        if _k not in _gf_key_index:
            _gf_key_index[_k] = _dr

    insert_log: list = []
    for _d in discrepancies:
        if _d.get("flag_type") != "MISSING_IN_GF_TRUE":
            continue

        _sheet = str(_d.get("sheet_name", "") or "")
        _num   = str(_d.get("numero", "") or "")
        _ind   = str(_d.get("indice", "") or "")

        # Join with GED row (doc metadata)
        _doc_row = _gf_key_index.get((_sheet, _num, _ind))
        _doc_id  = str(_doc_row.get("doc_id", "") or "") if _doc_row is not None else ""
        _emetteur = str(_doc_row.get("emetteur", "") or "") if _doc_row is not None else ""
        _date_reception = _doc_row.get("created_at") if _doc_row is not None else None

        # DATE CONTRACTUELLE: Case B if SAS passed, else Case A
        _sas_entry  = gf_sas_lookup.get(_doc_id, {})
        _sas_result = str(_sas_entry.get("sas_result") or "")
        _sas_passed = _sas_result in ("VSO-SAS", "VAO-SAS", "VSO", "VAO")
        _sas_date   = _sas_entry.get("sas_date") if _sas_passed else None
        _base_date  = _sas_date if _sas_date is not None else _date_reception
        _date_contract = None
        if _base_date is not None:
            try:
                _date_contract = pd.to_datetime(_base_date) + _dt.timedelta(days=15)
            except Exception:
                pass

        # VISA GLOBAL + Date réel de visa — derived from same MOEX entry
        _visa_global    = None
        _date_reel_visa = None
        if _doc_id:
            _visa_global, _date_reel_visa = wf_engine.compute_visa_global_with_date(_doc_id)

        insert_log.append({
            "Sheet":             _sheet,
            "Emetteur":          _emetteur,
            "Lot":               str(_d.get("gfi_lot", "") or ""),
            "Numero":            _num,
            "Indice":            _ind,
            "Titre":             str(_d.get("gfi_titre", "") or ""),
            "Type Doc":          str(_d.get("gfi_type_doc", "") or ""),
            "Date réception":    _date_reception,
            "Date contract":     _date_contract,
            "Date réel de visa": _date_reel_visa,
            "VISA Global":       _visa_global or "",
            "Reason":            "MISSING_IN_GF_TRUE",
            "Confidence":        "HIGH",
        })

    log(f"  INSERT LOG: {len(insert_log)} newly inserted rows")
    write_insert_log(str(OUTPUT_INSERT_LOG), insert_log)
    log(f"  → {OUTPUT_INSERT_LOG}")

    # ── NEW SUBMITTAL ANALYSIS ───────────────────────────────────────────────
    log("Building new submittal analysis...")
    dernier_df_excluded_for_ns = dernier_df[dernier_df["is_excluded_config"]].copy()
    ns_rows = _build_new_submittal_analysis(
        dernier_df_for_gf=dernier_df_for_gf,
        dernier_df_excluded=dernier_df_excluded_for_ns,
        discrepancies=discrepancies,
        sas_lookup=gf_sas_lookup,
        data_date=data_date,
    )
    log(f"  New submittal analysis: {len(ns_rows)} docs analysed")
    _ns_by_status = {}
    for _r in ns_rows:
        _s = _r.get("new_submittal_status", "?")
        _ns_by_status[_s] = _ns_by_status.get(_s, 0) + 1
    for _s, _c in sorted(_ns_by_status.items(), key=lambda x: -x[1]):
        log(f"    {_s}: {_c}")
    write_new_submittal_analysis(
        analysis_path=str(OUTPUT_NEW_SUBMITTAL_ANALYSIS),
        summary_path=str(OUTPUT_NEW_SUBMITTAL_SUMMARY),
        rows=ns_rows,
    )
    log(f"  → {OUTPUT_NEW_SUBMITTAL_ANALYSIS}")
    log(f"  → {OUTPUT_NEW_SUBMITTAL_SUMMARY}")

    # Write additional debug artifacts
    log("Writing debug artifacts...")
    write_all_debug(
        debug_dir=str(DEBUG_DIR),
        versioned_df=versioned_df,
        discrepancies=discrepancies,
    )
    log(f"  → debug/coarse_groups.xlsx")
    log(f"  → debug/family_clusters.xlsx")
    log(f"  → debug/lifecycle_resolution.xlsx")
    log(f"  → debug/discrepancy_sample.xlsx")

    # ── RUN HISTORY: copy artifacts + register ───────────────
    if _run_number is not None and _run_dir is not None:
        try:
            # Map: (source_path, artifact_type, is_debug)
            _artifacts_to_register = [
                # Primary outputs
                (OUTPUT_GF,                    "FINAL_GF",                    False),
                (OUTPUT_DISCREPANCY,           "DISCREPANCY_REPORT",          False),
                (OUTPUT_DISCREPANCY_REVIEW,    "DISCREPANCY_REVIEW_REQUIRED", False),
                (OUTPUT_ANOMALY,               "ANOMALY_REPORT",              False),
                (OUTPUT_AUTO_RESOLUTION,       "AUTO_RESOLUTION_LOG",         False),
                (OUTPUT_IGNORED,               "IGNORED_ITEMS_LOG",           False),
                (OUTPUT_INSERT_LOG,            "INSERT_LOG",                  False),
                (OUTPUT_RECONCILIATION_LOG,    "RECONCILIATION_LOG",          False),
                (OUTPUT_MISSING_GED_DIAGNOSIS, "MISSING_IN_GED_DIAGNOSIS",    False),
                (OUTPUT_MISSING_GED_TRUE,      "MISSING_IN_GED_TRUE",         False),
                (OUTPUT_MISSING_GF_DIAGNOSIS,  "MISSING_IN_GF_DIAGNOSIS",     False),
                (OUTPUT_MISSING_GF_TRUE,       "MISSING_IN_GF_TRUE",          False),
                (OUTPUT_NEW_SUBMITTAL_ANALYSIS,"NEW_SUBMITTAL_ANALYSIS",      False),
                (OUTPUT_CONSULTANT_REPORTS_WB, "CONSULTANT_REPORTS",          False),
                (CONSULTANT_MATCH_REPORT,      "CONSULTANT_MATCH_REPORT",     False),
                (OUTPUT_GF_TEAM_VERSION,       "GF_TEAM_VERSION",             False),
                (OUTPUT_GF_STAGE1,             "CONSULTANT_ENRICHED_STAGE1",  False),
                (OUTPUT_GF_STAGE2,             "CONSULTANT_ENRICHED_STAGE2",  False),
                (OUTPUT_SUSPICIOUS_ROWS,       "SUSPICIOUS_ROWS_REPORT",      False),
                # Debug outputs
                (OUTPUT_RECONCILIATION_SUMMARY,       "DEBUG_RECONCILIATION_SUMMARY", True),
                (OUTPUT_NEW_SUBMITTAL_SUMMARY,         "DEBUG_NEW_SUBMITTAL_SUMMARY",  True),
                (DEBUG_DIR / "routing_summary.xlsx",   "DEBUG_ROUTING_SUMMARY",        True),
                (DEBUG_DIR / "exclusion_summary.xlsx", "DEBUG_EXCLUSION_SUMMARY",      True),
                (DEBUG_DIR / "missing_in_ged_summary.xlsx", "DEBUG_MISSING_IN_GED_SUMMARY", True),
                (DEBUG_DIR / "missing_in_gf_summary.xlsx",  "DEBUG_MISSING_IN_GF_SUMMARY",  True),
                (DEBUG_DIR / "coarse_groups.xlsx",     "DEBUG_COARSE_GROUPS",          True),
                (DEBUG_DIR / "family_clusters.xlsx",   "DEBUG_FAMILY_CLUSTERS",        True),
                (DEBUG_DIR / "lifecycle_resolution.xlsx","DEBUG_LIFECYCLE_RESOLUTION",  True),
                (DEBUG_DIR / "discrepancy_sample.xlsx","DEBUG_DISCREPANCY_SAMPLE",      True),
                (DEBUG_DIR / "gf_duplicates.xlsx",     "DEBUG_GF_DUPLICATES",           True),
                (DEBUG_DIR / "gf_sheet_schema.xlsx",   "DEBUG_GF_SHEET_SCHEMA",         True),
            ]
            _arts_registered = 0
            from run_memory import sha256_file as _sha256_art
            for _ap, _atype, _is_debug in _artifacts_to_register:
                _ap = Path(_ap)
                if not _ap.exists():
                    continue
                _subfolder = "debug" if _is_debug else None
                _dest = copy_artifact_to_run_dir(str(_ap), _run_dir, subfolder=_subfolder)
                if _dest is None:
                    continue
                try:
                    _ahash = _sha256_art(str(_ap))
                except Exception:
                    _ahash = None
                register_run_artifact(
                    db_path       = RUN_MEMORY_DB,
                    run_number    = _run_number,
                    artifact_type = _atype,
                    artifact_name = _ap.name,
                    file_path     = _dest,
                    file_hash     = _ahash,
                    format        = "xlsx",
                )
                _arts_registered += 1

            # Register report_memory.db as artifact
            _rdb_path = Path(REPORT_MEMORY_DB)
            if _rdb_path.exists():
                _rdb_dest = copy_artifact_to_run_dir(str(_rdb_path), _run_dir)
                try:
                    _rdb_hash = _sha256_art(str(_rdb_path))
                except Exception:
                    _rdb_hash = None
                register_run_artifact(
                    db_path       = RUN_MEMORY_DB,
                    run_number    = _run_number,
                    artifact_type = "REPORT_MEMORY_DB",
                    artifact_name = "report_memory.db",
                    file_path     = _rdb_dest or str(_rdb_path),
                    file_hash     = _rdb_hash,
                    format        = "sqlite",
                )
                _arts_registered += 1

            mark_run_current(RUN_MEMORY_DB, _run_number)
            log(f"Run history: run {_run_number} registered ({_arts_registered} artifacts)")

        except Exception as _ra_err:
            _safe_console_print(f"  [WARN] Run history artifact registration error (non-fatal): {_ra_err}")

    _safe_console_print("\n✅ Pipeline complete!")
    _safe_console_print(f"   Output folder: {OUTPUT_DIR}")

    if _run_number is not None:
        try:
            _run_summary = {
                "docs_total": int(len(docs_df)),
                "responses_total": int(len(responses_df)),
                "final_gf_rows": int(len(dernier_df_for_gf) + len(ancien_df)),
                "discrepancies_count": int(len(discrepancies)),
                "artifacts_registered_count": int(_arts_registered),
                "consultant_report_memory_rows_loaded": int(len(persisted_df)),
            }
            if _RUN_CONTROL_CONTEXT:
                _run_summary["run_mode"] = _RUN_CONTROL_CONTEXT.get("run_mode")
                _run_summary["input_files"] = _RUN_CONTROL_CONTEXT.get("input_files", {})

            update_run_metadata(
                RUN_MEMORY_DB,
                _run_number,
                core_version=RUN_MEMORY_CORE_VERSION,
                summary_json=json.dumps(_run_summary, sort_keys=True),
            )
            finalize_run_success(RUN_MEMORY_DB, _run_number)
            _ACTIVE_RUN_FINALIZED = True
        except Exception as _run_finalize_err:
            _safe_console_print(f"  [WARN] Run history finalize error (non-fatal): {_run_finalize_err}")

    return {
        "total_versions": total,
        "dernier_indices": int(dernier_count),
        "anomalies": int(anomaly_count),
        "routed_ok": int(ok_count),
        "routed_ambiguous": int(ambiguous_count),
        "unrouted": int(unrouted),
        "emetteur_mismatch": int(mismatch_count),
        "discrepancies_total": len(discrepancies),
        "discrepancies_review_required": len(disc_review),
        "discrepancies_cosmetic": len(disc_cosmetic),
        "recon_mid_before": mid_before,
        "recon_mid_after": mid_after,
        "recon_mif_before": mif_before,
        "recon_mif_after": mif_after,
        "recon_events": len(recon_log),
    }


# ─────────────────────────────────────────────────────────────
# DISCREPANCY COMPUTATION (Patch B+C+D)
# ─────────────────────────────────────────────────────────────

def _parse_gf_sheet_data(ws, struct: dict) -> dict:
    """
    Part 1 / Patch E: Parse a GF sheet's data rows.

    Returns {(numero_clean, indice_clean): [row_dict, row_dict, ...]}
    — a LIST per key to handle duplicate rows correctly.

    Patch E additions:
      - gf_visa_global: raw value from VISA GLOBAL column (col 15)
      - gf_has_sas_ref: True if any approver column contains "SAS" + "REF"
        (or any "SAS" variant used as a status in the approbateurs section)
    """
    col_map    = struct.get("col_map", {})
    data_start = struct.get("data_start_row", 10)
    approvers  = struct.get("approvers", [])   # [{name, date_col, num_col, statut_col}, ...]

    num_col = col_map.get("numero", 7)
    ind_col = col_map.get("indice", 8)
    doc_col = col_map.get("document", 0)
    tit_col = col_map.get("titre", 1)
    dat_col = col_map.get("date_diffusion", 2)
    lot_col = col_map.get("lot", 3)
    typ_col = col_map.get("type_doc", 4)
    vg_col  = col_map.get("visa_global", 15)   # Patch E

    # Build set of all approver data columns (num + statut) for SAS scan
    approver_data_cols: set = set()
    for ap in approvers:
        approver_data_cols.add(ap.get("num_col", 999))
        approver_data_cols.add(ap.get("statut_col", 999))
    # Also scan all columns from approbateurs start onwards (robust fallback)
    app_start = col_map.get("approbateurs", 16)

    gf_docs: dict = _defaultdict(list)
    max_row = ws.max_row or 5000
    for row in ws.iter_rows(min_row=data_start, max_row=max_row, values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip().upper() in ("", "DOCUMENT", "NONE"):
            continue

        numero_raw = row[num_col] if len(row) > num_col else None
        if numero_raw is None:
            continue

        numero_clean = normalize_numero_for_compare(numero_raw)
        indice_raw   = row[ind_col] if len(row) > ind_col else None
        indice_clean = normalize_indice_for_compare(indice_raw)

        # Patch E: visa_global and gf_has_sas_ref
        visa_global_raw = row[vg_col] if len(row) > vg_col else None

        # Scan all approver-section columns (app_start onwards) for SAS text
        gf_has_sas_ref = False
        for col_idx in range(app_start, len(row)):
            val = row[col_idx]
            if val is None:
                continue
            val_str = str(val).lower().replace("\n", " ").strip()
            if "sas" in val_str and "ref" in val_str:
                gf_has_sas_ref = True
                break

        key = (numero_clean, indice_clean)
        gf_docs[key].append({
            "document":          row[doc_col] if len(row) > doc_col else None,
            "titre":             row[tit_col] if len(row) > tit_col else None,
            "date_diffusion":    row[dat_col] if len(row) > dat_col else None,
            "lot":               row[lot_col] if len(row) > lot_col else None,
            "type_doc":          row[typ_col] if len(row) > typ_col else None,
            "numero":            numero_raw,
            "indice":            indice_raw,
            "numero_normalized": numero_clean,
            "indice_normalized": indice_clean,
            # Patch E
            "gf_visa_global":  visa_global_raw,
            "gf_has_sas_ref":  gf_has_sas_ref,
        })
    for key, rows in gf_docs.items():
        gf_docs[key] = sorted(rows, key=_gf_row_stable_key)
    return dict(gf_docs)


def _is_excluded_sheet_for_discrepancy(
    sheet_name: str,
    excluded_sheets: set,
    sheet_year_filters: dict,
) -> bool:
    """
    Patch C: Return True if an entire sheet should be skipped in discrepancy.
    We skip fully excluded sheets and let year-filtered sheets pass through
    (they are handled at document level).
    """
    return sheet_name in excluded_sheets


def _build_ged_families(
    ged_sheet_docs: pd.DataFrame,
    gf_numeros_in_sheet: set = None,
) -> list:
    """
    Part 2: Group GED rows into document FAMILIES for comparison against GF.

    A FAMILY = a set of GED rows that represent ONE logical document.

    Algorithm (two-pass):
      Pass 1 — Group by (numero_normalized).
        All GED rows with the same numero are the same logical document,
        regardless of indice.  The family representative is the row with
        the dernier indice (picked deterministically: last alpha order).

      Pass 2 — Merge proto-families with different numeros that share
        (lot_normalized, type_de_doc) AND title_similarity >= 0.75.
        This catches cases where GED submitted the same document under
        slightly different numbers.

    Returns: list of family dicts:
      {
        "family_id": str,
        "numero"   : str,    ← representative numero (may be list for merged)
        "indice"   : str,
        "title"    : str,    ← representative title (best normalized title)
        "date"     : any,
        "doc_code" : str,
        "all_numeros": set[str],
        "all_keys" : set[(num, ind)],
        "member_count": int,
      }
    """
    if ged_sheet_docs.empty:
        return []

    # ── Pass 1: group by numero_normalized ───────────────────────────────────
    proto_families: dict = {}  # num_clean → {rows: [...], best_row: ...}

    for _, row in ged_sheet_docs.iterrows():
        num = normalize_numero_for_compare(
            row.get("numero_normalized") or row.get("numero")
        )
        ind     = normalize_indice_for_compare(row.get("indice"))
        title   = str(row.get("libelle_du_document", "") or "")
        date    = row.get("date_diffusion") or row.get("cree_le")
        doc_code = str(row.get("libelle_du_document", "") or "")[:80]
        doc_id  = str(row.get("doc_id", "") or "")

        if num not in proto_families:
            proto_families[num] = {
                "best_row":    {"num": num, "ind": ind, "title": title, "date": date, "doc_code": doc_code},
                "all_keys":    set(),
                "all_doc_ids": set(),   # ← Patch E: track doc_ids for SAS lookup
                "member_count": 0,
                "lot":  str(row.get("lot_normalized", "") or ""),
                "type_doc": str(row.get("type_de_doc", "") or ""),
            }
        pf = proto_families[num]
        pf["all_keys"].add((num, ind))
        pf["member_count"] += 1
        if doc_id:
            pf["all_doc_ids"].add(doc_id)

        # Choose best representative: prefer longer, more descriptive title
        cur_norm = normalize_title_for_compare(pf["best_row"]["title"])
        new_norm = normalize_title_for_compare(title)
        if len(new_norm) > len(cur_norm):
            pf["best_row"] = {"num": num, "ind": ind, "title": title, "date": date, "doc_code": doc_code}

    # ── Pass 2: merge proto-families with same (lot, type_doc) + title_sim ≥ 0.75
    # CRITICAL CONSTRAINT: only merge numeros that have NO exact GF match.
    # Rows whose numero IS in GF should each remain their own family so they can
    # independently match their own GF row. Merging GF-present numeros would
    # cause their GF rows to become "unmatched" → false MISSING_IN_GED_HISTORICAL.
    gf_nums = gf_numeros_in_sheet or set()

    # Union-Find approach for merging
    parent = {num: num for num in proto_families}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[ry] = rx

    nums_list = list(proto_families.keys())
    for i in range(len(nums_list)):
        for j in range(i + 1, len(nums_list)):
            na, nb = nums_list[i], nums_list[j]
            # Only merge if NEITHER numero has a direct GF match.
            # If either is in GF, it must match its own GF row independently.
            if na in gf_nums or nb in gf_nums:
                continue
            pfa, pfb = proto_families[na], proto_families[nb]
            # Same (lot, type_doc) is prerequisite
            if pfa["lot"] != pfb["lot"] or pfa["type_doc"] != pfb["type_doc"]:
                continue
            # Check title similarity
            sim = title_similarity(pfa["best_row"]["title"], pfb["best_row"]["title"])
            if sim >= 0.75:
                union(na, nb)

    # ── Build final families from merged groups ───────────────────────────────
    groups: dict = {}  # root → merged family dict
    for num in nums_list:
        root = find(num)
        if root not in groups:
            groups[root] = {
                "family_id":   root,
                "all_numeros": set(),
                "all_keys":    set(),
                "all_doc_ids": set(),   # Patch E
                "member_count": 0,
                "best_row": proto_families[root]["best_row"],
                "lot":      proto_families[root]["lot"],
                "type_doc": proto_families[root]["type_doc"],
            }
        g = groups[root]
        g["all_numeros"].add(num)
        g["all_keys"].update(proto_families[num]["all_keys"])
        g["all_doc_ids"].update(proto_families[num]["all_doc_ids"])   # Patch E
        g["member_count"] += proto_families[num]["member_count"]

        # Keep best representative (longest normalized title)
        cur_n  = normalize_title_for_compare(g["best_row"]["title"])
        cand_n = normalize_title_for_compare(proto_families[num]["best_row"]["title"])
        if len(cand_n) > len(cur_n):
            g["best_row"] = proto_families[num]["best_row"]

    # ── Convert to list of family dicts ──────────────────────────────────────
    families = []
    for root, g in groups.items():
        br = g["best_row"]
        families.append({
            "family_id":   g["family_id"],
            "numero":      br["num"],
            "indice":      br["ind"],
            "title":       br["title"],
            "date":        br["date"],
            "doc_code":    br["doc_code"],
            "all_numeros": g["all_numeros"],
            "all_keys":    g["all_keys"],
            "all_doc_ids": g["all_doc_ids"],   # Patch E
            "member_count": g["member_count"],
            "lot":      g["lot"],
            "type_doc": g["type_doc"],
        })

    return families


def _compute_discrepancies(
    dernier_df: pd.DataFrame,
    gf_filepath: str,
    debug_dir: str = None,
    excluded_sheets: set = None,
    sheet_year_filters: dict = None,
    responses_df: pd.DataFrame = None,   # Patch E: SAS data for GED docs
    data_date: _dt.date = None,          # Patch E: reference date for recency
) -> tuple:  # returns (discrepancies_list, gf_by_sheet_dict)
    """
    Compare GED dernier indices against existing GF data.

    Patch B: Uses normalize_date_for_compare / normalize_title_for_compare
             so cosmetic differences don't create false mismatches.
    Patch C: excluded_sheets and sheet_year_filters are passed in so GF rows
             from excluded populations are skipped entirely.

    Discrepancy types:
      MISSING_IN_GF      – GED dernier indice not present in GF at all
      MISSING_IN_GED     – GF has a row whose numero+indice is not in GED dernier
      INDICE_MISMATCH    – GF has this numero but with a different indice
      TITRE_MISMATCH     – numero+indice matched but title significantly differs
      DATE_MISMATCH      – numero+indice matched but date_diffusion differs
      SHEET_MISMATCH     – numero found in GF but on a different sheet than routed

    Returns list of discrepancy dicts (each has a 'severity' key after
    classify_discrepancy is applied in run_pipeline).
    """
    import openpyxl
    from routing import read_all_gf_sheet_structures

    excluded_sheets    = excluded_sheets    or set()
    sheet_year_filters = sheet_year_filters or {}

    # Patch E: Build SAS lookup once for all sheets
    sas_lookup: dict = {}
    if responses_df is not None:
        sas_lookup = _build_sas_lookup(responses_df)
        _safe_console_print(f"  [Patch E] SAS lookup built: {len(sas_lookup)} doc_ids with SAS data")
    if data_date is None:
        data_date = _dt.date.today()

    discrepancies = []

    try:
        wb = openpyxl.load_workbook(gf_filepath, read_only=True, data_only=True)
    except Exception as e:
        _safe_console_print(f"  WARNING: Could not read existing GF for discrepancy check: {e}")
        return discrepancies, {}

    sheet_names = [s for s in wb.sheetnames if not s.upper().startswith("OLD")]

    # Remove fully-excluded sheets (Patch C)
    sheet_names_for_compare = [
        s for s in sheet_names
        if not _is_excluded_sheet_for_discrepancy(s, excluded_sheets, sheet_year_filters)
    ]
    excluded_skipped = [s for s in sheet_names if s not in sheet_names_for_compare]
    if excluded_skipped:
        _safe_console_print(f"  [Patch C] Skipping {len(excluded_skipped)} excluded sheets: {excluded_skipped}")

    structures = read_all_gf_sheet_structures(gf_filepath, sheet_names_for_compare)

    if debug_dir:
        _write_gf_schema_debug(debug_dir, structures)

    # ── Build GF indexes ──────────────────────────────────────────────────────
    # gf_by_sheet[sheet] = {(num, ind): [row, row, ...]}   ← LIST per key (Part 1)
    # full_gf_index[num] = [(sheet, ind, row), ...]
    gf_by_sheet: dict = {}
    full_gf_index: dict = {}
    # For gf_duplicates debug: (sheet, num, ind) → count
    gf_dup_counts: dict = {}

    for sheet_name in sheet_names_for_compare:
        ws = wb[sheet_name]
        struct = structures.get(sheet_name, {})
        gf_docs = _parse_gf_sheet_data(ws, struct)  # → dict[(num,ind)] → [rows]
        gf_by_sheet[sheet_name] = gf_docs

        for (num_c, ind_c), rows in gf_docs.items():
            for row in rows:
                full_gf_index.setdefault(num_c, [])
                full_gf_index[num_c].append((sheet_name, ind_c, row))
            if len(rows) > 1:
                gf_dup_counts[(sheet_name, num_c, ind_c)] = len(rows)

    total_dup_keys = len(gf_dup_counts)
    total_dup_excess = sum(v - 1 for v in gf_dup_counts.values())
    if total_dup_keys:
        _safe_console_print(f"  [Part 1] GF duplicate keys on active sheets: "
              f"{total_dup_keys} keys, {total_dup_excess} excess rows")

    # Write gf_duplicates.xlsx debug artifact
    if debug_dir:
        _write_gf_duplicates_debug(debug_dir, gf_dup_counts)

    # ── Per-sheet comparison ─────────────────────────────────────────────────
    for sheet_name in sheet_names_for_compare:
        gf_docs = gf_by_sheet.get(sheet_name, {})
        ged_sheet_docs = dernier_df[dernier_df["gf_sheet_name"] == sheet_name]
        min_year = sheet_year_filters.get(sheet_name)

        # ── Part 2: Build GED families ──────────────────────────────────────
        # Group GED rows into logical document families so that multiple GED
        # submissions for the same document don't generate N discrepancies.
        # Pass GF numeros so Pass 2 merge skips numeros that have a GF match
        # (those rows must remain independent to match their own GF row).
        gf_numeros_in_sheet = {k[0] for k in gf_docs.keys()}
        ged_families = _build_ged_families(ged_sheet_docs, gf_numeros_in_sheet)

        # Collect RAW GED numeros directly from original rows (NOT from families)
        # This is used for the MISSING_IN_GED_HISTORICAL check:
        # "GED has this numero at some indice → old GF row is historical"
        # Using family["all_numeros"] would contaminate this with merged numeros
        # from different logical documents.
        all_ged_numeros_in_sheet: set = set()
        for _, raw_row in ged_sheet_docs.iterrows():
            n = normalize_numero_for_compare(
                raw_row.get("numero_normalized") or raw_row.get("numero")
            )
            if n:
                all_ged_numeros_in_sheet.add(n)

        # Track which GF keys have been matched (exact or fuzzy) — used to
        # suppress MISSING_IN_GED for GF rows that ARE covered by a GED family.
        matched_gf_keys: set = set()

        # Track GF keys where DUPLICATE_ACTIVE_IN_GF has already been emitted
        dup_flagged_gf_keys: set = set()

        for family in ged_families:
            numero    = family["numero"]
            indice    = family["indice"]
            ged_titre_raw = family["title"]
            ged_date_raw  = family["date"]
            doc_code  = family["doc_code"]
            all_fam_keys = sorted(family["all_keys"])  # all (num, ind) in this family

            # ── Step 1: Exact-key match: try every (num, ind) in the family ──
            exact_gf_key = None
            exact_cands  = None
            # CRITICAL: initialize to -1.0, not 0.0.
            # When GED libelle is a raw doc-code (e.g. "P17_T2_..._130027_A.pdf") and the
            # GF row has a meaningful title ("Platelage bois MOSO BAMBOO"), title_similarity
            # returns 0.0.  Using 0.0 as the floor means "0.0 > 0.0" is False and the exact
            # key match is never registered, causing a false MISSING_IN_GF_TRUE.
            # Initializing to -1.0 ensures ANY exact (numero, indice) match is accepted.
            exact_best_sim = -1.0

            for fam_key in all_fam_keys:
                if fam_key in gf_docs:
                    cands = gf_docs[fam_key]
                    for cand in cands:
                        sim = title_similarity(ged_titre_raw, cand.get("titre", ""))
                        if sim > exact_best_sim:
                            exact_best_sim = sim
                            exact_gf_key   = fam_key
                            exact_cands    = cands

            if exact_gf_key is not None:
                # ── Exact key match: full comparison ──────────────────────
                matched_gf_keys.add(exact_gf_key)
                scored = sorted(
                    [(title_similarity(ged_titre_raw, c.get("titre", "")), c)
                     for c in exact_cands],
                    key=lambda x: (-x[0], _gf_row_stable_key(x[1])),
                )
                best_sim, best_cand = scored[0]

                # Flag structural duplicates (once per GF key)
                if exact_gf_key not in dup_flagged_gf_keys and len(exact_cands) > 1:
                    uniq_titles = set(
                        normalize_title_for_compare(c.get("titre", ""))
                        for c in exact_cands
                    )
                    if len(uniq_titles) > 1:
                        dup_flagged_gf_keys.add(exact_gf_key)
                        discrepancies.append({
                            "sheet_name": sheet_name,
                            "document_code": doc_code,
                            "numero": numero,
                            "indice": indice,
                            "field": "DUPLICATE",
                            "ged_value": ged_titre_raw[:80],
                            "gf_value": f"{len(exact_cands)} GF rows with same (num,ind)",
                            "ged_value_normalized": normalize_title_for_compare(ged_titre_raw),
                            "gf_value_normalized": str(uniq_titles),
                            "flag_type": "DUPLICATE_ACTIVE_IN_GF",
                            "title_similarity": best_sim,
                            "date_diff_days": None,
                            "is_excluded_population": False,
                        })

                # Title comparison
                gf_titre_raw = str(best_cand.get("titre", "") or "")
                if ged_titre_raw and gf_titre_raw and best_sim < 0.85:
                    discrepancies.append({
                        "sheet_name": sheet_name,
                        "document_code": doc_code,
                        "numero": numero,
                        "indice": indice,
                        "field": "TITRE",
                        "ged_value": ged_titre_raw[:100],
                        "gf_value": gf_titre_raw[:100],
                        "ged_value_normalized": normalize_title_for_compare(ged_titre_raw),
                        "gf_value_normalized": normalize_title_for_compare(gf_titre_raw),
                        "flag_type": "TITRE_MISMATCH",
                        "title_similarity": round(best_sim, 3),
                        "date_diff_days": None,
                        "is_excluded_population": False,
                    })

                # Date comparison
                gf_date_raw = best_cand.get("date_diffusion")
                if ged_date_raw is not None and gf_date_raw is not None:
                    diff = date_diff_days(ged_date_raw, gf_date_raw)
                    if diff is not None and diff > 0:
                        discrepancies.append({
                            "sheet_name": sheet_name,
                            "document_code": doc_code,
                            "numero": numero,
                            "indice": indice,
                            "field": "DATE_DIFFUSION",
                            "ged_value": str(ged_date_raw),
                            "gf_value": str(gf_date_raw),
                            "ged_value_normalized": normalize_date_for_compare(ged_date_raw),
                            "gf_value_normalized": normalize_date_for_compare(gf_date_raw),
                            "flag_type": "DATE_MISMATCH",
                            "title_similarity": None,
                            "date_diff_days": diff,
                            "is_excluded_population": False,
                        })

            else:
                # ── No exact key match: try cross-key title fuzzy match ────
                # If a GF row exists with title_similarity >= 0.75 to this
                # family's representative title, suppress MISSING_IN_GF.
                # We do NOT emit TITRE/DATE discrepancies for fuzzy matches
                # (different numero = different document reference).
                fuzzy_gf_key = None
                fuzzy_best_key = None
                fuzzy_best_sim = 0.0
                for gf_key, gf_cands in gf_docs.items():
                    for gf_cand in gf_cands:
                        sim = title_similarity(ged_titre_raw, gf_cand.get("titre", ""))
                        cand_key = (gf_key, _gf_row_stable_key(gf_cand))
                        if (
                            sim > fuzzy_best_sim or
                            (
                                fuzzy_best_key is not None and
                                sim == fuzzy_best_sim and
                                cand_key < fuzzy_best_key
                            ) or
                            (fuzzy_best_key is None and sim == fuzzy_best_sim and sim >= 0.75)
                        ):
                            fuzzy_best_sim = sim
                            fuzzy_best_key = cand_key
                            if sim >= 0.75:
                                fuzzy_gf_key = gf_key

                if fuzzy_gf_key is not None:
                    # Fuzzy match found → suppress MISSING_IN_GF, mark GF key covered
                    matched_gf_keys.add(fuzzy_gf_key)
                    # No TITRE or DATE discrepancies for cross-key fuzzy matches
                else:
                    # No match at all — check INDICE/SHEET/MISSING_IN_GF
                    if numero in full_gf_index:
                        all_occ = full_gf_index[numero]
                        same_sheet_other_ind = [
                            (sn, ind2, row) for sn, ind2, row in all_occ
                            if sn == sheet_name and ind2 != indice
                        ]
                        other_sheet_same_ind = [
                            (sn, ind2, row) for sn, ind2, row in all_occ
                            if sn != sheet_name and ind2 == indice
                        ]

                        if same_sheet_other_ind:
                            same_sheet_other_ind = sorted(
                                same_sheet_other_ind,
                                key=lambda x: (str(x[1]), _gf_row_stable_key(x[2])),
                            )
                            gf_ind = same_sheet_other_ind[0][1]
                            # Patch D: include date/title info so the relaxation
                            # pass can accept INDICE_MISMATCH when dates match
                            _gf_row_ind  = same_sheet_other_ind[0][2]
                            _gf_date_ind = _gf_row_ind.get("date_diffusion")
                            _ind_date_diff = date_diff_days(ged_date_raw, _gf_date_ind)
                            _ind_title_sim = title_similarity(
                                ged_titre_raw,
                                str(_gf_row_ind.get("titre") or ""),
                            )
                            discrepancies.append({
                                "sheet_name": sheet_name,
                                "document_code": doc_code,
                                "numero": numero,
                                "indice": indice,
                                "field": "INDICE",
                                "ged_value": indice,
                                "gf_value": gf_ind,
                                "ged_value_normalized": indice,
                                "gf_value_normalized": gf_ind,
                                "flag_type": "INDICE_MISMATCH",
                                "title_similarity": round(_ind_title_sim, 3),
                                "date_diff_days": _ind_date_diff,
                                "is_excluded_population": False,
                                # For Patch D reconstruction
                                "ged_date_raw": str(ged_date_raw or ""),
                                "gf_date_raw":  str(_gf_date_ind or ""),
                            })
                        elif other_sheet_same_ind:
                            other_sheet_same_ind = sorted(
                                other_sheet_same_ind,
                                key=lambda x: (str(x[0]), str(x[1]), _gf_row_stable_key(x[2])),
                            )
                            gf_sheet = other_sheet_same_ind[0][0]
                            discrepancies.append({
                                "sheet_name": sheet_name,
                                "document_code": doc_code,
                                "numero": numero,
                                "indice": indice,
                                "field": "SHEET",
                                "ged_value": sheet_name,
                                "gf_value": gf_sheet,
                                "ged_value_normalized": sheet_name,
                                "gf_value_normalized": gf_sheet,
                                "flag_type": "SHEET_MISMATCH",
                                "title_similarity": None,
                                "date_diff_days": None,
                                "is_excluded_population": False,
                            })
                        else:
                            # Patch E: Classify MISSING_IN_GF using SAS timing
                            mig_subtype, _ = _classify_missing_in_gf(
                                family, sas_lookup, data_date
                            )
                            sas_info = {}
                            for did in _sorted_family_doc_ids(family):
                                if did in sas_lookup:
                                    sas_info = sas_lookup[did]
                                    break
                            discrepancies.append({
                                "sheet_name":    sheet_name,
                                "document_code": doc_code,
                                "numero":        numero,
                                "indice":        indice,
                                "field":         "DOCUMENT",
                                "ged_value":     f"{numero}/{indice}",
                                "gf_value":      "NOT FOUND IN GF (numero on other sheet)",
                                "ged_value_normalized": "",
                                "gf_value_normalized":  "",
                                "flag_type":     mig_subtype,
                                "title_similarity": None,
                                "date_diff_days":   None,
                                "is_excluded_population": False,
                                # Patch E extras
                                "sas_status_type":  sas_info.get("sas_status_type"),
                                "sas_result":       sas_info.get("sas_result"),
                                "sas_date":         str(sas_info.get("sas_date") or ""),
                                # GF_INSERTION_QUEUE fields (Part 9)
                                "gfi_numero":  numero,
                                "gfi_indice":  indice,
                                "gfi_titre":   ged_titre_raw,
                                "gfi_date":    str(ged_date_raw or ""),
                                "gfi_sheet":   sheet_name,
                                "gfi_lot":     family.get("lot", ""),
                                "gfi_type_doc": family.get("type_doc", ""),
                            })
                    else:
                        # Patch E: Classify MISSING_IN_GF using SAS timing
                        mig_subtype, _ = _classify_missing_in_gf(
                            family, sas_lookup, data_date
                        )
                        sas_info = {}
                        for did in _sorted_family_doc_ids(family):
                            if did in sas_lookup:
                                sas_info = sas_lookup[did]
                                break
                        discrepancies.append({
                            "sheet_name":    sheet_name,
                            "document_code": doc_code,
                            "numero":        numero,
                            "indice":        indice,
                            "field":         "DOCUMENT",
                            "ged_value":     f"{numero}/{indice}",
                            "gf_value":      "NOT FOUND IN GF",
                            "ged_value_normalized": "",
                            "gf_value_normalized":  "",
                            "flag_type":     mig_subtype,
                            "title_similarity": None,
                            "date_diff_days":   None,
                            "is_excluded_population": False,
                            # Patch E extras
                            "sas_status_type":  sas_info.get("sas_status_type"),
                            "sas_result":       sas_info.get("sas_result"),
                            "sas_date":         str(sas_info.get("sas_date") or ""),
                            # GF_INSERTION_QUEUE fields (Part 9)
                            "gfi_numero":   numero,
                            "gfi_indice":   indice,
                            "gfi_titre":    ged_titre_raw,
                            "gfi_date":     str(ged_date_raw or ""),
                            "gfi_sheet":    sheet_name,
                            "gfi_lot":      family.get("lot", ""),
                            "gfi_type_doc": family.get("type_doc", ""),
                        })

        # ── MISSING_IN_GED: GF rows not covered by any GED family ─────────
        # Part 4 / Patch E: classify into subtypes
        for (gf_num, gf_ind), gf_rows in gf_docs.items():
            if (gf_num, gf_ind) not in matched_gf_keys:
                gf_row    = sorted(gf_rows, key=_gf_row_stable_key)[0]
                doc_label = str(gf_row.get("document", "") or "")[:80]

                # Historical: GED has this numero but at a newer indice
                is_historical = gf_num in all_ged_numeros_in_sheet

                # Year-filter exclusion (Patch C)
                gf_date_raw = gf_row.get("date_diffusion")
                is_excluded_by_year = False
                if not is_historical and min_year and gf_date_raw:
                    try:
                        gf_year = pd.to_datetime(gf_date_raw).year
                        if gf_year < min_year:
                            is_excluded_by_year = True
                    except Exception:
                        pass

                # Patch E: refined subtype classification
                flag_type_val = _classify_missing_in_ged(
                    gf_num, gf_ind, gf_rows,
                    is_historical, is_excluded_by_year,
                    gf_dup_counts, sheet_name,
                )

                discrepancies.append({
                    "sheet_name":    sheet_name,
                    "document_code": doc_label,
                    "numero":        gf_num,
                    "indice":        gf_ind,
                    "field":         "DOCUMENT",
                    "ged_value":     "NOT IN GED DERNIER",
                    "gf_value":      f"{gf_num}/{gf_ind}",
                    "ged_value_normalized": "",
                    "gf_value_normalized":  "",
                    "flag_type":     flag_type_val,
                    "title_similarity": None,
                    "date_diff_days":   None,
                    "is_excluded_population": is_excluded_by_year,
                    # Patch E: GF SAS context
                    "gf_visa_global":  gf_row.get("gf_visa_global"),
                    "gf_has_sas_ref":  gf_row.get("gf_has_sas_ref", False),
                })

    return discrepancies, gf_by_sheet


def _write_missing_in_ged_diagnosis(
    diag_path: str,
    true_only_path: str,
    summary_path: str,
    discrepancies: list,
):
    """
    Part 7 / Patch E: Write MISSING_IN_GED diagnosis outputs.

    Columns written: sheet, numero, indice, document_code,
                     flag_type (subtype), gf_visa_global, gf_has_sas_ref, severity
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    mid_types = {
        "MISSING_IN_GED_TRUE", "MISSING_IN_GED_HISTORICAL",
        "MISSING_IN_GED_GF_SAS_REF", "MISSING_IN_GED_GF_DUPLICATE_ROW",
        "MISSING_IN_GED_EXCLUDED", "MISSING_IN_GED_AMBIGUOUS",
        "MISSING_IN_GED_FAMILY_MATCH_MISSED",
        # legacy
        "MISSING_IN_GED",
    }

    mid_recs = [d for d in discrepancies if d.get("flag_type") in mid_types]
    if not mid_recs:
        _safe_console_print("    [Patch E] No MISSING_IN_GED records found — skipping diagnosis files")
        return

    headers = [
        "Sheet", "Numéro", "Indice", "Document Code",
        "Subtype", "GF Visa Global", "GF Has SAS REF",
        "Severity",
    ]
    colour_map = {
        "MISSING_IN_GED_TRUE":            "FFCCCC",
        "MISSING_IN_GED_GF_SAS_REF":      "FFE8CC",
        "MISSING_IN_GED_HISTORICAL":      "E8F5E9",
        "MISSING_IN_GED_GF_DUPLICATE_ROW":"FFF9C4",
        "MISSING_IN_GED_EXCLUDED":        "E3F2FD",
        "MISSING_IN_GED_AMBIGUOUS":       "F3E5F5",
        "MISSING_IN_GED_FAMILY_MATCH_MISSED": "FFF3E0",
    }

    def _make_wb(recs):
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "MISSING_IN_GED"
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        for ri, rec in enumerate(recs, 2):
            ft = rec.get("flag_type", "")
            sev = classify_discrepancy(rec)
            row_data = [
                rec.get("sheet_name", ""),
                rec.get("numero", ""),
                rec.get("indice", ""),
                rec.get("document_code", "")[:80],
                ft,
                rec.get("gf_visa_global", ""),
                "YES" if rec.get("gf_has_sas_ref") else "",
                sev,
            ]
            fill_color = colour_map.get(ft, "FFFFFF")
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
        for col_idx, width in zip(range(1, len(headers) + 1),
                                   [45, 12, 8, 60, 38, 15, 15, 18]):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        return wb2

    Path(diag_path).parent.mkdir(parents=True, exist_ok=True)
    _make_wb(mid_recs).save(diag_path)

    true_recs = [d for d in mid_recs if d.get("flag_type") == "MISSING_IN_GED_TRUE"]
    _make_wb(true_recs).save(true_only_path)

    # Summary counts
    from collections import Counter
    counts = Counter(d.get("flag_type") for d in mid_recs)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.cell(row=1, column=1, value="Subtype").font = Font(bold=True)
    ws_s.cell(row=1, column=2, value="Count").font   = Font(bold=True)
    for ri, (ft, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        ws_s.cell(row=ri, column=1, value=ft)
        ws_s.cell(row=ri, column=2, value=cnt)
    ws_s.column_dimensions["A"].width = 42
    ws_s.column_dimensions["B"].width = 10
    wb_s.save(summary_path)

    _safe_console_print(f"    [Patch E] MISSING_IN_GED breakdown: {dict(counts)}")
    _safe_console_print(f"    [Patch E] TRUE only: {len(true_recs)}")


def _write_missing_in_gf_diagnosis(
    diag_path: str,
    true_only_path: str,
    summary_path: str,
    discrepancies: list,
):
    """
    Part 7 / Patch E: Write MISSING_IN_GF diagnosis outputs.

    Columns: sheet, numero, indice, document_code, subtype,
             sas_status_type, sas_result, sas_date, severity,
             GFI fields (for future auto-insertion queue, Part 9)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    mig_types = {
        "MISSING_IN_GF_TRUE", "MISSING_IN_GF_PENDING_SAS",
        "MISSING_IN_GF_RECENT_SAS_REMINDER", "MISSING_IN_GF_RECENT_REFUSAL",
        "MISSING_IN_GF_RECENT_ACCEPTED_SAS", "MISSING_IN_GF_SAME_KEY_COLLISION",
        "MISSING_IN_GF_AMBIGUOUS",
        # legacy
        "MISSING_IN_GF",
    }

    mig_recs = [d for d in discrepancies if d.get("flag_type") in mig_types]
    if not mig_recs:
        _safe_console_print("    [Patch E] No MISSING_IN_GF records found — skipping diagnosis files")
        return

    headers = [
        "Sheet", "Numéro", "Indice", "Document Code",
        "Subtype", "SAS Status Type", "SAS Result", "SAS Date",
        "Severity",
        # GF_INSERTION_QUEUE (Part 9)
        "GFI Titre", "GFI Date", "GFI Lot", "GFI Type Doc",
    ]
    colour_map = {
        "MISSING_IN_GF_TRUE":                 "FFCCCC",
        "MISSING_IN_GF_PENDING_SAS":          "E8F5E9",
        "MISSING_IN_GF_RECENT_SAS_REMINDER":  "FFF9C4",
        "MISSING_IN_GF_RECENT_REFUSAL":       "FFE8CC",
        "MISSING_IN_GF_RECENT_ACCEPTED_SAS":  "E3F2FD",
        "MISSING_IN_GF_SAME_KEY_COLLISION":   "F3E5F5",
        "MISSING_IN_GF_AMBIGUOUS":            "EEEEEE",
    }

    def _make_wb(recs):
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "MISSING_IN_GF"
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="7B2D00")
        for ri, rec in enumerate(recs, 2):
            ft  = rec.get("flag_type", "")
            sev = classify_discrepancy(rec)
            row_data = [
                rec.get("sheet_name", ""),
                rec.get("numero", ""),
                rec.get("indice", ""),
                rec.get("document_code", "")[:80],
                ft,
                rec.get("sas_status_type", ""),
                rec.get("sas_result", ""),
                rec.get("sas_date", ""),
                sev,
                # GFI queue
                rec.get("gfi_titre", "")[:80],
                rec.get("gfi_date", ""),
                rec.get("gfi_lot", ""),
                rec.get("gfi_type_doc", ""),
            ]
            fill_color = colour_map.get(ft, "FFFFFF")
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
        widths = [45, 12, 8, 60, 38, 18, 12, 12, 18, 70, 14, 12, 14]
        for col_idx, width in zip(range(1, len(headers) + 1), widths):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        return wb2

    Path(diag_path).parent.mkdir(parents=True, exist_ok=True)
    _make_wb(mig_recs).save(diag_path)

    true_recs = [d for d in mig_recs if d.get("flag_type") == "MISSING_IN_GF_TRUE"]
    _make_wb(true_recs).save(true_only_path)

    # Summary counts
    from collections import Counter
    counts = Counter(d.get("flag_type") for d in mig_recs)
    Path(summary_path).parent.mkdir(parents=True, exist_ok=True)
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Summary"
    ws_s.cell(row=1, column=1, value="Subtype").font = Font(bold=True)
    ws_s.cell(row=1, column=2, value="Count").font   = Font(bold=True)
    for ri, (ft, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        ws_s.cell(row=ri, column=1, value=ft)
        ws_s.cell(row=ri, column=2, value=cnt)
    ws_s.column_dimensions["A"].width = 42
    ws_s.column_dimensions["B"].width = 10
    wb_s.save(summary_path)

    _safe_console_print(f"    [Patch E] MISSING_IN_GF breakdown: {dict(counts)}")
    _safe_console_print(f"    [Patch E] TRUE only: {len(true_recs)}")


def _write_gf_schema_debug(debug_dir: str, structures: dict):
    """
    Write debug/gf_sheet_schema.xlsx from parsed GF sheet structures.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    Path(debug_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(debug_dir) / "gf_sheet_schema.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GF Sheet Schema"

    headers = [
        "sheet_name",
        "header_row",
        "data_start_row",
        "total_cols",
        "base_col_count",
        "approver_count",
        "col_map",
        "approvers",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row_idx, sheet_name in enumerate(sorted(structures.keys()), 2):
        struct = structures.get(sheet_name, {}) or {}
        approvers = struct.get("approvers", []) or []
        approver_names = ", ".join(
            str(a.get("name", "")).strip()
            for a in approvers
            if str(a.get("name", "")).strip()
        )
        col_map = struct.get("col_map", {}) or {}
        col_map_str = ", ".join(
            f"{key}:{value}" for key, value in sorted(col_map.items(), key=lambda kv: str(kv[0]))
        )
        row = [
            sheet_name,
            struct.get("header_row"),
            struct.get("data_start_row"),
            struct.get("total_cols"),
            struct.get("base_col_count"),
            len(approvers),
            col_map_str,
            approver_names,
        ]
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [42, 12, 14, 10, 14, 14, 60, 80]
    for col_idx, width in zip(range(1, len(headers) + 1), widths):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)


def _write_gf_duplicates_debug(debug_dir: str, gf_dup_counts: dict):
    """
    Part 1: Write debug/gf_duplicates.xlsx.

    Lists every (sheet_name, numero, indice) key that appeared more than once
    in the existing GF, along with the duplicate count.

    Columns: sheet_name, numero, indice, count_of_duplicates
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    Path(debug_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(debug_dir) / "gf_duplicates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GF Duplicate Keys"

    headers = ["sheet_name", "numero", "indice", "count_of_duplicates"]
    header_fill = PatternFill("solid", fgColor="C00000")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    # Sort by count desc, then sheet, num, ind for readability
    rows_sorted = sorted(
        gf_dup_counts.items(),
        key=lambda kv: (-kv[1], kv[0][0], kv[0][1], kv[0][2]),
    )

    row_fill = PatternFill("solid", fgColor="FFE0E0")
    for r_idx, ((sheet_name, num, ind), count) in enumerate(rows_sorted, 2):
        ws.cell(row=r_idx, column=1, value=sheet_name)
        ws.cell(row=r_idx, column=2, value=num)
        ws.cell(row=r_idx, column=3, value=ind)
        c = ws.cell(row=r_idx, column=4, value=count)
        if count >= 3:
            for col in range(1, 5):
                ws.cell(row=r_idx, column=col).fill = row_fill

    # Summary row
    total_keys = len(rows_sorted)
    total_excess = sum(v - 1 for v in gf_dup_counts.values())
    summary_row = total_keys + 2
    ws.cell(row=summary_row, column=1, value="TOTAL duplicate keys").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=total_keys).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="TOTAL excess rows").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=total_excess).font = Font(bold=True)

    for col_idx, width in zip(range(1, len(headers) + 1), [34, 18, 10, 20]):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)


if __name__ == "__main__":
    try:
        from src.run_orchestrator import RUN_MODE_FULL, run_pipeline_controlled

        result = run_pipeline_controlled(
            run_mode=RUN_MODE_FULL,
            ged_path=str(GED_FILE),
            gf_path=str(GF_FILE),
            reports_dir=str(CONSULTANT_REPORTS_ROOT) if CONSULTANT_REPORTS_ROOT.exists() else None,
        )
        if not result.get("success"):
            raise RuntimeError("; ".join(result.get("errors", []) or ["Controlled execution failed"]))
    except Exception as exc:
        if _ACTIVE_RUN_NUMBER is not None and not _ACTIVE_RUN_FINALIZED:
            try:
                finalize_run_failure(
                    RUN_MEMORY_DB,
                    _ACTIVE_RUN_NUMBER,
                    f"{type(exc).__name__}: {exc}",
                )
            except Exception:
                pass
        traceback.print_exc()
        raise
